import os,glob
import time
import pyautogui
import pyperclip
import shutil
import PyPDF2
import img2pdf
import subprocess
import zipfile
import xlwt
import chromedriver_autoinstaller
from openpyxl import Workbook
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains

import natsort ## 숫자 정렬용 라이브러리
import pandas as pd

from datetime import datetime
from django.db import connection
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support import expected_conditions as EC
from pywinauto.keyboard import send_keys
from pywinauto import keyboard    # 단축키 기능을 이용하기 위해서
import pywinauto
from pywinauto import Desktop, Application
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException

from app.models import MemUser
from app.models import MemAdmin
from app.models import MemDeal

def ChromeDriver(isInVisable):
  chrome_ver = chromedriver_autoinstaller.get_chrome_version()
  print(f'크롬 현재버전:{chrome_ver}')
  options = webdriver.ChromeOptions()
  options.add_argument("--disable-proxy-certificate-handler")
  driver = webdriver.Chrome(ChromeDriverManager().install())
  # driver = webdriver.Chrome()

  # 마우스 이동하기
  width, height = pyautogui.size()
  pyautogui.moveTo(width/2, height/2)
  return driver

  # try:
  #   if isInVisable:
  #     options = webdriver.ChromeOptions()
  #     options.add_argument('headless')
  #     options.add_argument('window-size=1920x1080')
  #     options.add_argument("disable-gpu")
  #     options.add_argument('--kiosk-printing')
  #     options.add_argument('--disable-print-preview')
  #     options.add_argument('--disable-dev-shm-usage')
  #     options.add_argument('--no-sandbox')
  #     options.add_argument('--disable-software-rasterizer')
  #     options.add_argument('--disable-extensions')
  #     options.add_argument('--disable-infobars')
  #     options.add_experimental_option('excludeSwitches', ['enable-logging']) 
  #     driver = webdriver.Chrome(executable_path=ChromeDriverManager(version="114.0.5735.90").install(), options=options)
  #   else:
  #     driver = webdriver.Chrome(executable_path=ChromeDriverManager(version="114.0.5735.90").install())
    
    # driver.implicitly_wait(20)
    # print('크롬드라이버실행')
    # return driver
  # except Exception as e:
  #   print('크롬webdriver드라이버 오류:', e)
  #   return None


# def ChromeDriver(isInVisable):

#   # try:
#     driver = webdriver.Chrome(ChromeDriverManager(version="114.0.5735.90").install())
#     options = webdriver.ChromeOptions() 
#     # driver = webdriver.Chrome('C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe')
#     if isInVisable:
#       options.add_argument('headless')
#       options.add_argument('window-size=1920x1080')
#       options.add_argument("disable-gpu")
#       options.add_argument('--kiosk-printing')
#       options.add_argument('--disable-print-preview')
#       options.add_argument('--disable-gpu')
#       options.add_argument('--disable-dev-shm-usage')
#       options.add_argument('--no-sandbox')
#       options.add_argument('--disable-software-rasterizer')
#       options.add_argument('--disable-extensions')
#       options.add_argument('--disable-infobars')
#       options.add_experimental_option('excludeSwitches', ['enable-logging']) 
#     driver = webdriver.Chrome(options=options)
#     driver.implicitly_wait(20)
#     print('크롬드라이버실행')
#   # except:
#   #   print('크롬webdriver드라이버 오류')
#   #   return JsonResponse({'data':' webdriver드라이버 오류'},safe=False)
#     return driver




# 홈택스 인증서로그인  사용자 로그인
def conHometaxLogin_Personal(result,isNotShowChrome):
  driver = ChromeDriver(isNotShowChrome)
  driver.get('http://www.hometax.go.kr/websquare/websquare_cdn.html?w2xPath=/ui/pp/index.xml')
  driver.implicitly_wait(3)
  win = pyautogui.getActiveWindow();win.maximize();time.sleep(0.5) #윈도우 최대화
  loginbtn = driver.find_element(By.XPATH,r'/html/body/div[1]/div[1]/div[2]/div/div[1]/div[1]/div[2]/div[1]/ul/li[3]/a/span');    time.sleep(1);    loginbtn.click();    time.sleep(2)                                            
  # 팝업뜨는 경우 없애기
  # 현재 윈도우 핸들 가져오기
  main_window_handle = driver.current_window_handle
  # 버튼 클릭 등으로 새로운 팝업 윈도우가 열리는 이벤트 발생
  # 팝업 윈도우 핸들 가져오기
  popup_window_handle = None
  while not popup_window_handle:
    if len(driver.window_handles)>1:
      for handle in driver.window_handles:
          if handle != main_window_handle:
              popup_window_handle = handle
              # 팝업 윈도우로 전환
              driver.switch_to.window(popup_window_handle)
              # 팝업 윈도우에서 필요한 동작 수행
              group2152 = driver.find_element(By.CSS_SELECTOR,'#group2152')
              group218 =  driver.find_element(By.ID,'group218')#보안카드 취소버튼
              if group2152:group2152.click(); 
              elif group218:
                group218.click()
                WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(0.5);print('취소합니다 알람')
              else:
                trigger4 = driver.find_element(By.CSS_SELECTOR,'#trigger4')
                if trigger4:trigger4.click();  
              # 메인 윈도우로 전환
              driver.switch_to.window(main_window_handle) ; print('메인 윈도우로 전환')             
              break
    else:
      popup_window_handle = True

  print('홈택스 접속')
  iframe = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, r'/html/body/div[1]/div[2]/iframe')));  driver.switch_to.frame(iframe)
  idLoginBtn = driver.find_element(By.ID,'anchor15');    time.sleep(1);    idLoginBtn.click();    time.sleep(3)
  driver.find_element(By.ID,'group91882156').click();    time.sleep(1)
  driver.find_element(By.ID,'iptUserId').send_keys(result[8]);    time.sleep(0.5)
  driver.find_element(By.ID,'iptUserPw').send_keys(result[9]);    time.sleep(0.5)
  idpwLoginBtn = driver.find_element(By.ID,'group91882177').click();print('로그인버튼');    time.sleep(1)  
  popup_window_handle = None
  while not popup_window_handle:
    if len(driver.window_handles)>1:
      for handle in driver.window_handles:
          if handle != main_window_handle:
              popup_window_handle = handle
              # 팝업 윈도우로 전환
              driver.switch_to.window(popup_window_handle)
              # 팝업 윈도우에서 필요한 동작 수행
              group218 =  driver.find_element(By.ID,'group218')#보안카드 취소버튼
              if group218:
                group218.click()
                WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(0.5);print('취소합니다 알람')
              # 메인 윈도우로 전환
              driver.switch_to.window(main_window_handle) ; print('메인 윈도우로 전환')             
              break
    else:
      popup_window_handle = True
  if  result[4]>=4:#개인사업자인 경우
    try:
      driver.find_element(By.ID,'btnApply').click();time.sleep(1.5);print('사업장선택')
      iframe = driver.find_element(By.CSS_SELECTOR,'#UTXPPAAA24_iframe');    driver.switch_to.frame(iframe);    time.sleep(1);print('iframe UTXPPAAA24_iframe1')
      driver.find_element(By.ID,'iptBsno').clear();driver.find_element(By.ID,'iptBsno').send_keys(result[0].replace('-',''));time.sleep(0.25)
      driver.find_element(By.ID,'trigger85').click();time.sleep(0.5);print('조회하기 버튼')
      table = driver.find_element(By.ID,'grid1_body_table')
      tbody = table.find_element(By.ID,'grid1_body_tbody')  
      if len(tbody.find_elements(By.TAG_NAME,'tr')) >0 : 
        driver.find_element(By.XPATH,'//*[@id="G_grid1___radio_radSelect_0"]').click()         
      driver.find_element(By.ID,'group1818').click();time.sleep(1.5);print('사업장 변경하기 버튼')
      WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1);print('선택한 사업장으로 변경합니까 알람')
      main = driver.window_handles;    print('보안카드 닫기');    driver.switch_to.window(main[1]); time.sleep(1)
      driver.find_element(By.ID,'group218').click();time.sleep(1.5);print('취소')
      WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(0.5);print('취소합니다 알람')
      main = driver.window_handles;    print(main);    driver.switch_to.window(main[0]);
    except:
      try:
        iframe = driver.find_element(By.CSS_SELECTOR,'#UTXPPAAA24_iframe');    driver.switch_to.frame(iframe);    time.sleep(1);print('iframe UTXPPAAA24_iframe2')
        WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1);print('선택한 사업장으로 변경버튼을 누르면 바로 변경됩니다.')
        driver.find_element(By.ID,'group1818').click();time.sleep(1.5);print('사업장 변경하기 버튼2')
        WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1);print('선택한 사업장으로 변경합니까 알람2')
      except:
        try:
          driver.find_element(By.ID,'group1818').click();time.sleep(1.5);print('사업장 변경하기 버튼3')
          WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1);print('선택한 사업장으로 변경합니까 알람3')    
          # 선택한 사업장으로 변경버튼을 누르면 바로 변경 알럿에서 확인을 누른 후 보안카드 입력창이 발생한 경우
          main_window_handle = driver.current_window_handle
          popup_window_handle = None
          while not popup_window_handle:
            if len(driver.window_handles)>1:
              for handle in driver.window_handles:
                  if handle != main_window_handle:
                      popup_window_handle = handle
                      # 팝업 윈도우로 전환
                      driver.switch_to.window(popup_window_handle)
                      # 팝업 윈도우에서 필요한 동작 수행
                      group218 =  driver.find_element(By.ID,'group218')#보안카드 취소버튼
                      if group218:
                        group218.click()
                        WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(0.5);print('취소합니다 알람')
                      # 메인 윈도우로 전환
                      driver.switch_to.window(main_window_handle) ; print('메인 윈도우로 전환')             
                      break
            else:
              popup_window_handle = True                  
        except:
          print(result[6]+' : 사업장 고유아이디')

  return driver



def conHometaxLogin(loginID,isNotShowChrome):
    context = {}
    # 1.크롬접속
    driver = ChromeDriver(isNotShowChrome)
  #try:
    driver.get('http://www.hometax.go.kr/websquare/websquare_cdn.html?w2xPath=/ui/pp/index.xml');driver.implicitly_wait(1)
    win = pyautogui.getActiveWindow();win.maximize(); #윈도우 최대화
    WebDriverWait(driver, 61).until(EC.element_to_be_clickable((By.XPATH, r'/html/body/div[1]/div[1]/div[2]/div/div[1]/div[1]/div[2]/div[1]/ul/li[3]/a/span'))).click()
    # loginbtn = driver.find_element(By.XPATH,r'/html/body/div[1]/div[1]/div[2]/div/div[1]/div[1]/div[2]/div[1]/ul/li[3]/a/span').click();    time.sleep(2)
                                              
    # 팝업뜨는 경우 없애기
    # 현재 윈도우 핸들 가져오기
    main_window_handle = driver.current_window_handle

    # 버튼 클릭 등으로 새로운 팝업 윈도우가 열리는 이벤트 발생
    # 팝업 윈도우 핸들 가져오기
    popup_window_handle = None
    while not popup_window_handle:
      if len(driver.window_handles)>1:
        for handle in driver.window_handles:
            if handle != main_window_handle:
                popup_window_handle = handle
                # 팝업 윈도우로 전환
                driver.switch_to.window(popup_window_handle)
                # 팝업 윈도우에서 필요한 동작 수행
                group2152 = driver.find_element(By.CSS_SELECTOR,'#group2152')
                if group2152:group2152.click(); 
                else:
                  trigger4 = driver.find_element(By.CSS_SELECTOR,'#trigger4')
                  if trigger4:trigger4.click();  
                # 메인 윈도우로 전환
                driver.switch_to.window(main_window_handle)                
                break
      else:
        popup_window_handle = True

    print('홈택스 접속')
    iframe = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, r'/html/body/div[1]/div[2]/iframe')));driver.switch_to.frame(iframe);
    time.sleep(1)
    # WebDriverWait(driver, 61).until(EC.element_to_be_clickable((By.ID, 'anchor15'))).click()#아이디로그인 탭이동 group91882156
    #idLoginBtn = driver.find_element(By.ID,'anchor15');        idLoginBtn.click();    time.sleep(1)#아이디로그인 탭이동 group91882156
    # target_element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'anchor15')))
    # ActionChains(driver).double_click(target_element).perform()
    # WebDriverWait(driver, 11).until(EC.element_to_be_clickable((By.ID, 'group91882156'))).click(); #li 접근
    # WebDriverWait(driver, 11).until(EC.element_to_be_clickable((By.ID, 'anchor15'))).click(); #a태그 접근


    # idLoginBtn = driver.find_element(By.ID,'anchor15');        idLoginBtn.click();    time.sleep(1)#아이디로그인 탭이동 group91882156

    start_time = time.time()
    timeout = 61  # Set your timeout limit

    while True:
      try:
        # Attempt to click the element with ID 'group91882156'
        driver.find_element(By.ID, 'group91882156').click()

        # Wait for the element with ID 'iptUserId' to be clickable
        WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.ID, 'iptUserId'))).send_keys(loginID)
        break  # If successful, exit the loop
      except Exception as e:
        # If 'iptUserId' is not clickable yet, wait for 1 second before trying again
        time.sleep(1)

        # Check if the timeout has been reached
        if time.time() - start_time > timeout:
            print("Timed out waiting for page to load")
            break  # Exit the loop if the timeout is reached




    # driver.find_element(By.ID,'group91882156').click();    time.sleep(1)    
    # WebDriverWait(driver, 61).until(EC.element_to_be_clickable((By.ID, 'iptUserId'))).send_keys(loginID)
    WebDriverWait(driver, 61).until(EC.element_to_be_clickable((By.ID, 'iptUserPw'))).send_keys('daeseung@1128')
    WebDriverWait(driver, 61).until(EC.element_to_be_clickable((By.ID, 'group91882177'))).click();print('로그인버튼 클릭');
    # idpwLoginBtn = driver.find_element(By.ID,'group91882177');idpwLoginBtn.click();print('로그인버튼 클릭');    time.sleep(1)
    iframe = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#dscert')));driver.switch_to.frame(iframe)
    # iframe = driver.find_element(By.CSS_SELECTOR,'#dscert');    driver.switch_to.frame(iframe); time.sleep(2)
    #인증서버튼
    WebDriverWait(driver, 61).until(EC.element_to_be_clickable((By.XPATH, r'/html/body/div[9]/div[2]/div[1]/div/div[5]/div/div[2]/div/div[4]/div[2]/div/table/tbody/tr/td[1]/a/span'))).click()
    # certBtn = driver.find_element(By.XPATH,r'/html/body/div[9]/div[2]/div[1]/div/div[5]/div/div[2]/div/div[4]/div[2]/div/table/tbody/tr/td[1]/a/span')
    # if certBtn:certBtn.click()
    # else:
    #    time.sleep(2)
    #    certBtn:certBtn.click()
    # time.sleep(1)
    passwd = 'daeseung@1128'
    driver.find_element(By.ID,'input_cert_pw').send_keys(passwd);    time.sleep(0.5)
    driver.find_element(By.ID,'btn_confirm_iframe').click();    print('공인인증서 로그인');    time.sleep(1)
    WebDriverWait(driver, 5).until(EC.alert_is_present());  al = Alert(driver);  al.accept();  time.sleep(2)#필수
    try:
      iframe = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, r'/html/body/div[1]/div[2]/iframe')));driver.switch_to.frame(iframe)
    except:
      print('아이디패스워드 로그인 재시도')
      driver.get('http://www.hometax.go.kr/websquare/websquare_cdn.html?w2xPath=/ui/pp/index.xml');driver.implicitly_wait(3)
      driver.find_element(By.XPATH,r'/html/body/div[1]/div[1]/div[2]/div/div[1]/div[1]/div[2]/div[1]/ul/li[3]/a/span').click();    time.sleep(2)
      iframe = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, r'/html/body/div[1]/div[2]/iframe')));driver.switch_to.frame(iframe)
      idLoginBtn = driver.find_element(By.ID,'anchor15').click();    time.sleep(1)
      driver.find_element(By.ID,'group91882156').click();    time.sleep(1)
      driver.find_element(By.ID,'iptUserId').send_keys(loginID);    time.sleep(0.31)
      driver.find_element(By.ID,'iptUserPw').send_keys('daeseung@1128');    time.sleep(0.31)
      driver.find_element(By.ID,'group91882177').click();print('로그인버튼2');    time.sleep(1)
      iframe = driver.find_element(By.CSS_SELECTOR,'#dscert');    driver.switch_to.frame(iframe); time.sleep(2);print('공인인증서 팝업')
      certBtn = driver.find_element(By.XPATH,r'/html/body/div[9]/div[2]/div[1]/div/div[5]/div/div[2]/div/div[4]/div[2]/div/table/tbody/tr/td[1]/a/span')
      if certBtn:certBtn.click()
      else:
        time.sleep(2)
        certBtn:certBtn.click()   
      time.sleep(1)   
      driver.find_element(By.ID,'input_cert_pw').send_keys('daeseung@1128');    time.sleep(0.5)
      driver.find_element(By.ID,'btn_confirm_iframe').click();    print('공인인증서 로그인2');    time.sleep(1)
      WebDriverWait(driver, 5).until(EC.alert_is_present());  al = Alert(driver);  al.accept();  time.sleep(2)#필수
      iframe = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, r'/html/body/div[1]/div[2]/iframe')));driver.switch_to.frame(iframe)
    driver.find_element(By.ID,'input1').send_keys('P30447')
    driver.find_element(By.ID,'input2').send_keys('daeseung@1128');time.sleep(1)
    # pyautogui.press('tab',presses=2,interval=0.3);pyautogui.press('enter',presses=3,interval=1)
    WebDriverWait(driver, 31).until(EC.element_to_be_clickable((By.ID, 'group1669'))).click();print('관리번호 로그인')
    time.sleep(0.5)
    return driver  

def conWetaxLogin(isNotShowChrome):
    context = {}

    driver = ChromeDriver(False)   
    driver.get('https://www.wetax.go.kr/login.do');time.sleep(1)
    win = pyautogui.getActiveWindow();win.maximize();time.sleep(0.5) #윈도우 최대화
    driver.find_element(By.XPATH,r'//*[@id="contents"]/div[2]/div[1]/a[1]').click();    time.sleep(1)
    # idLoginBtn = driver.find_element(By.XPATH,'//*[@id="tab-1"]/div[1]/button[2]');    time.sleep(1);    idLoginBtn.click();    time.sleep(3)
    loginPass = 'daeseung@1128'
    #인증서버튼
    # certBtn = driver.find_element(By.XPATH,r'//*[@id="tab-1_2"]/div/div[1]/a[1]').click();    time.sleep(3)
    certBtn = driver.find_element(By.ID,'btnCertLogin').click();    time.sleep(3)
    passwd = 'daeseung@1128'
    print('공인인증서 로그인')
    handle = find_window_with_retry('전자 서명 작성', '#32770', 10, 1) 
    # handle = pywinauto.findwindows.find_windows(title='전자 서명 작성', class_name='#32770')[0]
    app = Application().connect(handle=handle)
    w_open = app.window(handle=handle)
    w_open.Edit.type_keys(passwd)
    w_open['확인Button'].click()

    return driver  

#세무포털 로그인
def conSemuportalLogin():
  driver = ChromeDriver(False)
  driver.get('https://www.semuportal.com/bizon/index.do');  driver.implicitly_wait(3)	
  driver.find_element(By.XPATH,'/html/body/div/div[1]/div[1]/ul[2]/li[1]/a').click();time.sleep(0.5)
  driver.find_element(By.ID,'userId').send_keys('simplebook');time.sleep(0.5)
  driver.find_element(By.ID,'userPassword').send_keys('just1928');time.sleep(0.5)
  driver.find_element(By.XPATH,'//*[@id="form1"]/ul[4]/button[1]').click();time.sleep(1.5)

  # 팝업뜨는 경우 없애기
  # 현재 윈도우 핸들 가져오기
  main_window_handle = driver.current_window_handle
  # 팝업 윈도우 핸들 가져오기
  popup_window_handle = None
  while not popup_window_handle:
    if len(driver.window_handles)>1:
      for handle in driver.window_handles:
          if handle != main_window_handle:
              popup_window_handle = handle
              driver.switch_to.window(popup_window_handle)
              # x 버튼 클릭
              driver.find_element(By.XPATH,'/html/body/div/form/div/div[2]/ul[2]/a/span').click();time.sleep(0.5) 
              driver.switch_to.window(main_window_handle)                
              break
    else:
      popup_window_handle = True

  #포커스 가져오기
  procs = pywinauto.findwindows.find_elements();handle=''
  for proc in procs: 
    if proc.name[:5]== '한길TIS':handle = proc.handle;break
  app = Application().connect(handle=handle)
  w_open = app.window(handle=handle)
  w_open.set_focus()  

  return driver
	

# 고용토탈 사무대행기관 인증서로그인까지
def conTotalComwelLogin(bizno):
   #True:브라우저 없이 실행, False:브라우서띄우기
  chrome_options = Options()
  chrome_options.add_argument('--ignore-certificate-errors')
  chrome_options.add_argument('--allow-running-insecure-content')

  driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)  
  driver.get('http://total.kcomwel.or.kr');  time.sleep(3)  
  procs = pywinauto.findwindows.find_elements();handle=''
  for proc in procs: 
    if proc.name[:6]== '근로복지공단':handle = proc.handle;break
  app = Application().connect(handle=handle)
  w_open = app.window(handle=handle)
  w_open.set_focus()    
  loginbtn = driver.find_element(By.ID,'mf_wfm_header_btn_login').is_displayed();  time.sleep(0.5)
  if loginbtn:
    driver.find_element(By.ID,'mf_wfm_header_btn_login').click();time.sleep(0.5);print('id클릭')
    
  if bizno=='220-85-33586':
    bizno = bizno.split('-')
    driver.find_element(By.ID,'mf_wfm_content_TabContent1_tab_tabs3_tabHTML').click();    time.sleep(0.5);    print('사무대행탭클릭')
    loginbtn = driver.find_element(By.XPATH,'//*[@id="mf_wfm_content_u_group_fg3"]/span[2]/label').click();    time.sleep(0.5);    print('사업장명의인증서 클릭')  
    driver.find_element(By.ID,'mf_wfm_content_drno3_1').send_keys(bizno[0])
    driver.find_element(By.ID,'mf_wfm_content_drno3_2').send_keys(bizno[1])
    driver.find_element(By.ID,'mf_wfm_content_drno3_3').send_keys(bizno[2])
    driver.implicitly_wait(5);time.sleep(1)
    driver.find_element(By.ID,'mf_wfm_content_wq_uuid_638').click() ;    driver.implicitly_wait(5);time.sleep(1)
    driver.find_element(By.XPATH,'//*[@id="xwup_media_hdd"]').click();    driver.implicitly_wait(5);time.sleep(1);    print('하드디스크선택')
    # 테이블내 인증서찾기 개선필요
    driver.find_element(By.XPATH,'//*[@id="xwup_cert_table"]/table/tbody/tr[1]/td[2]/div').click()
    passwd = 'daeseung@1128'
    loginbtn = driver.find_element(By.XPATH,'//*[@id="xwup_certselect_tek_input1"]').is_displayed()
    if loginbtn:
      print('인증서암호')
      driver.find_element(By.XPATH,'//*[@id="xwup_certselect_tek_input1"]').send_keys(passwd)
    driver.implicitly_wait(5)
    driver.find_element(By.ID,'xwup_OkButton').click()
    print('공인인증서 로그인')
    time.sleep(1)  
    try:  
      notToday = driver.find_element(By.ID,'mf_wfm_content_samuInfoPopup_wframe_chk_today').is_displayed()
      driver.find_element(By.ID,'mf_wfm_content_samuInfoPopup_wframe_chk_today').click()  ;time.sleep(1) ;print('오늘 그만열기 해제')
    except:
      print('오늘 그만열기 없음')
  return driver

def find_window_with_retry(title, class_name, max_retries=10, retry_interval=1):
    for _ in range(max_retries):
        handle_list = pywinauto.findwindows.find_windows(title=title, class_name=class_name)
        if handle_list:
            return handle_list[0]
        print("재시도 : "+title)
        print(_)
        time.sleep(retry_interval)
    return None  # 창이 나타나지 않을 경우 None 반환

# 홈택스 출력팝업 - 민원서류
def Htx_Popup_Print_Minwon(driver,printBtnName,fileName,directory,directPrint):
  driver.find_element(By.XPATH,printBtnName).click();time.sleep(0.25)
  WebDriverWait(driver, 30).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(0.25)
  time.sleep(3);print('일괄출력 후 6초대기'+printBtnName)
  main = driver.window_handles;        print(main);        driver.switch_to.window(main[1]) 
  driver.find_element(By.XPATH,'/html/body/div/div/div[1]/table/tbody/tr/td/div/nobr/button[6]').click();time.sleep(3);print('최초인쇄')
  # driver.find_element(By.XPATH,"/html/body/div/div[2]/div/button[1]").click();time.sleep(3);print('최종인쇄')
  # main = driver.window_handles;    print(main);    driver.switch_to.window(main[2]); time.sleep(5);print('필수대기시간')
  # pyautogui.moveTo(0,0)
  # pyautogui.click(x=580,y=100);time.sleep(0.5);    
  pyautogui.press('enter');time.sleep(1);print('인쇄버튼')
  createDirectory(directory)    
  fullFileName = directory +"/"+ fileName +".pdf"
  if  os.path.isfile(fullFileName.replace('/',"\\")):os.remove(fullFileName.replace('/',"\\"));print(fullFileName+" 삭제완료") ;time.sleep(1)
  Report_save(fileName,directory);print('파일저장완료');time.sleep(1)
  pyautogui.hotkey('alt','f4')
  try:
    main = driver.window_handles;    print(main);    driver.switch_to.window(main[0]); 
  except Exception as e:
    print('에러 발생:', e)
  try:
    iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe');      driver.switch_to.frame(iframe);      time.sleep(1); print('iframe 전환txppIframe')
  except Exception as e:
    print('에러 발생:', e)

  return 1


# 홈택스 출력팝업
def Htx_Popup_Print(driver,printBtnName,fileName,directory,directPrint):
  if printBtnName[0:3]=='//*':
    driver.find_element(By.XPATH,printBtnName).click();time.sleep(0.25)
    if printBtnName!='//*[@id="grdList_cell_0_16"]/span/button' :WebDriverWait(driver, 30).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(0.25)
  else:
    try:
        print('관련버튼ID:'+printBtnName)
        saleBtn = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, printBtnName)))
        print('Htx_Popup_Print '+printBtnName + " trying")
        driver.find_element(By.ID,printBtnName).click()
        # saleBtn.click();print('Htx_Popup_Print '+printBtnName + " click")
    except NoSuchElementException:
        print(f"{printBtnName} 요소를 찾을 수 없습니다.")
    except TimeoutException:
        print(f"{printBtnName} 요소가 클릭 가능한 상태가 되지 않았습니다.")
  print('일괄출력 후 6초대기'+printBtnName);time.sleep(6)
  if not directPrint:
    main = driver.window_handles;        print(main);        driver.switch_to.window(main[1]) 
    driver.find_element(By.XPATH,'/html/body/div/div/div[1]/table/tbody/tr/td/div/nobr/button[6]').click();time.sleep(0.5);print('최초인쇄')
    driver.find_element(By.XPATH,"/html/body/div/div[2]/div/button[1]").click();time.sleep(3);print('최종인쇄')
    main = driver.window_handles;    print(main);    driver.switch_to.window(main[2]); time.sleep(5);print('필수대기시간')
    pyautogui.moveTo(0,0)
    pyautogui.click(x=580,y=100);time.sleep(0.5);    
    pyautogui.press('enter');time.sleep(1);print('인쇄버튼')
    createDirectory(directory)    
    Report_save(fileName,directory);print('파일저장완료')
    pyautogui.hotkey('alt','f4')
    try:
      main = driver.window_handles;    print(main);    driver.switch_to.window(main[0]); 
      current_url = driver.current_url
      print(current_url)
    except Exception as e:
      print('에러 발생:', e)
    try:
      iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe');      driver.switch_to.frame(iframe);      time.sleep(1); print('iframe 전환txppIframe')
      # iframe = driver.find_element(By.CSS_SELECTOR,'#UTERNAAT71_iframe');      driver.switch_to.frame(iframe);      time.sleep(1); print('iframe 전환txppIframe')
    except Exception as e:
      print('에러 발생:', e)
  else:#일반출력창이 아닌 경우(프로세스가 멈춰버린다)
    main = driver.window_handles;       driver.switch_to.window(main[1]);   print(main);time.sleep(0.5)
    # 캡처할 팝업창의 좌표와 크기 설정
    x = 10
    y = 100
    width = 850
    height = 900
    # 스크린샷 캡처
    pyautogui.screenshot('popup.png', region=(x, y, width, height))
    # 이미지를 PDF로 변환하여 저장
    with open(directory+"/"+fileName+".pdf", "wb") as f:
        f.write(img2pdf.convert('popup.png'))
    print("변환완료")
    pyautogui.hotkey('alt','f4')
    try:
      main = driver.window_handles;    print(main);    driver.switch_to.window(main[0]); 
      current_url = driver.current_url
      print(current_url)
    except Exception as e:
      print('에러 발생:', e)
    try:
      iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe');      driver.switch_to.frame(iframe);      time.sleep(1); print('iframe 전환txppIframe')
      # iframe = driver.find_element(By.CSS_SELECTOR,'#UTERNAAT71_iframe');      driver.switch_to.frame(iframe);      time.sleep(1); print('iframe 전환txppIframe')
    except Exception as e:
      print('에러 발생:', e)
  return 1


#다운된 파일 이름바꿔저장하기
def FileSave_Downloaded_PDF(directory,file_Origin_name,file_Purpose_name,biz_no):
  memuser = MemUser.objects.get(biz_no=biz_no)
  memdeal = MemDeal.objects.get(seq_no=memuser.seq_no.strip())
  if file_Origin_name=='':
    downloads = 'C:\\Users\\Administrator\\Downloads'
    file_Origin_name = max([downloads + "\\" + f for f in os.listdir(downloads)],key=os.path.getctime)
    print(f'{file_Origin_name}')  
  savePath = directory.replace('/',"\\")
  if memdeal.biz_manager=='화물':savePath = directory.replace('D:\\',"D:\\화물연대\\")
  print(savePath)
  if os.path.isdir(savePath):
    if os.path.isdir(savePath) == False:      os.mkdir(savePath)
  else:
    os.mkdir(savePath)
  shutil.move(file_Origin_name,os.path.join(savePath,file_Purpose_name))
  print('파일저장성공')
  return 1

# 디렉토리생성
def createDirectory(directory):
  try:
    if not os.path.exists(directory):
      os.makedirs(directory)
      return "none"
    else:
      file_list = os.listdir(directory)
      return file_list
  except OSError:
    print("Error: 폴더생성실패 - "+directory)

def Report_save(fileName,directory):  
  pyautogui.hotkey('alt','d')
  pyperclip.copy(directory.replace('/',"\\"));   time.sleep(0.25);pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  # pyperclip.copy("c:/".replace('/',"\\"));    time.sleep(0.25);pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  pyautogui.press('enter');time.sleep(0.5);pyautogui.press('tab',presses=3,interval=0.5)
  pyautogui.hotkey('alt','n')
  pyperclip.copy(fileName);    time.sleep(0.5)
  pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  pyautogui.hotkey('alt','s');    time.sleep(5)  
  return 1    

def delete_CorpSheet_998_files(flag,directory):
  delFile = ['1','3','4','3-1','3-2','3-4','15','47-갑','48','50-갑','92-지-43-2','92-지-43','101','102','103','104','105','106','300']
  if flag=="2": delFile = ['1','40-6','40-7','40-8','40-9','46','47','69','71','71-1','73','91-1','91-2','91-2-2','101','102','103','104','105','106','300']
  if os.path.exists(directory):
    for file in os.scandir(directory):
      title = file.name.split(".")[0]
      if title  in delFile: os.remove(file.path);print('지운파일:'+file.path)
    print('법인세 결산서 삭제완료')
  else:
    print('Directory Not Found')  

def delete_CorpSheet_files(flag,directory):
  notDel = ['0','19-갑','58','98','99','100']
  if os.path.exists(directory):
    if os.path.isfile(directory+"/Thumbs.db"): os.remove(directory+"/Thumbs.db")
    for file in os.scandir(directory):
      print(file.name)
      title = file.name.split(".")[0]
      titleDetail = "110"
      if title.find("-")!=-1: #신고서외 보조파일인 경우
        titleDetail = title.split("-")[0]
        if title not in notDel and int(titleDetail)<110: os.remove(file.path);print('지운파일:'+file.path)
      elif title not in notDel and int(title)<110: os.remove(file.path);print('지운파일:'+file.path)
      if title=='300':os.remove(file.path)
    print('법인세 신고서 삭제완료')
  else:
    print('Directory Not Found')  
# 폴더내 파일 전체 삭제
def delete_all_files(download_path):
  if os.path.exists(download_path):
    for file in os.scandir(download_path):
      os.remove(file.path)
    print('Remove All File')
  else:
    print('Directory Not Found')

# 경로의 파일 갯수 체크
def dir_search(self, download_path):
    file_len = len(os.listdir(download_path))
    print('파일갯수: ' + str(file_len))
    return file_len
# 경로내의 지정된 단어를 가진 파일리스트 리턴
def find_target_files(folder_path,txtTarget):
    if not os.path.exists(folder_path):
        raise ValueError(f"{folder_path} : 경로가 존재하지 않습니다")

    target_files = []
    search_pattern = os.path.join(folder_path, f'*{txtTarget}*')
    for root, _, files in os.walk(folder_path):
        for file_name in files:
            if txtTarget in file_name:
              target_files.append([root, file_name])
                # target_files.append(os.path.join(root, file_name))

    return target_files


# 다운로드
def download_excel(self, download_path):
    driver = self.driver
    # 폴더가 없으면 생성
    if os.path.isdir(download_path) == False:
        os.mkdir(download_path)
    else:  print(f'이미 폴더가 있습니다.')
    # 폴더안에 이미 존재하는 파일 삭제
    print(f'{self.delete_all_files(download_path)}')
    # 폴더안의 파일 갯수 (다운로드 확인에 사용)
    before_file_len = self.dir_search(download_path)
    # 다운로드 버튼 클릭 (실패 시 무한반복)
    downloaded = False
    while downloaded == False:
        try:
            # 다운로드 클릭
            download_button = driver.find_element(By.CSS_SELECTOR, 'div[onclick*="javascript:doExcelDownload()"]')
            download_button.click()
            # 확인 클릭
            ok_button = driver.find_element(By.CSS_SELECTOR, '#button-1006')
            ok_button.click()
            downloaded = True
        except Exception as e:
            downloaded = False
    # 다운로드 완료 여부를 5초에 한번씩 체크
    after_file_len = self.dir_search(download_path)
    while before_file_len == after_file_len:
        print('대기중...')
        after_file_len = self.dir_search(download_path)
        time.sleep(5)
    try: # 가장 최근에 다운받은 파일 이름을 변경
        new_file_name = max([download_path + "\\" + f for f in os.listdir(download_path)],key=os.path.getctime)
        shutil.move(new_file_name,os.path.join(download_path,r"myfile.xlsx"))
        new_file_name = download_path + "\\" + r"myfile.xlsx"
        print(f'{new_file_name}')
    except Exception as e:
        print(e)
        self.log_append(f'{download_path} 파일 찾기 실패')
        print('파일명 변경 실패')

    return new_file_name
def check_alert(driver):
    from selenium.common.exceptions import NoAlertPresentException
    try:
        driver.switch_to.alert
        return True
    except NoAlertPresentException:
        return False
def PDF_Merge(fileName,directory):
  merger = PyPDF2.PdfMerger()
  filst = os.listdir(directory)
  for file in natsort.natsorted(filst):
    merger.append(os.path.join(directory,file))
  merger.write(f"{directory}/{fileName}.pdf")
  merger.close()
  print(directory+" 폴더에 "+fileName+".PDF로 병합 성공")
  return 1

def unicode_slice(s, start, end):
  """유니코드 문자열 s에서 start부터 end-1까지의 문자열을 슬라이싱합니다."""
  sliced = ""
  count = 0
  for char in s:
    if not is_hangul(char) :            count += 1
    else:count += 2
    if count >= start and count <= end:
      sliced += char
    if count >= end:
      break
  print(sliced)
  return sliced
def is_hangul(char):
    """문자 char가 한글인지 여부를 반환합니다."""
    cp = ord(char)
    if 0xAC00 <= cp <= 0xD7AF:
        return True
    return False
def find_similar_strings(matched_files, target_string):
  for file_name in matched_files:
    if file_name.find(target_string) != -1:
      print(file_name)
      print(target_string)      
      return True
  return False
def get_quarter(strmonth):
  month = int(strmonth)
  if month in [1, 2, 3]:
      return '1'
  elif month in [4, 5, 6]:
      return '2'
  elif month in [7, 8, 9]:
      return '3'
  elif month in [10, 11, 12]:
      return '4'
  else:
      return None
  

  
def is_program_running(process_name):
  windows = Desktop(backend='uia').windows()
  for window in windows:
    # print(window.window_text())
    if window.window_text()==process_name:
      return True
  return False

def semusarang_LaunchProgram_App(strID):
  global isSemusarangOn
  print('semusarang_LaunchProgram')
  app = None
  program_path = "C:\\NewGen\\Rebirth\\Rebirth.exe"

  # 프로그램 실행 여부 확인
  if is_program_running("세무사랑 Pro"):
    print('이미 실행 중인 경우, 프로그램에 연결')
    app = Application(backend='uia').connect(path=program_path)
    subprocess.Popen(r'C:/NewGen/Rebirth/Rebirth.exe');  time.sleep(2)
    dig = app['세무사랑 Pro']
  else:
    print('실행되지 않은 경우, 프로그램을 실행')
    app = Application(backend='uia').start(program_path)  ;time.sleep(5)
    dig = app['세무사랑 Pro']
    # Edit 컨트롤이 활성화될 때까지 대기 (예: 최대 10초 동안 대기)
    timeout = 10  # 대기 시간 초 단위로 설정
    start_time = time.time()
    while time.time() - start_time < timeout:
      if dig.Edit.exists():  break
      print(time.time() - start_time)
      time.sleep(1)  # 1초 대기

    # Edit 컨트롤이 활성화되었는지 확인
    if dig.Edit.exists():
      # Edit 컨트롤에 엔터 키 입력
      dig.Edit.type_keys('{ENTER}')  #Edit 속성은 맨위로 가게한다(서버-1)
      # dig.type_keys(strID)  
      dig.type_keys('{ENTER}')    
      dig.type_keys('1122')  ;dig.type_keys('{ENTER}') 
      dig.type_keys('102')  ;dig.type_keys('{ENTER}') ;  time.sleep(1)
      dig.type_keys('112837');dig.type_keys('{ENTER}')  ;time.sleep(2)   # dig['Button4'].click()  이것도 가능
      if pyautogui.locateOnScreen('K:/noa-django/Noa/static/assets/images/autowork/semusarang_diff_Update.png',confidence=0.7): 
        pyautogui.press('enter');print('서버와 업데이트일자가 다릅니다') 
      time.sleep(3)
      isSemusarangOn = True
    else:
      print("Edit 컨트롤이 활성화되지 않았습니다.")
      isSemusarangOn = False
  return dig

def semusarang_LaunchProgram(strID):
  global isSemusarangOn
  print('semusarang_LaunchProgram')
  subprocess.Popen(r'C:/NewGen/Rebirth/Rebirth.exe');  time.sleep(3)
  print('세무사랑ID가 이미 등록된 경우는 아래 3개를 지워야해')
  pyautogui.write(' ');  time.sleep(1)
  pyautogui.write(strID);  time.sleep(1)
  pyautogui.press('enter');  time.sleep(1)
  pyautogui.write('1122');  time.sleep(0.25)
  pyautogui.press('enter');  time.sleep(0.25) 
  pyautogui.press('f2')
  isSemusarangOn = True
  return isSemusarangOn
def semusarang_ChangeCompany_ID(DuzonID):
  print('회사변경')
  time.sleep(0.5) 
  pyautogui.press('f3');  time.sleep(1)  
  print(DuzonID)
  pyautogui.write(str(DuzonID));  time.sleep(1)  
  if len(str(DuzonID))==3:pyautogui.press('enter');  time.sleep(1.5)#필수 
  if str(DuzonID)=='102' or str(DuzonID)=='2078':
    print('비번입력')      
    pyautogui.write('112837');    time.sleep(0.1)
    pyautogui.press('enter');    time.sleep(2) #필수
  pyautogui.press('enter');  time.sleep(1.5)#필수 
  return 1  
def semusarang_ChangeCompany_ID_App(dig,DuzonID):
  print('회사변경')
  time.sleep(0.5) 
  # pyautogui.press('f5');  time.sleep(1)  
  dig['사업자등록번호'].click();time.sleep(0.5)
  keyboard.send_keys(DuzonID) 
  if len(str(DuzonID))==3:keyboard.send_keys('{ENTER}');  time.sleep(1.5)#필수 
  if str(DuzonID)=='102' or str(DuzonID)=='2078':
    print('비번입력')      
    keyboard.send_keys('112837');    time.sleep(0.1)
    keyboard.send_keys('{ENTER}');    time.sleep(2) #필수
  keyboard.send_keys('{ENTER}');  time.sleep(2)#필수 
  keyboard.send_keys('{ENTER}');print('업데이트일자가 다릅니다. 팝업제거')
  return 1  

def semusarang_ChangeCompany(bizNo):
  print('회사변경')
  time.sleep(0.5) 
  pyautogui.press('f3');  time.sleep(1)  
  pyautogui.press('f2');  time.sleep(0.5)  
  bizNo = bizNo.replace('-','')
  pyautogui.write(str(bizNo));  time.sleep(1)  
  pyautogui.press('enter');  time.sleep(1.5)#필수  
  if str(bizNo)=='2208533586' or str(bizNo)=='1448127161':
    print('비번입력')      
    pyautogui.write('112837');    time.sleep(0.1)
    pyautogui.press('enter');    time.sleep(2) #필수
  return 1
def semusarang_ChangeFiscalYearAll(Kisu,strYear):
  print('회계 기수변경')
  time.sleep(1)
  pyautogui.press('enter')
  pyautogui.press('f4') ; print('f4')
  time.sleep(0.5)
  print('기수 : '+ Kisu)
  pyautogui.write(Kisu);  time.sleep(0.5) 
  pyautogui.press('enter');time.sleep(0.25)
  pyautogui.press('down');print('연도세팅 : '+strYear)
  pyautogui.write(strYear);    time.sleep(2)#필수 
  pyautogui.press('tab')
  time.sleep(1.5)     
  return 1

def semusarang_ChangeFiscalYear(flag,strYear):
  time.sleep(2)
  if flag=='insa':
    print('인사급여 기수변경')
    pyautogui.press('enter')
    pyautogui.press('f4');time.sleep(1); print('f4')
    print('다운버튼 1회')
    pyautogui.press('down');    time.sleep(0.25) 
    print('연도세팅 : '+strYear)
    pyautogui.write(strYear);    time.sleep(2)#필수 
  else:
    print('회계 기수변경')
    time.sleep(1)
    pyautogui.press('enter')
    pyautogui.press('f4') ; print('f4')
    time.sleep(1)
    print('기수 : '+ strYear)
    pyautogui.write(strYear)
    time.sleep(0.5) 
    pyautogui.press('enter')
    time.sleep(2)#필수 
  pyautogui.press('tab')
  time.sleep(1.5)     
  return 1

def semusarang_ChangeFiscalYear_App(flag,strYear):
  time.sleep(2)
  if flag=='insa':
    print('인사급여 기수변경');time.sleep(1)
    keyboard.send_keys('{ENTER}')
    keyboard.send_keys('{F4}');time.sleep(1); print('다운버튼 1회')
    keyboard.send_keys('{DOWN}');    time.sleep(0.25) ;    print('연도세팅 : '+strYear)
    keyboard.send_keys(strYear);    time.sleep(2)#필수 
  else:
    print('회계 기수변경');time.sleep(1)
    keyboard.send_keys('{F2}') 
    keyboard.send_keys('{ENTER}')
    keyboard.send_keys('{F4}') ; time.sleep(1)
    
    print('기수 : '+ strYear)
    keyboard.send_keys(strYear);    time.sleep(0.5) 
    keyboard.send_keys('{ENTER}');    time.sleep(2)#필수 
  keyboard.send_keys('{TAB}');  time.sleep(1.5)     
  return 1

def semusarang_Login(semusarangID,duzon_id,biz_no,reg_date,workyear):
  dig = semusarang_LaunchProgram_App(semusarangID)
  if duzon_id.strip()=='1':  semusarang_ChangeCompany(biz_no.replace("-",""))
  else:               semusarang_ChangeCompany_ID_App(dig,duzon_id.strip())          
  finalKi = workyear - reg_date.year + 1    
  semusarang_ChangeFiscalYear_App('vat',str(finalKi))   
  return True

def semusarang_Print(fileName,directory):
  createDirectory(directory)    
  fullFileName = directory+"/"+fileName +".pdf"
  if  os.path.isfile(fullFileName.replace('/',"\\")):os.remove(fullFileName.replace('/',"\\"));print(fullFileName+" 삭제완료") 
  pyautogui.press('f9');time.sleep(3); print(fileName + '출력버튼') #낮은사양에서는 필수
  pyautogui.hotkey('alt','1');time.sleep(3)    
  pyperclip.copy(fileName);    time.sleep(0.5)
  pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  pyautogui.hotkey('alt','d');    time.sleep(0.25)
  pyperclip.copy(directory.replace('/',"\\"))
  pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  pyautogui.hotkey('alt','s');    time.sleep(2)
  pyautogui.press('esc',presses=4,interval=0.25);print('열린 세무사랑 하부메뉴 닫기');    time.sleep(0.5)        
  return 1 
def semusarang_Print_Sheet(AppTitle,txtSelect,fileName_Month,directory):
  fileName = fileName_Month + txtSelect[:5]+".pdf"
  fullFileName = directory+"/"+fileName_Month + txtSelect[:5]+".pdf"
  if  os.path.isfile(fullFileName.replace('/',"\\")):os.remove(fullFileName.replace('/',"\\"));print(fullFileName+" 삭제완료") 

  pyautogui.press('f9');time.sleep(3); print(fileName + ' 출력') #낮은사양에서는 필수
  try:
    handle = pywinauto.findwindows.find_windows(title=AppTitle, class_name='#32770')[0]
  except:
    SET_FOCUS(AppTitle)
    pyautogui.press('f9');time.sleep(3); print(fileName + ' 출력') #낮은사양에서는 필수
    handle = pywinauto.findwindows.find_windows(title=AppTitle, class_name='#32770')[0]
  app = Application().connect(handle=handle)
  w_open = app.window(handle=handle)
  w_open.ComboBox3.Select(txtSelect);time.sleep(1)    

  pyautogui.hotkey('alt','1');time.sleep(3)    
  pyperclip.copy(fileName);    time.sleep(0.5)
  pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  pyautogui.hotkey('alt','d');    time.sleep(0.25)
  pyperclip.copy(directory.replace('/',"\\"))
  pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  pyautogui.hotkey('alt','s');    time.sleep(2)
  #pyautogui.press('esc');print('열린 세무사랑 하부메뉴 닫기');    => 저장후 출력창은 바로 닫힌다        
  return 1 


def semusarang_Print_F11(fileName,directory):
    pyautogui.press('f11');time.sleep(3); print(fileName + '출력버튼') #낮은사양에서는 필수
    pyautogui.hotkey('alt','1');time.sleep(3)    
    pyperclip.copy(fileName);    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
    pyautogui.hotkey('alt','d');    time.sleep(0.25)
    pyperclip.copy(directory.replace('/',"\\"))
    pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
    pyautogui.hotkey('alt','s');    time.sleep(2)
    pyautogui.press('esc',presses=4,interval=0.25);print('열린 세무사랑 하부메뉴 닫기');    time.sleep(0.5)        
    return 1  
def semusarang_Print_TI_Summary(fileName,directory):
    createDirectory(directory)
    if fileName=='2.pdf' or fileName=='4.pdf':
      center = pyautogui.locateCenterOnScreen('K:/noa-django/Noa/static/assets/images/autowork/vatCostSummation.png',confidence=0.9)
      #center = pyautogui.locateCenterOnScreen('K:/noa-django/Noa/static/assets/images/autowork/vatCostSummation.png',confidence=0.7)
      if center:pyautogui.click(center);    time.sleep(1) 
    pyautogui.hotkey('ctrl','f11');time.sleep(3); print(fileName + '출력버튼') #낮은사양에서는 필수
    pyautogui.hotkey('alt','1');time.sleep(3)    
    pyperclip.copy(fileName);    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
    pyautogui.hotkey('alt','d');    time.sleep(0.25)
    pyperclip.copy(directory.replace('/',"\\"))
    pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
    pyautogui.hotkey('alt','s');    time.sleep(2)     
    return 1 
def semusarang_Print_NotExit(fileName,directory):
    pyautogui.press('f9');time.sleep(3); print('출력버튼') #낮은사양에서는 필수
    center = pyautogui.locateCenterOnScreen('K:/noa-django/Noa/static/assets/images/autowork/semusarang_PDF_save.png')
    pyautogui.click(center)
    time.sleep(0.5) 
    pyautogui.press('tab',presses=2,interval=0.25)
    pyautogui.press('enter');print('찾기버튼');    time.sleep(0.5)      
    pyperclip.copy(fileName);    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
    pyautogui.hotkey('alt','d');    time.sleep(0.25)
    pyperclip.copy(directory.replace('/',"\\"))
    pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
    pyautogui.hotkey('alt','s');    time.sleep(2)
    pyautogui.press('tab',presses=2,interval=0.25)
    pyautogui.press('enter');print('확인버튼')  ;    time.sleep(2)   
    pyautogui.hotkey('alt','f4');print('열린 PDF닫기');    time.sleep(1)     
    return 1 
def semusarang_Excelclose_Openmenuclose():
  time.sleep(1)
  pyautogui.hotkey('alt','f4')
  time.sleep(1)
  pyautogui.press('esc',presses=3, interval=1) 
def semusarang_Menu_Popup_App(menuName):
  time.sleep(1)
  keyboard.send_keys('^{ENTER}');time.sleep(0.5);pyperclip.copy(menuName)    #메뉴검색
  keyboard.send_keys('^v');time.sleep(0.5);keyboard.send_keys('{ENTER}');  time.sleep(1)
  return 1

#세무사랑 메뉴띄우기
def semusarang_Menu_Popup(menuName):
  time.sleep(1)
  keyboard.send_keys('^{ENTER}');time.sleep(0.5) #메뉴검색
  pyperclip.copy(menuName);  keyboard.send_keys('^v');time.sleep(1.5)
  pyautogui.press('enter');time.sleep(2)   
  if menuName in ('간이지급명세서전자신고','신용카드매출전표등수령명세서(갑)(을)','지급명세서전자신고(연말정산)','원천징수이행상황신고서','거래처등록'):
    pyautogui.press('enter');time.sleep(3)  
  intlen = len(menuName)
  procs = pywinauto.findwindows.find_elements();handle=''
  for proc in procs: 
    if proc.name[:intlen]== menuName:handle = proc.handle;break
  app = Application().connect(handle=handle)
  w_open = app.window(handle=handle)
  w_open.set_focus()  
  return 1

def SET_FOCUS(menuName):
  intlen = len(menuName)
  procs = pywinauto.findwindows.find_elements();handle=''
  for proc in procs: 
    if proc.name[:intlen]== menuName:handle = proc.handle;break
  app = Application().connect(handle=handle)
  w_open = app.window(handle=handle)
  w_open.set_focus()  
  return True

#디비데이터로 엑셀파일 만들어 driver에 저장하기
def MakeExcel_FromDB(driver,manageNo,flag):
  totalPath = driver
  wb = Workbook()  # 새 워크북 만들기  
  ws = wb.active   # 워크북의 첫 번째 워크시트 가져오기
  fileName = f"{flag}.xlsx" 
  if flag=="엑셀자료 일반전표전송":
    ws['A1'] = flag
    ws['O3'] = manageNo[1]
    ws['M4'] = 'v1.2'           
    MIDTITLE = ["년도월일","구분","코드","계정과목","코드","거래처명","코드","적요명","금액"]   
    for i, value in enumerate(MIDTITLE):
      ws.cell(row=10, column=i+1, value=value)    
    strsql = "select Work_YY+Tran_Dt YYMMDD,CrDr,Acnt_cd,Acnt_Nm,Trader_Code,Trader_Name,'' remk_cd,Remk,TranAmt "
    strsql += f"from DS_SlipLedgr_Ecount a where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and left(Tran_Stat,4)<>'매입전표' and left(Tran_Stat,4)<>'매출전표' order by Tran_Dt,slip_no,prt_ord "
    print(strsql)
    cursor = connection.cursor()
    cursor.execute(strsql)
    results = cursor.fetchall()
    connection.commit()
    if results:
      for i, row in enumerate(results):
        for j, value in enumerate(row):
          ws.cell(row=i+11, column=j+1, value=value)
  elif flag=="신용카드 매입 자료 입력":
    ws['A1'] = flag
    ws['c4'] = manageNo[1]       
    MIDTITLE = ["카드종류","신용카드사명","신용카드번호","승인일자","사업자등록번호","거래처명","거래처유형","공급가액","세액","봉사료","합계금액","부가세공제여부","부가세유형","품목"]   
    for i, value in enumerate(MIDTITLE):
      ws.cell(row=8, column=i+1, value=value) 
    #--카드전표: 매입전표(상대계정이 하나인 경우) 
    strsql = f"select '3', trader_name,trader_bizno,work_yy+tran_dt,"
    strsql +=f"(select case when trader_bizno='' then Trader_Code_ORG else Trader_Bizno end from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and acnt_cd<>253 and Acnt_cd<>135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"(select case when trader_name='' then Trader_name_ORG else Trader_name end from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and acnt_cd<>253 and Acnt_cd<>135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"'일반',"
    strsql +=f"(select tranamt_cr from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"(select tranamt from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd=135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"'0',TranAmt,'공제','57',"
    strsql +=f"(select acnt_cd from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"(select remk from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ) "
    strsql +=f"from DS_SlipLedgr_Ecount a where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and left(Tran_Stat,4)='매입전표' and acnt_cd=253"
    strsql +=f"and (select COUNT(slip_no) from DS_SlipLedgr_Ecount where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt )>2"
    strsql +=f" union all "
    #--카드전표: 일반전표(상대계정이 하나인 경우)
    strsql +=f"select '3', trader_name,trader_bizno,work_yy+tran_dt,"
    strsql +=f"(select case when trader_bizno='' then Trader_Code_ORG else Trader_Bizno end from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and acnt_cd<>253 and Acnt_cd<>135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"(select case when trader_name='' then Trader_name_ORG else Trader_name end from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and acnt_cd<>253 and Acnt_cd<>135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"'일반',"
    strsql +=f"(select tranamt_cr from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"'0','0',TranAmt,'불공제','',"
    strsql +=f"(select acnt_cd from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"(select remk from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt )"
    strsql +=f"from DS_SlipLedgr_Ecount a where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and left(Tran_Stat,4)='매입전표' and acnt_cd=253 "
    strsql +=f"and (select COUNT(slip_no) from DS_SlipLedgr_Ecount where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt)=2    "
    strsql +=f" union all "
    #--카드전표: 일반전표(상대계정이 둘 이상인 경우)
    strsql +=f"select '3', trader_name,trader_bizno,work_yy+tran_dt,"
    strsql +=f"(select case when trader_bizno='' then Trader_Code_ORG else Trader_Bizno end from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and acnt_cd<>253 and Acnt_cd<>135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"(select case when trader_name='' then Trader_name_ORG else Trader_name end from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and acnt_cd<>253 and Acnt_cd<>135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"'일반',"
    strsql +=f"(select tranamt_cr from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ), "
    strsql +=f"'0','0',TranAmt,'불공제','',"
    strsql +=f"(select acnt_cd from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ), "
    strsql +=f"(select remk from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ) "
    strsql +=f"from DS_SlipLedgr_Ecount a where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and left(Tran_Stat,4)='매입전표' and acnt_cd=253 "
    strsql +=f"and (select COUNT(slip_no) from DS_SlipLedgr_Ecount where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt)>3"
    print(strsql)
    cursor = connection.cursor()
    cursor.execute(strsql)
    results = cursor.fetchall()
    connection.commit()
    if results:
      for i, row in enumerate(results):
        for j, value in enumerate(row):
          ws.cell(row=i+10, column=j+1, value=value)    
  ##########################
  wholeName = totalPath + fileName
  if os.path.exists(wholeName):        os.remove(wholeName)
  wb.save(wholeName)
  # openpyxl로 읽어온 xlsx 데이터를 xlwt로 생성한 xls 파일에 쓰기
  wb_xls = xlwt.Workbook()
  for sheet_name in wb.sheetnames:
    ws_xlsx = wb[sheet_name]
    ws_xls = wb_xls.add_sheet(sheet_name)
    for row in ws_xlsx.rows:
      for cell in row:
        ws_xls.write(cell.row-1, cell.col_idx-1, cell.value) 
  if os.path.exists(wholeName.split('.')[0]+'.xls'):os.remove(wholeName.split('.')[0]+'.xls')         
  wb_xls.save(wholeName.split('.')[0]+'.xls')    #xlsx를 xls로 변환한다  
  os.remove(wholeName)      
  return True

#다운된 파일 열고 디비로 저장하기
def DBSave_Downloaded_xlsx(driver,manageNo,flag):
  workyear = datetime.now().year-1
  today = str(datetime.now())[:10].replace("-","")
  file_name=""
  if flag in ("보수총액신고대상자업데이트",'거래처등록'):
    downloads = 'C:\\NewGen\\Rebirth\\리버스문서보관함'
    file_name = max([downloads + "\\" + f for f in os.listdir(downloads)],key=os.path.getctime)    
  else:
    downloads = 'C:\\Users\\Administrator\\Downloads'
    try:
      file_name = max([downloads + "\\" + f for f in os.listdir(downloads)],key=os.path.getctime)
      extension = os.path.splitext(file_name)[1].lower()  # Get the extension and convert it to lower case
      if extension in ['.xls', '.xlsx']:
          print("This is an Excel file.")
      else:
          print("엑셀파일이 저장되지 않았습니다.")      
          return False
    except:
      print(manageNo[6] + "의 엑셀파일 없음")
      return False
  print(f'{file_name}')
  if flag=='부가세예정고지(확정)':
    df = pd.read_excel(file_name,header=1, engine='openpyxl')
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
    for row in df.itertuples():
      bizno=str(row[5])
      try:
        memuser = MemUser.objects.get(biz_no=bizno[:3]+"-"+bizno[3:5]+"-"+bizno[5:10])
        # if int(row[8])>0 and not (row[2]=='예정고지 취소' or row[2]=='예정고지 기준금액 미만' or row[2]=='예정고지 기준금액 미만 및 폐업'):
        if int(row[8]) > 0 and row[2].strip() not in {'예정고지 취소', '예정고지 기준금액 미만', '예정고지 기준금액 미만 및 폐업'}:
          sql = "merge tbl_vat  as A Using (select '"+memuser.seq_no+"' as seq_no , '"+manageNo[0]+"' as work_yy, '"+manageNo[1]+"' as work_qt) as B "
          sql += "On A.seq_no = B.seq_no and A.work_yy = B.work_yy  and A.work_qt = B.work_qt "
          sql += "WHEN Matched Then  	Update set YN_15=" + str(row[8])
          sql += "	When Not Matched Then  "
          sql += "	INSERT 	values('"+memuser.seq_no+"','"+manageNo[0]+"','"+manageNo[1]+"',0,0,0,0,0,0,0,0,0,0,0,0,'',-1,'"+str(row[8])+"',0,0);"
          print(sql)
          cursor = connection.cursor()
          cursor.execute(sql)
          cursor.commit()
          print(str(row[0])+" : "+row[6]+" : "+'디비저장성공')
        else:
          print(str(row[0])+" : "+row[6]+" : "+row[2]+" : "+str(row[8]))
      except:
        print(str(row[0])+" : "+row[6]+" : "+str(row[5])+" : Mem_User DB 사업자번호 중복 또는 없음")
  elif flag=='사무수임사업장':
    strdel = "delete from EDI_COMWEL"
    cursor = connection.cursor()
    cursor.execute(strdel)    
    df = pd.read_excel(file_name,header=1, engine='openpyxl')
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
    for row in df.itertuples():  
      sql = f"insert into EDI_COMWEL values('{row[1]}','{row[2]}','{row[3]}','{row[4]}','{row[5]}','{row[6]}','{row[7]}','{row[8]}','{row[9]}','{row[10]}'"
      sql += f",'{row[11]}','{row[12]}','{row[13]}','{row[14]}','{row[15]}','{row[16]}','{row[17]}','{row[18]}')"
      print(sql)
      cursor = connection.cursor()
      cursor.execute(sql)
      cursor.commit()
    print('디비저장성공')  
  elif flag=='근로자고용정보현황조회':
    df = pd.read_excel(file_name,header=1, engine='openpyxl')
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
    strdel = "delete from tbl_Employ where seq_no='"+ manageNo[2]  + "'"
    cursor = connection.cursor()
    cursor.execute(strdel)
    for row in df.itertuples():
      if row[5]!="" and row[11]!="":
        Birth_Dt=""
        if    int(row[3][7:8])<3 : Birth_Dt	= "19" + row[3][:2] + "-" + row[3][2:4] +"-"+ row[3][4:6]
        elif  int(row[3][7:8])<5 :	Birth_Dt	= "20"+ row[3][:2] + "-" + row[3][2:4] +"-"+ row[3][4:6]
        elif int(row[3][7:8])<7 : Birth_Dt	= "19"+ row[3][:2] + "-" + row[3][2:4] +"-"+ row[3][4:6]
        else:			                Birth_Dt	= "20"+ row[3][:2] + "-" + row[3][2:4] +"-"+ row[3][4:6]
        # sql = "Merge tbl_employ as A Using (select '"+manageNo[2]+"' as seqno , '"+row[3][:8]+"' as juNum, '"+row[5]+"' as EnterDate) as B "
        # sql += "On A.seq_no = B.seqno and left(A.empJumin,8)=B.juNum and A.empRegDate=B.EnterDate "
        # sql += "when not matched then insert values('"+manageNo[2]

        sql = f"insert into Tbl_Employ values('{manageNo[2]}','','{row[3]}','{row[4]}','{Birth_Dt}','{row[5]}','{row[6]}','" 
        sql += "','','','','','','0');"
        print(sql)
        cursor = connection.cursor()
        cursor.execute(sql)
        cursor.commit()
        print('디비저장성공')
    time.sleep(1)
    driver.find_element(By.XPATH,'//*[@id="mf_wfm_content_edtGwanriNo"]').clear()
    time.sleep(0.5)
  elif flag=='보수총액신고대상자':
    df = pd.read_excel(file_name, header=0,dtype = {'주민(외국인)등록번호':str})
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환    c.관리번호, trim(a.biz_name), a.seq_no
    for row in df.itertuples():
      sql = "Merge 보수총액신고_고용 as A Using (select '"+str(manageNo[2])+"' as seqno , '"+str(workyear)+"' as work_yy, '"+str(row[2])+"' as ssn, '"+str(row[5])[:8]+"' as 산재취득일 , '"+str(row[9])[:8]+"' as 고용취득일) as B "
      sql += "On A.seq_no = B.seqno and A.seq_no=B.seqno and A.work_yy=B.work_yy and A.주민번호=B.ssn and A.산재취득일=B.산재취득일 and A.고용취득일=B.고용취득일 when not matched then "
      sql += "insert values('"+str(manageNo[2])
      sql += "','" + str(workyear)  #귀속년
      sql += "','" + str(row[1])  #성명
      sql += "','" + str(row[2]) #주번
      sql += "','" + str(row[3]) #부과구분
      sql += "','" + str(row[4]).replace(".0","") #건설업
      sql += "','" + str(row[5]).replace(".0","") #산재취득일
      sql += "','" + str(row[6]).replace(".0","") #산재싱실일
      sql += "',0"  #산재보수총액
      sql += ",'" + str(row[8]).replace(".0","") #근무지코드
      sql += "','" + str(row[9])[:8] #고용취득일
      sql += "','" + str(row[10])[:8] #고용상
      sql += "',0"  #총액상반기
      sql += ",0"  #총액하반기
      sql += "); "
      print(sql)
      cursor = connection.cursor()
      cursor.execute(sql)
      cursor.commit()
      print('디비저장성공')      
  elif flag=='보수총액신고대상자업데이트':
    df = pd.read_excel(file_name, header=2, engine='openpyxl',dtype = {'주민(외국인)등록번호':str})
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환    c.관리번호, trim(a.biz_name), a.seq_no
    print(df)
    for row in df.itertuples():
      if str(row[6])!='':
        sql = "Merge 보수총액신고_고용 as A Using (select '"+str(manageNo)+"' as seqno , '"+str(workyear)+"' as work_yy, '"+str(row[2]).replace("-","")+"' as ssn, '"+str(row[4]).replace("-","")+"' as 산재취득일 , '"+str(row[8]).replace("-","")+"' as 고용취득일) as B "
        sql += "On A.seq_no = B.seqno and A.seq_no=B.seqno and A.work_yy=B.work_yy and A.주민번호=B.ssn and A.산재취득일=B.산재취득일 and A.고용취득일=B.고용취득일 when  matched then "
        sql += "update set "        
        sql += " 산재연간보수총액="+str(row[6])+";"
        cursor = connection.cursor()
        cursor.execute(sql)
        cursor.commit()
      if str(row[10])!='':
        sql = "Merge 보수총액신고_고용 as A Using (select '"+str(manageNo)+"' as seqno , '"+str(workyear)+"' as work_yy, '"+str(row[2]).replace("-","")+"' as ssn, '"+str(row[4]).replace("-","")+"' as 산재취득일 , '"+str(row[8]).replace("-","")+"' as 고용취득일) as B "
        sql += "On A.seq_no = B.seqno and A.seq_no=B.seqno and A.work_yy=B.work_yy and A.주민번호=B.ssn and A.산재취득일=B.산재취득일 and A.고용취득일=B.고용취득일 when  matched then "
        sql += "update set "                
        sql += " 고용연간보수총액_상반기 = " + str(row[10])+";"
        print(sql)
        cursor = connection.cursor()
        cursor.execute(sql)
        cursor.commit()        
      if str(row[11])!='':
        sql = "Merge 보수총액신고_고용 as A Using (select '"+str(manageNo)+"' as seqno , '"+str(workyear)+"' as work_yy, '"+str(row[2]).replace("-","")+"' as ssn, '"+str(row[4]).replace("-","")+"' as 산재취득일 , '"+str(row[8]).replace("-","")+"' as 고용취득일) as B "
        sql += "On A.seq_no = B.seqno and A.seq_no=B.seqno and A.work_yy=B.work_yy and A.주민번호=B.ssn and A.산재취득일=B.산재취득일 and A.고용취득일=B.고용취득일 when  matched then "
        sql += "update set "                
        sql += " 고용연간보수총액_하반기 = " + str(row[11])+";"
        print(sql)
        cursor = connection.cursor()
        cursor.execute(sql)
        cursor.commit()
      print('디비저장성공')            
  elif flag=='은행':
    if manageNo[2]=='0003':   #기업은행
      df = pd.read_excel(file_name, header=None)
      df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
      print(df)
      for row in df.itertuples():
        sql = "Merge 은행거래내역 as A Using (select '"+str(manageNo[0])+"' as seqno , '"+str(manageNo[2])+"' as bankCode, '"+str(manageNo[3].replace(" ",""))+"' as bankNum, '"+str(row[1])+"' as tradeDatetime) as B "
        sql += "On A.seq_no = B.seqno and A.은행코드=B.bankCode and A.계좌번호=B.bankcode and A.거래일시=B.tradeDatetime when not matched then "
        sql += "insert values('"+str(manageNo[0])
        sql += "','" + str(manageNo[2])  #은행코드
        sql += "','" + str(manageNo[3].replace(" ",""))  #계좌번호
        sql += "'," + "(SELECT ISNULL(MAX(일련번호), 0) + 1 FROM 은행거래내역  Where Seq_No = '"+str(manageNo[0])+"' AND 은행코드 = '"+str(manageNo[2].replace(" ",""))+"' AND 계좌번호  = '"+str(manageNo[3].replace(" ",""))+"')"  #일련번호
        sql += ",'" + str(row[1]) #거래일시
        sql += "'," + str(row[2]) #출금
        sql += "," + str(row[3]) #입금
        sql += "," + str(row[4]) #잔액
        sql += ",'" + str(row[5]) #거래내용1
        sql += "','" + "" #거래내용2
        sql += "','" + "" #거래내용3
        sql += "','" + str(row[6]) #상대계좌
        sql += "','" + str(row[7]) #상대은행
        sql += "','" + str(row[8]) #cms코드
        sql += "','" + str(row[9]) #거래구분
        sql += "',0"  #미결재
        sql += ",'" + str(datetime.now())[:10].replace("-","") #crtdt
        sql += "','" + "" #slipdt
        sql += "','" + "" #slipAcntCd
        sql += "','" + "" #slipAcntCd
        sql += "N');" #slipYN
        print(sql)
        cursor = connection.cursor()
        cursor.execute(sql)
        cursor.commit()
      print('디비저장성공')    
    elif manageNo[2]=='0081':   #하나은행
      df = pd.read_excel(file_name, header=5)
      df = df.fillna('0')  # 3. excel 의 nill 값을 '' 로 변환
      print(df)      
      for row in df.itertuples():
        row5 = str(row[5]).replace(" ","")
        if row5=='':row5="0"
        row4 = str(row[4]).replace(" ","")
        if row4=='':row4="0"
        row6 = str(row[6]).replace(" ","")
        if row6=='':row6="0"
        sql = "Merge 은행거래내역 as A Using (select '"+str(manageNo[0])+"' as seqno , '"+str(manageNo[2])+"' as bankCode, '"+str(manageNo[3].replace(" ",""))+"' as bankNum, '"+str(row[1])+" "+str(row[7])+"' as tradeDatetime) as B "
        sql += "On A.seq_no = B.seqno and A.은행코드=B.bankCode and A.계좌번호=B.bankcode and A.거래일시=B.tradeDatetime when not matched then "
        sql += "insert values('"+str(manageNo[0])
        sql += "','" + str(manageNo[2])  #은행코드
        sql += "','" + str(manageNo[3].replace(" ",""))  #계좌번호
        sql += "'," + "(SELECT ISNULL(MAX(일련번호), 0) + 1 FROM 은행거래내역  Where Seq_No = '"+str(manageNo[0])+"' AND 은행코드 = '"+str(manageNo[2].replace(" ",""))+"' AND 계좌번호  = '"+str(manageNo[3].replace(" ",""))+"')"  #일련번호
        sql += ",'" + str(row[1])+" "+str(row[7]) #거래일시
        sql += "'," + row5 #출금
        sql += "," + row4 #입금
        sql += "," + row6.replace(".0","") #잔액
        sql += ",'" + str(row[3]) #거래내용1
        sql += "','" + str(row[2]) #거래내용2
        sql += "','" + str(row[8]) #거래내용3
        sql += "','" + "" #상대계좌
        sql += "','" + "" #상대은행
        sql += "','" + "" #cms코드
        sql += "','" + "" #거래구분
        sql += "',0"  #미결재
        sql += ",'" + str(datetime.now())[:10].replace("-","") #crtdt
        sql += "','" + "" #slipdt
        sql += "','" + "" #slipAcntCd
        sql += "','" + "" #slipAcntCd
        sql += "N');" #slipYN
        print(sql)
        cursor = connection.cursor()
        cursor.execute(sql)
        cursor.commit()      
  elif flag=='거래처등록':
    strsql = f"select trim(TrdCd_Semu) from Financial_Trdst_Ecnt where seq_no={manageNo} order by TrdCd_Semu/1"
    cursor = connection.cursor()
    cursor.execute(strsql);print(strsql)
    TrdCd_Semus = cursor.fetchall()
    connection.commit()
     
    df = pd.read_excel(file_name,header=0)
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
    for row in df.itertuples():
      TrdCd_Semu = str(row[1])
      if len(TrdCd_Semu)==3:TrdCd_Semu = "00"+TrdCd_Semu 
      elif len(TrdCd_Semu)==4:TrdCd_Semu = "0"+TrdCd_Semu 
      if TrdCd_Semu not in [row[0] for row in TrdCd_Semus]:
        strIns = f"insert into Financial_Trdst_Ecnt values({TrdCd_Semu},'{manageNo}','{TrdCd_Semu}','{row[2]}','{row[3].replace('-','')}','{row[2]}','{row[3]}','{driver}')"
        print(strIns)
        cursor.execute(strIns)
  elif flag=='CMS회원목록':
    if f'{today}_회원목록' in file_name :
      strsql = f"select trim(seq_CMS) from CMS_MEMBER  order by seq_CMS"
      cursor = connection.cursor()
      cursor.execute(strsql);print(strsql)
      seq_CMSs = cursor.fetchall()
      connection.commit()

      df = pd.read_excel(file_name,header=0)
      df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
      
      for row in df.itertuples():  
        seq_MEMBER = seq_BIZNO = seq_SEMU = ''
        strsql = f"select a.seq_no,biz_no,biz_type from mem_user a,mem_deal b  where a.seq_no=b.seq_no and trim(biz_name) = '{row[2].strip()}'  "
        cursor = connection.cursor()
        cursor.execute(strsql);print(strsql)
        mem_user = cursor.fetchone()
        connection.commit()
        if mem_user:
          seq_MEMBER = mem_user[0];seq_BIZNO = mem_user[1]
          strsql = f"select trim(seq_SEMU) from cms_semusarang_trdst where semu_bizno='{seq_BIZNO}'"
          cursor = connection.cursor()
          cursor.execute(strsql);print(strsql)
          seq_SEMUs = cursor.fetchone()
          connection.commit()
          if seq_SEMUs:seq_SEMU = seq_SEMUs[0]
        if row[1].strip() not in [row[0] for row in seq_CMSs]:
          strIns = f"insert into CMS_MEMBER values('{row[1]}','{row[2]}','{row[3]}','{row[4]}','{row[7]}','{row[8]}','{row[9]}','{row[10]}','{seq_MEMBER}','{seq_SEMU}')"
          print(strIns)
          cursor.execute(strIns)        
  elif flag=='CMS출금내역업로드':
    if f'{today}_CMS출금조회내역' in file_name :
      df = pd.read_excel(file_name,header=0)
      df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
      for row in df.itertuples():
        strsql = f"select trim(seq_SEMU),trim(seq_MEMUSER) from CMS_MEMBER where trim(seq_CMS)='{row[3].strip()}'"
        cursor = connection.cursor()
        cursor.execute(strsql);print(strsql)
        CMS_MEMBER = cursor.fetchone()
        connection.commit()
        if CMS_MEMBER and CMS_MEMBER[1]:
          if row[9] == '출금실패' or  row[9] == '재출금실패':
            str_mg = f"Merge tbl_mng_jaroe as A Using (select '{CMS_MEMBER[1]}' as seq_no, '{row[1][:4]}' as work_yy, '{int(row[1][4:6])}' as work_mm  ) as B "
            str_mg += "On A.seq_no = B.seq_no and A.work_yy = B.work_yy  and A.work_mm = B.work_mm  "
            str_mg += f"WHEN Matched Then  Update set YN_10=1 "
            str_mg += "	When Not Matched Then  "
            str_mg += "	INSERT (seq_no,work_YY,work_MM,YN_1,YN_2,YN_3,YN_4,YN_5,YN_6,YN_7,YN_8,YN_9,YN_10,YN_11,YN_12,YN_13,YN_14,bigo) "
            str_mg += f"	values({CMS_MEMBER[1]},{row[1][:4]},{int(row[1][4:6])},0,0,0,0,0,0,0,0,0,1,0,0,0,0,'');"          
            print(str_mg)
            cursor.execute(str_mg)  
      for row in df.itertuples():
        strsql = f"select trim(seq_SEMU),trim(seq_MEMUSER) from CMS_MEMBER where trim(seq_CMS)='{row[3].strip()}'"
        cursor = connection.cursor()
        cursor.execute(strsql);print(strsql)
        CMS_MEMBER = cursor.fetchone()
        connection.commit()
        if CMS_MEMBER and CMS_MEMBER[1]:
          if '성공' in row[9]:    
            str_mg = f"Merge tbl_mng_jaroe as A Using (select '{CMS_MEMBER[1]}' as seq_no, '{row[1][:4]}' as work_yy, '{int(row[1][4:6])}' as work_mm  ) as B "
            str_mg += "On A.seq_no = B.seq_no and A.work_yy = B.work_yy  and A.work_mm = B.work_mm  "
            str_mg += f"WHEN Matched Then  Update set YN_10=0 ;"         
            print(str_mg)
            cursor.execute(str_mg)             
  elif flag=='이카운트분개장업로드':
    str_del = f"delete from Ds_slipledgr_Ecount where seq_no='{manageNo[0]}' and work_YY='{manageNo[1]}' and work_QT = '{manageNo[2]}'"
    cursor = connection.cursor()
    cursor.execute(str_del)    
    connection.close()  
    file_name = driver;print(file_name)
    df = pd.read_excel(file_name,header=0)
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
    for i,row in enumerate(df.itertuples()):
      if row[7]!='':#마지막 공란에러를 발생시켜야 원본파일이 삭제되지 않는다
        strsql_trd = "select AcntCd_Semu,AcntNm_Semu, "
        if row[5].strip()=="" :
          strsql_trd += " '' TrdCd_Semu ,'' TrdNm_Semu ,'' Trdst_Num  "
        else:
          strsql_trd += f" isnull((select trdcd_semu from financial_Trdst_Ecnt where seq_no={manageNo[0]} and trdCd_ecnt=rtrim('{row[5].strip()}')) ,'') TrdCd_Semu, "
          strsql_trd += f" isnull((select trdnm_semu from financial_Trdst_Ecnt where seq_no={manageNo[0]} and trdCd_ecnt=rtrim('{row[5].strip()}')) ,'') TrdNm_Semu,"
          strsql_trd += f" isnull((select trdst_Num from financial_Trdst_Ecnt where seq_no={manageNo[0]} and trdCd_ecnt=rtrim('{row[5].strip()}')) ,'') Trdst_Num "			
        strsql_trd += f" from Financial_AcntCd_Ecnt where AcntCd_Ecnt like '{int(row[7])}%'";print(strsql_trd)
        cursor = connection.cursor()
        cursor.execute(strsql_trd)
        result = cursor.fetchone()
        columns = [col[0] for col in cursor.description]
        connection.commit()
        if result:
          rs2 = dict(zip(columns, result))
          TrdCd_Semu = rs2["TrdCd_Semu"].strip()
          TrdNm_Semu = rs2["TrdNm_Semu"].strip()
          Trn_Num = rs2["Trdst_Num"].strip()
          Trn_Date = row[2].replace("/","").strip()
          if TrdCd_Semu == "" :
            tmpBizNo = row[5].strip()
            if len(tmpBizNo)<=12:tmpBizNo = f"{row[5][:3]}-{row[5][3:5]}-{row[5][5:10]}"
            strsql = f"select trdcd_semu,trdNm_semu from financial_trdst_ecnt where seq_no={manageNo[0]} and trdst_num='{tmpBizNo}'"
            cursor.execute(strsql)
            result2 = cursor.fetchone()
            connection.commit()
            if result2:
              TrdCd_Semu=result2[0]
              TrdNm_Semu=result2[1]
              strUpt = f"update  Financial_Trdst_Ecnt set TrdCd_Ecnt='{row[5].strip()}' where seq_no='{manageNo[0]}' and trdcd_semu='{TrdCd_Semu}'";print(strUpt)
              cursor.execute(strUpt)
          AcntCd_Semu = rs2["AcntCd_Semu"].strip()
          AcntNm_Semu = rs2["AcntNm_Semu"].strip()
          if TrdCd_Semu=="" : TrdNm_Semu=row[6].strip()
          CrDr = "";TranAmt_Cr = TranAmt_Dr = 0
          if row[9] is not None and row[9]!='':
            strtmp = str(row[9]).split(".");  str_amt = strtmp[0].replace(",","") 
            CrDr = "3";TranAmt_Cr = int(str_amt);TranAmt_Dr = 0
          if row[10] is not None and row[10]!='':
            strtmp = str(row[10]).split(".");  str_amt = strtmp[0].replace(",","")
            CrDr = "4";TranAmt_Dr = int(str_amt);TranAmt_Cr = 0
          str_ins = f"insert into Ds_slipledgr_Ecount values('{manageNo[0]}','{manageNo[1]}','{manageNo[2]}','{row[1].strip()}"
          str_ins += f"','{AcntCd_Semu}','{AcntNm_Semu}','{Trn_Date}','{row[11].strip()}','{TrdCd_Semu}','{TrdNm_Semu}','{Trn_Num}'"
          str_ins += f",'{int(row[3])}','{row[4].strip()}','{CrDr}',{int(TranAmt_Cr)+int(TranAmt_Dr)},'{TranAmt_Cr}','{TranAmt_Dr}'" 
          str_ins += f",'{int(row[7])}','{row[8].strip()}','{row[5].strip()}','{row[6].strip()}','{today}')" 
          cursor.execute(str_ins)
        else:
          print(f"이카운트 계정코드 미등록으로 업로드 안됨 : {i},  sql :{strsql_trd}")
      else:
        print(f'{i}: 계정코드가 공란입니다 END OF FILE')
  elif flag=='연말정산대상근로자':
    try:
      df = pd.read_excel(file_name, header=0,dtype = {'주민등록번호':str})
    except:
      return False
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환    c.관리번호, trim(a.biz_name), a.seq_no
    for row in df.itertuples():
      sql = f"Merge 연말정산일괄제공근로자 as A Using (select '{manageNo[2]}' as seqno , '{manageNo[10]}' as work_yy, '{row[2]}' as ssn) as B "
      sql += "On A.seq_no = B.seqno and A.work_yy=B.work_yy and A.주민등록번호=B.ssn   "
      sql += f" when matched then update set 동의일자='{row[6]}',동의취소일자='{row[7]}',비고='조회일 : " + str(datetime.now())[:10].replace("-","") +"'"
      sql += f"when not matched then insert values('{manageNo[2]}','{manageNo[10]}','{row[1]}','{row[2]}','{row[3]}','{row[4]}','{row[5]}','{row[6]}','{row[7]}','0"   #PDFDownYN
      sql += "','" + "조회일 : " + str(datetime.now())[:10].replace("-","")  #비고 - 조회일
      sql += "'); "
      print(sql)
      cursor = connection.cursor()
      cursor.execute(sql)
      cursor.commit()
      print('연말정산일괄제공근로자 디비저장성공') 
  elif flag=='간이지급명세서':
    try:
      df = pd.read_excel(file_name, header=0)
    except:
      print('간이지급명세서 전자신고 결과 엑셀파일 읽기 실패')
      return False
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환    c.관리번호, trim(a.biz_name), a.seq_no
    for row in df.itertuples():
      sql = "Merge 지급조서간이소득 as A Using (select '"+manageNo[0]+"' as 과세년도 , '"+manageNo[1]+"' as 신고서종류, '"+manageNo[2]+"' as 사업자번호) as B "
      sql += "On A.과세년도 = B.과세년도 and A.신고서종류=B.신고서종류 and A.사업자번호=B.사업자번호  "
      sql += "when matched then update set 신고구분='"+ str(row[8])+"',접수자='"+str(row[2])+"',제출건수='"+str(row[9])+"',제출금액='"+str(row[10])+"',접수번호='"+str(row[1])+"',접수일시='"+str(row[11])+"',작성일자=getdate() "
      sql += "when not matched then insert values('"+str(row[3])  #신고서종류
      sql += "','" + str(row[4])  #사업자번호
      sql += "','" + str(row[5])  #상호
      sql += "','" + str(row[6]) #과세년월
      sql += "','" + str(row[8]) #정기/기한후
      sql += "','" + str(row[2]) #접수자
      sql += "','" + str(row[9]) #제출건수
      sql += "','" + str(row[10]) #제출금액
      sql += "','" + str(row[11]) #접수번호
      sql += "','" + str(row[1]) #접수번호
      sql += "',getdate()" #접수일
      sql += "); "
      print(sql)
      cursor = connection.cursor()
      cursor.execute(sql)
      cursor.commit()
      print('간이지급명세서 전자신고 결과 디비저장성공') 
  if flag!='이카운트분개장업로드':os.remove(file_name)
  return True

#zip파일 압축해제
def Extract_Download_File(zip_path, destination_folder, password=None):
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    # Open the zip file
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        # If a password was provided, use it to extract the file
        if password:
            zip_ref.setpassword(password.encode())

        # Extract all the contents into the destination folder
        zip_ref.extractall(destination_folder)

    return True