import json
import datetime,calendar
import math
import os, glob,traceback
import natsort
import re
import PyPDF2

from django.conf import settings
from datetime import datetime as dt, date, timedelta
from django.utils import timezone
from django.db import connection
from urllib.parse import unquote
from django.views.decorators.csrf import csrf_exempt
from django.db import connection
from django.http import JsonResponse, HttpResponseBadRequest
from django.core.serializers import serialize
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect

from django.views.decorators.http import require_GET


from app.models import MemAdmin
from app.models import MemDeal
from app.models import MemUser  # KijangMember는 "기장회원관리"와 관련된 모델이라고 가정합니다.
from app.models import userProfile


from django.db.models import Q

from admins.utils import send_kakao_notification,kijang_member_popup,tbl_mng_jaroe_update,getSentMails,sendMail,mid_union,PDF_Merge


@login_required(login_url="/login/")
def index(request):
  context = {}
  admin_grade     = request.session.get('Admin_Grade')
  admin_biz_level = request.session.get('Admin_Biz_Level')
  admin_biz_area  = request.session.get('Admin_Area')
  ADID = request.session.get('ADID')
  flag = request.GET.get("flag")

  arr_ADID = []

  if admin_grade != "SA":
      if admin_biz_level == "세무사":
          # Query for admin_id
          arr_ADID = MemAdmin.objects.filter(
              ~Q(grade="SA"),
              ~Q(biz_level="세무사"),
              ~Q(del_yn="y"),
              admin_biz_area = admin_biz_area
          ).order_by('admin_id').values_list('admin_id', flat=True)
  else:  # admin_grade == "SA"
      arr_ADID = list(MemAdmin.objects.filter(
          ~Q(grade="SA"),
          ~Q(biz_level="세무사"),
          ~Q(del_yn="y")
      ).order_by("admin_id").values_list('admin_id', flat=True))
      arr_ADID.insert(0, "전체")
  if not ADID:
    ADID = arr_ADID[0] if arr_ADID else ""

  work_YY = request.GET.get('work_YY', '')
  work_MM = request.GET.get('work_MM', '')
  today = dt.now()
  
  if not work_MM:
    work_MM = request.session.get('workmonth')
    if not work_MM:
      work_MM = today.month
  corpYear = today.year
  if int(work_MM) <= 4 :
    corpYear = today.year-1
  if not work_YY:
    work_YY = request.session.get('workyearStat')
    if not work_YY:
      if int(work_MM) <= 4 :
        work_YY = today.year - 1
      else:
        work_YY = today.year
    else:
      work_YY = int(work_YY)

  today = dt.now()
  cur_date = today.strftime("%Y%m%d")
  work_QT = request.session.get("work_QT") 
  print(f"a:{work_QT}")
  if not work_QT:
      if f"{work_YY}0401" <= cur_date <= f"{work_YY}0630":
        work_QT = 1;sql_QT ='(1,2,3)'
      elif f"{work_YY}0701" <= cur_date <= f"{work_YY}0930":
        work_QT = 2;sql_QT ='(4,5,6)'
      elif f"{work_YY}1001" <= cur_date <= f"{work_YY}1231":
        work_QT = 3;sql_QT ='(7,8,9)'
      elif f"{int(work_YY) + 1}0101" <= cur_date <= f"{int(work_YY) + 1}0331":
        work_QT = 4;sql_QT ='(10,11,12)'
      else:
        work_QT = 4;sql_QT ='(10,11,12)'
  else:
      work_QT = int(work_QT)
  print(f"b:{work_QT}")
  request.session['work_QT'] = work_QT      
  request.session['workyearStat'] = work_YY
  request.session['workmonth'] = work_MM
  
  corpYears = list(range(corpYear, corpYear - 6, -1))
  # print(corpYears)
  context['corpYears'] = corpYears
  context['admin_grade'] = admin_grade
  context['admin_biz_level'] = admin_biz_level
  context['arr_ADID'] = json.dumps(list(arr_ADID))
  context['flag'] = flag
  context['ADID'] = ADID
  request.session['ADID'] = ADID  

  request.session.save()

  
  gridTitle="기장보고서·가결산 : 법인"    
  if flag=="income":        
    gridTitle="기장보고서·가결산 : 개인"    
  templateMenu = 'admin/mng_statement.html'; 
  context['gridTitle'] = gridTitle  
  return render(request, templateMenu,context)

#대상자 리스트
# 대상자 리스트
def mng_statement(request):
    flag = request.GET.get('flag')
    str_biztype = "('1','2','3')"
    if flag == "income":
        str_biztype = "('4','5','6')"

    ADID = request.GET.get('ADID')
    if not ADID:
        ADID = request.session.get('ADID')  # 전체 선택시 ADID=""상태가 된다
    request.session['ADID'] = ADID

    work_YY = request.GET.get('work_YY', '')
    request.session['workyearStat'] = work_YY
    request.session.save()

    work_QT = request.GET.get('work_QT', '')
    sql_QT = ""
    if int(work_QT) == 1:
        sql_QT = '(1,2,3)'
    elif int(work_QT) == 2:
        sql_QT = '(4,5,6)'
    elif int(work_QT) == 3:
        sql_QT = '(7,8,9)'
    elif int(work_QT) == 4:
        sql_QT = '(10,11,12)'

    if request.method == 'GET':
        s_sql = ""
        if ADID != "전체":
            s_sql = f" AND b.biz_manager = N'{ADID}'"

        sql_query = f"""
          DECLARE @year    char(4)      = '{work_YY}';
          DECLARE @quarter nvarchar(10) = N'{work_QT}분기';

          DECLARE @reportLike nvarchar(100) = N'%' + @year + N'년 ' + @quarter + N'%';
          DECLARE @kakaoLike  nvarchar(200) = N'%' + @year + N'년 ' + @quarter 
                                              + N' 재무제표 및 예상세액 등에 대한 기장현황보고서%';

          ;WITH J AS (  -- 연도별 최신 전표일자: 'MM-DD' → 'YYYY-MM-DD'로 보정
            SELECT 
                d.seq_no,
                MAX(
                  CASE 
                    -- 1) 'MM-DD' 또는 'M-D' 형태: @year 붙여서 'YYYY-MM-DD'로 변환
                    WHEN d.tran_dt LIKE '%-%' 
                        AND CHARINDEX('-', d.tran_dt) > 0
                        AND LEN(d.tran_dt) BETWEEN 4 AND 5
                        AND ISDATE(@year + '-' + d.tran_dt) = 1
                      THEN CAST(@year + '-' + d.tran_dt AS date)

                    -- 2) 'YYYY-MM-DD' 같은 표준 문자열
                    WHEN ISDATE(d.tran_dt) = 1
                      THEN CAST(d.tran_dt AS date)

                    -- 3) 'YYYYMMDD' 순수 숫자 8자리
                    WHEN LEN(d.tran_dt) = 8 AND d.tran_dt NOT LIKE '%[^0-9]%'
                        AND ISDATE(STUFF(STUFF(d.tran_dt,5,0,'-'),8,0,'-')) = 1
                      THEN CAST(STUFF(STUFF(d.tran_dt,5,0,'-'),8,0,'-') AS date)

                    ELSE NULL
                  END
                ) AS lastTranDt
            FROM ds_slipledgr2 AS d
            WHERE (ISNUMERIC(d.work_yy) = 1 AND CONVERT(int, d.work_yy) = @year)
            GROUP BY d.seq_no
          )
          SELECT
            b.biz_manager as groupManager,
            biz_type,
            a.seq_no,
            a.biz_name,
            r.mail_subject       AS isMail,        -- 메일 여부
            k.Kakao_tempCode     AS isKakaoSend,   -- 카카오톡 여부
            ISNULL(j.lastTranDt, CONVERT(date, '1900-01-01')) AS jaroeCount,
            n.MH_Name,
            n.sum_MH_Amt,
            ISNULL(ex.execs_cnt, 0)        AS execs_cnt,        -- ★ 현직 임원 수
            ISNULL(sh.stockholders_cnt, 0) AS stockholders_cnt  -- ★ 주주 수
          FROM mem_user  AS a
          JOIN mem_deal  AS b ON a.seq_no = b.seq_no
          LEFT JOIN J AS j
            ON j.seq_no = a.seq_no
          OUTER APPLY (
              SELECT TOP (1) m.mail_subject
              FROM Tbl_Mail AS m
              WHERE m.seq_no = a.seq_no
                AND m.mail_class = 'report'
                AND m.mail_subject LIKE @reportLike
          ) AS r
          OUTER APPLY (
              SELECT TOP (1) s.Kakao_tempCode
              FROM Tbl_OFST_KAKAO_SMS AS s
              WHERE s.seq_user    = a.seq_no
                AND s.send_result = 'Y'
                /* send_dt가 datetime이면: AND YEAR(s.send_dt) = CAST(@year AS int) */
                AND LEFT(s.send_dt,4) = @year
                AND s.contents LIKE @kakaoLike
              ORDER BY s.send_dt DESC
          ) AS k
          OUTER APPLY (
              SELECT TOP (1) 
                H.MH_Name,
                SUM(H.MH_Amt) OVER (PARTITION BY H.SEQ_NO) AS sum_MH_Amt
              FROM Diag_capital AS H
              WHERE H.SEQ_NO = a.seq_no
              AND (
                    (ISNUMERIC(H.MH_DcRate) = 1 AND CAST(H.MH_DcRate AS decimal(18,6)) < 2)
                    OR (ISNUMERIC(H.MH_DcRate) = 0 AND H.MH_DcRate < '2')
                  )
              ORDER BY H.MH_Amt DESC
          ) AS n
          -- ★ 추가: 임원수 카운트
          OUTER APPLY (
              SELECT COUNT(*) AS execs_cnt
              FROM lawregistration WITH(NOLOCK)
              WHERE seq_no = a.seq_no
                AND execflag IN (N'대표이사', N'사내이사', N'감사')
                AND ISNULL(fireDate,'') = ''
          ) AS ex
          -- ★ 추가: 주주수 카운트
          OUTER APPLY (
              SELECT COUNT(*) AS stockholders_cnt
              FROM Tbl_StckHolderList WITH(NOLOCK)
              WHERE seq_no = a.seq_no
          ) AS sh
          WHERE ISNULL(a.duzon_ID, '') <> ''
            AND b.keeping_YN = 'Y'
            AND ISNULL(a.Del_YN,'N') <> 'Y'
            --AND a.Biz_Type IN {str_biztype}
            {s_sql}
          ORDER BY  a.Biz_Name ASC;  
        """

        recordset = []
        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            columns = [col[0] for col in cursor.description]  # 컬럼명 가져오기
            results = cursor.fetchall()
            for row in results:
                # 문자열 값이면 strip() 적용, 아니면 그대로 유지
                row_dict = {columns[i]: (value.strip() if isinstance(value, str) else value)
                            for i, value in enumerate(row)}

                # 파일 경로 구성 (biz_name 기준)
                folder_path = os.path.join('static/cert_DS/', row_dict["biz_name"], work_YY, "기장보고서", f"{work_QT}분기")
                file_name = "300.pdf"
                file_path = os.path.join(folder_path, file_name)

                # 파일 존재 여부 체크
                if os.path.exists(file_path):
                    row_dict["file_exists"] = True
                    # 마지막 수정 시각 가져오기 (timestamp → datetime 변환)
                    mtime = os.path.getmtime(file_path)
                    dt_aware = datetime.datetime.fromtimestamp(  # ← 클래스 경로를 풀로
                        mtime, tz=timezone.get_current_timezone()
                    )
                    dt_local = timezone.localtime(dt_aware)
                    row_dict["file_date"] = dt_local.strftime("%Y-%m-%d")
                else:
                    row_dict["file_exists"] = False
                    row_dict["file_date"] = ""

                recordset.append(row_dict)

        return JsonResponse(list(recordset), safe=False)
    else:
        return JsonResponse({'error': 'Invalid request method.'}, status=400)

  
# 0.표지데이터
def section0_data(request):
    """필요 필드만 조회해서 JSON으로 반환"""
    try:
        seq_no = int(request.GET.get("seq_no", "0"))
    except ValueError:
        return HttpResponseBadRequest("invalid seq_no")

    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT a.biz_name, a.ceo_name, a.email, a.biz_no, a.biz_type,a.Reg_date,  c.admin_name, fiscalMM as FiscalMM
              FROM mem_user a
              JOIN mem_deal b ON a.seq_no = b.seq_no
              JOIN mem_admin c ON b.biz_manager = c.admin_id
             WHERE a.seq_no = %s
            """,
            [seq_no],
        )
        row = fetchone_dict(cur)

    if not row:
        return JsonResponse({"ok": False, "error": "not_found"}, status=404)

    # 직렬화-friendly (날짜는 문자열로)
    reg_date = row.get("Reg_date")
    if reg_date is not None:
        reg_date = str(reg_date)[:19]  # "YYYY-MM-DD HH:MM:SS" 정도로

    data = {
        "ok": True,
        "seq_no": seq_no,
        "biz_name": row.get("biz_name") or "",
        "ceo_name": row.get("ceo_name") or "",
        "email": row.get("email") or "",
        "biz_no": row.get("biz_no") or "",
        "biz_type": row.get("biz_type") or 0,
        "FiscalMM": row.get("FiscalMM") or 12,
        "Reg_date": reg_date or "",
        "admin_name": row.get("admin_name") or "",
    }
    return JsonResponse(data, json_dumps_params={"ensure_ascii": False})  
def biz_accounts_section(request):
    """
    A4 보고서용: 사업용계좌 개설현황 섹션 (server-side partial)
    - 쿼리스트링: ?seq_no=...
    - 최신순 3줄만 표시, 총 건수는 헤더에 노출
    """
    try:
        seq_no = int(request.GET.get("seq_no") or 0)
    except Exception:
        return HttpResponseBadRequest("seq_no is required")

    memuser = MemUser.objects.filter(seq_no=seq_no).only("seq_no", "biz_type").first()
    if not memuser:
        return HttpResponseBadRequest("invalid seq_no")

    sql = r"""
        SELECT 등록번호, 납세계좌구분, 은행명, 계좌번호, 등록일자
        FROM 사업용계좌신고현황 WITH(NOLOCK)
        WHERE Seq_No = %s
        ORDER BY 등록번호 DESC
    """
    all_rows = []
    with connection.cursor() as cur:
        cur.execute(sql, [seq_no])
        for 등록번호, 납세계좌구분, 은행명, 계좌번호, 등록일자 in cur.fetchall():
            all_rows.append({
                "reg_no": (등록번호 or "").strip(),
                "acc_type": (납세계좌구분 or "").strip(),
                "bank": (은행명 or "").strip(),
                "acct": (계좌번호 or "").strip(),
                "reg_dt": (등록일자 or "").strip(),
            })

    total_cnt = len(all_rows)

    # 최신 3줄만 화면에, 부족하면 공백행으로 패딩
    MAX_ROWS = 3
    rows = all_rows[:MAX_ROWS]
    while len(rows) < MAX_ROWS:
        rows.append({"reg_no": "", "acc_type": "", "bank": "", "acct": "", "reg_dt": ""})

    ctx = {
        "rows": rows,
        "total_cnt": total_cnt,
    }
    return render(request, "report/sections/biz_accounts.html", ctx)

# 2. 매출 - 재무데이터조회
def _to_int(v, default=None):
    try:
        return int(v)
    except (TypeError, ValueError):
        return default
def _json_error(msg, status=400):
    return JsonResponse({"ok": False, "error": msg}, status=status)
def _today_str():
    return timezone.localdate().strftime("%Y-%m-%d")
def _month_end(year: int, month: int) -> datetime.date:
    last = calendar.monthrange(year, month)[1]
    return datetime.date(year, month, last)
def _fmt_mmdd(fiscalMM_plus1: int) -> str:
    """MM-DD 문자열 ('07-01' 등)"""
    mm = fiscalMM_plus1
    if mm == 13:  # 12월 결산의 +1 은 사용하지 않음(Else 분기에서만 사용)
        mm = 1
    return f"{mm:02d}-01"
def _period_clause(year:int, fiscalMM:int):
    """
    VB Else 분기의 기간 조건:
      (Work_YY = year AND tran_dt < '0(fiscalMM+1)-01')
      OR
      (Work_YY = year-1 AND tran_dt >= '0(fiscalMM+1)-01')
    """
    cutoff = _fmt_mmdd(fiscalMM + 1)
    clause = """
      (
        (Work_YY = %s AND tran_dt < %s)
        OR
        (Work_YY = %s AND tran_dt >= %s)
      )
    """
    params = [year, cutoff, year-1, cutoff]
    return clause, params
def _fmt_cutoff_mmdd(fiscalMM_plus1: int) -> str:
    mm = fiscalMM_plus1
    if mm == 13: mm = 1
    return f"{mm:02d}-01"

def _local_today():
    """서버 타임존 기준 '오늘' (Django USE_TZ 환경 고려)"""
    try:
        return timezone.localdate()
    except Exception:
        return datetime.date.today()
def _compute_enddate_and_flags(biz_type:int, work_yy:int, today:date|None=None)-> tuple[str, str, bool]:

    d = today or _local_today()
    cur = int(d.strftime("%Y%m%d"))

    y = int(work_yy)
    # 경계값(정수) 생성
    y_0101 = int(f"{y}0101")
    y_0416 = int(f"{y}0416")
    y_0601 = int(f"{y}0601")
    y_0701 = int(f"{y}0701")
    y_0930 = int(f"{y}0930")
    y_1001 = int(f"{y}1001")
    y_1231 = int(f"{y}1231")
    y1_0415 = int(f"{y+1}0415")

    endDate = "12-31"
    work_mm = "12"
    if (cur > y_0416) and (cur < y_0601):
        endDate, work_mm = "03-31", "3"
    elif (cur > y_0701) and (cur < y_0930):
        endDate, work_mm = "06-30", "6"
    elif (cur > y_1001) and (cur < y_1231):
        endDate, work_mm = "09-30", "9"
    elif (cur > y_0101) and (cur < y1_0415):
        endDate, work_mm = "12-31", "12"

    # VB: If biz_type<=3 Then isDisabled="false" Else "true"
    # => 법인/일반사업자(<=3)는 입력가능(False), 개인(>3)은 비활성(True)로 해석
    is_disabled = not (biz_type <= 3)
    return endDate, work_mm, is_disabled
def _period_clause_params(work_yy:int, fiscalMM:int):
    """
    (Work_YY = work_yy AND tran_dt < '0(fiscalMM+1)-01')
     OR (Work_YY = work_yy-1 AND tran_dt >= '0(fiscalMM+1)-01')
    """
    cutoff = _fmt_cutoff_mmdd(fiscalMM + 1)
    clause = """
      (
        (Work_YY = %s AND tran_dt < %s)
        OR
        (Work_YY = %s AND tran_dt >= %s)
      )
    """
    params = [work_yy, cutoff, work_yy-1, cutoff]
    return clause, params

def _get_base_pl_aggregates(seq_no:int, work_yy:int, fiscalMM:int):
    """
    PL 핵심 집계: 매출, 매출원가, 급여, 판관비, 영업외수익, 영업외비용, 법인세
    + 집계 검증용 로그 출력
    """
    with connection.cursor() as cur:
        if fiscalMM == 12:
            sql = """
            SELECT
              SUM(CASE WHEN acnt_cd BETWEEN 401 AND 430 THEN tranAmt_dr - tranAmt_cr ELSE 0 END) AS sales,
              SUM(CASE WHEN acnt_cd BETWEEN 451 AND 470 THEN tranAmt_cr ELSE 0 END)             AS cogs,
              SUM(CASE WHEN acnt_cd BETWEEN 801 AND 810 THEN tranAmt_cr - tranAmt_dr ELSE 0 END) AS salary,
              SUM(CASE WHEN acnt_cd BETWEEN 811 AND 900 THEN tranAmt_cr - tranAmt_dr ELSE 0 END) AS sga,
              SUM(CASE WHEN acnt_cd BETWEEN 901 AND 950 THEN tranAmt_dr - tranAmt_cr ELSE 0 END) AS nonOpInc,
              SUM(CASE WHEN acnt_cd BETWEEN 951 AND 997 THEN tranAmt_cr - tranAmt_dr ELSE 0 END) AS nonOpExp,
              SUM(CASE WHEN acnt_cd BETWEEN 998 AND 999 THEN tranAmt_cr - tranAmt_dr ELSE 0 END) AS tax
            FROM DS_SlipLedgr2
            WHERE seq_no=%s
              AND work_yy=%s
              AND ((acnt_cd BETWEEN 401 AND 999) OR (acnt_cd BETWEEN 146 AND 253))
              AND acnt_cd<>150
              AND Remk<>N'손익계정에 대체'
              AND tran_dt<>'00-00'
            """
            params=[seq_no, work_yy]
        else:
            clause, p = _period_clause_params(work_yy, fiscalMM)
            sql = f"""
            SELECT
              SUM(CASE WHEN acnt_cd BETWEEN 401 AND 430 THEN tranAmt_dr - tranAmt_cr ELSE 0 END),
              SUM(CASE WHEN acnt_cd BETWEEN 451 AND 470 THEN tranAmt_cr ELSE 0 END),
              SUM(CASE WHEN acnt_cd BETWEEN 801 AND 810 THEN tranAmt_cr - tranAmt_dr ELSE 0 END),
              SUM(CASE WHEN acnt_cd BETWEEN 811 AND 900 THEN tranAmt_cr - tranAmt_dr ELSE 0 END),
              SUM(CASE WHEN acnt_cd BETWEEN 901 AND 950 THEN tranAmt_dr - tranAmt_cr ELSE 0 END),
              SUM(CASE WHEN acnt_cd BETWEEN 951 AND 997 THEN tranAmt_cr - tranAmt_dr ELSE 0 END),
              SUM(CASE WHEN acnt_cd BETWEEN 998 AND 999 THEN tranAmt_cr - tranAmt_dr ELSE 0 END)
            FROM DS_SlipLedgr2
            WHERE seq_no=%s
              AND {clause}
              AND ((acnt_cd BETWEEN 401 AND 999) OR (acnt_cd BETWEEN 146 AND 253))
              AND acnt_cd<>150
              AND Remk<>N'손익계정에 대체'
              AND tran_dt<>'00-00'
            """
            params=[seq_no]+p

        row = _fetchone(cur, sql, params) or [0,0,0,0,0,0,0]
        nums = [ _to_int(x) for x in row ]

        # 🔥 디버그 출력
        # print("\n[DEBUG] _get_base_pl_aggregates")
        # print(f" seq_no={seq_no} work_yy={work_yy} fiscalMM={fiscalMM}")
        # print(f" 매출      : {nums[0]:,}")
        # print(f" 매출원가  : {nums[1]:,}")
        # print(f" 급여      : {nums[2]:,}")
        # print(f" 기타판관비: {nums[3]:,}")
        # print(f" 영업외수익: {nums[4]:,}")
        # print(f" 영업외비용: {nums[5]:,}")
        # print(f" 법인세    : {nums[6]:,}")
        # print("----------------------------------------------------")

        return nums

def _get_latest_worktax_row(base_amount: float):
    """WorkTax 최신연도 구간 찾아 산출세액 요소(prgrs_ddct_amt, taxat_stand_min, taxrat) 반환"""
    with connection.cursor() as cur:
        sql = """
        SELECT TOP 1 prgrs_ddct_amt, taxat_stand_min, taxrat
          FROM WorkTax WITH (NOLOCK)
         WHERE taxat_stand_min <= %s
           AND taxat_stand_max  > %s
         ORDER BY YY DESC
        """
        row = _fetchone(cur, sql, [base_amount, base_amount])
        if row:
            return _to_int(row[0]), _to_int(row[1]), _to_int(row[2])
        # 구간이 없으면 0% 처리
        return 0.0, 0.0, 0.0

def _calc_progressive_by_worktax(base_amount: float):
    prgrs_ddct_amt, min_base, taxrat = _get_latest_worktax_row(base_amount)
    if taxrat <= 0:
        return 0.0, 0.0
    # VB: Fix( prgrs_ddct_amt + ((base - min)/100 * taxrat) )
    sanse = float(int(prgrs_ddct_amt + ((base_amount - min_base) / 100.0 * taxrat)))
    return sanse, taxrat

def _tax_piecewise_corp(base_amount: float):
    """법인세 산출세액 구간 (VB 로직 그대로)"""
    if base_amount > 20_000_000_000:
        return base_amount * 0.21 - 420_000_000, 21
    elif base_amount > 200_000_000:
        return base_amount * 0.19 - 20_000_000, 19
    else:
        return base_amount * 0.09, 9

def _query_ar_by_year(seq_no: int, work_yy: int, fiscalMM: int):
    # (기본 회계연도와 상관 없이 연도별 잔액 집계; 필요시 fiscalMM 기준 윈도우로 재구성)
    sql = """
    SELECT 
      MAX(work_yy) AS work_yy,
      (SUM(tranAmt_Cr) - SUM(tranAmt_Dr)) AS balance
    FROM DS_SlipLedgr2 WITH (NOLOCK)
    WHERE seq_no = %s
      AND acnt_cd = 108
      AND ISNULL(cncl_Dt, '') = ''
    GROUP BY work_yy
    ORDER BY work_yy DESC;
    """
    rows = _fetchall(sql, [seq_no])
    return [{"year": int(y or 0), "balance": bal or 0} for (y, bal) in rows]

def _zero_fill_quarters(rows, year_curr, year_prev):
    """
    rows: [{'year': y, 'quarter': q, 'sales': x}, ...]
    → 두 연도 각각 1..4분기를 모두 채워 반환
    """
    m = {(int(r["year"]), int(r["quarter"])): float(r.get("sales") or 0) for r in rows}
    out = []
    for y in (year_prev, year_curr):
        for q in range(1, 5):
            out.append({"year": int(y), "quarter": q, "sales": float(m.get((y, q), 0.0))})
    return out

def _query_quarterly_two_years(seq_no: int, work_yy: int, fiscalMM: int):
    """
    올해(work_yy) vs 전년(work_yy-1) 분기 매출 집계 (401~430, Remk<>'손익계정에 대체')
    - fiscalMM=12: 캘린더 분기(1~3,4~6,7~9,10~12)
    - fiscalMM!=12: 회계 시작월(= fiscalMM+1) 기준 3개월씩 4분기
    반환: [{'year': 2025, 'quarter': 1..4, 'sales': 금액}, ...] (두 연도 모두)
    """
    if fiscalMM == 12:
        sql = """
        SELECT work_yy AS year,
               CEILING(CAST(LEFT(tran_dt,2) AS INT) / 3.0) AS quarter,
               SUM(tranamt_dr - tranamt_cr) AS sales
        FROM ds_slipledgr2
        WHERE seq_no = %s
          AND acnt_cd BETWEEN 401 AND 430
          AND Remk <> N'손익계정에 대체'
          AND work_yy IN (%s, %s)
        GROUP BY work_yy, CEILING(CAST(LEFT(tran_dt,2) AS INT) / 3.0)
        ORDER BY work_yy, CEILING(CAST(LEFT(tran_dt,2) AS INT) / 3.0)
        """
        params = [seq_no, work_yy, work_yy - 1]
    else:
        startM = (fiscalMM % 12) + 1  # 회계 시작월
        # 분기식: 1 + (((월 + 12 - 시작월) % 12) / 3)
        sql = """
        SELECT * FROM (
          -- 당해 회계연도
          SELECT %s AS year,
                 1 + (((CAST(LEFT(a.tran_dt,2) AS INT) + 12 - %s) % 12) / 3) AS quarter,
                 SUM(a.tranamt_dr - a.tranamt_cr) AS sales
          FROM ds_slipledgr2 a
          JOIN ds_FiscalMM_V b ON a.seq_no = b.seq_no
          WHERE a.seq_no = %s
            AND a.acnt_cd BETWEEN 401 AND 430
            AND a.Remk <> N'손익계정에 대체'
            AND (
                 (a.Work_YY = %s            AND a.tran_dt <  b.시작일)
              OR (a.Work_YY = %s + b.기준년 AND a.tran_dt >= b.시작일)
            )
          GROUP BY 1 + (((CAST(LEFT(a.tran_dt,2) AS INT) + 12 - %s) % 12) / 3)

          UNION ALL

          -- 직전 회계연도
          SELECT %s AS year,
                 1 + (((CAST(LEFT(a.tran_dt,2) AS INT) + 12 - %s) % 12) / 3) AS quarter,
                 SUM(a.tranamt_dr - a.tranamt_cr) AS sales
          FROM ds_slipledgr2 a
          JOIN ds_FiscalMM_V b ON a.seq_no = b.seq_no
          WHERE a.seq_no = %s
            AND a.acnt_cd BETWEEN 401 AND 430
            AND a.Remk <> N'손익계정에 대체'
            AND (
                 (a.Work_YY = %s-1            AND a.tran_dt <  b.시작일)
              OR (a.Work_YY = %s-1 + b.기준년 AND a.tran_dt >= b.시작일)
            )
          GROUP BY 1 + (((CAST(LEFT(a.tran_dt,2) AS INT) + 12 - %s) % 12) / 3)
        ) S
        ORDER BY year, quarter
        """
        params = [
            work_yy, startM, seq_no, work_yy, work_yy, startM,
            work_yy - 1, startM, seq_no, work_yy, work_yy, startM
        ]

    rows = _fetchall(sql, params)
    norm = []
    for r in rows:
        if isinstance(r, (list, tuple)):
            y, q, s = r[0], r[1], r[2]
        else:
            y, q, s = r.get("year"), r.get("quarter"), r.get("sales")
        norm.append({"year": int(y), "quarter": int(q), "sales": float(s or 0)})
    return _zero_fill_quarters(norm, work_yy, work_yy - 1)

def _query_monthly_two_years(seq_no: int, work_yy: int | None, fiscalMM: int):
    """
    올해 vs 직전년도 월별 매출(12개월, 회계월 반영)
    반환: [{year: 2025, month: 1..12, sales: 금액}, {year: 2024, ...}]
    """
    if not work_yy:
        raise ValueError("work_yy is required for PL_MONTHLY")

    params = [seq_no, work_yy, seq_no, work_yy - 1]
    if fiscalMM == 12:
        sql = """
        -- 12월 결산: 캘린더 그대로
        SELECT work_yy AS year,
               CAST(LEFT(tran_dt,2) AS INT) AS month,
               SUM(tranamt_dr - tranamt_cr) AS sales
        FROM ds_slipledgr2
        WHERE seq_no = %s
          AND acnt_cd BETWEEN 401 AND 430
          AND Remk <> N'손익계정에 대체'
          AND work_yy IN (%s, %s)
        GROUP BY work_yy, LEFT(tran_dt,2)
        ORDER BY work_yy, CAST(LEFT(tran_dt,2) AS INT)
        """
        params = [seq_no, work_yy, work_yy - 1]
    else:
        # 비 12월 결산: 귀사 로직(시작일/기준년)을 그대로 적용
        sql = """
        SELECT * FROM (
          -- 당해 회계연도 분
          SELECT %s AS year,
                 CAST(LEFT(a.tran_dt,2) AS INT) AS month,
                 SUM(a.tranamt_dr - a.tranamt_cr) AS sales
          FROM ds_slipledgr2 a
          JOIN ds_FiscalMM_V b ON a.seq_no = b.seq_no
          WHERE a.seq_no = %s
            AND a.acnt_cd BETWEEN 401 AND 430
            AND a.Remk <> N'손익계정에 대체'
            AND (
                 (a.Work_YY = %s            AND a.tran_dt <  b.시작일)
              OR (a.Work_YY = %s + b.기준년 AND a.tran_dt >= b.시작일)
            )
          GROUP BY LEFT(a.tran_dt,2)

          UNION ALL

          -- 직전 회계연도 분
          SELECT %s AS year,
                 CAST(LEFT(a.tran_dt,2) AS INT) AS month,
                 SUM(a.tranamt_dr - a.tranamt_cr) AS sales
          FROM ds_slipledgr2 a
          JOIN ds_FiscalMM_V b ON a.seq_no = b.seq_no
          WHERE a.seq_no = %s
            AND a.acnt_cd BETWEEN 401 AND 430
            AND a.Remk <> N'손익계정에 대체'
            AND (
                 (a.Work_YY = %s            AND a.tran_dt <  b.시작일)
              OR (a.Work_YY = %s + b.기준년 AND a.tran_dt >= b.시작일)
            )
          GROUP BY LEFT(a.tran_dt,2)
        ) S
        ORDER BY year, month
        """
        # params: [year, seq_no, work_yy, work_yy, year-1, seq_no, work_yy-1, work_yy-1]
        params = [work_yy, seq_no, work_yy, work_yy, work_yy - 1, seq_no, work_yy - 1, work_yy - 1]

    with connection.cursor() as cur:
        cur.execute(sql, params)
        cols = [c[0] for c in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    # None -> 0, 정수화
    for r in rows:
        r["year"] = int(r["year"])
        r["month"] = int(r["month"])
        r["sales"] = float(r.get("sales") or 0)
    return rows
def _get_real_tax_for_year(biz_type: int, biz_no, ssn, year: int) -> float:
    """
    ASP 원본 로직을 그대로 이식한 '실제 세액' 계산용 함수.
    - 개인(biz_type > 3): elec_income에서 종합소득/지방소득세/농특세 결정세액 합계
    - 법인(biz_type <= 3): tbl_EquityEval에서 법인세 + 지방세 * 0.1
    """
    print(f"[real_tax] enter: biz_type={biz_type}, biz_no={biz_no}, ssn={ssn}, year={year}")

    try:
        # 개인사업자 / 종합소득세
        if biz_type > 3:
            sql = """
                SELECT 종합소득_결정세액, 지방소득세_결정세액, 농특세_결정세액
                FROM elec_income
                WHERE ssn = %s
                  AND work_YY = %s
            """
            params = [ssn, str(year)]
            print(f"[real_tax] 개인 SQL: {sql.strip()} / params={params}")

            rows = _fetchall(sql, params)
            print(f"[real_tax] 개인 rows={rows!r}")

            if rows:
                gi, local, special = rows[0]
                return float(gi or 0) + float(local or 0) + float(special or 0)
            return 0.0

        # 법인 / 법인세
        if biz_type <= 3:
            sql = """
                SELECT 법인세, 지방세
                FROM tbl_EquityEval
                WHERE 사업자번호 = %s
                  AND LEFT(사업연도말, 4) = %s
            """
            params = [biz_no, str(year)]
            print(f"[real_tax] 법인 SQL: {sql.strip()} / params={params}")

            rows = _fetchall(sql, params)
            print(f"[real_tax] 법인 rows={rows!r}")

            if rows:
                corp, local = rows[0]
                return float(corp or 0) + float(local or 0) * 0.1
            return 0.0

        # biz_type이 이상한 경우
        print(f"[real_tax] biz_type not matched: {biz_type}")
        return 0.0

    except Exception as e:
        import traceback
        print(f"[real_tax][ERROR] biz_type={biz_type}, biz_no={biz_no}, ssn={ssn}, year={year}")
        print("[real_tax][ERROR]", e)
        traceback.print_exc()
        return 0.0


def _build_annual_payload(memuser, work_yy, fiscalMM):
    now_year = timezone.localdate().year
    years = [now_year - i for i in range(0, 6, 1)]  # 최근 6년

    # 연도별 컨테이너
    by_year = {
        y: {
            "A00": 0, "A10": 0, "A20": 0,
            "B00": 0, "B10": 0, "B20": 0,
            "C00": 0, "C10": 0,
            "Z108": 0, "Z260": 0, "Z293": 0,
            "E10": 0, "F10": 0, "H10": 0,
            "J10": 0, "L10": 0, "N10": 0,
            "O10": 0, "Q10": 0,
            # ★ 여기부터는 _get_base_pl_aggregates에서 채울 예정
            "BASE_SALES": 0,
            "BASE_COGS": 0,
            "BASE_SALARY": 0,
            "BASE_SGA": 0,
            "BASE_NONOP_EXP": 0,
            "BASE_TAX": 0,
        }
        for y in years
    }

    # 1) BS: 각 연도
    for y in years:
        rows = _fetchall("EXEC up_Act_BSInquiry %s, %s", [str(y), memuser.seq_no])
        for acnt_cd, _, rightAmt, rightAmt_pre, *_ in rows:
            if   acnt_cd == "A00": by_year[y]["A00"] = rightAmt
            elif acnt_cd == "A10": by_year[y]["A10"] = rightAmt
            elif acnt_cd == "A20": by_year[y]["A20"] = rightAmt
            elif acnt_cd == "B00": by_year[y]["B00"] = rightAmt
            elif acnt_cd == "B10": by_year[y]["B10"] = rightAmt
            elif acnt_cd == "B20": by_year[y]["B20"] = rightAmt
            elif acnt_cd == "C10": by_year[y]["C10"] = rightAmt or 1
            elif acnt_cd == "108": by_year[y]["Z108"] = rightAmt
            elif acnt_cd == "260": by_year[y]["Z260"] = rightAmt
            elif acnt_cd == "293": by_year[y]["Z293"] = rightAmt

        by_year[y]["C00"] = (by_year[y]["A00"] or 0) - (by_year[y]["B00"] or 0) or 1

    # 2) PL: 각 연도
    for y in years:
        rows = _fetchall("EXEC up_Act_PLInquiry %s, %s", [str(y), memuser.seq_no])
        for acnt_cd, _, rightAmt, rightAmt_pre, *_ in rows:
            if acnt_cd in ("E10", "F10", "H10", "J10", "L10", "N10"):
                by_year[y][acnt_cd] = rightAmt
            elif acnt_cd == "Q10":
                by_year[y]["Q10"] = rightAmt

    # 3) 세액 O10 집계 (법인/개인) ← 기존 그대로
    # ... (여기 부분은 기존 코드 유지) ...

    # Q10 보정(N10 - O10)
    for y in years:
        if not by_year[y]["Q10"]:
            by_year[y]["Q10"] = (by_year[y]["N10"] or 0) - (by_year[y]["O10"] or 0)

    # ★ 4) _get_base_pl_aggregates로 급여·판관비·영업외비용 분리
    for y in years:
        sales, cogs, salary, sga, nonOpInc, nonOpExp, corpTax_pl  = \
            _get_base_pl_aggregates(memuser.seq_no, y, fiscalMM)
        
        # ASP 원본과 동일한 방식으로 '실제 세액' 계산
        corpTax_real = _get_real_tax_for_year(memuser.biz_type,memuser.biz_no,memuser.ssn, year=y)

        # real이 우선, 없으면 PL 값 사용
        corpTax = corpTax_real if corpTax_real is not None else corpTax_pl


        by_year[y]["BASE_SALES"]     = sales
        by_year[y]["BASE_COGS"]      = cogs
        by_year[y]["BASE_SALARY"]    = salary
        by_year[y]["BASE_SGA"]       = sga
        by_year[y]["BASE_NONOP_EXP"] = nonOpExp
        by_year[y]["BASE_TAX"]       = corpTax

    def _safe_pct(a, b):
        try:
            if not b:
                return None
            return float(a or 0) / float(b) * 100.0
        except Exception:
            return None

    saleCostYears = []
    currentRatioA = {}
    currentRatioR = {}
    totalData = []

    for y in years:
        A00 = by_year[y]["A00"]; A10 = by_year[y]["A10"]; B00 = by_year[y]["B00"]; B10 = by_year[y]["B10"]; C00 = by_year[y]["C00"]
        E10 = by_year[y]["E10"]; N10 = by_year[y]["N10"]

        curRatioA = _safe_pct(A10, B10)                     # 유동비율
        curRatioR = _safe_pct(B00, C00)                     # 부채비율
        curRatio1 = _safe_pct(A10, A00)                     # 유동자산비율
        curRatio2 = _safe_pct(C00, (C00 or 0) + (B00 or 0)) # 자기자본비율
        curRatio3 = _safe_pct((by_year[y]["Z260"] or 0) + (by_year[y]["Z293"] or 0), A00)  # 차입금의존도

        if (y-1) in by_year and by_year[y-1]["C00"]:
            curRatio5 = _safe_pct(C00 - by_year[y-1]["C00"], by_year[y-1]["C00"])  # 자기자본증가율
            curRatio6 = _safe_pct(A00 - by_year[y-1]["A00"], by_year[y-1]["A00"])  # 총자산증가율
        else:
            curRatio5 = None; curRatio6 = None

        totalData.append({
            "year": y,
            "A00": A00, "A10": A10, "A20": by_year[y]["A20"],
            "B00": B00, "B10": B10, "B20": by_year[y]["B20"],
            "C00": C00, "C10": by_year[y]["C10"],
            "Z108": by_year[y]["Z108"], "Z260": by_year[y]["Z260"], "Z293": by_year[y]["Z293"],
            "E10": E10, "F10": by_year[y]["F10"], "H10": by_year[y]["H10"],
            "J10": by_year[y]["J10"], "L10": by_year[y]["L10"], "N10": N10,
            "O10": by_year[y]["O10"], "Q10": by_year[y]["Q10"],

            # ★ 여기서 프런트용으로 바로 내려줌
            "salary": by_year[y]["BASE_SALARY"],
            "sga":    by_year[y]["BASE_SGA"],
            "nonOp":  by_year[y]["BASE_NONOP_EXP"],
            "tax":    by_year[y]["BASE_TAX"],

            "curRatioA": curRatioA, "curRatioR": curRatioR,
            "curRatio1": curRatio1, "curRatio2": curRatio2, "curRatio3": curRatio3,
            "curRatio5": curRatio5, "curRatio6": curRatio6,
        })

        currentRatioA[str(y)] = [
            {"sector": "유동자산", "size": A10},
            {"sector": "유동부채", "size": B10},
        ]
        currentRatioR[str(y)] = [
            {"sector": "부채총액", "size": B00},
            {"sector": "자본총액", "size": C00},
        ]
        saleCostYears.append({
            "year": y,
            "income": E10,
            "expenses": (E10 or 0) - (N10 or 0),
        })

    return {
        "current": 1,
        "totalData": totalData,
        "currentRatioA": currentRatioA,
        "currentRatioR": currentRatioR,
        "saleCostYears": saleCostYears,
    }

def _months_between(d1: datetime.date, d2: datetime.date) -> int:
    return (d2.year - d1.year) * 12 + (d2.month - d1.month)

def _query_ar_top_list(seq_no: int, work_yy: int, fiscalMM: int, as_of: datetime.date):
    """
    ASP 코드의 로직을 그대로 재현:
    1) 거래처별 기말잔액(전기이월+차변-대변) > 0 인 상위 N(기본 8)
    2) 각 거래처에 대해 acnt_cd=108, tran_dt<>'00-00' 내역을 '연도 desc, 월-일 desc' 로 훑으면서
       endAmt(잔액)에서 tranamt_cr를 차감 → 0 이하가 되는 시점의 전표일자를 '발생년월'로 사용
    3) as_of 기준 개월수로 미수기간/부실여부 산정
    """
    # 기간 필터
    if fiscalMM == 12:
        period_where = "a.work_yy = %s"
        period_params = [work_yy]
        period_where_carry = "work_yy = %s"
        period_params_carry = [work_yy]
    else:
        # 회계시작월 = fiscalMM + 1 (MM은 2자리)
        start_mm = (fiscalMM % 12) + 1
        start_mm_str = f"{start_mm:02d}-01"
        period_where = "( (a.work_yy = %s AND a.tran_dt < %s) OR (a.work_yy = %s AND a.tran_dt >= %s) )"
        period_params = [work_yy, start_mm_str, work_yy - 1, start_mm_str]
        period_where_carry = period_where.replace("a.", "")
        period_params_carry = period_params[:]

    # 1) 거래처별 잔액 상위
    sql_top = f"""
    SELECT  AA.Trader_Code,
            MAX(AA.Trader_Name) AS trader_name,
            MAX(AA.Trader_Bizno) AS bizno,
            SUM(AA.carry) AS carry,
            SUM(AA.cr)    AS cr,
            SUM(AA.dr)    AS dr,
            SUM(AA.carry) + SUM(AA.cr) - SUM(AA.dr) AS end_amt
    FROM (
        SELECT  a.Trader_Code, MAX(a.Trader_Name) Trader_Name, MAX(a.Trader_Bizno) Trader_Bizno,
                0 AS carry,
                SUM(a.tranAmt_Cr) AS cr,
                SUM(a.tranAmt_Dr) AS dr
        FROM DS_SlipLedgr2 a WITH (NOLOCK)
        WHERE {period_where} AND a.seq_no=%s AND a.acnt_cd=108 AND a.cncl_Dt='' AND a.tran_dt<>'00-00'
        GROUP BY a.Trader_Code

        UNION ALL
        SELECT  Trader_Code, MAX(Trader_Name), MAX(Trader_Bizno),
                SUM(CASE WHEN Acnt_cd BETWEEN 101 AND 250 THEN tranAmt_Cr - tranAmt_Dr
                         WHEN Acnt_cd BETWEEN 251 AND 330 THEN tranAmt_Dr - tranAmt_Cr
                         ELSE 0 END) AS carry,
                0 AS cr, 0 AS dr
        FROM DS_SlipLedgr2 WITH (NOLOCK)
        WHERE {period_where_carry} AND seq_no=%s AND acnt_cd=108 AND cncl_Dt='' AND tran_dt='00-00' AND Trader_Code<>''
        GROUP BY Trader_Code
    ) AA
    GROUP BY AA.Trader_Code
    HAVING SUM(AA.carry) + SUM(AA.cr) - SUM(AA.dr) > 0
    ORDER BY end_amt DESC
    """
    params_top = period_params + [seq_no] + period_params_carry + [seq_no]
    top_rows = _fetchall(sql_top, params_top)

    result = []
    # 2) 거래처별 발생년월(잔액을 커버하는 전표일자) 탐색
    for trader_code, name, bizno, carry, cr, dr, end_amt in top_rows[:8]:
        end_amt = end_amt or 0

        sql_tx = """
        SELECT work_yy, tran_dt, tranamt_cr
        FROM DS_SlipLedgr2 WITH (NOLOCK)
        WHERE seq_no=%s AND acnt_cd=108 AND tran_dt<>'00-00' AND Trader_Code=%s
        ORDER BY work_yy DESC, tran_dt DESC
        """
        tx = _fetchall(sql_tx, [seq_no, trader_code])

        occur_date = None
        remain = int(end_amt)
        for yy, mmdd, cr_amt in tx:
            cr_amt = int(cr_amt or 0)
            remain -= cr_amt
            if remain <= 0:
                # 'YYYY' + '-' + 'MM-DD' → 날짜
                try:
                    occur_date = datetime.date(int(yy), int(str(mmdd)[:2]), int(str(mmdd)[-2:]))
                except Exception:
                    # 포맷이 다르면 안전하게 1일로
                    try:
                        occur_date = datetime.date(int(yy), int(str(mmdd)[:2]), 1)
                    except Exception:
                        occur_date = as_of
                break

        months = _months_between(occur_date or as_of, as_of)
        if months >= 60:
            grade = "악성채권"
        elif months >= 36:
            grade = "대손가능채권"
        elif months >= 24:
            grade = "부실채권"
        else:
            grade = ""

        result.append({
            "trader_code": trader_code,
            "trader_name": name,
            "bizno": bizno,
            "end_amount": int(end_amt),
            "occur_ym": occur_date.strftime("%Y-%m") if occur_date else "",
            "months": months,
            "grade": grade,
        })
    return result

def _query_sales_top(seq_no: int, work_yy: int, fiscalMM: int, biz_type: int):
    """
    DS_SlipLedgr2에서 당기(회계연도) 주요 매출처 TOP 리스트를 반환
    return: [ {거래처명, 사업자번호, 차변, 비율}, ... ]
    """
    # 회계연도 경계: 결산월이 12가 아니면 (fiscalMM+1)-01 이 회계시작월
    start_mm = ((int(fiscalMM) % 12) + 1)
    mmdd = f"{start_mm:02d}-01"  # 'MM-01' 형식

    # 기간절 조건/파라미터
    if int(fiscalMM) == 12:
        period_cond = "work_yy = %s"
        period_params = [str(work_yy)]
    else:
        # (해당연도 시작월 전 < …) OR (전년도 시작월 이상 …)
        period_cond = "( (Work_YY = %s AND tran_dt < %s) OR (Work_YY = %s AND tran_dt >= %s) )"
        period_params = [str(work_yy), mmdd, str(work_yy - 1), mmdd]

    base_cond = (
        f"{period_cond} AND seq_no = %s "
        "AND (acnt_cd >= 401 AND acnt_cd <= 430) "
        "AND cncl_Dt = '' AND tran_dt <> '00-00'"
    )

    # 분모(해당 기간 총 매출) 서브쿼리
    denom_sql = f"SELECT CASE WHEN SUM(tranAmt_Dr)=0 THEN 1 ELSE SUM(tranAmt_Dr) END FROM DS_SlipLedgr2 WHERE {base_cond}"
    denom_params = period_params + [seq_no]

    # 개인/법인 구분에 따른 GROUP BY
    group_col = "Trader_Code" if (int(biz_type or 0) <= 3) else "Trader_Name"

    # 메인 쿼리 (TOP 10)
    sql = f"""
    SELECT TOP 10
        MAX(Trader_Name)  AS Trader_Name,
        MAX(Trader_Bizno) AS Trader_Bizno,
        SUM(tranAmt_Dr)   AS amount,
        ROUND( (SUM(tranAmt_Dr) / ({denom_sql})) * 100.0, 2) AS ratio
    FROM DS_SlipLedgr2
    WHERE {base_cond}
    GROUP BY {group_col}
    ORDER BY amount DESC
    """

    # 플레이스홀더 순서: (메인 WHERE) + (denom 서브쿼리) + (메인 WHERE 재사용 아님)
    params = period_params + [seq_no] + denom_params
    rows = _fetchall(sql, params)

    # 튜플 → dict
    out = []
    for name, bizno, amt, ratio in rows:
        out.append({
            "거래처명":   (name or "").strip(),
            "사업자번호": (bizno or "").strip(),
            "차변":       int(amt or 0),
            "비율":       float(ratio or 0.0),
        })
    return out

def _query_purchase_top(seq_no: int, work_yy: int, fiscalMM: int, biz_type: int):
    seq_no = int(seq_no)
    work_yy = int(work_yy)
    fiscalMM = int(fiscalMM)

    rows = []
    with connection.cursor() as cur:
        if biz_type <= 3:
            # ── 개인/간편
            if fiscalMM == 12:
                # 당해연도 전체
                sql = """
                    SELECT TOP 10
                           LEFT(MAX(Trader_Name), 12)         AS 거래처명,
                           LEFT(MAX(Trader_Bizno),12)         AS 사업자번호,
                           SUM(tranAmt_Dr)          AS 차변,      -- 원
                           ROUND( (SUM(tranAmt_Dr) / NULLIF((
                                    SELECT SUM(tranAmt_Dr)
                                      FROM DS_SlipLedgr2
                                     WHERE work_yy = %s
                                       AND seq_no  = %s
                                       AND acnt_cd IN (101,251,253)
                                       AND cncl_Dt = ''
                                       AND tran_dt <> '00-00'
                                   ), 0)) * 100, 0)            AS 비율
                      FROM DS_SlipLedgr2
                     WHERE work_yy = %s
                       AND seq_no  = %s
                       AND acnt_cd IN (101,251,253)
                       AND cncl_Dt = ''
                       AND tran_dt <> '00-00'
                     GROUP BY Trader_Code
                     ORDER BY 차변 DESC
                """
                params = [work_yy, seq_no, work_yy, seq_no]
            else:
                # 전기 4/1 ~ 당기 3/31
                prev_yy = work_yy - 1
                sql = """
                    SELECT TOP 10
                           LEFT(MAX(Trader_Name), 12)         AS 거래처명,
                           LEFT(MAX(Trader_Bizno),12)         AS 사업자번호,
                           SUM(tranAmt_Dr)          AS 차변,      -- 원
                           ROUND( (SUM(tranAmt_Dr) / NULLIF((
                                    SELECT SUM(tranAmt_Dr)
                                      FROM DS_SlipLedgr2
                                     WHERE ( (Work_YY = %s AND tran_dt < '04-01')
                                          OR (Work_YY = %s AND tran_dt >= '04-01') )
                                       AND seq_no  = %s
                                       AND acnt_cd IN (101,251,253)
                                       AND cncl_Dt = ''
                                       AND tran_dt <> '00-00'
                                   ), 0)) * 100, 0)            AS 비율
                      FROM DS_SlipLedgr2
                     WHERE ( (Work_YY = %s AND tran_dt < '04-01')
                          OR (Work_YY = %s AND tran_dt >= '04-01') )
                       AND seq_no  = %s
                       AND acnt_cd IN (101,251,253)
                       AND cncl_Dt = ''
                       AND tran_dt <> '00-00'
                     GROUP BY Trader_Code
                     ORDER BY 차변 DESC
                """
                params = [work_yy, prev_yy, seq_no, work_yy, prev_yy, seq_no]

            cur.execute(sql, params)
            cols = ["거래처명", "사업자번호", "차변", "비율"]
            rows = _rows(cur, cols)

        else:
            # ── 법인/복잡
            sql = """
                SELECT TOP 10
                       RANK() OVER (ORDER BY TMP.차변 DESC) AS 금액순,
                       TMP.*
                  FROM (
                        SELECT
                               LEFT(MAX(Trader_Name), 12)            AS 거래처명,
                               LEFT(MAX(Trader_Bizno),12)            AS 사업자번호,
                               CAST(COUNT(*) AS VARCHAR(10)) + N'회' AS 거래횟수,
                               SUM(tranAmt_Dr)   AS 차변,     -- 원
                               CEILING(
                                 ROUND(
                                   ( SUM(tranAmt_Dr) / NULLIF((
                                       SELECT SUM(tranAmt_Dr)
                                         FROM DS_SlipLedgr2
                                        WHERE Work_YY = %s
                                          AND Seq_no  = %s
                                          AND Acnt_cd IN (101,103,251,253)
                                          AND Cncl_Dt = ''
                                          AND Tran_dt <> '00-00'
                                   ), 0)) * 100, 0)
                               )                                     AS 비율
                          FROM DS_SlipLedgr2
                         WHERE Work_YY = %s
                           AND Seq_no  = %s
                           AND Acnt_cd IN (101,103,251,253)
                           AND Cncl_Dt = ''
                           AND Tran_dt <> '00-00'
                         GROUP BY Trader_Name
                        HAVING LEFT(MAX(Trader_Name),8) <> ''
                  ) TMP
            """
            params = [work_yy, seq_no, work_yy, seq_no]
            cur.execute(sql, params)
            cols = ["금액순", "거래처명", "사업자번호", "거래횟수", "차변", "비율"]
            rows = _rows(cur, cols)

    for r in rows:
        if "차변" in r:
            r["차변"] = _to_int(r["차변"], 2)
        if "비율" in r:
            r["비율"] = _to_int(r["비율"], 0)

    return rows

def _query_purchase_evidence(seq_no:int, work_yy:int, fiscalMM:int, biz_type:int, biz_no:str, ssn:str):
    """
    매입내역 증빙분석 (최근 work_yy부터 최대 5개 연도)
    반환: [{year, 검증대상금액, 세금계산서, 계산서, 신용카드등, 원천징수, 증빙불비}]
    * 모든 금액은 '원' 단위
    """
    out_rows = []

    # 조회 연도: work_yy, work_yy-1, ... 최대 5개
    years = [work_yy - i for i in range(5)]

    with connection.cursor() as cur:
        for yy in years:
            # ---------------------------
            # 1) DS_SlipLedgr2 기본 합계
            # ---------------------------
            if fiscalMM == 12:
                # 해당 회계연도(yy)만 필터
                base_sql = """
                    SELECT
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 401 AND 430 THEN tranAmt_dr - tranAmt_cr ELSE 0 END),0) AS 매출액,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 451 AND 470 THEN tranAmt_cr ELSE 0 END),0)           AS 매출원가,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 146 AND 149 THEN tranAmt_cr ELSE 0 END),0)           AS 상품당기매입,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 501 AND 800 AND Remk NOT LIKE '%%원가로 대체%%'
                                      THEN tranAmt_cr - tranAmt_dr ELSE 0 END),0)                               AS 제조당기매입,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 801 AND 810  THEN tranAmt_cr - tranAmt_dr ELSE 0 END),0) AS 급여,
                      ISNULL(SUM(CASE WHEN acnt_cd = 186 THEN tranAmt_cr - tranAmt_dr ELSE 0 END),0)             AS 퇴직연금운용자산,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 811 AND 900  THEN tranAmt_cr - tranAmt_dr ELSE 0 END),0) AS 기타판관비,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 901 AND 950 THEN tranAmt_dr - tranAmt_cr ELSE 0 END),0) AS 영업외수익,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 951 AND 997 THEN tranAmt_cr - tranAmt_dr ELSE 0 END),0) AS 영업외비용,
                      ISNULL(SUM(CASE WHEN acnt_cd IN (518,618,668,718,768,818) THEN tranAmt_cr ELSE 0 END),0)      AS 감가상각비,
                      ISNULL(SUM(CASE WHEN acnt_cd = 253 AND LEFT(Trader_Code,1)='9' THEN tranAmt_dr ELSE 0 END),0) AS 신용카드
                    FROM DS_SlipLedgr2
                    WHERE seq_no = %s
                      AND work_yy = %s
                      AND ((acnt_cd BETWEEN 401 AND 999) OR (acnt_cd BETWEEN 146 AND 253))
                      AND acnt_cd <> 150
                      AND Remk <> N'손익계정에 대체'
                      AND tran_dt <> '00-00'
                """
                base_params = [seq_no, yy]
            else:
                # 전기 (fiscalMM+1 ~ 12월) + 당기 (1월 ~ fiscalMM)
                period_clause, period_params = _period_clause(yy, fiscalMM)
                base_sql = f"""
                    SELECT
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 401 AND 430 THEN tranAmt_dr - tranAmt_cr ELSE 0 END),0) AS 매출액,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 451 AND 470 THEN tranAmt_cr ELSE 0 END),0)           AS 매출원가,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 146 AND 149 THEN tranAmt_cr ELSE 0 END),0)           AS 상품당기매입,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 501 AND 800 AND Remk NOT LIKE '%%원가로 대체%%'
                                      THEN tranAmt_cr - tranAmt_dr ELSE 0 END),0)                               AS 제조당기매입,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 801 AND 810  THEN tranAmt_cr - tranAmt_dr ELSE 0 END),0) AS 급여,
                      ISNULL(SUM(CASE WHEN acnt_cd = 186 THEN tranAmt_cr - tranAmt_dr ELSE 0 END),0)             AS 퇴직연금운용자산,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 811 AND 900  THEN tranAmt_cr - tranAmt_dr ELSE 0 END),0) AS 기타판관비,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 901 AND 950 THEN tranAmt_dr - tranAmt_cr ELSE 0 END),0) AS 영업외수익,
                      ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 951 AND 997 THEN tranAmt_cr - tranAmt_dr ELSE 0 END),0) AS 영업외비용,
                      ISNULL(SUM(CASE WHEN acnt_cd IN (518,618,668,718,768,818) THEN tranAmt_cr ELSE 0 END),0)      AS 감가상각비,
                      ISNULL(SUM(CASE WHEN acnt_cd = 253 AND LEFT(Trader_Code,1)='9' THEN tranAmt_dr ELSE 0 END),0) AS 신용카드
                    FROM DS_SlipLedgr2
                    WHERE seq_no = %s
                      AND {period_clause}
                      AND ((acnt_cd BETWEEN 401 AND 999) OR (acnt_cd BETWEEN 146 AND 253))
                      AND acnt_cd <> 150
                      AND Remk <> N'손익계정에 대체'
                      AND tran_dt <> '00-00'
                """
                base_params = [seq_no] + period_params

            cur.execute(base_sql, base_params)
            base = cur.fetchone() or [0]*11
            (saleAmt, cogs, goodsPurchase, manufPurchase, salary, pensionAsset,
             sga, nonOpInc, nonOpExp, depr, cardFromLedger) = [float(x or 0) for x in base]

            # ---------------------------
            # 2) 기초 원재료(153, tran_dt='00-00') - 당기제품제조원가 차감 항목
            # ---------------------------
            kicho_sql = """
                SELECT ISNULL(SUM(tranamt_cr),0)
                  FROM ds_slipledgr2
                 WHERE acnt_cd = 153
                   AND tran_dt = '00-00'
                   AND work_yy = %s
                   AND seq_no  = %s
            """
            amt153kicho = _fetchone_scalar(cur, kicho_sql, [yy, seq_no], 0)

            # ---------------------------
            # 3) 유형자산 증가액(당기)  ※ VB: 전기 비교도 구하지만 현재식엔 증가액만 반영
            # ---------------------------
            tangible_sql = """
                SELECT ISNULL(SUM(tranamt_cr - tranamt_dr),0)
                  FROM ds_slipledgr2
                 WHERE seq_no = %s
                   AND acnt_cd IN (197,199,202,204,206,208,210,212,214,215,217,225,226,227,228,229)
                   AND tran_dt <> '00-00'
                   AND work_yy = %s
            """
            tangibleAsset = _fetchone_scalar(cur, tangible_sql, [seq_no, yy], 0)

            # ---------------------------
            # 4) 부가세 전자신고(매입) - 세금계산서/기타공제/면세
            # ---------------------------
            vat_where = ""
            vat_params = [biz_no]
            if fiscalMM == 12:
                vat_where = " AND LEFT(과세기간,4) = %s "
                vat_params += [str(yy)]
            elif fiscalMM == 6:
                vat_where = " AND LEFT(과세기간,4) <= %s AND LEFT(과세기간,4) >= %s AND NOT (과세기간 = %s) "
                vat_params += [str(yy), str(yy-1), f"{yy-1}년 1기"]
            elif fiscalMM == 3:
                vat_where = " AND LEFT(과세기간,4) <= %s AND LEFT(과세기간,4) >= %s AND NOT (과세기간 = %s AND 과세유형='C17') AND NOT (과세기간 = %s) "
                vat_params += [str(yy), str(yy-1), f"{yy-1}년 1기", f"{yy}년 2기"]

            vat_sql = f"""
                SELECT
                  ISNULL(SUM(매입세금계산서수취일반금액 + 매입세금계산서수취고정자산금액 + 예정누락매입신고세금계산서금액 + 매입자발행세금계산서매입금액),0) AS taxInvoice,
                  ISNULL(SUM(그밖의공제매입명세합계금액),0) AS otherDeductible,
                  ISNULL(SUM(계산서수취금액),0) AS exemptInvoice
                  FROM 부가가치세전자신고3
                 WHERE 사업자등록번호 = %s
                 {vat_where}
            """
            cur.execute(vat_sql, vat_params)
            row = cur.fetchone() or [0,0,0]
            taxInvoice, otherDeductible, exemptInvoice = [float(x or 0) for x in row]

            # ---------------------------
            # 5) 신용카드 등 (biz_type<=10 이면 원장 253/Trader_Code='9*', 아니면 부가세 기타매입)
            #    0이면 기타매입으로 대체
            # ---------------------------
            if biz_type <= 10:
                if fiscalMM == 12:
                    card_sql = """
                        SELECT ISNULL(SUM(tranAmt_dr),0)
                          FROM DS_SlipLedgr2
                         WHERE seq_no = %s
                           AND work_yy = %s
                           AND acnt_cd = 253
                           AND LEFT(Trader_Code,1) = '9'
                           AND tran_dt <> '00-00'
                    """
                    card_val = _fetchone_scalar(cur, card_sql, [seq_no, yy], 0)
                else:
                    period_clause, period_params = _period_clause(yy, fiscalMM)
                    card_sql = f"""
                        SELECT ISNULL(SUM(tranAmt_dr),0)
                          FROM DS_SlipLedgr2
                         WHERE seq_no = %s
                           AND {period_clause}
                           AND acnt_cd = 253
                           AND LEFT(Trader_Code,1) = '9'
                           AND tran_dt <> '00-00'
                    """
                    card_val = _fetchone_scalar(cur, card_sql, [seq_no] + period_params, 0)
            else:
                card_val = otherDeductible

            if card_val == 0:
                card_val = otherDeductible  # VB 폴백

            # ---------------------------
            # 6) 원천징수 + 급여총액
            # ---------------------------
            empTot = 0.0
            withHold = 0.0

            if fiscalMM == 12:
                # 고용총액
                empTot_sql = """
                    SELECT ISNULL(SUM(EmpTot),0)
                      FROM Tbl_EmployTotSalary
                     WHERE Seq_no = %s AND Work_yy = %s
                """
                empTot = _fetchone_scalar(cur, empTot_sql, [seq_no, yy], 0)

                if empTot != 0:
                    with_sql = """
                        SELECT ISNULL(SUM(A20)+SUM(A30)+SUM(A40)+SUM(A50),0)
                          FROM 원천세전자신고
                         WHERE 사업자등록번호 = %s
                           AND LEFT(과세연월,4) = %s
                    """
                    withHold = _fetchone_scalar(cur, with_sql, [biz_no, str(yy)], 0)
                else:
                    with_sql = """
                        SELECT ISNULL(SUM(A01)+SUM(A03)+SUM(A20)+SUM(A30)+SUM(A40)+SUM(A50)+SUM(A60),0)
                          FROM 원천세전자신고
                         WHERE 사업자등록번호 = %s
                           AND LEFT(과세연월,4) = %s
                    """
                    withHold = _fetchone_scalar(cur, with_sql, [biz_no, str(yy)], 0)
            else:
                # 예: yy-1 (fiscalMM+1) ~ yy (fiscalMM)
                startYm = f"{yy-1}{(fiscalMM+1):02d}"
                endYm   = f"{yy}{fiscalMM:02d}"
                with_sql = """
                    SELECT ISNULL(SUM(A99),0)
                      FROM 원천세전자신고
                     WHERE 사업자등록번호 = %s
                       AND 과세연월 >= %s
                       AND 과세연월 <= %s
                """
                withHold = _fetchone_scalar(cur, with_sql, [biz_no, startYm, endYm], 0)

            # ---------------------------
            # 7) 검증대상금액/증빙불비
            #    검증대상금액 = 유형자산 증가액 + 당기매입원가(상품+제조-153기초+급여+기타판관비 - 감가상각비 - 퇴직연금운용자산)
            # ---------------------------
            current_cost = (goodsPurchase + manufPurchase - amt153kicho
                            + salary + sga + tangibleAsset - depr - pensionAsset)
            target_amt = max(current_cost, 0.0)

            # 증빙 합: 세금계산서 + 계산서(면세) + 신용카드 등 + 원천징수(=withHold+empTot)
            evidence_sum = (taxInvoice + exemptInvoice + card_val + (withHold + empTot))
            lack_amt = target_amt - evidence_sum
            if lack_amt < 0:
                lack_amt = 0.0

            out_rows.append({
                "year": yy,
                "검증대상금액": _to_int(target_amt, 0),
                "세금계산서": _to_int(taxInvoice, 0),
                "계산서": _to_int(exemptInvoice, 0),
                "신용카드등": _to_int(card_val, 0),
                "원천징수": _to_int(withHold + empTot, 0),
                "증빙불비": _to_int(lack_amt, 0),
                # 참고용(디버그): 필요 없으면 주석
                # "기초원재료153": _to_float(amt153kicho, 0),
                # "유형자산증가": _to_float(tangibleAsset, 0),
                # "급여": _to_float(salary, 0),
                # "판관비": _to_float(sga, 0),
                # "감가상각비": _to_float(depr, 0),
                # "퇴직연금운용자산": _to_float(pensionAsset, 0),
            })

    # 최신연도 우선 정렬 보장
    out_rows.sort(key=lambda r: r["year"], reverse=True)
    return out_rows

def _query_tax_estimate(
    seq_no: int,
    work_yy: int,
    work_qt: int,
    fiscalMM: int,
    biz_type: int,
    biz_no: str,
    ssn: str,
    reg_date: date | None,
):
    """
    예상세액 산출 (법인세/종합소득세)
    의존함수: _get_base_pl_aggregates, _tax_piecewise_corp, _calc_progressive_by_worktax
    - None 안전 처리 포함
    """

    # ---- 로컬 유틸(이 함수 안에서만 사용) -----------------------------------------
    def _nz(x, default=0.0):
        """None/문자 → float, 실패 시 default"""
        if x is None:
            return float(default)
        try:
            return float(x)
        except Exception:
            return float(default)

    def _safe_sub(a, b):
        return _nz(a) - _nz(b)

    def _safe_add(*args):
        return sum(_nz(x) for x in args)

    # ---- 1) PL 집계 가져오기 ------------------------------------------------------
    # sales, cogs, salary, sga, nonOpInc, nonOpExp, corpTaxEtc
    sales, cogs, salary, sga, nonOpInc, nonOpExp, corpTaxEtc = _get_base_pl_aggregates(seq_no, work_yy, fiscalMM)
    # 세차감전이익 (pre_tax_profit)
    # = 매출 - 매출원가 - 급여 - 판관비 + 영업외수익 - 영업외비용
    pre_tax_profit = _safe_sub(_safe_sub(_safe_sub(_nz(sales), _nz(cogs)), _nz(salary)), _nz(sga))
    pre_tax_profit = _safe_add(pre_tax_profit, _nz(nonOpInc))
    pre_tax_profit = _safe_sub(pre_tax_profit, _nz(nonOpExp))

    # ---- 2) 등록일·분기 보정 값 ---------------------------------------------------
    if not reg_date:
        # 등록일이 없으면 해당연도 1/1로 가정
        reg_date = datetime.date(int(work_yy or datetime.date.today().year), 1, 1)
    startMM = int(getattr(reg_date, "month", 1) or 1)

    work_qt = int(work_qt or 4)

    # ---- 3) 분기: 법인/개인 -------------------------------------------------------
    values = {}
    tax_name = "법인세" if int(biz_type or 0) <= 3 else "종합소득세"
    kind     = "CORP"    if int(biz_type or 0) <= 3 else "PERSONAL"

    with connection.cursor() as cur:
        if kind == "CORP":
            # ── 전년도 법인 데이터
            cur.execute(
                """
                SELECT 각사업연도소득, 결손금누계, 최저한세적용대상, 최저한세적용제외
                  FROM tbl_EquityEval
                 WHERE 사업자번호=%s
                   AND LEFT(사업연도말,4)=%s
                """,
                [biz_no, str(int(_nz(work_yy)) - 1)],
            )
            row = cur.fetchone()
            kacksa         = _nz(row[0]) if row else 0.0
            valDefict      = _nz(row[1]) if row else 0.0
            targetLimit    = _nz(row[2]) if row else 0.0
            nontargetLimit = _nz(row[3]) if row else 0.0

            # 최저한세 대비 공제율(전년 과세표준으로 산출세액 유사치 산정)
            sanchul_prev, _rate_prev = _tax_piecewise_corp(_nz(kacksa))
            gongjeRate = (targetLimit / sanchul_prev) if sanchul_prev else 0.0
            if not targetLimit and nontargetLimit and sanchul_prev:
                gongjeRate = nontargetLimit / sanchul_prev

            # 세무조정(퇴직연금운용자산: acnt_cd=186)
            cur.execute(
                """
                SELECT ISNULL(SUM(CASE WHEN acnt_cd=186 THEN tranAmt_cr - tranAmt_dr ELSE 0 END),0)
                  FROM DS_SlipLedgr2
                 WHERE seq_no=%s AND work_yy=%s AND tran_dt<>'00-00'
                """,
                [seq_no, work_yy],
            )
            row = cur.fetchone()
            val_SEMUJJ = _nz(row[0]) if row else 0.0

            # 세차감전이익
            valKackRev = _nz(pre_tax_profit)
            PREvalKackRev = (valKackRev * 4.0 / work_qt) if (work_qt and work_qt != 4) else valKackRev

            # 과세표준
            valKwase    = max(_safe_sub(_safe_sub(valKackRev, valDefict), val_SEMUJJ), 0.0)
            PREvalKwase = max(_safe_sub(PREvalKackRev, valDefict), 0.0)

            # 등록년도 보정
            tmpValKwase = valKwase
            if str(int(_nz(work_yy))) == str(reg_date.year) and tmpValKwase > 0:
                tmpValKwase = valKwase * 12.0 / max(1, (12 - startMM + 1))

            # 산출세액(법인 누진)
            valSanse, valTaxRate = _tax_piecewise_corp(_nz(tmpValKwase))
            PREvalSanse, _       = _tax_piecewise_corp(_nz(PREvalKwase))

            # 등록년도 산출세액 비례보정
            if str(int(_nz(work_yy))) == str(reg_date.year) and tmpValKwase > 0:
                valSanse = valSanse * max(1, (12 - startMM + 1)) / 12.0

            # 감면/공제
            valGongje   = _nz(valSanse) * _nz(gongjeRate)
            valBubTax   = max(_safe_sub(valSanse, valGongje), 0.0)
            valRegalTax = _nz(valBubTax) * 0.1               # 지방세(10%)
            valTotalTax = _safe_add(valBubTax, valRegalTax)

            # 분기 추정치
            PREvalGongje   = _nz(PREvalSanse) * _nz(gongjeRate)
            PREvalBubTax   = max(_safe_sub(PREvalSanse, PREvalGongje), 0.0)
            PREvalRegalTax = _nz(PREvalBubTax) * 0.1
            PREvalTotalTax = _safe_add(PREvalBubTax, PREvalRegalTax)

            values = dict(
                valKackRev      = round(_nz(valKackRev)),
                val_SEMUJJ      = round(_nz(val_SEMUJJ)),
                valDefict       = round(_nz(valDefict)),
                valKwase        = round(_nz(valKwase)),
                valTaxRate      = _nz(valTaxRate),
                valSanse        = round(_nz(valSanse)),
                valGongje       = round(_nz(valGongje)),
                valBubTax       = round(_nz(valBubTax)),
                valRegalTax     = round(_nz(valRegalTax)),
                valTotalTax     = round(_nz(valTotalTax)),
                PREvalKackRev   = round(_nz(PREvalKackRev)),
                PREvalTotalTax  = round(_nz(PREvalTotalTax)),
            )

        else:
            # ── 전년도 개인(종소) 데이터
            cur.execute(
                """
                SELECT 소득공제, 종합소득_산출세액, 종합소득_세액감면, 종합소득_세액공제
                  FROM elec_income
                 WHERE ssn=%s AND work_yy=%s
                """,
                [ssn, str(int(_nz(work_yy)) - 1)],
            )
            row = cur.fetchone()
            if row:
                valDefict    = _nz(row[0])   # 소득공제
                sanchul_prev = _nz(row[1])   # 전년도 산출세액
                tax_kammyun  = _nz(row[2])   # 세액감면
                tax_gongje   = _nz(row[3])   # 세액공제
                gongjeRate   = (tax_kammyun / sanchul_prev) if sanchul_prev else 0.0
            else:
                valDefict    = 1_500_000.0
                tax_gongje   = 70_000.0
                gongjeRate   = 0.0

            # 세무조정(퇴직연금운용자산: acnt_cd=186)
            cur.execute(
                """
                SELECT ISNULL(SUM(CASE WHEN acnt_cd=186 THEN tranAmt_cr - tranAmt_dr ELSE 0 END),0)
                  FROM DS_SlipLedgr2
                 WHERE seq_no=%s AND work_yy=%s AND tran_dt<>'00-00'
                """,
                [seq_no, work_yy],
            )
            row = cur.fetchone()
            val_SEMUJJ = _nz(row[0]) if row else 0.0

            # 세차감전이익 → 과세표준
            valKackRev = _nz(pre_tax_profit)
            valKwase   = max(_safe_sub(_safe_sub(valKackRev, valDefict), val_SEMUJJ), 0.0)

            # 산출세액(근거: WorkTax 최신 연도)
            valSanse, valTaxRate = _calc_progressive_by_worktax(_nz(valKwase))

            # 분기 추정
            PREvalKackRev = (valKackRev * 4.0 / work_qt) if (work_qt and work_qt != 4) else valKackRev
            PREvalKwase   = max(_safe_sub(PREvalKackRev, valDefict), 0.0)
            PREvalSanse, PREvalTaxRate = _calc_progressive_by_worktax(_nz(PREvalKwase))

            # 세액감면/공제 (한도: 산출세액)
            valGongje    = min(_nz(valSanse), _nz(valSanse) * _nz(gongjeRate) + _nz(tax_gongje))
            PREvalGongje = min(_nz(PREvalSanse), _nz(PREvalSanse) * _nz(gongjeRate) + _nz(tax_gongje))

            valBubTax    = max(_safe_sub(valSanse, valGongje), 0.0)
            valRegalTax  = _nz(valBubTax) * 0.1
            valTotalTax  = _safe_add(valBubTax, valRegalTax)

            PREvalBubTax   = max(_safe_sub(PREvalSanse, PREvalGongje), 0.0)
            PREvalRegalTax = _nz(PREvalBubTax) * 0.1
            PREvalTotalTax = _safe_add(PREvalBubTax, PREvalRegalTax)

            values = dict(
                valKackRev      = round(_nz(valKackRev)),
                val_SEMUJJ      = round(_nz(val_SEMUJJ)),
                valDefict       = round(_nz(valDefict)),
                valKwase        = round(_nz(valKwase)),
                valTaxRate      = _nz(valTaxRate),
                valSanse        = round(_nz(valSanse)),
                valGongje       = round(_nz(valGongje)),
                valBubTax       = round(_nz(valBubTax)),
                valRegalTax     = round(_nz(valRegalTax)),
                valTotalTax     = round(_nz(valTotalTax)),
                PREvalKackRev   = round(_nz(PREvalKackRev)),
                PREvalTotalTax  = round(_nz(PREvalTotalTax)),
            )

    return dict(kind=kind, tax_name=tax_name, values=values)

def _to_float_safe(v):
    try:
        f = float(v or 0)
        from math import isfinite
        return f if isfinite(f) else 0.0
    except Exception:
        return 0.0

def _safe_div(a, b):
    a = _to_float_safe(a)
    b = _to_float_safe(b)
    return (a / b) if b not in (0, None) else 0.0

def _fetch_pl(year: int, seq_no: int):
    """
    up_Act_PLInquiry 결과를 {계정코드: (당기잔액1, 전기잔액1)} 딕셔너리로 반환
    """
    sql = "EXEC up_Act_PLInquiry %s, %s"
    rows = {}
    with connection.cursor() as cur:
        cur.execute(sql, [str(year), str(seq_no)])
        # 예상 컬럼: 계정코드, 당기잔액1, 전기잔액1
        cols = [c[0] for c in cur.description]
        idx_cd = cols.index("계정코드")
        idx_now = cols.index("당기잔액1")
        idx_prev = cols.index("전기잔액1")
        for r in cur.fetchall():
            acnt = str(r[idx_cd]).strip()
            rows[acnt] = (_to_float_safe(r[idx_now]), _to_float_safe(r[idx_prev]))
    return rows

def _fetch_bs(year: int, seq_no: int):
    """
    up_Act_BSInquiry 결과를 {계정코드: (당기잔액1, 전기잔액1)} 딕셔너리로 반환
    """
    sql = "EXEC up_Act_BSInquiry %s, %s"
    rows = {}
    with connection.cursor() as cur:
        cur.execute(sql, [str(year), str(seq_no)])
        # 예상 컬럼: 계정코드, 당기잔액1, 전기잔액1
        cols = [c[0] for c in cur.description]
        idx_cd = cols.index("계정코드")
        idx_now = cols.index("당기잔액1")
        idx_prev = cols.index("전기잔액1")
        for r in cur.fetchall():
            acnt = str(r[idx_cd]).strip()
            rows[acnt] = (_to_float_safe(r[idx_now]), _to_float_safe(r[idx_prev]))
    return rows

from typing import Dict, Tuple, Any, List, Callable, Optional

Number = float  # 필요 시 Decimal로 교체 가능

def _compute_issue_table(
    seq_no: int,
    work_yy: int,
    *,
    # (선택) 외부에서 직접 주입(테스트/캐시용)
    pl_data: Optional[Dict[int, Dict[str, Tuple[Number, Number]]]] = None,
    bs_data: Optional[Dict[int, Dict[str, Tuple[Number, Number]]]] = None,
    # (선택) 외부 fetcher 주입(DB/ORM 조회용)
    pl_fetcher: Optional[Callable[[int, int], Dict[str, Tuple[Number, Number]]]] = None,
    bs_fetcher: Optional[Callable[[int, int], Dict[str, Tuple[Number, Number]]]] = None,
    # 디버그 출력
    debug: bool = False,
) -> Dict[str, Any]:
    """
    VBScript 로직을 1:1에 가깝게 이식한 '재무이슈표' 생성기.
    years: [work_yy-2, work_yy-1, work_yy]
    내부 계산 인덱스: 0=Y, 1=Y-1, 2=Y-2, 3=Y-3
    반환 key_values 는 years 와 정확히 매칭되도록 [2,1,0] 순서.
    """

    # -----------------------
    # 내부 유틸
    # -----------------------
    def _safe_div(a: Number, b: Number) -> Number:
        try:
            return (a or 0.0) / (b or 0.0) if b not in (0, 0.0, None) else 0.0
        except Exception:
            return 0.0

    def dprint(*args, **kwargs):
        if debug:
            print("[ISSUE_TABLE]", *args, **kwargs)

    # 공통: 안전 fetch (예외 던지지 않음)
    def _get_pl(year: int) -> Dict[str, Tuple[Number, Number]]:
        # 1) 직접 주입 데이터
        if pl_data and year in pl_data:
            return pl_data[year]
        # 2) 주입 fetcher
        if pl_fetcher:
            try:
                return pl_fetcher(year, seq_no) or {}
            except Exception as e:
                dprint(f"[WARN] pl_fetcher 실패: year={year}, seq={seq_no}, err={e}")
        # 3) 동일 모듈 내 함수 자동 탐색(fetch_pl/_fetch_pl)
        for name in ("fetch_pl", "_fetch_pl"):
            fn = globals().get(name)
            if callable(fn):
                try:
                    return fn(year, seq_no) or {}
                except Exception as e:
                    dprint(f"[WARN] {name} 실패: year={year}, seq={seq_no}, err={e}")
        # 4) 최종 안전 fallback
        dprint(f"[WARN] PL 데이터 소스 없음 → year={year}, seq={seq_no} 0값 사용")
        return {}

    def _get_bs(year: int) -> Dict[str, Tuple[Number, Number]]:
        if bs_data and year in bs_data:
            return bs_data[year]
        if bs_fetcher:
            try:
                return bs_fetcher(year, seq_no) or {}
            except Exception as e:
                dprint(f"[WARN] bs_fetcher 실패: year={year}, seq={seq_no}, err={e}")
        for name in ("fetch_bs", "_fetch_bs"):
            fn = globals().get(name)
            if callable(fn):
                try:
                    return fn(year, seq_no) or {}
                except Exception as e:
                    dprint(f"[WARN] {name} 실패: year={year}, seq={seq_no}, err={e}")
        dprint(f"[WARN] BS 데이터 소스 없음 → year={year}, seq={seq_no} 0값 사용")
        return {}

    # -----------------------
    # 내부 작업 배열 (0..3): 0=Y, 1=Y-1, 2=Y-2, 3=Y-3
    # -----------------------
    E10 = [0,0,0,0]; F10=[0,0,0,0]; J10=[0,0,0,0]
    N10 = [0,0,0,0]; O10=[0,0,0,0]; Q10=[0,0,0,0]
    A00 = [1,1,1,1]; A10=[0,0,0,0]; A20=[0,0,0,0]
    B00 = [0,0,0,0]; B10=[0,0,0,0]; B20=[0,0,0,0]
    C00 = [1,1,1,1]; C10=[1,1,1,1]
    Z108= [1,1,1,1]; Z260=[0,0,0,0]; Z293=[0,0,0,0]
    Z951= [0,0,0,0]  # 이자비용

    # 초기값 (원본과 동일)
    for k in range(4):
        Z108[k]=1; Z260[k]=0; Z293[k]=0; C00[k]=1

    # -----------------------
    # 주입 (Y, Y-2 기준으로 cur/prev 한 번에)
    # -----------------------
    for i in (0, 2):
        year = work_yy - i

        # ----- P/L -----
        pl = _get_pl(year)
        def getp(ac: str) -> Tuple[Number, Number]:
            return pl.get(ac, (0.0, 0.0))

        cur, prev = getp("E10"); E10[i], E10[i+1] = cur, prev
        cur, prev = getp("F10"); F10[i], F10[i+1] = cur, prev
        cur, prev = getp("J10"); J10[i], J10[i+1] = cur, prev
        cur, prev = getp("N10"); N10[i], N10[i+1] = cur, prev

        cur, prev = getp("O10")
        O10[i]   = (1.0 if cur  == 0 else cur)
        O10[i+1] = (1.0 if prev == 0 else prev)

        cur, prev = getp("Q10")
        if i == 0:
            Q10[i]   = N10[i]   - O10[i]
            Q10[i+1] = N10[i+1] - O10[i+1]
        else:
            Q10[i], Q10[i+1] = cur, prev

        # 951 이자비용
        cur, prev = getp("951"); Z951[i], Z951[i+1] = cur, prev

        # ----- B/S -----
        bs = _get_bs(year)
        def gets(ac: str) -> Tuple[Number, Number]:
            return bs.get(ac, (0.0, 0.0))

        cur, prev = gets("A00"); A00[i], A00[i+1] = cur, prev
        cur, prev = gets("A10"); A10[i], A10[i+1] = cur, prev
        cur, prev = gets("A20"); A20[i], A20[i+1] = cur, prev

        cur, prev = gets("B00"); B00[i], B00[i+1] = cur, prev
        cur, prev = gets("B10"); B10[i], B10[i+1] = cur, prev
        cur, prev = gets("B20"); B20[i], B20[i+1] = cur, prev

        # C00 = A00 - B00 (0이면 1로)
        C00[i]   = A00[i]   - B00[i]
        C00[i+1] = A00[i+1] - B00[i+1]
        if C00[i]   == 0: C00[i]   = 1
        if C00[i+1] == 0: C00[i+1] = 1

        # C10 (0이면 1로)
        cur, prev = gets("C10")
        C10[i]   = (1.0 if cur  == 0 else cur)
        C10[i+1] = (1.0 if prev == 0 else prev)

        # 108,260,293
        cur, prev = gets("108"); Z108[i], Z108[i+1] = cur, prev
        cur, prev = gets("260"); Z260[i], Z260[i+1] = cur, prev
        cur, prev = gets("293"); Z293[i], Z293[i+1] = cur, prev

    dprint(f"[SEQ={seq_no}][YY={work_yy}] RAW arrays filled.")
    dprint("E10:", E10)
    dprint("J10:", J10, "Z951:", Z951)
    dprint("Z260:", Z260, "Z293:", Z293)
    dprint("A00:", A00, "B00:", B00, "C00:", C00, "C10:", C10)
    dprint("Z108:", Z108)

    # -----------------------
    # 경고지표 계산 (3 x 13)
    # -----------------------
    financeIssue: List[List[str]] = [[""]*13 for _ in range(3)]
    warningCount = 0

    def mark(row: int, col: int):
        nonlocal warningCount
        if financeIssue[row][col] != "√":
            financeIssue[row][col] = "√"
            warningCount += 1

    # 인덱스: 0=Y,1=Y-1,2=Y-2,3=Y-3
    # 화면 rowIdx: 0=Y-2,1=Y-1,2=Y
    pairs = [(3,2,0), (2,1,1), (1,0,2)]  # (과거→현재, row)

    # ─ 매출액 감소(전년 -30%) : col 0
    for pa, pb, row in pairs:
        if E10[pa] > 0 and _safe_div(E10[pb] - E10[pa], E10[pa]) < -0.30:
            mark(row, 0)

    # ─ 영업이익 감소(전년 -30%) : col 1
    for pa, pb, row in pairs:
        if J10[pa] > 1 and _safe_div(J10[pb] - J10[pa], J10[pa]) < -0.30:
            mark(row, 1)

    # ─ 매출채권/매출액 > 50% : col 3
    check_trip = [(2,0), (1,1), (0,2)]  # (지표 idx, row)
    for idx, row in check_trip:
        if E10[idx] > 0 and _safe_div(Z108[idx], E10[idx]) > 0.5:
            mark(row, 3)

    # ─ 영업손실(J10 < 0) : col 5
    for idx, row in [(2,0), (1,1), (0,2)]:
        if J10[idx] < 0:
            mark(row, 5)

    # ─ 차입금 증가 30%↑ : col 7
    def _rise_over_30(prev_idx: int, now_idx: int, row: int):
        prev_sum = Z260[prev_idx] + Z293[prev_idx]
        now_sum  = Z260[now_idx] + Z293[now_idx]
        if prev_sum > 1 and _safe_div(now_sum - prev_sum, prev_sum) > 0.30:
            mark(row, 7)

    _rise_over_30(3,2,0)  # Y-3 → Y-2
    _rise_over_30(2,1,1)  # Y-2 → Y-1
    _rise_over_30(1,0,2)  # Y-1 → Y

    # ─ 차입금/총자산 > 50% : col 8
    for idx, row in [(2,0),(1,1),(0,2)]:
        if A00[idx] != 0 and _safe_div(Z260[idx] + Z293[idx], A00[idx]) > 0.5:
            mark(row, 8)

    # ─ 단기차입/총차입 > 90% (총차입/총자산>50%) : col 9
    for idx, row in [(2,0),(1,1),(0,2)]:
        sum_z = Z260[idx] + Z293[idx]
        if A00[idx] != 0 and _safe_div(sum_z, A00[idx]) > 0.5:
            if sum_z != 0 and _safe_div(Z260[idx], sum_z) > 0.9:
                mark(row, 9)

    # ─ 부채비율 300%↑ : col 10
    for idx, row in [(2,0),(1,1),(0,2)]:
        if C00[idx] > 0 and (_safe_div(B00[idx], C00[idx]) * 100) > 300:
            mark(row, 10)

    # ─ 일부자본잠식(C00 < C10 이면서 자본총액 양수) : col 11
    for idx, row in [(2,0),(1,1),(0,2)]:
        if (C00[idx] < C10[idx]) and ((A00[idx] - B00[idx]) > 0):
            mark(row, 11)

    # ─ 완전자본잠식(C00 < 0) : col 12
    for idx, row in [(2,0),(1,1),(0,2)]:
        if C00[idx] < 0:
            mark(row, 12)

    # 메시지
    if warningCount <= 3:
        txt = "※ 영업활동 및 재무활동에 대한 경고표시가 3개 이하 발생하여 재무능력 및 기업신용도 상황이 우수합니다."
    elif warningCount <= 5:
        txt = "※ 영업활동 및 재무활동에 대한 경고표시가 5개 이하 발생하여 재무능력 및 기업신용도 상황이 양호합니다."
    else:
        txt = "※ 영업활동 및 재무활동에 대한 경고표시가 6개 이상 발생하여 재무능력 및 기업신용도 상황이 건전하지 않습니다."

    # 표시에 쓸 연도(오름차순)
    years = [work_yy-2, work_yy-1, work_yy]

    # 반환 key_values: years = [Y-2, Y-1, Y] ↔ 내부 인덱스 [2,1,0]
    key_values = {
        "J10":  [J10[2],  J10[1],  J10[0]],   # 영업이익
        "Z951": [Z951[2], Z951[1], Z951[0]],  # 이자비용
        "Z260": [Z260[2], Z260[1], Z260[0]],  # 단기차입금
        "Z293": [Z293[2], Z293[1], Z293[0]],  # 장기차입금

        # 필요 시 추가 전달
        "E10":  [E10[2],  E10[1],  E10[0]],
        "Z108": [Z108[2], Z108[1], Z108[0]],
        "A00":  [A00[2],  A00[1],  A00[0]],
        "B00":  [B00[2],  B00[1],  B00[0]],
        "C00":  [C00[2],  C00[1],  C00[0]],
    }

    # 빠른 검증 로그
    dprint(f"years = {years}")
    for i, y in enumerate(years):
        dprint(f"  [{i}] year={y}  J10={key_values['J10'][i]}  Z951={key_values['Z951'][i]}  "
               f"Z260={key_values['Z260'][i]}  Z293={key_values['Z293'][i]}")

    return {
        "ok": True,
        "years": years,                        # [Y-2, Y-1, Y]
        "financeIssue": financeIssue,          # 3 x 13, "√" 또는 ""
        "warningCount": warningCount,
        "message": txt,
        "keyValues": key_values,
        "columns": [
            "매출액 감소(전년 대비 -30%↓)",     # 0
            "영업이익 감소(전년 대비 -30%↓)",   # 1
            "",                                   # 2 (미사용)
            "매출채권/매출액 > 50%",             # 3
            "",                                   # 4 (미사용)
            "영업손실(영업이익<0)",              # 5
            "",                                   # 6 (미사용)
            "차입금 증가 30%↑",                  # 7
            "차입금/총자산 > 50%",               # 8
            "단기차입금/총차입금 > 90%",         # 9
            "부채비율 300%↑",                    # 10
            "일부자본잠식",                      # 11
            "완전자본잠식"                       # 12
        ]
    }


PL_KEYS = {"E10","F10","J10","N10","O10","Q10"}
BS_KEYS = {"A00","A10","B00","B10","C10","108","260","293"}  # 108/260/293 -> Z108/Z260/Z293
def _to_float(v):
    try:
        return float(v or 0)
    except Exception:
        try:
            return float(str(v).replace(",",""))
        except Exception:
            return 0.0
def _blank_arrays():
    return {
        "E10":[0,0,0], "F10":[0,0,0], "J10":[0,0,0], "N10":[0,0,0], "O10":[0,0,0], "Q10":[0,0,0],
        "A00":[0,0,0], "A10":[0,0,0], "B00":[0,0,0], "B10":[0,0,0], "C10":[0,0,0],
        "C00":[0,0,0],
        "Z108":[0,0,0], "Z260":[0,0,0], "Z293":[0,0,0],
    }

def _exec_sp(cursor, sp_name, work_yy, seq_no):
    cursor.execute(f"EXEC {sp_name} %s, %s", [str(work_yy), str(seq_no)])
    cols = [c[0] for c in cursor.description] if cursor.description else []
    rows = cursor.fetchall()
    out = []
    for row in rows:
        rec = dict(zip(cols, row))
        acnt = rec.get("계정코드") or rec.get("account_code") or rec.get("acnt_cd") or rec.get("ACNT_CD")
        cur  = rec.get("당기잔액1") or rec.get("cur_amt") or rec.get("RIGHTAMT") or rec.get("CUR")
        prev = rec.get("전기잔액1") or rec.get("prev_amt") or rec.get("RIGHTAMT_PRE") or rec.get("PREV")
        out.append({"acnt_cd": str(acnt or "").strip(), "cur": _to_float(cur), "prev": _to_float(prev)})
    return out

def _fill_pl(arr, records, idx_cur, idx_prev):
    for r in records:
        a = r["acnt_cd"]
        if a in PL_KEYS:
            arr[a][idx_cur]  = r["cur"]
            arr[a][idx_prev] = r["prev"]

def _fill_bs(arr, records, idx_cur, idx_prev):
    for r in records:
        a = r["acnt_cd"]
        if a in BS_KEYS:
            key = {"108":"Z108","260":"Z260","293":"Z293"}.get(a, a)
            arr[key][idx_cur]  = r["cur"]
            arr[key][idx_prev] = r["prev"]
#기업진단
def _build_diagnosis_payload(seq_no: int, work_yy: int, work_qt: int):
    memdeal = MemDeal.objects.get(seq_no=seq_no)
    end_dt    = getLastFiscalDate(seq_no, work_yy, memdeal.fiscalmm)
    diag_mmdd = end_dt.strftime("%m-%d")     # 'MM-DD'
    quarter_start = f"{end_dt.month:02d}-01" 
    from_30 = end_dt
    # diag_mmdd     = QT_END.get(work_qt, "12-31")
    print(f"end_dt:{end_dt}")
    print(f"diag_mmdd:{diag_mmdd}")
    print(f"quarter_start:{quarter_start}")
    print(f"from_30:{from_30}")

    # “월” 평균 산출용 시작/끝 날짜(실제 계산에 사용)
    month_start   = end_dt.replace(day=1)                       # date(2025, 11, 1)
    avg_start_dt  = month_start                                 # 평균 시작
    avg_end_dt    = end_dt 

    # 0) 기준자본금/업종
    with connection.cursor() as cur:
        cur.execute(
            "SELECT COALESCE(SUM(MH_Amt),0) FROM Diag_Capital WHERE Seq_No=%s AND MH_DcRate<'2'",
            [seq_no]
        )
        stndCapital = float(cur.fetchone()[0] or 0)

        cur.execute("""
            SELECT MH_Name, MH_Amt, MH_DcRate
            FROM Diag_Capital
            WHERE Seq_No=%s  AND MH_DcRate<'2'
            ORDER BY MH_Name
        """, [seq_no])
        capitalRows = [
            {"MH_Name": r[0], "MH_Amt": float(r[1] or 0), "MH_DcRate": str(r[2] or "")}
            for r in cur.fetchall()
        ]

    # 1) BSInquiry로 A00/B00/101/138 + ★ 부실자산(BAD_ASSET_CODES/이름패턴) 동시 수집
    amt_A00=amt_B00=amt_101=amt_108=amt_109=amt_138=0.0
    bad_list = []   # ← [{code,name,amt}]
    bad_sum  = 0.0

    with connection.cursor() as cur:
        cur.execute("EXEC up_Act_BSInquiry %s,%s", [str(work_yy), str(seq_no)])
        cols  = [c[0].lower() for c in cur.description]
        i_cd  = cols.index('financial_acntcd') if 'financial_acntcd' in cols else 0
        i_nm  = cols.index('financial_trnacntnm') if 'financial_trnacntnm' in cols else 1
        i_bal = cols.index('당기잔액1') if '당기잔액1' in cols else (cols.index('bal') if 'bal' in cols else 2)

        for row in cur.fetchall():
            code = str(row[i_cd]).strip()
            name = str(row[i_nm]).strip()
            bal  = _to_float(row[i_bal])

            # 합계/대손충당금 등은 여기선 각각 목적대로만 사용
            if   code == 'A00': amt_A00 = bal
            elif code == 'B00': amt_B00 = bal
            elif code == '101': amt_101 = bal
            elif code == '108': amt_108 = bal
            elif code == '109': amt_109 = bal
            elif code == '138': amt_138 = bal

            # ★ 부실자산 후보: 코드셋 or 이름패턴 (A00/B00/109 제외)
            if code not in ('A00','B00','109'):
                if str(code) == '249' and int(seq_no) == 3650:
                    pass  # 도경개발의 출자금은 부실자산에서 제외
                elif (code in BAD_ASSET_CODES or BAD_NAME_REGEX.search(name or '')):
                    if bal != 0:
                        bad_list.append({"code": str(code), "name": name, "amt": float(bal)})
                        bad_sum += float(bal)

    # 1%) 현금·전도금(1%) 계산
    c00         = max(0.0, amt_A00 - max(0.0, amt_B00))
    cash_limit  = c00 * 0.01
    cash_target = max(0.0, amt_101) + max(0.0, amt_138)
    cash_over   = max(0.0, cash_target - cash_limit)
    over_101    = min(max(0.0, amt_101), cash_over)
    over_138    = max(0.0, cash_over - over_101)

    # 2) 103 보통예금: 30일 평균 vs 기말잔액
    dep_last, dep_avg30 = _calc_deposit_avg_last(
        seq_no,
        work_yy,
        #avg_start_dt,   # 2025-11-01
        avg_end_dt        # 2025-11-30
    )
    dep_over = max(0.0, dep_last - dep_avg30)

    # 3) 108/109: 부실, 충당(1%)
    bond_bad, bond_target, bond_allow, bond_after = _calc_bond_and_allowance(seq_no, work_yy, end_dt)
    bondOver = (amt_108 + amt_109)  - bond_after
    # 4) 장부 + 진단분개 합성표
    lines, realCapital_gross = _compose_lines_and_capital(
        seq_no, work_yy, work_qt, diag_mmdd, quarter_start, amt_A00, amt_B00
    )

    # ★ 순자산 - 부실자산 - 현금평정 - 보통예금평정 - 외상매출금평정
    realCapital_net = float(realCapital_gross) - float(bad_sum) - float(cash_over) - float(dep_over) - float(bondOver)

    return {
        "ok": True,
        "diag_date": diag_mmdd,
        "stndCapital": round(stndCapital, 2),
        "realCapital": round(realCapital_net, 2),          # ← 순자산 - 부실자산 차감 후
        "realCapitalGross": round(realCapital_gross, 2),   # ← 차감 전 참고치
        "capitalRows": capitalRows,
        "lines": lines,
        "diagExtra": {
            "c00": c00,
            "cashLimit": cash_limit,
            "cash101": amt_101,
            "adv138":  amt_138,
            "over101": over_101,
            "over138": over_138,
            "cashOver": cash_over,

            "depAvg30": dep_avg30,
            "depLast":  dep_last,
            "depOver":  dep_over,

            "bondBad":    bond_bad,#부실채권
            "bondTarget": bond_target,#설정대상 충당금
            "bondAllow":  bond_allow,#설정대상 충당금 * 1%
            "bondAfter":  bond_after,#평정후 채권잔액
            "book108": amt_108,
            "book109": amt_109,
            "bondOver": bondOver,

            # ★ 프런트에서 즉시 사용
            "badAssetsList": bad_list,  # [{code,name,amt}]
            "badAssets":     bad_sum
        }
    }

def _calc_deposit_avg_last(seq_no: int, work_yy: int, end_dt, dump_limit:int=200):
    """
    보통예금(103) 30일 평잔 계산 + 디버그용 일자/거래처별 잔액 출력:
      - start_dt = end_dt - 29일
      - opening_bal = start_dt 전일까지 누계(차변-대변)
      - 날짜별 delta(차변-대변) 누적 → 일자별 최종잔액
      - dep_last  = end_dt 총 잔액
      - dep_avg30 = 30일 총잔액(일별 합)의 평균
      - 결과가 모두 음수인 경우(크기는 맞고 부호만 반대) 최종 단계에서 +로 뒤집어 교정
      - dump_limit: 디버그 출력 행수 제한
    """
    try:
        start_dt   = end_dt - timedelta(days=29)
        start_mmdd = start_dt.strftime("%m-%d")
        end_mmdd   = end_dt.strftime("%m-%d")
        prev_mmdd  = (start_dt - timedelta(days=1)).strftime("%m-%d")

        print("[_calc_deposit_avg_last] seq_no=", seq_no,
              " work_yy=", work_yy,
              " start_dt=", start_dt,
              " end_dt=", end_dt,
              " prev_mmdd=", prev_mmdd,
              " start_mmdd=", start_mmdd,
              " end_mmdd=", end_mmdd)

        # --- 메인 계산 (dep_last, dep_avg30) ---
        sql_main = r"""
            ;WITH D AS (
                SELECT CAST(%s AS date) AS d
                UNION ALL
                SELECT DATEADD(day, 1, d) FROM D WHERE d < %s
            ),
            ACC AS (
                SELECT DISTINCT Trader_Code
                FROM ds_slipledgr2
                WHERE seq_no=%s AND work_yy=%s AND acnt_cd='103'
            ),
            OPEN_BAL AS (
                SELECT a.Trader_Code,
                    COALESCE(SUM(CASE WHEN x.acnt_cd='103' THEN (x.tranamt_dr - x.tranamt_cr) ELSE 0 END), 0) AS opening_bal
                FROM ACC a
                LEFT JOIN ds_slipledgr2 x
                ON x.seq_no=%s
                AND x.work_yy=%s
                AND x.acnt_cd='103'
                AND x.Trader_Code=a.Trader_Code
                AND x.tran_dt <= %s
                GROUP BY a.Trader_Code
            ),
            DAY_MOV AS (
                SELECT t.Trader_Code,
                    t.dconv AS d,
                    SUM(t.delta) AS delta
                FROM (
                    SELECT
                        x.Trader_Code,
                        DATEFROMPARTS(%s, CAST(SUBSTRING(x.tran_dt,1,2) AS int), CAST(SUBSTRING(x.tran_dt,4,2) AS int)) AS dconv,
                        (x.tranamt_dr - x.tranamt_cr) AS delta
                    FROM ds_slipledgr2 x
                    WHERE x.seq_no=%s
                    AND x.work_yy=%s
                    AND x.acnt_cd='103'
                    AND x.tran_dt BETWEEN %s AND %s
                ) t
                GROUP BY t.Trader_Code, t.dconv
            ),
            GRID AS (
                SELECT a.Trader_Code, d.d, COALESCE(m.delta, 0) AS delta
                FROM ACC a
                CROSS JOIN D d
                LEFT JOIN DAY_MOV m ON m.Trader_Code=a.Trader_Code AND m.d=d.d
            ),
            BAL AS (
                SELECT g.Trader_Code, g.d,
                    (ob.opening_bal
                        + SUM(g.delta) OVER (PARTITION BY g.Trader_Code ORDER BY g.d ROWS UNBOUNDED PRECEDING)
                    ) AS daily_end_bal
                FROM GRID g
                JOIN OPEN_BAL ob ON ob.Trader_Code=g.Trader_Code
            )
            SELECT
                (SELECT COALESCE(SUM(daily_end_bal),0) FROM BAL WHERE d=%s) AS dep_last,
                (SELECT AVG(CAST(day_sum AS float))
                FROM (SELECT d, SUM(daily_end_bal) AS day_sum FROM BAL GROUP BY d) S
                ) AS dep_avg30
            OPTION (MAXRECURSION 400);
            """
        params_main = [
            start_dt, end_dt,
            seq_no, work_yy,
            seq_no, work_yy, prev_mmdd,
            work_yy, seq_no, work_yy, start_mmdd, end_mmdd,
            end_dt
        ]
        print("[_calc_deposit_avg_last] executing main. param count=", len(params_main))
        with connection.cursor() as cur:
            cur.execute(sql_main, params_main)
            row = cur.fetchone()

        dep_last  = float(row[0] or 0.0)
        dep_avg30 = float(row[1] or 0.0)

        # --- 디버그: 거래처×일자별 잔액 덤프 ---
        sql_dump = r"""
            ;WITH D AS (
                SELECT CAST(%s AS date) AS d
                UNION ALL
                SELECT DATEADD(day, 1, d) FROM D WHERE d < %s
            ),
            ACC AS (
                SELECT DISTINCT Trader_Code
                FROM ds_slipledgr2
                WHERE seq_no=%s AND work_yy=%s AND acnt_cd='103'
            ),
            OPEN_BAL AS (
                SELECT a.Trader_Code,
                    COALESCE(SUM(CASE WHEN x.acnt_cd='103' THEN (x.tranamt_dr - x.tranamt_cr) ELSE 0 END), 0) AS opening_bal
                FROM ACC a
                LEFT JOIN ds_slipledgr2 x
                ON x.seq_no=%s
                AND x.work_yy=%s
                AND x.acnt_cd='103'
                AND x.Trader_Code=a.Trader_Code
                AND x.tran_dt <= %s
                GROUP BY a.Trader_Code
            ),
            DAY_MOV AS (
                SELECT t.Trader_Code,
                    t.dconv AS d,
                    SUM(t.delta) AS delta
                FROM (
                    SELECT
                        x.Trader_Code,
                        DATEFROMPARTS(%s, CAST(SUBSTRING(x.tran_dt,1,2) AS int), CAST(SUBSTRING(x.tran_dt,4,2) AS int)) AS dconv,
                        (x.tranamt_dr - x.tranamt_cr) AS delta
                    FROM ds_slipledgr2 x
                    WHERE x.seq_no=%s
                    AND x.work_yy=%s
                    AND x.acnt_cd='103'
                    AND x.tran_dt BETWEEN %s AND %s
                ) t
                GROUP BY t.Trader_Code, t.dconv
            ),
            GRID AS (
                SELECT a.Trader_Code, d.d, COALESCE(m.delta, 0) AS delta
                FROM ACC a
                CROSS JOIN D d
                LEFT JOIN DAY_MOV m ON m.Trader_Code=a.Trader_Code AND m.d=d.d
            ),
            BAL AS (
                SELECT g.Trader_Code, g.d,
                    (ob.opening_bal
                        + SUM(g.delta) OVER (PARTITION BY g.Trader_Code ORDER BY g.d ROWS UNBOUNDED PRECEDING)
                    ) AS daily_end_bal
                FROM GRID g
                JOIN OPEN_BAL ob ON ob.Trader_Code=g.Trader_Code
            )
            SELECT TOP (%s)
                Trader_Code,
                CONVERT(varchar(10), d, 120) AS d,
                daily_end_bal
            FROM BAL
            ORDER BY Trader_Code, d
            OPTION (MAXRECURSION 400);
            """
        params_dump = [
            start_dt, end_dt,
            seq_no, work_yy,
            seq_no, work_yy, prev_mmdd,
            work_yy, seq_no, work_yy, start_mmdd, end_mmdd,
            dump_limit
        ]
        with connection.cursor() as cur:
            cur.execute(sql_dump, params_dump)
            dump_rows = cur.fetchall()

        # 보기 좋게 로그로 뿌리기
        print("[_calc_deposit_avg_last][DUMP] Trader_Code | date | daily_end_bal  (top", dump_limit, ")")
        last_trader = None
        for trader, dstr, bal in dump_rows:
            if trader != last_trader:
                print(f"  ─ Trader: {trader}")
                last_trader = trader
            print(f"    {dstr}  =>  {bal}")

        # --- 부호 교정(필요 시) ---
        # 크기는 맞고 부호만 반대인 경우 → 양수로 교정
        if dep_last < 0 and dep_avg30 < 0:
            dep_last  = -dep_last
            dep_avg30 = -dep_avg30

        print("[_calc_deposit_avg_last] RESULT dep_last=", dep_last, " dep_avg30=", dep_avg30)
        return dep_last, dep_avg30

    except Exception as e:
        print("[_calc_deposit_avg_last][ERROR]", repr(e))
        print("[_calc_deposit_avg_last] seq_no=", seq_no, "work_yy=", work_yy, "end_dt=", end_dt)
        return 0.0, 0.0

def _safe_int(x, default=None):
    try:
        return int(str(x).strip())
    except Exception:
        return default

def _parse_mmdd(mmdd: str):
    """
    'MM-DD' 또는 'MMDD' → (mm:int, dd:int)
    '00-00'/'0000'/잘못된 값이면 None
    """
    if not mmdd:
        return None
    s = str(mmdd).strip()
    if s in ('00-00', '0000'):
        return None
    try:
        if '-' in s:            # 'MM-DD'
            mm, dd = s.split('-')
        else:                   # 'MMDD'
            s = s.zfill(4)
            mm, dd = s[:2], s[2:]
        mm_i, dd_i = int(mm), int(dd)
        if 1 <= mm_i <= 12 and 1 <= dd_i <= 31:
            return mm_i, dd_i
    except Exception:
        pass
    return None
def _calc_bond_and_allowance(seq_no:int, work_yy:int, end_dt):
    """
    108 외상매출금 잔액 → 최근 매출일 역산해 24개월↑ 부실분 합계
    충당금 1% 산출
    """
    # 거래처별 기말잔액
    with connection.cursor() as cur:
        cur.execute("""
            SELECT AA.Trader_Code,
                   SUM(AA.전기이월)+SUM(AA.차변)-SUM(AA.대변) AS 기말잔액
            FROM(
              SELECT Trader_Code,
                     0 AS 전기이월,
                     SUM(tranAmt_Cr) AS 차변,
                     SUM(tranAmt_Dr) AS 대변
              FROM DS_SlipLedgr2
              WHERE work_yy=%s AND seq_no=%s AND acnt_cd=108
                AND ISNULL(cncl_Dt,'')='' AND tran_dt<>'00-00'
              GROUP BY Trader_Code
              UNION ALL
              SELECT Trader_Code,
                     SUM(CASE WHEN Acnt_cd BETWEEN 101 AND 250 OR (Acnt_cd>=451 AND NOT(Acnt_cd BETWEEN 901 AND 950))
                              THEN tranAmt_Cr-tranAmt_Dr
                              WHEN Acnt_cd BETWEEN 251 AND 330
                              THEN tranAmt_Dr-tranAmt_Cr END) AS 전기이월,
                     0,0
              FROM DS_SlipLedgr2
              WHERE work_yy<%s AND seq_no=%s AND acnt_cd=108
                AND ISNULL(cncl_Dt,'')='' AND tran_dt<>'00-00'
              GROUP BY Trader_Code
            ) AA
            GROUP BY AA.Trader_Code
            HAVING SUM(AA.전기이월)+SUM(AA.차변)-SUM(AA.대변) <> 0
        """, [work_yy, seq_no, work_yy, seq_no])
        rows = cur.fetchall()

    bond_bad = 0.0
    bond_target = 0.0
    for trader_code, end_bal in rows:
        end_bal = _to_float(end_bal)

        # 최근 매출일 역산
        with connection.cursor() as cur2:
            cur2.execute("""
                SELECT
                    CAST(work_yy AS int)      AS work_yy,
                    CASE
                    WHEN LEN(tran_dt)=4 THEN STUFF(tran_dt,3,0,'-')   -- 'MMDD' -> 'MM-DD'
                    ELSE tran_dt
                    END                       AS tran_dt,
                    tranamt_cr
                FROM DS_SlipLedgr2
                WHERE seq_no=%s AND acnt_cd=108 AND tran_dt<>'00-00' AND COALESCE(cncl_Dt,'')=''
                AND Trader_Code=%s
                ORDER BY work_yy DESC, tran_dt DESC                       
            """, [seq_no, trader_code])
            tx = cur2.fetchall()

        remain = end_bal
        last_sale_date = None

        for wyy, tdt, amt_cr in tx:
            remain -= _to_float(amt_cr)
            if remain > 0:
                continue

            # 이 거래일까지 소진 → 이 날짜가 미수발생 기준
            wy = _safe_int(wyy)
            mmdd = _parse_mmdd(tdt)
            if wy is None or mmdd is None:
                # 디버깅 로그 (원인 추적용)
                print(f"[DIAG][WARN] last_sale_date parse fail: wyy={wyy!r}, tran_dt={tdt!r}")
                continue

            mm, dd = mmdd
            try:
                last_sale_date = date(wy, mm, dd)
            except Exception as e:
                print(f"[DIAG][WARN] invalid date composed: {wy}-{mm}-{dd} ({e})")
                last_sale_date = None
            break

        if last_sale_date:
            diff_months = (end_dt.year - last_sale_date.year)*12 + (end_dt.month - last_sale_date.month)
            if diff_months >= 24:
                bond_bad += end_bal
                continue
        bond_target += end_bal

    bond_allow = max(0.0, bond_target * 0.01)
    bond_after = max(0.0, bond_target - bond_allow)
    return bond_bad, bond_target, bond_allow, bond_after
BAD_ASSET_CODES = {
    # ASP searchTxt 와 동일 세트
    '107','114','116','120','123','124','125','131','133','134','137','139','140','146',
    '150','153','156','159','162','166','167','168','169','170','171','172','179','181','182',
    '218','219','220','221','222','223','224','225','226','227','228','229','230','231',
    '235','236','237','238','239','240','241','242','243','244','245','246','247','248','249','250',
}
BAD_NAME_REGEX = re.compile(r'(단기대여금|가지급금|미수금|미수수익|선급금|선납세금|재고|투자|지분법|무형|영업권|개발비|창업비|출자금|회수의문|부실)')
def _compose_lines_and_capital(seq_no:int, work_yy:int, work_qt:int, diag_mmdd:str, quarter_start:str,
                               amt_A00:float, amt_B00:float):
    # ── 0) 기간 필터
    if work_qt == 4:
        period_filter = "x.work_yy=%s AND x.Tran_Dt BETWEEN '00-00' AND %s"
        params        = [work_yy, diag_mmdd]
    else:
        period_filter = "x.work_yy=%s AND x.Tran_Dt BETWEEN %s AND %s"
        params        = [work_yy, quarter_start, diag_mmdd]

    # ── 1) 장부 집계 (계정별 book_amt)
    with connection.cursor() as cur:
        cur.execute(f"""
            SELECT x.acnt_cd,
                   MAX(COALESCE(z.Financial_AcntnM, x.acnt_nm)) AS prt_nm,
                   SUM(CASE WHEN x.acnt_cd<=250 OR (x.acnt_cd>=451 AND NOT(x.acnt_cd BETWEEN 901 AND 950))
                            THEN (x.tranamt_cr - x.tranamt_dr)
                            ELSE (x.tranamt_dr - x.tranamt_cr) END) AS book_amt,
                   MIN(nm.prt_ord) AS prt_ord
            FROM ds_slipledgr2 x
            LEFT JOIN financial_SpcAcnt z
              ON x.acnt_cd=z.Financial_AcntCd AND x.seq_no=z.seq_no
            LEFT JOIN financial_acntnm3 nm
              ON nm.Financial_AcntCd=x.acnt_cd
            WHERE x.seq_no=%s
              AND x.acnt_cd<=330
              AND x.acnt_cd NOT IN ('C00','C10','375','377','376')
              AND ({period_filter})
            GROUP BY x.acnt_cd
            ORDER BY MIN(nm.prt_ord), x.acnt_cd
        """, [seq_no, *params])
        base = cur.fetchall()

    # ── 2) 진단분개 (Diag_Total)
    with connection.cursor() as cur:
        cur.execute("""
            SELECT acnt_cd, acnt_nm, tranamt_dr, tranamt_cr, prt_ord
            FROM Diag_Total
            WHERE seq_no=%s AND work_yy=%s AND Diag_Date=%s
        """,[seq_no, work_yy, diag_mmdd])
        diag = cur.fetchall()

    # ── 3) rows 초기화 (book + diag 반영)
    rows = {}
    for acnt_cd, prt_nm, book_amt, prt_ord in base:
        k = str(acnt_cd).strip()
        rows[k] = {
            "acnt_cd": k,
            "prt_nm": prt_nm or "",
            "bookAmt": _to_float(book_amt),
            "DiagDr":  0.0,
            "DiagCr":  0.0,
            "afterDiag": _to_float(book_amt),
            "prt_ord": int(prt_ord or 0)
        }

    for acnt_cd, acnt_nm, dr, cr, prt_ord in diag:
        k = str(acnt_cd).strip()
        if k not in rows:
            rows[k] = {
                "acnt_cd": k,
                "prt_nm": acnt_nm or "",
                "bookAmt": 0.0,
                "DiagDr":  0.0,
                "DiagCr":  0.0,
                "afterDiag": 0.0,
                "prt_ord": int(prt_ord or 9999),
            }
        r = rows[k]
        r["DiagDr"]    = _to_float(dr)
        r["DiagCr"]    = _to_float(cr)
        r["afterDiag"] = r["bookAmt"] - r["DiagDr"] + r["DiagCr"]

    # ── 4) BAD 계정군 강제 포함(없으면 0원으로 추가)
    #     prt_nm은 financial_acntnm3에서 가져오되, 없으면 fallback
    missing_codes = [cd for cd in BAD_ASSET_CODES if cd not in rows]
    if missing_codes:
        with connection.cursor() as cur:
            cur.execute(f"""
                SELECT Financial_AcntCd, MAX(COALESCE(Financial_TrnAcntNm,'')) AS nm,
                       MIN(COALESCE(prt_ord, 9999)) AS ord
                FROM financial_acntnm3
                WHERE Financial_AcntCd IN ({','.join(['%s']*len(missing_codes))})
                GROUP BY Financial_AcntCd
            """, missing_codes)
            name_rows = {str(cd): (nm or "", int(ord or 9999)) for cd, nm, ord in cur.fetchall()}

        for cd in missing_codes:
            nm, ordv = name_rows.get(cd, (f"계정 {cd}", 9999))
            rows[cd] = {
                "acnt_cd": cd,
                "prt_nm": nm,
                "bookAmt": 0.0,
                "DiagDr":  0.0,
                "DiagCr":  0.0,
                "afterDiag": 0.0,
                "prt_ord": ordv,
            }

    # ── 5) A00/B00 가상행 주입(프런트 안정성↑)
    A00_after = _to_float(rows.get("A00",{}).get("afterDiag", amt_A00))
    B00_after = _to_float(rows.get("B00",{}).get("afterDiag", amt_B00))
    if B00_after < 0:
        B00_after = 0.0

    rows["A00"] = {
        "acnt_cd":"A00","prt_nm":"자산총액(A00)","bookAmt":A00_after,"DiagDr":0.0,"DiagCr":0.0,
        "afterDiag":A00_after,"prt_ord":0
    }
    rows["B00"] = {
        "acnt_cd":"B00","prt_nm":"부채총액(B00)","bookAmt":B00_after,"DiagDr":0.0,"DiagCr":0.0,
        "afterDiag":B00_after,"prt_ord":1
    }

    realCapital = A00_after - B00_after

    # ── 6) 정렬 및 반환
    lines = sorted(rows.values(), key=lambda x:(x["prt_ord"], x["acnt_cd"]))
    return lines, realCapital

from datetime import date

from datetime import date, datetime as dt

def _get_fiscalmm(seq_no: int) -> int:
    with connection.cursor() as cur:
        cur.execute("SELECT TOP 1 ISNULL(fiscalMM, 12) FROM mem_user WHERE seq_no=%s", [seq_no])
        row = cur.fetchone()
    try:
        fm = int(row[0]) if row and row[0] else 12
        return fm if fm in (3, 6, 9, 12) else 12
    except Exception:
        return 12

def getLastFiscalDate(seq_no: int, work_yy: int, fiscalmm: int):
    """
    반환: 마지막 기준일을 나타내는 datetime.date 한 개만 리턴
    (레코드가 없거나 파싱 실패 시 결산월 말일로 대체)
    """
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 MAX(tran_dt)
            FROM DS_SlipLedgr2
            WHERE seq_no=%s AND work_yy=%s
            GROUP BY tran_dt
            ORDER BY tran_dt DESC
        """, [seq_no, work_yy])
        row = cur.fetchone()

    # 결산월 말일 기본값
    if fiscalmm == 12:
        default_end = dt.strptime(f"{work_yy}-12-31", "%Y-%m-%d").date()
    elif fiscalmm == 9:
        default_end = dt.strptime(f"{work_yy}-09-30", "%Y-%m-%d").date()
    elif fiscalmm == 6:
        default_end = dt.strptime(f"{work_yy}-06-30", "%Y-%m-%d").date()
    else:  # 3
        default_end = dt.strptime(f"{work_yy}-03-31", "%Y-%m-%d").date()

    if row and row[0]:
        mmdd = str(row[0])  # 예: '09-30'
        try:
            return dt.strptime(f"{work_yy}-{mmdd}", "%Y-%m-%d").date()
        except Exception:
            return default_end
    return default_end


# 파일 상단 임포트 (한 스타일로 통일 권장)
from datetime import datetime as dt, timedelta

def fetch_from_to_bal(cur, work_yy, seq_no, trader_code, from_dt, to_dt):
    """
    up_Act_FromToBal / up_act_fromtobal 호출.
    - 인자: work_yy, seq_no, trader_code, from_dt('YYYY-MM-DD'), to_dt('YYYY-MM-DD')
    - 1) 저장프로시저, 2) TVF, 3) 대체 SQL 순서로 시도
    - 반환: 커서.fetchall() 결과(일자별 합계 시계열; 첫 컬럼을 합계로 사용)
    """
    # 날짜 정규화
    if isinstance(from_dt, str):
        from_dt = dt.strptime(from_dt, "%Y-%m-%d").date()
    if isinstance(to_dt, str):
        to_dt = dt.strptime(to_dt, "%Y-%m-%d").date()

    params = [str(work_yy), str(seq_no), str(trader_code), from_dt, to_dt]

    # 1) 저장프로시저 시도
    for sql in (
        "EXEC dbo.up_Act_FromToBal ?,?,?,?,?",
        "EXEC up_Act_FromToBal ?,?,?,?,?",
        "EXEC dbo.up_act_fromtobal ?,?,?,?,?",
        "EXEC up_act_fromtobal ?,?,?,?,?",
    ):
        try:
            cur.execute(sql, params)
            return cur.fetchall()
        except Exception:
            pass

    # 2) TVF 시도
    for sql in (
        "SELECT * FROM dbo.up_Act_FromToBal(?,?,?,?,?)",
        "SELECT * FROM up_Act_FromToBal(?,?,?,?,?)",
        "SELECT * FROM dbo.up_act_fromtobal(?,?,?,?,?)",
        "SELECT * FROM up_act_fromtobal(?,?,?,?,?)",
    ):
        try:
            cur.execute(sql, params)
            return cur.fetchall()
        except Exception:
            pass

    # 3) 대체 SQL (103 계정, 거래처별 기간 합계 시계열)
    fallback_sql = """
    ;WITH Dates AS (
      SELECT CAST(%s AS date) AS d
      UNION ALL
      SELECT DATEADD(day, 1, d) FROM Dates WHERE d < %s
    ),
    Tx AS (
      SELECT CAST(CONCAT(x.work_yy,'-',x.tran_dt) AS date) AS dt,
             SUM(x.tranAmt_Cr - x.tranAmt_Dr) AS delta
        FROM DS_SlipLedgr2 x
       WHERE x.seq_no = %s
         AND x.acnt_cd = '103'
         AND COALESCE(x.cncl_Dt,'') = ''
         AND x.tran_dt <> '00-00'
         AND CAST(CONCAT(x.work_yy,'-',x.tran_dt) AS date) BETWEEN %s AND %s
         AND x.Trader_Code = %s
       GROUP BY CAST(CONCAT(x.work_yy,'-',x.tran_dt) AS date)
    ),
    DayAgg AS (
      SELECT d.d,
             COALESCE(t.delta, 0) AS day_sum
        FROM Dates d
        LEFT JOIN Tx t ON t.dt = d.d
    )
    SELECT day_sum
      FROM DayAgg
      ORDER BY d
    OPTION (MAXRECURSION 32767);
    """
    # 파라미터: from_dt, to_dt, seq_no, from_dt, to_dt, trader_code
    cur.execute(fallback_sql, [from_dt, to_dt, str(seq_no), from_dt, to_dt, str(trader_code)])
    return cur.fetchall()

def getFinancialData_Report(request):
    seq_no   = _to_int(request.GET.get("seq_no"), 0)
    flag     = (request.GET.get("flag") or "PL").upper()
    work_yy  = _to_int(request.GET.get("work_yy") or request.GET.get("work_YY")) \
               or timezone.localdate().year
    work_qt  = _to_int(request.GET.get("work_qt") or request.GET.get("work_QT"))
    fiscalMM = _to_int(request.GET.get("fiscalMM") or request.GET.get("FiscalMM"), 12)
    endDate  = (request.GET.get("endDate") or "").strip()

    memuser = MemUser.objects.filter(seq_no=seq_no).only(
        "seq_no","biz_type","biz_no","ssn","reg_date"
    ).first()

    if not memuser:
        return JsonResponse({"error": "사용자 정보가 없습니다."}, status=400)
    # ── 당기 주요매출처(도넛)
    if flag in {"SALES_TOP", "TOP_SALES", "SALES_DONUT"}:
        top_rows = _query_sales_top(seq_no, work_yy, fiscalMM, memuser.biz_type)
        return JsonResponse({
            "ok": True,
            "work_yy": work_yy,
            "fiscalMM": fiscalMM,
            "topCustomers": top_rows  # salesTopDonut()에서 normalize
        })
    # ── 당기 주요매입처(도넛) ← 여기 추가
    if flag in {"PURCHASE_TOP", "TOP_PURCHASE", "PURCHASE_DONUT"}:
        vendor_rows = _query_purchase_top(seq_no, work_yy, fiscalMM, memuser.biz_type)
        return JsonResponse({
            "ok": True,
            "work_yy": work_yy,
            "fiscalMM": fiscalMM,
            "topVendors": vendor_rows  # renderTopPurchaseDonut()에서 normalize하여 사용
        })
    # ── 연도별(기존 PL 집계)
    if flag in {"PL", "ANNUAL"}:
        # ★ 여기서 기초 집계 + 디버그 출력 실행
        base_sales, base_cogs, base_salary, base_sga, base_nonOpInc, base_nonOpExp, base_tax = \
            _get_base_pl_aggregates(seq_no, work_yy, fiscalMM)

        # (원래 하던 연도별 payload 구성)
        rtnJson = _build_annual_payload(memuser, work_yy, fiscalMM)

        # 원하면 프론트에서 쓰도록 같이 내려도 됨
        rtnJson["baseAgg"] = {
            "sales":   base_sales,
            "cogs":    base_cogs,
            "salary":  base_salary,
            "sga":     base_sga,
            "nonOpInc": base_nonOpInc,
            "nonOpExp": base_nonOpExp,
            "tax":     base_tax,
        }
        return JsonResponse(rtnJson)
    # ── 매출채권 상위 리스트
    if flag in {"AR_LIST", "AR_TOP", "AR_TABLE"}:
        as_of_str = (request.GET.get("as_of") or "").strip()
        try:
            as_of = datetime.date.fromisoformat(as_of_str) if as_of_str else timezone.localdate()
        except Exception:
            as_of = timezone.localdate()
        rows = _query_ar_top_list(seq_no, work_yy, fiscalMM, as_of)
        return JsonResponse({"ok": True, "as_of": as_of.isoformat(), "rows": rows})
    # ── 분기별 매출(전년 동분기 비교)
    if flag in {"PL_QUARTERLY", "QUARTERLY"}:
        quarterlyData = _query_quarterly_two_years(seq_no, work_yy, fiscalMM)
        return JsonResponse({"ok": True, "quarterlyData": quarterlyData})
    # ── 월별 매출(전년동월 비교)
    if flag in {"PL_MONTHLY", "MONTHLY"}:
        monthlyData = _query_monthly_two_years(seq_no, work_yy, fiscalMM)
        return JsonResponse({"ok": True, "monthlyData": monthlyData})
    # ── 주요 매입처
    if flag in {"TOP_PURCHASE"}:
        purchaseData = _query_purchase_top(seq_no, work_yy, fiscalMM, memuser.biz_type)
        return JsonResponse({"ok": True, "purchaseData": purchaseData})
    # ── 매입내역 증빙분석 (A4 표)
    if flag in {"PURCHASE_EVID", "EVID_PURCHASE"}:
        rows = _query_purchase_evidence(seq_no,work_yy,fiscalMM,memuser.biz_type,memuser.biz_no,memuser.ssn)
        return JsonResponse({"ok": True, "receiptAnalysis": rows})
    # ───────────────── TAX 저장(옵션) ─────────────────
    if flag in {"SET_ADVICE","ADVICE_SET"}:
        txt = (request.POST.get("text") or request.GET.get("text") or "").strip()
        if not txt:
            return JsonResponse({"ok": False, "msg":"text가 비어 있습니다."}, status=400)

        endDate_calc, work_mm, is_disabled = _compute_enddate_and_flags(
            int(getattr(memuser, "biz_type", 0) or 0),
            work_yy
        )

        with connection.cursor() as cur:
            # upsert (endDate는 계산값 사용)
            cur.execute("""
                IF EXISTS (
                    SELECT 1 FROM tbl_report_advice
                    WHERE seq_no=%s AND work_yy=%s AND (ISNULL(endDate,'') = ISNULL(%s,''))
                )
                    UPDATE tbl_report_advice
                    SET txtAdvice=%s
                    WHERE seq_no=%s AND work_yy=%s AND (ISNULL(endDate,'') = ISNULL(%s,''))
                ELSE
                    INSERT INTO tbl_report_advice(seq_no, work_yy, endDate, txtAdvice)
                    VALUES(%s, %s, %s, %s)
            """, [
                seq_no, str(work_yy), endDate_calc, txt,
                seq_no, str(work_yy), endDate_calc,
                seq_no, str(work_yy), endDate_calc, txt
            ])

        return JsonResponse({
            "ok": True,
            "endDate": endDate_calc,
            "work_mm": work_mm,
            "is_disabled": is_disabled
        })
    # ───────────────── TAX 조회 ─────────────────
    if flag in {"TAX_ESTIMATE","TAX"}:

        # endDate/flags 계산
        endDate_calc, work_mm, is_disabled = _compute_enddate_and_flags(
            int(getattr(memuser, "biz_type", 0) or 0),
            work_yy
        )

        payload = _query_tax_estimate(
            seq_no=seq_no, work_yy=work_yy, work_qt=work_qt, fiscalMM=fiscalMM,
            biz_type=int(memuser.biz_type or 0),
            biz_no=str(memuser.biz_no or "").strip(),
            ssn=str(memuser.ssn or "").strip(),
            reg_date=getattr(memuser, "reg_date", None)
        )

        # 담당 의견 조회 (계산된 endDate 기준)
        with connection.cursor() as cur:
            cur.execute(
                "SELECT txtAdvice FROM tbl_report_advice "
                " WHERE seq_no=%s AND work_yy=%s AND (ISNULL(endDate,'') = ISNULL(%s,''))",
                [seq_no, str(work_yy), endDate_calc]
            )
            row = cur.fetchone()
            advice = (row[0] if row and row[0] else "")

        payload.update({
            "advice": advice,
            "work_yy": work_yy,
            "work_qt": work_qt,
            "fiscalMM": fiscalMM,
            # ↓ 계산 결과도 함께 반환
            "endDate": endDate_calc,
            "work_mm": work_mm,
            "is_disabled": is_disabled,
        })
        return JsonResponse({"ok": True, **payload})
    # ──── 4. 재무건전성 ─────────────────
    if flag == "ISSUE_TABLE":
        data = _compute_issue_table(seq_no, work_yy)
        return JsonResponse(data)
    # ──── 5. 재무비율 ─────────────────
    if flag in ("RATIOS", "FIN_RATIO", "FIN_RATIOS"):
        years = [work_yy - 2, work_yy - 1, work_yy]
        arr = _blank_arrays()

        with connection.cursor() as cur:
            # (1) work_yy: 당기→idx=2, 전기→idx=1
            pl_now = _exec_sp(cur, "up_Act_PLInquiry", work_yy, seq_no)
            _fill_pl(arr, pl_now, idx_cur=2, idx_prev=1)

            bs_now = _exec_sp(cur, "up_Act_BSInquiry", work_yy, seq_no)
            _fill_bs(arr, bs_now, idx_cur=2, idx_prev=1)

            # (2) work_yy-2: 당기만 idx=0
            y2 = work_yy - 2
            pl_y2 = _exec_sp(cur, "up_Act_PLInquiry", y2, seq_no)
            for r in pl_y2:
                if r["acnt_cd"] in PL_KEYS:
                    arr[r["acnt_cd"]][0] = r["cur"]

            bs_y2 = _exec_sp(cur, "up_Act_BSInquiry", y2, seq_no)
            for r in bs_y2:
                a = r["acnt_cd"]
                if a in BS_KEYS:
                    key = {"108":"Z108","260":"Z260","293":"Z293"}.get(a, a)
                    arr[key][0] = r["cur"]

        # C00 = A00 - B00
        arr["C00"] = [
            (arr["A00"][0] - arr["B00"][0]),
            (arr["A00"][1] - arr["B00"][1]),
            (arr["A00"][2] - arr["B00"][2]),
        ]

        data = {"years": years, **arr}
        return JsonResponse({"ok": True, "data": data})
    # ──── 6 재무제표 ─────────────────
    if flag == "STATEMENT":
        flag2 = request.GET.get("flag2")[-2:]
        rows = []
        proc_name = f"up_Act_{flag2}Inquiry"
        with connection.cursor() as cur:
            sql = f"EXEC {proc_name} %s, %s"
            cur.execute(sql, [work_yy, str(seq_no)])
            data = cur.fetchall()

            if flag2 == "CS":
                # SP가 CS일 때 컬럼 위치가 다르면 그대로 맞춰 매핑
                # 예상: [.., acnt_cd(2), acnt_nm(3), this(4), last(5)]
                for r in data:
                    rows.append({
                        "acnt_cd": str(r[2]).strip(),
                        "acnt_nm": str(r[3]).strip(),
                        "amt_now": str(r[4]),
                        "amt_bef": str(r[5]),
                    })
            else:
                # BS/PL 공통: [acnt_cd(0), acnt_nm(1), this(2), last(3)]
                for r in data:
                    rows.append({
                        "acnt_cd": str(r[0]).strip(),
                        "acnt_nm": str(r[1]).strip(),
                        "amt_now": str(r[2]),
                        "amt_bef": str(r[3]),
                    })

        # --- 기간 정보 (있으면) ---
        # fiscalMM 구하는 로직이 따로 있다면 그대로 사용
        try:
            row_duration = getLastFiscalDate(seq_no, work_yy, fiscalMM)
        except Exception:
            row_duration = {}

        # 그대로 반환 (정렬/섹션/깊이 등 추가 가공 없음)
        return JsonResponse({"rowDuration": row_duration, "rows": rows}, safe=False)
    if flag == "STATEMENTTR":
        try:
            seq_no  = int(request.GET.get("seq_no"))
            work_yy = int(request.GET.get("work_yy"))
        except Exception:
            return JsonResponse({"ok": False, "error": "Invalid params (seq_no/work_yy)"},
                                status=400, json_dumps_params={"ensure_ascii": False})

        # 선택 파라미터(없으면 자동 추정)
        try:
            # 클라이언트가 주면 우선 사용
            end_mm = request.GET.get("end_month")
            end_mm = int(end_mm) if end_mm is not None else None
        except Exception:
            end_mm = None

        try:
            start_mm = request.GET.get("start_mm")
            start_mm = int(start_mm) if start_mm is not None else 1
        except Exception:
            start_mm = 1
        
        # mountID 는 echo 용(로깅/디버깅)
        mount_id = request.GET.get("flag2") or "statementTR"

        # 제외 계정 (원본과 동일)
        excluded = ('101','135','138','254','255')

        # 0) end_mm 자동 추정: 해당년도 DS_SlipLedgr2 중 tran_dt='MM-DD' 최대 MM
        if end_mm is None:
            with connection.cursor() as cur:
                cur.execute(r"""
                    SELECT MAX(CAST(LEFT(tran_dt,2) AS INT)) AS max_mm
                    FROM DS_SlipLedgr2
                    WHERE seq_no=%s
                    AND work_yy=%s
                    AND tran_dt <> '00-00'
                    AND LEFT(tran_dt,2) LIKE '[0-9][0-9]'
                """, [seq_no, work_yy])
                row = cur.fetchone()
                end_mm = int(row[0] or 12)
            if end_mm < 1 or end_mm > 12:
                end_mm = 12

        # 1) 계정 목록(원본 rs2): 잔액>0 인 계정만
        accounts_sql = r"""
            WITH F AS (
            SELECT CAST(a.acnt_cd AS INT) AS acnt_cd, 
                    MAX(a.acnt_nm) AS acnt_nm,
                    SUM(CASE WHEN a.acnt_cd<251 THEN a.tranAmt_CR - a.tranAmt_DR
                            ELSE a.tranAmt_DR - a.tranAmt_CR END) AS bal
            FROM DS_SlipLedgr2 a
            JOIN Financial_AcntCd b ON a.Acnt_cd = b.acnt_cd
            WHERE a.Seq_No = %s
                AND a.Work_YY = %s
                AND LEFT(a.tran_dt,2) LIKE '[0-9][0-9]'
                AND CAST(LEFT(a.tran_dt, 2) AS INT) <= %s         -- ← TRY_CONVERT 제거
                AND (a.acnt_cd <= 145 OR (a.acnt_cd BETWEEN 176 AND 194) OR (a.acnt_cd BETWEEN 231 AND 330))
                AND a.acnt_cd NOT IN (%s,%s,%s,%s,%s)
            GROUP BY a.acnt_cd
            )
            SELECT acnt_cd, acnt_nm
            FROM F
            WHERE bal > 0
            ORDER BY acnt_cd
        """

        # 2) 계정별 거래처 집계(원본 rs): 전기이월/당기 차·대/기말
        trader_sql = r"""
            WITH CURR AS (
            SELECT acnt_cd,
                    Trader_Code AS trader_code,
                    MAX(trader_name) AS trader_name,
                    0 AS begin_bal,
                    ISNULL(SUM(CASE WHEN tran_dt <> '00-00' THEN tranAmt_Cr END), 0) AS dr_amt,
                    ISNULL(SUM(CASE WHEN tran_dt <> '00-00' THEN tranAmt_Dr END), 0) AS cr_amt
            FROM DS_SlipLedgr2
            WHERE work_yy = %s
                AND seq_no = %s
                AND acnt_cd = %s
                AND cncl_Dt = ''
                AND tran_dt <> '00-00'
                AND LEFT(tran_dt,2) LIKE '[0-9][0-9]'            -- ← 숫자 월만
                AND CAST(LEFT(tran_dt,2) AS INT) BETWEEN %s AND %s  -- ← TRY_CONVERT 제거
            GROUP BY acnt_cd, Trader_Code
            ),
            BEGINNING AS (
            SELECT acnt_cd,
                    Trader_Code AS trader_code,
                    MAX(trader_name) AS trader_name,
                    SUM(CASE
                        WHEN acnt_cd <= 250 OR (acnt_cd >= 451 AND NOT (acnt_cd BETWEEN 901 AND 950))
                                THEN tranAmt_Cr - tranAmt_Dr
                        ELSE tranAmt_Dr - tranAmt_Cr
                        END) AS begin_bal,
                    0 AS dr_amt, 0 AS cr_amt
            FROM DS_SlipLedgr2
            WHERE work_yy <= %s
                AND seq_no = %s
                AND acnt_cd = %s
                AND cncl_Dt = ''
                AND tran_dt <> '00-00'
            GROUP BY acnt_cd, Trader_Code
            )
            SELECT X.trader_code,
                MAX(X.trader_name) AS trader_name,
                SUM(X.begin_bal) AS begin_bal,
                SUM(X.dr_amt)    AS dr_amt,
                SUM(X.cr_amt)    AS cr_amt,
                SUM(CASE WHEN X.acnt_cd<=250 OR (X.acnt_cd>=451 AND NOT (X.acnt_cd BETWEEN 901 AND 950))
                            THEN X.begin_bal + X.dr_amt - X.cr_amt
                            ELSE X.begin_bal + X.cr_amt - X.dr_amt
                    END) AS end_bal
            FROM (
            SELECT * FROM CURR
            UNION ALL
            SELECT * FROM BEGINNING
            ) X
            GROUP BY X.trader_code
            HAVING SUM(CASE WHEN X.acnt_cd<=250 OR (X.acnt_cd>=451 AND NOT (X.acnt_cd BETWEEN 901 AND 950))
                            THEN X.begin_bal + X.dr_amt - X.cr_amt
                            ELSE X.begin_bal + X.cr_amt - X.dr_amt
                    END) <> 0
            ORDER BY trader_code
            """


        result = {
            "ok": True,
            "flag": "statementTR",
            "flag2": mount_id,
            "meta": {
                "seq_no": seq_no,
                "work_yy": work_yy,
                "start_mm": start_mm,
                "end_mm": end_mm,
                "period_text": f"({start_mm}월 1일 ~ {end_mm}월 말)"
            },
            "accounts": []
        }

        with connection.cursor() as cur:
            # 계정 목록
            cur.execute(accounts_sql, [seq_no, work_yy, end_mm, *excluded])
            accounts = cur.fetchall()  # [(acnt_cd, acnt_nm), ...]

            for acnt_cd, acnt_nm in accounts:
                cur.execute(
                    trader_sql,
                    [
                        work_yy, seq_no, acnt_cd,
                        start_mm, end_mm,
                        work_yy - 1, seq_no, acnt_cd
                    ]
                )
                rows = cur.fetchall()

                sb = sd = sc = se = 0.0
                items = []
                for trader_code, trader_name, begin_bal, dr_amt, cr_amt, end_bal in rows:
                    begin_bal = float(begin_bal or 0)
                    dr_amt    = float(dr_amt or 0)
                    cr_amt    = float(cr_amt or 0)
                    end_bal   = float(end_bal or 0)
                    sb += begin_bal; sd += dr_amt; sc += cr_amt; se += end_bal
                    inc = dr_amt
                    dec = cr_amt
                    # === 자산 / 부채 구분 ===
                    # if 100 <= acnt_cd <= 250:        # 자산
                        # inc = dr_amt
                        # dec = cr_amt
                    # elif 251 <= acnt_cd <= 450:      # 부채
                    #     inc = cr_amt
                    #     dec = dr_amt
                    # else:
                    #     # 자본(451~), 수익/비용(501~) 등은 증감 계산 제외 또는 별도 처리
                    #     inc = dec = 0.0

                    items.append({
                        'trader_code': trader_code,
                        'trader_name': trader_name,
                        'begin_bal': begin_bal,
                        'dr_amt': dr_amt,
                        'cr_amt': cr_amt,
                        'end_bal': end_bal,
                        'inc': inc,
                        'dec': dec
                    })

                is_asset_like = (acnt_cd <= 250) or (acnt_cd >= 451 and not (901 <= acnt_cd <= 950))
                columns = ["거래처명", "기초잔액"] + \
                          (["당기증가(+)", "당기감소(-)"] if is_asset_like else ["당기감소(-)", "당기증가(+)"]) + \
                          ["기말잔액"]

                totals_inc = sd #if is_asset_like else sc
                totals_dec = sc #if is_asset_like else sd

                footnote = None
                if acnt_cd == 251:
                    footnote = "* 외상매입금 기말잔액이 (-)인 경우 대금은 지급하였으나 세금계산서를 받지 못한 경우입니다."
                elif acnt_cd == 108:
                    footnote = "* 외상매출금 기말잔액이 (-)인 경우 대금은 수령하였으나 세금계산서를 발행하지 않은 경우입니다."

                result["accounts"].append({
                    "acnt_cd": acnt_cd,
                    "acnt_nm": acnt_nm,
                    "is_asset_like": is_asset_like,
                    "period_text": result["meta"]["period_text"],
                    "columns": columns,
                    "rows": items,
                    "totals": {
                        "begin": sb,
                        "inc": totals_inc,
                        "dec": totals_dec,
                        "end": se
                    },
                    "footnote": footnote
                })

        return JsonResponse(result, json_dumps_params={"ensure_ascii": False})
    # 7 주식가치평가
    if flag == "EV":
        # ---------- helpers ----------
        def _q(sql, params=None, one=False):
            with connection.cursor() as cur:
                cur.execute(sql, params or [])
                if cur.description is None:
                    return None if one else []
                cols = [c[0] for c in cur.description]
                rows = [dict(zip(cols, r)) for r in cur.fetchall()]
            return (rows[0] if rows else None) if one else rows

        def _num(x):
            try:
                return float(x or 0)
            except Exception:
                return 0.0

        trace = []
        def dbg(tag, payload):
            trace.append({"tag": tag, "data": payload})
            try:
                print(f"[EV][{tag}] {payload}")
            except Exception:
                pass

        # ---------- inputs ----------
        try:
            seq_no = int(request.GET.get("seq_no") or request.POST.get("seq_no"))
        except Exception:
            return JsonResponse({"ok": False, "error": "seq_no required"}, status=400)

        today = date.today()

        # 회사 기본정보 (mem_user)
        memrow = _q("""
            SELECT a.biz_name, a.ceo_name, a.email, a.biz_no, a.biz_type, a.reg_date
            FROM mem_user a
            WHERE a.seq_no = %s
        """, [seq_no], one=True)
        if not memrow:
            return JsonResponse({"ok": False, "error": "company not found"}, status=404)

        biz_no   = str(memrow["biz_no"] or "").strip()
        biz_type = int(memrow["biz_type"] or 0)     # <-- 쉼표 금지
        reg_date = memrow.get("reg_date")
        createdYear = int(str(reg_date)[:4]) if reg_date else today.year
        dbg("input.meta", {
            "seq_no": seq_no, "biz_no": biz_no, "biz_type": biz_type,
            "reg_date": str(reg_date), "createdYear": createdYear, "today": str(today)
        })

        # work_yy
        work_yy_raw = (request.GET.get("work_yy") or request.POST.get("work_yy") or "").strip()
        try:
            work_yy = int(work_yy_raw)
        except Exception:
            # ASP와 동일한 자동 보정
            if (today.month < 4 and biz_type < 4) or (today.month < 6 and biz_type >= 4):
                work_yy = today.year - 1
            else:
                work_yy = today.year
        dbg("input.work_yy", {"work_yy_raw": work_yy_raw, "work_yy": work_yy})

        # 기준 구간 및 k(기말 보정)
        curDate = int(today.strftime("%Y%m%d"))
        y = work_yy
        if int(f"{y}0211") <= curDate <= int(f"{y}0415"):
            endDate = "12-31"; k = 0
        elif int(f"{y}0416") <= curDate <= int(f"{y}0731"):
            endDate = "03-31"; k = 1
        elif int(f"{y}0801") <= curDate <= int(f"{y}1015"):
            endDate = "06-30"; k = 1
        elif int(f"{y}1016") <= curDate <= int(f"{y}1231"):
            endDate = "09-30"; k = 1
        elif curDate >= int(f"{y+1}0101"):
            endDate = "12-31"; k = 0
        else:
            endDate = "12-31"; k = 0
        dbg("period", {"curDate": curDate, "endDate": endDate, "k": k})

        # ---------- 유상증자액 보정(ASP의 component 37) ----------
        def _zj_for_year(yy: int) -> float:
            rec = _q("""
            SELECT TOP 1 Tran_Dt, StckH_FEquityGP
            FROM tbl_StckHListTrn
            WHERE StckH_TY='A2' AND seq_no=%s
            ORDER BY Tran_Dt ASC
            """, [seq_no], one=True)
            if not rec: return 0.0
            gp = _num(rec.get("StckH_FEquityGP"))
            td = str(rec.get("Tran_Dt") or "")[:10]
            try:
                t_year = int(td[:4]); t_month = int(td[5:7])
            except Exception:
                return 0.0

            if yy == t_year:
                return gp * 0.1 / 12.0 * t_month      # 당해연도: 월수 비례
            elif yy > t_year:
                return gp * 0.1                        # 이후 연도: 10% 전액
            else:
                return 0.0                             # 이전 연도: 0

        # ---------- 연도별 컴포넌트 수집(최신부터 n=7) ----------
        def _build_components_for_years(base_year: int, n=7):
            comp = []
            for off in range(n):
                yy = int(base_year) - off

                # 주식수/액면가
                st = _q("""
                    SELECT
                    ISNULL(SUM(CASE WHEN A.StckH_TranGB='B' THEN A.StckH_FEquityNum*-1 ELSE A.StckH_FEquityNum END),0) AS shares,
                    ISNULL(MAX(A.StckH_FEquityFP),0) AS par
                    FROM Tbl_StckHolderList B WITH (NOLOCK)
                    JOIN Tbl_StckHListTrn A WITH (NOLOCK)
                    ON B.Seq_No = A.Seq_No AND B.StckH_Num = A.StckH_Num
                    WHERE B.Seq_No = %s AND REPLACE(A.TRAN_DT,'-','') <= %s
                """, [seq_no, f"{yy}1231"], one=True) or {"shares": 0, "par": 0}
                dbg("sql.st", {"year": yy, "st": st})

                # EquityEval(있으면 우선)
                ee = _q("""
                    SELECT TOP 1 E.*, R.retirement_amt
                    FROM tbl_EquityEval AS E
                    LEFT JOIN tbl_retirement AS R
                    ON E.[사업자번호] = R.biz_no AND LEFT(E.[사업연도말], 4) = R.work_yy
                    WHERE E.[사업자번호] = %s AND LEFT(E.[사업연도말], 4) = %s
                """, [biz_no, str(yy)], one=True)
                dbg("sql.ee", {"year": yy, "has_ee": bool(ee)})

                zj = _zj_for_year(yy)

                if not ee:
                    # ds_slipledgr2 집계 (ASP의 no-EE 분기와 동일 필드 포함)
                    bs = _q("""
                        SELECT
                        ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 101 AND 250 THEN tranAmt_cr-tranAmt_dr ELSE 0 END),0) AS 자산총액,
                        ISNULL((SELECT SUM(tranamt_dr)
                                FROM ds_slipledgr2
                                WHERE seq_no = %s AND Tran_Dt <> '00-00' AND acnt_cd = 147 AND Work_YY < %s),0) AS 매입에누리,
                        ISNULL(SUM(CASE WHEN acnt_cd = 133 THEN tranAmt_cr-tranAmt_dr ELSE 0 END),0) AS 선급비용,
                        ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 251 AND 330 THEN tranAmt_dr-tranAmt_cr ELSE 0 END),0) AS 부채총액,
                        -- 부동산 비중 판정용
                        ISNULL(SUM(CASE WHEN acnt_cd = 201 THEN tranAmt_cr-tranAmt_dr ELSE 0 END),0) AS 토지,
                        ISNULL(SUM(CASE WHEN acnt_cd = 202 THEN tranAmt_cr-tranAmt_dr ELSE 0 END),0) AS 건물,
                        ISNULL(SUM(CASE WHEN acnt_cd = 203 THEN tranAmt_dr-tranAmt_cr ELSE 0 END),0) AS 건물감가누계,

                        ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 401 AND 430 THEN tranAmt_dr ELSE 0 END),0) AS 매출액,
                        ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 451 AND 470 THEN tranAmt_cr ELSE 0 END),0) AS 매출원가,
                        ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 801 AND 810 THEN tranAmt_cr-tranAmt_dr ELSE 0 END),0) AS 급여,
                        ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 811 AND 900 THEN tranAmt_cr-tranAmt_dr ELSE 0 END),0) AS 기타판관비,
                        ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 901 AND 950 THEN tranAmt_dr-tranAmt_cr ELSE 0 END),0) AS 영업외수익,
                        ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 951 AND 997 THEN tranAmt_cr-tranAmt_dr ELSE 0 END),0) AS 영업외비용,
                        ISNULL(SUM(CASE WHEN acnt_cd BETWEEN 998 AND 999 THEN tranAmt_cr-tranAmt_dr ELSE 0 END),0) AS 법인세등,
                        ISNULL(SUM(CASE WHEN acnt_cd = 383 THEN tranAmt_dr-tranAmt_cr ELSE 0 END),0) AS 자기주식
                        FROM ds_slipledgr2
                        WHERE seq_no = %s AND Remk <> '손익계정에 대체' AND work_yy = %s
                    """, [seq_no, work_yy, seq_no, yy], one=True) or {}
                    dbg("sql.bs", {"year": yy, "bs": bs})

                    # ⬇️ 추가: 해당 연도 퇴직충당금
                    ret = _q("""
                        SELECT TOP 1 retirement_amt
                        FROM tbl_retirement
                        WHERE work_yy=%s AND seq_no=%s
                    """, [str(yy), seq_no], one=True) or {"retirement_amt": 0}

                    valKack = (_num(bs.get("매출액")) - _num(bs.get("매출원가")) - _num(bs.get("급여")) - _num(bs.get("기타판관비"))
                            + _num(bs.get("영업외수익")) - _num(bs.get("영업외비용")))

                    # 자산/부채 구성요소를 변수로 분해 (검증용)
                    asset_base         = _num(bs.get("자산총액")) - _num(bs.get("매입에누리"))
                    asset_minus_prepa  = _num(bs.get("선급비용"))
                    asset_plus_treas   = -_num(bs.get("자기주식"))   # ASP: 자기주식 * -1 을 자산가산으로
                    liab_base          = _num(bs.get("부채총액"))
                    liab_corp_tax      = 0.0   # no-EE 에선 0
                    liab_nong_tax      = 0.0
                    liab_local_tax     = 0.0
                    liab_dividend_add  = 0.0
                    liab_retirement    = _num(ret.get("retirement_amt"))

                    asset_total = asset_base + asset_plus_treas - asset_minus_prepa
                    liab_total  = liab_base + liab_corp_tax + liab_nong_tax + liab_local_tax + liab_dividend_add + liab_retirement

                    # 🔎 프린트(서버 콘솔 + trace)
                    dbg("liab.noEE.breakdown", {
                        "year": yy,
                        "liab_base(부채총액)": liab_base,
                        "법인세": liab_corp_tax,
                        "농특세": liab_nong_tax,
                        "지방세": liab_local_tax,
                        "부채가산배당": liab_dividend_add,
                        "retirement_amt": liab_retirement,
                        "liab_total": liab_total
                    })
                    dbg("asset.noEE.breakdown", {
                        "year": yy,
                        "자산총액": _num(bs.get("자산총액")),
                        "매입에누리": _num(bs.get("매입에누리")),
                        "선급비용": _num(bs.get("선급비용")),
                        "자기주식(자산가산)": asset_plus_treas,
                        "asset_total": asset_total
                    })

                    row = {
                        "year": yy,
                        "shares": int(_num(st["shares"])),
                        "par": int(_num(st["par"])),
                        "asset_total": asset_total,
                        "liab_total": liab_total,
                        "ga": valKack, "adj_profit": 0, "deduct": 0,
                        "corp_type": "",
                        "land": _num(bs.get("토지")), "building": _num(bs.get("건물")), "depr_acc": _num(bs.get("건물감가누계")),
                        "zj": zj,
                        # ⬇️ 검증용 원재료도 함께 보관
                        "_liab_parts": {
                            "liab_base": liab_base,
                            "corp_tax": liab_corp_tax, "nong_tax": liab_nong_tax, "local_tax": liab_local_tax,
                            "dividend_add": liab_dividend_add, "retirement": liab_retirement
                        }
                    }
                    dbg("comp.row.noEE", row)
                    comp.append(row)
                else:
                    # EquityEval 있는 해
                    nongtax = _num(str(ee.get("농특세") or "0")[:15])
                    raw_corp_tax = _num(ee.get("차감납부세액_법인세"))  # 부호 확인용
                    liab_corp_tax = abs(raw_corp_tax)                  # 회계상 음수여도 '부채 가산' 의미이므로 abs

                    liab_base         = _num(ee.get("부채가액"))
                    liab_nong_tax     = _num(nongtax)
                    liab_local_tax    = _num(ee.get("지방세"))
                    liab_dividend_add = _num(ee.get("부채가산배당"))
                    liab_retirement   = _num(ee.get("retirement_amt"))

                    liab_total        = (liab_base + liab_corp_tax + liab_nong_tax +
                                        liab_local_tax + liab_dividend_add + liab_retirement)

                    asset_total = (_num(ee.get("자산가액")) + _num(ee.get("자산가산1")) + _num(ee.get("자산가산2"))
                                - _num(ee.get("자산차감1")) - _num(ee.get("자산차감2")))

                    dbg("normalize.corp_tax", {
                        "year": yy,
                        "raw_corp_tax": raw_corp_tax,
                        "liab_corp_tax(abs)": liab_corp_tax
                    })

                    dbg("liab.EE.breakdown", {
                        "year": yy,
                        "liab_base(부채가액)": liab_base,
                        "법인세(차감납부세액_법인세, abs)": liab_corp_tax,
                        "농특세": liab_nong_tax,
                        "지방세": liab_local_tax,
                        "부채가산배당": liab_dividend_add,
                        "retirement_amt": liab_retirement,
                        "liab_total": liab_total
                    })

                    row = {
                        "year": yy,
                        "shares": int(_num(st["shares"])),
                        "par": int(_num(st["par"])),
                        "asset_total": asset_total,
                        "liab_total": liab_total,
                        "ga": _num(ee.get("각사업연도소득")),
                        "adj_profit": (_num(ee.get("업무용승용차손금산입")) + _num(ee.get("소득가산배당")) + _num(ee.get("소득가산기부추인"))),
                        "deduct": (_num(ee.get("소득공제벌금")) + _num(ee.get("소득공제공과금")) + _num(ee.get("소득공제업무무관")) +
                                _num(ee.get("업무용승용차손금불산입")) + _num(ee.get("소득공제기부금")) + _num(ee.get("소득공제접대비")) +
                                _num(ee.get("외화환산손실")) + _num(ee.get("소득공제지급이자")) + _num(ee.get("소득공제감비추인")) +
                                _num(ee.get("법인세")) + _num(nongtax) + _num(ee.get("지방세"))),
                        "corp_type": (ee.get("회사종류") or ""),
                        "land": 0, "building": 0, "depr_acc": 0,
                        "zj": zj,
                        "_liab_parts": {
                            "liab_base": liab_base,
                            "corp_tax": liab_corp_tax, "nong_tax": liab_nong_tax, "local_tax": liab_local_tax,
                            "dividend_add": liab_dividend_add, "retirement": liab_retirement
                        }
                    }
                    dbg("comp.row.EE", row)
                    comp.append(row)
            return comp

        # ▼ 7개 조회(최근 5개를 출력용으로 사용)
        comp = _build_components_for_years(work_yy, n=7)
        dbg("comp.all", comp)

        # ---------- 예외 플래그 ----------
        is3YearMinus = "√" if comp[k]["ga"] < 0 and comp[k+1]["ga"] < 0 and comp[k+2]["ga"] < 0 else ""
        is3YearLow   = "√" if (work_yy - k - createdYear) < 3 else ""
        overAsset80 = overAsset50 = ""
        if comp and comp[0]["asset_total"] > 0:
            real_estate = comp[0]["land"] + comp[0]["building"] - comp[0]["depr_acc"]
            if real_estate > comp[0]["asset_total"] * 0.8:
                overAsset80 = "√"
            elif real_estate > comp[0]["asset_total"] * 0.5:
                overAsset50 = "√"
        dbg("flags", {
            "is3YearMinus": is3YearMinus, "is3YearLow": is3YearLow,
            "overAsset80": overAsset80, "overAsset50": overAsset50
        })

        # ---------- 순손익가치(연도별 k_local 적용) ----------
        def _k_local_for_year(y_val: int) -> int:
            """해당 연도가 이미 결산 완료면 0, 아니면 전역 k."""
            return 0 if y_val <= (today.year - 1) else k

        def _profit_block(y_idx: int) -> float:
            """
            대상 연도마다 k를 동적으로 적용:
            - 이미 결산이 끝난 과거 연도(오늘 기준 y <= today.year-1): k_local = 0
            - 그 외: 전역 k 유지
            (3년 가중합: 3:2:1)
            """
            y_target = int(comp[y_idx]["year"])
            k_local = _k_local_for_year(y_target)

            # 인덱스 계산(가중 3개 연도)
            i0 = y_idx + 0 + k_local
            i1 = y_idx + 1 + k_local
            i2 = y_idx + 2 + k_local

            # 범위 체크
            if i2 >= len(comp):
                # 계산 불성립
                dbg("profit_block.skip", {"y_idx": y_idx, "k_local": k_local, "need_indices": [i0, i1, i2], "len": len(comp)})
                return 0.0

            r0, r1, r2 = comp[i0], comp[i1], comp[i2]
            # S = (가산 포함 손익 - 공제 + 유상증자보정) 의 가중합
            S0 = r0["ga"] + r0["adj_profit"] - r0["deduct"] + r0.get("zj", 0.0)
            S1 = r1["ga"] + r1["adj_profit"] - r1["deduct"] + r1.get("zj", 0.0)
            S2 = r2["ga"] + r2["adj_profit"] - r2["deduct"] + r2.get("zj", 0.0)

            # 🔎 디버그: 창에 어떤 연도가 가중에 들어갔는지 출력
            dbg("profit_block.window", {
                "y_idx(base)": y_idx, "k_local": k_local,
                "years": [r0["year"], r1["year"], r2["year"]],
                "indices": [i0, i1, i2],
                "partials": {"S0": S0, "S1": S1, "S2": S2},
                "weighted_total": S0*3 + S1*2 + S2*1
            })
            return S0*3 + S1*2 + S2*1

        PLUS_ASSET_RATE = 0.8 if today >= date(2018, 4, 1) else 0.7
        dbg("PLUS_ASSET_RATE", PLUS_ASSET_RATE)

        # --- 3년 블록이 성립하는 범위 산정(연도별 k_local 반영해서 안전하게) ---
        def _can_compute_income_idx(y_idx: int) -> bool:
            y_target = int(comp[y_idx]["year"])
            k_local = _k_local_for_year(y_target)
            return (y_idx + 2 + k_local) < len(comp)

        # ── 각 연도의 "주당 순손익가치"(ASP의 totArr(y,4))를 최근 5개만 산출
        income_per_share_list = []
        S_trace = []
        for y_idx in range(0, min(5, len(comp))):
            shares_y = max(1, int(_num(comp[y_idx]["shares"])))
            if _can_compute_income_idx(y_idx):
                S_total = _profit_block(y_idx)     # 내부에서 k_local 적용 + 창 구성 로그
            else:
                S_total = 0.0
            per_share_S = math.floor((S_total / 6.0) / shares_y)  # fix(...)
            income_per_share = int(round(max(0.0, per_share_S / 0.1)))  # /0.1 => ×10
            income_per_share_list.append(income_per_share)
            S_trace.append({
                "year": comp[y_idx]["year"], "shares_y": shares_y,
                "S_total": S_total, "per_share_S": per_share_S,
                "income_per_share": income_per_share
            })
        dbg("income_per_share_list", S_trace)

        # ── 최근 5개 연도의 최종 결과 산출
        years_out = []
        for y_idx in range(0, min(5, len(comp))):
            yrow = comp[y_idx]
            yv = int(yrow["year"])
            shares = max(1, int(_num(yrow["shares"])))

            # S_total: _profit_block이 k_local 사용 (불가 시 0)
            if _can_compute_income_idx(y_idx):
                S_total = _profit_block(y_idx)
            else:
                S_total = 0.0
            S = S_total / 6.0

            net_assets = yrow["asset_total"] - yrow["liab_total"]
            goodwill = max(0.0, (S * 0.5 - net_assets * 0.1) * 3.79079)
            asset_per_share = int(round((net_assets + goodwill) / shares))

            # ▼ 가중평균용 순손익 인덱스: ASP 로직(y>0 이면 y-k, else y)과 동일하게 연도별 k_local 반영
            k_local = _k_local_for_year(yv)
            idx_for_income = (y_idx - k_local) if (y_idx > 0) else y_idx
            idx_for_income = max(0, min(idx_for_income, len(income_per_share_list)-1))
            income_for_weight = income_per_share_list[idx_for_income]

            # 예외 규정
            if is3YearLow or is3YearMinus or overAsset80:
                valuation_per_share = asset_per_share
                rule = "exception(asset)"
            else:
                tmp = (income_for_weight * 3 + asset_per_share * 2) / 5.0
                tmp = max(tmp, asset_per_share * PLUS_ASSET_RATE)
                valuation_per_share = int(round(tmp))
                rule = "weighted/max80"

            corp_type = str(yrow.get("corp_type") or "11")  # 기본 중소기업(미기재 대비)
            premium = 0 if corp_type in {"11", "21", "30"} else int(round(valuation_per_share * 0.2))

            row_out = {
                "year": yv,
                "shares": shares,
                "par": int(_num(yrow["par"])),
                "S": S,
                "net_assets": net_assets,
                "goodwill": goodwill,
                "asset_per_share": asset_per_share,
                "income_per_share": income_per_share_list[y_idx] if y_idx < len(income_per_share_list) else 0,
                "valuation_per_share": valuation_per_share,
                "major_holder_premium": premium,
                "rule": rule,
                "corp_type": corp_type
            }
            dbg(f"year.calc.{yv}", row_out)
            years_out.append(row_out)

        # ▼ (표/차트 정렬 방향은 프론트에서 처리. 필요시 아래 주석 해제)
        # years_out.sort(key=lambda r: r["year"])  # 오름차순(과거→최근)
        years_out = years_out[-5:]  # 방어적 슬라이싱

        out = {
            "ok": True,
            "meta": {
                "biz_no": biz_no, "work_yy": work_yy, "k": k, "endDate": endDate,
                "queried_years": 7,     # ← 7년 조회
                "delivered_years": 5    # ← 5년 전달
            },
            "flags": {
                "is3YearLow": is3YearLow,
                "is3YearMinus": is3YearMinus,
                "overAsset80": overAsset80,
                "overAsset50": overAsset50,
            },
            "years": years_out[:5],     # 최근 5개
        }

        # debug=1 이면 trace 포함
        debug_flag = (request.GET.get("debug") or request.POST.get("debug") or "").lower()
        if debug_flag in {"1", "true", "y", "yes"}:
            out["trace"] = trace

        return JsonResponse(out)
    if flag == "DIAGNOSIS":
        payload = _build_diagnosis_payload(seq_no, work_yy, work_qt)
        return JsonResponse(payload)

    def _safe_filename(name: str) -> str:
        # Windows/일반 파일명 금지문자 제거 
        cleaned = re.sub(r'[\\/:*?"<>|]+', '_', name).strip(' .')
        return cleaned or "merged"
    def _safe_dirpart(name: str) -> str:
        # 경로 세그먼트용(폴더명) 금지문자 제거
        return re.sub(r'[\\/:*?"<>|]+', '_', str(name)).strip(' .') or "_"
    if flag == "MERGE":
        filename = (request.GET.get("filename") or "").strip()

        # ── 1) 저장 디렉터리 구성: STATIC_DIR/static/cert_DS/{biz}/{yy}/기장보고서/{qt}분기
        static_dir  = os.path.normpath(getattr(settings, "STATIC_DIR", ""))  # e.g., BASE_DIR/static
        static_url  = getattr(settings, "STATIC_URL", "/static/")
        if not static_dir or not os.path.isdir(static_dir):
            return JsonResponse({"ok": False, "msg": f"STATIC_DIR이 올바르지 않습니다: {static_dir}"}, status=500)

        biz   = _safe_dirpart(memuser.biz_name)
        yy    = _safe_dirpart(work_yy)
        qtseg = _safe_dirpart(f"{work_qt}분기")

        directory = os.path.join(static_dir, "cert_DS", biz, str(yy), "기장보고서", qtseg)
        os.makedirs(directory, exist_ok=True)

        try:
            # ── 2) 파일명/출력경로
            filename = _safe_filename(filename) if filename else _safe_filename(f"{work_yy}_{work_qt}Q_합본")
            out_path = os.path.join(directory, f"{filename}.pdf")

            # ── 3) 병합 대상 수집(결과파일 제외) + 자연정렬
            files = [f for f in os.listdir(directory)
                    if f.lower().endswith(".pdf") and f.lower() != f"{filename.lower()}.pdf"]
            if not files:
                return JsonResponse({"ok": False, "msg": "병합할 PDF가 없습니다."}, status=400)
            files = natsort.natsorted(files)

            # ── 4) 병합 실행
            merger = PyPDF2.PdfMerger()
            try:
                for f in files:
                    src = os.path.join(directory, f)
                    merger.append(src)
                with open(out_path, "wb") as fout:
                    merger.write(fout)
            finally:
                try:
                    merger.close()
                except Exception:
                    pass

            # ── 5) STATIC_URL로 접근 가능한 URL 계산
            # out_path 기준으로 STATIC_DIR 이후의 상대경로를 STATIC_URL에 붙인다.
            rel_from_static = os.path.relpath(out_path, static_dir).replace(os.sep, "/")
            url = (static_url.rstrip("/") + "/" + rel_from_static.lstrip("/"))

            return JsonResponse({
                "ok": True,
                "msg": f"병합 완료: {directory}",
                "path": out_path,
                "url": url,
                "count": len(files),
            })

        except Exception as e:
            return JsonResponse({"ok": False, "msg": str(e)}, status=500)
    if flag == "MERGE_STATUS":
        try:
            filename = (request.GET.get("filename") or "").strip()
            static_dir = os.path.normpath(getattr(settings, "STATIC_DIR", ""))
            static_url = getattr(settings, "STATIC_URL", "/static/")

            biz   = _safe_dirpart(getattr(memuser, "biz_name", ""))
            yy    = _safe_dirpart(work_yy)
            qtseg = _safe_dirpart(f"{work_qt}분기")
            directory = os.path.join(static_dir, "cert_DS", biz, str(yy), "기장보고서", qtseg)

            print(f"[DBG][STATUS] dir={directory}, filename={filename}")
            if not os.path.isdir(directory):
                return JsonResponse({"ok": False, "msg": "PDF 디렉터리가 없습니다."}, status=404)

            target_path = None
            if filename:
                safe = _safe_filename(filename)  # 하이픈 유지됨
                exact = os.path.join(directory, f"{safe}.pdf")
                if os.path.exists(exact):
                    target_path = exact
                else:
                    # ▶ 접두 일치(예: 6-3_무엇.pdf) 허용 — 숫자 오인 fallback은 여전히 금지
                    cand_list = glob.glob(os.path.join(directory, f"{safe}*.pdf"))
                    if cand_list:
                        target_path = max(cand_list, key=os.path.getmtime)
                    else:
                        return JsonResponse({"ok": False, "msg": f"지정 파일 없음: {safe}.pdf"}, status=404)
            else:
                # filename 미지정일 때만 최신본 fallback
                pdfs = glob.glob(os.path.join(directory, "*.pdf"))
                if not pdfs:
                    return JsonResponse({"ok": False, "msg": "표시할 PDF가 없습니다."}, status=404)
                prefer = [p for p in pdfs if "합본" in os.path.basename(p)]
                pool = prefer if prefer else pdfs
                target_path = max(pool, key=os.path.getmtime)

            rel = os.path.relpath(target_path, static_dir).replace(os.sep, "/")
            mtime = int(os.path.getmtime(target_path))
            url = f"{static_url.rstrip('/')}/{rel.lstrip('/')}?v={mtime}"
            print(f"[DBG][STATUS] OK -> {url}")
            return JsonResponse({"ok": True, "url": url, "path": target_path})

        except Exception as e:
            print("[DBG][STATUS] EXC:", e)
            return JsonResponse({"ok": False, "msg": str(e)}, status=500)

    # ── 알 수 없는 플래그 방어
    return JsonResponse({"ok": False, "error": f"Unknown flag: {flag}"}, status=400)

# 유틸
def _safe_biz_name(name: str) -> str:
    # 폴더명에 안전하게 쓰도록 정리
    return re.sub(r'[\\/:*?"<>|]+', '_', (name or '').strip())

def _normalize_number(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return v
    s = str(v).replace(',', '').strip()
    try:
        return float(s) if '.' in s else int(s)
    except Exception:
        return 0

def _is_excel_date(cell_value) -> bool:
    # ASP: If IsDate(RsSheet(0)) Then … 과 동일
    # openpyxl은 엑셀 날짜를 datetime으로 읽어줄 수 있음. 문자열일 때도 처리
    if isinstance(cell_value, (datetime.date, datetime.datetime)):
        return True
    if isinstance(cell_value, str):
        try:
            # 유연 파싱
            datetime.datetime.strptime(cell_value.strip(), '%Y-%m-%d')
            return True
        except Exception:
            pass
        for fmt in ('%Y.%m.%d', '%Y/%m/%d', '%Y%m%d', '%m/%d/%Y', '%m-%d-%Y'):
            try:
                datetime.datetime.strptime(cell_value.strip(), fmt)
                return True
            except Exception:
                continue
    return False

def _to_date(cell_value) -> datetime.date:
    if isinstance(cell_value, datetime.datetime):
        return cell_value.date()
    if isinstance(cell_value, datetime.date):
        return cell_value
    if isinstance(cell_value, str):
        s = cell_value.strip()
        for fmt in ('%Y-%m-%d','%Y.%m.%d','%Y/%m/%d','%Y%m%d','%m/%d/%Y','%m-%d-%Y'):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except Exception:
                continue
    # 실패 시 오늘
    return datetime.date.today()

def _parse_sheet_title(title: str):
    """
    ASP 원본:
      sheetName = Split(oTable.Name,")")
      sheetCode = Right(sheetName(0),3)
      acnt_nm   = Replace(sheetName(1),"'","") / "$" 제거
    엑셀 시트명 예) "101)현금$", "249)예수금$", "110)외상매출금$" 등
    """
    title = (title or '').replace("'", "")
    parts = title.split(")")
    code = ''
    name = ''
    if len(parts) >= 2:
        left = parts[0]
        code = left[-3:] if len(left) >= 3 else left
        name = parts[1].replace('$', '')
    else:
        # 예외: 괄호 없으면 숫자 3자리 후행 추출 시도
        m = re.search(r'(\d{3})', title)
        if m:
            code = m.group(1)
        name = title.replace('$', '')
    return code.strip(), name.strip()

# ─────────────────────────────────────────────────────────
from django.views.decorators.http import require_POST
from openpyxl import load_workbook
from django.db import connection, transaction
from openpyxl.utils.datetime import from_excel, CALENDAR_WINDOWS_1900  # 직렬날짜 처리
@csrf_exempt
@require_POST
def upload_slip_ledger_excel(request):
    """
    ASP 원본 흐름을 보존한 DS_SlipLedgr2 업로드 (디버그 풍부)
      - 파일 저장 → (fiscalMM==12 ? 당해년도 전체 : 회계말 기준 경계) 삭제
      - Excel 파싱 (prog_value: 세무사랑=1, 더존=0)
      - 금액(우선: col5=CR, col6=DR), 보완(6=CR,8=DR,9=잔액), CR/DR 및 tran_stat/적요/코드 매핑
      - acnt_cd: 뒤에서 3자리 숫자 추출(부족시 zfill)
      - Tran_Dt: 'MM-DD'
      - 스키마 조회로 NOT NULL & default 없는 컬럼 자동 포함/기본값 채움, 길이 클램프
      - up_Act_PreBSInquiry 실행
      - biz_type<4 → tbl_corporate2, else → tbl_income2 업데이트
      - 월별 업로드 현황(tbl_mng_jaroe) 업데이트
      - ★ fiscalMM='12'이면 첫 유효 거래일 연도와 work_yy 불일치 시 400 반환
    """
    import os, re, uuid, calendar, datetime, decimal
    from django.db import connection, transaction
    from django.http import JsonResponse
    from django.conf import settings

    # xls/xlsx 파서
    try:
        import xlrd
    except Exception:
        xlrd = None
    try:
        from openpyxl import load_workbook
    except Exception:
        load_workbook = None

    try:
        print("\n[upload_slip_ledger_excel] ================== START ==================")
        seq_no    = (request.POST.get('seq_no') or '').strip()
        work_yy   = (request.POST.get('work_yy') or '').strip()
        biz_name  = (request.POST.get('biz_name') or '').strip()
        prog_value= (request.POST.get('prog_value') or '1').strip()  # '1': 세무사랑, '0': 더존
        fiscalMM  = (request.POST.get('fiscalMM') or '12').strip()   # '12' or '06' 등
        print(f"[params] seq_no='{seq_no}', work_yy='{work_yy}', biz_name='{biz_name}', prog_value='{prog_value}', fiscalMM='{fiscalMM}'")

        if not seq_no or not work_yy:
            return JsonResponse({"ok": False, "msg": "필수 파라미터 누락(seq_no, work_yy)."}, status=400)

        f = request.FILES.get('uploadFile')
        if not f:
            return JsonResponse({"ok": False, "msg": "업로드 파일이 없습니다."}, status=400)

        print(f"[request.method] {request.method}, [FILES keys] {list(request.FILES.keys())}")
        print(f"[file] name={f.name}, size={getattr(f, 'size', '?')}, content_type={getattr(f, 'content_type', '?')}")

        # 저장 경로
        static_dir = getattr(settings, 'STATIC_DIR', getattr(settings, 'STATIC_ROOT', settings.BASE_DIR))
        root_dir   = os.path.join(static_dir, 'upload')
        os.makedirs(root_dir, exist_ok=True)
        print(f"[path] static_dir={static_dir}, root_dir={root_dir}, exists={os.path.isdir(root_dir)}")

        name, ext = os.path.splitext(f.name)
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_name = f"{name}_{ts}_{uuid.uuid4().hex[:6]}{ext}"
        save_path = os.path.join(root_dir, safe_name)
        with open(save_path, "wb+") as dest:
            for i, chunk in enumerate(f.chunks()):
                dest.write(chunk)
                print(f"[save] wrote chunk #{i}")
        print(f"[save] file saved. size={os.path.getsize(save_path)} bytes")
        ext = ext.lower()
        print(f"[ext] {ext}")
        # ───────────────────────── 날짜 파싱 유틸 (구 ASP 로직에 가깝게) ─────────────────────────
        def _coerce_date(raw):
            """
            - datetime/date면 그대로
            - 그 외에는 문자열로 바꿔서:
            1) 앞 10글자만 잘라 다양한 포맷 시도 (YYYY-MM-DD, YYYY/MM/DD, YYYY.MM.DD)
            2) 숫자만 뽑아서 YYYYMMDD 형식 시도
            """
            if isinstance(raw, datetime.datetime):
                return raw.date()
            if isinstance(raw, datetime.date):
                return raw

            if raw is None:
                return None

            s = str(raw).strip()
            if not s:
                return None

            # 1) 옛날 코드처럼 "앞 10글자"만 먼저 본다
            if len(s) >= 10:
                s10 = s[:10]
                for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d'):
                    try:
                        return datetime.datetime.strptime(s10, fmt).date()
                    except:
                        pass

            # 2) 숫자만 추출해서 YYYYMMDD 형태 시도
            digits = ''.join(ch for ch in s if ch.isdigit())
            if len(digits) >= 8:
                try:
                    y = int(digits[0:4])
                    m = int(digits[4:6])
                    d = int(digits[6:8])
                    return datetime.date(y, m, d)
                except:
                    pass

            return None

        # ───────────────────────── 스키마 조회 유틸 ─────────────────────────
        def _fetch_table_columns(table='DS_SlipLedgr2'):
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT COLUMN_NAME
                    FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_NAME=%s
                    ORDER BY ORDINAL_POSITION
                """, [table])
                return [r[0] for r in cur.fetchall()]

        def _fetch_schema_detail(table='DS_SlipLedgr2'):
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT
                        c.COLUMN_NAME,
                        c.IS_NULLABLE,
                        c.DATA_TYPE,
                        c.CHARACTER_MAXIMUM_LENGTH,
                        c.NUMERIC_PRECISION,
                        c.NUMERIC_SCALE,
                        COLUMNPROPERTY(OBJECT_ID(c.TABLE_SCHEMA + '.' + c.TABLE_NAME), c.COLUMN_NAME, 'IsIdentity') AS IsIdentity,
                        col.column_id,
                        o.object_id
                    FROM INFORMATION_SCHEMA.COLUMNS c
                    JOIN sys.objects o
                      ON o.name = c.TABLE_NAME AND o.type IN ('U','V')
                    JOIN sys.columns col
                      ON col.object_id = o.object_id AND col.name = c.COLUMN_NAME
                    WHERE c.TABLE_NAME=%s
                    ORDER BY c.ORDINAL_POSITION
                """, [table])
                base = cur.fetchall()

            with connection.cursor() as cur:
                cur.execute("""
                    SELECT
                        o.object_id, col.column_id, dc.definition
                    FROM sys.objects o
                    JOIN sys.columns col
                      ON col.object_id=o.object_id
                    LEFT JOIN sys.default_constraints dc
                      ON dc.parent_object_id=o.object_id
                     AND dc.parent_column_id=col.column_id
                    WHERE o.name=%s
                """, [table])
                defs = cur.fetchall()

            has_def = {(r[0], r[1]): (r[2] is not None) for r in defs}
            detail = {}
            for (colname, is_null, dtype, charlen, nump, nums, is_ident, col_id, obj_id) in base:
                detail[colname] = {
                    "is_nullable": (is_null == 'YES'),
                    "data_type"  : (dtype or '').lower(),
                    "char_len"   : charlen,
                    "num_p"      : nump, "num_s": nums,
                    "is_identity": (is_ident == 1),
                    "has_default": has_def.get((obj_id, col_id), False)
                }
            return detail

        db_cols = _fetch_table_columns()
        schema  = _fetch_schema_detail()
        print(f"[db] table=DS_SlipLedgr2 columns({len(db_cols)}): {db_cols}")

        def _ci_in(target, *cands):
            idx = {s.lower(): s for s in target}
            for c in cands:
                if c and c.lower() in idx: return idx[c.lower()]
            return None

        colmap = {
            'seq_no'      : _ci_in(db_cols, 'Seq_No'),
            'work_yy'     : _ci_in(db_cols, 'Work_YY'),
            'acnt_cd'     : _ci_in(db_cols, 'Acnt_cd'),
            'acnt_nm'     : _ci_in(db_cols, 'Acnt_Nm'),
            'tran_dt'     : _ci_in(db_cols, 'Tran_Dt'),
            'remk'        : _ci_in(db_cols, 'Remk'),
            'trader_code' : _ci_in(db_cols, 'Trader_Code'),
            'trader_name' : _ci_in(db_cols, 'Trader_Name'),
            'trader_bizno': _ci_in(db_cols, 'Trader_Bizno'),
            'slip_no'     : _ci_in(db_cols, 'Slip_No'),
            'tran_stat'   : _ci_in(db_cols, 'Tran_Stat'),
            'crdr'        : _ci_in(db_cols, 'CrDr'),
            'tran_cr'     : _ci_in(db_cols, 'TranAmt_Cr'),
            'tran_dr'     : _ci_in(db_cols, 'TranAmt_Dr'),
            'cncl_dt'     : _ci_in(db_cols, 'Cncl_Dt'),
            'reg_ymd'     : _ci_in(db_cols, 'Crt_Dt'),
        }

        # 필수 체크
        for req in ('seq_no','work_yy','acnt_cd','acnt_nm','tran_dt','remk','trader_code','trader_name','slip_no','crdr','tran_cr','tran_dr','reg_ymd'):
            if not colmap.get(req):
                return JsonResponse({"ok": False, "msg": f"테이블 필수 컬럼 누락: {req}"}, status=500)

        # INSERT 컬럼 구성(우선순위)
        key_order = [
            'seq_no','work_yy','acnt_cd','acnt_nm','tran_dt','remk',
            'trader_code','trader_name','trader_bizno','slip_no',
            'tran_stat','crdr','tran_cr','tran_dr','cncl_dt','reg_ymd'
        ]
        ins_cols, ins_keys = [], []
        for k in key_order:
            col = colmap[k]
            if not schema[col]["is_identity"]:
                ins_cols.append(col)
                ins_keys.append(k)

        # NOT NULL & default 없음 컬럼 자동 추가(스키마 기반)
        def _default_for(col):
            info = schema[col]
            dt = info["data_type"]
            if 'char' in dt or 'text' in dt or 'nchar' in dt or 'varchar' in dt or 'nvarchar' in dt:
                return ''
            if dt in ('numeric','decimal','int','bigint','smallint','tinyint','float','real','money','smallmoney'):
                return 0
            return ''

        extra_fill_cols = []
        for c in db_cols:
            if c in ins_cols: continue
            info = schema[c]
            if info["is_identity"]: continue
            if (not info["is_nullable"]) and (not info["has_default"]):
                extra_fill_cols.append(c)
        ins_cols += extra_fill_cols
        if extra_fill_cols:
            print(f"[fillers] add NOT NULL no-default cols: {[(c, _default_for(c)) for c in extra_fill_cols]}")

        # ───────────────────────── 보조 유틸 ─────────────────────────
        def _normalize_number(v):
            if v is None or v == '': return 0
            if isinstance(v, (int, float, decimal.Decimal)): return float(v)
            s = str(v).replace(',', '').strip()
            try:
                return float(s)
            except Exception:
                return 0

        def _is_numeric_str(s):
            if s is None: return False
            try:
                float(str(s).replace(',', '').strip())
                return True
            except: return False

        def _tail3_digits(s):
            s = str(s or '').strip()
            m = re.search(r'(\d+)\D*$', s)
            if not m: return ''
            return m.group(1)[-3:].zfill(3)

        def _dt_to_mmdd(dt: datetime.date) -> str:
            return f"{dt.month:02d}-{dt.day:02d}"

        def _last_day_of_month(year: int, month: int) -> int:
            return calendar.monthrange(year, month)[1]

        # fiscal cutoff(MM-DD)
        try:
            fmm = int(fiscalMM)
        except:
            fmm = 12
        cutoff_dd = _last_day_of_month(int(work_yy), fmm)
        fiscal_cutoff = f"{fmm:02d}-{cutoff_dd:02d}"
        print(f"[fiscal cutoff] fiscalMM={fmm}, cutoff={fiscal_cutoff}")

        # 집계용
        amt_sales = 0.0
        amt_cost  = 0.0

        payloads = []
        last_tran_dt = None
        first_dt_year = None  # ★ 연도 불일치 검사용: 첫 유효 거래일의 연도
        bad_sheets, parsed_rows, skip_no_date = [], 0, 0

        # ───────────────────────── 기존데이터 삭제 ─────────────────────────
        with transaction.atomic():
            with connection.cursor() as cur:
                if fmm == 12:
                    # 해당 년도 전체(단, '00-00' 제외)
                    print(f"[tx] delete range: year={work_yy}, Tran_Dt != '00-00'")
                    cur.execute(
                        "DELETE FROM DS_SlipLedgr2 WHERE Seq_No=%s AND Work_YY=%s AND Tran_Dt<>'00-00'",
                        [seq_no, work_yy]
                    )
                else:
                    # (작년 & Tran_Dt > cutoff) OR (당해 & Tran_Dt <= cutoff)
                    prev_yy = str(int(work_yy)-1)
                    print(f"[tx] delete range: (prev={prev_yy} & >{fiscal_cutoff}) OR (curr={work_yy} & <={fiscal_cutoff}), Tran_Dt!='00-00'")
                    cur.execute(
                        """
                        DELETE FROM DS_SlipLedgr2
                         WHERE Seq_No=%s
                           AND Tran_Dt<>'00-00'
                           AND (
                                (Work_YY=%s AND Tran_Dt > %s)
                             OR (Work_YY=%s AND Tran_Dt <= %s)
                           )
                        """,
                        [seq_no, prev_yy, fiscal_cutoff, work_yy, fiscal_cutoff]
                    )
                # 피드백 테이블도 삭제(ASP 동일)
                print("[tx] delete DS_SlipLedgr2_Feedback for this year")
                cur.execute(
                    "DELETE FROM DS_SlipLedgr2_Feedback WHERE Seq_No=%s AND Work_YY=%s",
                    [seq_no, work_yy]
                )

        # ───────────────────────── Excel 파싱 ─────────────────────────
        def _push_payload(dt, rowvals, from_x='xls', row_idx=None):
            """ASP 매핑 1:1 구성 → payload append + amt_sales/amt_cost 집계"""
            nonlocal last_tran_dt, first_dt_year, amt_sales, amt_cost

            # ★ 첫 유효 거래일 연도 기록
            if first_dt_year is None and isinstance(dt, datetime.date):
                first_dt_year = dt.year
                print(f"[first_dt_year] detected={first_dt_year}")

            # prog_value 분기
            pv = int(prog_value) if prog_value.isdigit() else 1

            # 기본 매핑
            acnt_cd_raw = rowvals.get('acnt_3')  # col3
            acnt_nm     = rowvals.get('acnt_4')  # col4

            # 세무사랑: acnt_cd = Left(col3, 4), 더존: acnt_cd = col3
            acnt_cd_src = str(acnt_cd_raw or '')
            if pv == 1:
                acnt_cd_src = acnt_cd_src[:4]
            # 더존 보정(ASP): 931 → 951
            if pv != 1 and acnt_cd_src == '931':
                acnt_cd_src = '951'

            acnt_cd = _tail3_digits(acnt_cd_src) or '000'

            # 날짜
            tran_dt = _dt_to_mmdd(dt)
            tran_dt_year = str(dt.year)

            # 적요/거래처/전표유형
            if pv == 1:
                remk        = rowvals.get('c7')   # col7
                trader_code = rowvals.get('c8')   # col8
                trader_name = rowvals.get('c9')   # col9
                bill_kind   = rowvals.get('c19')  # col19
            else:
                remk        = rowvals.get('c8')   # col8
                trader_code = rowvals.get('c11')  # col11
                trader_name = rowvals.get('c7')   # col7
                bill_kind   = ''                  # 빈값
            trader_bizno = rowvals.get('c10')      # col10
            slip_no      = rowvals.get('c1')       # col1

            # 적요 숫자형 방지
            remk = '' if _is_numeric_str(remk) else (str(remk or '').strip())

            # 금액 컬럼 숫자화
            amt_c5 = _normalize_number(rowvals.get('c5'))  # 세무사랑: Tran_Cr
            amt_c6 = _normalize_number(rowvals.get('c6'))  # 세무사랑: Tran_Dr
            amt_c8 = _normalize_number(rowvals.get('c8'))  # 일부 양식 보완용

            crdr_text = (str(rowvals.get('c2') or '').strip())

            # ───────────────── 세무사랑(prog_value=1): ASP와 최대한 동일 ─────────────────
            if pv == 1:
                # 원본: CRDR=row[3], Tran_Cr=row[6], Tran_Dr=row[7]
                tran_cr = amt_c5   # 음수 포함 그대로
                tran_dr = amt_c6

                # CRDR 텍스트가 비어 있으면 금액 방향으로 보정(참고용)
                if crdr_text not in ('차변', '대변'):
                    if tran_dr != 0 and tran_cr == 0:
                        crdr_text = '차변'   # DR만 있으면 차변
                    elif tran_cr != 0 and tran_dr == 0:
                        crdr_text = '대변'   # CR만 있으면 대변

            # ───────────────── 더존/기타 포맷(pv != 1): 휴리스틱 + abs ─────────────────
            else:
                # “구분” 텍스트 우선, 없으면 숫자 위치로 추론
                if crdr_text not in ('차변', '대변'):
                    if amt_c5 != 0 and amt_c6 == 0 and amt_c8 == 0:
                        crdr_text = '차변'
                    elif amt_c6 != 0 and amt_c5 == 0 and amt_c8 == 0:
                        crdr_text = '대변'
                    elif amt_c8 != 0 and amt_c5 == 0:
                        crdr_text = '차변'
                    else:
                        crdr_text = ''

                # 하나의 금액만 선택하되, 음수도 인정 → 절대값 사용
                amount_detected = 0.0
                for cand in (amt_c5, amt_c6, amt_c8):
                    if cand is None:
                        continue
                    if cand != 0:
                        amount_detected = abs(cand)
                        break

                tran_cr = 0.0
                tran_dr = 0.0
                if crdr_text == '대변':
                    tran_dr = amount_detected
                elif crdr_text == '차변':
                    tran_cr = amount_detected

            crdr = crdr_text  # 저장값
            tran_stat = (str(bill_kind or '').strip())
            cncl_dt   = ''  # ASP는 공란

            if (last_tran_dt is None) or (dt > last_tran_dt):
                last_tran_dt = dt

            # payload
            p = {
                'seq_no': seq_no, 'work_yy': tran_dt_year,
                'acnt_cd': acnt_cd, 'acnt_nm': str(acnt_nm or '')[:100],
                'tran_dt': tran_dt,
                'remk': (remk or '')[:500],
                'trader_code': str(trader_code or '')[:5],
                'trader_name': str(trader_name or '')[:100],
                'trader_bizno': str(trader_bizno or '')[:20],
                'slip_no': str(slip_no or '')[:5],
                'tran_stat': tran_stat[:50] if colmap.get('tran_stat') else '',
                'crdr': str(crdr or '')[:4],
                'tran_cr': tran_cr, 'tran_dr': tran_dr,
                'cncl_dt': cncl_dt, 'reg_ymd': datetime.date.today().strftime('%Y%m%d'),
            }
            payloads.append(p)

            # 집계(ASP 동일 규칙)
            try:
                cd_num = int(acnt_cd)
            except:
                cd_num = 0
            if 401 <= cd_num <= 430:
                amt_sales += (tran_dr - tran_cr)
            elif (451 <= cd_num <= 470) or (501 < cd_num <= 999):
                amt_cost  += (tran_cr - tran_dr)

            if row_idx is not None and row_idx <= 6:
                print(
                    f"[row{row_idx}] acnt_cd_raw='{acnt_cd_raw}' → acnt_cd='{acnt_cd}', "
                    f"acnt_nm='{acnt_nm}', dt='{tran_dt}', remk='{remk}', slip_no='{slip_no}', "
                    f"trader_code='{trader_code}', trader_name='{trader_name}', "
                    f"tran_stat='{tran_stat}', CR={tran_cr}, DR={tran_dr}, CRDR='{crdr}'"
                )

        # 실제 파싱
        if ext == '.xls':
            if not xlrd:
                return JsonResponse({"ok": False, "msg": "서버에 xlrd가 설치되어 있지 않습니다."}, status=500)
            print(f"[xls] loading workbook: {save_path}")
            try:
                book = xlrd.open_workbook(save_path, formatting_info=False)
            except Exception as e:
                print(f"[xls] open error: {e}")
                raise
            print(f"[xls] sheets: {book.sheet_names()}")

            for sh in book.sheets():
                ncols, nrows = sh.ncols, sh.nrows
                if nrows <= 1:
                    bad_sheets.append(sh.name); continue
                print(f"[xls] sheet='{sh.name}' ncols={ncols}, nrows={nrows}")

                def cellv(r, c):
                    if c is None or c < 0 or c >= ncols: return None
                    return sh.cell_value(r, c)

                def to_dt(r, c0=0):
                    cell = sh.cell(r, c0)
                    from xlrd import XL_CELL_DATE
                    # 1) 엑셀 날짜 타입이면 먼저 시도
                    if cell.ctype == XL_CELL_DATE:
                        try:
                            return xlrd.xldate.xldate_as_datetime(cell.value, book.datemode).date()
                        except:
                            pass

                    # 2) 그래도 안 되면, 구 로직과 유사한 느슨한 파서 사용
                    v = cellv(r, c0)
                    dt = _coerce_date(v)
                    return dt

                # ASP: 첫줄은 헤더. 두번째 줄부터 데이터
                for r in range(1, nrows):
                    dt = to_dt(r, 0)
                    if not dt:
                        skip_no_date += 1
                        if r <= 6: print(f"[skip/xls] no-date row={r} raw={cellv(r,0)}")
                        continue

                    rowvals = {
                        'c0' : cellv(r,0),  'c1' : cellv(r,1),  'c2' : cellv(r,2),  'acnt_3': cellv(r,3),
                        'acnt_4': cellv(r,4),'c5' : cellv(r,5),  'c6' : cellv(r,6),  'c7' : cellv(r,7),
                        'c8' : cellv(r,8),  'c9' : cellv(r,9),  'c10': cellv(r,10), 'c11': cellv(r,11),
                        'c12': cellv(r,12), 'c13': cellv(r,13), 'c14': cellv(r,14), 'c19': cellv(r,19) if ncols>19 else ''
                    }
                    _push_payload(dt, rowvals, from_x='xls', row_idx=r)
                    parsed_rows += 1

        elif ext in ('.xlsx', '.xlsm'):
            if not load_workbook:
                return JsonResponse({"ok": False, "msg": "서버에 openpyxl이 설치되어 있지 않습니다."}, status=500)
            print(f"[xlsx] loading workbook: {save_path}")
            wb = load_workbook(save_path, data_only=True)
            for ws in wb.worksheets:
                print(f"[xlsx] sheet='{ws.title}' max_row={ws.max_row}, max_col={ws.max_column}")
                if ws.max_row <= 1:
                    bad_sheets.append(ws.title); continue

                for rix, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    def v(ci):
                        if ci is None or ci<0: return None
                        if ci >= len(row): return None
                        return row[ci]

                    raw_date = v(0)
                    dt = _coerce_date(raw_date)

                    if not dt:
                        skip_no_date += 1
                        if rix <= 6: print(f"[skip/xlsx] no-date row={rix} raw={raw_date}")
                        continue

                    rowvals = {
                        'c0' : v(0),  'c1' : v(1),  'c2' : v(2),  'acnt_3': v(3),
                        'acnt_4': v(4),'c5' : v(5),  'c6' : v(6),  'c7' : v(7),
                        'c8' : v(8),  'c9' : v(9),  'c10': v(10), 'c11': v(11),
                        'c12': v(12), 'c13': v(13), 'c14': v(14), 'c19': v(19) if ws.max_column>19 else ''
                    }
                    _push_payload(dt, rowvals, from_x='xlsx', row_idx=rix)
                    parsed_rows += 1
        else:
            return JsonResponse({"ok": False, "msg": f"지원하지 않는 형식: {ext}"}, status=400)

        print(f"[parse done] parsed_rows={parsed_rows}, payloads_len={len(payloads)}, skip_no_date={skip_no_date}, bad_sheets={bad_sheets}")
        if payloads[:3]:
            for i, p in enumerate(payloads[:3]):
                print(f"[payload sample #{i+1}] {p}")

        # ───────────────────────── ★ 연도 불일치 검사 (ASP 동작 동일) ─────────────────────────
        # fiscalMM == '12' 인 경우에만 강제. 첫 유효 거래일 연도가 work_yy와 다르면 에러 반환.
        if str(fiscalMM) == '12':
            if first_dt_year is not None and str(first_dt_year) != str(work_yy):
                msg = f"작업연도와 분개장 연도가 다릅니다. (파일:{first_dt_year}, 작업:{work_yy})"
                print(f"[YEAR MISMATCH] {msg}")
                return JsonResponse({"ok": False, "msg": "작업연도와 분개장 연도가 다릅니다."}, status=400)

        # ───────────────────────── INSERT 실행 ─────────────────────────
        values_matrix = []
        for p in payloads:
            row = []
            for k in ins_keys:
                v = p.get(k, '')
                if k in ('tran_cr','tran_dr'): v = _normalize_number(v)
                if v is None: v = ''
                row.append(v)
            for c in extra_fill_cols:
                row.append(_default_for(c))
            values_matrix.append(tuple(row))

        if values_matrix[:2]:
            print(f"[values_matrix sample #1] {list(values_matrix[0])}")
            if len(values_matrix) > 1:
                print(f"[values_matrix sample #2] {list(values_matrix[1])}")

        placeholders = ", ".join(["%s"]*len(ins_cols))
        col_list     = ", ".join(f"[{c}]" for c in ins_cols)
        sql_insert   = f"INSERT INTO DS_SlipLedgr2 ({col_list}) VALUES ({placeholders})"
        print(f"[sql] {sql_insert}")
        print(f"[insert] rows={len(values_matrix)}, cols={len(ins_cols)}")

        with transaction.atomic():
            with connection.cursor() as cur:
                try:
                    if values_matrix:
                        cur.executemany(sql_insert, values_matrix)
                except Exception as ex:
                    import traceback
                    print(f"[insert ERROR] {type(ex).__name__}: {ex}\n{traceback.format_exc()}")
                    for j in range(min(5, len(values_matrix))):
                        print(f"[bad row sample #{j+1}] {values_matrix[j]}")
                    raise

        # ───────────────────────── 전기이월 프로시저 ─────────────────────────
        with connection.cursor() as cur:
            print(f"[proc] Exec up_Act_PreBSInquiry '{work_yy}','{seq_no}'")
            cur.execute("EXEC up_Act_PreBSInquiry %s, %s", [work_yy, seq_no])

        # ───────────────────────── 손익 집계 반영(tbl_corporate2 / tbl_income2) ─────────────────────────
        with connection.cursor() as cur:
            cur.execute("SELECT biz_type FROM mem_user WHERE seq_no=%s", [seq_no])
            row = cur.fetchone()
            biz_type = row[0] if row else None
            print(f"[biz_type] {biz_type}")

        tableName = "tbl_corporate2" if (biz_type is not None and int(biz_type) < 4) else "tbl_income2"
        print(f"[p&l table] {tableName} (YN_2=cost, YN_3=sales-cost)  sales={amt_sales}, cost={amt_cost}, profit={amt_sales-amt_cost}")

        with transaction.atomic():
            with connection.cursor() as cur:
                cur.execute(f"SELECT seq_no FROM {tableName} WHERE seq_no=%s AND work_YY=%s", [seq_no, work_yy])
                exists = cur.fetchone() is not None
                if exists:
                    sql_upd = f"""
                        UPDATE {tableName}
                        SET YN_2=%s, YN_3=%s
                        WHERE seq_no=%s AND work_YY=%s
                    """
                    print(f"[update {tableName}] YN_2, YN_3")
                    cur.execute(sql_upd, [amt_cost, (amt_sales-amt_cost), seq_no, work_yy])
                else:
                    # ✅ 먼저 파라미터 리스트를 만들고
                    params = [
                        int(seq_no),              # 1
                        int(work_yy),             # 2
                        amt_sales,                # 3
                        amt_cost,                 # 4
                        (amt_sales - amt_cost),   # 5
                        0, 0, 0, 0, 0,            # 6~10
                        '', '',                   # 11~12
                        0,                        # 13
                        '',                       # 14
                        0, 0, 0                   # 15~17
                    ]
                    # ✅ 파라미터 개수에 맞춰 placeholder 생성
                    placeholders_pl = ", ".join(["%s"] * len(params))
                    sql_ins = f"INSERT INTO {tableName} VALUES ({placeholders_pl})"
                    print(f"[insert {tableName}] sql={sql_ins}, params={params}")
                    cur.execute(sql_ins, params)

        # ───────────────────────── 월별 업로드 현황(tbl_mng_jaroe) ─────────────────────────
        with connection.cursor() as cur:
            cur.execute("""
                SELECT LEFT(tran_dt,2) AS work_mm
                  FROM DS_SlipLedgr2
                 WHERE seq_no=%s AND Work_YY=%s
                 GROUP BY LEFT(tran_dt,2)
                 ORDER BY LEFT(tran_dt,2)
            """, [seq_no, work_yy])
            months = [r[0] for r in cur.fetchall()]
        print(f"[months] {months}")

        with transaction.atomic():
            with connection.cursor() as cur:
                for mm in months:
                    cur.execute("""
                        SELECT work_mm FROM tbl_mng_jaroe
                         WHERE seq_no=%s AND work_YY=%s AND work_MM=%s
                    """, [seq_no, work_yy, int(mm)])
                    exists = cur.fetchone() is not None
                    if exists:
                        print(f"[jaroe update] seq={seq_no}, yy={work_yy}, mm={mm}")
                        cur.execute("""
                            UPDATE tbl_mng_jaroe
                               SET YN_5='1', YN_6='1', YN_7='1', YN_8='1', YN_9='1'
                             WHERE seq_no=%s AND work_YY=%s AND work_MM=%s
                        """, [seq_no, work_yy, int(mm)])
                    else:
                        print(f"[jaroe insert] seq={seq_no}, yy={work_yy}, mm={mm}")
                        flags = [('1' if 5 <= j <= 9 else '0') for j in range(1, 15)]
                        sql_ins = f"""
                            INSERT INTO tbl_mng_jaroe
                                (seq_no, work_YY, work_MM, {", ".join(f"YN_{k}" for k in range(1,15))}, bigo)
                            VALUES (%s,%s,%s,{",".join(["%s"]*14)},%s)
                        """
                        cur.execute(sql_ins, [seq_no, work_yy, int(mm), *flags, ''])

        resp = {
            "ok": True,
            "count": len(payloads),
            "last_tran_dt": last_tran_dt.isoformat() if last_tran_dt else None,
            "filename": os.path.basename(save_path),
            "amt_sales": amt_sales,
            "amt_cost" : amt_cost,
            "profit"   : (amt_sales - amt_cost),
        }
        if bad_sheets:
            resp["warn"] = f"헤더 인식 실패 시트: {', '.join(sorted(bad_sheets))}"
        print(f"[DONE] rows={resp['count']}  sales={amt_sales}  cost={amt_cost}  profit={resp['profit']}")
        return JsonResponse(resp)

    except Exception as ex:
        import traceback
        tb = traceback.format_exc()
        print(f"[EXCEPTION] {type(ex).__name__} : {ex}\n{tb}")
        return JsonResponse({"ok": False, "msg": f"{type(ex).__name__}: {ex}"}, status=500)



def getCompanyInfo(request):
    """
    회사정보 섹션 전용 JSON API
    - flag:
        EXECS         : 임원 등기현황
        WORKERS       : 직원현황(인원)
        SALARYS       : 급여(금액)
        STOCKHOLDERS  : 주주현황
        SUMMARY       : (선택) 간단 회사 요약 등 확장 포인트
    """
    seq_no   = _to_int(request.GET.get("seq_no"), 0)
    flag     = (request.GET.get("flag") or "SUMMARY").upper()
    work_yy  = _to_int(request.GET.get("work_yy") or request.GET.get("work_YY")) \
               or timezone.localtime().year
    fiscalMM = _to_int(request.GET.get("fiscalMM") or request.GET.get("FiscalMM"), 12)

    memuser = MemUser.objects.filter(seq_no=seq_no).only("seq_no","biz_type","biz_name","biz_no").first()
    memdeal = MemDeal.objects.get(seq_no=seq_no)

    user_img_url = ""

    qs = userProfile.objects.filter(title=memuser.seq_no) \
            .exclude(image__isnull=True).exclude(image="")

    obj = qs.order_by('-description', '-id').first()   # 최신 1건

    if obj and getattr(obj, 'image', None):
        try:
            # 절대 URL로 변환 (iframe/인쇄에서도 안전)
            user_img_url = request.build_absolute_uri(obj.image.url)
        except Exception:
            user_img_url = ""

    if not memuser:
        return JsonResponse({"ok": False, "error": "사용자 정보가 없습니다."}, status=400)

    # ─────────────────────────────────────────
    # 1) 임원 등기현황
    # ─────────────────────────────────────────
    if flag == "EXECS":
        sql = r"""
        SELECT
            d.execflag, d.execName, d.regDate AS exec_regDate,
            ISNULL(d.extentDate,'') AS extentDate,
            CASE WHEN d.execflag = N'감사'
                 THEN CONVERT(char(4), CONVERT(int, LEFT(CASE WHEN d.extentDate<>'' THEN d.extentDate ELSE d.regDate END,4))+3) + N'-03-31'
                 ELSE CONVERT(char(4), CONVERT(int, LEFT(CASE WHEN d.extentDate<>'' THEN d.extentDate ELSE d.regDate END,4))+3)
                      + N'-' + RIGHT(CASE WHEN d.extentDate<>'' THEN d.extentDate ELSE d.regDate END,5)
            END AS duedate,
            CASE d.execflag WHEN N'감사' THEN 'warning' WHEN N'대표이사' THEN 'primary' ELSE 'success' END AS colorProgress,
            -- totalDD
            CASE WHEN d.execflag = N'감사' THEN
                DATEDIFF(DAY, LEFT(CASE WHEN d.extentDate<>'' THEN d.extentDate ELSE d.regDate END,10),
                              CONVERT(char(4), CONVERT(int, LEFT(CASE WHEN d.extentDate<>'' THEN d.extentDate ELSE d.regDate END,4))+3) + N'-03-31')
            ELSE
                DATEDIFF(DAY, LEFT(CASE WHEN d.extentDate<>'' THEN d.extentDate ELSE d.regDate END,10),
                              CONVERT(char(4), CONVERT(int, LEFT(CASE WHEN d.extentDate<>'' THEN d.extentDate ELSE d.regDate END,4))+3)
                              + N'-' + RIGHT(CASE WHEN d.extentDate<>'' THEN d.extentDate ELSE d.regDate END,5))
            END AS totalDD,
            -- passDD
            CASE WHEN d.extentDate<>'' THEN DATEDIFF(DAY, d.extentDate, GETDATE())
                 ELSE DATEDIFF(DAY, d.regDate, GETDATE()) END AS passDD
        FROM Mem_User a
        JOIN mem_deal b ON a.seq_no=b.seq_no
        JOIN lawregistration d ON a.seq_no=d.seq_no
        WHERE a.seq_no=%s
          AND b.keeping_YN='Y'
          AND a.biz_type IN ('1','2','3')
          AND d.execflag IN (N'대표이사', N'사내이사', N'감사')
          AND ISNULL(d.fireDate,'')=''
        ORDER BY d.regDate;
        """
        rows = []
        with connection.cursor() as cur:
            cur.execute(sql, [seq_no])
            for execflag, execName, exec_regDate, extentDate, duedate, colorProgress, totalDD, passDD in cur.fetchall():
                totalDD = int(totalDD or 0)
                passDD  = int(passDD or 0)
                passedRate = max(0, min(100, int(round((passDD/totalDD*100) if totalDD>0 else 0))))
                remainDD = max(0, totalDD - passDD)
                txtPass = "경과" if passDD > totalDD else "미경과"
                badgeColorPass = "danger" if txtPass=="경과" else "light"
                rows.append({
                    "execflag": execflag,
                    "execName": execName,
                    "exec_regDate": exec_regDate,
                    "extentDate": extentDate,
                    "duedate": duedate,
                    "colorProgress": colorProgress,
                    "txtPass": txtPass,
                    "passedRate": passedRate,
                    "remainDD": remainDD,
                    "badgeColorPass": badgeColorPass,
                })
        return JsonResponse({"ok": True, "execs": rows})
    # ─────────────────────────────────────────
    # 직원현황공통 유틸
    MONTH_KEYS = ["JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"]

    def _spread_halfyear_amount(records):
        """
        records: [(mm_code, value), ...]
        예) 과세년도 2025-03 → mm_code='03' → 3월에만 value 저장
            과세년도 2025-07 → mm_code='07' → 7월에만 value 저장
        금액은 해당 월에만 넣고 나머지 월은 0 (value는 만원 단위라고 가정).
        """
        monthly = {m: 0.0 for m in range(1, 13)}

        for mm_code, val in records:
            if mm_code is None:
                continue
            try:
                mm = int(str(mm_code)[-2:])
            except ValueError:
                continue
            if 1 <= mm <= 12:
                monthly[mm] += float(val or 0)

        # 소수점 이하는 절사
        for m in monthly:
            monthly[m] = int(monthly[m])

        return monthly
    def _spread_halfyear_count(records):
        """
        records: [(mm_code, cnt), ...]
        제출건수(인원수)를 해당 월에만 그대로 사용.
        예) 제출건수 1명, 과세년도 2025-03 → 3월만 1명
            제출건수 2명, 과세년도 2025-07 → 7월만 2명
        """
        monthly = {m: 0.0 for m in range(1, 13)}

        for mm_code, cnt in records:
            if mm_code is None:
                continue
            try:
                mm = int(str(mm_code)[-2:])
            except ValueError:
                continue
            if 1 <= mm <= 12:
                monthly[mm] += float(cnt or 0)

        for m in monthly:
            monthly[m] = int(monthly[m])

        return monthly
    def _build_worker_row(title, color, monthly_dict, half_based=False):
        """
        half_based=True  → 값이 있는 '월 개수'로 평균 내기 (지급조서 기반 반기용)
        half_based=False → 1월~마지막 발생월까지 개월수로 평균 (정규직 인원용)
        """
        vals = {m: int(float(monthly_dict.get(m, 0) or 0)) for m in range(1, 13)}
        total = sum(vals.values())

        if half_based:
            months_count = sum(1 for v in vals.values() if v != 0)
        else:
            last_month = max((m for m, v in vals.items() if v != 0), default=0)
            months_count = last_month if last_month > 0 else 0

        avg = int(total / months_count) if months_count else 0

        row = {"TITLE": title, "COLOR": color, "TOT": avg}
        for idx, key in enumerate(MONTH_KEYS, start=1):
            row[key] = vals[idx]
        return row
    def _build_money_row(title, color, monthly_dict):
        vals = {m: int(float(monthly_dict.get(m, 0) or 0)) for m in range(1, 13)}
        total = sum(vals.values())
        row = {"TITLE": title, "COLOR": color, "TOT": int(total)}
        for idx, key in enumerate(MONTH_KEYS, start=1):
            row[key] = vals[idx]
        return row    # 2) 직원현황(인원): workers
    # ─────────────────────────────────────────
    if flag == "WORKERS":
        is_half = memdeal.goyoung_banki == "Y"

        rows = []

        if is_half:
            # ===== 반기 사업장 (goyoung_banki='Y') =====
            with connection.cursor() as cur:
                # 1) 정규직 인원: 급여지급현황
                cur.execute("""
                    SELECT work_mm, COUNT(DISTINCT empNo) AS cnt
                    FROM 급여지급현황
                    WHERE seq_no = %s
                      AND work_yy = %s
                    GROUP BY work_mm
                """, [seq_no, work_yy])
                reg_monthly = {int(mm): float(cnt) for mm, cnt in cur.fetchall()}

                base_where = """
                    FROM 지급조서간이소득
                    WHERE 사업자번호 = %s
                    AND LEFT(접수일시,4) = %s
                    AND LEFT(과세년도,4) = %s
                """
                base_params = [memuser.biz_no, str(work_yy), str(work_yy)]

                # === 사업소득자 인원 ===
                sql_biz = f"""
                    SELECT RIGHT(과세년도,2) AS mm_code,
                        SUM(
                            CASE WHEN ISNUMERIC(제출건수) = 1
                                THEN CONVERT(int, 제출건수)
                                ELSE 0 END
                        ) AS cnt
                    {base_where}
                    AND 신고서종류 = N'간이지급명세서(거주자의 사업소득)'
                    GROUP BY RIGHT(과세년도,2)
                """
                cur.execute(sql_biz, base_params)
                biz_records = cur.fetchall()
                biz_monthly = _spread_halfyear_count(biz_records)

                # === 기타소득자 인원 ===
                sql_etc = f"""
                    SELECT RIGHT(과세년도,2) AS mm_code,
                        SUM(
                            CASE WHEN ISNUMERIC(제출건수) = 1
                                THEN CONVERT(int, 제출건수)
                                ELSE 0 END
                        ) AS cnt
                    {base_where}
                    AND 신고서종류 = N'간이지급명세서(거주자의 기소득)'
                    GROUP BY RIGHT(과세년도,2)
                """
                cur.execute(sql_etc, base_params)
                etc_records = cur.fetchall()
                etc_monthly = _spread_halfyear_count(etc_records)

                # === 일용직 인원 ===
                sql_daily = f"""
                    SELECT RIGHT(과세년도,2) AS mm_code,
                        SUM(
                            CASE WHEN ISNUMERIC(제출건수) = 1
                                THEN CONVERT(int, 제출건수)
                                ELSE 0 END
                        ) AS cnt
                    {base_where}
                    AND 신고서종류 = N'일용근로소득 지급명세서'
                    GROUP BY RIGHT(과세년도,2)
                """
                cur.execute(sql_daily, base_params)
                daily_records = cur.fetchall()
                daily_monthly = _spread_halfyear_count(daily_records)

            rows = [
                _build_worker_row("정규직",   "blue",    reg_monthly, False),  # half_based=False
                _build_worker_row("사업소득", "primary", biz_monthly, True),   # half_based=True
                _build_worker_row("기타소득", "danger",  etc_monthly, True),
                _build_worker_row("일용직",   "warning", daily_monthly, True),
            ]
            order_map = {"정규직": 1, "사업소득": 2, "기타소득": 3, "일용직": 4}
            rows.sort(key=lambda r: order_map.get(r["TITLE"], 9))
            return JsonResponse({"ok": True, "workers": rows})

        else:
            # ===== 기존 월별 원천세 로직 (질문에 주신 SQL) =====
            sql = f"""
            SELECT * FROM (
                SELECT N'정규직' AS TITLE,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='01' THEN a01m END),0) AS JAN,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='02' THEN a01m END),0) AS FEB,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='03' THEN a01m END),0) AS MAR,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='04' THEN a01m END),0) AS APR,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='05' THEN a01m END),0) AS MAY,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='06' THEN a01m END),0) AS JUN,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='07' THEN a01m END),0) AS JUL,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='08' THEN a01m END),0) AS AUG,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='09' THEN a01m END),0) AS SEP,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='10' THEN a01m END),0) AS OCT,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='11' THEN a01m END),0) AS NOV,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='12' THEN a01m END),0) AS [DEC],
                  CASE WHEN (SELECT CONVERT(int, RIGHT(MAX(과세연월),2))
                             FROM 원천세전자신고
                             WHERE LEFT(과세연월,4)=%s AND 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE User_ID=%s)) = 0
                       THEN 0
                       ELSE ISNULL(SUM(a01m),0) /
                            NULLIF((SELECT CONVERT(int, RIGHT(MAX(과세연월),2))
                                    FROM 원천세전자신고
                                    WHERE LEFT(과세연월,4)=%s AND 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE User_ID=%s)), 0)
                  END AS TOT,
                  'blue' AS COLOR
                FROM 원천세전자신고
                WHERE 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE user_id=%s) AND LEFT(과세연월,4)=%s
                GROUP BY LEFT(과세연월,4)
                UNION ALL
                SELECT N'사업소득',
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='01' THEN a30m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='02' THEN a30m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='03' THEN a30m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='04' THEN a30m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='05' THEN a30m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='06' THEN a30m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='07' THEN a30m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='08' THEN a30m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='09' THEN a30m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='10' THEN a30m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='11' THEN a30m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='12' THEN a30m END),0),
                  CASE WHEN (SELECT CONVERT(int, RIGHT(MAX(과세연월),2)) FROM 원천세전자신고
                             WHERE LEFT(과세연월,4)=%s AND 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE User_ID=%s)) = 0
                       THEN 0
                       ELSE ISNULL(SUM(a30m),0) /
                            NULLIF((SELECT CONVERT(int, RIGHT(MAX(과세연월),2)) FROM 원천세전자신고
                                    WHERE LEFT(과세연월,4)=%s AND 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE User_ID=%s)), 0)
                  END,
                  'primary'
                FROM 원천세전자신고
                WHERE 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE user_id=%s) AND LEFT(과세연월,4)=%s
                GROUP BY LEFT(과세연월,4)
                UNION ALL
                SELECT N'기타소득',
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='01' THEN a40m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='02' THEN a40m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='03' THEN a40m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='04' THEN a40m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='05' THEN a40m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='06' THEN a40m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='07' THEN a40m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='08' THEN a40m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='09' THEN a40m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='10' THEN a40m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='11' THEN a40m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='12' THEN a40m END),0),
                  CASE WHEN (SELECT CONVERT(int, RIGHT(MAX(과세연월),2)) FROM 원천세전자신고
                             WHERE LEFT(과세연월,4)=%s AND 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE User_ID=%s)) = 0
                       THEN 0 ELSE ISNULL(SUM(a40m),0) /
                            NULLIF((SELECT CONVERT(int, RIGHT(MAX(과세연월),2)) FROM 원천세전자신고
                                    WHERE LEFT(과세연월,4)=%s AND 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE User_ID=%s)), 0)
                  END,
                  'danger'
                FROM 원천세전자신고
                WHERE 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE user_id=%s) AND LEFT(과세연월,4)=%s
                GROUP BY LEFT(과세연월,4)
                UNION ALL
                SELECT N'일용직',
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='01' THEN a03m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='02' THEN a03m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='03' THEN a03m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='04' THEN a03m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='05' THEN a03m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='06' THEN a03m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='07' THEN a03m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='08' THEN a03m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='09' THEN a03m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='10' THEN a03m END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='11' THEN a03m END),0), ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='12' THEN a03m END),0),
                  CASE WHEN (SELECT CONVERT(int, RIGHT(MAX(과세연월),2)) FROM 원천세전자신고
                             WHERE LEFT(과세연월,4)=%s AND 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE User_ID=%s)) = 0
                       THEN 0 ELSE ISNULL(SUM(a03m),0) /
                            NULLIF((SELECT CONVERT(int, RIGHT(MAX(과세연월),2)) FROM 원천세전자신고
                                    WHERE LEFT(과세연월,4)=%s AND 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE User_ID=%s)), 0)
                  END,
                  'warning'
                FROM 원천세전자신고
                WHERE 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE user_id=%s) AND LEFT(과세연월,4)=%s
                GROUP BY LEFT(과세연월,4)
            ) t
            ORDER BY CASE TITLE WHEN N'정규직' THEN 1 WHEN N'사업소득' THEN 2 WHEN N'기타소득' THEN 3 WHEN N'일용직' THEN 4 ELSE 9 END;
            """
            p = [str(work_yy), memuser.user_id, str(work_yy), memuser.user_id,
                 memuser.user_id, str(work_yy),
                 str(work_yy), memuser.user_id, str(work_yy), memuser.user_id, memuser.user_id, str(work_yy),
                 str(work_yy), memuser.user_id, str(work_yy), memuser.user_id, memuser.user_id, str(work_yy),
                 str(work_yy), memuser.user_id, str(work_yy), memuser.user_id, memuser.user_id, str(work_yy)]
            with connection.cursor() as cur:
                cur.execute(sql, p)
                for r in cur.fetchall():
                    rows.append({
                        'TITLE': r[0], 'JAN': r[1], 'FEB': r[2], 'MAR': r[3], 'APR': r[4], 'MAY': r[5], 'JUN': r[6],
                        'JUL': r[7], 'AUG': r[8], 'SEP': r[9], 'OCT': r[10], 'NOV': r[11], 'DEC': r[12],
                        'TOT': round(float(r[13] or 0), 2), 'COLOR': r[14],
                    })
            return JsonResponse({"ok": True, "workers": rows})

    # ─────────────────────────────────────────
    # 3) 급여(금액): salarys
    # ─────────────────────────────────────────
    if flag == "SALARYS":
        is_half = memdeal.goyoung_banki == "Y"
        rows = []

        if is_half:
            # ===== 반기 사업장 (goyoung_banki='Y') =====
            with connection.cursor() as cur:
                # 1) 정규직급여: 급여지급현황 지급총액 (원 → 만원)
                cur.execute("""
                    SELECT work_mm, SUM(지급총액) AS amt
                    FROM 급여지급현황
                    WHERE seq_no = %s
                      AND work_yy = %s
                    GROUP BY work_mm
                """, [seq_no, work_yy])
                reg_monthly = {int(mm): float(amt) / 10000.0 for mm, amt in cur.fetchall()}

                base_where = """
                    FROM 지급조서간이소득
                    WHERE 사업자번호 = %s
                      AND LEFT(접수일시,4) = %s
                      AND LEFT(과세년도,4) = %s
                """
                base_params = [memuser.biz_no, str(work_yy), str(work_yy)]

                # 2) 사업소득급여: 제출금액 (char → numeric) 반기→월별
                sql_biz = f"""
                    SELECT RIGHT(과세년도,2) AS mm_code,
                           SUM(
                             CASE WHEN ISNUMERIC(제출금액) = 1
                                  THEN CONVERT(NUMERIC(18,0), 제출금액)
                                  ELSE 0 END
                           ) AS amt
                    {base_where}
                      AND 신고서종류 LIKE '%%거주자의 사업소득%%'
                    GROUP BY RIGHT(과세년도,2)
                """
                cur.execute(sql_biz, base_params)
                biz_records = [(mm, float(amt) / 10000.0) for mm, amt in cur.fetchall()]
                biz_monthly = _spread_halfyear_amount(biz_records)

                # 3) 기타소득급여
                sql_etc = f"""
                    SELECT RIGHT(과세년도,2) AS mm_code,
                           SUM(
                             CASE WHEN ISNUMERIC(제출금액) = 1
                                  THEN CONVERT(NUMERIC(18,0), 제출금액)
                                  ELSE 0 END
                           ) AS amt
                    {base_where}
                      AND 신고서종류 = N'간이지급명세서(거주자의 기소득)'
                    GROUP BY RIGHT(과세년도,2)
                """
                cur.execute(sql_etc, base_params)
                etc_records = [(mm, float(amt) / 10000.0) for mm, amt in cur.fetchall()]
                etc_monthly = _spread_halfyear_amount(etc_records)

                # 4) 일용직급여
                sql_daily = f"""
                    SELECT RIGHT(과세년도,2) AS mm_code,
                           SUM(
                             CASE WHEN ISNUMERIC(제출금액) = 1
                                  THEN CONVERT(NUMERIC(18,0), 제출금액)
                                  ELSE 0 END
                           ) AS amt
                    {base_where}
                      AND 신고서종류 = N'일용근로소득 지급명세서'
                    GROUP BY RIGHT(과세년도,2)
                """
                cur.execute(sql_daily, base_params)
                daily_records = [(mm, float(amt) / 10000.0) for mm, amt in cur.fetchall()]
                daily_monthly = _spread_halfyear_amount(daily_records)

            rows = [
                _build_money_row("정규직급여",   "blue",    reg_monthly),
                _build_money_row("사업소득급여", "primary", biz_monthly),
                _build_money_row("기타소득급여", "danger",  etc_monthly),
                _build_money_row("일용직급여",   "warning", daily_monthly),
            ]
            order_map = {"정규직급여": 1, "사업소득급여": 2, "기타소득급여": 3, "일용직급여": 4}
            rows.sort(key=lambda r: order_map.get(r["TITLE"], 9))
            return JsonResponse({"ok": True, "salarys": rows})

        else:
            # ───────── 기존 SALARYS 쿼리 그대로 ─────────
            sql = r"""
            SELECT * FROM (
                -- 1) 정규직급여 (a01)
                SELECT N'정규직급여' AS TITLE,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='01' THEN ROUND(a01/10000,0) END),0) AS JAN,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='02' THEN ROUND(a01/10000,0) END),0) AS FEB,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='03' THEN ROUND(a01/10000,0) END),0) AS MAR,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='04' THEN ROUND(a01/10000,0) END),0) AS APR,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='05' THEN ROUND(a01/10000,0) END),0) AS MAY,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='06' THEN ROUND(a01/10000,0) END),0) AS JUN,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='07' THEN ROUND(a01/10000,0) END),0) AS JUL,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='08' THEN ROUND(a01/10000,0) END),0) AS AUG,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='09' THEN ROUND(a01/10000,0) END),0) AS SEP,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='10' THEN ROUND(a01/10000,0) END),0) AS OCT,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='11' THEN ROUND(a01/10000,0) END),0) AS NOV,
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='12' THEN ROUND(a01/10000,0) END),0) AS [DEC],
                  ISNULL(SUM(ROUND(a01/10000,0)),0) AS TOT,
                  'blue' AS COLOR
                FROM 원천세전자신고
                WHERE 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE user_id=%s)
                  AND LEFT(과세연월,4)=%s
                GROUP BY LEFT(과세연월,4)

                UNION ALL

                -- 2) 사업소득급여 (a30)
                SELECT N'사업소득급여',
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='01' THEN ROUND(a30/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='02' THEN ROUND(a30/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='03' THEN ROUND(a30/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='04' THEN ROUND(a30/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='05' THEN ROUND(a30/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='06' THEN ROUND(a30/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='07' THEN ROUND(a30/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='08' THEN ROUND(a30/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='09' THEN ROUND(a30/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='10' THEN ROUND(a30/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='11' THEN ROUND(a30/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='12' THEN ROUND(a30/10000,0) END),0),
                  ISNULL(SUM(ROUND(a30/10000,0)),0),
                  'primary'
                FROM 원천세전자신고
                WHERE 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE user_id=%s)
                  AND LEFT(과세연월,4)=%s
                GROUP BY LEFT(과세연월,4)

                UNION ALL

                -- 3) 기타소득급여 (a40)
                SELECT N'기타소득급여',
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='01' THEN ROUND(a40/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='02' THEN ROUND(a40/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='03' THEN ROUND(a40/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='04' THEN ROUND(a40/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='05' THEN ROUND(a40/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='06' THEN ROUND(a40/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='07' THEN ROUND(a40/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='08' THEN ROUND(a40/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='09' THEN ROUND(a40/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='10' THEN ROUND(a40/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='11' THEN ROUND(a40/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='12' THEN ROUND(a40/10000,0) END),0),
                  ISNULL(SUM(ROUND(a40/10000,0)),0),
                  'danger'
                FROM 원천세전자신고
                WHERE 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE user_id=%s)
                  AND LEFT(과세연월,4)=%s
                GROUP BY LEFT(과세연월,4)

                UNION ALL

                -- 4) 일용직급여 (a03)
                SELECT N'일용직급여',
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='01' THEN ROUND(a03/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='02' THEN ROUND(a03/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='03' THEN ROUND(a03/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='04' THEN ROUND(a03/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='05' THEN ROUND(a03/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='06' THEN ROUND(a03/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='07' THEN ROUND(a03/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='08' THEN ROUND(a03/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='09' THEN ROUND(a03/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='10' THEN ROUND(a03/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='11' THEN ROUND(a03/10000,0) END),0),
                  ISNULL(MAX(CASE WHEN RIGHT(과세연월,2)='12' THEN ROUND(a03/10000,0) END),0),
                  ISNULL(SUM(CONVERT(NUMERIC(13,0), ROUND(a03/10000,0))),0),
                  'warning'
                FROM 원천세전자신고
                WHERE 사업자등록번호=(SELECT Biz_No FROM mem_user WHERE user_id=%s)
                  AND LEFT(과세연월,4)=%s
                GROUP BY LEFT(과세연월,4)
            ) t
            ORDER BY CASE TITLE
                WHEN N'정규직급여' THEN 1
                WHEN N'사업소득급여' THEN 2
                WHEN N'기타소득급여' THEN 3
                WHEN N'일용직급여' THEN 4
                ELSE 9 END;
            """
            params = [
                memuser.user_id, str(work_yy),
                memuser.user_id, str(work_yy),
                memuser.user_id, str(work_yy),
                memuser.user_id, str(work_yy),
            ]
            with connection.cursor() as cur:
                cur.execute(sql, params)
                for r in cur.fetchall():
                    rows.append({
                        "TITLE": r[0],
                        "JAN":   r[1],  "FEB": r[2],  "MAR": r[3],  "APR": r[4],
                        "MAY":   r[5],  "JUN": r[6],  "JUL": r[7],  "AUG": r[8],
                        "SEP":   r[9],  "OCT": r[10], "NOV": r[11], "DEC": r[12],
                        "TOT":   r[13],
                        "COLOR": r[14],
                    })
            return JsonResponse({"ok": True, "salarys": rows})
    # ─────────────────────────────────────────
    # 4) 주주현황: stockHolders
    # ─────────────────────────────────────────
    if flag == "STOCKHOLDERS":
        now_str = timezone.localtime().strftime('%Y-%m-%d')
        sql = r"""
        SELECT
            B.StckH_Num,
            MAX(B.StckH_Nm),
            CASE B.StckH_RS
                WHEN 0 THEN N'지배주주' WHEN 1 THEN N'배우자' WHEN 2 THEN N'자녀' WHEN 3 THEN N'부모'
                WHEN 4 THEN N'형제자매' WHEN 5 THEN N'손자' WHEN 6 THEN N'조부모' WHEN 7 THEN N'친족의 배우자'
                WHEN 8 THEN N'기타 친족' WHEN 9 THEN N'기타' WHEN 10 THEN N'특수관계법인' ELSE N''
            END,
            MIN(A.tran_Dt),
            SUM(CASE WHEN A.StckH_TranGB='B' THEN A.StckH_FEquityNum*-1 ELSE A.StckH_FEquityNum END),
            MAX(A.StckH_FEquityFP),
            SUM(CASE WHEN A.StckH_TranGB='B' THEN A.StckH_FEquityGP*-1 ELSE A.StckH_FEquityGP END),
            SUM(CASE WHEN A.StckH_TranGB='B' THEN A.StckH_FEquityNum*-1 ELSE A.StckH_FEquityNum END) * 100.0 /
            (
              SELECT SUM(CASE WHEN D.StckH_TranGB='B' THEN D.StckH_FEquityNum*-1 ELSE D.StckH_FEquityNum END)
              FROM Tbl_StckHolderList C WITH(NOLOCK)
              JOIN Tbl_StckHListTrn D WITH(NOLOCK) ON C.Seq_No=D.Seq_No AND C.StckH_Num=D.StckH_Num
              WHERE C.Seq_No=%s AND D.TRAN_DT<=%s
            )
        FROM Tbl_StckHolderList B WITH(NOLOCK)
        JOIN Tbl_StckHListTrn A WITH(NOLOCK) ON B.Seq_No=A.Seq_No AND B.StckH_Num=A.StckH_Num
        WHERE B.Seq_No=%s AND A.TRAN_DT<=%s
        GROUP BY B.StckH_Num, B.StckH_RS
        HAVING SUM(CASE WHEN A.StckH_TranGB='B' THEN A.StckH_FEquityNum*-1 ELSE A.StckH_FEquityNum END) > 0
        ORDER BY 5 DESC;
        """
        rows, total = [], 0
        with connection.cursor() as cur:
            cur.execute(sql, [seq_no, now_str, seq_no, now_str])
            for sthNum, sthName, sthRelation, sthGetDate, sthCnt, sthFaceValue, sthTotalValue, sthRate in cur.fetchall():
                colorSpec = "warning" if sthRelation == "기타" else "info"
                try: rate = round(float(sthRate), 2)
                except Exception: rate = 0.0
                rate = max(0.0, min(100.0, rate))
                rows.append({
                    "sthName": sthName, "sthRelation": sthRelation, "sthGetDate": sthGetDate,
                    "sthCnt": int(sthCnt or 0), "sthFaceValue": sthFaceValue,
                    "sthTotalValue": int(sthTotalValue or 0), "sthRate": rate, "colorSpec": colorSpec
                })
                total += int(sthTotalValue or 0)
        return JsonResponse({"ok": True, "stockHolders": rows, "stockTotal": total})
    # ─────────────────────────────────────────
    # 5) (개인) 사업용계좌 개설현황
    # ─────────────────────────────────────────
            
    def _s(v):
        if v is None:
            return ""
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        return str(v).strip()
    if flag == "BIZ_ACCOUNTS":
      sql = r"""
          SELECT 등록번호, 납세계좌구분, 은행명, 계좌번호, 등록일자
          FROM 사업용계좌신고현황 WITH(NOLOCK)
          WHERE Seq_No = %s
          ORDER BY 등록번호 DESC
      """
      rows = []
      with connection.cursor() as cur:
          cur.execute(sql, [seq_no])
          for 등록번호, 납세계좌구분, 은행명, 계좌번호, 등록일자 in cur.fetchall():
              rows.append({
                  "reg_no":  _s(등록번호),
                  "acc_type": _s(납세계좌구분),
                  "bank":    _s(은행명),
                  "acct":    _s(계좌번호),
                  "reg_dt":  _s(등록일자),   # ← 날짜 안전 처리
              })
      return JsonResponse({"ok": True, "total_cnt": len(rows), "rows": rows})
    if flag == "CASH":
        # 1) 최신 가맹점 가입의무 현황 (TOP 1)
        sql_duty = r"""
            SELECT TOP 1 기준연도, 가입의무대상, 가입의무기한시작일자, 가입기한, 가맹일자
            FROM 현금영수증가맹점가입의무현황 WITH(NOLOCK)
            WHERE Seq_No=%s
            ORDER BY 기준연도 DESC
        """
        
        cash_data = {
            "year": "",
            "is_target": False, # 의무대상 여부
            "target_txt": "비대상",
            "join_date": "-",
            "limit_date": "-",
            "refuse": None,     # 발급거부 내역
            "unissued": None    # 미발급 내역
        }

        with connection.cursor() as cur:
            cur.execute(sql_duty, [seq_no])
            row = cur.fetchone()
            if row:
                기준연도, 가입의무대상, 기한시작, 가입기한, 가맹일자 = row
                cash_data["year"] = _s(기준연도)
                cash_data["target_txt"] = _s(가입의무대상)
                cash_data["is_target"] = (True if _s(가입의무대상) == "대상" else False)
                cash_data["join_date"] = _s(가맹일자).replace("1900-01-01", "")
                cash_data["limit_date"] = _s(가입기한).replace("1900-01-01", "")

        # 2) 가산세 조회 (의무대상일 경우에만 조회하거나, 데이터 있으면 표시)
        #    발급거부 / 미발급 각각 조회
        if cash_data["is_target"]:
            sql_penalty = r"""
                SELECT TOP 1 가산세사유, 거래일자, 가산세, 확정일
                FROM 가산세내역 WITH(NOLOCK)
                WHERE Seq_No=%s AND YEAR(거래일자)=%s AND 가산세사유=%s
                ORDER BY 거래일자 DESC
            """
            with connection.cursor() as cur:
                # 발급거부
                cur.execute(sql_penalty, [seq_no, work_yy, '발급거부'])
                p_row = cur.fetchone()
                if p_row:
                    cash_data["refuse"] = {
                        "date": _s(p_row[1]),
                        "amt": _to_int(p_row[2], 0)
                    }
                
                # 미발급
                cur.execute(sql_penalty, [seq_no, work_yy, '미발급'])
                p_row = cur.fetchone()
                if p_row:
                    cash_data["unissued"] = {
                        "date": _s(p_row[1]),
                        "amt": _to_int(p_row[2], 0)
                    }

        return JsonResponse({
            "ok": True,
            "data": cash_data
        })
    if flag == "CARDS":
      sql = r"""
          SELECT MAX(CrcmClNm) AS 카드사,
                busnCrdCardNoEncCntn AS 카드번호,
                COUNT(*) AS 건수,
                SUM(totaTrsAmt) AS 사용금액
          FROM tbl_hometax_scrap WITH(NOLOCK)
          WHERE seq_no=%s AND Tran_YY=%s
          GROUP BY busnCrdCardNoEncCntn
          ORDER BY SUM(totaTrsAmt) DESC
      """
      rows = []
      with connection.cursor() as cur:
          cur.execute(sql, [seq_no, work_yy])
          for 카드사, 카드번호, 건수, 사용금액 in cur.fetchall():
              rows.append({
                  "card_co": (카드사 or "").strip(),
                  "card_no": (카드번호 or "").strip(),
                  "use_cnt": int(건수 or 0),
                  "use_amt": int(사용금액 or 0),
              })

      return JsonResponse({
          "ok": True,
          "cards": {
              "year": work_yy,
              "total_cnt": len(rows),
              "rows": rows,   # ← 패딩 없이 그대로 전달
          }
      })
    
    # ─────────────────────────────────────────
    # 7) 기본 (요약)
    # ─────────────────────────────────────────
    user_pwd = memuser.user_pwd
    if len(user_pwd) > 4:
        user_pwd =  user_pwd[:-4] + "****"
    return JsonResponse({
        "ok": True,
        "summary": {
            "biz_name": memuser.biz_name,
            "ceo_name": memuser.ceo_name,
            "biz_no": memuser.biz_no,
            "regDate": f"{memuser.reg_date.year}년 {memuser.reg_date.month}월 {memuser.reg_date.day}일",
            "createdDate": f"{memdeal.createddate.year}년 {memdeal.createddate.month}월 {memdeal.createddate.day}일",
            "uptae":memuser.uptae,
            "jongmok":memuser.jongmok,
            "userID":memuser.user_id,
            "userPW":user_pwd,
            "fiscalMM": fiscalMM,
            "isrnd":memuser.isrnd,
            "isventure":memuser.isventure,
            "userImg": user_img_url,
            "addr":memuser.biz_addr1 +" "+memuser.biz_addr2
        }
    })


@require_GET
def diag_capital_list(request):
    seq_no = request.GET.get('seq_no')
    if not seq_no:
        return HttpResponseBadRequest('missing seq_no')
    with connection.cursor() as cur:
        cur.execute("""
            SELECT Seq_No, MH_Name, MH_Amt, MH_DcRate
            FROM Diag_capital
            WHERE Seq_No = %s
              AND (CASE WHEN ISNUMERIC(CAST(MH_DcRate AS VARCHAR(10)))=1 THEN CAST(MH_DcRate AS DECIMAL(18,6)) ELSE 9 END) < 2
            ORDER BY MH_Amt DESC, MH_Name ASC
        """, [seq_no])
        rows = _dictfetchall(cur)
    # 그대로 반환(프론트가 동일 키로 사용)
    return JsonResponse(rows, safe=False)

@require_POST
def diag_capital_upsert(request):
    try:
        data = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return HttpResponseBadRequest('invalid json')

    seq_no   = data.get('Seq_No')
    mh_name  = (data.get('MH_Name') or '').strip()
    mh_amt   = data.get('MH_Amt') or 0
    dc_rate  = data.get('MH_DcRate') or 1  # 1 or 0.5

    if not seq_no or not mh_name:
        return HttpResponseBadRequest('missing Seq_No or MH_Name')

    with transaction.atomic():
        with connection.cursor() as cur:
            # 존재 여부
            cur.execute("""
                SELECT COUNT(*) FROM Diag_capital
                WHERE Seq_No=%s AND MH_Name=%s
            """, [seq_no, mh_name])
            exists = cur.fetchone()[0] > 0

            if exists:
                cur.execute("""
                    UPDATE Diag_capital
                    SET MH_Amt=%s, MH_DcRate=%s
                    WHERE Seq_No=%s AND MH_Name=%s
                """, [mh_amt, dc_rate, seq_no, mh_name])
            else:
                cur.execute("""
                    INSERT INTO Diag_capital(Seq_No, MH_Name, MH_Amt, MH_DcRate)
                    VALUES(%s, %s, %s, %s)
                """, [seq_no, mh_name, mh_amt, dc_rate])

    return JsonResponse({'ok': True, 'Seq_No': seq_no, 'MH_Name': mh_name})

@require_POST
def diag_capital_delete(request):
    try:
        data = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return HttpResponseBadRequest('invalid json')

    seq_no  = data.get('Seq_No')
    mh_name = (data.get('MH_Name') or '').strip()
    if not seq_no or not mh_name:
        return HttpResponseBadRequest('missing Seq_No or MH_Name')

    # 디버그 로그
    print('[diag_capital_delete] payload:', {'Seq_No': seq_no, 'MH_Name': mh_name})

    with transaction.atomic():
        with connection.cursor() as cur:
            cur.execute("""
                DELETE FROM Diag_capital
                WHERE Seq_No=%s AND MH_Name=%s
            """, [seq_no, mh_name])
            deleted = cur.rowcount

    print('[diag_capital_delete] deleted rows:', deleted)
    return JsonResponse({'ok': True, 'deleted': deleted})


@require_GET
def diag_capital_summary(request):
    seq_no = request.GET.get('seq_no')
    if not seq_no:
        return HttpResponseBadRequest('missing seq_no')

    with connection.cursor() as cur:
        # 합계
        cur.execute("""
            SELECT 
                CAST(%s AS varchar(50)) AS Seq_No,
                ISNULL(SUM(CAST(MH_Amt AS decimal(18,2))),0) AS sum_MH_Amt
            FROM Diag_capital
            WHERE Seq_No=%s
              AND (CASE WHEN ISNUMERIC(CAST(MH_DcRate AS VARCHAR(10)))=1 THEN CAST(MH_DcRate AS DECIMAL(18,6)) ELSE 9 END) < 2
        """, [seq_no, seq_no])
        sum_row = _dictfetchone(cur) or {'Seq_No': seq_no, 'sum_MH_Amt': 0}

        # 최대 금액 면허명 (없으면 빈 값)
        cur.execute("""
            SELECT TOP 1 MH_Name
            FROM Diag_capital
            WHERE Seq_No=%s
              AND (CASE WHEN ISNUMERIC(CAST(MH_DcRate AS VARCHAR(10)))=1 THEN CAST(MH_DcRate AS DECIMAL(18,6)) ELSE 9 END) < 2
            ORDER BY CAST(MH_Amt AS decimal(18,2)) DESC, MH_Name ASC
        """, [seq_no])
        top_row = _dictfetchone(cur) or {'MH_Name': ''}

    return JsonResponse({
        'Seq_No': sum_row['Seq_No'],
        'MH_Name': top_row.get('MH_Name') or '',
        'sum_MH_Amt': float(sum_row['sum_MH_Amt'] or 0)
    })


def _dictfetchone(cur):
    row = cur.fetchone()
    if not row:
        return None
    cols = [c[0] for c in cur.description]
    return dict(zip(cols, row))
def _dictfetchall(cur):
    cols = [col[0] for col in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]
def _rows(cursor, cols):
    out = []
    for row in cursor.fetchall():
        d = {}
        for i, c in enumerate(cols):
            d[c] = row[i]
        out.append(d)
    return out
def _fetchall(sql, params=None):
    with connection.cursor() as cur:
        cur.execute(sql, params or [])
        return cur.fetchall()
    
def _fetchone(cursor, sql, params=()):
    cursor.execute(sql, params)
    return cursor.fetchone()

def fetchone_dict(cur):
    cols = [c[0] for c in cur.description]
    row = cur.fetchone()
    if not row:
        return None
    return {cols[i]: row[i] for i in range(len(cols))}

def _fetchone_scalar(cursor, sql, params, default=0):
    cursor.execute(sql, params)
    row = cursor.fetchone()
    if not row:
        return default
    val = row[0]
    return int(val or 0)