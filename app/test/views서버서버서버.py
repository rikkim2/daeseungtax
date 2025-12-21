from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.db import connection

from app.models import MemUser
from app.models import MemAdmin
from app.models import MemDeal
from app.models import 전자세금계산서
from app.models import userProfile

from app.test import utils
from app.test import jsHometax
from app.test import jsHometax_Screen_UTXPPAAA24 #개인TIN값
from app.test import jsHometax_Screen_UTECRCB023 #사업용카드
from app.test import jsHometax_Screen_UTECRCB024 #화물복지카드
from app.test import jsHometax_Screen_UTECRCB013 #현금영수증 매출
from app.test import jsHometax_Screen_UTECRCB005 #현금영수증 매출
from app.test import elecResult_Save

import socket
import os,glob

import pandas as pd
from pandas import DataFrame
import tkinter
from tkinter.filedialog import askopenfilename
import pyautogui
import pyperclip
import img2pdf
import mouse
import subprocess
import psutil
import pywinauto
from pywinauto.keyboard import send_keys
from pywinauto import keyboard    # 단축키 기능을 이용하기 위해서
from pywinauto import Desktop, Application, timings
import shutil
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException,NoAlertPresentException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.alert import Alert
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import UnexpectedAlertPresentException

import time
import math
import xlwt
from openpyxl import Workbook
from selenium.webdriver.support.ui import Select
from datetime import datetime
from datetime import timedelta 
from selenium.webdriver.chrome.service import Service
import multiprocessing
from glob import glob

pyautogui.FAILSAFE = False#마우스가 모서리로 이동하여 에러발생시키는 것을 방지
semusarangID='daeseung4';htxLoginID = 'daeseung20'
isSemusarangOn = False
myIP = socket.gethostbyname(socket.gethostname())
if myIP=='115.88.70.130':  semusarangID='daeseung6';  htxLoginID = "daeseung23"  #서버                    
if myIP=='123.142.103.148': semusarangID='daeseung1'; htxLoginID = "daeseung20"  #디비서버                    
if myIP=='192.168.0.3':   semusarangID='daeseung15';  htxLoginID = "daeseung29"  #i9-13900   - 세무사랑 전용
if myIP=='192.168.0.22':  semusarangID='daeseung20';   htxLoginID = "daeseung37"  #i9-10900-M2 - 상시온

if myIP=='123.142.103.149':  semusarangID='daeseung21';  htxLoginID = "daeseung36"  #i9-10900-DB
if myIP=='192.168.0.20':  semusarangID='daeseung21';  htxLoginID = "daeseung35"  #i7-7700 - 스크래핑 전용 /  / 312설치로 escape literal err 발생

if myIP=='192.168.0.19':  semusarangID='daeseung10';  htxLoginID = "daeseung22"  #i7-4790-2  - 세무사랑 전용 / 312설치로 escape literal err 발생

if myIP=='192.168.0.10':  semusarangID='daeseung21';  htxLoginID = "daeseung31"  #i7-8700 - 스크래핑 전용               
if myIP=='192.168.0.16':  semusarangID='daeseung3';   htxLoginID = "daeseung32"  #i5-6600 - 스크래핑 전용


if myIP=='192.168.0.18':  semusarangID='daeseung3';  htxLoginID = "daeseung27"  #i7-4790-1  - 세무사랑 안됨 / 원격 안켜짐 / 세무사랑 법인세 인식오류
      
print(myIP)
@login_required(login_url="/login/")
def index(request):
  context = {}
  memuser = MemUser.objects.get(user_id=request.user.username)
  if memuser.biz_type<4:
    context['isCorp'] = True
  userprofile = userProfile.objects.filter(title=memuser.seq_no)
  if userprofile:
    userprofile = userprofile.latest('description')
  if userprofile is not None:
    context['userProfile'] = userprofile
  context['memuser'] = memuser
  print(str(datetime.now())[:10].replace("-",""))
  return render(request, "test/test.html",context)


def Htx_Scrap(request):
  workyear = datetime.now().year
  workmm  = datetime.now().month
  today = str(datetime.now())[:10].replace("-","")
  flag = request.GET.get('flag',False)

  flagGroup = "";flagBizNo="";flagSeqNo='';flagYear='';flagPeriod='';flagPeriod2='';flagManagerName =''
  txtRangeStart="";txtRangeEnd="";strChkDate="";strChkDelFlag="Q";ks2="확정";ks1="1기"
  directory = "D:/"
  if flag in ('1','2'):flagPeriod = pyautogui.confirm(' 스크래핑 작업할 형태를 선택하세요',buttons=['무조건 루핑','통합조회와 비교'])  #카드매입인 경우
  flagFirst = pyautogui.confirm(' 작업할 형태를 선택하세요',buttons=['관리자별','단일사업자','전체(기장만)','전체(신고대리포함)'])
  flagSecond = "전체" #'개인','법인','전체'
  if flagFirst=="관리자별":
    sqladmin = "select admin_Name,Htx_ID,admin_id,manage_YN from mem_admin  order by admin_name"
    cursor = connection.cursor()
    cursor.execute(sqladmin)
    admins = cursor.fetchall()
    connection.commit()

    arrAdmin = []
    for admin in admins:
      if admin[3]=="Y" or admin[3]=="S":
        arrAdmin.append(admin[2])
    flagManagerName = pyautogui.confirm('그룹작업 진행할 관리자를 선택하세요. ',buttons=arrAdmin)
    for admin in admins:
      if admin[0]==flagManagerName:  
        flagGroup += "'"+admin[2]+"'," 
  elif flagFirst=="단일사업자":
    flagSeqNo = pyautogui.prompt('SEQNO를 입력하세요.');
  if flagFirst!="단일사업자": flagSecond = pyautogui.confirm('스크랩 대상을 선택하세요',buttons=['개인','법인','전체'])
  if flag in ('1','2') :#1:홈택스스크랩-카드, 2:홈택스 스크랩-현영
    # 증명원 발급은 1월은 당해, 나머지는 작년
    if datetime.now().month<=1 : workyear = datetime.now().year -1
    arrYear = [workyear,workyear-1,workyear-2]    
    flagYear = pyautogui.confirm('작업할 연도를 선택하세요. ',buttons=arrYear)
    flagPeriod2 = pyautogui.confirm('스크랩할 분기를 선택하세요',buttons=['1분기','2분기','3분기','4분기'])
    if flagPeriod2=='1분기' : txtRangeStart = flagYear+'0101';txtRangeEnd = flagYear+'0331';strChkDate="1";strChkDelFlag="Q";ks2="예정"
    elif flagPeriod2=='2분기' : txtRangeStart = flagYear+'0401';txtRangeEnd = flagYear+'0630';strChkDate="2";strChkDelFlag="Q"
    elif flagPeriod2=='3분기' : txtRangeStart = flagYear+'0701';txtRangeEnd = flagYear+'0930';strChkDate="3";strChkDelFlag="Q";ks2="예정";ks1="2기"
    elif flagPeriod2=='4분기' : txtRangeStart = flagYear+'1001';txtRangeEnd = flagYear+'1231';strChkDate="4";strChkDelFlag="Q";ks1="2기"    
  elif flag=='3':#증명원-부가세과표 -> 최종신고부터 소급3년
    txtRangeStart="1기";txtRangeEnd="2기"
    flagPeriod = workyear-3;  flagPeriod2 = workyear
    if workmm>1 and workmm<=4 : flagPeriod2 -= 1;flagPeriod = flagYear-3;txtRangeEnd="2기"
    if workmm>4 and workmm<10 : txtRangeEnd="1기"
    if workmm>=10 and workmm<=12 : txtRangeEnd="2기"
    if workmm==1: txtRangeEnd="2기"    
  elif flag=='5':#증명원-소득금액증명원 -> 다중연도
    flagPeriod = workyear-2;  flagPeriod2 = workyear-1  
    if workmm<6:  flagPeriod2 = workyear-2

  strsql = " select duzon_id,year(reg_date),a.seq_no,fiscalMM,biz_type,biz_name,biz_manager,ceo_name,ssn,HometaxID,HometaxPW,biz_no,hometaxAgree,fiscalMM,CONVERT(varchar(10), reg_date, 120),biz_addr1,hp_no from mem_user a,mem_deal b where a.seq_no=b.seq_no  "
  if flagFirst=="관리자별":     
    if flagSecond=="개인" : strsql += " and a.biz_type>=4 "
    elif flagSecond=="법인" : strsql += " and a.biz_type<4 "
    strsql += " and biz_manager ='" + flagManagerName + "' "
    strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
  elif flagFirst=="단일사업자": 
    strsql += " and a.seq_no ='"+flagSeqNo+"'   "      
  elif flagFirst=="전체(기장만)":   
    sqladmin = "select admin_Name,Htx_ID,admin_id,manage_YN from mem_admin where manage_YN='Y' order by admin_name"
    cursor = connection.cursor()
    cursor.execute(sqladmin)
    admins = cursor.fetchall()
    connection.commit()
    connection.close()  
    for admin in admins:
      flagGroup += "'"+admin[2]+"'," 
    flagGroup = flagGroup[:-1]      
    strsql += " and biz_manager  in ("+flagGroup+") " 
  elif  flagFirst=="전체":
    if flagSecond=="개인" : strsql += " and a.biz_type>=4 "
    elif flagSecond=="법인" : strsql += " and a.biz_type<4 "
    if flag in ('1','2'):strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') "  
    else :              strsql += " and biz_manager not in ('강남A','강남B','강남C','강남D','강남E','강남F','강남G','분당A','분당B','서울을','화물','종소세','종소세2','종소세3','오피스텔','환급1') "  
  strsql += " and a.seq_no not in ('3037') " 
  strsql += " and HometaxID<>'' and HometaxPW<>'' and duzon_ID <> '' and b.keeping_YN = 'Y' and a.Del_YN <> 'Y'   order by a.seq_no/1 "
  print(strsql)    
  cursor = connection.cursor()
  cursor.execute(strsql)
  results = cursor.fetchall()
  connection.commit()
  
  if flag=='1' or flag=='2' :#1:홈택스스크랩-카드, 2:홈택스 스크랩-현영
    result_LoginSystemCtx = jsHometax.LoginSystemContext()
    for result in results:    
      txtStnd_GB = " stnd_gb  "; arrChkDate = []
      strsql_f = "SELECT * FROM 부가가치세전자신고3 WHERE 사업자등록번호 = %s AND 과세기간 = %s AND 과세유형 = 'C17'"#c17은 고정임 수정말 것
      if strChkDate== '2':    
        with connection.cursor() as cursor:
          cursor.execute(strsql_f, (result[11], str(workyear)+"년 1기"))
          result_f = cursor.fetchall()
        connection.commit() 
        if result_f: txtStnd_GB += " ='"+strChkDate+"'";                                  arrChkDate=['2']
        else       : txtStnd_GB += " in ('"+str(int(strChkDate)-1)+"','"+strChkDate+"')"; arrChkDate=['1','2']
      elif strChkDate=='4': 
        with connection.cursor() as cursor:
          cursor.execute(strsql_f, (result[11], str(workyear)+"년 2기"))
          result_f = cursor.fetchall()
        connection.commit() 
        if result_f: txtStnd_GB += " ='"+strChkDate+"'";                                  arrChkDate=['4']
        else       : txtStnd_GB += " in ('"+str(int(strChkDate)-1)+"','"+strChkDate+"')"; arrChkDate=['3','4']
      else: txtStnd_GB += " ='"+strChkDate+"'";arrChkDate.append(strChkDate)

      #1:홈택스스크랩-카드
      if flag=='1' and len(result[11])>10:  
        loopYN = True
        if flagPeriod=="통합조회와 비교":
          strsql = "select isnull((select top 1 매입신용카드세액 from 부가가치세통합조회 where  사업자등록번호='"+result[11]+"' and 과세기수='"+flagYear+"년 "+ks1+"'  order by 신고구분 desc),0), "
          strsql += " case when "+str(result[4])+"<3 then Isnull((select sum(vaTxamt) from Tbl_HomeTax_Scrap  Where seq_no ='"+result[2]+"' and tran_yy='" + flagYear + "' and stnd_gb='"+strChkDate+"' and   ddcYnNm='공제'), 0) "
          strsql += " else Isnull((select sum(vaTxamt) from Tbl_HomeTax_Scrap Where seq_no ='"+result[2]+"' and tran_yy='" + flagYear + "' and "+txtStnd_GB+" and   ddcYnNm='공제'), 0) end"
          cursor = connection.cursor()
          print(strsql)
          cursor.execute(strsql)   
          scraps = cursor.fetchall()
          connection.commit() 
          if scraps[0][1]==scraps[0][0]:  loopYN = False
        HometaxID = result[9].strip()
        HometaxPW = result[10].strip() 
        if loopYN and HometaxPW!="" and HometaxPW[:4]!="비번틀림":
          rst = jsHometax.JsHometax.login_personal(HometaxID,HometaxPW,result_LoginSystemCtx)
          if rst==1:
            print("로그인성공 : "+ result[5])
            if result[4] > 3:#개인사업자인 경우 사업자번호별 TIN값을 가져온다
              screen24 = jsHometax_Screen_UTXPPAAA24.JsHometax_Screen_UTXPPAAA24()
              rst24 = screen24.requestPermission()
              print("UTXPPAAA24 requestPermission 결과 : " + str(rst24) + "(" + jsHometax.getErrorString(rst24) + ")")
              if rst24!=1: continue
              result_LoginSystemCtx.m_tin = screen24.action_ATXPPAAA003R01_getTin(result[11], result_LoginSystemCtx.m_tin)
              print("개인_ATXPPAAA003R01_getTin 결과 : " + result_LoginSystemCtx.m_tin)

            screen = jsHometax_Screen_UTECRCB023.JsHometax_Screen_UTECRCB023()
            actionctx = jsHometax_Screen_UTECRCB023.ActionCtx_ATECRCCA001R06()
            print(f'screen : {screen}')
            print(f'actionctx : {actionctx}')
            rst = screen.requestPermission()
            print("UTECRCB023 requestPermission 결과 : " + str(rst) + "(" + jsHometax.getErrorString(rst) + ")")
            if rst!=1: continue
            print(arrChkDate)
            for chkPeriod in arrChkDate:
              if chkPeriod=='1' : txtRangeStart = flagYear+'0101';txtRangeEnd = flagYear+'0331'
              elif  chkPeriod=='2' : txtRangeStart = flagYear+'0401';txtRangeEnd = flagYear+'0630'
              elif  chkPeriod=='3' : txtRangeStart = flagYear+'0701';txtRangeEnd = flagYear+'0930'
              elif  chkPeriod=='4' : txtRangeStart = flagYear+'1001';txtRangeEnd = flagYear+'1231'
              rst = screen.action_ATECRCCA001R06(actionctx, txtRangeStart, txtRangeEnd)
              print("ATECRCCA001R06 : " + str(rst) + "(" + jsHometax.getErrorString(rst) + ") - "+ "txtRangeStart:"+txtRangeStart+", txtRangeEnd"+txtRangeEnd)
              if rst!=1: continue
              print("조회기간 : " + actionctx.m_trsDtRngStrt + " ~ " + actionctx.m_trsDtRngEnd)
              print("공급대가 : " + str(actionctx.m_sumTotaTrsAmt) + " , 공급가액 : " + str(actionctx.m_sumSplCft)+" , 개수 : " + str(actionctx.m_totalCount))
              if result[6] != "화물" and actionctx.m_totalCount<=0: continue
              # //22.1.18 변경 - 복지카드를 사업용카드로 등록하는 경우가 많아서 사업용카드와 중복이 발생
              # //사업용카드 돌리고 사업용카드가 없는 경우 복지카드를 돌리자
              rst2Oil = -1
              screenOil = jsHometax_Screen_UTECRCB024.JsHometax_Screen_UTECRCB024()
              actionctxOil = jsHometax_Screen_UTECRCB024.ActionCtx_ATECRCCA001R07()
              print("화물공급대가 : " + str(actionctxOil.m_sumTotaTrsAmt) + " , 화물공급가액 : " + str(actionctxOil.m_sumSplCft)+" , 화물개수 : " + str(actionctxOil.m_totalCount))
              if result[6] == "화물" and actionctxOil.m_totalCount > 0:
                rst2Oil = screenOil.requestPermission()
                print("UTECRCB024 requestPermission: " + str(rst2Oil) + "(" + jsHometax.getErrorString(rst2Oil) + ")")
                rst2Oil = screenOil.action_ATECRCCA001R07(actionctxOil, txtRangeStart, txtRangeEnd)
              if result_LoginSystemCtx.m_code != "F":
                strdel = "delete from Tbl_HomeTax_Scrap Where seq_no ='"+result[2]+"' and tran_yy='" + flagYear + "' and stnd_GB='"+chkPeriod+"'"
                cursor = connection.cursor()
                cursor.execute(strdel)   
                if rst == 1:
                  rst2 = 0; page2 = 0
                  for page2 in range(1, actionctx.m_pageCount+1):
                    rst2 = screen.action_ATECRCCA001R06_getItems(actionctx,  page2,cursor,workyear,chkPeriod,HometaxID,HometaxPW,result[2],result[11],result[5])
                    print("신용카드_getItems (page " + str(page2) + " / " + str(actionctx.m_pageCount) + ") : " + str(rst2) + "(" + jsHometax.getErrorString(rst) + ")")
                    if rst2==-9:
                      rst2 = screen.action_ATECRCCA001R06_getItems(actionctx,  page2,cursor,workyear,chkPeriod,HometaxID,HometaxPW,result[2],result[11],result[5])
                      print("신용카드_getItems (page " + str(page2) + " / " + str(actionctx.m_pageCount) + ") : " + str(rst2) + "(" + jsHometax.getErrorString(rst) + ")")                      
                  strMNG = "Merge  tbl_mng_jaroe as A using(select '"+result[2]+"' as seq_no,'"+str(workyear)+"' as work_yy, '"+ str(13+int(chkPeriod))+"' as work_mm) as B  "
                  strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_mm=B.work_mm "
                  strMNG += " when matched then update set YN_7='1'  "
                  strMNG += " when not matched then insert values('"+result[2]+"','"+str(workyear)+"',"+ str(13+int(chkPeriod))+",'0','0','0','0','0','0','1','0','0','0','0','0','0','0',''); " 
                  cursor.execute(strMNG)                       
                if rst2Oil == 1:
                  rst2Oil = 0; page2Oil = 0
                  for page2Oil in range(1, actionctxOil.m_pageCount+1):
                    rst2Oil = screenOil.action_ATECRCCA001R07_getItems(actionctxOil, page2Oil,cursor,workyear,chkPeriod,HometaxID,HometaxPW,result[2],result[11],result[5])
                    print("ATECRCCA001R07 (page " + str(page2Oil) + " / " + str(actionctx.m_pageCount) + ") : " + str(rst2Oil) + "(" + jsHometax.getErrorString(rst2Oil) + ")")
                    if rst2Oil==-9:
                      rst2Oil = screenOil.action_ATECRCCA001R07_getItems(actionctxOil, page2Oil,cursor,workyear,chkPeriod,HometaxID,HometaxPW,result[2],result[11],result[5])
                      print("ATECRCCA001R07 (page " + str(page2Oil) + " / " + str(actionctx.m_pageCount) + ") : " + str(rst2Oil) + "(" + jsHometax.getErrorString(rst2Oil) + ")")
              strsql = "exec SP_HomeTax_Scrap '" + result[2] + "','" + str(workyear) + "','" + chkPeriod + "' "
              cursor.execute(strsql);print(strsql)
          else:
            print(' 비번이 맞지 않음 : '+result[5])
            if HometaxPW[:4]!="비번틀림":
              strUdt = "update mem_deal set HometaxPw = '비번틀림"+HometaxPW+"' where seq_no='"+result[2]+"'";print(strUdt)
              cursor = connection.cursor()
              cursor.execute(strUdt)
      # 2:홈택스 스크랩-현영
      elif flag=='2' and len(result[11])>10:
        #현영매출
        strsql = "select * from 스크래핑관리 where crt_dt='" + today + "' and seqno_lastTry='"+result[2]+"' and bigo='현금영수증매출"+strChkDate+"'"
        cursor = connection.cursor()
        cursor.execute(strsql)   
        scraps = cursor.fetchall()
        connection.commit() 
        if not scraps:   
          HometaxID = result[9].strip()
          HometaxPW = result[10].strip()
          if HometaxPW!="" and HometaxPW[:4]!="비번틀림":
            rst = jsHometax.JsHometax.login(HometaxID,HometaxPW,result_LoginSystemCtx)
            print(rst)
            if rst==1:
              print("로그인성공 : "+ result[5])
              if result[4] > 3:#개인사업자인 경우 사업자번호별 TIN값을 가져온다
                screen24 = jsHometax_Screen_UTXPPAAA24.JsHometax_Screen_UTXPPAAA24()
                rst24 = screen24.requestPermission()
                print("UTXPPAAA24 requestPermission 결과 : " + str(rst24) + "(" + jsHometax.getErrorString(rst24) + ")")
                if rst24!=1: continue
                result_LoginSystemCtx.m_tin = screen24.action_ATXPPAAA003R01_getTin(result[11], result_LoginSystemCtx.m_tin)
                print("개인_ATXPPAAA003R01_getTin 결과 : " + result_LoginSystemCtx.m_tin)

              screen = jsHometax_Screen_UTECRCB013.JsHometax_Screen_UTECRCB013()
              actionctx = jsHometax_Screen_UTECRCB013.ActionCtx_ATECRCBA001R03()
              rst = screen.requestPermission()          
              print("UTECRCB013 requestPermission 결과 : " + str(rst) + "(" + jsHometax.getErrorString(rst) + ")")
              if rst!=1: continue

              rst = screen.action_ATECRCBA001R03(actionctx, txtRangeStart, txtRangeEnd)
              print("ATECRCBA001R03 현영매출집계 : " + str(rst) + "(" + jsHometax.getErrorString(rst) + ")")
              if rst!=1: continue
              print("조회기간 : " + actionctx.m_trsDtRngStrt + " ~ " + actionctx.m_trsDtRngEnd)
              print("공급대가 : " + str(actionctx.m_sumTotaTrsAmt) + " , 공급가액 : " + str(actionctx.m_sumSplCft)+" , 개수 : " + str(actionctx.m_totalCount)) 
              if result_LoginSystemCtx.m_code != "F":
                if rst == 1:
                  del_sql = ""
                  if strChkDelFlag == "Q":
                    del_sql = "DELETE FROM Tbl_HomeTax_CashSale WHERE seq_no="+result[2]+" and left(tran_dt,4)="+str(workyear)+" and stnd_gb='"+strChkDate+"'"
                  elif strChkDelFlag == "M":
                    del_sql = "DELETE FROM Tbl_HomeTax_CashSale WHERE seq_no="+result[2]+" and left(tran_dt,4)="+str(workyear)+" and stnd_gb='"+strChkDate+"' and left(Tran_Dt,7)='"+txtRangeStart[:4]+"-"+txtRangeStart[4:6]+"'"
                  elif strChkDelFlag == "B":
                    del_sql = "DELETE FROM Tbl_HomeTax_CashSale WHERE seq_no="+result[2]+" and left(tran_dt,4)="+str(workyear)+" and  replace(left(Tran_Dt,10),'-','')>='"+txtRangeStart+"' and replace(left(Tran_Dt,10),'-','')<='"+txtRangeEnd+"'"
                  # print(del_sql)
                  cursor.execute(del_sql)

                  rst2 = 0; page2 = 0
                  # print(actionctx.m_pageCount+1)
                  for page2 in range(1, actionctx.m_pageCount+1):
                    rst2 = screen.action_ATECRCBA001R03_getItems(actionctx,  page2,cursor,workyear,strChkDate,txtRangeStart,txtRangeEnd,HometaxID,HometaxPW,result[2],result[11],result[5])
                    print("현영매출 (page " + str(page2) + " / " + str(actionctx.m_pageCount) + ") : " + str(rst2) + "(" + jsHometax.getErrorString(rst) + ")")  
                    if rst2==-9:
                      rst2 = screen.action_ATECRCBA001R03_getItems(actionctx,  page2,cursor,workyear,strChkDate,txtRangeStart,txtRangeEnd,HometaxID,HometaxPW,result[2],result[11],result[5])
                      print("현영매출 (page " + str(page2) + " / " + str(actionctx.m_pageCount) + ") : " + str(rst2) + "(" + jsHometax.getErrorString(rst) + ")")  
                  strdel = "delete from tbl_hometax_salecard  where seq_no="+result[2]+" and tran_YY="+str(workyear)+" and saleGubun='현금영수증'"       
                  cursor.execute(strdel)  
                  strins = "INSERT INTO tbl_hometax_salecard (Tran_YY,Seq_No,SaleGubun,Stnd_GB,Tran_MM,MM_Scnt,Tot_StlAmt,Etc_StlAmt,PurcEuCardAmt,TipAmt,Crt_Dt,Acnt_cd) "
                  strins += "SELECT LEFT(max(tran_dt), 4) AS Tran_YY,max(seq_no) as Seq_No,'현금영수증' as SaleGubun," 
                  strins += "case when month(max(tran_dt))<4 then '1' when month(max(tran_dt))<7 then '2'	when month(max(tran_dt))<10 then '3' when month(max(tran_dt))<13 then '4' end as Stnd_GB," 
                  strins += "replace(left(tran_dt,7),'-','') as Tran_MM,COUNT(*) AS MM_Scnt,SUM(totaTrsAmt) AS Tot_StlAmt, SUM(splCft) AS Etc_StlAmt, SUM(vaTxamt) AS PurcEuCardAmt,SUM(tip) AS TipAmt, getdate(),"
                  strins += "isnull((SELECT TOP 1 acnt_cd  FROM ds_slipledgr2 where seq_no="+result[2]+" and work_yy>"+str(int(flagYear)-5)+" and acnt_cd>=401 and acnt_cd<430 "
                  strins += "GROUP BY acnt_cd ORDER BY COUNT(acnt_cd) DESC),'') as Acnt_Cd "
                  strins += "FROM tbl_hometax_cashsale WHERE seq_no = "+result[2]+" AND LEFT(tran_dt, 4) = "+str(workyear)+" GROUP BY LEFT(tran_dt, 7);"
                  cursor.execute(strins)                   
                  strsql = "insert into 스크래핑관리 values('"+today+"','"+result[2]+"','"+result[5]+"','현금영수증매출"+strChkDate+"')"       
                  cursor.execute(strsql)
            else:
              print(result[5] + " : 홈택스 로그인하지 못했습니다 아이디/패스워드를 확인하시기 바랍니다2")
        #현영매입
        strsql = "select Isnull((select top 1 매입현금영수증세액 from 부가가치세통합조회 Where 사업자등록번호='"+result[11]+"' and 과세기수='"+flagYear+"년 "+ks1+"' and 신고구분='"+ks2+"'),0)  YN_TongCashCost,"
        strsql += " case when "+str(result[4])+"<3 then Isnull((select sum(vaTxamt) from Tbl_HomeTax_CashCost Where seq_no ='"+result[2]+"' and left(tran_dt,4)='" + flagYear + "' and stnd_gb='"+strChkDate+"' and   ddcYnNm='공제'), 0) else Isnull((select sum(vaTxamt) from Tbl_HomeTax_CashCost Where seq_no ='"+result[2]+"' and left(tran_dt,4)='" + flagYear + "' and "+txtStnd_GB+" and   ddcYnNm='공제'), 0) end"
        # cursor = connection.cursor()
        print(strsql)
        cursor.execute(strsql)   
        scraps = cursor.fetchall()
        connection.commit() 
        if scraps[0][1] != scraps[0][0]:    
          HometaxID = result[9].strip()
          HometaxPW = result[10].strip()
          rst = jsHometax.JsHometax.login(HometaxID,HometaxPW,result_LoginSystemCtx)
          if rst==1:
            print("로그인성공 : "+ result[5])
            if result[4] > 3:#개인사업자인 경우 사업자번호별 TIN값을 가져온다
              screen24 = jsHometax_Screen_UTXPPAAA24.JsHometax_Screen_UTXPPAAA24()
              rst24 = screen24.requestPermission()
              print("UTXPPAAA24 requestPermission 결과 : " + str(rst24) + "(" + jsHometax.getErrorString(rst24) + ")")
              if rst24!=1: continue
              result_LoginSystemCtx.m_tin = screen24.action_ATXPPAAA003R01_getTin(result[11], result_LoginSystemCtx.m_tin)
              print("개인_ATXPPAAA003R01_getTin 결과 : " + result_LoginSystemCtx.m_tin)

            screen = jsHometax_Screen_UTECRCB005.JsHometax_Screen_UTECRCB005()
            actionctx = jsHometax_Screen_UTECRCB005.ActionCtx_ATECRCBA001R02()
            rst = screen.requestPermission()          
            print("UTECRCB005 requestPermission 결과 : " + str(rst) + "(" + jsHometax.getErrorString(rst) + ")")
            if rst!=1: continue

            rst = screen.action_ATECRCBA001R02(actionctx, txtRangeStart, txtRangeEnd)
            print("현영매입집계 : " + str(rst) + "(" + jsHometax.getErrorString(rst) + ")")
            if rst!=1: continue
            print("조회기간 : " + actionctx.m_trsDtRngStrt + " ~ " + actionctx.m_trsDtRngEnd)
            print("공급대가 : " + str(actionctx.m_sumTotaTrsAmt) + " , 공급가액 : " + str(actionctx.m_sumSplCft)+" , 개수 : " + str(actionctx.m_totalCount)) 
            if result_LoginSystemCtx.m_code != "F":
              if rst == 1:
                del_sql = ""
                if strChkDelFlag == "Q":
                  del_sql = "DELETE FROM Tbl_HomeTax_CashCost WHERE seq_no="+result[2]+" and left(tran_dt,4)="+str(workyear)+" and stnd_gb='"+strChkDate+"'"
                elif strChkDelFlag == "M":
                  del_sql = "DELETE FROM Tbl_HomeTax_CashCost WHERE seq_no="+result[2]+" and left(tran_dt,4)="+str(workyear)+" and stnd_gb='"+strChkDate+"' and left(Tran_Dt,7)='"+txtRangeStart[:4]+"-"+txtRangeStart[4:6]+"'"
                elif strChkDelFlag == "B":
                  del_sql = "DELETE FROM Tbl_HomeTax_CashCost WHERE seq_no="+result[2]+" and left(tran_dt,4)="+str(workyear)+" and  replace(left(Tran_Dt,10),'-','')>='"+txtRangeStart+"' and replace(left(Tran_Dt,10),'-','')<='"+txtRangeEnd+"'"
                # print(del_sql)
                cursor.execute(del_sql)

                rst2 = 0; page2 = 0
                # print(actionctx.m_pageCount+1)
                for page2 in range(1, actionctx.m_pageCount+1):
                  rst2 = screen.action_ATECRCBA001R02_getItems(actionctx,  page2,cursor,workyear,strChkDate,txtRangeStart,txtRangeEnd,HometaxID,HometaxPW,result[2],result[11],result[5])
                  print("현영매입 (page " + str(page2) + " / " + str(actionctx.m_pageCount) + ") : " + str(rst2) + "(" + jsHometax.getErrorString(rst) + ")")  
                  if rst2==-9:  
                    rst2 = screen.action_ATECRCBA001R02_getItems(actionctx,  page2,cursor,workyear,strChkDate,txtRangeStart,txtRangeEnd,HometaxID,HometaxPW,result[2],result[11],result[5])
                    print("현영매입 (page " + str(page2) + " / " + str(actionctx.m_pageCount) + ") : " + str(rst2) + "(" + jsHometax.getErrorString(rst) + ")")                                    
                # strsql = "insert into 스크래핑관리 values('"+today+"','"+result[2]+"','"+result[5]+"','현금영수증매입')"       
                # cursor.execute(strsql)
            if strChkDelFlag=="B" :
              if strChkDate=="6":
                strsql = "exec SP_HomeTax_ScrapCash '" + result[2] + "','" + str(workyear) + "','1' "
                cursor.execute(strsql)          
                strsql = "exec SP_HomeTax_ScrapCash '" + result[2] + "','" + str(workyear) + "','2' "
                cursor.execute(strsql)               
              elif strChkDate=="12":
                strsql = "exec SP_HomeTax_ScrapCash '" + result[2] + "','" + str(workyear) + "','3' "
                cursor.execute(strsql)          
                strsql = "exec SP_HomeTax_ScrapCash '" + result[2] + "','" + str(workyear) + "','4' "
                cursor.execute(strsql)                         
            else:
              strsql = "exec SP_HomeTax_ScrapCash '" + result[2] + "','" + str(workyear) + "','" + strChkDate + "' ";print(strsql)
              cursor.execute(strsql)    
          else:
            print(result[5] + " : 홈택스 로그인하지 못했습니다 아이디/패스워드를 확인하시기 바랍니다1")                          
    connection.close()
  elif flag=='3' and results:#증명원-부가세과표 -> 최종신고부터 소급3년
    driver = utils.conHometaxLogin(htxLoginID,False);time.sleep(1)
    for result in results: 
      driver.get('https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/rn/a/b/a/e/UTERNAA150.xml');time.sleep(1)
      strsql = "select * from 스크래핑관리 where crt_dt='" + today + "' and seqno_lastTry='"+result[2]+"' and bigo='부가가치세과세표준증명원'"
      cursor = connection.cursor()
      cursor.execute(strsql)   
      scraps = cursor.fetchall()
      connection.commit()         
      if result[12]=='Y' and result[4]!=6 and result[6]!='화물' and not scraps:#홈택스 수임동의
        bizNo = result[11].split('-')
        # driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo1').clear();time.sleep(0.5)
        driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo1').send_keys(bizNo[0]);time.sleep(0.25);    
        # driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo2').clear();time.sleep(0.5)
        driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo2').send_keys(bizNo[1]);time.sleep(0.25);          
        # driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo3').clear();time.sleep(0.5)
        driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo3').send_keys(bizNo[2]);time.sleep(0.5);          
        driver.find_element(By.ID,'groupBtnCvaAplnBscClntFnmTnm').click();print(result[11]+' 사업자번호로 조회 : 확인버튼')
        WebDriverWait(driver, 5).until(EC.alert_is_present()); al = Alert(driver); al.accept(); time.sleep(1)  
        agree = driver.find_element(By.CSS_SELECTOR,'#rbChkClntAfaYn_input_0');    agree.click(); print('수임여부 : 예') ;time.sleep(0.5) 
        print('신청내용')
        select1 = Select(driver.find_element(By.ID,'cmbCvaDcumUseUsgCd'));        select1.select_by_visible_text('기타');time.sleep(0.5)
        select2 = Select(driver.find_element(By.ID,'cmbCvaDcumSbmsOrgnClCd'));    select2.select_by_visible_text('기타');time.sleep(0.5)
        print('과세기간')
        select3 = Select(driver.find_element(By.ID,'selectTxnrmStrtYear'));       select3.select_by_visible_text(str(flagPeriod));time.sleep(0.5)
        select4 = Select(driver.find_element(By.ID,'selectStrtHt'));              select4.select_by_visible_text('1기');time.sleep(0.5)
        select5 = Select(driver.find_element(By.ID,'selectTxnrmEndYear'));        select5.select_by_visible_text(str(flagPeriod2));time.sleep(0.5)
        select6 = Select(driver.find_element(By.ID,'selectEndHt'));               select6.select_by_visible_text(txtRangeEnd);time.sleep(0.5)        
        driver.find_element(By.ID,'trigger85').click(); print('조회')
        WebDriverWait(driver, 10).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1);print('의뢰인수임여부확인알람')
        try:
          WebDriverWait(driver, 10).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(5);print('신청하시겠습니까')
          table = driver.find_element(By.ID,'grdCvaAlfaBrkdInqr_body_table')
          tbody = table.find_element(By.ID,'grdCvaAlfaBrkdInqr_body_tbody')  
          if len(tbody.find_elements(By.TAG_NAME,'tr')) >0 and driver.find_element(By.XPATH,'//*[@id="grdCvaAlfaBrkdInqr_cell_0_10"]').text=='출력': 
            printBtnName = '//*[@id="grdCvaAlfaBrkdInqr_cell_0_10"]'
            fileName = "부가가치세과세표준증명원"
            dirProof = directory + result[5]+"/"+str(workyear)+"/홈택스민원서류"
            utils.Htx_Popup_Print_Minwon(driver,printBtnName,fileName,dirProof,False)   
            strsql = "insert into 스크래핑관리 values('"+today+"','"+result[2]+"','"+result[5]+"','부가가치세과세표준증명원')"       
            cursor.execute(strsql)   
        except TimeoutException:   
          print('오류 : 신청하시겠습니까? 팝업이 발생하지 않음' )
  elif flag=='4' and results:#증명원-재무제표증명원 -> 단일연도
    driver = utils.conHometaxLogin(htxLoginID,False);time.sleep(1)
    for result in results: 
      driver.get('https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/rn/a/a/a/d/UTERNAAT17.xml');time.sleep(1)
      strsql = "select * from 스크래핑관리 where crt_dt='" + today + "' and seqno_lastTry='"+result[2]+"' and bigo='표준재무제표증명원'"
      cursor = connection.cursor()
      cursor.execute(strsql)   
      scraps = cursor.fetchall()
      connection.commit()         
      if result[12]=='Y' and result[4]!=5 and result[6]!='화물' and not scraps and result[11]!="--":#홈택스 수임동의
        
        bizNo = result[11].split('-')        
        # driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo1').clear();time.sleep(0.5)
        driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo1').send_keys(bizNo[0]);time.sleep(0.25);    
        # driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo2').clear();time.sleep(0.5)
        driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo2').send_keys(bizNo[1]);time.sleep(0.25);          
        # driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo3').clear();time.sleep(0.5)
        driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo3').send_keys(bizNo[2]);time.sleep(0.5);          
        driver.find_element(By.ID,'groupBtnCvaAplnBscClntFnmTnm').click();print(result[11]+' 사업자번호로 조회 : 확인버튼')
        WebDriverWait(driver, 5).until(EC.alert_is_present()); al = Alert(driver); al.accept(); time.sleep(1)  
        agree = driver.find_element(By.CSS_SELECTOR,'#rbChkClntAfaYn_input_0');    agree.click(); print('수임여부 : 예') ;time.sleep(0.5) 
        print('신청내용')
        select1 = Select(driver.find_element(By.ID,'cmbCvaDcumUseUsgCd'));        select1.select_by_visible_text('기타');time.sleep(0.5)
        select2 = Select(driver.find_element(By.ID,'cmbCvaDcumSbmsOrgnClCd'));    select2.select_by_visible_text('기타');time.sleep(0.5)    
        print('과세기간')
        flagPeriod = str(workyear-1)+"12"
        if result[4]<4:
          if result[13]!="12": flagPeriod =  str(workyear-1)+"0"+result[13]
          # driver.find_element(By.ID,'inputBsEndYear2').clear();       
          driver.find_element(By.ID,'inputBsEndYear2').send_keys(flagPeriod);time.sleep(0.5);  
        else:
          # driver.find_element(By.ID,'inputBsEndYear').clear();       
          driver.find_element(By.ID,'inputBsEndYear').send_keys(str(workyear-1));time.sleep(0.5);   
          agree = driver.find_element(By.CSS_SELECTOR,'#radio1_indx_input_0');    agree.click(); print('소득구분 : 사업소득') ;time.sleep(0.5)         
        driver.find_element(By.ID,'trigger85').click(); print('조회')
        WebDriverWait(driver, 10).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1);print('의뢰인수임여부확인알람')
        WebDriverWait(driver, 10).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(3);print('신청하시겠습니까')
        try:
          WebDriverWait(driver, 10).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(3);print('정상접수되었습니다.')
          table = driver.find_element(By.ID,'grdCvaAlfaBrkdInqr_body_table')
          tbody = table.find_element(By.ID,'grdCvaAlfaBrkdInqr_body_tbody')  
          if len(tbody.find_elements(By.TAG_NAME,'tr')) >0 and driver.find_element(By.XPATH,'//*[@id="grdCvaAlfaBrkdInqr_cell_0_10"]').text=='출력': 
            printBtnName = '//*[@id="grdCvaAlfaBrkdInqr_cell_0_10"]'
            dirProof = directory + result[5]+"/"+str(workyear)+"/홈택스민원서류"
            utils.Htx_Popup_Print_Minwon(driver,printBtnName,"표준재무제표증명원("+flagPeriod+")",dirProof,False)   
            strsql = "insert into 스크래핑관리 values('"+today+"','"+result[2]+"','"+result[5]+"','표준재무제표증명원')"       
            cursor.execute(strsql)
        except TimeoutException:   
          print('오류 : 정상접수되었습니다. 팝업이 발생하지 않음' )  
          driver.get('http://www.hometax.go.kr/websquare/websquare_cdn.html?w2xPath=/ui/pp/index.xml');driver.implicitly_wait(3)
          continue
      else:
        continue  
  elif flag=='5' and results:#증명원-소득금액증명원 -> 다중연도
    driver = utils.conHometaxLogin(htxLoginID,False);time.sleep(1) 
    for result in results: 
      driver.get('https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/sf/a/a/UTESFAAE35.xml') ;time.sleep(1)  
      if result[4]>3:
        strsql = "select * from 스크래핑관리 where crt_dt='" + today + "' and seqno_lastTry='"+result[2]+"' and bigo='소득금액증명원'"
        cursor = connection.cursor()
        cursor.execute(strsql)   
        scraps = cursor.fetchall()
        connection.commit()         
        if result[12]=='Y' and result[4]>=4 and result[6]!='화물' and not scraps:#홈택스 수임동의
          try:
            WebDriverWait(driver, 5).until(EC.alert_is_present()); al = Alert(driver); al.accept(); time.sleep(1);print('원활한홈택스권장사항')
          except:
            print('원활한홈택스권장사항 알람 없음')
          ssn = result[8];print(ssn)
          if ssn:
            # driver.find_element(By.ID,'inputCvaAplnBscClntResRgtNo1').clear();time.sleep(0.5)
            driver.find_element(By.XPATH,'//*[@id="inputCvaAplnBscClntResRgtNo1"]').send_keys(ssn[:6]);time.sleep(0.25);    
            # driver.find_element(By.ID,'inputCvaAplnBscClntResRgtNo2').clear();time.sleep(0.5)
            driver.find_element(By.XPATH,'//*[@id="inputCvaAplnBscClntResRgtNo2"]').send_keys(ssn[6:13]);time.sleep(0.25);          
            driver.find_element(By.ID,'groupBtnCvaAplnBscClntFnmTnm').click();print(ssn +' 주민번호로 조회 : 확인버튼')
            WebDriverWait(driver, 5).until(EC.alert_is_present()); al = Alert(driver)
            if al.text=='확인되었습니다':
              al.accept(); time.sleep(1);print('확인되었습니다.')
              agree = driver.find_element(By.CSS_SELECTOR,'#rbChkClntAfaYn_input_0');    agree.click(); print('수임여부 : 예') ;time.sleep(0.5) 
              agree = driver.find_element(By.CSS_SELECTOR,'#rdoEnglCvaAplnYn_input_0');    agree.click(); print('한글증명') ;time.sleep(0.5) 
              agree = driver.find_element(By.CSS_SELECTOR,'#rdoIncOcplOpYn_input_0');    agree.click(); print('소득발생처공개여부') ;time.sleep(0.5) 
              print('과세기간')
              select3 = Select(driver.find_element(By.ID,'sbTxnrmStrtYr'));       select3.select_by_visible_text(str(flagPeriod));time.sleep(0.5)
              select5 = Select(driver.find_element(By.ID,'sbTxnrmEndYr'));        select5.select_by_visible_text(str(flagPeriod2));time.sleep(0.5)        
              print('신청내용')
              select1 = Select(driver.find_element(By.ID,'sbCvaDcumUseUsgCd'));        select1.select_by_visible_text('기타');time.sleep(0.5)
              select2 = Select(driver.find_element(By.ID,'sbCvaDcumSbmsOrgnClCd'));    select2.select_by_visible_text('기타');time.sleep(0.5)    
              print('신청하기')
              driver.find_element(By.ID,'trigger167').click(); 
              WebDriverWait(driver, 10).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1);print('의뢰인수임여부확인알람')
              try:
                WebDriverWait(driver, 10).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(3);print('신청하시겠습니까')
                WebDriverWait(driver, 10).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(3);print('정상접수되었습니다.')
                table = driver.find_element(By.ID,'grdCvaAlfaBrkdInqr_body_table')
                tbody = table.find_element(By.ID,'grdCvaAlfaBrkdInqr_body_tbody')  
                if len(tbody.find_elements(By.TAG_NAME,'tr')) >0 and driver.find_element(By.XPATH,'//*[@id="grdCvaAlfaBrkdInqr_cell_0_10"]').text=='출력': 
                  printBtnName = '//*[@id="grdCvaAlfaBrkdInqr_cell_0_10"]'
                  dirProof = directory + result[5]+"/"+str(workyear)+"/홈택스민원서류"
                  utils.Htx_Popup_Print_Minwon(driver,printBtnName,"소득금액증명원",dirProof,False)   
                  strsql = "insert into 스크래핑관리 values('"+today+"','"+result[2]+"','"+result[5]+"','소득금액증명원')"       
                  cursor.execute(strsql)
              except TimeoutException:   
                print('오류 : 정상접수되었습니다. 팝업이 발생하지 않음' )   
            else:
              print('수임자료를 확인하세요(미등록 또는 미동의시에 증명발급이 불가합니다')
              continue
          else:
            print('회원등록정보에 주민번호가 누락되었습니다')                   
  elif flag=='6' and results:#국세완납증명원
    driver = utils.conHometaxLogin(htxLoginID,False);time.sleep(1)
    for result in results: 
      driver.get('https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/rd/a/a/UTERDAAA04.xml');time.sleep(1)      
      strsql = "select * from 스크래핑관리 where crt_dt='" + today + "' and seqno_lastTry='"+result[2]+"' and bigo='국세완납증명원'"
      cursor = connection.cursor()
      cursor.execute(strsql)   
      scraps = cursor.fetchall()
      connection.commit()         
      if result[12]=='Y' and result[6]!='화물' and not scraps and result[11]!="--":#홈택스 수임동의
        bizNo = result[11].split('-') ; ssn1 = result[8][:6];ssn2 = result[8][7:13]  
        if result[4]<4: 
          driver.find_element(By.ID,'rbCvaAplnBscClntBmanClCd_input_1').click();  
          # driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo1').clear();time.sleep(0.25)
          driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo1').send_keys(bizNo[0]);time.sleep(0.25);    
          # driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo2').clear();time.sleep(0.25)
          driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo2').send_keys(bizNo[1]);time.sleep(0.25);          
          # driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo3').clear();time.sleep(0.25)
          driver.find_element(By.ID,'inputCvaAplnBscClntBmanRgtNo3').send_keys(bizNo[2]);time.sleep(0.25);          
        else:
          driver.find_element(By.ID,'rbCvaAplnBscClntBmanClCd_input_0').click();  
          # driver.find_element(By.ID,'inputCvaAplnBscClntResRgtNo1').clear();time.sleep(0.25)
          driver.find_element(By.ID,'inputCvaAplnBscClntResRgtNo1').send_keys(ssn1);time.sleep(0.25);    
          # driver.find_element(By.ID,'inputCvaAplnBscClntResRgtNo2').clear();time.sleep(0.25)
          driver.find_element(By.ID,'inputCvaAplnBscClntResRgtNo2').send_keys(ssn2);time.sleep(0.25);                 
        driver.find_element(By.ID,'btnCvaAplnBscClntFnmTnm').click();print(result[11]+' 사업자번호로 조회 : 확인버튼');time.sleep(0.5)
        WebDriverWait(driver, 10).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1);print('확인하엿습니다')
        driver.find_element(By.ID,'rbChkClntAfaYn_input_0').click();print('의뢰인 수임동의');time.sleep(0.3)
        driver.find_element(By.ID,'rbCvaAplnRecptResNoOpYn_input_0').click();print('주민번호 공개 라디오버튼');time.sleep(0.3)
        select = Select(driver.find_element(By.ID,'cmbCvaDcumSbmsOrgnClCd'));select.select_by_visible_text('금융기관'); driver.implicitly_wait(5)
        driver.find_element(By.ID,'btnApln').click();time.sleep(0.8);print('신청하기')
        WebDriverWait(driver, 5).until(EC.alert_is_present()); al = Alert(driver); al.accept(); time.sleep(0.3)  ;print('의뢰인수임여부확인알람')
        try:
          WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1);print('신청하시겠습니까')
          WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(3);print('정상접수되었습니다.')
          table = driver.find_element(By.ID,'grdCvaAlfaBrkdInqr_body_table')
          tbody = table.find_element(By.ID,'grdCvaAlfaBrkdInqr_body_tbody')  
          if len(tbody.find_elements(By.TAG_NAME,'tr')) >0 and driver.find_element(By.XPATH,'//*[@id="grdCvaAlfaBrkdInqr_cell_0_10"]').text=='출력': 
            printBtnName = '//*[@id="grdCvaAlfaBrkdInqr_cell_0_10"]'
            dirProof = directory + result[5]+"/"+str(workyear)+"/홈택스민원서류"
            utils.Htx_Popup_Print_Minwon(driver,printBtnName,"국세완납증명서",dirProof,False)   
            strsql = "insert into 스크래핑관리 values('"+today+"','"+result[2]+"','"+result[5]+"','국세완납증명원')"       
            cursor.execute(strsql)
          else:
            print(result[5]+' : 국세를 완납하지 않음')
            strsql = "insert into 스크래핑관리 values('"+today+"','"+result[2]+"','"+result[5]+"','국세완납증명원')"       
            cursor.execute(strsql)                      
            # driver.get('http://www.hometax.go.kr/websquare/websquare_cdn.html?w2xPath=/ui/pp/index.xml');time.sleep(3)
        except TimeoutException:   
          print('오류 : 정상접수되었습니다. 팝업이 발생하지 않음 - 수임동의 안됨' )  
          strsql = "insert into 스크래핑관리 values('"+today+"','"+result[2]+"','"+result[5]+"','국세완납증명원')"       
          cursor.execute(strsql)          
          # driver.get('http://www.hometax.go.kr/websquare/websquare_cdn.html?w2xPath=/ui/pp/index.xml');time.sleep(3)
          continue
      else:
        print('대상자 아님')
        continue 
  elif flag=='7':#지방세 완납증명원
    driver = utils.conWetaxLogin(False);time.sleep(1)
    for result in results: 
      if len(result[11])>=10:
        strsql = "select * from 스크래핑관리 where crt_dt='" + today + "' and seqno_lastTry='"+result[2]+"' and bigo='지방세 납세증명서'"
        cursor = connection.cursor()
        cursor.execute(strsql)   
        scraps = cursor.fetchall()
        connection.commit()         
        if result[12]=='Y' and result[4]!=5 and result[6]!='화물' and not scraps:#홈택스 수임동의
          driver.find_element(By.XPATH,'//*[@id="header"]/div[2]/div/div[3]/a/I').click(); time.sleep(2); print('전체메뉴')
          driver.find_element(By.XPATH,'//*[@id="contentsWrap"]/div[2]/ul[5]/li[1]/a').click();time.sleep(1); print('문서발급 ')
          driver.execute_script("javascript:moveScrn('G0102');return false;");time.sleep(1)   
          wait = WebDriverWait(driver, 10)  # Wait up to 10 seconds before timing out
          element = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="btnDlgpLst"]')))
          driver.execute_script("arguments[0].click();", element)  
          iframe = driver.find_element(By.XPATH,'//*[@id="cmnPopup_dlgp"]/div/div/iframe'); driver.switch_to.frame(iframe);    time.sleep(1)
          driver.find_element(By.XPATH,'//*[@id="condDlgpBrno"]').send_keys(result[11].replace("-",""));time.sleep(1)
          driver.find_element(By.XPATH,'//*[@id="btnDlgpSearch"]').click();time.sleep(1);print(f'{result[5]} : 검색버튼 클릭')
          table = driver.find_element(By.ID, 'tbl_listDlgpUser')
          tbody = table.find_element(By.XPATH, '//*[@id="tbl_listDlgpUser"]/tbody')
          for tr in tbody.find_elements(By.TAG_NAME, 'tr'):
            cols = tr.find_elements(By.TAG_NAME, 'td')
            if len(cols)>2:
              cols[0].click();print('라디오버튼 선택')
              choicebtn = driver.find_element(By.XPATH,'//*[@id="btnDlgpSelect"]');print('선택버튼')
              WebDriverWait(driver, 10).until(EC.element_to_be_clickable(choicebtn)).click();time.sleep(2)
              driver.switch_to.default_content()#메인으로 복귀

              table = driver.find_element(By.XPATH,'//*[@id="contents"]/div[3]/table')
              tbody = table.find_element(By.XPATH,'//*[@id="tblAot"]')
              for tr in tbody.find_elements(By.TAG_NAME, 'tr'):
                cols = tr.find_elements(By.TAG_NAME, 'td')
                if "납세증명서 발급 신청이 가능합니다." in cols[0].text:
                  #주소체크
                  if driver.find_element(By.XPATH,'//*[@id="divTxpInfo"]/div[2]/table/tbody/tr[4]/td/div/div/input').text == "":
                    print(f'{result[5]} : 주소없음')
                    element = driver.find_element(By.ID, "srchTxpAddrBtn")
                    driver.execute_script("arguments[0].click();", element)
                    iframe = driver.find_element(By.XPATH,'//*[@id="cmnPopup_addr"]/div/div/iframe'); driver.switch_to.frame(iframe);    time.sleep(1)
                    biz_addr1 = result[15].split(',')[0]
                    # if len(biz_addr1)<8:
                    biz_addr1 = "강남대로84길 15"
                    driver.find_element(By.XPATH,'//*[@id="ibx_search"]').send_keys(biz_addr1);time.sleep(1);print(biz_addr1)
                    driver.find_element(By.XPATH,'//*[@id="ibx_search"]').send_keys(Keys.ENTER);time.sleep(1)
                    table = driver.find_element(By.XPATH,'//*[@id="tblAddr"]')
                    tbody = table.find_element(By.XPATH, '//*[@id="tblAddr"]/tbody')
                    for tr in tbody.find_elements(By.TAG_NAME, 'tr'):
                      cols = tr.find_elements(By.TAG_NAME, 'td')
                      cols[0].click();time.sleep(0.3)
                      break
                    driver.find_element(By.XPATH,'//*[@id="btnConfirm"]').click();time.sleep(0.5)
                    driver.switch_to.default_content()#메인으로 복귀
                  # driver.find_element(By.XPATH,'//*[@id="divTxpInfo"]/div[2]/table/tbody/tr[3]/td/div/div[1]/input').clear();time.sleep(0.2)
                  # driver.find_element(By.XPATH,'//*[@id="divTxpInfo"]/div[2]/table/tbody/tr[3]/td/div/div[2]/input').clear();print("이메일클리어");time.sleep(0.2)
                  hpno = driver.find_element(By.XPATH,'//*[@id="divTxpInfo"]/div[2]/table/tbody/tr[2]/td[2]/input')
                  if hpno.text == '':
                    hpno.send_keys(result[16])
                  select = Select(driver.find_element(By.XPATH,'//*[@id="ctpv"]'));select.select_by_visible_text('서울특별시');time.sleep(0.5)
                  select = Select(driver.find_element(By.XPATH,'//*[@id="sgg"]'));select.select_by_visible_text('강남구');time.sleep(0.5)
                  select = Select(driver.find_element(By.XPATH,'//*[@id="issuUsgCd"]'));select.select_by_visible_text('기타목적');time.sleep(0.5)
                  driver.find_element(By.XPATH,'//*[@id="issuUsePurCn"]').send_keys("기타");time.sleep(0.5)
                  element = driver.find_element(By.XPATH,'//*[@id="btnTxpcIssu"]')
                  driver.execute_script("arguments[0].click();", element)
                  WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(2);print('발급신청하시겠습니까')
                  WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1);print('신청완료')
                  #문서발급 신청내역으로 이동
                  xpath = '//*[@id="spn_totCnt"]'
                  success = utils.wait_for_value(driver,xpath)
                  if success:
                    table = driver.find_element(By.XPATH,'//*[@id="tblListIssuDtcnInq"]')
                    tbody = table.find_element(By.XPATH, '//*[@id="tblListIssuDtcnInq"]/tbody')
                    for tr in tbody.find_elements(By.TAG_NAME, 'tr'):
                      cols = tr.find_elements(By.TAG_NAME, 'td')
                      cols[9].click();print('메인 출력버튼 클릭');time.sleep(5)
                      main = driver.window_handles;    print(main);    driver.switch_to.window(main[1]); time.sleep(3)
                      wetaxProof = pyautogui.locateOnScreen('K:/noa-django/Noa/static/assets/images/autowork/wetaxProof_wanap.png',confidence=0.8)
                      if wetaxProof:
                        pyautogui.click(wetaxProof)
                        pyautogui.hotkey('ctrl','p');time.sleep(3)
                        pyautogui.press('enter');time.sleep(3);print('인쇄 엔터')
                        dirProof = directory + result[5]+"/"+str(workyear)+"/홈택스민원서류";print(dirProof)
                        fileName = "지방세납세증명서"
                        utils.createDirectory(dirProof)    
                        fullFileName = dirProof +"/"+fileName+".pdf"
                        if  os.path.isfile(fullFileName.replace('/',"\\")):os.remove(fullFileName.replace('/',"\\"));print(fullFileName+" 삭제완료") ;time.sleep(1)
                        utils.Report_save(fileName,dirProof);print('파일저장완료');time.sleep(1)
                        pyautogui.hotkey('alt','f4')                      
                        strsql = "insert into 스크래핑관리 values('"+today+"','"+result[2]+"','"+result[5]+"','지방세 납세증명서')"       
                        cursor.execute(strsql)                      
                      main = driver.window_handles;    print(main);    driver.switch_to.window(main[0]);  
                      break 
                  else:
                    print(f"{result[5]} : 문서발급 신청내역 검색결과 없음")
                    break
                else:
                  print(result[5] + " : " + cols[0].text)    
                  break          
            else:
              print(f'{result[5]} : 위임자검색결과 없음')
              driver.switch_to.default_content()#메인으로 복귀
              driver.find_element(By.XPATH,'//*[@id="cmnPopup_dlgp"]/div/button').click();time.sleep(0.3)
            break
  elif flag=='8' or flag=='9':#고지/체납내역조회
    driver = utils.conHometaxLogin(htxLoginID,False);time.sleep(1)
    txtBigo = '고지세액조회';tableName = 'tbl_goji';linkAddr = 'https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/nf/a/a/UTENFAAA07.xml&menuId=48&subMenuId=03'
    if flag=='9':  txtBigo = '체납세액조회';tableName = 'tbl_chenap';linkAddr = 'https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/nf/a/a/UTENFAAA08.xml&menuId=48&subMenuId=03'
    for result in results: 
      driver.get(linkAddr);time.sleep(1)
      if len(result[11])>=10:
        strsql = "select * from 스크래핑관리 where crt_dt='" + today + "' and seqno_lastTry='"+result[2]+"' and bigo='"+txtBigo+"'"
        cursor = connection.cursor()
        cursor.execute(strsql)   
        scraps = cursor.fetchall()
        connection.commit()         
        if result[12]=='Y' and result[4]!=5 and result[6]!='화물' and not scraps:#홈택스 수임동의
          # driver.find_element(By.ID,'edtTxprNo').clear();time.sleep(0.5)
          driver.find_element(By.ID,'edtTxprNo').send_keys(result[11].replace("-",""));time.sleep(0.3); 
          button = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, 'trigger15')))
          driver.find_element(By.ID,'trigger15').click();      time.sleep(2) #조회하기 후 대기시간 필수
          try:
              WebDriverWait(driver, 1).until(EC.alert_is_present()); al = Alert(driver); al.accept();time.sleep(0.25)
              print(result[5] + ' - 조회시 알람 창이 발생하여 다음 루프로 진행 - 수임동의안됨')
              continue  
          except:
              pass              
          table = driver.find_element(By.ID,'grdList_body_table')
          tbody = table.find_element(By.ID,'grdList_body_tbody')  
          for i,tr in enumerate(tbody.find_elements(By.TAG_NAME,'tr')):
            cols = tr.find_elements(By.TAG_NAME,'td')
            if len(cols)>1:
              if i==0:
                str_del = "delete from "+tableName+" where seq_no="+ result[2]  + " and work_yy='" + str(workyear)  + "'  and work_mm='" + str(workmm)  + "'" 
                cursor.execute(str_del) 
              titleTax		= cols[1].get_attribute("innerText")#과세기간세목명
              elec_no			= cols[5].text
              due_date		= cols[8].text
              amt_Tax			= cols[9].text.replace(",","")
              if amt_Tax.strip()=='':amt_Tax='0'
              taxoffice		= cols[12].text
              if flag=='9':
                titleTax		= cols[0].text
                elec_no			= cols[4].text
                due_date		= cols[7].text
                amt_Tax			= cols[9].text.replace(",","")
                taxoffice		= cols[10].text               
              if titleTax == "종합소득세" and (due_date[:7]==str(workyear)+"-11" or due_date[:7]==str(workyear)+"-12" ) :
                if amt_Tax!='0' and amt_Tax.strip()!='':
                  str_mg = "Merge Tbl_income2 as A Using (select '"+result[2]+"' as seq_no, '"+str(workyear)+"' as work_yy ) as B "
                  str_mg += "On A.seq_no = B.seq_no and A.work_yy = B.work_yy  "
                  str_mg += "WHEN Matched Then   "
                  str_mg += "	Update set YN_4=" + amt_Tax
                  str_mg += "	When Not Matched Then  "
                  str_mg += "	INSERT (seq_no,work_YY,YN_1,YN_2,YN_3,YN_4,YN_5,YN_6,YN_7,YN_8,YN_9) "
                  str_mg += "	values('"+result[2]+"','"+str(workyear)+"',0,0,0,0,0,0,0,0,'');"
                  cursor = connection.cursor();print(str_mg)
                  cursor.execute(str_mg)				
              else:
                if amt_Tax!='0' and amt_Tax.strip()!='':
                  str_ins = "insert into "+tableName+" values('"+result[2]+"','"+result[5]+"','"+str(workyear)+"','"+str(workmm)+"','"+str(workyear)+"-"+ ("0"+str(workmm))[-2:] +"-"+ today[-2:]+"','','"+titleTax+"',"+amt_Tax+",'"+elec_no+"','"+taxoffice+"','"+due_date+"')"
                  cursor = connection.cursor();print(str_ins)
                  cursor.execute(str_ins)
            strsql = "insert into 스크래핑관리 values('"+today+"','"+result[2]+"','"+result[5]+"','"+txtBigo+"')"       
            cursor.execute(strsql)
  print(f'정상종료 : {flagFirst}:{flagManagerName} - {flag} - {flagSeqNo}' )
  return JsonResponse({'data':'성공 : 신고서작성'},safe=False) 


def SS_Corptax(request):
  workyear = datetime.now().year-1
  flag = request.GET.get('flag',False)
  flagGroup = "";flagBizNo="";flagSeqNo='';flagManagerName =''
  flagFirst = pyautogui.confirm('[통장정리 > 결산분개 체크된 업체가 진행됨] 작업형태 선택',buttons=['관리자별','단일사업자','전체'])
  if flagFirst=="관리자별":
    sqladmin = "select admin_Name,Htx_ID,admin_id,manage_YN from mem_admin  order by admin_name"
    cursor = connection.cursor()
    cursor.execute(sqladmin)
    admins = cursor.fetchall()
    connection.commit()
    connection.close()  
    arrAdmin = []
    for admin in admins:
      if admin[3]=="Y" or admin[3]=="S":
        arrAdmin.append(admin[2])
    flagManagerName = pyautogui.confirm('그룹작업 진행할 관리자를 선택하세요. ',buttons=arrAdmin)
    for admin in admins:
      if admin[0]==flagManagerName:  
        flagGroup += "'"+admin[2]+"'," 
  elif flagFirst=="단일사업자":
    flagSeqNo = pyautogui.prompt('SEQNO를 입력하세요.'); 
  #법인세/소득세마감 아닌업체 체크업체 리스트
  strsql = " select duzon_id,year(reg_date),a.seq_no,fiscalMM,biz_type,biz_name,biz_manager,ceo_name,ssn,replace(biz_no,'-','') from mem_user a,mem_deal b,tbl_corporate2 d "
  strsql += f" where a.seq_no=b.seq_no  and b.seq_no=d.seq_no  and work_yy={workyear}   and duzon_ID <> '' and b.keeping_YN = 'Y' and a.Del_YN <> 'Y' "
  if flagFirst=="관리자별":      
    strsql += " and biz_manager ='" + flagManagerName + "' and YN_6<>'1' "
  elif flagFirst=="단일사업자": 
    strsql += " and a.seq_no ='"+flagSeqNo+"'   "         
  else: 
    if flag=='2' or flag=='5':  strsql += " and biz_manager not in ('오피스텔','환급1') " 
    else        :               strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') "          
  if flag=='1' or flag=='3':#법인세신고 안내문 또는 부가세 예정신고인 경우 법인만
    strsql += " and a.biz_type<4 "
  elif flag=='2'  or flag=='5':#종합소득세 신고대상
    strsql += " and a.biz_type>=4 "
  strsql += "  order by a.seq_no/1 "
  print(strsql)    
  cursor = connection.cursor()
  cursor.execute(strsql)
  results = cursor.fetchall()
  connection.commit()
  connection.close()    
  for result in results:
    memuser = MemUser.objects.get(seq_no=result[2])
    
    strsql = "with  "
    strsql += " Total AS ( Select a.biz_no,a.reg_date,a.biz_name,a.seq_no,admin_id,d.YN_1,d.YN_2        from mem_user a , mem_deal  b  , mem_admin  c , tbl_mng_jaroe d "
    strsql += "			where a.seq_no=b.seq_no and b.biz_manager=c.admin_id  and b.seq_no=d.seq_no and keeping_yn='Y'  and fiscalMM=12 and year(a.Reg_Date)<='"+str(workyear)+"' and admin_id<>'환급1' "
    if flagFirst=="단일사업자":
      strsql += "			  and a.seq_no ='"+result[2]+"' )  "
    else:#단일사업자 아닌 경우는 결산분개 체크된 경우 진행
      strsql += "			 and work_yy='"+str(workyear)+"' and work_mm=13 and yn_9=1 and a.seq_no ='"+result[2]+"' )  "   
                           
    strsql += ", SALE AS ( select seq_no, sum(tranamt_Dr) as 금액 from ds_slipledgr2  WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd>=401 and acnt_cd<=430 GROUP BY seq_no )  "
    strsql += ", BENE as (SELECT  seq_no,sum(tranamt_dr-tranamt_cr) as '금액'  from ds_slipledgr2 where work_yy = '"+str(workyear)+"' and (acnt_cd>400) and Remk <> '손익계정에 대체' group by seq_no) "
    strsql += ", TAA AS ( select seq_no, sum(tranamt_cr) as 금액 from ds_slipledgr2  WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd=817   GROUP BY seq_no )  "
    strsql += ", SAA AS ( select seq_no, sum(tranamt_cr) as 금액 from ds_slipledgr2  WITH (NOLOCK) where  work_yy='"+str(workyear)+"'  and (remk like '%가산%' or remk like '%벌금%' or remk like '%과태%' or remk like '%경찰%' or remk like '%법인세%') GROUP BY seq_no )  "
    strsql += ", BB AS  ( select seq_no, sum(tranamt_cr) as 금액 from ds_slipledgr2 WITH (NOLOCK) where work_yy='"+str(workyear)+"' and (acnt_cd=813 or acnt_cd=613 or acnt_cd=513 or acnt_cd=713) and tranamt_cr>0 GROUP BY seq_no  )  "
    strsql += ", CC AS  ( select seq_no, isnull(sum(tranamt_cr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd=953   GROUP BY seq_no  )  "
    strsql += ", DD AS  ( select seq_no, isnull(sum(tranamt_cr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd=835   GROUP BY seq_no  )  "
    strsql += ", EE AS  ( select seq_no, isnull(sum(tranamt_cr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd=951  GROUP BY seq_no  )  "
    strsql += ", FF AS  ( select seq_no, isnull(sum(tranamt_Dr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and (acnt_cd=260 or acnt_cd=293)  GROUP BY seq_no  )  "
    strsql += ", GG AS  ( select seq_no, isnull((sum(tranamt_Cr)-sum(tranamt_Dr)),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd=134  GROUP BY seq_no  )  "
    strsql += ", HH AS  ( select seq_no, isnull(sum(tranamt_Dr)-sum(tranamt_Cr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd=257  GROUP BY seq_no  )  "
    strsql += ", JJ AS  ( select seq_no, isnull(sum(tranamt_Cr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd=136  GROUP BY seq_no  )  "
    strsql += ", KK AS  ( select seq_no, isnull(sum(tranamt_Cr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd=116  GROUP BY seq_no  ) "
    strsql += ", LL AS  ( select seq_no, isnull(sum(tranamt_Cr)-sum(tranamt_Dr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd>=195 and acnt_cd<=230  GROUP BY seq_no) "
    strsql += ", MM AS  ( select seq_no, isnull(sum(tranamt_Cr)-sum(tranamt_Dr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd=208  GROUP BY seq_no  ) "
    strsql += ", NN AS  ( select seq_no, isnull(sum(tranamt_Cr)-sum(tranamt_Dr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd>=146 and acnt_cd<=175  GROUP BY seq_no  ) "
    strsql += ", PP AS  ( select seq_no, isnull(sum(tranamt_Cr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd=201  GROUP BY seq_no  ) "
    strsql += ", QQ AS  ( select seq_no, isnull(sum(tranamt_Cr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and acnt_cd=818  GROUP BY seq_no  ) "
    strsql += ", TT as (select seq_no, isnull(sum(a60),0) as '금액'  from mem_user a,원천세전자신고 b where a.biz_no=b.사업자등록번호  and left(과세연월,4)='"+str(workyear)+"' and a60>0 group by a.seq_no) "
    strsql += ", VV as ( select seq_no,max(Tran_Dt) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"'   GROUP BY seq_no  )  "
    strsql += ", RR AS  ( select seq_no, isnull(sum(tranamt_Cr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"' and Tran_dt='12-31' and acnt_cd='400'  GROUP BY seq_no  ) "
    strsql += ", SS as (select  M.seq_no,isnull(sum(StckH_FEquityGP),0) as '금액' from Tbl_StckHListTrn T,mem_user M where M.seq_no=T.seq_no and year(tran_dt)='"+str(workyear)+"' and year(M.reg_date)<"+str(workyear)+" group by M.seq_no) "
    strsql += ", WW as ( select seq_no,isnull(sum(tranamt_Dr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"'  and acnt_cd=901  GROUP BY seq_no  )  "
    strsql += ", XX as ( select seq_no,isnull(sum(tranamt_Dr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"'  and acnt_cd=903  GROUP BY seq_no  )  "
    strsql += ", YY as ( select seq_no,isnull(sum(tranamt_Dr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"'  and acnt_cd=910  GROUP BY seq_no  )  "
    strsql += ", ZZ as ( select seq_no,isnull(sum(tranamt_Cr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"'  and acnt_cd=955  GROUP BY seq_no  )  "
    strsql += ", AB as ( select seq_no,isnull(sum(tranamt_Cr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"'  and acnt_cd=998  GROUP BY seq_no  )  "
    strsql += ",AC as (select 사업자번호,isnull(trim(결손금누계),0) as '금액' from tbl_EquityEval where  left(사업연도말,4)='"+str(workyear-1)+"' )  "
    strsql += ",AD as ( select seq_no,isnull(sum(tranamt_Cr),0) as '금액' from ds_slipledgr2 WITH (NOLOCK) where  work_yy='"+str(workyear)+"'  and acnt_cd>500 and acnt_cd<800  GROUP BY seq_no) "
    strsql += "SELECT "
    strsql += "replace(Total.biz_no,'-',''),  year(getdate())-year(Total.reg_date),  Total.biz_name ,Total.seq_no "
    strsql += ",ISNULL(SALE.금액, '0')  AS '매출' " #4
    strsql += ",ISNULL(BENE.금액, '0') AS '순손익' " #5
    strsql += ",ISNULL(TAA.금액, '0')  AS 'T817' " #6
    strsql += ",ISNULL(SAA.금액, '0')  AS 'S817' " #7
    strsql += ",ISNULL(BB.금액, '0')  '접대비'  " #8
    strsql += ",ISNULL(CC.금액, '0')  '기부금' " #9
    strsql += ",ISNULL(DD.금액, '0')  '대손상각' " #10
    strsql += ",ISNULL(EE.금액, '0')  '이자비용'  "  #11
    strsql += ",ISNULL(FF.금액, '0')  '차입금' " #12
    strsql += ",ISNULL(GG.금액, '0')  '가지급' " #13
    strsql += ",ISNULL(HH.금액, '0')  '가수금' " #14
    strsql += ",ISNULL(JJ.금액, '0')  '선납세금' " #15
    strsql += ",ISNULL(KK.금액, '0')  '미수수익' " #16
    strsql += ",ISNULL(LL.금액, '0')  '고정자산' " #17
    strsql += ",ISNULL(QQ.금액, '0')  '감가상각' " #18
    strsql += ",ISNULL(MM.금액, '0')  '차량' " #19
    strsql += ",ISNULL(NN.금액, '0')  '재고자산' "  #20
    strsql += ",ISNULL(PP.금액, '0')  '토지' " #21
    strsql += ",ISNULL(TT.금액, '0')  '중간배당' " #22
    strsql += ",ISNULL(VV.금액, '0')  '최종분개' " #23
    strsql += ",ISNULL(RR.금액, '0')  '결산분개' " #24
    strsql += ",ISNULL(SS.금액, '0')  '주식변동' " #25
    strsql += ",ISNULL(WW.금액, '0')  '이자수익' " #26  
    strsql += ",ISNULL(XX.금액, '0')  '배당금수익' " #27 
    strsql += ",ISNULL(YY.금액, '0')  '외화환산이익' " #28
    strsql += ",ISNULL(ZZ.금액, '0')  '외화환산손실' " #29
    strsql += ",ISNULL(AB.금액, '0')  '법인세비용' " #30
    strsql += ",ISNULL(AC.금액, '0')  '이월결손금'" # 31
    strsql += ",ISNULL(trim(Total.YN_1), '0')  '고용증가인원'" # 32
    strsql += ",ISNULL(trim(Total.YN_2), '0')  '세액감면율'" # 33
    strsql += ",ISNULL(AD.금액, '0')  '원가' "#34
    strsql += " FROM Total  "
    strsql += "  Left outer join SALE   On Total.seq_no = SALE.seq_no "
    strsql += "  Left outer join BENE On Total.seq_no = BENE.seq_no "
    strsql += "  Left outer join TAA   On Total.seq_no = TAA.seq_no "
    strsql += "  Left outer join SAA   On Total.seq_no = SAA.seq_no "
    strsql += "  Left outer join BB  On Total.seq_no = BB.seq_no "
    strsql += "  Left outer join CC  On Total.seq_no = CC.seq_no "
    strsql += "  Left outer join DD  On Total.seq_no = DD.seq_no "
    strsql += "  Left outer join EE  On Total.seq_no = EE.seq_no "
    strsql += "  Left outer join FF  On Total.seq_no = FF.seq_no "
    strsql += "  Left outer join GG  On Total.seq_no = GG.seq_no "
    strsql += "  Left outer join HH  On Total.seq_no = HH.seq_no "
    strsql += "  Left outer join JJ  On Total.seq_no = JJ.seq_no "
    strsql += "  Left outer join KK  On Total.seq_no = KK.seq_no "
    strsql += "  Left outer join LL  On Total.seq_no = LL.seq_no "
    strsql += "  Left outer join MM  On Total.seq_no = MM.seq_no "
    strsql += "  Left outer join NN  On Total.seq_no = NN.seq_no "
    strsql += "  Left outer join PP  On Total.seq_no = PP.seq_no "
    strsql += "  Left outer join QQ  On Total.seq_no = QQ.seq_no "
    strsql += "  Left outer join RR  On Total.seq_no = RR.seq_no "
    strsql += "  Left outer join SS  On Total.seq_no = SS.seq_no "
    strsql += "  Left outer join TT  On Total.seq_no = TT.seq_no "
    strsql += "  Left outer join VV  On Total.seq_no = VV.seq_no "
    strsql += "  Left outer join WW  On Total.seq_no = WW.seq_no "  
    strsql += "  Left outer join XX  On Total.seq_no = XX.seq_no "
    strsql += "  Left outer join YY  On Total.seq_no = YY.seq_no "  
    strsql += "  Left outer join ZZ  On Total.seq_no = ZZ.seq_no " 
    strsql += "  Left outer join AB  On Total.seq_no = AB.seq_no " 
    strsql += "  Left outer join AD  On Total.seq_no = AD.seq_no  "
    strsql += "  Left outer join AC  On Total.biz_no = AC.사업자번호 "
    strsql += " ORDER BY Total.admin_id, Total.biz_name    "
    cursor = connection.cursor()
    cursor.execute(strsql)
    member = cursor.fetchall()
    connection.commit()
    connection.close()       
    if member:
      print('조정대상회사 : ' + member[0][2])
      preFilelist = []
      directory = "D:/"+member[0][2]+"/"+str(workyear)+"/세무조정계산서"
      filelist = utils.createDirectory(directory)

      isAcntChange="N"#삭제/분개장업로드
      confirm = pyautogui.confirm(f'[{memuser.biz_name}]\n\n 회계처리 변동이 있는 경우 분개장을 업로드 하세요. 작성된 법인세 신고서도 삭제됩니다.\n\n[중단] 현 상태로 세무조정 중단',buttons=['모두삭제(표지등제외)/분개장업로드','주요재무제표삭제/분개장업로드','[계속] 분개장 업로드 없이 신고서 작성','중단(현재업체)'])
      if confirm=='모두삭제(표지등제외)/분개장업로드': 
        isAcntChange="Y" ;time.sleep(1);utils.delete_CorpSheet_files(flag,directory)
      elif confirm=='주요재무제표삭제/분개장업로드'  : 
        isAcntChange="Y" ;time.sleep(1);utils.delete_CorpSheet_998_files(flag,directory)
      elif confirm=='중단(현재업체)':continue

      
      dig = utils.semusarang_LaunchProgram_App(semusarangID)
      if result[0]=='1':  utils.semusarang_ChangeCompany(result[9])
      else:               utils.semusarang_ChangeCompany_ID_App(dig,result[0])          
      finalKi = workyear - memuser.reg_date.year + 1
      
      utils.semusarang_ChangeFiscalYear_App('vat',str(finalKi))
      if isAcntChange=="Y" :  #삭제/분개장업로드
        if  flag=='2'  or flag=='5':#종합소득세 신고대상인 경우 집합손익 회계처리
          utils.semusarang_Menu_Popup("손익계산서")
          pyautogui.press('enter');time.sleep(0.25)
          pyautogui.hotkey('ctrl','f5');pyautogui.press('enter',presses=2,interval=0.3);time.sleep(0.25);pyautogui.press('esc',presses=2,interval=0.3);time.sleep(0.25)
        print("분개장 진입");
        elecResult_Save.ACCOUNT.semusarang_Acct_Excel(result[2],result[1]+finalKi-1,result[3],result[4],'분개장') 

      if finalKi>1:
        preDirectory = "D:/"+member[0][2]+"/"+str(workyear-1)+"/세무조정계산서"
        preFilelist = utils.createDirectory(preDirectory)

      elecResult_Save.CORP.semusarang_Make_CorpSheet(flag,member[0],filelist,workyear,directory,preFilelist,finalKi,htxLoginID)
      if flag=='2':#개인
        print('종소세 전자신고 ')
        # elecResult_Save.SS_ElecIssue('income',result[8],workyear,result[2],1)             
    else:
      print(f'[{memuser.biz_name}]\n\n결산분개에 체크되지 않아 종료합니다.\n 1. 부가세예수금/대급금\n 2. 감가상각, 대손상각\n 3. 가지급금 인정이자')
  print(f'정상종료 : {flagFirst}:{flagManagerName} - {flag} - {flagSeqNo}' )
  return JsonResponse({'data':'성공 : 신고서작성'},safe=False) 


def Elec_Issue(request):
  workyear = datetime.now().year-1#연도는 일단 고정
  workmonth =  datetime.today().month
  flagYear = str(workyear)
  isJungKi = 1;#정기신고

  flagFirst = pyautogui.confirm(' 작업할 형태를 선택하세요',buttons=['관리자별','단일사업자','전체'])
  if flagFirst=="관리자별":
    sqladmin = "select admin_Name,Htx_ID,admin_id,manage_YN from mem_admin  order by admin_name"
    cursor = connection.cursor()
    cursor.execute(sqladmin)
    admins = cursor.fetchall()
    connection.commit()
    connection.close()  
    arrAdmin = []
    for admin in admins:
      if admin[3]=="Y" or admin[3]=="S":
        arrAdmin.append(admin[2])
    flagManagerName = pyautogui.confirm('그룹작업 진행할 관리자를 선택하세요. ',buttons=arrAdmin)
    for admin in admins:
      if admin[2]==flagManagerName:  
        flagGroup += "'"+admin[2]+"'," 
  elif flagFirst=="단일사업자":
    flagSeqNo = pyautogui.prompt('인트라넷 SEQNO를 입력하세요. ') 

  #단체인 경우는 어떻게 할래


  flag = request.GET.get('flag',False)
  if flag in ('1','2'):#법인세 / 법인지방세
    strsql = "select a.seq_no,biz_Name,replace(biz_No,'-',''),CONVERT(varchar(10), reg_date, 120),duzon_id,biz_no,YN_6,fiscalMM from  mem_user a, mem_deal b, tbl_corporate2 d "
    strsql += " where a.seq_no=b.seq_no and d.seq_no=b.seq_no and a.biz_no<>'' and keeping_yn='Y' and biz_manager<>'' and biz_type <4 "
    strsql += f" and work_yy={flagYear}  "
    if flagFirst=="관리자별":           strsql += " and biz_manager ='" + flagManagerName + "' " 
    elif flagFirst=="단일사업자":       strsql += " and a.seq_no ='"+flagSeqNo+"'   "  
    else:                              strsql += " and biz_manager not in ('오피스텔','환급1') " 
    print(strsql)  
    cursor = connection.cursor()
    cursor.execute(strsql)
    result = cursor.fetchall()
    connection.commit()

    dig = utils.semusarang_LaunchProgram_App(semusarangID)  
    for member in result:
      memuser = MemUser.objects.get(seq_no=member[0].strip())
      memdeal = MemDeal.objects.get(seq_no=member[0].strip())      
      if member[7]=="12" and workmonth>4: isJungKi=3#기한후신고
      kwaseKikan = f'{flagYear}년{member[7]}월';kwaseYouhyung = ""
      directory = "D:\\"+member[1]+"\\"+flagYear+"\\세무조정계산서"      
      #결재완료에 체크 안돼 있으면 전자신고 되어 있어도 다시 전자신고 진행한다...
      if member[6]=='1' or flagFirst=="단일사업자":
        finalKi = int(flagYear)-int(member[3][:4])+1
        if member[4]=='1':  utils.semusarang_ChangeCompany(member[2])
        else:               utils.semusarang_ChangeCompany_ID_App(dig,member[4])        
        utils.semusarang_ChangeFiscalYear_App('vat',str(finalKi))    
        if flag=='1':
          if  os.path.exists(directory+"\\1.pdf"):
            #신고서파일이 있고(인트라넷 신고서칸에 체크) 결재완료에 체크 안된 경우 전자신고 진행, 결재완료는 진행제외시킬 때 사용
            print(f'{member[1]} 법인세 전자신고 : 1.pdf 있음')
            strsql = f"select 사용자ID from 법인세전자신고2 where 사업자번호='{member[5]}' and 과세년월='{flagYear}년{member[7]}월'"
            print(strsql)
            cursor = connection.cursor()
            cursor.execute(strsql)
            userID = cursor.fetchone()
            connection.commit()
            connection.close()    
            tmpLoginID = htxLoginID
            if userID:          tmpLoginID = userID
            elecResult_Save.ElecIssue.SS_MakeElecFile('corp',member[0],member[4],"","",kwaseKikan,kwaseYouhyung,isJungKi,tmpLoginID)     
            elecResult_Save.SS_ElecIssue('corp',memuser.seq_no.replace("-",""),workyear,flagSeqNo,1) 

            # 지방세 전자신고
            if  os.path.exists(directory+"\\92-지-43.pdf"):            
              elecResult_Save.ElecIssue.SS_MakeElecFile('corp-wetax',member[0],member[4],"","",kwaseKikan,kwaseYouhyung,isJungKi,tmpLoginID)     
              elecResult_Save.SS_ElecIssue('corp-wetax',memuser.seq_no.replace("-",""),workyear,f"{workyear}{memdeal.fiscalmm}",1)             
            else:
              print('92-지-43.pdf가 없습니다.')   	 
          else:
            print('1.pdf가 없습니다.')           
        elif flag=='2':
          if  os.path.exists(directory+"\\92-지-43.pdf"):            
            elecResult_Save.ElecIssue.SS_MakeElecFile('corp-wetax',member[0],member[4],"","",kwaseKikan,kwaseYouhyung,isJungKi,1)     
            elecResult_Save.SS_ElecIssue('corp-wetax',memuser.seq_no.replace("-",""),workyear,f"{workyear}{memdeal.fiscalmm}",1)             
          else:
            print('92-지-43.pdf가 없습니다.')   
      else:
        print('결재완료 체크바랍니다.')
  elif flag in ('3','4'):#종소세 / 지방소득세
    strsql = "select a.seq_no,biz_Name,replace(biz_No,'-',''),CONVERT(varchar(10), reg_date, 120),duzon_id,biz_no,YN_6,fiscalMM from  mem_user a, mem_deal b, tbl_income2 d "
    strsql += " where a.seq_no=b.seq_no and d.seq_no=b.seq_no and  keeping_yn='Y' and biz_manager<>'' and biz_type >=4 "
    strsql += f" and work_yy={flagYear}  "
    if flagFirst=="관리자별":           strsql += " and biz_manager ='" + flagManagerName + "' " 
    elif flagFirst=="단일사업자":       strsql += " and a.seq_no ='"+flagSeqNo+"'   "  
    else:                              strsql += " and biz_manager not in ('오피스텔','환급1') " 
    print(strsql)  
    cursor = connection.cursor()
    cursor.execute(strsql)
    result = cursor.fetchall()
    connection.commit()

    dig = utils.semusarang_LaunchProgram_App(semusarangID)  
    for member in result:
      memuser = MemUser.objects.get(seq_no=member[0].strip())
      memdeal = MemDeal.objects.get(seq_no=member[0].strip())   
      if workmonth>6: isJungKi=3#기한후신고
      kwaseKikan = f'{flagYear}년{member[7]}월';kwaseYouhyung = ""
      directory = "D:/"#+member[1]+"\\"+flagYear+"\\세무조정계산서" ;print(directory)
      savefile=fileName=""
      if memdeal.biz_manager=='화물' or memdeal.biz_manager[:3]=='종소세': 
        directory += "AAA/종합소득세/" + str(workyear) +"/" + memdeal.biz_manager 
        savefile = directory + "/"+memuser.ceo_name+"("+memuser.ssn[:6]+")-";fileName = memuser.ceo_name+"("+memuser.ssn[:6]+")-"
      else:                                           
        directory += memuser.biz_name+"/" + str(workyear) +"/세무조정계산서"   
        savefile = directory + "/"        
      #결재완료에 체크 안돼 있으면 전자신고 되어 있어도 다시 전자신고 진행한다...
      print(directory)
      print(savefile)
      if member[6]=='1' or flagFirst=="단일사업자":
        finalKi = int(flagYear)-int(member[3][:4])+1
        if member[4]=='1':  utils.semusarang_ChangeCompany(member[2])
        else:               utils.semusarang_ChangeCompany_ID_App(dig,member[4])        
        utils.semusarang_ChangeFiscalYear_App('vat',str(finalKi))    
        if flag=='3':
          if  os.path.exists(savefile+"1.pdf"):
            #신고서파일이 있고(인트라넷 신고서칸에 체크) 결재완료에 체크 안된 경우 전자신고 진행, 결재완료는 진행제외시킬 때 사용
            print(f'{member[1]} 종소세 전자신고 : 1.pdf 있음')
            strsql = f"select 제출자 from 종합소득세전자신고2 where 이름='{memuser.ceo_name}' and left(주민번호,6)='{memuser.ssn[:6]}' and 과세년월='{flagYear}년1월'"
            print(strsql)
            cursor = connection.cursor()
            cursor.execute(strsql)
            userID = cursor.fetchone()
            connection.commit()
            connection.close()    
            tmpLoginID = htxLoginID
            if userID:          tmpLoginID = userID
            elecResult_Save.ElecIssue.SS_MakeElecFile('income',member[0],member[4],"","",kwaseKikan,kwaseYouhyung,isJungKi,tmpLoginID)        
            elecResult_Save.SS_ElecIssue('income',memuser.seq_no.replace("-",""),workyear,flagSeqNo,1) 
            # 소득세 신고시 지방세 전자신고는 소득세 신고와 연동하여 진행
          else:
            print(f"{member[1]} : 신고서파일이 없습니다.")
        elif flag=='4':
          elecResult_Save.ElecIssue.SS_MakeElecFile('income',member[0],member[4],"","",kwaseKikan,kwaseYouhyung,isJungKi,htxLoginID)        
          elecResult_Save.ElecIssue.SS_MakeElecFile('income-wetax',member[0],member[4],"","",kwaseKikan,kwaseYouhyung,isJungKi,1)                 
          elecResult_Save.SS_ElecIssue('income-wetax',memuser.seq_no.replace("-",""),workyear,flagSeqNo,1)
  elif flag=='6' or flag=='7':
    print('원천세 전자신고는 SS_ElecIssue에서 진행')
  return JsonResponse({'data':'성공'},safe=False)

#홈택스 조회서비스
def Htx_Report(request):
    workyear = datetime.now().year
    workmm = datetime.now().month
    work_qt = '';work_yy=str(workyear) 
    today = str(datetime.now())[:10].replace("-","")
    flag = request.GET.get('flag',False)
    flagGroup = flagSeqNo=flagManagerName =flagFirst = flagBigo = ''
    switcher = {
        1: "법인세신고안내(98.pdf)",
        2: "종소세신고안내(400.pdf)",
        3: "부가세신고안내(300.pdf)",
        5: "종소세대비 연금건강고용산재보험료조회",
        6: "종소세예정고지(204.pdf)",
        7: "부가세예정고지(200.pdf)",
    }
    flagBigo = switcher.get(int(flag),"")
    flagFirst = pyautogui.confirm(' 스크래핑 작업할 형태를 선택하세요',buttons=['관리자별','단일사업자','전체'])
    if flagFirst=="관리자별":
      sqladmin = "select admin_Name,Htx_ID,admin_id,manage_YN from mem_admin  order by admin_name"
      cursor = connection.cursor()
      cursor.execute(sqladmin)
      admins = cursor.fetchall()
      connection.commit()
      connection.close()  
      arrAdmin = []
      for admin in admins:
        if admin[3]=="Y" or admin[3]=="S":
          arrAdmin.append(admin[2])
      flagManagerName = pyautogui.confirm('그룹작업 진행할 관리자를 선택하세요. ',buttons=arrAdmin)
      for admin in admins:
        if admin[0]==flagManagerName:  
          flagGroup += "'"+admin[2]+"'," 
    elif flagFirst=="단일사업자":
      flagSeqNo = pyautogui.prompt('사업자등록번호를 입력하세요. (-)포함')

    if flag=='3' and (workmm==7 or workmm==1): flag='4'  #부가세신고안내 : 부가세 예정기간이 아니면 flag변경
    if flag=='7' and (workmm==7 or workmm==1): flag='8'  #부가세예정신고 : 부가세 예정기간이 아니면 flag변경

    if flag=='1':#법인세신고 안내문 - 팝업창 해결
      main = driver.window_handles;    print(main);    driver.switch_to.window(main[1]); time.sleep(1)
      driver.find_element(By.CSS_SELECTOR,'#group2152').click()
      driver.switch_to.window(main[0])
  # try:
    strsql = " select  biz_no,biz_name,year(reg_date),a.seq_no,fiscalMM,biz_type,biz_manager,ceo_name,ssn,hometaxAgree from mem_user a,mem_deal b "
    strsql += " where a.seq_no=b.seq_no  and keeping_yn='Y'  "
    if flagFirst=="관리자별":     
      strsql += " and biz_manager ='" + flagManagerName + "' "
    elif flagFirst=="단일사업자": 
      strsql += " and a.seq_no ='"+flagSeqNo+"'   "      
    else: 
      if flag=='2' or flag=='5' or flag=='6':  strsql += " and biz_manager not in ('오피스텔','환급1') " 
      else        :               strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') "   

    if flag=='1' or flag=='3':#법인세신고 안내문 또는 부가세 예정신고인 경우 법인만
      strsql += " and a.biz_type<4  and kijang_yn='Y' "
    elif flag=='2'  or flag=='5' or flag=='6':#종합소득세 신고대상
      strsql += " and a.biz_type>=4 "
      if flag=='6': strsql += " and kijang_yn='Y' "
    elif flag=='7'  or flag=='8':#부가세 예정고지 조회 - 개인사업자만,, 법인 고지는 어떻게 하나
      strsql += " and a.biz_type>=4  and kijang_yn='Y' "
    strsql += "  order by a.seq_no/1 "
    print(strsql)
    cursor = connection.cursor()
    cursor.execute(strsql)
    result = cursor.fetchall()
    connection.commit()
    connection.close()
    loginID = htxLoginID
    driver = utils.conHometaxLogin(loginID,False);time.sleep(2)
    driver.find_element(By.ID,'mf_wfHeader_hdGrp920').click();print('세무대리')
    if flagBigo== "법인세신고안내(98.pdf)":
      driver.find_element(By.ID,'span_a_4803080000').click();time.sleep(0.5);print(flagBigo) 
    elif flagBigo== "종소세신고안내(400.pdf)":
      driver.find_element(By.ID,'span_a_4803050000').click();time.sleep(0.5);print(flagBigo) 
    elif flagBigo== "부가세신고안내(300.pdf)":
      driver.find_element(By.ID,'span_a_4801050000').click();time.sleep(0.5);print(flagBigo) 
    elif flagBigo== "종소세대비 연금건강고용산재보험료조회":
      driver.find_element(By.ID,'span_a_4803060000').click();time.sleep(0.5);print(flagBigo) 
    elif flagBigo== "종소세예정고지(204.pdf)":
      driver.find_element(By.ID,'span_a_4803040000').click();time.sleep(0.5);print(flagBigo) 
    elif flagBigo== "부가세예정고지(200.pdf)":
      driver.find_element(By.ID,'span_a_4801030000').click();time.sleep(0.5);print(flagBigo) 
       
    for member in result:
      strsql = f"select * from 스크래핑관리 where crt_dt='{today}' and seqno_lastTry='{member[3]}' and bigo='{flagBigo}'";print(strsql)
      cursor = connection.cursor()
      cursor.execute(strsql)   
      scraps = cursor.fetchall()
      connection.commit()  
      if  not scraps:  
        directory = "D:/"
        if flag!='2' and  flag!='5':  directory += member[1]+"/"
        file = directory;fileName="";printBtnName = ""
        print("진입 : "+directory)
        if flag=='1':
          directory += str(workyear-1) + "/세무조정계산서"   ;print(directory)
          file = directory + "/98.pdf";fileName='98';printBtnName = 'group7980'
        elif flag=='2' or flag=='5' or flag=='6':
          if member[6]=='화물' or member[6][:3]=='종소세': 
            directory += "AAA/종합소득세/" + str(workyear-1) +"/" + member[6] 
            file = directory + "/"+member[7]+"("+member[8][:6]+")-"+"400.pdf"  ;fileName=member[7]+"("+member[8][:6]+")-"+"400";printBtnName = 'trigger4'#미리보기group6755
          else:                                           
            if flag=='2' or flag=='5' :    
              directory += member[1]+"/" + str(workyear-1)+ "/세무조정계산서"        
              file = directory + "/400.pdf"   ;fileName='400';printBtnName = 'trigger4' #group6755
            elif flag=='6'            :    
              directory += str(workyear) +"/세무조정계산서"        
              file = directory + "/204.pdf"   ;fileName='204'
        elif flag=='3' or flag=='4' or flag=='7' or flag=='8': #부가세 신고안내문(3,4) 부가세 예정고지(7,8)
          if workmm==1:    directory += str(workyear-1) + "/부가세/2기확정" ;work_qt = '4';work_yy=str(workyear-1) 
          elif workmm<=4:    directory += str(workyear) + "/부가세/1기예정" ;work_qt = '1'
          elif workmm<=7:  directory += str(workyear) + "/부가세/1기확정" ;work_qt = '2'
          elif workmm<=10: directory += str(workyear) + "/부가세/2기예정" ;work_qt = '3'
          elif workmm<=12: directory += str(workyear) + "/부가세/2기확정" ;work_qt = '4'
          if flag=='3' or flag=='4':file = directory + "/300.pdf" ;print(file);fileName='300';printBtnName = 'group9010749'  
          elif flag=='7':           file = directory + "/200.pdf" ;print(file);fileName='200';printBtnName = 'group9010749'  
        print("진입2 : "+directory+file)
        if not os.path.isfile(file.replace('/',"\\")):#파일이 없으면 조회한다
            bizNo=member[0].split('-');print('진입3 : '+member[1]+str(member[3]));print(bizNo)
            ssn = member[8]
            if flag=='1': #법인세 신고도움 서비스              
              driver.get('https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/rn/a/b/a/b/UTERNAAA98.xml&menuId=48&subMenuId=03');      time.sleep(1)
              # driver.find_element(By.ID,'edtTxprDscmNo1').clear();time.sleep(0.3)
              driver.find_element(By.ID,'edtTxprDscmNo1').send_keys(bizNo[0]);time.sleep(0.5)
              # driver.find_element(By.ID,'edtTxprDscmNo2').clear();time.sleep(0.3)
              driver.find_element(By.ID,'edtTxprDscmNo2').send_keys(bizNo[1]);time.sleep(0.5)
              # driver.find_element(By.ID,'edtTxprDscmNo3').clear();time.sleep(0.3)
              driver.find_element(By.ID,'edtTxprDscmNo3').send_keys(bizNo[2]);time.sleep(1)   
              inputyear=""
              while inputyear==str(workyear-1):
                inputyear = driver.find_element(By.XPATH,'//*[@id="input3"]').text;time.sleep(1)
              time.sleep(2)
              driver.find_element(By.ID,'group2371').click();time.sleep(2);print('조회하기')     
            elif flag=='2':   #종소세 신고도움 서비스
              driver.get('https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/rn/a/a/a/a/UTERNAAT31.xml&menuId=48&subMenuId=03');time.sleep(1)
              WebDriverWait(driver, 60).until(EC.alert_is_present());al = Alert(driver);al.accept()
              prtTxt = "신고안내문 발송대상이 아닌 납세자의 수입금액 조회도 가능합니다."; print("alert 해제: "+member[1] + " - "+prtTxt) 
              iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe');    driver.switch_to.frame(iframe);    time.sleep(1);print('iframe txppIframe')
              # driver.find_element(By.ID,'input5').clear();time.sleep(0.3)
              driver.find_element(By.ID,'input5').send_keys(ssn[:6]);time.sleep(0.5)
              # driver.find_element(By.ID,'input6').clear();time.sleep(0.3)
              driver.find_element(By.ID,'input6').send_keys(ssn[6:13]);time.sleep(0.5)   
              driver.find_element(By.ID,'group6263').click();time.sleep(3);print('조회하기') 
              try:              
                WebDriverWait(driver, 25).until(EC.alert_is_present());  al = Alert(driver);  al.accept()
                prtTxt = "수임동의가 안된 경우 ";  print("alert 해제: "+member[1] + " - "+prtTxt)  
              except TimeoutException: 
                # wait = WebDriverWait(driver, 10)
                # clickable_element = wait.until(EC.element_to_be_clickable((By.ID, printBtnName)))
                WebDriverWait(driver, 61).until(EC.element_to_be_clickable((By.ID, printBtnName))).click();print('미리보기 버튼 나올때까지 기다려') 
                utils.Htx_Popup_Print(driver,printBtnName,fileName,directory,False)
                elecResult_Save.PAY.SS_Zkzs(workyear,driver,directory,fileName)#지급조서(원천징수영수증) 출력
            elif flag=='3' or flag=='4':#부가세 신고도움 서비스 
              driver.get('https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/rn/a/b/a/e/UTERNAA130.xml&menuId=48&subMenuId=01');      time.sleep(1)
              # driver.find_element(By.ID,'input15').clear();time.sleep(0.3)
              if workmm==1:driver.find_element(By.ID,'input15').send_keys(workyear-1);time.sleep(0.5)
              else        :driver.find_element(By.ID,'input15').send_keys(workyear);time.sleep(0.5)
              if workmm==4 or workmm==7:driver.find_element(By.XPATH,r'/html/body/div[1]/div[2]/div[4]/div/dl[1]/dd/select/option[1]').click();print('1기 선택')
              if workmm==10 or workmm==12:driver.find_element(By.XPATH,r'/html/body/div[1]/div[2]/div[4]/div/dl[1]/dd/select/option[2]').click();print('2기 선택')
              if workmm==4 or workmm==10:period = driver.find_element(By.CSS_SELECTOR,'#radio8_input_0');time.sleep(0.5);period.click() ;print('예정 선택')
              if workmm==1 or workmm==7:period = driver.find_element(By.CSS_SELECTOR,'#radio8_input_1');time.sleep(0.5);period.click() ;print('확정 선택')
              # driver.find_element(By.ID,'inputBsno').clear();time.sleep(0.3)
              driver.find_element(By.ID,'inputBsno').send_keys(bizNo);time.sleep(2)  
              driver.find_element(By.ID,'group901072').click();time.sleep(3);print('조회하기')                            
            elif flag=='6':#종소세 예정고지 조회/저장
              txtBigo = '종소세예정고지';tableName = 'Tbl_income2';linkName = 'a_a_4803010000'
              if len(ssn)>=13:
                strsql = "select * from 스크래핑관리 where crt_dt='" + today + "' and seqno_lastTry='"+member[3]+"' and bigo='"+txtBigo+"'"; print(strsql)
                cursor = connection.cursor()
                cursor.execute(strsql)   
                scraps = cursor.fetchall()
                connection.commit()         
                if member[9]=='Y'  and member[6]!='화물' and not scraps:#홈택스 수임동의       
                  driver.find_element(By.ID,'span_a_4803010000').click();time.sleep(0.5);print('고지내역조회')
                  driver.find_element(By.ID,'mf_txppWframe_edtTxprNo').send_keys(ssn);time.sleep(0.3); 
                  driver.find_element(By.ID,'mf_txppWframe_trigger15').click();      time.sleep(2) #조회하기 후 대기시간 필수
                  try:
                      WebDriverWait(driver, 1).until(EC.alert_is_present()); al = Alert(driver); al.accept();time.sleep(0.25)
                      print(member[1] + ' - 조회시 알람 창이 발생하여 다음 루프로 진행 - 수임동의안됨')
                      continue  
                  except:
                      pass              
                  table = driver.find_element(By.ID,'mf_txppWframe_grdList_body_table')
                  tbody = table.find_element(By.ID,'mf_txppWframe_grdList_body_tbody')  
                  for i,tr in enumerate(tbody.find_elements(By.TAG_NAME,'tr')):
                    cols = tr.find_elements(By.TAG_NAME,'td');print(len(tr.size))
                    if len(tr.size)>=2 and len(cols)>=30:
                      kyuljung    = cols[0].text;print(kyuljung)
                      titleTax		= cols[1].get_attribute("innerText")#과세기간세목명
                      elec_no			= cols[5].text;print(elec_no)
                      due_date		= cols[8].text;print(due_date)
                      amt_Tax			= cols[9].text.replace(",","");print(amt_Tax)
                      taxoffice		= cols[12].text;print(taxoffice)             
                      if kyuljung =='중간예납/예정고지' and titleTax == str(workyear)+"년 종합소득세" and (due_date[:7]==str(workyear)+"-11" or due_date[:7]==str(workyear)+"-12" ) :
                        str_mg = "Merge Tbl_income2 as A Using (select '"+member[3]+"' as seq_no, '"+str(workyear)+"' as work_yy ) as B "
                        str_mg += "On A.seq_no = B.seq_no and A.work_yy = B.work_yy  "
                        str_mg += "WHEN Matched Then  	Update set YN_4=" + amt_Tax
                        str_mg += "	When Not Matched Then  "
                        str_mg += "	INSERT (seq_no,work_YY,YN_1,YN_2,YN_3,YN_4,YN_5,YN_6,YN_7,YN_8,YN_9) "
                        str_mg += "	values('"+member[3]+"','"+str(workyear)+"',0,0,0,0,0,0,0,0,'');"
                        print(str_mg)
                        cursor.execute(str_mg)		
                        utils.Htx_Popup_Print(driver,'//*[@id="mf_txppWframe_grdList_cell_0_16"]',fileName,directory,False)		
                        break
                    strsql = "insert into 스크래핑관리 values('"+today+"','"+member[3]+"','"+member[1]+"','"+txtBigo+"')"    ;print(strsql)   
                    cursor.execute(strsql)
            elif flag=='7':#개인-부가세 예정고지 예정기간 납부서(200.pdf) 저장
              txtBigo = '고지세액조회';tableName = 'tbl_goji';linkName = 'a_a_4803010000'
              if len(member[0])==12:
                strsql = "select * from 스크래핑관리 where crt_dt='" + today + "' and seqno_lastTry='"+member[3]+"' and bigo='"+txtBigo+"'"; print(strsql)
                cursor = connection.cursor()
                cursor.execute(strsql)   
                scraps = cursor.fetchall()
                connection.commit()         
                if member[9]=='Y'  and member[6]!='화물' and not scraps:#홈택스 수임동의     
                  driver.get('https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/nf/a/a/UTENFAAA07.xml&menuId=48&subMenuId=03');time.sleep(1);print('고지내역조회')
                  # driver.find_element(By.ID,'edtTxprNo').clear();time.sleep(0.5)
                  driver.find_element(By.ID,'edtTxprNo').send_keys(member[0].replace("-",""));time.sleep(0.3); 
                  driver.find_element(By.ID,'trigger15').click();      time.sleep(2) #조회하기 후 대기시간 필수
                  try:
                      WebDriverWait(driver, 1).until(EC.alert_is_present()); al = Alert(driver); al.accept();time.sleep(0.25)
                      print(member[1] + ' - 조회시 알람 창이 발생하여 다음 루프로 진행 - 수임동의안됨')
                      continue  
                  except:
                      pass              
                  table = driver.find_element(By.ID,'grdList_body_table')
                  tbody = table.find_element(By.ID,'grdList_body_tbody')  
                  for i,tr in enumerate(tbody.find_elements(By.TAG_NAME,'tr')):
                    cols = tr.find_elements(By.TAG_NAME,'td');print(len(tr.size))
                    if len(tr.size)>=2 and len(cols)>=30:
                      kyuljung    = cols[0].text;print(kyuljung)
                      titleTax		= cols[1].get_attribute("innerText")#과세기간세목명
                      elec_no			= cols[5].text;print(elec_no)
                      due_date		= cols[8].text;print(due_date)
                      amt_Tax			= cols[9].text.replace(",","");print(amt_Tax)
                      taxoffice		= cols[12].text;print(taxoffice)     
                      strKi        = "1"
                      if workmm>6 : strKi        = "2"
                      if kyuljung =='중간예납/예정고지' and titleTax == work_yy+"년"+strKi+"기분 부가가치세"  :
                        str_mg = "Merge Tbl_vat as A Using (select '"+member[3]+"' as seq_no, '"+work_yy+"' as work_yy , '"+strKi+"' as work_qt ) as B "
                        str_mg += "On A.seq_no = B.seq_no and A.work_yy = B.work_yy  and A.work_qt = B.work_qt "
                        str_mg += "WHEN Matched Then  	Update set YN_15=" + amt_Tax
                        str_mg += "	When Not Matched Then  "
                        str_mg += "	INSERT (seq_no,work_YY,work_QT,YN_1,YN_2,YN_3,YN_4,YN_5,YN_6,YN_7,YN_8,YN_9,YN_10,YN_11,YN_12,YN_13,YN_14,YN_15,YN_16,YN_17) "
                        str_mg += f"	values('{member[3]}','{work_yy}','{strKi}',0,0,0,0,0,0,0,0,0,0,0,0,0,0,{amt_Tax},0,0);"
                        print(str_mg)
                        cursor.execute(str_mg)		
                        utils.Htx_Popup_Print(driver,'//*[@id="grdList_cell_0_16"]/span/button',fileName,directory,False)		
                        break
                    strsql = "insert into 스크래핑관리 values('"+today+"','"+member[3]+"','"+member[1]+"','"+txtBigo+"')"    ;print(strsql)   
                    cursor.execute(strsql)
            elif  flag=='8':#부가세 확정기간 :예정고지금액 저장 => 세무대리인:부가가치세 예정고지(부과) 조회
              driver.get('https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/rn/a/b/a/e/UTERNAA807.xml&menuId=48&subMenuId=01');time.sleep(1) 
              driver.find_element(By.ID,'trigger46').click();      time.sleep(2) #조회하기 버튼 클릭 후 대기시간 필수
              driver.find_element(By.ID,'trigger47').click();      time.sleep(2) #내려받기 버튼 클릭 ==> 엑셀파일 다운로드
              driver.close()
              isOK = utils.DBSave_Downloaded_xlsx(driver,[work_yy,work_qt],'부가세예정고지(확정)') 
              return 1
            try:
              WebDriverWait(driver, 1.5).until(EC.alert_is_present());  al = Alert(driver);  al.accept()
              prtTxt = "조회결과 없음.";  print("alert 해제: "+member[1] + " - "+prtTxt) 
            except TimeoutException:  
              print("조회됨")
              if flag!='6' and flag!='7' : utils.Htx_Popup_Print(driver,printBtnName,fileName,directory,False)
        else:#400파일이 있으면 원천세파일 제작
          print('400파일이 있는 경우')
          if flag=='2':  
            bizNo=member[0].split('-');print('진입:'+member[1]+str(member[3]));print(bizNo)
            ssn = member[8]            
            driver.get('https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/rn/a/a/a/a/UTERNAAT31.xml&menuId=48&subMenuId=03');  print('종소세 신고도움 서비스 클릭');time.sleep(1)
            WebDriverWait(driver, 60).until(EC.alert_is_present());al = Alert(driver);al.accept()
            prtTxt = "신고안내문 발송대상이 아닌 납세자의 수입금액 조회도 가능합니다."; print("alert 해제: "+member[1] + " - "+prtTxt) ;time.sleep(0.5)
            iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe');    driver.switch_to.frame(iframe);    time.sleep(1);print('iframe txppIframe')
            # driver.find_element(By.ID,'input5').clear();time.sleep(0.3)
            driver.find_element(By.ID,'input5').send_keys(ssn[:6]);time.sleep(0.5)
            # driver.find_element(By.ID,'input6').clear();time.sleep(0.3)
            driver.find_element(By.ID,'input6').send_keys(ssn[6:13]);time.sleep(0.5)   
            driver.find_element(By.ID,'group6263').click();time.sleep(3);print('조회하기') 
            try:              
              WebDriverWait(driver, 1.5).until(EC.alert_is_present());  al = Alert(driver);  al.accept()
              prtTxt = "안내문이 있는 경우 : 수임동의가 안된 경우 ";  print("alert 해제: "+member[1] + " - "+prtTxt)  
              WebDriverWait(driver, 1.5).until(EC.alert_is_present());  al = Alert(driver);  al.accept()
            except TimeoutException: 
              WebDriverWait(driver, 61).until(EC.element_to_be_clickable((By.ID, printBtnName)));print('미리보기 버튼 나올때까지 기다려') 
              elecResult_Save.PAY.SS_Zkzs(workyear,driver,directory,fileName)
              
          elif flag=='5':  #종소세대비 연금건강고용산재보험료조회
            driver.get('https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/rn/a/a/a/a/UTERNAAD71.xml&menuId=48&subMenuId=03');time.sleep(1)
            bizNo=member[0].replace('-','');print('진입:'+member[1]+str(member[3]));print(bizNo)
            ssn = member[8]              
            # driver.find_element(By.ID,'input8').clear();time.sleep(0.3)
            driver.find_element(By.ID,'input8').send_keys(ssn[:6]);time.sleep(0.5)
            # driver.find_element(By.ID,'secret1').clear();time.sleep(0.3)
            driver.find_element(By.ID,'secret1').send_keys(ssn[6:13]);time.sleep(0.5)   
            driver.find_element(By.ID,'group461').click();time.sleep(1);print('조회하기') 
            WebDriverWait(driver, 1.5).until(EC.alert_is_present());  al = Alert(driver);  al.accept()    
            prtTxt = "안내문이 있는 경우 :팝업1 - 조회하기후 사업자번호가 조회된경우/안된경우 ";  print("alert 해제: "+member[1] + " - "+prtTxt)     
            try:
              select = Select(driver.find_element(By.ID, 'selectbox1'))
              options = select.options
              for option in options:
                  print(option.text)
                  if option.text == bizNo:
                      select.select_by_visible_text('{}'.format(bizNo))
                      break   
              driver.implicitly_wait(5);driver.find_element(By.ID,'group461').click();time.sleep(1);print('조회하기') 
            except Exception as e:
              print('에러 발생:', e) 
            WebDriverWait(driver, 1.5).until(EC.alert_is_present());  al = Alert(driver);  al.accept();print('alert해제 - 조회결과')  
            
            max_num = 0;index_400 = 0;target_name = fileName.split('-')[0];print(target_name)
            if directory.find('AAA')!=-1:index_400 = 1
            matched_files = []
            for files2 in os.listdir(directory):
              if index_400 == 0:
                matched_files.append(files2)
              else:
                if files2.find(target_name) != -1:              matched_files.append(files2)
            for files in matched_files:
              fname, extension = os.path.splitext(files)
              if fname.split('-')[index_400].isdigit():
                if index_400==1 and fname.split('-')[0]==target_name:
                    num = int(fname.split('-')[index_400])
                    max_num = max(max_num, num)
                elif index_400==0:
                    num = int(fname.split('-')[index_400])
                    max_num = max(max_num, num)  
            try:
              table = driver.find_element(By.CSS_SELECTOR,'#pplHifeRgnSbsr_body_table');  print("테이블획득");  driver.implicitly_wait(30)    
              tbody = table.find_element(By.CSS_SELECTOR,'#pplHifeRgnSbsr_body_tbody'); print("티바디획득");print('조회할테이블수:'+str(len(tbody.find_elements(By.TAG_NAME,'tr'))))
              rows = tbody.find_elements(By.TAG_NAME, 'tr')
              for j, tr in enumerate(rows):
                td = tr.find_elements(By.TAG_NAME,"td")
                txtYear="";txtTitle="";txtBizName="";txtBizNo="";txtID="";saveName=str(max_num+1)
                if index_400==1:saveName=fileName.split('-')[0]+"-"+str(max_num+j+1)
                for i, tag in enumerate(td):
                  print(i)
                  id_td = tag.get_attribute('id');print(id_td)
                  txt_td = tag.get_attribute("innerText");print(txt_td)                
                  if i==0:txtTitle = txt_td.replace(' ','')
                  if i==1:txtID = id_td;
                if txtTitle=="지역가입자건강보험료":
                  if index_400==1:saveName=fileName.split('-')[0]+"-"+str(max_num+1)
                  saveName += "-연금건강고용산재보험료조회";print(saveName)
                  if not utils.find_similar_strings(matched_files, "-연금건강고용산재보험료조회"):      utils.Htx_Popup_Print(driver,txtID,saveName,directory,True)
            except NoSuchElementException:
              print(" 보험료 테이블이 없습니다")                   
            try:
              table = driver.find_element(By.CSS_SELECTOR,'#pplHifeBman_body_table');  print("테이블획득");  driver.implicitly_wait(30)    
              tbody = table.find_element(By.CSS_SELECTOR,'#pplHifeBman_body_tbody'); print("티바디획득");print('조회할테이블수:'+str(len(tbody.find_elements(By.TAG_NAME,'tr'))))
              rows = tbody.find_elements(By.TAG_NAME, 'tr')
              for j, tr in enumerate(rows):
                td = tr.find_elements(By.TAG_NAME,"td")
                txtYear="";txtTitle="";txtBizName="";txtBizNo="";txtID="";saveName=str(max_num+1)
                if index_400==1:saveName=fileName.split('-')[0]+"-"+str(max_num+j+1)
                for i, tag in enumerate(td):
                  print(i)
                  id_td = tag.get_attribute('id');print(id_td)
                  txt_td = tag.get_attribute("innerText");print(txt_td)                
                  if i==0:txtTitle = txt_td.replace(' ','')
                  if i==1:txtID = id_td;
                if  txtTitle=="사업장사용자부담건강보험료":
                  if index_400==1:saveName=fileName.split('-')[0]+"-"+str(max_num+1)
                  saveName += "-연금건강고용산재보험료조회";print(saveName)
                  if not utils.find_similar_strings(matched_files, "-연금건강고용산재보험료조회"):      utils.Htx_Popup_Print(driver,txtID,saveName,directory,True)
            except NoSuchElementException:
              print(" 보험료 테이블이 없습니다")
        strsql = f"insert into 스크래핑관리 values('{today}','{member[3]}','{member[1]}','{flagBigo}')"       
        cursor.execute(strsql)  
  # except:
  #   return JsonResponse({'data':'실패:홈택스 내부조회 중 에러'},safe=False)
  # finally:
  #   driver.quit() 
    print(f'정상종료 : {flagFirst}:{flagManagerName} - {flag} - {flagSeqNo}' ) 
    return JsonResponse({'data':'성공'},safe=False)

#기장보고서
def Ibotax_Report(request):
  work_dd = datetime.now().day
  work_mm = datetime.now().month
  workyear = datetime.now().year-1
  if work_mm>=4  :#4월 부터는 1분기 기장보고서 작성
    workyear = datetime.now().year
  
  work_qt=1;endMM = "1"
  if work_mm<4 : work_qt = 4;endMM = "12"
  elif work_mm==4 and work_dd>15: work_qt = 1;endMM = "03"
  elif work_mm<7 : work_qt = 1;endMM = "03"
  elif work_mm<10 : work_qt = 2;endMM = "06"
  elif work_mm<=12 : work_qt = 3;endMM = "09"
  flag = request.GET.get('flag',False)
  flagGroup = flagSeqNo=flagManagerName =flagFirst = ''
  #기장보고서
  if flag=="1":
    flagProc = pyautogui.confirm('기장보고서 작업을 시작할 단계를 지정하세요',buttons=['1.급여회계처리부터','2.분개장업로드부터','3.보고서만작성'])
    strsql = "select replace(biz_no,'-',''),year(reg_date),a.seq_no,fiscalMM,biz_type,biz_name "
    strsql += " ,(select isnull(sum(tranamt_cr),0) from ds_slipledgr2 d where d.seq_no=a.seq_no and work_yy="+str(workyear)+"  and acnt_cd>500 and acnt_cd<800) "
    strsql += " ,isnull((SELECT top 1 e.seq_no FROM Diag_Total e where a.seq_no=e.seq_no and work_yy="+str(workyear)+" ),'')"
    strsql += " ,Duzon_ID from mem_user a, mem_deal b where a.seq_no=b.seq_no and  keeping_yn='Y' "    
    flagFirst = pyautogui.confirm('기장보고서를 제작할 형태를 선택하세요',buttons=['관리자별','단일사업자','전체'])   
    if flagFirst=="관리자별":
      sqladmin = "select admin_ID from mem_admin where manage_YN='Y'  order by admin_id"
      cursor = connection.cursor()
      cursor.execute(sqladmin)
      admins = cursor.fetchall()
      connection.commit()
      connection.close()  
      arrAdmin = []
      for admin in admins:
        arrAdmin.append(admin[0])
      flagManagerName = pyautogui.confirm('그룹작업 진행할 관리자를 선택하세요. ',buttons=arrAdmin)  
      strsql += " and biz_manager='"+flagManagerName+"' "   
      flagSecond = pyautogui.confirm('기장보고서를 제작할 사업자 형태를 선택하세요. ',buttons=['법인','개인','전체'])  
      if flagSecond=='법인':  strsql += " and a.biz_type<4 "  
      elif flagSecond=='개인':  strsql += " and a.biz_type>=4 "  
    elif flagFirst=="단일사업자":
      flagSeqNo = pyautogui.prompt('제작할 회사의 SEQNO를 입력하세요.')   
      strsql += " and a.seq_no='"+flagSeqNo+"' "    
    cursor = connection.cursor()
    cursor.execute(strsql);print(strsql)
    members = cursor.fetchall()
    connection.commit()
    connection.close()
    for result in members:
      memuser = MemUser.objects.get(seq_no=result[2])
      finalKi = workyear - memuser.reg_date.year + 1
      if flagProc[:1]!="3": 
        dig = utils.semusarang_LaunchProgram_App(semusarangID)
        if result[8]=='1':  utils.semusarang_ChangeCompany(result[0])
        else:               utils.semusarang_ChangeCompany_ID_App(dig,result[8])          
        utils.semusarang_ChangeFiscalYear_App('vat',str(finalKi))
        if int(flagProc[:1])<2:
          #인사급여 회계처리
          elecResult_Save.ACCOUNT.semusarang_MakeAccount_Insa_B(result[2],workyear,endMM);time.sleep(1)
          #결산분개
          elecResult_Save.ACCOUNT.semusarang_KyulsanAccnt(endMM);time.sleep(1)
        else: 
          #분개장업로드
          elecResult_Save.ACCOUNT.semusarang_Acct_Excel(result[2],workyear,result[3],result[4],endMM);time.sleep(1)
      #저장디렉토리 설정   
      directory = "D:/"+result[5]+"/"+str(workyear)+"/기장보고서/"+str(work_qt)+"분기";
      if os.path.isdir(directory): utils.delete_all_files(directory);      print(directory+" 폴더내 파일 지우기")
      else: utils.createDirectory(directory) 
      #0.pdf
      driver = utils.ChromeDriver(False) ;time.sleep(1)
      driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/Report_cover.asp?seq_no='+result[2]);time.sleep(2)
      pyautogui.moveTo(0,0);time.sleep(1);pyautogui.click(x=501,y=223);time.sleep(1);print('프린트 스크립트 0');
      pyautogui.press('enter');time.sleep(1);print('모달에서 인쇄버튼')
      utils.Report_save('0',directory)  
      #1.pdf
      driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/Report_companyInfo.asp?seq_no='+result[2]);time.sleep(2)
      pyautogui.moveTo(0,0);time.sleep(1.5);pyautogui.click(x=501,y=243);time.sleep(1.5);print('프린트 스크립트 0');
      pyautogui.press('enter');time.sleep(1);print('모달에서 인쇄버튼')
      utils.Report_save('1',directory)  
      #2.pdf
      driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/Report_SalesInfo.asp?seq_no='+result[2]);time.sleep(2)
      pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=501,y=223);time.sleep(1.5);print('프린트 스크립트 0');
      pyautogui.press('enter');time.sleep(1);print('모달에서 인쇄버튼')
      utils.Report_save('2',directory)     
      #3.pdf
      driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/Report_CostRevenue.asp?seq_no='+result[2]);time.sleep(3)
      pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=501,y=223);time.sleep(1.5);print('프린트 스크립트 0');
      pyautogui.press('enter');time.sleep(1);print('모달에서 인쇄버튼')
      utils.Report_save('3',directory)      
      #4.pdf 재무능력평가
      driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/Report_AbleIssue.asp?seq_no='+result[2]);time.sleep(2)
      pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=501,y=223);time.sleep(1.5);print('프린트 스크립트 0');
      pyautogui.press('enter');time.sleep(1);print('모달에서 인쇄버튼')
      utils.Report_save('4',directory)  
      #5.pdf 법인이면 재무비율
      if result[4]<4:
        driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/Report_FinancialRatio.asp?seq_no='+result[2]);time.sleep(2)
        pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=501,y=223);time.sleep(1.5);print('프린트 스크립트 0');
        pyautogui.press('enter');time.sleep(1);print('모달에서 인쇄버튼')
        utils.Report_save('5',directory)    
      #6-1.pdf 재무상태표
      driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/Report_Statement_BS.asp?seq_no='+result[2]);time.sleep(2)
      pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=501,y=223);time.sleep(1.5);print('프린트 스크립트 0');
      pyautogui.press('enter');time.sleep(1);print('모달에서 인쇄버튼')
      utils.Report_save('6-1',directory)  
      #6-2.pdf 손익계산서
      driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/Report_Statement_IS.asp?seq_no='+result[2]);time.sleep(2)
      pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=501,y=223);time.sleep(1.5);print('프린트 스크립트 0');
      pyautogui.press('enter');time.sleep(1);print('모달에서 인쇄버튼')
      utils.Report_save('6-2',directory)                              
      #원가가 있으면
      if result[6]>0 :
        driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/Report_Statement_MC.asp?seq_no='+result[2]);time.sleep(2)
        pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=501,y=223);time.sleep(1.5);print('프린트 스크립트 0')
        pyautogui.press('enter');time.sleep(1);print('모달에서 인쇄버튼')
        utils.Report_save('6-3',directory)   
      # 법인인 경우
      #6-4.pdf 부속명세서
      if result[4]<4:
        driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/Report_Statement_ETC.asp?seq_no='+result[2]);time.sleep(2)
        pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=501,y=223);time.sleep(1.5);print('프린트 스크립트 0');
        pyautogui.press('enter');time.sleep(1);print('모달에서 인쇄버튼')
        utils.Report_save('6-4',directory)          
        #7.pdf  주식가치평가/기업진단

        driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/Report_EquityValuation.asp?seq_no='+result[2]);time.sleep(2)
        pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=501,y=223);time.sleep(1.5);print('프린트 스크립트 0');
        pyautogui.press('enter');time.sleep(1);print('모달에서 인쇄버튼')
        utils.Report_save('7',directory)   
        #8.pdf 기업진단
        if result[7] !='':
          print('기업진단:'+str(result[7]))
          driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/Report_Diagnosis.asp?seq_no='+result[2]);time.sleep(2)
          pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=501,y=223);time.sleep(1.5);print('프린트 스크립트 0');
          pyautogui.press('enter');time.sleep(1);print('모달에서 인쇄버튼')
          utils.Report_save('8',directory)   
      pyautogui.hotkey('alt','f4')
      utils.PDF_Merge('300',directory)                     
  #법인세무조정표지
  elif flag=="2":
    strsql = " select  biz_no,biz_name,year(reg_date),a.seq_no,fiscalMM,biz_type from mem_user a,mem_deal b where a.seq_no=b.seq_no  and a.biz_type<4 and keeping_yn='Y'"
    # strsql += "  and a.seq_no/1>=241 "
    strsql += "  order by a.seq_no/1"
    cursor = connection.cursor()
    cursor.execute(strsql)
    result = cursor.fetchall()
    connection.commit()
    connection.close()
    if result:
      driver = utils.ChromeDriver(False) ;time.sleep(1)
      for member in result:
        directory = "D:/"+member[1]+"/"+str(workyear)+"/세무조정계산서";utils.createDirectory(directory) 
        file_0 = directory + "/0.pdf"
        file_100 = directory + "/100.pdf"
        print(os.path.isfile(file_0.replace('/',"\\")))
        if not os.path.isfile(file_0.replace('/',"\\")):   
          driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/cover_0.asp?seq_no='+member[3]);time.sleep(2)
          pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=560,y=300);time.sleep(0.5);print('프린트 스크립트 0')
          # pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=560,y=300);time.sleep(0.5);print('프린트 스크립트 1')
          pyautogui.press('enter');time.sleep(1);print('모달에서 인쇄버튼')
          utils.Report_save('0',directory)      
        if not os.path.isfile(file_100.replace('/',"\\")):   
          driver.get('http://www.simplebook.co.kr/ibotax_admin/mn3sch/cover_100.asp?seq_no='+member[3]);time.sleep(2)
          pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=560,y=300);time.sleep(0.5);print('프린트 스크립트 100')
          # pyautogui.moveTo(0,0);time.sleep(0.5);pyautogui.click(x=560,y=300);time.sleep(0.5);print('프린트 스크립트 101')
          pyautogui.press('enter');time.sleep(1);print('인쇄버튼')
          utils.Report_save('100',directory)        
  print(f'정상종료 : {flagFirst}:{flagManagerName} - {flag} - {flagSeqNo}' )
  return JsonResponse({'data':'성공 : 법인세신고서작성'},safe=False) 

def File_Manage(request):
  workyear = datetime.now().year-1
  flag = request.GET.get('flag',False)
  flagGroup = flagSeqNo=flagManagerName =flagFirst = ''
  if flag=='2':
    workyear = pyautogui.confirm('세무법인대승/연도/세무조정계산서/99.pdf를 복사합니다. 원본파일 유무를 확인하세요. 작업대상 연도를 선택하세요',buttons=[workyear,(workyear+1)])
    print(workyear)
    if   workyear != "": 
      originFile = "D:/세무법인대승/"+str(workyear)+"/세무조정계산서/99.pdf"
      if os.path.isfile(originFile):
        #법인/개인 기장 사업자 모두 
        strsql = " select biz_no,biz_name,year(a.reg_date),a.seq_no,fiscalMM,biz_type from mem_user a,mem_deal b,mem_admin c where a.seq_no=b.seq_no and b.biz_manager=c.admin_id "
        # strsql += "  and a.seq_no/1>=241 "
        strsql += " and c.manage_YN='Y' and  keeping_yn='Y' and kijang_YN='Y' "
        strsql += "  order by a.seq_no/1"
        cursor = connection.cursor()
        cursor.execute(strsql)
        result = cursor.fetchall()
        connection.commit()
        connection.close()
        for member in result:
          directory = "D:/"+member[1]+"/"+str(workyear)+"/세무조정계산서";utils.createDirectory(directory) 
          file_99 = directory + "/99.pdf"  
          if not os.path.isfile(file_99):shutil.copyfile(originFile,file_99);print(file_99+" 성공")
          else:print(file_99+" 이미 존재")
      else: print(f'{originFile} : 조정반 파일(99.pdf)가 없습니다.')
  print(f'정상종료 : {flagFirst}:{flagManagerName} - {flag} - {flagSeqNo}' )
  return JsonResponse({'data':'성공 : 조정반 배포'},safe=False) 

# 급여대장 작성
def SS_MakePayDaejang(request):
  
  # 인사 작업 연도 결정을 위한 계산
  current_date = datetime.now().date()
  start_date = datetime(current_date.year, 1, 10).date() #1월10일
  end_date = datetime(current_date.year, 2, 15).date()  ##2월15일

  flag = request.GET.get('flag',False) # 1: 급여대장 2: 임금명세서
  flagSeqNo="";flagManagerName=""
  flagFirst = pyautogui.confirm(' 급여대장 작업할 형태를 선택하세요',buttons=['관리자별','단일사업자'])
  if flagFirst=="관리자별":  
    sqladmin = "select admin_ID from mem_admin where manage_YN='Y' order by admin_id"
    cursor = connection.cursor()
    cursor.execute(sqladmin)
    admins = cursor.fetchall()
    connection.commit()
    connection.close()  
    arrAdmin = []
    for admin in admins:
      arrAdmin.append(admin[0])
    flagManagerName = pyautogui.confirm('그룹작업 진행할 관리자를 선택하세요. [확정완료] 체크된 업체들만 작업됩니다.',buttons=arrAdmin) 
  elif flagFirst=="단일사업자":
    flagSeqNo = pyautogui.prompt('인트라넷 SEQNO를 입력하세요. [확정완료] 체크된 업체만 작업됩니다.')   

  workyear = datetime.now().year 
  if datetime.now().month==1 and datetime.now().day<12:workyear = workyear -1#12월분 정기신고 익년 11일까지는 전년도로 변경
  haveToIssue_MM = datetime.now().month
  if datetime.now().month==2:
    if datetime.now().day<14: haveToIssue_MM =haveToIssue_MM-1 #2월 설날인 경우 정기신고월 결정 매월 14일까지는 전월로...
  else:
    if datetime.now().day<12: haveToIssue_MM =haveToIssue_MM-1 # 정기신고월 결정 매월 14일까지는 전월로...
  flagPeriod2 = pyautogui.confirm('전자신고할 귀속월을 선택하세요',buttons=['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'])
  work_mm = flagPeriod2.replace("월",'')
  text_mm = work_mm
  if int(work_mm)<10 : text_mm = "0" + work_mm    

  isJungKi = 1;#정기신고
  flagYear = str(workyear)
  if int(work_mm)!=haveToIssue_MM:
    isJungKi=3#기한후신고
    arrYear = [workyear,workyear-1,workyear-2]
    flagYear = pyautogui.confirm('작업할 연도를 선택하세요. ',buttons=arrYear)

  strsql = " select replace(biz_no,'-',''),year(getdate())-year(reg_date),a.seq_no,fiscalMM,biz_name,trim(ediid),edipw,Duzon_ID " # 0부터 7까지
  strsql += " ,isnull(trim(yn_11),'') as ord_8,isnull(trim(yn_12),'') as ord_9,isnull(trim(yn_13),'') as ord_10,isnull(trim(yn_14),'') as ord_11,goyoung_banki  as ord_12 "
  strsql += " from mem_user a,mem_deal b left outer join tbl_mng_jaroe c on  b.seq_no=c.seq_no and work_yy="+str(workyear)+" and work_mm="+work_mm
  strsql += " where a.seq_no=b.seq_no  and keeping_yn='Y' "#and kijang_yn='Y'   "
  if flagFirst=="관리자별":
    flagfix = pyautogui.confirm('급여 지급 형태를 선택하세요',buttons=['고정','변동'])
    if   flagfix=="고정":  strsql += " and b.edipw like '%고정%' and goyoung_jungkyu='Y' "   
    elif flagfix=="변동":  strsql += " and b.edipw not like '%고정%' and YN_14='1' " 
    strsql += " and biz_manager ='" + flagManagerName + "' " 
  elif flagFirst=="단일사업자":  strsql += " and a.seq_no ='"+flagSeqNo+"'   "    
  strsql += "  order by a.seq_no/1 " 
  cursor = connection.cursor();print(strsql)
  cursor.execute(strsql)
  members = cursor.fetchall()
  connection.commit()
  connection.close()    
  for member in members:
    print('업체명 : ' + member[4])
    fileName_Month = str(work_mm) + "월 "
    fileName_Daejang = fileName_Month + "급여대장"

    fileName_Wonchun = str(work_mm) + "월 원천징수이행상황신고서"
    directory = "D:/"+member[4]+"/"+str(workyear)+"/인건비";utils.createDirectory(directory)
    if isJungKi==3: directory = "D:/"+member[4]+"/"+flagYear+"/인건비";utils.createDirectory(directory)
    print('Duzon_ID : '+member[7])
    print('급여대장 : '+fileName_Daejang)
    print(os.path.isfile(directory+"/"+fileName_Daejang+".pdf"))

    if flag=='1':#급여대장 선택 =>
      #급여변경란에 체크 안됐고 급여대장/명세서 파일이 없는 경우
      if member[8]!="1" and not ( os.path.isfile(directory+"/"+fileName_Month+"급여대장.pdf") and  os.path.isfile(directory+"/"+fileName_Month+"임금명세서.pdf") ): 
        dig = utils.semusarang_LaunchProgram_App(semusarangID)
        if member[7]=='1':  utils.semusarang_ChangeCompany(member[0])
        else:               utils.semusarang_ChangeCompany_ID_App(dig,member[7])  

        if work_mm=='1':#1월이면 전년도에서 마감이월 받아온다
          utils.semusarang_ChangeFiscalYear('insa',str(workyear-1))
          elecResult_Save.semusarang_MagamEwall('insa')
        utils.semusarang_ChangeFiscalYearAll(str(member[1]+1),str(workyear)) 
        elecResult_Save.PAY.semusarang_MakePayDaejang(text_mm,member[5],fileName_Month,directory)
        elecResult_Save.ACCOUNT.semusarang_MakeAccount_ThisMonth()   #급여 회계처리
        print("확정완료:"+member[11])#원천징수이행상황신고서 작성
        # 확정완료체크 됐거나 단일사업자인 경우, 반기아닌 경우, 반기인 경우 6월과 12월에 
        if (member[11]=="1" or flagFirst=="단일사업자") and ( (member[12]=="Y" and flagPeriod2 in ('6월','12월')) or member[12]!="Y"):
          isJungKi=1     
          elecResult_Save.PAY.semusarang_Issue_Wonchun(text_mm,fileName_Wonchun,directory,isJungKi,flagYear)           
          elecResult_Save.ElecIssue.SS_MakeElecFile('wonchun',member[2],member[0],str(workyear),text_mm,'kwasekikan','kwaseyuhyung',1,htxLoginID) 
          if not os.path.isfile(directory+"/"+fileName_Month+" 납부서(지방소득세).pdf"): 
            elecResult_Save.PAY.semusarang_Issue_Wonchun_wetax(text_mm)
            elecResult_Save.ElecIssue.SS_MakeElecFile('wonchun-wetax',member[2],member[0],str(workyear),text_mm,'kwasekikan','kwaseyuhyung',1,htxLoginID)  
          else : print('지방소득세 납부서가 존재합니다')
        if work_mm=='12':#12월이면 마감한다
          elecResult_Save.semusarang_MagamEwall('insa') 
      else:
        if member[8]=="1":
          print("급여변동이 있거나 입퇴사자가 존재합니다. 급여대장을 작성하고 체크해제해 주세요")        
          return 1
        if  os.path.isfile(directory+"/"+fileName_Daejang+".pdf"):
          confirmProcess = pyautogui.confirm("이미 급여대장이 존재합니다. 원천세 신고를 진행합니까?") 
          if confirmProcess=='OK':
            dig = utils.semusarang_LaunchProgram_App(semusarangID)
            if member[7]=='1':  utils.semusarang_ChangeCompany(member[0])
            else:               utils.semusarang_ChangeCompany_ID_App(dig,member[7])         
            isJungKi=1     
            elecResult_Save.PAY.semusarang_Issue_Wonchun(text_mm,fileName_Wonchun,directory,isJungKi,flagYear) 
            elecResult_Save.ElecIssue.SS_MakeElecFile('wonchun',member[2],member[0],str(workyear),text_mm,'kwasekikan','kwaseyuhyung',1,htxLoginID)  
            if not os.path.isfile(directory+"/"+fileName_Month+" 납부서(지방소득세).pdf"): 
              elecResult_Save.PAY.semusarang_Issue_Wonchun_wetax(text_mm)  
              elecResult_Save.ElecIssue.SS_MakeElecFile('wonchun-wetax',member[2],member[0],str(workyear),text_mm,'kwasekikan','kwaseyuhyung',1,htxLoginID)  
            else : print('지방소득세 납부서가 존재합니다')              
    # elif flag=='2':#원천징수이행상황신고서 -> 급여작업 사업 등 개별진행시
      #확정완료체크 됐거나 단일사업자인 경우, 반기아닌 경우, 반기인 경우 6월과 12월에  지우고 다시 제작
      # if (member[11]=="1" or flagFirst=="단일사업자") and ( (member[12]=="Y" and flagPeriod2 in ('6월','12월')) or member[12]!="Y"):
      #   dig = utils.semusarang_LaunchProgram_App(semusarangID)
      #   if member[7]=='1':  utils.semusarang_ChangeCompany(member[0])
      #   else:               utils.semusarang_ChangeCompany_ID_App(dig,member[7])          
      #   elecResult_Save.PAY.semusarang_Issue_Wonchun(text_mm,fileName_Wonchun,directory)  
      #   if not os.path.isfile(directory+"/"+fileName_Month+" 납부서(지방소득세).pdf"):   elecResult_Save.PAY.semusarang_Issue_Wonchun_wetax(text_mm)  
      #   else : print('지방소득세 납부서가 존재합니다')
    elif flag=='3':#원천세 전자신고    (확정완료 또는 단일사업자) and {(반기이고 신고월이 6,12월) 이거나 반기가 아닌 경우}
      if (member[11]=="1" or flagFirst=="단일사업자") and ( (member[12]=="Y" and flagPeriod2 in ('6월','12월')) or member[12]!="Y"):
        dig = utils.semusarang_LaunchProgram_App(semusarangID)
        if member[7]=='1':  utils.semusarang_ChangeCompany(member[0])
        else:               utils.semusarang_ChangeCompany_ID_App(dig,member[7])   
        if flagPeriod2 in ('1월','2월') and (start_date <= current_date <= end_date):  #1월10일보다 크고 2월 15일보다 작으면 기수변경시킨다
          utils.semusarang_ChangeFiscalYear('insa',str(workyear))

        elecResult_Save.PAY.semusarang_Issue_Wonchun(text_mm,fileName_Wonchun,directory,isJungKi,flagYear)    
        elecResult_Save.ElecIssue.SS_MakeElecFile('wonchun',member[2],member[0],str(workyear),text_mm,flagYear,'KwaseyuHyung',isJungKi,htxLoginID)  
        if not os.path.isfile(directory+"/"+fileName_Month+" 납부서(지방소득세).pdf"): 
          elecResult_Save.PAY.semusarang_Issue_Wonchun_wetax(text_mm)  
          elecResult_Save.ElecIssue.SS_MakeElecFile('wonchun-wetax',member[2],member[0],str(workyear),text_mm,flagYear,'KwaseyuHyung',isJungKi,htxLoginID)  
        else : print('지방소득세 납부서가 존재합니다') 
    elif flag=='4':#지방세 전자신고 
      if (member[11]=="1" or flagFirst=="단일사업자") and ( (member[12]=="Y" and flagPeriod2 in ('6월','12월')) or member[12]!="Y") and  not os.path.isfile(directory+"/"+fileName_Month+" 납부서(지방소득세).pdf") :#확정완료 체크
        dig = utils.semusarang_LaunchProgram_App(semusarangID)
        if member[7]=='1':  utils.semusarang_ChangeCompany(member[0])
        else:               utils.semusarang_ChangeCompany_ID_App(dig,member[7])          
        time.sleep(1)
        elecResult_Save.PAY.semusarang_Issue_Wonchun_wetax(text_mm)  
        elecResult_Save.ElecIssue.SS_MakeElecFile('wonchun-wetax',member[2],member[0],str(workyear),text_mm,'kwasekikan','kwaseyuhyung',1,htxLoginID)                  
      else:
        if  member[11]!="1" : print('확정완료에 체크되지 않았습니다')
        if  os.path.isfile(directory+"/"+fileName_Month+" 납부서(지방소득세).pdf") : print('지방소득세 납부서가 존재합니다')
  print(f'정상종료 : {flagFirst}:{flagManagerName} - {flag} - {flagSeqNo}' )
  return JsonResponse({'data':'급여대장 : '+str(work_mm)+'월 '+str(workyear)},safe=False) 

#세무사랑 - 마감이월
def SS_MagamEwall(request):
  flag = request.GET.get('flag',False)
  if flag=='1': flag='vat'
  else        :flag='insa'
  work_yy = pyautogui.confirm('회계마감할 연도를 선택하세요. ',buttons=[(datetime.now().year -1),datetime.now().year ])
  flagGroup = "";flagSeqNo='';flagManagerName =''
  flagFirst = pyautogui.confirm('간이지급명세서 제작 대상을 선택하세요',buttons=['관리자별','단일사업자','전체'])
  if flagFirst=="관리자별":
    sqladmin = "select admin_Name,Htx_ID,admin_id,manage_YN from mem_admin  order by admin_id"
    cursor = connection.cursor()
    cursor.execute(sqladmin)
    admins = cursor.fetchall()
    connection.commit()
    arrAdmin = []
    for admin in admins:
      if admin[3]=="Y" :
        arrAdmin.append(admin[2])
    flagManagerName = pyautogui.confirm('그룹작업 진행할 관리자를 선택하세요. ',buttons=arrAdmin)
    for admin in admins:
      if admin[0]==flagManagerName:  
        flagGroup += "'"+admin[2]+"'," 
  elif flagFirst=="단일사업자":
    flagSeqNo = pyautogui.prompt('SEQNO를 입력하세요.');    
  if flagFirst!="단일사업자" : flagTxt = pyautogui.confirm('마감할 대상을 선택하세요. ',buttons=['법인','개인','전체'])
  strsql = " select replace(biz_no,'-',''),"+work_yy+"-year(reg_date)+1,a.seq_no,fiscalMM,biz_type,duzon_id from mem_user a,mem_deal b "
  strsql += "where a.seq_no=b.seq_no  and keeping_yn='Y' and kijang_yn='Y'   "
  if flagFirst=="관리자별":     
    strsql += " and biz_manager ='" + flagManagerName + "' "
    strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
  elif flagFirst=="단일사업자": 
    strsql += " and a.seq_no ='"+flagSeqNo+"'   "           
  elif  flagFirst=="전체":
    strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') "   
  if flagTxt=="법인": strsql += " and a.biz_type<4 "  #법인
  elif flagTxt=="개인": strsql += " and a.biz_type>=4 "  #개인
  strsql += " order by a.seq_no/1"
  cursor = connection.cursor()
  cursor.execute(strsql)
  members = cursor.fetchall()
  connection.commit()
  connection.close()    
  dig = utils.semusarang_LaunchProgram_App(semusarangID)
  for member in members:
    print('seqNo : ' + member[2])
    finalKi = member[1]
    if int(finalKi)>1: 
      if member[5]=='1':  utils.semusarang_ChangeCompany(member[0])
      else:               utils.semusarang_ChangeCompany_ID_App(dig,member[5])            
      utils.semusarang_ChangeFiscalYear_App(flag,str(finalKi)) 
      elecResult_Save.semusarang_MagamEwall(flag)
      utils.semusarang_ChangeFiscalYear_App(flag,str(int(finalKi)+1)) 
  print(f'정상종료 : {flagFirst}:{flagManagerName} - {flag}' )
  return JsonResponse({'data':'마감이월 : 성공'},safe=False) 

# 회계처리 - 인사
def SS_MakeAccount_MonthlyPay(request):
  flagSeqNo="";flagManagerName="";flag=""
  flagFirst = pyautogui.confirm(' 급여대장 작업할 형태를 선택하세요',buttons=['관리자별','단일사업자'])
  if flagFirst=="관리자별":  
    sqladmin = "select admin_ID from mem_admin where manage_YN='Y' order by admin_id"
    cursor = connection.cursor()
    cursor.execute(sqladmin)
    admins = cursor.fetchall()
    connection.commit()
    connection.close()  
    arrAdmin = []
    for admin in admins:
      arrAdmin.append(admin[0])
    flagManagerName = pyautogui.confirm('그룹작업 진행할 관리자를 선택하세요. ',buttons=arrAdmin) 
  elif flagFirst=="단일사업자":
    flagSeqNo = pyautogui.prompt('인트라넷 SEQNO를 입력하세요. ')   

  work_dd = datetime.now().day
  work_mm = datetime.now().month
  work_yy = datetime.now().year
  if work_mm<=2:work_yy = datetime.now().year - 1
  if work_dd<=10:work_mm = datetime.now().month - 1
  text_mm = str(work_mm)
  if datetime.now().month<10 : text_mm = "0" + str(datetime.now().month)

  strsql = " select  max(biz_no),max(year(getdate())-year(reg_date)),a.seq_no,max(fiscalMM),max(biz_type),max(duzon_id),max(과세연월),"
  strsql +=" sum(a01) a01,sum(a03) a03,sum(a20) a20,sum(a30) a30,sum(a40) a40,sum(a50) a50,sum(a60) a60 "
  # strsql = "select biz_no,year(getdate())-year(reg_date),a.seq_no,fiscalMM,biz_type,duzon_id,과세연월, a01,a03, a20, a30, a40, a50, a60 "
  strsql += " from mem_user a,mem_deal b,원천세전자신고 c where a.seq_no=b.seq_no and a.biz_no=c.사업자등록번호 and keeping_yn='Y' and kijang_yn='Y'   "
  strsql += " and  left(과세연월,4)='"+str(work_yy)+"'  "
  if flagFirst=="관리자별":
    strsql += " and biz_manager ='" + flagManagerName + "' and a.seq_no not in ('2209')  " 
  elif flagFirst=="단일사업자":  strsql += " and a.seq_no ='"+flagSeqNo+"'   "    
  strsql += " group by a.seq_no,biz_name order by a.seq_no/1,biz_name " 
  cursor = connection.cursor();print(strsql)
  cursor.execute(strsql)
  members = cursor.fetchall()
  connection.commit()
  connection.close()    
  preSeqNo='';nowSeqNo=''
  for member in members:
    print('seqNo : ' + member[2])
    print('과세연월 : ' + member[6])
    nowSeqNo = member[2]
    if preSeqNo!=nowSeqNo:
      dig = utils.semusarang_LaunchProgram_App(semusarangID)
      if member[5]=='1':  utils.semusarang_ChangeCompany(member[0].replace('-',''))
      else:               utils.semusarang_ChangeCompany_ID_App(dig,member[5])        
      if text_mm=='01':   
        utils.semusarang_ChangeFiscalYear('vat',str(member[1]))  
        elecResult_Save.semusarang_MagamEwall('vat')
        utils.semusarang_ChangeFiscalYearAll(str(member[1]),str(work_yy)) 
      else:                       
        utils.semusarang_ChangeFiscalYearAll(str(member[1]+1),str(work_yy)) 
      
    text_mm = member[6][4:6]
    if member[7]>0: elecResult_Save.semusarang_MakeAccount_Insa_AA('a01',member[0],member[4],str(work_yy),text_mm) #a01
    if member[8]>0: elecResult_Save.semusarang_MakeAccount_Insa_AA('a03',member[0],member[4],str(work_yy),text_mm) #a03
    if member[9]>0: elecResult_Save.semusarang_MakeAccount_Insa_AA('a20',member[0],member[4],str(work_yy),text_mm) #a20
    if member[10]>0: elecResult_Save.semusarang_MakeAccount_Insa_AA('a30',member[0],member[4],str(work_yy),text_mm)  #a30
    if member[11]>0: elecResult_Save.semusarang_MakeAccount_Insa_AA('a40',member[0],member[4],str(work_yy),text_mm)  #a40 - 기타
    if member[12]>0: elecResult_Save.semusarang_MakeAccount_Insa_AA('a50',member[0],member[4],str(work_yy),text_mm)
    if member[13]>0: elecResult_Save.semusarang_MakeAccount_Insa_AA('a60',member[0],member[4],str(work_yy),text_mm)
    preSeqNo=nowSeqNo
  print(f'정상종료 : {flagFirst}:{flagManagerName} - {flag} - {flagSeqNo}' )
  return JsonResponse({'data':'마감이월 : '+str(work_mm)+'월 '+str(work_yy)},safe=False) 


#세무사랑 - 급여지급현황 - 수당공제합계 - 엑셀업로드
def SS_Payroll(request):
  flag=request.GET.get('flag',False)
  workyear = datetime.now().year 
  today = str(datetime.now())[:10].replace("-","")
  if datetime.now().month<2:    workyear = workyear - 1
  arrYear = [workyear,workyear-1,workyear-2]

  flagGroup = "";flagSeqNo='';flagYear='';flagManagerName ='';flagBiztype = ''
  flagFirst = pyautogui.confirm('급여지급현황 업로드 대상을 선택하세요',buttons=['관리자별','단일사업자','전체'])
  if flagFirst=="관리자별":
    sqladmin = "select admin_Name,Htx_ID,admin_id,manage_YN from mem_admin  order by admin_id"
    cursor = connection.cursor()
    cursor.execute(sqladmin)
    admins = cursor.fetchall()
    connection.commit()

    arrAdmin = []
    for admin in admins:
      if admin[3]=="Y" :
        arrAdmin.append(admin[2])
    flagManagerName = pyautogui.confirm('그룹작업 진행할 관리자를 선택하세요. ',buttons=arrAdmin)
    for admin in admins:
      if admin[0]==flagManagerName:  
        flagGroup += "'"+admin[2]+"'," 
    flagBiztype = pyautogui.confirm('그룹작업 진행할 형태를 선택하세요. ',buttons=['법인','개인','전체'])
  elif flagFirst=="단일사업자":
    flagSeqNo = pyautogui.prompt('SEQNO를 입력하세요.');  
  # flagPeriod = pyautogui.confirm('작업 기간을 선택하세요',buttons=['연도별','월별'])  #,'반기별'조회기간은 최대 3개월 까지 가능합니다.조회기간을 확인하시기 바랍니다
  flagYear = pyautogui.confirm('작업할 연도를 선택하세요. ',buttons=arrYear)

  strsql = " select biz_no,year(reg_date),a.seq_no,fiscalMM,biz_type,duzon_id,biz_name,goyoung_banki,trim(HometaxID),trim(HometaxPW) "
  strsql += "from mem_user a,mem_deal b where a.seq_no=b.seq_no  "
  if flagFirst=="관리자별":     
    strsql += " and biz_manager ='" + flagManagerName + "' "
    strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
  elif flagFirst=="단일사업자": 
    strsql += " and a.seq_no ='"+flagSeqNo+"'   "           
  elif  flagFirst=="전체":
    strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') "   
  if flagBiztype == '법인':    strsql += " and biz_type<4 "
  elif flagBiztype == '개인':  strsql += " and biz_type>=4 "
  strsql += " and (select sum(a01) from 원천세전자신고 c where c.사업자등록번호=a.biz_no and  left(지급연월,4)='"+flagYear+"' )>0 "
  strsql += " and HometaxID<>'' and HometaxPW<>'' and duzon_ID <> '' and b.keeping_YN = 'Y' and a.Del_YN <> 'Y'   order by a.seq_no/1 "
  print(strsql)    
  cursor = connection.cursor()
  cursor.execute(strsql)
  results = cursor.fetchall()
  connection.commit()   
  if flag=='1': #급여지급현황 업로드
    for result in results:   
      print(result[6])   
      strsql = "select (select right(max(과세연월),2) from 원천세전자신고 where 사업자등록번호='"+result[0]+"' and a01>0),* from 스크래핑관리 where crt_dt='" + today + "' and seqno_lastTry='"+result[2]+"' and bigo='급여지급현황'"
      cursor.execute(strsql)   
      scraps = cursor.fetchall()
      connection.commit() 
      print(strsql)
      if not scraps or flagFirst=="단일사업자":   
        strsql2 = "select right(min(지급연월),2),right(max(지급연월),2) from 원천세전자신고 where 사업자등록번호='"+result[0]+"' and left(과세연월,4)='"+flagYear+"' and a01>0"
        cursor.execute(strsql2)   
        scraps2 = cursor.fetchall()
        connection.commit() 
        print(scraps2[0][0])  
        dig = utils.semusarang_LaunchProgram_App(semusarangID)
        if result[5]=='1':  utils.semusarang_ChangeCompany(result[0].replace('-',''))
        else:               utils.semusarang_ChangeCompany_ID_App(dig,result[5])      
        utils.semusarang_ChangeFiscalYear('insa',flagYear)
        startMM = scraps2[0][0];print(startMM)
        endMM   = scraps2[0][1];print(endMM)
        if result[7]=="Y":#반기
          if int(startMM)<=6:startMM = '1' 
          endMM = '12' #연말정산 및 보수총액 신고시는 무조건 12월까지 조회해본다
          # if int(endMM)>6:endMM = '12'
        elecResult_Save.PAY.semusarang_Pay_Monthly_Excel(result[2],flagYear,startMM,endMM,'급여지급현황',cursor)
        # strsql = "insert into 스크래핑관리 values('"+today+"','"+result[2]+"','"+result[6]+"','급여지급현황')"  
        strsql = "Merge 스크래핑관리 as A Using (select '"+today+"' as crt_dt,'"+result[2]+"' as seq_no,'급여지급현황' as bigo) as B "
        strsql += "On A.crt_dt=B.crt_dt  and A.seqno_lastTry=B.seq_no and A.bigo=B.bigo "
        strsql += "when matched then update set seqno_lastTry='"+result[2]+"' "
        strsql += "When Not Matched Then insert values('"+today+"','"+result[2]+"','"+result[6]+"','급여지급현황');"
        print(strsql)  
        cursor.execute(strsql)
  elif flag=='2':#연말정산명단등록
    for result in results:  
      strsql = "select (select right(max(과세연월),2) from 원천세전자신고 where 사업자등록번호='"+result[0]+"' and a01>0),* from 스크래핑관리  "
      strsql += "where crt_dt='" + flagYear + "' and seqno_lastTry='"+result[2]+"' and bigo='연말정산대상 근로자 명단등록'"
      cursor.execute(strsql)   
      scraps = cursor.fetchall()
      connection.commit() 
      print(strsql)
      if (not scraps and result[9]!="" and result[9][:4]!="비번틀림")  or flagFirst=="단일사업자":          
        destinationDirectory = "D:\\"+result[6]+"\\"+flagYear+"\\연말정산\\"
        downDirectory = "C:\\Users\\Administrator\\Downloads\\"
        utils.createDirectory(destinationDirectory)      
        excelFile = "일괄제공 동의 신청 근로자 명세_"+flagYear+".xlsx"
        if os.path.isfile(downDirectory+excelFile): os.remove(downDirectory+excelFile)
        strsql2 = "select '','' union select trim(empJumin),trim(empName) from 급여지급현황 where seq_no='"+result[2]+"' and work_yy='"+flagYear+"' and work_mm in(11,12) ";print(strsql2)
        cursor.execute(strsql2)
        result2 = cursor.fetchall()
        connection.commit()
        df = DataFrame(result2)
        df.columns = ["주민등록번호","성명"]
        excel_writer = pd.ExcelWriter(downDirectory+excelFile)
        df.to_excel(excel_writer,index=False)
        excel_writer.close()
        print(downDirectory+excelFile+" 파일저장성공")   
        if os.path.isfile(downDirectory+excelFile): 
          shutil.copyfile(downDirectory+excelFile, destinationDirectory+excelFile)  
          driver = utils.conHometaxLogin_Personal(result,False);time.sleep(1)
          try:
            driver.find_element(By.ID,'hdGroup920').click();    time.sleep(1); print('연말정산 ')
          except:
            strUdt = "update mem_deal set HometaxPw = '비번틀림"+result[9]+"' where seq_no='"+result[2]+"'";print(strUdt)
            cursor = connection.cursor()
            cursor.execute(strUdt)            
            continue
          iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe');    driver.switch_to.frame(iframe);    time.sleep(1);print('iframe txppIframe')
          driver.find_element(By.ID,'sub_a_4504070200').click() ;driver.implicitly_wait(30);   time.sleep(1);print('연말정산대상 근로자 명단등록')            
          try:
            WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(0.5);print(result[6]+' : 연말정산 부서사용자 있음')
            driver.find_element(By.ID,'textbox915').click();time.sleep(1);print('로그아웃')
            WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1);print('로그아웃 합니다 알람')
            continue
          except:
            print('연말정산 부서사용자 없음')
          iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe');    driver.switch_to.frame(iframe);    time.sleep(1);print('iframe txppIframe')         
          # driver.find_element(By.ID,'edtYsboChrgNm').clear();
          driver.find_element(By.ID,'edtYsboChrgNm').send_keys('김기현');time.sleep(0.25);     
          # driver.find_element(By.ID,'edtTelno1').clear();
          driver.find_element(By.ID,'edtTelno1').send_keys('02');time.sleep(0.25)
          # driver.find_element(By.ID,'edtTelno2').clear();
          driver.find_element(By.ID,'edtTelno2').send_keys('501');time.sleep(0.25)
          # driver.find_element(By.ID,'edtTelno3').clear();
          driver.find_element(By.ID,'edtTelno3').send_keys('1732');time.sleep(0.25)
          # driver.find_element(By.ID,'edtYsboFlePswd').clear();
          driver.find_element(By.ID,'edtYsboFlePswd').send_keys('1111');time.sleep(0.25)
          driver.find_element(By.ID,'btnInsertChrgInfr').click();time.sleep(0.5);print('신청대상자 저장하기')
          WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(0.5);print('저장하시겠습니까')
          WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1.5);print('간소화 일괄제공 업무 수행자 정보가 저장되었습니다')      
          agree = driver.find_element(By.CSS_SELECTOR,'#edtYsboFleTxaaUseYn_input_1');    agree.click(); print('일괄제공 : 동의') ;time.sleep(0.5) 
          try:
            select1 = Select(driver.find_element(By.ID,'bkpPrxTxaa'));        select1.select_by_visible_text('세무법인 대승');time.sleep(0.5)
          except:
            print(result[6]+' : 세무대리인이 없습니다')
            continue
          radio = driver.find_element(By.CSS_SELECTOR,'#cmbSbmsMthd_input_1');    radio.click(); print('제출방법 : 엑셀파일') ;time.sleep(0.5) 
          driver.find_element(By.ID,'trigger116').click();time.sleep(1)  ; print('찾아보기')
          menuName='열기'
          intlen = len(menuName)
          procs = pywinauto.findwindows.find_elements();handle=''
          for proc in procs: 
            if proc.name[:intlen]== menuName:handle = proc.handle;break
          app = Application().connect(handle=handle)
          w_open = app.window(handle=handle)
          w_open.set_focus()        
          pyautogui.hotkey('alt','d');time.sleep(0.5);pyperclip.copy(downDirectory); pyautogui.hotkey('ctrl', 'v'); 
          pyautogui.press('enter');time.sleep(0.5);pyautogui.press('tab',presses=3,interval=0.25)
          pyautogui.hotkey('alt','n');print('연말정산파일선택');time.sleep(0.5);pyperclip.copy(excelFile); pyautogui.hotkey('ctrl', 'v');time.sleep(0.5);pyautogui.hotkey('alt','o'); time.sleep(2)
          pyautogui.press('esc',presses=3,interval=0.25)#잘못된 팝업 털기 
          driver.find_element(By.ID,'btnExcelUpload').click();time.sleep(1)  ; print('엑셀명단 제출하기')       
          WebDriverWait(driver, 50).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1.5);print('간소화 일괄제공 업무 수행자 정보 등록완료')      
          strsql = "insert into 스크래핑관리 values('"+flagYear+"','"+result[2]+"','"+result[6]+"','연말정산대상 근로자 명단등록')"       
          cursor.execute(strsql)       
          os.remove(downDirectory+excelFile) 
  elif flag=='3':
    for result in results:  
      strsql = "select * from 스크래핑관리 where crt_dt='" + str(datetime.now())[:10].replace("-","") + "' and seqno_lastTry='"+result[2]+"' and bigo='연말정산대상 근로자 동의조회'"
      cursor.execute(strsql)   
      scraps = cursor.fetchall()
      connection.commit() 
      print(strsql)
      if (not scraps and result[9]!="" and result[9][:4]!="비번틀림")  or flagFirst=="단일사업자":    
        driver = utils.conHometaxLogin_Personal(result,False);time.sleep(1)
        try:
          driver.find_element(By.ID,'hdGroup920').click();    time.sleep(1); print('연말정산 ')
        except:
          strUdt = "update mem_deal set HometaxPw = '비번틀림"+result[9]+"' where seq_no='"+result[2]+"'";print(strUdt)
          cursor = connection.cursor()
          cursor.execute(strUdt)    
          driver.close()          
          continue
        iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe');    driver.switch_to.frame(iframe);    time.sleep(1);print('iframe txppIframe')
        driver.find_element(By.ID,'sub_a_4504070300').click() ;driver.implicitly_wait(30);   time.sleep(1);print('연말정산대상 근로자 관리화면')            
        try:
          WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(0.5);print(result[6]+' : 연말정산 부서사용자 있음')
          driver.find_element(By.ID,'textbox915').click();time.sleep(1);print('로그아웃')
          WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1);print('로그아웃 합니다 알람')
          continue
        except:
          print('연말정산 부서사용자 없음') 
        iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe');    driver.switch_to.frame(iframe);    time.sleep(1);print('iframe txppIframe')  
        driver.find_element(By.ID,'btnSearch').click();time.sleep(1);print('조회하기')
        driver.find_element(By.ID,'btnDownload').click();time.sleep(15);print('명단내려받기')
        try:
          WebDriverWait(driver, 2).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(0.5);print(result[6]+' : 조회내역이 없습니다')
          strsql = "merge 스크래핑관리  as A using(select '"+result[2]+"' as seqno_lastTry,'"+result[6]+"' as biz_name, '연말정산대상 근로자 동의조회' as bigo) as B  "   
          strsql += " on A.seqno_lastTry=B.seqno_lastTry  and A.biz_name=B.biz_name and A.bigo=B.bigo "
          strsql += " when matched then update set crt_dt='"+str(datetime.now())[:10].replace("-","")+"'"   
          strsql += " when not matched then insert values('"+str(datetime.now())[:10].replace("-","")+"','"+result[2]+"','"+result[6]+"','연말정산대상 근로자 동의조회');" 
          print(strsql)
          cursor.execute(strsql)   
          driver.close()    
        except:
          print('조회내역이 있음')         
        vList = list(result)
        vList.append(flagYear)
        result = tuple(vList)
        isOK = utils.DBSave_Downloaded_xlsx(driver,result,'연말정산대상근로자') 
        if isOK:
          strsql = "merge 스크래핑관리  as A using(select '"+result[2]+"' as seqno_lastTry,'"+result[6]+"' as biz_name, '연말정산대상 근로자 동의조회' as bigo) as B  "   
          strsql += " on A.seqno_lastTry=B.seqno_lastTry  and A.biz_name=B.biz_name and A.bigo=B.bigo "
          strsql += " when matched then update set crt_dt='"+str(datetime.now())[:10].replace("-","")+"'"   
          strsql += " when not matched then insert values('"+str(datetime.now())[:10].replace("-","")+"','"+result[2]+"','"+result[6]+"','연말정산대상 근로자 동의조회');" 
          print(strsql)
          cursor.execute(strsql)   
        driver.close()         
  elif flag=='4':#간소화 자료제공동의 근로자 파일 다운로드 압축해제  
    for result in results:             
      strsql = f"select seq_no,isnull((select seqno_lastTry from 스크래핑관리 where crt_dt='{flagYear}' and bigo='간소화 자료제공동의 파일저장' and seqno_lastTry=a.seq_no),'') from 연말정산일괄제공근로자 a where work_yy='{flagYear}' and 동의일자<>'' and seq_no='{result[2]}'"
      cursor.execute(strsql)   
      scraps = cursor.fetchone()
      connection.commit() 
      if (scraps and scraps[0].strip()!=scraps[1].strip() and result[9]!="" and result[9][:4]!="비번틀림")  or flagFirst=="단일사업자":    
        driver = utils.conHometaxLogin_Personal(result,False);time.sleep(1)
        try:
          driver.find_element(By.ID,'hdGroup920').click();    time.sleep(1); print('연말정산 ')
        except:
          strUdt = "update mem_deal set HometaxPw = '비번틀림"+result[9]+"' where seq_no='"+result[2]+"'";print(strUdt)
          cursor = connection.cursor()
          cursor.execute(strUdt)    
          driver.close()          
          continue
        iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe');    driver.switch_to.frame(iframe);    time.sleep(1);print('iframe txppIframe')
        driver.find_element(By.ID,'sub_a_4504070400').click() ;driver.implicitly_wait(30);   time.sleep(1);print('일괄제공파일 내려받기')            
        try:
          WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(0.5);print(result[6]+' : 연말정산 부서사용자 있음')
          driver.find_element(By.ID,'textbox915').click();time.sleep(1);print('로그아웃')
          WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1);print('로그아웃 합니다 알람')
          continue
        except:
          print('연말정산 부서사용자 없음') 
        iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe');    driver.switch_to.frame(iframe);    time.sleep(1);print('iframe txppIframe')  
        driver.find_element(By.ID,'btnSearch').click();time.sleep(3);print('조회하기')
        try:
          WebDriverWait(driver, 3).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1)
          print(f"{result[6]}:아직 파일이 생성되지 않았습니다.")
          driver.close()
          continue
        except:
          print(f"{result[6]}:파일 생성됨")
        driver.find_element(By.ID,'btnAllFleDwld').click();time.sleep(5);print('명단내려받기')
        directory = f"d:\\{result[6]}\\{workyear}\\연말정산"
        fileName = f"{result[0].replace('-','')}_{result[8]}_pdf.zip"
        utils.SET_FOCUS('다른 이름으로 저장')
        pyautogui.hotkey('alt','d');print(f'{directory} 선택');time.sleep(0.5);pyperclip.copy(directory); pyautogui.hotkey('ctrl', 'v');
        pyautogui.press('enter');time.sleep(1);pyautogui.hotkey('alt','s'); time.sleep(3)
        isOK = utils.Extract_Download_File(f"{directory}\\{fileName}",directory,'1111') 
        if isOK:
          strsql = "merge 스크래핑관리  as A using(select '"+result[2]+"' as seqno_lastTry,'"+result[6]+"' as biz_name, '간소화 자료제공동의 파일저장' as bigo) as B  "   
          strsql += " on A.seqno_lastTry=B.seqno_lastTry  and A.biz_name=B.biz_name and A.bigo=B.bigo "
          strsql += f" when matched then update set crt_dt='{flagYear}'"   
          strsql += f" when not matched then insert values('{flagYear}','{result[2]}','{result[6]}','간소화 자료제공동의 파일저장');" 
          print(strsql)
          cursor.execute(strsql)   
        driver.close()      
      else:
        print(f'{result[6]}:동의 근로자 없음')
  print(f'정상종료 : {flagFirst}:{flagManagerName} - {flag} - {flagSeqNo}' )
  return JsonResponse({'data':'성공 : 급여지급현황 - 수당공제합계업로드'},safe=False) 

#세무사랑-분개장  업로드
def SS_SlipLedgr(request):
  biz_no = request.GET.get('bizNo',False)
  flag = request.GET.get('flag',False)
  workyear = datetime.now().year 
  work_mm = datetime.now().month
  today = str(datetime.now())[:10].replace("-","")
  if work_mm<4:    workyear = workyear - 1

  work_qt=4;endMM = "12";startMM = "01";vatKikan = ''
  if work_mm<4 : work_qt = 4;endMM = "12";startMM = "10";vatKikan = '2기확정'
  elif work_mm<7 : work_qt = 1;endMM = "03";startMM = "01";vatKikan = '1기예정'
  elif work_mm<10 : work_qt = 2;endMM = "06";startMM = "04";vatKikan = '1기확정'
  elif work_mm<=12 : work_qt = 3;endMM = "09";startMM = "07";vatKikan = '2기예정'
  if flag=='2':#혜솔 계열사 분개장 업무
    memuser = MemUser.objects.get(biz_no=biz_no) 
    flagProc = pyautogui.confirm('분개장 작업을 시작할 단계를 지정하세요',buttons=['스크래핑처리부터','거래처코드등록부터','원시분개장업로드부터','업로드엑셀생성','세무사랑반영'])
    
    if flagProc=="스크래핑처리부터":    
      utils.semusarang_Login(semusarangID,memuser.duzon_id,memuser.biz_no,memuser.reg_date,workyear)
      elecResult_Save.ECOUNT.semusarang_Scrapping(str(workyear),startMM,endMM);time.sleep(15)#이미지 "실시간"이 뜨면 다음으로 넘어가도록
      elecResult_Save.ECOUNT.TraderUpload(memuser.seq_no)
      elecResult_Save.ECOUNT.TransactionExcelUpDnload(memuser.seq_no,memuser.biz_name,workyear,vatKikan,work_qt)
      elecResult_Save.ECOUNT.MakeResultExcel(memuser.seq_no,memuser.biz_no,memuser.biz_name,workyear,vatKikan,work_qt)
      elecResult_Save.ECOUNT.semusarang_TransactionExcelUpload()      
    elif flagProc=="거래처코드등록부터":
      utils.semusarang_Login(semusarangID,memuser.duzon_id,memuser.biz_no,memuser.reg_date,workyear)
      elecResult_Save.ECOUNT.TraderUpload(memuser.seq_no)
      elecResult_Save.ECOUNT.TransactionExcelUpDnload(memuser.seq_no,memuser.biz_name,workyear,vatKikan,work_qt)
      elecResult_Save.ECOUNT.MakeResultExcel(memuser.seq_no,memuser.biz_no,memuser.biz_name,workyear,vatKikan,work_qt)
      elecResult_Save.ECOUNT.Semusarang_TransactionExcelUpload()
    elif flagProc=="원시분개장업로드부터":
      elecResult_Save.ECOUNT.TransactionExcelUpDnload(memuser.seq_no,memuser.biz_name,workyear,vatKikan,work_qt)
      elecResult_Save.ECOUNT.MakeResultExcel(memuser.seq_no,memuser.biz_no,memuser.biz_name,workyear,vatKikan,work_qt)
      elecResult_Save.ECOUNT.Semusarang_TransactionExcelUpload()
    elif flagProc=="업로드엑셀생성":      
      elecResult_Save.ECOUNT.MakeResultExcel(memuser.seq_no,memuser.biz_no,memuser.biz_name,workyear,vatKikan,work_qt)
      elecResult_Save.ECOUNT.Semusarang_TransactionExcelUpload()      
    elif flagProc=="세무사랑반영":
      elecResult_Save.ECOUNT.Semusarang_TransactionExcelUpload()

  elif flag=='1':#일반 분개장 업무
    arrYear = [workyear,workyear-1,workyear-2,workyear-3,workyear-4,workyear-5,workyear-6,workyear-7,workyear-8,workyear-9,workyear-10]
    flagGroup = "";flagSeqNo='';flagYear='';flagManagerName ='';flagPeriod = '연도별'
    flagFirst = pyautogui.confirm('분개장 업로드 대상을 선택하세요',buttons=['관리자별','단일사업자','전체'])
    if flagFirst=="관리자별":
      sqladmin = "select admin_Name,Htx_ID,admin_id,manage_YN from mem_admin  order by admin_id"
      cursor = connection.cursor()
      cursor.execute(sqladmin)
      admins = cursor.fetchall()
      connection.commit()

      arrAdmin = []
      for admin in admins:
        if admin[3]=="Y" :
          arrAdmin.append(admin[2])
      flagManagerName = pyautogui.confirm('그룹작업 진행할 관리자를 선택하세요. ',buttons=arrAdmin)
      for admin in admins:
        if admin[0]==flagManagerName:  
          flagGroup += "'"+admin[2]+"',"         
    elif flagFirst=="단일사업자":
      flagSeqNo = pyautogui.prompt('SEQNO를 입력하세요.');  
      flagPeriod = pyautogui.confirm('작업 기간을 선택하세요',buttons=['연도별','1기부터'])  
    if flagPeriod=="연도별":flagYear = pyautogui.confirm('작업할 연도를 선택하세요. ',buttons=arrYear)

    strsql = " select biz_no,year(reg_date),a.seq_no,fiscalMM,biz_type,duzon_id,biz_name,goyoung_banki from mem_user a,mem_deal b where a.seq_no=b.seq_no  "
    if flagFirst=="관리자별":     
      strsql += " and biz_manager ='" + flagManagerName + "' "
      strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
      flagSecond = pyautogui.confirm('분개장 제작할 사업자 형태를 선택하세요. ',buttons=['법인','개인','전체'])  
      if flagSecond=='법인':  strsql += " and a.biz_type<4 "  
      elif flagSecond=='개인':  strsql += " and a.biz_type>=4 "      
    elif flagFirst=="단일사업자": 
      strsql += " and a.seq_no ='"+flagSeqNo+"'   "           
    elif  flagFirst=="전체":
      strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') "   
    # strsql += " and (select sum(a01) from 원천세전자신고 c where c.사업자등록번호=a.biz_no and  left(지급연월,4)='"+flagYear+"' )>0 "
    strsql += " and  duzon_ID <> '' and b.keeping_YN = 'Y' and a.Del_YN <> 'Y'   order by a.seq_no/1 "
    print(strsql)
    cursor = connection.cursor()
    cursor.execute(strsql)
    result = cursor.fetchall()
    connection.commit()
    connection.close()
    dig = utils.semusarang_LaunchProgram_App(semusarangID)
    for member in result:
      if member[5]=='1':  utils.semusarang_ChangeCompany(member[0])
      else:               utils.semusarang_ChangeCompany_ID_App(dig,member[5])  
      startKi = 0;finalKi=workyear - member[1] 
      if flagPeriod=="연도별": 
        finalKi = int(flagYear) - member[1] 
        startKi = finalKi
      print(finalKi)
      for i in range(startKi + 1, finalKi + 2):
        print(member[6]+" 분개장 작업기수:"+str(i))
        utils.semusarang_ChangeFiscalYear_App('vat',str(i));time.sleep(1)
        if (member[1]+i-1)!=workyear:endMM="12"
        elecResult_Save.ACCOUNT.semusarang_Acct_Excel(member[2],member[1]+i-1,member[3],member[4],endMM);time.sleep(1)
        
  print(f'정상종료 : {flagFirst}:{flagManagerName} - {flag} - {flagSeqNo}' )
  return JsonResponse({'data':'성공 : 분개장업로드'},safe=False) 


# 세무사랑작업 - 간이지급명세서-근로
def SS_Kani(request):
  work_mm = datetime.now().month - 1  #매월 신고의 경우 귀속월의 다음달 신고하므로 1현재월에서 1을 뺀다
  text_mm = str(work_mm)              # 귀속월임
  workyear = datetime.now().year 
  if work_mm<10 : text_mm = "0" + str(work_mm)
  if work_mm==0: workyear = datetime.now().year - 1; work_mm = 12;text_mm='12' # 현재1월인 경우 12월 귀속분 신고함

  flag=request.GET.get('flag',False)

  flagGroup = "";flagSeqNo='';flagManagerName =''
  flagFirst = pyautogui.confirm('간이지급명세서 제작 대상을 선택하세요',buttons=['관리자별','단일사업자','전체'])
  if flagFirst=="관리자별":
    sqladmin = "select admin_Name,Htx_ID,admin_id,manage_YN from mem_admin  order by admin_id"
    cursor = connection.cursor()
    cursor.execute(sqladmin)
    admins = cursor.fetchall()
    connection.commit()
    arrAdmin = []
    for admin in admins:
      if admin[3]=="Y" :
        arrAdmin.append(admin[2])
    flagManagerName = pyautogui.confirm('그룹작업 진행할 관리자를 선택하세요. ',buttons=arrAdmin)
    for admin in admins:
      if admin[0]==flagManagerName:  
        flagGroup += "'"+admin[2]+"'," 
  elif flagFirst=="단일사업자":
    flagSeqNo = pyautogui.prompt('SEQNO를 입력하세요.');  

  text_mm_eng = ""
  if work_mm<=9 :    text_mm_eng = str(workyear) +"-0" + str(work_mm)
  else : text_mm_eng = str(workyear) +"-"+ str(work_mm)

  if flag=='2':#사업소득 매월
    #전체 중 신고안된 업체 리스트업
    strsql = " select replace(a.biz_no,'-',''),a.seq_no,a.biz_name,duzon_id from mem_user a,mem_deal b,mem_admin c,원천세전자신고 d "
    strsql += " where a.seq_no=b.seq_no and b.biz_manager=c.admin_id   and a.biz_no=d.사업자등록번호 "
    strsql += " and a.biz_no not in(select 사업자번호 from 지급조서간이소득 where 과세년도='"+text_mm_eng+"' and 신고서종류 like '%거주자의 사업소득%' )"
    strsql += " and  과세연월='"+str(workyear)+text_mm+"' and a30>0  "
    if flagFirst=="관리자별":     
      strsql += " and biz_manager ='" + flagManagerName + "' "
      strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    elif flagFirst=="단일사업자": 
      strsql += " and a.seq_no ='"+flagSeqNo+"'   "           
    elif  flagFirst=="전체":
      strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    strsql += " union "#반기사업자 합치기
    # strsql += " select  replace(a.biz_no,'-',''),a.seq_no,a.biz_name,duzon_id from mem_user a,mem_deal b,mem_admin c, tbl_mng_jaroe e "
    strsql += " select  replace(a.biz_no,'-',''),a.seq_no,a.biz_name,duzon_id from mem_user a,mem_deal b,mem_admin c "
    # strsql += " where  a.seq_no=b.seq_no and b.biz_manager=c.admin_id and a.seq_no=e.seq_no " 
    strsql += " where  a.seq_no=b.seq_no and b.biz_manager=c.admin_id " 
    # strsql += " and work_yy='"+str(workyear)+"' and work_mm='"+str(work_mm)+"' and yn_12=1 and goyoung_banki ='Y' and kijang_YN='Y' and keeping_YN='Y'  "
    strsql += " and  goyoung_banki ='Y' and kijang_YN='Y' and keeping_YN='Y'  "
    if flagFirst=="관리자별":     
      strsql += " and biz_manager ='" + flagManagerName + "' "
      strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    elif flagFirst=="단일사업자": 
      strsql += " and a.seq_no ='"+flagSeqNo+"'   "           
    elif  flagFirst=="전체":
      strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    #strsql += " order by a.seq_no/1"
    cursor = connection.cursor()
    cursor.execute(strsql)
    members = cursor.fetchall()
    connection.commit()
    connection.close()    
    print(strsql)   
    dig = utils.semusarang_LaunchProgram_App(semusarangID)
    for member in members:
      if member[3]=='1':  utils.semusarang_ChangeCompany(member[0])
      else:               utils.semusarang_ChangeCompany_ID_App(dig,member[3])       
      utils.semusarang_ChangeFiscalYear_App('insa',str(workyear))
      pyautogui.hotkey('ctrl', 'enter') ;      time.sleep(0.5) 
      pyperclip.copy('사업소득간이지급명세서') ;      pyautogui.hotkey('ctrl', 'v');      time.sleep(1);      pyautogui.press('enter')  
      time.sleep(1) #필수
      # print('지급년 설정');pyautogui.press('enter');time.sleep(0.5)#환경등록이 지급기준인 경우 연도세팅은 하지 않음
      pyautogui.write(text_mm);      time.sleep(1) 
      if pyautogui.locateOnScreen('K:/noa-django/Noa/static/assets/images/autowork/semusarang_ZKJS_NoData_Condition.png', confidence=0.9):
        pyautogui.press('esc',presses=2, interval=0.3);        time.sleep(0.25)  ;print('조회결과 없음')
      else:#결과값이 있으면
        pyautogui.press('f8');        time.sleep(0.5) 
        if  pyautogui.locateOnScreen('K:/noa-django/Noa/static/assets/images/autowork/semusarang_MagamCancelYN.png'):          
          pyautogui.press('N');print('이미 마감됨- 마감풀까요 -> 아니오')
        else:          pyautogui.press('f8')
        time.sleep(0.5) ;        pyautogui.press('esc', presses=2, interval=1);        time.sleep(1)   
        elecResult_Save.PAY.electric_Issue_KNZK(flag,member[0],workyear,text_mm,htxLoginID)      #전자신고 파일작성 > 홈택스 전자신고까지
  elif flag=='1': #간이지급 - 근로 - 26년부터 매월
    startMM="01";endMM="06"
    if work_mm==12:text_mm_eng = str(workyear) + " 하반기";startMM="07";endMM="12"
    else:text_mm_eng = str(workyear) + " 상반기"
    strsql = " select replace(a.biz_no,'-',''),duzon_id from mem_user a, mem_deal b,mem_admin c "
    strsql += " where a.seq_no=b.seq_no and b.biz_manager=c.admin_id  and a.biz_no<>'' and keeping_yn='Y' and kijang_yn='Y' and hometaxAgree='Y'  and taltyoedate='' "
    strsql += f" and a.biz_no not in(select 사업자번호 from 지급조서간이소득 where 과세년도='{text_mm_eng}' and 신고서종류 like '%근로소득%' )"    
    strsql += f" and a.biz_no in(select 사업자등록번호 from 원천세전자신고 where 과세연월>='{workyear}{startMM}' and 과세연월<='{workyear}{endMM}' GROUP BY 사업자등록번호  HAVING SUM(a01) > 0 )  "    
    if flagFirst=="관리자별":     
      strsql += " and biz_manager ='" + flagManagerName + "' "
      strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    elif flagFirst=="단일사업자": 
      strsql += " and a.seq_no ='"+flagSeqNo+"'   "           
    elif  flagFirst=="전체":
      strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') "     
    strsql += " and biz_name <>'다온기업' " # 귀속/지급 다른 업체
    strsql += " order by a.seq_no/1";print(strsql)
    cursor = connection.cursor()
    cursor.execute(strsql)
    members = cursor.fetchall()
    connection.commit()
    connection.close()    
    if members:
      dig = utils.semusarang_LaunchProgram_App(semusarangID)
      for member in members:
        if member[1]=='1':  utils.semusarang_ChangeCompany(member[0])
        else:               utils.semusarang_ChangeCompany_ID_App(dig,member[1])      
        utils.semusarang_ChangeFiscalYear_App('insa',str(workyear))
        pyautogui.hotkey('ctrl', 'enter');      time.sleep(0.5) 
        pyperclip.copy('근로소득간이지급명세서') ;   pyautogui.hotkey('ctrl', 'v');  time.sleep(1);  pyautogui.press('enter')  
        time.sleep(1) #필수
        # 하반기 선택
        period='1';print(f"work_mm:{work_mm}")
        if work_mm==12:         pyautogui.press('2');    period='2'#하반기
        else          :       pyautogui.press('1')
        time.sleep(1) 
        pyautogui.press('enter');print("저장된 데이터가 없습니다 새로 불러오시겠습니까: 예"); time.sleep(1.5) 
        if  pyautogui.locateOnScreen('K:/noa-django/Noa/static/assets/images/autowork/semusarang_Insa_NoData(Doen).png',confidence=0.8):
          pyautogui.press('enter');print('조회된 데이터가 없습니다.확인')
          pyautogui.press('esc', presses=4, interval=0.1);print('esc 4회');time.sleep(0.5)
          continue
        else:
          print('팝업없음 : 조회된 데이터가 없습니다.')       
        pyautogui.press('f4');      time.sleep(1) 
        pyautogui.press('Y');      time.sleep(1) 
        pyautogui.press('f8');      time.sleep(0.5) 
        if  pyautogui.locateOnScreen('K:/noa-django/Noa/static/assets/images/autowork/semusarang_MagamCancelYN.png',confidence=0.8):
          pyautogui.press('N');print('마감 취소되었습니다. 해제하시겠습니까? 아니오 선택')
        else:
          pyautogui.press('f8');print('마감 F8 선택')
        time.sleep(0.3) 
        pyautogui.press('esc', presses=4, interval=0.1);print('esc 4회');time.sleep(2) 
        elecResult_Save.PAY.electric_Issue_KNZK(flag,member[0],workyear,period,htxLoginID)
  elif flag=='3': #간이지급 - 일용
    #전체 중 신고안된 업체 리스트업
    strsql = " select replace(a.biz_no,'-',''),a.seq_no,a.biz_name,duzon_id from mem_user a,mem_deal b,mem_admin c,원천세전자신고 d "
    strsql += " where a.seq_no=b.seq_no and b.biz_manager=c.admin_id   and a.biz_no=d.사업자등록번호 and "
    strsql += f" ( ( a.biz_no not in(select 사업자번호 from 지급조서간이소득 where 과세년도='"+text_mm_eng+"' and 신고서종류 = '일용근로소 득지급명세서' )) or"
    strsql +=  f"((select bigo from tbl_mng_jaroe d where a.seq_no =d.seq_no and work_yy={workyear} and work_mm={work_mm}) is  null)"
    strsql += " ) "
    strsql += " and  과세연월='"+str(workyear)+text_mm+"' and a03>0  "
    if flagFirst=="관리자별":     
      strsql += " and biz_manager ='" + flagManagerName + "' "
      strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    elif flagFirst=="단일사업자": 
      strsql += " and a.seq_no ='"+flagSeqNo+"'   "           
    elif  flagFirst=="전체":
      strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    strsql += " union "#반기사업자 합치기
    strsql += " select  a.biz_no,a.seq_no,a.biz_name,duzon_id from mem_user a,mem_deal b,mem_admin c, tbl_mng_jaroe e "
    strsql += " where  a.seq_no=b.seq_no and b.biz_manager=c.admin_id and a.seq_no=e.seq_no " 
    strsql += " and work_yy='"+str(workyear)+"' and work_mm='"+str(work_mm)+"' and yn_12=1 and goyoung_banki ='Y' and kijang_YN='Y' and keeping_YN='Y'  "
    if flagFirst=="관리자별":     
      strsql += " and biz_manager ='" + flagManagerName + "' "
      strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    elif flagFirst=="단일사업자": 
      strsql += " and a.seq_no ='"+flagSeqNo+"'   "           
    elif  flagFirst=="전체":
      strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') "
    
    print(strsql) 
    cursor = connection.cursor()
    cursor.execute(strsql)
    members = cursor.fetchall()
    connection.commit() 
    if members:
      dig = utils.semusarang_LaunchProgram_App(semusarangID)
      for member in members:
        if member[3]=='1':  utils.semusarang_ChangeCompany(member[0])
        else:               utils.semusarang_ChangeCompany_ID_App(dig,member[3])       
        utils.semusarang_ChangeFiscalYear_App('insa',str(workyear))

        # 일용직간이지급명세서
        strsql = f"select 사업자번호 from 지급조서간이소득 where 과세년도='{text_mm_eng}' and 신고서종류 = '일용근로소득 지급명세서' and 사업자번호='{member[0][:3]}-{member[0][3:5]}-{member[0][5:10]}'";print(strsql)
        cursor.execute(strsql)
        result_Kani = cursor.fetchone()
        connection.commit() 
        if not result_Kani:
          pyautogui.hotkey('ctrl', 'enter') ;      time.sleep(0.5) 
          pyperclip.copy('일용근로소득지급명세서') ;      pyautogui.hotkey('ctrl', 'v');      time.sleep(1);      pyautogui.press('enter')  
          time.sleep(1) #필수
          pyautogui.write('1');time.sleep(0.25)#월별
          pyautogui.write(text_mm);      time.sleep(1) 
          if pyautogui.locateOnScreen('K:/noa-django/Noa/static/assets/images/autowork/semusarang_ZKJS_NoData_Condition.png', confidence=0.9):
            pyautogui.press('esc',presses=2, interval=0.3);        time.sleep(0.25)  ;print('조회결과 없음')
          else:#결과값이 있으면 마감
            pyautogui.press('f7');        time.sleep(0.5) 
            if  pyautogui.locateOnScreen('K:/noa-django/Noa/static/assets/images/autowork/semusarang_MagamCancelYN.png'):          
              pyautogui.press('N');print('이미 마감됨- 마감풀까요 -> 아니오')
            else:          
              pyautogui.press('f7');time.sleep(0.5) 
              pyautogui.press('Y');print('선택된 데이터가 없습니다 전체를 마감하시겠습니까 -> 예');time.sleep(0.5) 
              pyautogui.press('enter');print('데이터를 마감했습니다. -> 확인');
            time.sleep(0.5) ;        pyautogui.press('esc', presses=3, interval=1);        time.sleep(1)   
            jubsuNo = elecResult_Save.PAY.electric_Issue_KNZK(flag,member[0],workyear,text_mm,htxLoginID) 

        else:
          print(f"{member[1]} 일용직 간이지급명세서 제출함") 
        # 일용직 근로내용확인서
        # strsql = f"select bigo from tbl_mng_jaroe d where seq_no ={member[1]} and work_yy={workyear} and work_mm={work_mm}";print(strsql)
        # cursor.execute(strsql)
        # result_ilyoung = cursor.fetchone()
        # connection.commit() 
        # if  result_ilyoung is None or result_ilyoung[0]=='':   
        #   menuName='세무사랑 Pro'
        #   procs = pywinauto.findwindows.find_elements();handle=''
        #   for proc in procs: 
        #     if proc.name== menuName:handle = proc.handle;break
        #   app = Application().connect(handle=handle)
        #   w_open = app.window(handle=handle)
        #   w_open.set_focus() 
        #   pyautogui.hotkey('ctrl', 'enter') ;      time.sleep(0.5) 
        #   pyperclip.copy('근로내용확인신고서') ;      pyautogui.hotkey('ctrl', 'v');      time.sleep(1);      pyautogui.press('enter')  
        #   time.sleep(2) #필수
        #   cnt_enter = 2
        #   if member[1].strip()=='4168':cnt_enter = 3#대건건업
        #   pyautogui.press('enter',presses=cnt_enter,interval=0.5)#엔터 3회는 관리번호 2개 이상인 경우
        #   pyautogui.write(text_mm);      time.sleep(0.5) 
        #   if pyautogui.locateOnScreen('K:/noa-django/Noa/static/assets/images/autowork/semusarang_IsSaved_Ilyoung.png', confidence=0.9):
        #     pyautogui.press('N');        time.sleep(0.5)  ;print('기존에 저장된 데이터를 불러오시겠습니까? 아니오')
        #   pyautogui.hotkey('shift','f8');        time.sleep(0.5);pyautogui.press('tab');print('엑셀저장위치 C:\\NewGen\\Rebirth\\리버스문서보관함\\EDI');  time.sleep(1.5);        
        #   pyautogui.press('esc',presses=4,interval=0.5)
        #   jubsuNo = elecResult_Save.EDI.Total_IlyoungIssue(member[0],member[2],workyear,text_mm)
        #   if jubsuNo!='' and jubsuNo is not None:
        #     strsql3 =  f" Merge tbl_mng_jaroe as A Using (select '{member[1]}' as seq_no,'{workyear}' as work_YY, '{work_mm}' as work_MM) as B "
        #     strsql3 += " On A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_MM=B.work_MM "
        #     strsql3 += f" when matched then update  set bigo='{jubsuNo}' "
        #     strsql3 += f" When Not Matched Then insert  values('{member[1]}','{workyear}','{work_mm}'"
        #     for j in range(1,15):strsql3 = strsql3 + ",'0'" 
        #     strsql3 += ",'{jubsuNo}');"
        #     print(strsql3)
        #     cursor.execute(strsql3)          
        # else:
        #   print(f'{member[1]}일용직근로내용확인서 신고함')
  print(f'정상종료 : {flagFirst}:{flagManagerName} - {flag} - {flagSeqNo}' )
  return JsonResponse({'data':'성공 : 간이지급명세서-근로소득'},safe=False)


# 세무사랑작업 - 지급명세서
def SS_ZZMS(request):
  workyear = datetime.now().year-1
  flag = request.GET.get('flag',False)

  flagGroup = "";flagSeqNo='';flagManagerName =''
  flagFirst = pyautogui.confirm('지급명세서 제작 대상을 선택하세요',buttons=['관리자별','단일사업자','전체'])
  if flagFirst=="관리자별":
    sqladmin = "select admin_Name,Htx_ID,admin_id,manage_YN from mem_admin  order by admin_id"
    cursor = connection.cursor()
    cursor.execute(sqladmin)
    admins = cursor.fetchall()
    connection.commit()
    arrAdmin = []
    for admin in admins:
      if admin[3]=="Y" :
        arrAdmin.append(admin[2])
    flagManagerName = pyautogui.confirm('그룹작업 진행할 관리자를 선택하세요. ',buttons=arrAdmin)
    for admin in admins:
      if admin[0]==flagManagerName:  
        flagGroup += "'"+admin[2]+"'," 
  elif flagFirst=="단일사업자":
    flagSeqNo = pyautogui.prompt('SEQNO를 입력하세요.');  


  txtSingoseo=""
  if flag=='01':txtSingoseo = "근로소득지급명세서"
  elif flag=='20':txtSingoseo = "퇴직소득지급명세서"
  elif flag=='30':txtSingoseo = "거주자 사업소득지급명세서"
  elif flag=='40':txtSingoseo = "거주자 기타소득지급명세서"
  elif flag=='50':txtSingoseo = "이자소득지급명세서"
  elif flag=='60':txtSingoseo = "배당소득지급명세서"
  
  strsql = " select replace(a.biz_no,'-',''),duzon_id from mem_user a, mem_deal b,mem_admin c "
  strsql += " where a.seq_no=b.seq_no and b.biz_manager=c.admin_id  and a.biz_no<>'' and hometaxAgree='Y'   "
  strsql += f" and a.biz_no in(select 사업자등록번호 from 원천세전자신고 where left(과세연월,4)='{workyear}' GROUP BY 사업자등록번호  HAVING SUM(a{flag}) > 0 )  "    
  strsql += f" and a.biz_no not in(select 사업자번호 from 지급조서전자신고 where 과세년도='{workyear}' and 신고서종류='{txtSingoseo}' GROUP BY 사업자번호 )  "    
  if flagFirst=="관리자별":     
    strsql += "  and keeping_yn='Y' and kijang_yn='Y' and taltyoedate='' and  biz_manager ='" + flagManagerName + "' "
    strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
  elif flagFirst=="단일사업자": 
    strsql += " and a.seq_no ='"+flagSeqNo+"'   "           
  elif  flagFirst=="전체":
    strsql += "  and keeping_yn='Y' and kijang_yn='Y' and taltyoedate='' and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') "     
  strsql += " order by a.seq_no/1";print(strsql)
  cursor = connection.cursor()
  cursor.execute(strsql)
  members = cursor.fetchall()
  connection.commit()
  connection.close()    
  print(strsql)
  dig = utils.semusarang_LaunchProgram_App(semusarangID)
  for member in members:
    if member[1]=='1':  utils.semusarang_ChangeCompany(member[0])
    else:               utils.semusarang_ChangeCompany_ID_App(dig,member[1])       
    utils.semusarang_ChangeFiscalYear_App('insa',str(workyear))
    elecResult_Save.PAY.semusarang_Menu_ZZJS(flag,member[0],workyear,txtSingoseo,htxLoginID)#전자신고파일 제작 -> 홈택스신고
  print(f'정상종료 : {flagManagerName} - {flag} - {txtSingoseo}' )
  return JsonResponse({'data':'성공 : 지급명세서'},safe=False) 



# import requests
# import pypinksign
# import ssl
# from datetime import datetime
# import base64
# import hashlib
# import urllib.parse
# from urllib import parse
# from bs4 import BeautifulSoup
# from xml.etree import ElementTree

# def get_open():
#     ### 값 설정
#     url = 'https://hometax.go.kr/permission.do?screenId=index'
#     ### 요청
#     res = requests.post(url)
#     ### 지정 : TXPPsessionID
#     TXPPsessionID = res.cookies.get_dict()['TXPPsessionID']
#     print(f"1. 로그인 페이지에서 TXPPsessionID 값 얻기 get_open(): {TXPPsessionID}")
#     return TXPPsessionID
    
# ### Public Key Cryptography (공개키 암호화)
# def get_pkcEncSsn(TXPPsessionID):
#     ### 값 설정
#     url = 'https://hometax.go.kr/wqAction.do?actionId=ATXPPZXA001R01&screenId=UTXPPABA01'
#     headers = {'Cookie': f'TXPPsessionID={TXPPsessionID}'}
#     ### 요청
#     res = requests.get(url=url, headers=headers)
#     ### 지정 pkcEncSsn
#     root = ElementTree.fromstring(res.content) ### xml
#     pkcEncSsn = root.find("pkcEncSsn").text
#     ### 지정 WMONID
#     WMONID = res.cookies.get_dict()['WMONID'] ### cookie
#     print(f"2. 로그인 페이지에서 pkcEncSsn:{pkcEncSsn}, WMONID:{WMONID} 얻기 get_pkcEncSsn(TXPPsessionID)")
#     return pkcEncSsn , WMONID
    
# ### 로그 서명 (Log Signature)
# def get_logsgnt(pkcEncSsn, path, password, serialnum):
#     ### 비밀번호 인코드
#     password = bytes(password, 'utf-8')
#     ### 서명
#     p = pypinksign.PinkSign() ### PinkSign(한국정보인증) API를 사용하기 위한 객체를 생성
#     p.load_pubkey(pubkey_path=f"{path}/signCert.der") ### 공개키를 로드
#     p.load_prikey(prikey_path=f"{path}/signPri.key", prikey_password=password) ### 개인키를 로드
#     byte_pkcEncSsn = bytes(pkcEncSsn, 'utf-8')
#     sign = p.sign(byte_pkcEncSsn)  # 서명
#     pem_cert = ssl.DER_cert_to_PEM_cert(sign) ### SSL/TLS 인증서를 DER(Distinguished Encoding Rules) 형식에서 PEM(Privacy-Enhanced Mail) 형식으로 변환
#     sign = "".join(pem_cert.split('\n')[1:-2]) ### 양식 설정
#     ### 시간설정
#     now = datetime.now()
#     time = now.strftime('%Y%m%d%H%M%S')
    
#     ### logsgnt 생성
#     logsgnt = f'{pkcEncSsn}${serialnum}${time}${sign}' ### 서명값 이어붙이기
#     logsgnt = base64.b64encode(logsgnt.encode('utf-8')) ### base64 인코딩
#     logsgnt = parse.quote(logsgnt) ### 퍼센트 인코딩
#     print(f"3. 공인인증서 서명값 설정:{logsgnt}")
#     return logsgnt 
    
# def get_login(logsgnt, WMONID, TXPPsessionID, randomEnc, cert):
#     ### 값 설정
#     url = 'https://hometax.go.kr/pubcLogin.do?domain=hometax.go.kr&mainSys=Y'
#     headers = {'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
#                'Cookie': f'WMONID={WMONID}; TXPPsessionID={TXPPsessionID};'}
#     payload = f'logSgnt={logsgnt}&cert={urllib.parse.quote(cert)}&randomEnc={urllib.parse.quote(randomEnc)}&pkcLoginYnImpv=Y&pkcLgnClCd=04&ssoStatus=&portalStatus=&scrnId=UTXPPABA01&userScrnRslnXcCnt=1920&userScrnRslnYcCnt=1080'
#     ## 요청
#     res = requests.post(url = url, headers = headers, data = payload)
#     ### 지정 TXPPsessionID
#     TXPPsessionID = res.cookies.get_dict()['TXPPsessionID']
#     print(f"4. 공인 인증서 로그인 : {TXPPsessionID}")
#     return TXPPsessionID
    
# def get_ssotoken(session_url, TXPPsessionID):
#     ### 값 설정
#     headers = {'Cookie': f'TXPPsessionID={TXPPsessionID};'}
#     ## 요청
#     res = requests.get(url=session_url ,headers=headers)
#     ### 지정 ssotoken
#     soup = BeautifulSoup(res.text, 'html.parser')
#     ssotoken = soup.findAll('ssotoken')[0].text
#     print(f"5. 세션값 받기 : {ssotoken}")
#     return ssotoken 
    
# def get_session(ssotoken):
#     ### 값 설정
#     url = 'https://teet.hometax.go.kr/permission.do?screenId=UTEETBDA03&domain=hometax.go.kr'
#     payload = f"<map id='postParam'><ssoToken>{ssotoken}</ssoToken><userClCd>02</userClCd><popupYn>false</popupYn></map>"
#     ## 요청
#     res= requests.post(url = url, data = payload)
#     print(res.cookies)
#     ### 지정 TEETsessionID
#     TEETsessionID = res.cookies.get_dict()['TEETsessionID']
#     WMONID = res.cookies.get_dict()['WMONID']
#     print(f"6. 요청 세션 받기 : {TEETsessionID}")
#     return TEETsessionID,WMONID

# def login_IDPW(id, password):
#     url = "https://www.hometax.go.kr/pubcLogin.do"
#     data = {
#         'ssoLoginYn': 'Y',
#         'secCardLoginYn': '',
#         'secCardId': '',
#         'id': id,
#         'pswd': password,
#         'ssoStatus': '',
#         'portalStatus': '',
#         'scrnId': 'UTXPPABA01',
#         'userScrnRslnXcCnt': '1280',
#         'userScrnRslnYcCnt': '960'
#     }
#     session = requests.Session()
#     response = session.post(url, data=data)
#     return session, response.text
# def manageNO_IDPW(session,WMONID,TXPPsessionID,manageNo,pw):
#     password = parse.quote(pw)#urllib.parse.quote(pw) #parse.quote(pw)# bytes(pw, 'utf-8')
#     hashed_pw = hashlib.sha256(pw.encode()).hexdigest()
#     url = "https://www.hometax.go.kr/pubcLogin.do?operate=txaaLogin"
#     headers = {'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
#                'Cookie': f'WMONID={WMONID}; TXPPsessionID={TXPPsessionID};'}
#     payload = f'txaaAdmNo={manageNo}&txaaPswd={hashed_pw}'
#     response = session.post(url = url, headers = headers, data = payload)
#     response = requests.post(url = url, headers = headers, data = payload)
#     return  response

# def request_permission(session):
#     url = "https://www.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/pp/index.xml"
#     response = session.get(url)
#     return response.text

# def apply_cookies_to_selenium(driver, cookies):
#     for key, value in cookies.items():
#         driver.add_cookie({'name': key, 'value': value})

def MyWork(request):
  today = str(datetime.now())[:10].replace("-","")
  flag = request.GET.get('flag',False)
  if flag=='1':#로그인만
    flagPlatform = pyautogui.confirm(' 작업할 형태를 선택하세요',buttons=['홈택스(일반)','홈택스(아이디)','홈택스(인증서)','위택스','고용토탈']) 
    if flagPlatform=='홈택스(일반)':
      driver = utils.conHometaxLogin(htxLoginID,False) ;time.sleep(900)
    elif flagPlatform=='홈택스(아이디)':
      id = "daeseung21"
      # passwd = "daeseung@1128"
      # session, login_response = login_IDPW(id, passwd)
      # print(login_response)
      # permission_response = request_permission(session)
      # print(permission_response)
    elif flagPlatform=='홈택스(인증서)':
      id = "daeseung21"
      # path = "C:\\Users\\Administrator\\AppData\\LocalLow\\NPKI\\yessign\\USER\\cn=세무법인대승(DAESEUNG)008102820101129181000504,ou=DAESEUNG,ou=HNB,ou=corporation,o=yessign,c=kr"
      # password = "daeseung@1128"
      # serialnum = '32 15 09 2f' #공인인증서 폴더에 들어가면"signCert" 라는 파일이 있는데 더블크릭하면 인증서가 뜬다"자세히" 메뉴에 보면 일련 번호가 있는데 이게 serialnum 이다.
      # randomEnc = 'udhXw9AWjJrLqvCUQPRWHt+uVDs='
      # cert='-----BEGIN CERTIFICATE-----MIIF3TCCBMWgAwIBAgIEMhUJLzANBgkqhkiG9w0BAQsFADBSMQswCQYDVQQGEwJrcjEQMA4GA1UECgwHeWVzc2lnbjEVMBMGA1UECwwMQWNjcmVkaXRlZENBMRowGAYDVQQDDBF5ZXNzaWduQ0EgQ2xhc3MgMzAeFw0yMzEyMDYxNTAwMDBaFw0yNDEyMTIxNDU5NTlaMIGVMQswCQYDVQQGEwJrcjEQMA4GA1UECgwHeWVzc2lnbjEUMBIGA1UECwwLY29ycG9yYXRpb24xDDAKBgNVBAsMA0hOQjERMA8GA1UECwwIREFFU0VVTkcxPTA7BgNVBAMMNOyEuOustOuyleyduOuMgOyKuShEQUVTRVVORykwMDgxMDI4MjAxMDExMjkxODEwMDA1MDQwggEiMA0GCSqGSIb3DQEBAQUAA4IBDwAwggEKAoIBAQDrSQLDpO2CLal1t7HT58EBAEXn0kHLf+ouhjHuJttL/r+kYVpwhIf/XF3hyhJRimNbYyFV2PWuPXtUZJvgi3IzDbS3Oy+w5yeWQwY/TJ8O845pNg68daIVBBaKm1eRtW2K2M2O5j1AqIWXawJebQnGef56/FI/lf8IOiaO9WrrUy74BeA/Mp0CnNwGqoeAXlFYeKBOTjuBAYtbLvyNbpXZ2hGhXWSj1Uxfk4/TPIsVaksazchxdLaFHoxJc4WK0rYuiiLfRCRTIl8OnVLiVrO8ChfMR6H3UtWu7caEY30ksMD343TqwP+yFb83rwqPPTwUOJvNWZZ2agDdONzI29AzAgMBAAGjggJ1MIICcTCBjwYDVR0jBIGHMIGEgBTyh6Pm2V4WFnJO2MK8hTkDN1mQxKFopGYwZDELMAkGA1UEBhMCS1IxDTALBgNVBAoMBEtJU0ExLjAsBgNVBAsMJUtvcmVhIENlcnRpZmljYXRpb24gQXV0aG9yaXR5IENlbnRyYWwxFjAUBgNVBAMMDUtJU0EgUm9vdENBIDSCAhAoMB0GA1UdDgQWBBRr3C9U16/3ObuszuqlyKpewWYwtDAOBgNVHQ8BAf8EBAMCBsAwgYwGA1UdIAEB/wSBgTB/MH0GCSqDGoyaRQEBAjBwMEAGCCsGAQUFBwICMDQeMsd0ACDHeMmdwRyylAAgrgjHNaywyBzG0MXQwRwAILwcrgnVXAAgx3jJncEcx4WyyLLkMCwGCCsGAQUFBwIBFiBodHRwOi8vd3d3Lnllc3NpZ24ub3Iua3IvY3BzLmh0bTBxBgNVHREEajBooGYGCSqDGoyaRAoBAaBZMFcMEuyEuOustOuyleyduOuMgOyKuTBBMD8GCiqDGoyaRAoBAQEwMTALBglghkgBZQMEAgGgIgQgdGP5o1clsStHDdietqO90z+MvVlB4Pp+eI2+tNfvUrAwcgYDVR0fBGswaTBnoGWgY4ZhbGRhcDovL2RzLnllc3NpZ24ub3Iua3I6Mzg5L291PWRwNnAxODY2MixvdT1BY2NyZWRpdGVkQ0Esbz15ZXNzaWduLGM9a3I/Y2VydGlmaWNhdGVSZXZvY2F0aW9uTGlzdDA4BggrBgEFBQcBAQQsMCowKAYIKwYBBQUHMAGGHGh0dHA6Ly9vY3NwLnllc3NpZ24ub3JnOjQ2MTIwDQYJKoZIhvcNAQELBQADggEBAJH/rHFpUYXFvjfOhlb79lX2hl1gRS6F/1UZ2I8cF8pHb9a/GIv7NOXcUpkhNszJdsIXf1umf2LVisOR13tc1hf3dl3RvklcbEHDIv8cJA8ws4YTxxMBRdQDbZ5eWWEHtAvAh8LUQZjr1o0LR+av3oK3/eJH1g0yPlTUcW/cOMMYsAcH1lzp+Vg70TXJlJHOMVE5vBSysuJgu0/v3Dp9k3cWpoG1YTw17ahqKBmnfJfZS8eT4ZVvE4TOCPl12STNuCyX3fwrH9W9YY/A3QvwfI+zh+J3tLJ6FViiIzI8+PXGqRauAV5QpkY/AWuf4rCqRbs1MpPVbyYEDvF3s2IiNNc=-----END CERTIFICATE-----'
      # session_url='https://hometax.go.kr/token.do?query=_EUzMYrocZjvnXTVNLxJX&postfix=2024_09_20'
      # ### 1. 로그인 페이지에서 TXPPsessionID 값 얻기
      # session, login_response = login_IDPW(id, password)
      # print(f"1. 로그인 : {login_response}")
      # TXPPsessionID = session.cookies.get_dict()['TXPPsessionID']
      # ### 2. 로그인 페이지에서 pkcEncSsn, WMONID 얻기
      # pkcEncSsn , WMONID = get_pkcEncSsn(TXPPsessionID)

      # ### 3. 서명값 설정
      # logsgnt = get_logsgnt(pkcEncSsn, path, password, serialnum)
      # ### 4. 공인 인증서 로그인
      # TXPPsessionID = get_login(logsgnt, WMONID, TXPPsessionID, randomEnc, cert)
      # ### 5. 세션값 받기   
      # ssotoken = get_ssotoken(session_url, TXPPsessionID)
      # ### 6. 요청 세션 받기
      # TEETsessionID,WMONID = get_session(ssotoken)

      # driver = utils.ChromeDriver(False)
      # driver.get('https://www.hometax.go.kr/wqAction.do?actionId=ATXPPCBA001R020&screenId=index_pp&popupYn=false');time.sleep(3)
      # # Requests에서 얻은 쿠키 적용
      # #cookies = {'TXPPsessionID': TXPPsessionID, 'WMONID': WMONID}
      # cookies = {'TEETsessionID': TEETsessionID, 'WMONID': WMONID}
      # apply_cookies_to_selenium(driver, cookies)
      
      # win = pyautogui.getActiveWindow();win.maximize();time.sleep(5) #윈도우 최대화

      # manageLogin_url = "https://www.hometax.go.kr/websquare/websquare.wq?w2xPath=/ui/comm/a/b/UTEABHAA19.xml&amp;w2xHome=/ui/pp/&amp;w2xDocumentRoot="
      # driver.get(manageLogin_url)
      # time.sleep(10)


      # manage_response = manageNO_IDPW(session,WMONID,TEETsessionID,'P30447',password)
      # soup = BeautifulSoup(manage_response.content, 'html.parser')
      # print(f"7. 관리자로그인 : {soup}")

    elif flagPlatform== '위택스':
      driver = utils.conWetaxLogin(False);time.sleep(900)
    elif flagPlatform== '고용토탈':
      driver = utils.conTotalComwelLogin('220-85-33586');time.sleep(900)
  elif flag=='2':#CMS정리
    memuser = MemUser.objects.get(seq_no='3639') 
    flagProc = pyautogui.confirm('세무포털 처리할 단계를 지정하세요',buttons=['거래처등록부터','세무포털조회부터','결재결과작업(98002)'])
    if flagProc=='거래처등록부터':
      #0. 세무사랑 거래처등록
      utils.semusarang_Login(semusarangID,memuser.duzon_id,memuser.biz_no,memuser.reg_date,datetime.today().year)
      elecResult_Save.ECOUNT.TraderUpload(memuser.seq_no)
      #1. 세무포털로그인 - 출금조회 - 엑셀저장 - 디비 업로드
      driver = utils.conSemuportalLogin()
      elecResult_Save.SMPT.smpt_SearchExcract(driver)

    elif flagProc== '세무포털조회부터':
      #1. 세무포털로그인 - 출금조회 - 엑셀저장 - 디비 업로드
      driver = utils.conSemuportalLogin()
      elecResult_Save.SMPT.smpt_SearchExcract(driver)
    elif flagProc== '결재결과작업(98002)':
      #4. 엑셀자료 일반전표 전송자료 만들기
      # 파일 대화상자를 표시하여 파일 선택
      file_name = utils.select_file()      
      elecResult_Save.SMPT.smpt_SaveExcractResult(file_name)       
  elif flag=='3':#세무대리
    flagProc = pyautogui.confirm('세무포털 처리할 단계를 지정하세요',buttons=['수임등록','수임상태조회','수임해지'])  
    flagSeqNo = pyautogui.prompt('SEQNO를 입력하세요.')
    memuser = MemUser.objects.get(seq_no=flagSeqNo)
    driver = utils.conHometaxLogin(htxLoginID,False) 
    if flagProc=='수임등록':
      elecResult_Save.Htx_TotalMenu(driver,'tabs_08','기장대리수임납세자등록') 
      taPrxClntClCd_input = 'taPrxClntClCd_input_' #  0:개인사업자 1: 법인사업자 2:비사업자
      if memuser.biz_type<4: taPrxClntClCd_input += "1"
      elif memuser.biz_type<7: taPrxClntClCd_input += "0"
      elif memuser.biz_type==7: taPrxClntClCd_input += "2"
      WebDriverWait(driver, 61).until(EC.element_to_be_clickable((By.ID, taPrxClntClCd_input))).click();print('유형구분') 
      if memuser.biz_type!=7:
        bizNo = memuser.biz_no.split('-')
        driver.find_element(By.ID,'bsno1').send_keys(bizNo[0]);time.sleep(0.25);    
        driver.find_element(By.ID,'bsno2').send_keys(bizNo[1]);time.sleep(0.25);          
        driver.find_element(By.ID,'bsno3').send_keys(bizNo[2]);time.sleep(0.25); 
      driver.find_element(By.ID,'resno').send_keys(memuser.ssn);time.sleep(0.25);
      driver.find_element(By.ID,'telno1').send_keys('02');time.sleep(0.25);    
      driver.find_element(By.ID,'telno2').send_keys('501');time.sleep(0.25);          
      driver.find_element(By.ID,'telno3').send_keys('1732');time.sleep(0.25);      
      select1 = Select(driver.find_element(By.ID,'mp1'));        select1.select_by_visible_text('010');time.sleep(0.5)
      driver.find_element(By.ID,'mp2').send_keys('1111');time.sleep(0.25);          
      driver.find_element(By.ID,'mp3').send_keys('1111');time.sleep(0.25);   
      driver.find_element(By.ID,'afaDt_input').send_keys(today);time.sleep(0.25); 
      if memuser.biz_type>3 and memuser.biz_type<7:#정보제공범위 0:infrOfrRngCd_input_0 타소득포함 / 해당사업장
        driver.find_element(By.ID,'infrOfrRngCd_input_1').click()
      driver.find_element(By.ID,'trigger85').click();time.sleep(1);
      #수임동의신청하시겟습니까
      WebDriverWait(driver, 5).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(1)  
      print(f'{memuser.biz_name} : 수임동의 등록하기 완료');time.sleep(3)
    elif flagProc=='수임상태조회':
      elecResult_Save.Htx_TotalMenu(driver,'tabs_08','기장대리수임납세자조회')  
      WebDriverWait(driver, 61).until(EC.element_to_be_clickable((By.ID, 'edtCntn'))).send_keys(memuser.biz_no.replace('-',''))
      WebDriverWait(driver, 61).until(EC.element_to_be_clickable((By.ID, 'trigger43'))).click();time.sleep(1)
      table = driver.find_element(By.ID,'grid_afdsSttnInfr_body_table')
      tbody = table.find_element(By.ID,'grid_afdsSttnInfr_body_tbody')    
      print(len(tbody.find_elements(By.TAG_NAME,'tr')))
      for tr in tbody.find_elements(By.TAG_NAME,'tr'):
        cols = tr.find_elements(By.TAG_NAME,'td') 
        if cols[0].text==  memuser.biz_no and cols[6].text!='': print(f'{memuser.biz_name} 수임동의일 : {cols[6].text}')
        else : print(print(f'{memuser.biz_name}({cols[0].text}) 수임동의 안함 : 수임동의 신청일 {cols[5].text}'))
        break;
    elif flagProc=='수임해지':
      elecResult_Save.Htx_TotalMenu(driver,'tabs_08','기장대리수임납세자해지')  
  elif flag=='4':#키보드반복작업 - 거래처원장
    flagFirst = pyautogui.confirm('거래처원장 스타트 지점을 선택하고 확인을 누르세요, 하단 차변에서 시작')
    menuName = "거래처원장"
    intlen = len(menuName)
    procs = pywinauto.findwindows.find_elements();handle=''
    for proc in procs: 
      if proc.name[:intlen]== menuName:handle = proc.handle;break
    app = Application().connect(handle=handle)
    w_open = app.window(handle=handle)
    w_open.set_focus()  

    time.sleep(1)
    while True:
      pyautogui.press('right  ',presses=2,interval=0.5)
      pyautogui.write('00983');time.sleep(1)
      pyautogui.press('tab',presses=2,interval=0.5)
  return JsonResponse({'data':'성공'},safe=False)


def Bank_Scrap(request):
  flagSeqNo = pyautogui.prompt('통장정리할 업체의 SEQNO를 입력하세요.');
  sqlbank = f"select trader_code,trader_name from ds_slipledgr2 where seq_no={flagSeqNo} and acnt_cd=103 and trader_code<>'' group by trader_code,trader_name"
  cursor = connection.cursor()
  cursor.execute(sqlbank)
  banks = cursor.fetchall()
  connection.commit()

  arrBank = []
  for bank in banks:
    arrBank.append(f'{bank[0]}:{bank[1]}')
  flagBank = pyautogui.confirm('작업 은행코드를 선택하세요. ',buttons=arrBank)
  utils.BankTrade_xlsx(flagSeqNo,flagBank)

    
  return JsonResponse({'data':'성공'},safe=False)





# 고용토탈 - 근로자고용정보현황
def Total_Employ(request):
  workyear = datetime.now().year-1
  bizno="220-85-33586"
  flag=request.GET.get('flag',False)
  driver =  utils.conTotalComwelLogin(bizno);time.sleep(1.5)
  if    flag=='0':    elecResult_Save.EDI.Total_SamuDaehang_Save(driver)    #사무대행위탁사업장 저장
  elif  flag=='1':    elecResult_Save.EDI.Total_Employ_Searchlist(driver,workyear)   #근로자고용정보
  elif  flag=='2':    elecResult_Save.EDI.Total_Bosu_Singo_Prework(driver,workyear)  #보수총액신고대상자 엑셀저장
  elif  flag=='3':    elecResult_Save.EDI.Total_Bosu_Singo_SS_excelupdate(driver,workyear,semusarangID)  #세무사랑 보수총액신고자 엑셀저장하면서 토탈엑셀 수정
  elif  flag=='4':    elecResult_Save.EDI.Total_Bosu_Singo_excelWrite_upload(driver,workyear,'D:')  #보수총액신고대상자 엑셀작성 후 토탈에 업로드
  time.sleep(5)
  driver.quit()
  print(f'정상종료 : 근로자고용정보현황 {workyear}: {flag} ' )
  return JsonResponse({'data:성공 : 고용토탈 작업 '+flag},safe=False)


#부가가치세 신고서 저장 - 홈택스 신고서 : 미완성
def Issue_Vat(request):
  driver = utils.conHometaxLogin(htxLoginID,False);time.sleep(1)
    # 팝업뜨는 경우 없애기
  main = driver.window_handles;    print(main);    driver.switch_to.window(main[1]); time.sleep(1)
  driver.find_element(By.CSS_SELECTOR,'#group2152').click();  driver.switch_to.window(main[0]) 
  try:
    #신고/납부
    driver.switch_to.default_content()
    time.sleep(1)
    driver.get('http://www.hometax.go.kr/websquare/websquare.wq?w2xPath=/ui/pp/index_pp.xml&tmIdx=4&tm2lIdx=&tm3lIdx=')
    time.sleep(1)
    # iframe 전환txppIframe
    iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe')
    driver.switch_to.frame(iframe)
    time.sleep(1)
    #신고/납부  > 부가가치세 클릭
    driver.find_element(By.ID,'sub_a_0405010000').click()
    time.sleep(1)
    iframe = driver.find_element(By.CSS_SELECTOR,'#UTERNAA892_iframe')
    driver.switch_to.frame(iframe)
    time.sleep(1)
    # 팝업 닫기
    driver.find_element(By.ID,'btnClose2').click()  
    time.sleep(1)
    # # iframe 전환txppIframe
    iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe')
    driver.switch_to.frame(iframe)
    time.sleep(1)
    driver.find_element(By.XPATH,'/html/body/div[1]/div[5]/ul/li[2]/div[1]/a').click()
    time.sleep(1)
    #조회하기
    searchBizNo = '1304521777'
    driver.find_element(By.ID,'input_txprRgtNo_UTERNAAZ31').send_keys(searchBizNo)
    driver.find_element(By.ID,'trigger70_UTERNAAZ31').click()
    time.sleep(2)
    #팝업 컨트롤
    WebDriverWait(driver, 5).until(EC.alert_is_present())
    al = Alert(driver)
    al.accept()
    time.sleep(1)
    # 테이블 가져오기
    table = driver.find_element(By.ID,'ttirnam101DVOListDes_body_table')
    tbody = table.find_element(By.ID,'ttirnam101DVOListDes_body_tbody')
    # tbody > tr > td
    for tr in tbody.find_elements(By.TAG_NAME,'tr'):
      cnt=0
      for td in tr.find_elements(By.TAG_NAME,"td"):
        if cnt == 10: 
          #신고서팝업
          tr.find_element(By.TAG_NAME,"a").click()
          time.sleep(2)
          main = driver.window_handles
          print(main)
          driver.switch_to.window(main[2])
          time.sleep(2)
          driver.find_element(By.ID,'ntplInfpYn_input_0').click()
          time.sleep(1)
          driver.find_element(By.ID,'trigger1').click()
          time.sleep(3)
          
          driver.switch_to.window(main[1])
          time.sleep(2)
          iframe = driver.find_element(By.CSS_SELECTOR,'#iframe2_UTERNAAZ34')
          driver.switch_to.frame(iframe)
          time.sleep(1)
          print('iiiiiii')

          # table = driver.find_element(By.CLASS_NAME,'report_menu_table')
          # td = table.find_element(By.CLASS_NAME,'report_menu_table_td')
          # btn = td.find_element(By.ID,'re_printc4da56a20a02641d4ad3cdffe1201ea3fpt14')
          
          
          print('bbbbbbb')
          # driver.find_element(By.CSS_SELECTOR,r'/html/body/div/div/div[1]/table/tbody/tr/td/div/nobr/button[6]').click()
          # print('bbbbbbb')
          break;
        cnt = cnt + 1


    time.sleep(30)  
  except:
    return JsonResponse({'data':'실패:홈택스 내부조회 중 에러'},safe=False)
  finally:
    driver.quit()    
  return JsonResponse({'data':'성공'},safe=False)

 

#전자세금계산서 스크래핑/세무사랑 업로드 합계표작성
def TaxInvoice(request):
  workyear = datetime.today().year
  workmonth =  datetime.today().month-1
  if workmonth<2:workyear = workyear-1;workmonth=12
  arrYear = [workyear,workyear-1,workyear-2]
  work_qt = utils.get_quarter(str(workmonth))
  workPeriod = str(work_qt) +"기";workPeriod2 = "작성일자 >='";mmStart="";mmEnd=""; chgStartDate = "";isJungKi = 2

  today = str(datetime.now())[:10].replace("-","")
  flag = request.GET.get('flag',False)
  flagGroup = "";flagBizNo="";flagManagerName="";flagPeriod =""
  #합계표조회인 경우 월별조회 배제
  if flag!='5' or flag!='9':   flagPeriod = pyautogui.confirm(' 작업할 형태를 선택하세요',buttons=['분기별','반기별'])  
  # else:           flagPeriod = pyautogui.confirm(' 작업할 형태를 선택하세요',buttons=['월별','분기별','반기별'])  
  flagYear = pyautogui.confirm('작업할 연도를 선택하세요. ',buttons=arrYear)
  if flagPeriod=='분기별':  
    flagPeriod2 = pyautogui.confirm('스크랩할 분기를 선택하세요',buttons=['1분기','2분기','3분기','4분기'])
    if flagPeriod2=='1분기' : work_qt=1;  workPeriod = "1기 예정";  workPeriod2 += flagYear+"-01-01' and 작성일자<='"+flagYear+"-03-31'";mmStart="1";mmEnd="03";chgStartDate=flagYear+"-01-01";isJungKi = 1
    elif flagPeriod2=='2분기' : work_qt=2;workPeriod = "1기 확정";  workPeriod2 += flagYear+"-04-01' and 작성일자<='"+flagYear+"-06-30'";mmStart="4";mmEnd="06";chgStartDate=flagYear+"-04-01"
    elif flagPeriod2=='3분기' : work_qt=3;workPeriod = "2기 예정";  workPeriod2 += flagYear+"-07-01' and 작성일자<='"+flagYear+"-09-30'";mmStart="7";mmEnd="09";chgStartDate=flagYear+"-07-01";isJungKi = 1
    elif flagPeriod2=='4분기' : work_qt=4;workPeriod = "2기 확정";  workPeriod2 += flagYear+"-10-01' and 작성일자<='"+flagYear+"-12-31'" ;mmStart="10";mmEnd="12";chgStartDate=flagYear+"-10-01"
    if (work_qt in {2,4} and work_qt<int(utils.get_quarter(str(workmonth)))  ) or int(flagYear)<workyear:isJungKi = 4 #선택분기가 확정기간 현재 분기가 선택한 분기보다 크면 기한후 신고
  elif flagPeriod=='월별':  
    flagPeriod2 = pyautogui.confirm('스크랩할 월을 선택하세요',buttons=['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'])
    tmpfeb = '-02-28'
    if int(flagYear) % 4 == 0 and (int(flagYear) % 100 != 0 or int(flagYear) % 400 == 0):tmpfeb = '-02-29'
    if flagPeriod2=='1월' : workPeriod = "01월";  workPeriod2 += flagYear+"-01-01' and 작성일자<='"+flagYear+"-01-31'";mmStart="1";mmEnd="01";chgStartDate=flagYear+"-01-01"
    elif flagPeriod2=='2월' : workPeriod = "02월";  workPeriod2 += flagYear+"-02-01' and 작성일자<='"+flagYear+tmpfeb+"'";mmStart="2";mmEnd="02";chgStartDate=flagYear+"-02-01"
    elif flagPeriod2=='3월' : workPeriod = "03월";  workPeriod2 += flagYear+"-03-01' and 작성일자<='"+flagYear+"-03-31'";mmStart="3";mmEnd="03";chgStartDate=flagYear+"-03-01"
    elif flagPeriod2=='4월' : workPeriod = "04월";  workPeriod2 += flagYear+"-04-01' and 작성일자<='"+flagYear+"-04-30'";mmStart="4";mmEnd="04";chgStartDate=flagYear+"-04-01"
    elif flagPeriod2=='5월' : workPeriod = "05월";  workPeriod2 += flagYear+"-05-01' and 작성일자<='"+flagYear+"-05-31'";mmStart="5";mmEnd="05";chgStartDate=flagYear+"-05-01"
    elif flagPeriod2=='6월' : workPeriod = "06월";  workPeriod2 += flagYear+"-06-01' and 작성일자<='"+flagYear+"-06-30'";mmStart="6";mmEnd="06";chgStartDate=flagYear+"-06-01"
    elif flagPeriod2=='7월' : workPeriod = "07월";  workPeriod2 += flagYear+"-07-01' and 작성일자<='"+flagYear+"-07-31'";mmStart="7";mmEnd="07";chgStartDate=flagYear+"-07-01"
    elif flagPeriod2=='8월' : workPeriod = "08월";  workPeriod2 += flagYear+"-08-01' and 작성일자<='"+flagYear+"-08-31'";mmStart="8";mmEnd="08";chgStartDate=flagYear+"-08-01"
    elif flagPeriod2=='9월' : workPeriod = "09월";  workPeriod2 += flagYear+"-09-01' and 작성일자<='"+flagYear+"-09-30'";mmStart="9";mmEnd="09";chgStartDate=flagYear+"-09-01"
    elif flagPeriod2=='10월' : workPeriod = "10월";  workPeriod2 += flagYear+"-10-01' and 작성일자<='"+flagYear+"-10-31'";mmStart="10";mmEnd="10";chgStartDate=flagYear+"-10-01"
    elif flagPeriod2=='11월' : workPeriod = "11월";  workPeriod2 += flagYear+"-11-01' and 작성일자<='"+flagYear+"-11-30'";mmStart="11";mmEnd="11";chgStartDate=flagYear+"-11-01"
    elif flagPeriod2=='12월' : workPeriod = "12월";  workPeriod2 += flagYear+"-12-01' and 작성일자<='"+flagYear+"-12-31'";mmStart="12";mmEnd="12";chgStartDate=flagYear+"-12-01"
  elif flagPeriod=='반기별':  
    flagPeriod2 = pyautogui.confirm('스크랩할 분기를 선택하세요',buttons=['상반기','하반기'])
    if flagPeriod2=='상반기' :   workPeriod = "1기";  workPeriod2 += flagYear+"-01-01' and 작성일자<='"+flagYear+"-06-30'";mmStart="1";mmEnd="06";chgStartDate=flagYear+"-01-01"
    elif flagPeriod2=='하반기' : workPeriod = "2기";  workPeriod2 += flagYear+"-07-01' and 작성일자<='"+flagYear+"-12-31'" ;mmStart="7";mmEnd="12";chgStartDate=flagYear+"-07-01"
  flagFirst = pyautogui.confirm(' 작업할 형태를 선택하세요',buttons=['관리자별','단일사업자','전체(기장만)','전체(신고대리포함)'])
  if  flagFirst=="전체(기장만)":
    sqladmin = "select admin_Name,Htx_ID,admin_id,manage_YN from mem_admin where manage_YN='Y' order by admin_name"
    cursor = connection.cursor()
    cursor.execute(sqladmin)
    admins = cursor.fetchall()
    connection.commit()
    connection.close()  
    for admin in admins:
      flagGroup += "'"+admin[2]+"'," 
    flagGroup = flagGroup[:-1]
  elif flagFirst=="관리자별":
    sqladmin = "select admin_Name,Htx_ID,admin_id,manage_YN from mem_admin  order by admin_name"
    cursor = connection.cursor()
    cursor.execute(sqladmin)
    admins = cursor.fetchall()
    connection.commit()
    connection.close()  
    arrAdmin = []
    for admin in admins:
      if admin[3]=="Y" or admin[3]=="S":
        arrAdmin.append(admin[2])
    flagManagerName = pyautogui.confirm('그룹작업 진행할 관리자를 선택하세요. ',buttons=arrAdmin)
    for admin in admins:
      if admin[2]==flagManagerName:  
        flagGroup += "'"+admin[2]+"'," 
  elif flagFirst=="단일사업자":
    flagSeqNo = pyautogui.prompt('인트라넷 SEQNO를 입력하세요. ')   
  if flag=='9':#부가세전자신고
    strsql = "select a.seq_no,biz_Name,replace(biz_No,'-',''),CONVERT(varchar(10), reg_date, 120),duzon_id,biz_no "
    strsql += " ,YN_1,YN_2,YN_3,YN_4,YN_5,YN_6,YN_7,YN_8,YN_9,YN_10,YN_11,YN_12,YN_13,YN_14,bigo, fiscalMM,biz_manager,ceo_name "                        #[6]->YN_1
    strsql += " from  mem_user a, mem_deal b, tbl_mng_jaroe d "
    strsql += " where a.seq_no=b.seq_no and d.seq_no=b.seq_no and a.biz_no<>'' and keeping_yn='Y' and biz_manager<>'' and biz_type in(1,2,4,5) "
    strsql += " and work_yy="+flagYear+" and work_mm='"+str(13+int(work_qt))+"' "
    # strsql += " and yn_9 in ('0','1')"   #신고서 마감/미마감은 진행 / 전자신고(2)는 진행안함 ====> 23.10.19 tbl_vat 결재완료 여부로 신고대상 선택
    if flagFirst=="관리자별":           strsql += " and biz_manager ='" + flagManagerName + "' " 
    elif flagFirst=="전체(기장만)" :    strsql += " and biz_manager  in ("+flagGroup+") " 
    elif flagFirst=="단일사업자":       strsql += " and a.seq_no ='"+flagSeqNo+"'   "  
    else:                              strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    if flagFirst=="관리자별" or flagFirst[:2]=="전체":                                    #23.10.19 단체로 할때는 예정대 법인만 / 단일사업자는 예정신고가능
      if flagPeriod2=='1분기' or flagPeriod2=='3분기':strsql += " and biz_type<=3 "       
    print(strsql)  
    cursor = connection.cursor()
    cursor.execute(strsql)
    result = cursor.fetchall()
    connection.commit()

    dig = utils.semusarang_LaunchProgram_App(semusarangID)  
    for member in result:
      trd_1 = False;trd_2=False;trd_3=False;trd_4=False;trd_10=False;trd_200=False
      kwaseKikan = "";kwaseYouhyung = ""
      directory = "D:\\"+member[1]+"\\"+flagYear+"\\부가세\\"+workPeriod.replace(" ","")
      if member[22]=='화물':directory = "D:\\화물연대\\"+member[23]+"\\"+flagYear+"\\부가세\\"+workPeriod.replace(" ","")
      strsql_f = "SELECT * FROM 부가가치세전자신고3 WHERE 사업자등록번호 = %s AND 과세기간 = %s AND 과세유형 = 'C17'"#c17은 고정임 수정말 것
      if str(work_qt) == "2":    
        with connection.cursor() as cursor:
          cursor.execute(strsql_f, (member[5], flagYear+"년 1기"))
          result_f = cursor.fetchall()
        connection.commit() 
        if result_f:mmStart="4"; mmEnd="06"
        else        :
          if workyear<=datetime.strptime(member[3], '%Y-%m-%d').year and datetime.strptime(member[3], '%Y-%m-%d').month > 4: mmStart=str(datetime.strptime(member[3], '%Y-%m-%d').month); mmEnd="06"
          else:                                                   mmStart="1"; mmEnd="06"
        kwaseKikan = flagYear+"년 1기";kwaseYouhyung = "C07"
      elif str(work_qt) == "4": 
        with connection.cursor() as cursor:
          cursor.execute(strsql_f, (member[5], flagYear+"년 2기"))
          result_f = cursor.fetchall()
        connection.commit() 
        if result_f:mmStart="10"; mmEnd="12"
        else        :
          if workyear<=datetime.strptime(member[3], '%Y-%m-%d').year and datetime.strptime(member[3], '%Y-%m-%d').month > 10: mmStart=str(datetime.strptime(member[3], '%Y-%m-%d').month); mmEnd="12"
          else:                                                   mmStart="7"; mmEnd="12"              
        kwaseKikan = flagYear+"년 2기";kwaseYouhyung = "C07"
      elif str(work_qt) == "3": kwaseKikan = flagYear+"년 2기";kwaseYouhyung = "C17"
      elif str(work_qt) == "1": kwaseKikan = flagYear+"년 1기";kwaseYouhyung = "C17"
      if member[3][:4]==flagYear:  #설립연도가 신고연도와 같은 경우
        if int(mmStart)<int(member[3][5:7]):mmStart = member[3][5:7]      
      print("시작월:"+mmStart)  

      if  os.path.exists(directory+"\\0.pdf"):
        #신고서파일이 있고(인트라넷 신고서칸에 체크) 결재완료에 체크 안된 경우 전자신고 진행, 결재완료는 진행제외시킬 때 사용
        print(f'{member[1]} 부가세 전자신고 : 0.pdf 있음')
        strsql = "select YN_8 신고서,YN_10 결재완료,YN_14 전자신고여부 from tbl_vat where seq_no='"+member[0]+"' and work_yy='"+flagYear+"' and work_qt='"+str(work_qt)+"'"
        print(strsql)
        cursor = connection.cursor()
        cursor.execute(strsql)
        result = cursor.fetchall()
        connection.commit()
        connection.close()    

          #결재완료에 체크 안돼 있으면 전자신고 되어 있어도 다시 전자신고 진행한다...
        if (result and result[0] and result[0][1]=='0') or flagFirst=="단일사업자":
          finalKi = int(flagYear)-int(member[3][:4])+1
          if member[4]=='1':  utils.semusarang_ChangeCompany(member[2])
          else:               utils.semusarang_ChangeCompany_ID_App(dig,member[4])        
          utils.semusarang_ChangeFiscalYear_App('vat',str(finalKi))    
          elecResult_Save.ElecIssue.SS_MakeElecFile('vat',member[0],member[4],mmStart,mmEnd,kwaseKikan,kwaseYouhyung,isJungKi,htxLoginID)            # 전자신고파일은 한 업체만 만들어진다
          elecIssueResult = elecResult_Save.SS_ElecIssue('vat',member[0],workyear,workPeriod.replace(" ",""),isJungKi)      
          if elecIssueResult==1:
            strMNG = "Merge  tbl_mng_jaroe as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(13+int(work_qt))+"' as work_mm) as B  "
            strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_mm=B.work_mm "
            strMNG += " when matched then update set YN_9='2'  "
            strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','0','0','0','0','0','0','0','0','2','0','0','0','0','0',''); "
            print(strMNG)
            cursor = connection.cursor()
            cursor.execute(strMNG)  
        else:
          print(f'{member[1]} : 결재완료에 체크 안돼 있어서 중단됩니다.')
      else:
        print(f'{member[1]} : 신고서(0.pdf)가 없어서 중단됩니다.')
  elif flag=='5':#8.세사 VAT신고서
    strsql = "select a.seq_no,biz_Name,replace(biz_No,'-',''),CONVERT(varchar(10), reg_date, 120),duzon_id,biz_no " #[0]~[5]
    strsql += " ,YN_1,YN_2,YN_3,YN_4,YN_5,YN_6,YN_7,YN_8,YN_9,YN_10,YN_11,YN_12,YN_13,YN_14,bigo, fiscalMM,biz_manager,ceo_name,biz_type,uptae "    #[6]->YN_1,[15]->YN_10, 22-> biz_manager
    strsql += " from  mem_user a, mem_deal b, tbl_mng_jaroe d "
    strsql += " where a.seq_no=b.seq_no and d.seq_no=b.seq_no and a.biz_no<>'' and keeping_yn='Y' and biz_manager<>''  and biz_type in(1,2,4,5)  "
    strsql += " and work_yy="+flagYear+" and work_mm='"+str(13+int(work_qt))+"' "
    # strsql += " and yn_9 in ('0','1')"   #신고서 마감/미마감은 진행 / 전자신고(2)는 진행안함 ====> 23.10.19 tbl_vat 결재완료 여부로 신고대상 선택
    if flagFirst=="관리자별":           strsql += " and biz_manager ='" + flagManagerName + "' "  
    elif flagFirst=="전체(기장만)" :    strsql += " and biz_manager  in ("+flagGroup+") " 
    elif flagFirst=="단일사업자":       strsql += " and a.seq_no ='"+flagSeqNo+"'   "  
    else:                              strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    if flagFirst=="관리자별" or flagFirst[:2]=="전체":                                    #23.10.19 단체로 할때는 예정대 법인만 / 단일사업자는 예정신고가능
      if flagPeriod2=='1분기' or flagPeriod2=='3분기':strsql += " and biz_type<=3 "    
    print(strsql)  
    cursor = connection.cursor()
    cursor.execute(strsql)
    result = cursor.fetchall()
    connection.commit()

    dig = utils.semusarang_LaunchProgram_App(semusarangID)  
    for member in result:
      trd_1 = False;trd_2=False;trd_3=False;trd_4=False;trd_10=False;trd_200=False
      directory = "D:\\"+member[1]+"\\"+flagYear+"\\부가세\\"+workPeriod.replace(" ","")
      if member[22]=='화물':directory = "D:\\화물연대\\"+member[23]+"\\"+flagYear+"\\부가세\\"+workPeriod.replace(" ","")
      kwaseKikan = "";kwaseYouhyung = ""
      strsql_f = "SELECT * FROM 부가가치세전자신고3 WHERE 사업자등록번호 = %s AND 과세기간 = %s AND 과세유형 = 'C17'"#c17은 고정임 수정말 것
      if str(work_qt) == "2":    
        with connection.cursor() as cursor:
          cursor.execute(strsql_f, (member[5], flagYear+"년 1기"))
          result_f = cursor.fetchall()
        connection.commit() 
        if result_f:mmStart="4"; mmEnd="06"
        else        :
          if workyear<=datetime.strptime(member[3], '%Y-%m-%d').year and datetime.strptime(member[3], '%Y-%m-%d').month > 4: mmStart=str(datetime.strptime(member[3], '%Y-%m-%d').month); mmEnd="6"
          else:                                                   mmStart="1"; mmEnd="06"
        kwaseKikan = flagYear+"년 1기";kwaseYouhyung = "C07"
      elif str(work_qt) == "4": 
        with connection.cursor() as cursor:
          cursor.execute(strsql_f, (member[5], flagYear+"년 2기"))
          result_f = cursor.fetchall()
        connection.commit() 
        if result_f:mmStart="10"; mmEnd="12"
        else        :
          if workyear<=datetime.strptime(member[3], '%Y-%m-%d').year and datetime.strptime(member[3], '%Y-%m-%d').month > 10: mmStart=str(datetime.strptime(member[3], '%Y-%m-%d').month); mmEnd="12"
          else:                                                   mmStart="7"; mmEnd="12"    
        kwaseKikan = flagYear+"년 2기";kwaseYouhyung = "C07"       
      elif str(work_qt) == "3": kwaseKikan = flagYear+"년 2기";kwaseYouhyung = "C17"
      elif str(work_qt) == "1": kwaseKikan = flagYear+"년 1기";kwaseYouhyung = "C17"   
      if member[24]==5:   kwaseKikan = flagYear+"년 2기";kwaseYouhyung = "C03" ;     mmStart="1"; mmEnd="12"    #간이는 1년 전체로 잡는다
      if member[3][:4]==flagYear:  #설립연도가 신고연도와 같은 경우
        if int(mmStart)<int(member[3][5:7]):mmStart = member[3][5:7]
      print("시작월:"+mmStart)  
      #신고서 0번이 없으면 일단 신고마감까지
      if not os.path.exists(directory+"\\0.pdf"):
        print('0.pdf 없음')
        finalKi = int(flagYear)-int(member[3][:4])+1
        print('회계기간 : '+member[21])
        if member[21]!='12':
          if int(today[5:6])>int(member[21]):  finalKi += 1
        if member[4]=='1':  utils.semusarang_ChangeCompany(member[2])
        else:               utils.semusarang_ChangeCompany_ID_App(dig,member[4])        
        utils.semusarang_ChangeFiscalYear_App('vat',str(finalKi))  
        # strsel = 'select YN_5,YN_6,YN_7,YN_8 from tbl_vat where seq_no='+member[0]+' and work_yy='+flagYear+' and work_qt='+str(work_qt);print(strsel)   
        # cursor = connection.cursor()   
        # cursor.execute(strsel)
        # result_vat = cursor.fetchone()
        # connection.commit()
        # YN_5 = YN_6 = YN_7 = YN_8 = 0
        # if result_vat:
        #   YN_5 = result_vat[0];YN_6 = result_vat[1];YN_7 = result_vat[2];YN_8 = result_vat[3]
        # if (int(member[6])==5 and int(member[16])==2) or (int(member[6])==2 and int(YN_5)==1) : trd_1 = True   #YN_11
        # if (int(member[6])==5 and int(member[17])==2) or (int(member[17])==2 and int(YN_5)==1) : trd_2 = True   #YN_12
        # if (int(member[6])==5 and int(member[18])==2) or (int(member[18])==2 and int(YN_5)==1) : trd_3 = True   #YN_13
        # if (int(member[6])==5 and int(member[19])==2) or (int(member[19])==2 and int(YN_5)==1)  : trd_4 = True   #YN_14

        if int(member[16])==2 : trd_1 = True   #YN_11
        if int(member[17])==2 : trd_2 = True   #YN_12
        if int(member[18])==2 : trd_3 = True   #YN_13
        if int(member[19])==2 : trd_4 = True   #YN_14
        cursor = connection.cursor()  
        if (trd_1 and not os.path.exists(directory+"\\1.pdf")) or (trd_2 and not os.path.exists(directory+"\\2.pdf")):  
          elecResult_Save.VAT.SS_TI_Summation(directory,mmStart,mmEnd,trd_1,trd_2)   
          strMNG = "Merge  tbl_vat as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(work_qt)+"' as work_qt) as B  "
          strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_qt=B.work_qt "
          strMNG += " when matched then update set YN_5='1'  "  #부가세신고관리tbl_vat : YN_5 체크
          strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(work_qt)+"','0','0','0','0','1','0','0','0','0','0','0','0','',-1,0,0,0); "
          print(strMNG)
          cursor.execute(strMNG) 
        else:
          print(f'{member[1]} : 전자세금계산서합계표 작성오류')
        if (trd_3 and not os.path.exists(directory+"\\3.pdf")) or (trd_4 and not os.path.exists(directory+"\\4.pdf")):  
          elecResult_Save.VAT.SS_TI_Summation2(directory,mmStart,mmEnd,trd_3,trd_4)  
          strMNG = "Merge  tbl_vat as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(work_qt)+"' as work_qt) as B  "
          strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_qt=B.work_qt "
          strMNG += " when matched then update set YN_5='1'  "
          strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(work_qt)+"','0','0','0','0','1','0','0','0','0','0','0','0','',-1,0,0,0); "
          print(strMNG)
          cursor.execute(strMNG)  
          if (member[25][:1]=="음" or member[25][:1]=="숙") and trd_4: elecResult_Save.VAT.SS_YueJaeGongje(directory,mmStart,mmEnd,member[0],flagYear)  
        else:
          print(f'{member[1]} : 전자계산서합계표 작성오류')     
        #[15]->YN_10                 
        if int(member[15])==2 and not os.path.exists(directory+"\\10.pdf") : elecResult_Save.VAT.SS_TI_Bullgong(directory,mmStart,mmEnd)
        #[11]->YN_6  
        if int(member[11])==2 and not os.path.exists(directory+"\\11.pdf") : elecResult_Save.VAT.SS_TI_TangibleAsset(directory,mmStart,mmEnd)    
        if int(member[7])==2 or int(member[8])==2 : 
          if  not os.path.exists(directory+"\\5.pdf"):
            elecResult_Save.VAT.SS_SaleCard_Summation(directory,mmStart,mmEnd) 
            strMNG = "Merge  tbl_vat as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(work_qt)+"' as work_qt) as B  "
            strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_qt=B.work_qt "
            strMNG += " when matched then update set YN_6='1'  "
            strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(work_qt)+"','0','0','0','0','0','1','0','0','0','0','0','0','',-1,0,0,0); "
            print(strMNG)
            cursor.execute(strMNG)              
        if int(member[12])==2 or int(member[13])==2 : 
          if  not os.path.exists(directory+"\\6.pdf"):
            elecResult_Save.VAT.SS_Card_Summation(directory,mmStart,mmEnd)             
            strMNG = "Merge  tbl_vat as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(work_qt)+"' as work_qt) as B  "
            strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_qt=B.work_qt "
            strMNG += " when matched then update set YN_7='1'  "
            strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(work_qt)+"','0','0','0','0','0','0','1','0','0','0','0','0','',-1,0,0,0); "
            print(strMNG)
            cursor.execute(strMNG)          
        vatResult = elecResult_Save.SS_VatSingoMagam(directory,mmStart,mmEnd,member[0],isJungKi,flagYear + "년 " + workPeriod[:2]  ,workPeriod[-2:]);print(f'부가세 신고서 저장 결과 : {vatResult}') 
        if vatResult==1:
          strMNG = "Merge  tbl_vat as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(work_qt)+"' as work_qt) as B  "
          strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_qt=B.work_qt "
          strMNG += " when matched then update set YN_8='1'  "
          strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(work_qt)+"','0','0','0','0','0','0','0','1','0','0','0','0','',-1,0,0,0); "
          print(f'tbl_vat -YN_8:1 업데이트 {strMNG}')
          cursor.execute(strMNG)          
          strMNG = "Merge  tbl_mng_jaroe as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(13+int(work_qt))+"' as work_mm) as B  "
          strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_mm=B.work_mm "
          strMNG += " when matched then update set YN_9='1'  "
          strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','0','0','0','0','0','0','0','0','1','0','0','0','0','0',''); "
          print(f'tbl_mng_jaroe -YN_9:1 업데이트 {strMNG}')
          cursor.execute(strMNG)  
          ############################################################################
          #   0.pdf가 없는 경우는 신고서까지만 만든다 0.pdf가 있으면 전자신고가지 한다 => 일단 신고까지...
          ############################################################################
          elecResult_Save.ElecIssue.SS_MakeElecFile('vat',member[0],member[4],mmStart,mmEnd,kwaseKikan,kwaseYouhyung,isJungKi,htxLoginID)
          elecIssueResult = elecResult_Save.SS_ElecIssue('vat',member[0],workyear,workPeriod.replace(" ",""),isJungKi)      
          if elecIssueResult==1:
            strMNG = "Merge  tbl_mng_jaroe as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(13+int(work_qt))+"' as work_mm) as B  "
            strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_mm=B.work_mm "
            strMNG += " when matched then update set YN_9='2'  "
            strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','0','0','0','0','0','0','0','0','2','0','0','0','0','0',''); "
            print(strMNG)
            cursor = connection.cursor()
            cursor.execute(strMNG)              
        else:         
          print(f'{member[1]} : 부가세신고서 작성오류')
      else:
          ############################################################################
          #   0.pdf가 있으는 경우로 전자신고가지 한다
          ############################################################################         
        if flagFirst=="단일사업자":
          elecResult_Save.ElecIssue.SS_MakeElecFile('vat',member[0],member[4],mmStart,mmEnd,kwaseKikan,kwaseYouhyung,isJungKi,htxLoginID)
          elecIssueResult = elecResult_Save.SS_ElecIssue('vat',member[0],workyear,workPeriod.replace(" ",""),isJungKi)      
          if elecIssueResult==1:
            strMNG = "Merge  tbl_mng_jaroe as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(13+int(work_qt))+"' as work_mm) as B  "
            strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_mm=B.work_mm "
            strMNG += " when matched then update set YN_9='2'  "
            strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','0','0','0','0','0','0','0','0','2','0','0','0','0','0',''); "
            print(strMNG)
            cursor = connection.cursor()
            cursor.execute(strMNG)    
        else:
          print('단일사업자가 아닌 경우 : 부가세 신고서(0.pdf)가 있어서 중단합니다')
  elif flag=='4':#7.세사 신/현(매출/입)업로드
    strsql = " select a.seq_no,biz_Name,replace(biz_No,'-',''),CONVERT(varchar(10), reg_date, 120),duzon_id,biz_no,YN_2,YN_3,YN_4,YN_7,YN_8,biz_type from  mem_user a, mem_deal b, tbl_mng_jaroe d "
    # strsql = "select a.seq_no,biz_Name,replace(biz_No,'-',''),CONVERT(varchar(10), reg_date, 120),duzon_id,biz_no "
    # strsql += " ,YN_1,YN_2,YN_3,YN_4,YN_5,YN_6,YN_7,YN_8,YN_9,YN_10,YN_11,YN_12,YN_13,YN_14,bigo, fiscalMM,biz_manager,ceo_name,biz_type "    #[6]->YN_1,[15]->YN_10, 22-> biz_manager    
    # strsql += " from  mem_user a, mem_deal b, tbl_mng_jaroe d "
    strsql += " where work_yy="+flagYear+" and work_mm='"+str(13+int(work_qt))+"' and (yn_2='1' or yn_3='1'or yn_4='1' or yn_7='1' or yn_8='1') "
    strsql += " and a.seq_no=b.seq_no and d.seq_no=b.seq_no and a.biz_no<>'' and keeping_yn='Y' and biz_manager<>'' "
    strsql += " and a.seq_no not in (2209,4164,3432,3434)   "           #다온기업 , 스톤글로벌, 혜솔, 루키인터내셔널은 자체기장
    #if work_qt==1 or work_qt==3:strsql += " and biz_type<=3 "
    if flagFirst=="관리자별":           strsql += " and biz_manager ='" + flagManagerName + "' "  
    elif flagFirst=="전체(기장만)" :    strsql += " and biz_manager  in ("+flagGroup+") " 
    elif flagFirst=="단일사업자":       strsql += " and a.seq_no ='"+flagSeqNo+"'   "  
    else:                              strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    if flagFirst=="관리자별" or flagFirst[:2]=="전체":                                    #23.10.19 단체로 할때는 예정대 법인만 / 단일사업자는 예정신고가능
      if flagPeriod2=='1분기' or flagPeriod2=='3분기':strsql += " and biz_type<=3 "   
    print(strsql)  
    cursor = connection.cursor()
    cursor.execute(strsql)
    result = cursor.fetchall()
    connection.commit()
    for member in result:
      totalPath = "C:\\Users\\Administrator\\Documents\\"
      
      #작성일자 확정
      workPeriod3 = ""
      strsql_f = "SELECT * FROM 부가가치세전자신고3 WHERE 사업자등록번호 = %s AND 과세기간 = %s AND 과세유형 = 'C17'"#c17은 고정임 수정말 것
      if  str(work_qt) == "1":  
        workPeriod = " Tran_MM>='"+flagYear+"01' and Tran_MM<='"+flagYear+"03'"
        workPeriod2 = " AprvDt>='"+flagYear+"-01-01' and AprvDt<='"+flagYear+"-03-31'"
        workPeriod3 = " left(tran_dt,10)>='"+flagYear+"-01-01' and left(tran_dt,10)<='"+flagYear+"-03-31'"
      elif  str(work_qt) == "3":
        mmStart="07";
        if member[11]==5:   mmStart="01";          
        workPeriod = " Tran_MM>='"+flagYear+mmStart+"' and Tran_MM<='"+flagYear+"09'"
        workPeriod2 = " AprvDt>='"+flagYear+"-"+mmStart+"-01' and AprvDt<='"+flagYear+"-09-30'"
        workPeriod3 = " left(tran_dt,10)>='"+flagYear+"-07-01' and left(tran_dt,10)<='"+flagYear+"-09-30'"
      elif  str(work_qt) == "2":    
        with connection.cursor() as cursor:
          cursor.execute(strsql_f, (member[5],flagYear+"년 1기"))
          result_f = cursor.fetchall()
          connection.commit() 
          if result_f:  
            mmStart="04";
            if member[11]==5:   mmStart="01";                      
            workPeriod = " Tran_MM>='"+flagYear+mmStart+"' and Tran_MM<='"+flagYear+"06'"
            workPeriod2 = " AprvDt>='"+flagYear+"-"+mmStart+"-01' and AprvDt<='"+flagYear+"-06-30'";mmEnd="06";chgStartDate=flagYear+"-"+mmStart+"-01"
            workPeriod3 = " left(tran_dt,10)>='"+flagYear+"-"+mmStart+"-01' and left(tran_dt,10)<='"+flagYear+"-06-30'"
          else        : 
            mmStart="01";
            if member[11]==5:   mmStart="01";    
            workPeriod = " Tran_MM>='"+flagYear+mmStart+"' and Tran_MM<='"+flagYear+"06'"
            workPeriod2 = " AprvDt>='"+flagYear+"-"+mmStart+"-01' and AprvDt<='"+flagYear+"-06-30'";chgStartDate=flagYear+"-"+mmStart+"-01" 
            workPeriod3 = " left(tran_dt,10)>='"+flagYear+"-"+mmStart+"-01' and left(tran_dt,10)<='"+flagYear+"-06-30'"  
      elif str(work_qt)=="4":
        with connection.cursor() as cursor:
          cursor.execute(strsql_f, (member[5],flagYear+"년 2기"))
          result_f = cursor.fetchall()
          connection.commit() 
          if result_f:  
            mmStart="10";
            if member[11]==5:   mmStart="01";                
            workPeriod = " Tran_MM>='"+flagYear+mmStart+"' and Tran_MM<='"+flagYear+"12'"
            workPeriod2 = " AprvDt>='"+flagYear+"-"+mmStart+"-01' and AprvDt<='"+flagYear+"-12-31'";mmStart="10";mmEnd="12";chgStartDate=flagYear+"-"+mmStart+"-01"
            workPeriod3 = " left(tran_dt,10)>='"+flagYear+"-"+mmStart+"-01' and left(tran_dt,10)<='"+flagYear+"-12-31'"
          else        : 
            mmStart="07";
            if member[11]==5:   mmStart="01";  
            workPeriod = " Tran_MM>='"+flagYear+mmStart+"' and Tran_MM<='"+flagYear+"12'"
            workPeriod2 = " AprvDt>='"+flagYear+"-"+mmStart+"-01' and AprvDt<='"+flagYear+"-12-31'";chgStartDate=flagYear+"-"+mmStart+"-01"  
            workPeriod3 = " left(tran_dt,10)>='"+flagYear+"-"+mmStart+"-01' and left(tran_dt,10)<='"+flagYear+"-12-31'"
      fileName = member[1]
      dig = utils.semusarang_LaunchProgram_App(semusarangID)
      if member[4]=='1':  utils.semusarang_ChangeCompany(member[2])
      else:              utils.semusarang_ChangeCompany_ID_App(dig,member[4])      
      utils.semusarang_ChangeFiscalYear_App('vat',str(workyear-int(member[3][:4])+1))        
      for k in [6,7,8,9,10]:   #6번부터  YN_2(신카매출),YN_3(현영매출),YN_4(판매대행),YN_7(카드매입),YN_8(현영매입)
        if int(member[k])==1:
          wb = Workbook()  # 새 워크북 만들기
          ws = wb.active   # 워크북의 첫 번째 워크시트 가져오기
          MIDTITLE = []
          if k==6 or k==8: 
            fileName = member[1]+"_카드매출_" + str(work_qt) + "기.xlsx" 
            ws['A1'] = '신용카드 매출 자료 입력'
            ws['A4'] = member[1] 
            ws['C4'] = member[2]   
            ws['F1'] = '1.4'    
            MIDTITLE = ["년도월일","유형","신용카드사명","거래처명","사업자등록번호","품목","수량","단가","공급가액","세액","봉사료","합계금액","대표자", "업태","종목","사업장주소","부서/사원코드","현장코드","PJT코드"]    
          elif k==7: 
            fileName = member[1]+"_현영매출_" + str(work_qt) + "기.xlsx"      
            ws['A1'] = '매출 자료입력'
            ws['F1'] = '1.4'
            ws['A4'] = member[1] 
            ws['D4'] = member[2]      
            MIDTITLE = ["년도월일","유형","거래처명","사업자등록번호",        "신용카드사명",     "품목","수량","단가","공급가액","세액","봉사료","합계금액","기본계정코드",      "상대계정코드","영수/청구","대표자","업태","종목","사업장주소","비고","전자","부서/사원코드","현장코드","PJT코드","불공제사유"]            
          elif k==9: 
            fileName = member[1]+"_카드매입_" + str(work_qt) + "기.xlsx"   
            ws['A1'] = '신용카드 매입 자료 입력'
            ws['C4'] = member[2] 
            MIDTITLE = ["카드종류","신용카드사명","신용카드번호","승인일자","사업자등록번호","거래처명","거래처유형","공급가액","세액","봉사료","합계금액","부가세공제여부","부가세유형","계정과목","품목","수량","단가","대표자","업태","종목","사업장주소","부서/사원코드","현장코드","프로젝트코드"]
          elif k==10: 
            fileName = member[1]+"_현영매입_" + str(work_qt) + "기.xlsx"      
            ws['A1'] = '매입 자료입력'
            ws['F1'] = '1.4'
            ws['A4'] = member[1] 
            ws['D4'] = member[2]      
            MIDTITLE = ["년도월일","유형","거래처명","사업자등록번호","카드종류","신용카드사명","신용카드번호","부가세공제여부","품목","수량","단가","공급가액","세액","봉사료","합계금액","기본계정코드","의제, 재활용","상대계정코드","영수/청구","대표자","업태","종목","사업장주소","비고","전자","부서/사원코드","현장코드","PJT코드","불공제사유"]

          for i, value in enumerate(MIDTITLE):
            ws.cell(row=8, column=i+1, value=value)   
          strsql = ""
          if k==6 or k==8: #카드매출 
            strsql = "select Tran_MM+'25' as 년도월일,  '17' AS 유형,"
            strsql += " SaleGubun as 신용카드사명,'' as 상호,'' as 사업자등록번호,SaleGubun as 품목,'' as 수량,'' as 단가,round(Tot_StlAmt/1.1,0) as 공급가액,(Tot_StlAmt - round(Tot_StlAmt/1.1,0)) as 세액,0,Tot_StlAmt as 합계금액"
            strsql += " FROM Tbl_HomeTax_SaleCard where seq_no=" +member[0]+ " and tran_yy='"+flagYear+"' and saleGubun<>'현금영수증' and " +  workPeriod
            # 24.1.1부터 판매대행 매출은 신용카드발행세액공제 대상으로 산입한다. 부가법46조, 시행령 88조4항의4
            #strsql += " FROM Tbl_HomeTax_SaleCard where seq_no=" +member[0]+ " and tran_yy='"+flagYear+"' and saleGubun<>'신용카드 자료' and " +  workPeriod
            # strsql += " union all"
            # strsql += " select Tran_MM+'25' as 년도월일,  '14' AS 유형,"
            # strsql += " SaleGubun as 신용카드사명,'' as 상호,'' as 사업자등록번호,SaleGubun as 품목,'' as 수량,'' as 단가,round(Tot_StlAmt/1.1,0) as 공급가액,(Tot_StlAmt - round(Tot_StlAmt/1.1,0)) as 세액,0,Tot_StlAmt as 합계금액"
            # strsql += " FROM Tbl_HomeTax_SaleCard where seq_no=" +member[0]+ " and tran_yy='"+flagYear+"' and saleGubun not in ('현금영수증','신용카드 자료') and " +  workPeriod            
          elif k==7:  #현영매출
            strsql = "select Tran_MM+'25' as 년도월일,  '22' AS 유형,'현금영수증' as 거래처명,'000-00-00000','','현금영수증' as 품목,'','',"
            strsql += "round(Tot_StlAmt/1.1,0) as 공급가액,(Tot_StlAmt - round(Tot_StlAmt/1.1,0)) as 세액,0,Tot_StlAmt as 합계금액,"
            strsql += " isnull((select top 1 acnt_cd from ds_slipledgr2 where left(acnt_cd,1)=4 and seq_no="+member[0]+" and work_yy="+str(int(flagYear)-1)+" GROUP BY acnt_cd ),'401') as Acnt_Cd,'120'"
            strsql += " FROM Tbl_HomeTax_SaleCard where seq_no=" +member[0]+ " and tran_yy='"+flagYear+"' and saleGubun='현금영수증' and " +  workPeriod
          elif k==9:  #카드매입
            strsql = "SELECT '3',CrcmClNm,busnCrdCardNoEncCntn,AprvDt,mrntTxprDscmNoEncCntn,mrntTxprNm,vatDdcClNm,splCft,vaTxamt,tip,totaTrsAmt,File_DdctGB , case File_DdctGB when '공제' then '57' else '' end as VatChkTY,Acnt_Cd,mrntTxprNm "
            strsql += " FROM tbl_hometax_scrap where seq_no=" +member[0]+ " and tran_yy='"+flagYear+"' and " +  workPeriod2
          elif k==10:  #현영매입
            strsql = "SELECT left(tran_dt,10),"
            strsql +="CASE WHEN File_DdctGB='공제' THEN '61'        else '62' END AS 유형,mrntTxprNm,mrntTxprDscmNoEncCntn,'','','',"
            strsql +="CASE WHEN File_DdctGB='공제' THEN '공제' else '' END AS 공제여부,'','','',"
            strsql +="CASE WHEN File_DdctGB='공제' THEN splCft      else totaTrsAmt END AS 공급가액,"
            strsql +="CASE WHEN File_DdctGB='공제' THEN vaTxamt else 0 END AS 세액,tip, totaTrsAmt,Acnt_Cd,'','138'"
            strsql += " FROM tbl_hometax_cashcost where seq_no=" +member[0]+ "  and " +  workPeriod3            
          print(strsql)
          cursor = connection.cursor()
          cursor.execute(strsql)
          results = cursor.fetchall()
          connection.commit()
          if results:
            for i, row in enumerate(results):
              for j, value in enumerate(row):
                if k==8 and j==3 and value<chgStartDate:value=chgStartDate
                ws.cell(row=i+10, column=j+1, value=value)
            wholeName = totalPath + fileName
            if os.path.exists(wholeName):        os.remove(wholeName)
            wb.save(wholeName)
            # openpyxl로 읽어온 xlsx 데이터를 xlwt로 생성한 xls 파일에 쓰기
            wb_xls = xlwt.Workbook()
            for sheet_name in wb.sheetnames:
              ws_xlsx = wb[sheet_name]
              ws_xls = wb_xls.add_sheet(sheet_name)
              for row in ws_xlsx.rows:
                for cell in row:
                  ws_xls.write(cell.row-1, cell.col_idx-1, cell.value)          
            wb_xls.save(wholeName.split('.')[0]+'.xls')      
            os.remove(wholeName)          
            tmpNum = 0
            strMNG = "Merge  tbl_mng_jaroe as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(13+int(work_qt))+"' as work_mm) as B  "
            strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_mm=B.work_mm "            
            if  k==6 or k==8 : 
              tmpNum = elecResult_Save.VAT.SS_excelupload_Card(fileName.split('.')[0]+'.xls','매출')
              if tmpNum==1 :
                if   k==8:  
                  strMNG += " when matched then update set YN_2='2'  "
                  strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','0','2','0','0','0','0','0','0','0','0','0','0','0','0',''); "  
                elif k==8:  
                  strMNG += " when matched then update set YN_4='2'  "
                  strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','0','0','0','2','0','0','0','0','0','0','0','0','0','0',''); "              
            elif k==7 :
              tmpNum = elecResult_Save.VAT.SS_excelupload_Cash(fileName.split('.')[0]+'.xls','매출')
              if tmpNum==1 :
                strMNG += " when matched then update set YN_3='2'  "
                strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','0','0','2','0','0','0','0','0','0','0','0','0','0','0',''); "                              
            elif  k==9 : 
              tmpNum = elecResult_Save.VAT.SS_excelupload_Card(fileName.split('.')[0]+'.xls','매입')
              if tmpNum==1:
                strMNG += " when matched then update set YN_7='2'  "
                strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','0','0','0','0','0','0','2','0','0','0','0','0','0','0',''); "              
            elif  k==10 : 
              tmpNum = elecResult_Save.VAT.SS_excelupload_Cash(fileName.split('.')[0]+'.xls','매입')
              if tmpNum==1:
                strMNG += " when matched then update set YN_8='2'  "
                strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','0','0','0','0','0','0','0','2','0','0','0','0','0','0',''); "
            print(strMNG)
            if tmpNum == 1:cursor.execute(strMNG)   
            os.remove(wholeName.split('.')[0]+'.xls')  
  elif flag=='3':#6.세사 TI업로드
    strsql = "select a.seq_no,biz_Name,replace(biz_No,'-',''),CONVERT(varchar(10), reg_date, 120),duzon_id,biz_no "
    strsql += " ,YN_1,YN_2,YN_3,YN_4,YN_5,YN_6,YN_7,YN_8,YN_9,YN_10,YN_11,YN_12,YN_13,YN_14,bigo, fiscalMM,biz_manager,ceo_name,biz_type "    #[6]->YN_1,[15]->YN_10, 22-> biz_manager    
    strsql += " from  mem_user a, mem_deal b, tbl_mng_jaroe d "      
    strsql += " where a.seq_no=b.seq_no and d.seq_no=b.seq_no and a.biz_no<>'' and keeping_yn='Y' and  biz_manager<>'' and biz_type in(1,2,4,5)"
    strsql += " and a.seq_no not in (2209,4164,3432,3434)   "           #다온기업 , 스톤글로벌, 혜솔, 루키인터내셔널은 자체기장
    #if work_qt==1 or work_qt==3:strsql += " and biz_type<=3 "             #신고서 작성시만 예정에 법인사업자만 작성
    strsql += " and work_yy="+flagYear+" and work_mm='"+str(13+int(work_qt))+"' "

    if flagFirst=="관리자별":           strsql += " and biz_manager ='" + flagManagerName + "' "  
    elif flagFirst=="전체(기장만)" :    strsql += " and biz_manager  in ("+flagGroup+") " 
    elif flagFirst=="단일사업자":       strsql += " and a.seq_no ='"+flagSeqNo+"'   "  
    else:                              strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') and yn_1='3'" 
    if flagFirst=="관리자별" or flagFirst[:2]=="전체":                                    #23.10.19 단체로 할때는 예정대 법인만 / 단일사업자는 예정신고가능
      if flagPeriod2=='1분기' or flagPeriod2=='3분기':strsql += " and biz_type<=3 "         
    print(strsql)  
    cursor = connection.cursor()
    cursor.execute(strsql)
    result = cursor.fetchall()
    for member in result:
      trd_1 = False;trd_2=False;trd_3=False;trd_4=False;trd_10=False;trd_200=False
      totalPath = "C:\\Users\\Administrator\\Documents\\"
      fileNameSale = member[1]+"_기타매출전표전송_" + str(work_qt) + "기.xlsx"
      fileNameCost = member[1]+"_기타매입전표전송_" + str(work_qt) + "기.xlsx"
      #작성일자 확정
      strsql_f = "SELECT * FROM 부가가치세전자신고3 WHERE 사업자등록번호 = %s AND 과세기간 = %s AND 과세유형 = 'C17'"#c17은 고정임 수정말 것
      if  str(work_qt) == "2":    
        with connection.cursor() as cursor:
          cursor.execute(strsql_f, (member[5],flagYear+"년 1기"))
          result_f = cursor.fetchall()
        connection.commit() 
        if result_f:  workPeriod2 = " 작성일자>='"+flagYear+"-04-01' and 작성일자<='"+flagYear+"-06-30'";mmStart="4";mmEnd="06";chgStartDate=flagYear+"-04-01"
        else        : 
          workPeriod2 = " 작성일자>='"+flagYear+"-01-01' and 작성일자<='"+flagYear+"-06-30'";mmStart="1";mmEnd="06";chgStartDate=flagYear+"-01-01"   
          if workyear<=datetime.strptime(member[3], '%Y-%m-%d').year and datetime.strptime(member[3], '%Y-%m-%d').month > 4: mmStart=str(datetime.strptime(member[3], '%Y-%m-%d').month); mmEnd="06"
      elif str(work_qt)=="4":
        with connection.cursor() as cursor:
          cursor.execute(strsql_f, (member[5],flagYear+"년 2기"))
          result_f = cursor.fetchall()
        connection.commit() 
        if result_f:  workPeriod2 = " 작성일자>='"+flagYear+"-10-01' and 작성일자<='"+flagYear+"-12-31'";mmStart="10";mmEnd="12";chgStartDate=flagYear+"-10-01"
        else        : 
          workPeriod2 = " 작성일자>='"+flagYear+"-07-01' and 작성일자<='"+flagYear+"-12-31'";mmStart= "7";mmEnd="12";chgStartDate=flagYear+"-07-01"         
          if workyear<=datetime.strptime(member[3], '%Y-%m-%d').year and datetime.strptime(member[3], '%Y-%m-%d').month > 10: mmStart=str(datetime.strptime(member[3], '%Y-%m-%d').month); mmEnd="12"
      if member[24]==5:   kwaseKikan = flagYear+"년 2기";kwaseYouhyung = "C03" ;     mmStart="1"; mmEnd="12"    #간이는 1년 전체로 잡는다
      if member[3][:4]==flagYear:  #설립연도가 신고연도와 같은 경우
        if int(mmStart)<int(member[3][5:7]):mmStart = member[3][5:7]

      # strsel = f"select isnull(YN_5,0) as 세금계산서합계표,isnull(YN_6,0) as 카드매출합계표,isnull(YN_7,0) as 카드매입합계표,isnull(YN_8,0) as 부가세신고서 from tbl_vat where seq_no='{member[0]}' and work_yy='{flagYear}' and work_qt='{work_qt}'"    
      # cursor = connection.cursor()
      # cursor.execute(strsel)
      # result_vat = cursor.fetchone()
      # connection.commit()
      # print(f"전자세금계산서 업로드 상태 : {member[16]}")
      # print(f"인트라넷 체크여부 : {result_vat[0]}")
      if (int(member[16]) + int(member[17]) + int(member[18]) + int(member[19]))>0:# tbl_mng_jaroe : member[16](세금계산서) 1은 조회여부 2는 데이터 유무 
        dig = utils.semusarang_LaunchProgram_App(semusarangID)
        if member[4]=='1':  utils.semusarang_ChangeCompany(member[2])
        else:               utils.semusarang_ChangeCompany_ID_App(dig,member[4])       
        finalKi = int(flagYear)-int(member[3][:4])+1
        utils.semusarang_ChangeFiscalYear_App('vat',str(finalKi))         
        MIDTITLE = []
        for GB in ['매출','매입']:
          wb = Workbook()  # 새 워크북 만들기
          ws = wb.active   # 워크북의 첫 번째 워크시트 가져오기
          ws['A1'] = GB + ' 자료입력'
          ws['D4'] = member[2]   
          ws['F1'] = '1.4'     
          if GB=='매출':  MIDTITLE = ["년도월일","유형","거래처명","사업자등록번호","신용카드사명"                                         ,"품목","수량","단가","공급가액","세액","봉사료","합계금액","기본계정코드",              "상대계정코드","영수/청구","대표자","업태","종목","사업장주소","비고","전자","부서/사원코드","현장코드","PJT코드","영세율구분"]
          elif GB=='매입':MIDTITLE = ["년도월일","유형","거래처명","사업자등록번호","카드종류","신용카드사명","신용카드번호","부가세공제여부","품목","수량","단가","공급가액","세액","봉사료","합계금액","기본계정코드","의제, 재활용","상대계정코드","영수/청구","대표자","업태","종목","사업장주소","비고","전자","부서/사원코드","현장코드","PJT코드","불공제사유"]
          for i, value in enumerate(MIDTITLE):
            ws.cell(row=8, column=i+1, value=value) 
          # 타이틀 데이터 입력하기
          strsql = "select replace(작성일자,'-','') as 연도월일, "
          if GB=='매출':  
                          strsql += " CASE "
                          strsql += "   WHEN 매입매출구분='1' AND 전자세금계산서분류 IN ('일반','일반(수정)','위수탁','위수탁(수정)') THEN '11' "
                          strsql += "   WHEN 매입매출구분='1' AND 전자세금계산서분류 IN ('영세율','영세율(수정)') THEN '12' "
                          strsql += "   WHEN 매입매출구분='3' THEN '13' "
                          strsql += " END AS 유형,"
          elif GB=='매입':
                          strsql += " CASE "
                          strsql += "   WHEN 매입매출구분 = 2 AND 전자세금계산서분류 IN ('일반','일반(수정)', '수입','수입(수정)','위수탁','위수탁(수정)') THEN '51' " 
                          strsql += "   WHEN 매입매출구분 = 4 AND 전자세금계산서분류 IN ('일반','일반(수정)', '수입','수입(수정)','위수탁','위수탁(수정)') THEN '53'"
                          strsql += "   WHEN 매입매출구분 = 2 AND 전자세금계산서분류 IN ('영세율','영세율(수정)') THEN '52' "
                          strsql += " END AS 유형,"
          if GB=='매출':  strsql += " case when 공급받는자상호='' then 공급받는자대표자명 else 공급받는자상호 end as 상호,left(replace(공급받는자사업자등록번호,'-',''),10),'' as 신용카드사명,"
          elif GB=='매입':strsql += " case when 공급자상호='' then 공급자대표자명 else 공급자상호 end as 상호,left(replace(공급자사업자등록번호,'-',''),10),'','','',DdctGB,"
          strsql += " 품목명,품목수량,품목단가,공급가액,세액,0 as 봉사료,합계금액,Slip_Acnt_Cd,"
          if GB=='매출':  strsql += " '108',"
          elif GB=='매입':strsql += " case when 적요코드='12' then '6' when 적요코드='13' then '7' else '' end as 의제재활용, '251',"
          strsql += "영수청구구분,"
          if GB=='매출':  strsql += " 공급받는자대표자명,"
          elif GB=='매입':strsql += " 공급자대표자명,"        
          strsql += "'','','','','1' as 전자,'','','',"
          if GB=='매출':  strsql += "   CASE when 전자세금계산서분류='영세율' then '1' else '' end as 영세율구분"
          elif GB=='매입':strsql += "   CASE when 적요코드 in ('12','13') then '' else 적요코드 end as 불공제사유"            
          strsql += " FROM 전자세금계산서 WHERE SEQ_NO="+member[0] + " AND " + workPeriod2
          if GB=='매출':  strsql += " AND 매입매출구분 IN ('1','3')"
          elif GB=='매입':strsql += " AND 매입매출구분 IN ('2','4')"   
          strsql += " AND not 합계금액=0" 
          print(strsql)
          cursor = connection.cursor()
          cursor.execute(strsql)
          results = cursor.fetchall()
          if results:            
            for i, row in enumerate(results):
              for j, value in enumerate(row):
                ws.cell(row=i+10, column=j+1, value=value)
                if j==1:print(value)
                if j==1 and value=='11': trd_1=True
                if j==1 and value=='13': trd_3=True
                if j==1 and value=='51': trd_2=True
                if j==1 and value=='54': trd_10=True
                if j==1 and value=='53': trd_4=True 
                if GB=='매입' and j==15 and value[:1]=='2':trd_200=True
            wholeName = "";finalFileName = ""
            if GB=='매출':     wholeName = totalPath + fileNameSale;finalFileName = fileNameSale
            elif GB=='매입':   wholeName = totalPath + fileNameCost;finalFileName = fileNameCost   
            if os.path.exists(wholeName):        os.remove(wholeName)
            wb.save(wholeName)
            # openpyxl로 읽어온 xlsx 데이터를 xlwt로 생성한 xls 파일에 쓰기
            wb_xls = xlwt.Workbook()
            for sheet_name in wb.sheetnames:
              ws_xlsx = wb[sheet_name]
              ws_xls = wb_xls.add_sheet(sheet_name)
              for row in ws_xlsx.rows:
                for cell in row:
                  ws_xls.write(cell.row-1, cell.col_idx-1, cell.value)          
            wb_xls.save(wholeName.split('.')[0]+'.xls')      
            os.remove(wholeName)
            tmpNum = elecResult_Save.VAT.SS_excelupload_TI(finalFileName.split('.')[0]+'.xls',member[0],GB)
            if tmpNum==1:
              strMNG = "Merge  tbl_mng_jaroe as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(13+int(work_qt))+"' as work_mm) as B  "
              strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_mm=B.work_mm "
              strMNG += " when matched then update set YN_1='5'  "
              strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','5','0','0','0','0','0','0','0','0','0','0','0','0','0',''); "
              print(strMNG)
              cursor.execute(strMNG)   
              os.remove(wholeName.split('.')[0]+'.xls')       
          connection.commit()
        if trd_1: strUdt = "update tbl_mng_jaroe set YN_11='2' where seq_no='"+member[0]+"' and work_yy="+flagYear+" and work_mm="+str(13+int(work_qt));print(strUdt);cursor.execute(strUdt) 
        if trd_2: strUdt = "update tbl_mng_jaroe set YN_12='2' where seq_no='"+member[0]+"' and work_yy="+flagYear+" and work_mm="+str(13+int(work_qt));print(strUdt);cursor.execute(strUdt) 
        if trd_3: strUdt = "update tbl_mng_jaroe set YN_13='2' where seq_no='"+member[0]+"' and work_yy="+flagYear+" and work_mm="+str(13+int(work_qt));print(strUdt);cursor.execute(strUdt) 
        if trd_4: strUdt = "update tbl_mng_jaroe set YN_14='2' where seq_no='"+member[0]+"' and work_yy="+flagYear+" and work_mm="+str(13+int(work_qt));print(strUdt);cursor.execute(strUdt) 
        if trd_10:strUdt = "update tbl_mng_jaroe set YN_10='2' where seq_no='"+member[0]+"' and work_yy="+flagYear+" and work_mm="+str(13+int(work_qt));print(strUdt);cursor.execute(strUdt) 
        if trd_200:strUdt = "update tbl_mng_jaroe set YN_6='2' where seq_no='"+member[0]+"' and work_yy="+flagYear+" and work_mm="+str(13+int(work_qt));print(strUdt);cursor.execute(strUdt) 
        connection.close()  
  elif flag=='8':#5.부가세통합조회
    singoGB = workPeriod[-2:]
    workGi = workPeriod[:2]
    kawasekisu = flagYear + "년 " + workGi
    strsql = " select a.seq_no ,biz_name,biz_manager,biz_no,biz_type,CONVERT(varchar(10), reg_date, 120) "
    strsql += " from mem_user a, mem_deal b where a.seq_no=b.seq_no and a.biz_no<>'' and keeping_yn='Y' and biz_manager<>''  "
    if flagFirst=="관리자별":           strsql += " and biz_manager ='" + flagManagerName + "' "  
    elif flagFirst=="전체(기장만)" :    strsql += " and biz_manager  in ("+flagGroup+") " 
    elif flagFirst=="단일사업자":       strsql += " and a.seq_no ='"+flagSeqNo+"'   "  
    else:                              strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    if flagFirst=="관리자별" or flagFirst[:2]=="전체":                                    #23.10.19 단체로 할때는 예정대 법인만 / 단일사업자는 예정신고가능
      if flagPeriod2=='1분기' or flagPeriod2=='3분기':strsql += " and biz_type<=3 "   
    strsql += " order by a.seq_no/1"  
    print(strsql)
    cursor = connection.cursor()
    cursor.execute(strsql)
    members = cursor.fetchall()
    connection.commit()
    if members:
      driver = utils.conHometaxLogin(htxLoginID,False) ;time.sleep(2)
      driver.find_element(By.ID,'mf_wfHeader_hdGrp920').click();print('세무대리')
      driver.find_element(By.ID,'span_a_4801060000').click();time.sleep(0.5);print('부가가치세 신고자료 통합조회')      
      for result in members:
        if len(result[3])>=10 and (result[4]==1 or result[4]==4 or result[4]==5):
          workPer = '0'
          # 일단 분기별로 조회하고 보여준다
          if str(work_qt) == "2" or str(work_qt) == "4" :  workPer = '1'  ;singoGB = '확정'
          selectKi = Select(driver.find_element(By.ID,'mf_txppWframe_selectHt'));       selectKi.select_by_visible_text(workGi);time.sleep(0.5)#1기/2기 선택
          driver.find_element(By.ID,'mf_txppWframe_radioRtnClCd_input_'+workPer).click();#0:예정/1:확정 
          bizNo = result[3].replace("-","")
          driver.find_element(By.ID,'mf_txppWframe_inputBsno').clear();time.sleep(0.5)
          driver.find_element(By.ID,'mf_txppWframe_inputBsno').send_keys(bizNo);time.sleep(0.25);  
          driver.find_element(By.ID, 'mf_txppWframe_trigger113').click();time.sleep(3)
          try:
              WebDriverWait(driver, 1).until(EC.alert_is_present()); al = Alert(driver); al.accept();time.sleep(0.25)
              continue  # 알람 창이 발생하면 다음 루프로 진행
          except:
              pass
          print(driver.find_element(By.ID,'mf_txppWframe_edtEtxivSlsAmt').get_attribute("value"))
          sangho = driver.find_element(By.ID,'mf_txppWframe_textbox101711').get_attribute("innerText");print(sangho+' 사업자번호로 조회 : 확인버튼')
          kwasekikan    = driver.find_element(By.ID,'mf_txppWframe_txtTxnrm').get_attribute("innerText")
          singoyouhyng  = driver.find_element(By.ID,'mf_txppWframe_txtBmanClNm').get_attribute("innerText")
          bubinyejung  = driver.find_element(By.ID,'mf_txppWframe_txtSmscCrpBmanYn').get_attribute("innerText")
          kwanhanseo  = driver.find_element(By.ID,'mf_txppWframe_textbox101708').get_attribute("innerText")
          juupjongcode  = driver.find_element(By.ID,'mf_txppWframe_txtTfbCd').get_attribute("innerText")
          kanibukayul  = driver.find_element(By.ID,'mf_txppWframe_txtSmpVatRte').get_attribute("innerText")
          saleTI  = driver.find_element(By.ID,'mf_txppWframe_edtEtxivSlsAmt').get_attribute("value").replace(",","")
          saleTIvat  = driver.find_element(By.ID,'mf_txppWframe_edtEtxivSlsTxamt').get_attribute("value").replace(",","")
          costTI  = driver.find_element(By.ID,'mf_txppWframe_edtEtxivPrhAmt').get_attribute("value").replace(",","")
          costTIvat  = driver.find_element(By.ID,'mf_txppWframe_edtEtxivPrhTxamt').get_attribute("value").replace(",","")
          saleNTI  = driver.find_element(By.ID,'mf_txppWframe_edtEtivSlsTxamt').get_attribute("value").replace(",","")
          costNTI  = driver.find_element(By.ID,'mf_txppWframe_edtEtivPrhTxamt').get_attribute("value").replace(",","")
          saleSinca  = driver.find_element(By.ID,'mf_txppWframe_edtTotaCrdcSls').get_attribute("value").replace(",","")
          costSinca  = driver.find_element(By.ID,'mf_txppWframe_edtBusnCrdcPrhAmt').get_attribute("value").replace(",","")
          costSincavat  = driver.find_element(By.ID,'mf_txppWframe_edtBusnCrdcPrhTxamt').get_attribute("value").replace(",","")
          saleCash  = driver.find_element(By.ID,'mf_txppWframe_edtCshSlsSplCftCmttAmt').get_attribute("value").replace(",","")
          saleCashvat  = driver.find_element(By.ID,'mf_txppWframe_edtCshptSlsTxamt').get_attribute("value").replace(",","")
          costCash  = driver.find_element(By.ID,'mf_txppWframe_edtCshptPrhAmt').get_attribute("value").replace(",","")
          costCashvat  = driver.find_element(By.ID,'mf_txppWframe_edtCshptPrhTxamt').get_attribute("value").replace(",","")   
          saleDaehang = driver.find_element(By.ID,'mf_txppWframe_edtTotaSleVcexSlsAmt').get_attribute("value").replace(",","")   
          costBockji  = driver.find_element(By.ID,'mf_txppWframe_edtCargDrerWlfCardPrhAmt').get_attribute("value").replace(",","")
          costBockjivat  = driver.find_element(By.ID,'mf_txppWframe_edtCargDrerWlfCardPrhTxamt').get_attribute("value").replace(",","")         
          saleExport = '0'  
          pretax  = driver.find_element(By.ID,'mf_txppWframe_edtSchuNftam').get_attribute("value").replace(",","") 
          pretaxNot  = driver.find_element(By.ID,'mf_txppWframe_edtSchuRtnNrfnTxamt').get_attribute("value").replace(",","") 
          pretaxKani  = driver.find_element(By.ID,'mf_txppWframe_edtSchuImpsTxamt').get_attribute("value").replace(",","") 
          pretaxKaniSingo  = driver.find_element(By.ID,'mf_txppWframe_edtSchuRtnTxamt').get_attribute("value").replace(",","") 
          purchaseSpecial  = driver.find_element(By.ID,'mf_txppWframe_edtBuyePmtSpcsPpmTxamt').get_attribute("value").replace(",","") 
          salecashSheet  = driver.find_element(By.ID,'mf_txppWframe_textbox101640').get_attribute("innerText").replace(",","") 
          realestateSheet = driver.find_element(By.ID,'mf_txppWframe_txtRaltRmlSplCftSpecAdttxYn').get_attribute("innerText").replace(",","") 

          strsql_f = "SELECT * FROM 부가가치세전자신고3 WHERE 사업자등록번호 = %s AND 과세기간 = %s AND 과세유형 = 'C17'"#c17은 고정임 수정말 것
          with connection.cursor() as cursor:
            cursor.execute(strsql_f, (result[3],kawasekisu))
            result_f = cursor.fetchall()
          connection.commit() 
          str_del = ""
          if  result_f:   str_del = "delete from 부가가치세통합조회 where 과세기수='"+ kawasekisu  + "' and 과세기간='" + kwasekikan  + "' and 사업자등록번호='" + result[3]  + "'" 
          else:           str_del = "delete from 부가가치세통합조회 where 과세기수='"+ kawasekisu  + "' and 사업자등록번호='" + result[3]  + "'" 
          cursor = connection.cursor()
          print(str_del);cursor.execute(str_del)
          str_ins = "INSERT INTO 부가가치세통합조회 VALUES ('"+kawasekisu+"', '"+kwasekikan + "', '" + singoGB + "', '" + result[3] + "', '"
          str_ins = str_ins + result[1] + "', '" + singoyouhyng + "', '"+ bubinyejung +"', '"+ kwanhanseo +"','"+ juupjongcode +"', '"+ kanibukayul +"', "
          str_ins = str_ins + saleTI + ", "+ saleTIvat + ", "+ saleNTI + ", "+ saleSinca + ", 0, "
          str_ins = str_ins + saleCash + ", "+ saleCashvat + ", "+ saleExport + ", 0, "
          str_ins = str_ins + costTI +", " + costTIvat +", " + costNTI +", " + costSinca +", " + costSincavat +", "
          str_ins = str_ins + costCash +", " + costCashvat +", " + costBockji +", " + costBockjivat +", "
          str_ins = str_ins + pretax +", " + pretaxNot +", " + pretaxKani +", " + pretaxKaniSingo +", " + purchaseSpecial +", "
          str_ins = str_ins + "'"+ salecashSheet +"', '"+realestateSheet +"', '"+saleDaehang+"') "
          cursor.execute(str_ins);print(str_ins)
  elif flag=='7':#4.수출실적명세서
    strsql = " select a.seq_no ,biz_name,biz_manager,biz_no "
    strsql += " from mem_user a, mem_deal b where a.seq_no=b.seq_no and a.biz_no<>'' and keeping_yn='Y' and biz_manager<>''  "    
    if flagFirst=="관리자별":           strsql += " and biz_manager ='" + flagManagerName + "' "  
    elif flagFirst=="전체(기장만)" :    strsql += " and biz_manager  in ("+flagGroup+") " 
    elif flagFirst=="단일사업자":       strsql += " and a.seq_no ='"+flagSeqNo+"'   "  
    else:                              strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    # if flagFirst=="관리자별" or flagFirst[:2]=="전체":                                    #23.10.19 단체로 할때는 예정대 법인만 / 단일사업자는 예정신고가능
    #   if flagPeriod2=='1분기' or flagPeriod2=='3분기':strsql += " and biz_type<=3 "   
    strsql += " order by a.seq_no/1"
    print(strsql)
    cursor = connection.cursor()
    cursor.execute(strsql)
    members = cursor.fetchall()
    connection.commit()
    if members:
      driver = utils.conHometaxLogin(htxLoginID,False) ;time.sleep(2);    
      driver.find_element(By.ID,'mf_wfHeader_hdGrp920').click();print('세무대리')
      driver.find_element(By.ID,'span_a_4801200000').click();time.sleep(0.5);print('수출신고필증')           
      for result in members:
        strsql = "select * from 스크래핑관리 where crt_dt='" + today + "' and seqno_lastTry='"+result[0]+"' and bigo='수출실적명세서'"
        cursor = connection.cursor()
        cursor.execute(strsql)   
        scraps = cursor.fetchall()
        connection.commit()  
        if  not scraps:       
          if len(result[3])>=10:        
            bizNo = result[3].replace('-','');time.sleep(1)
            driver.find_element(By.ID,'mf_txppWframe_inputBsno').clear();time.sleep(0.5)
            driver.find_element(By.ID,'mf_txppWframe_inputBsno').send_keys(bizNo);time.sleep(1);  
            selectYear = Select(driver.find_element(By.ID,'mf_txppWframe_edtYear'));       selectYear.select_by_visible_text(str(flagYear));time.sleep(0.5)
            if flagPeriod=="월별":
              driver.find_element(By.ID,'mf_txppWframe_shpnYmGubun_input_0').click();time.sleep(0.25)
              selectPeriod = Select(driver.find_element(By.ID,'mf_txppWframe_edtMon'));       selectPeriod.select_by_visible_text(str(flagPeriod2));time.sleep(0.5)
            elif flagPeriod=="분기별":
              # 요소가 로드될 때까지 대기
              wait = WebDriverWait(driver, 10)
              # 클릭할 요소 찾기 (예: 버튼)
              button = wait.until(EC.presence_of_element_located((By.ID, 'mf_txppWframe_shpnYmGubun_input_1')))
              # JavaScript로 클릭
              driver.execute_script("arguments[0].click();", button)
              # 드롭다운 요소가 로드될 때까지 대기
              select_element = wait.until(EC.presence_of_element_located((By.ID, 'mf_txppWframe_edtQrt')))
              # 옵션을 JavaScript로 선택
              driver.execute_script(f"arguments[0].value = '{flagPeriod2}';", select_element)
              # 변경된 값을 선택하기 위해 change 이벤트 트리거
              driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", select_element)
            elif flagPeriod=="반기별":
              driver.find_element(By.ID,'mf_txppWframe_shpnYmGubun_input_2').click();time.sleep(0.25)
              selectPeriod = Select(driver.find_element(By.ID,'mf_txppWframe_edtHt'));       selectPeriod.select_by_visible_text(str(flagPeriod2));time.sleep(0.5)
            driver.find_element(By.ID,'mf_txppWframe_trigger93').click(); print(bizNo+' 조회하기 버튼')
            try:
                WebDriverWait(driver, 1).until(EC.alert_is_present()); al = Alert(driver); al.accept();time.sleep(0.25)
                continue  # 알람 창이 발생하면 다음 루프로 진행
            except:
                pass  
            totalCnt = int(driver.find_element(By.ID,'mf_txppWframe_textbox1713').text.replace("건",""));  print('총건수:'+str(totalCnt))
            if totalCnt>0:             
              print("조회페이지수:"+str(math.ceil(totalCnt/20)))
              for j in range(1,math.ceil(totalCnt/20)+1):
                tmpPaging = "pglNavi_page_" + str(j)
                print("페이지:"+str(j)+",ID"+tmpPaging)
                pagingBtn = driver.find_element(By.ID,tmpPaging)
                driver.implicitly_wait(10)
                pagingBtn.click()
                driver.implicitly_wait(30)    
                time.sleep(2) #필수
                print(tmpPaging+' 테이블 가져오기')
                table_export = driver.find_element(By.ID,'grdVatKcsvExprArstMate_body_table')
                tbody_export = table_export.find_element(By.ID,'grdVatKcsvExprArstMate_body_tbody') 
                rows_export = tbody_export.find_elements(By.TAG_NAME, 'tr')
                print(result[1]+' 신용카드조회결과 개수 :'+ str(len(rows_export)))        
                for k, tr_export in enumerate(rows_export):
                  td_export = tr_export.find_elements(By.TAG_NAME,"td")
                  if len(td_export)>1:
                    txtYYMM="";txtNum="";txtTCode="";txtTCode="";txtQRate="0";txtExportAmt="0";txtExchangeAmt="0";txtTotCnt='0';stnd_Gb=0
                    for i, tag_export in enumerate(td_export):
                      id_td = tag_export.get_attribute('id');#print(id_td)
                      txt_td_export = tag_export.get_attribute("innerText");#print(txt_td)      
                      if i==0:txtNum  = txt_td_export     ;print(txtNum)                #수출신고번호
                      if i==1:txtYYMM  = txt_td_export.replace("-","")    #선적일자
                      if i==2:txtTCode = txt_td_export                    #통화코드
                      if i==3:txtQRate = txt_td_export.replace(",","")      #환율
                      if i==4:txtExportAmt = txt_td_export.replace(",","")   #외화금액
                      if i==5:txtExchangeAmt = txt_td_export.replace(",","")#환산금액
                    if len(txtYYMM)>=6:  
                      if int(txtYYMM[4:6])<=3 :					stnd_Gb = '1'
                      elif int(txtYYMM[4:6])<=6 :				stnd_Gb = '2'
                      elif int(txtYYMM[4:6])<=9 :				stnd_Gb = '3'
                      elif int(txtYYMM[4:6])<=12 :				stnd_Gb = '4'
                      str_del = "delete from Tbl_HomeTax_SaleExport where seq_no="+ result[0]  + " and stnd_Gb='" + stnd_Gb  + "' and left(tran_dt,4)='" + str(flagYear)  + "' and 수출신고번호='"+txtNum+"'" 
                      print(str_del);cursor.execute(str_del)                       
                      str_ins = "INSERT INTO Tbl_HomeTax_SaleExport VALUES ('"+txtYYMM+"',"
                      str_ins += "isnull((select max(rowSeq)+1 from Tbl_HomeTax_SaleExport where  seq_no="+ result[0]  + " and left(tran_dt,4)="+str(flagYear)+" and stnd_GB='"+stnd_Gb+"'),0)"
                      str_ins += ", '" + stnd_Gb + "', '"+result[0]+"', '"+result[1]+"','"+ txtNum +"','"+txtTCode+"',"+txtQRate + "," + txtExportAmt + "," + txtExchangeAmt + ", GETDATE())"; print(str_ins)
                      print(str_ins);cursor.execute(str_ins) 
                if totalCnt>200 and divmod(j,10)[1]==0 and j>0:
                  print(str(j)+"개 이후 넥스트버튼 ")
                  nextBtn = driver.find_element(By.XPATH,'//*[@id="pglNavi_next_btn"]/a')
                  driver.implicitly_wait(10)
                  nextBtn.click()
                  print("넥스트 클릭")
                  driver.implicitly_wait(10)
                  time.sleep(1)
            strsql = "insert into 스크래핑관리 values('"+today+"','"+result[0]+"','"+result[1]+"','수출실적명세서')"       
            cursor.execute(strsql)
        # 내국신용장구매확인서
        strsql = "select * from 스크래핑관리 where crt_dt='" + today + "' and seqno_lastTry='"+result[0]+"' and bigo='내국신용장구매확인서'"
        cursor = connection.cursor()
        cursor.execute(strsql)   
        scraps = cursor.fetchall()
        connection.commit()  
        if  not scraps:           
          for iflag in ['0','1']:
            driver.get('https://teht.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/rn/a/b/a/e/UTERNAA134.xml&menuId=48&subMenuId=01');time.sleep(1)
            driver.find_element(By.ID,'edtDmsLcRcrdClCd_input_' + iflag).click();        print('내국신용장0/구매확인서1 클릭');      time.sleep(0.5)    
            if len(result[3])>=10:
              bizNo = result[3].replace('-','');time.sleep(1)
              # driver.find_element(By.ID,'inputBsno').clear();time.sleep(0.5)
              driver.find_element(By.ID,'inputBsno').send_keys(bizNo);time.sleep(0.25);  
              driver.find_element(By.ID,'edtYear').send_keys(flagYear);time.sleep(0.25);                  
              workKi = '1';workPer = '0'
              if flagPeriod=="월별":
                if   int(flagPeriod2.replace('월',''))<=3:workKi = '1';workPer = '0'
                elif int(flagPeriod2.replace('월',''))<=6:workKi = '1';workPer = '1'
                elif int(flagPeriod2.replace('월',''))<=9:workKi = '2';workPer = '0'
                elif int(flagPeriod2.replace('월',''))<=12:workKi = '2';workPer = '1'
              elif flagPeriod=="분기별":
                if   int(flagPeriod2.replace('분기',''))==1:workKi = '1';workPer = '0'
                elif int(flagPeriod2.replace('분기',''))==2:workKi = '1';workPer = '1'
                elif int(flagPeriod2.replace('분기',''))==3:workKi = '2';workPer = '0'
                elif int(flagPeriod2.replace('분기',''))==4:workKi = '2';workPer = '1'
              elif flagPeriod=="반기별":
                if flagPeriod2=='상반기':workKi = '1';workPer = '1'
                elif flagPeriod2=='하반기':workKi = '2';workPer = '1'
              selectKi = Select(driver.find_element(By.ID,'edtHt'));       selectKi.select_by_visible_text(workKi+'기');time.sleep(0.5)
              driver.find_element(By.ID,'edtRtnClCd_input_'+workPer).click();
              driver.find_element(By.ID,'trigger93').click(); print(bizNo+' 조회하기 버튼')
              try:
                  WebDriverWait(driver, 1).until(EC.alert_is_present()); al = Alert(driver); al.accept();time.sleep(0.25)
                  continue  # 알람 창이 발생하면 다음 루프로 진행
              except:
                  pass  
              totalCnt = int(driver.find_element(By.ID,'textbox1713').text.replace("건",""));  print('총건수:'+str(totalCnt))
              if totalCnt>0:
                print("조회페이지수:"+str(math.ceil(totalCnt/10)))

                table_export = driver.find_element(By.ID,'grdVatPurcCnftIsnBrkdMate_body_table')
                tbody_export = table_export.find_element(By.ID,'grdVatPurcCnftIsnBrkdMate_body_tbody') 
                rows_export = tbody_export.find_elements(By.TAG_NAME, 'tr')
                print(result[1]+' 신용카드조회결과 개수 :'+ str(len(rows_export)))        
                for k, tr_export in enumerate(rows_export):
                  td_export = tr_export.find_elements(By.TAG_NAME,"td")
                  if len(td_export)>1:
                    txtYYMM="";txtSeper="";txtNum='';txtSheetDate='';txtBizno="";txtSangho='';txtTCode="";txtQRate="0";txtExportAmt="0";txtExchangeAmt="0";txtWonwha='0';txtQC="";txtJaroe="";stnd_Gb=""
                    for i, tag_export in enumerate(td_export):
                      id_td = tag_export.get_attribute('id');#print(id_td)
                      txt_td_export = tag_export.get_attribute("innerText");#print(txt_td)      
                      if i==0:txtSeper  = txt_td_export    # ;print(txtSeper)                #서류구분
                      if i==1:txtNum  = txt_td_export    #서류번호
                      if i==2:txtSheetDate = txt_td_export                    #발급일
                      if i==3:txtBizno = txt_td_export
                      if i==4:txtSangho = txt_td_export
                      if i==5:txtExportAmt = txt_td_export.replace(",","")   #외화금액
                      if txtExportAmt=='':txtExportAmt='0'
                      if i==6:txtTCode = txt_td_export
                      if i==7:txtWonwha = txt_td_export.replace(",","")
                      if i==8:txtExchangeAmt = txt_td_export.replace(",","")
                      if i==9:txtYYMM = txt_td_export.replace("-","")
                      if i==10:txtQC = txt_td_export
                      if i==11:txtJaroe = txt_td_export
                      
                    if len(txtYYMM)>=6:  
                      if int(txtYYMM[4:6])<=3 :					stnd_Gb = '1'
                      elif int(txtYYMM[4:6])<=6 :				stnd_Gb = '2'
                      elif int(txtYYMM[4:6])<=9 :				stnd_Gb = '3'
                      elif int(txtYYMM[4:6])<=12 :				stnd_Gb = '4'
                      str_del = "delete from Tbl_HomeTax_SaleTT where seq_no="+ result[0]  + " and stnd_Gb='" + stnd_Gb  + "' and left(tran_dt,4)='" + str(flagYear)  + "' and 서류번호='"+txtNum+"'" 
                      print(str_del);cursor.execute(str_del)
                      str_ins = "INSERT INTO Tbl_HomeTax_SaleTT VALUES ('"+txtYYMM+"',"
                      str_ins += "isnull((select max(rowSeq)+1 from Tbl_HomeTax_SaleTT where seq_no="+ result[0]  + " and left(tran_dt,4)="+str(flagYear)+" and stnd_GB='"+stnd_Gb+"'),0)"
                      str_ins += ", '" + stnd_Gb + "', '"+result[0]+"', '"+result[1]+"','"+ txtSeper +"','"+txtNum+"','"+txtSheetDate +"','"+txtBizno +"','"+txtSangho +"',"+txtExportAmt +",'"+txtTCode 
                      str_ins += "'," + txtWonwha + "," + txtExchangeAmt  +",'"+txtQC+ "','"+txtJaroe + "', GETDATE())"; print(str_ins)
                      print(str_ins);cursor.execute(str_ins) 
          strsql = "insert into 스크래핑관리 values('"+today+"','"+result[0]+"','"+result[1]+"','내국신용장구매확인서')"       
          cursor.execute(strsql)
        # driver.switch_to.default_content() ;      print('세무대리 메뉴이동 하기 전에 프레임 원복');time.sleep(1)
  elif flag=='6':#3.신용카드/구매대행 매출조회
    strsql = " select a.seq_no ,biz_name,biz_manager,biz_no "
    strsql += " from mem_user a, mem_deal b where a.seq_no=b.seq_no and a.biz_no<>'' and keeping_yn='Y' and biz_manager<>''  "
    if flagFirst=="관리자별":           strsql += " and biz_manager ='" + flagManagerName + "' "  
    elif flagFirst=="전체(기장만)" :    strsql += " and biz_manager  in ("+flagGroup+") " 
    elif flagFirst=="단일사업자":       strsql += " and a.seq_no ='"+flagSeqNo+"'   "  
    else:                              strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    # if flagFirst=="관리자별" or flagFirst[:2]=="전체":                                    #23.10.19 단체로 할때는 예정대 법인만 / 단일사업자는 예정신고가능
    #   if flagPeriod2=='1분기' or flagPeriod2=='3분기':strsql += " and biz_type<=3 "   
    strsql += " order by a.seq_no/1"  
    print(strsql)
    cursor = connection.cursor()
    cursor.execute(strsql)
    members = cursor.fetchall()
    connection.commit()
    if members:
      driver = utils.conHometaxLogin(htxLoginID,False) ;time.sleep(3)
      driver.find_element(By.ID,'mf_wfHeader_hdGrp920').click();print('세무대리')
      driver.find_element(By.ID,'span_a_4801140000').click();time.sleep(0.5);print('신용카드/판매(결제)대행 매출자료 조회')        
      for result in members:
        strsql = "select * from 스크래핑관리 where crt_dt='" + today + "' and seqno_lastTry='"+result[0]+"' and bigo='신용카드판매대행매출조회'"
        cursor = connection.cursor()
        cursor.execute(strsql)   
        scraps = cursor.fetchall()
        connection.commit()  
        if  not scraps or flagFirst=="단일사업자":     
          if len(result[3])>=10:
            bizNo = result[3].split('-');time.sleep(1)
            driver.find_element(By.ID,'mf_txppWframe_edtTxprDscmNo1').clear();time.sleep(0.5)
            driver.find_element(By.ID,'mf_txppWframe_edtTxprDscmNo1').send_keys(bizNo[0]);time.sleep(0.25);    
            driver.find_element(By.ID,'mf_txppWframe_edtTxprDscmNo2').clear();time.sleep(0.5)
            driver.find_element(By.ID,'mf_txppWframe_edtTxprDscmNo2').send_keys(bizNo[1]);time.sleep(0.25);          
            driver.find_element(By.ID,'mf_txppWframe_edtTxprDscmNo3').clear();time.sleep(0.5)
            driver.find_element(By.ID,'mf_txppWframe_edtTxprDscmNo3').send_keys(bizNo[2]);time.sleep(0.5);   
            selectStlYr = Select(driver.find_element(By.ID,'mf_txppWframe_selectStlYr'));              selectStlYr.select_by_visible_text(flagYear);time.sleep(0.5)
            qtrStart="";qtrEnd=""
            print(f'작업분기:{work_qt}')
            if str(work_qt)=='1':qtrStart="1분기";qtrEnd="2분기"
            elif str(work_qt)=='2':qtrStart="1분기";qtrEnd="2분기"
            elif str(work_qt)=='3':qtrStart="3분기";qtrEnd="4분기"
            elif str(work_qt)=='4':qtrStart="3분기";qtrEnd="4분기"
            selectQrtFrom = Select(driver.find_element(By.ID,'mf_txppWframe_selectQrtFrom'));          selectQrtFrom.select_by_visible_text(qtrStart);time.sleep(0.5)        
            selectQrtTo = Select(driver.find_element(By.ID,'mf_txppWframe_selectQrtTo'));              selectQrtTo.select_by_visible_text(qtrEnd);time.sleep(0.5)        
            driver.find_element(By.ID,'mf_txppWframe_trigger163').click();print(result[1]+' 사업자번호로 조회 : 확인버튼');time.sleep(1.5)
            try:
                WebDriverWait(driver, 1).until(EC.alert_is_present()); al = Alert(driver); al.accept();time.sleep(0.25)
                continue  # 알람 창이 발생하면 다음 루프로 진행
            except:
                pass        
            #신용카드 제로페이 조회
            table_sinca = driver.find_element(By.ID,'mf_txppWframe_grdCrdcTrsBrkdMateAdmDVOList_body_table')
            tbody_sinca = table_sinca.find_element(By.ID,'mf_txppWframe_grdCrdcTrsBrkdMateAdmDVOList_body_tbody') 
            rows_sinca = tbody_sinca.find_elements(By.TAG_NAME, 'tr')
            print(result[1]+' 신용카드조회결과 개수 :'+ str(len(rows_sinca)))        
            for j, tr_sinca in enumerate(rows_sinca):
              td_sinca = tr_sinca.find_elements(By.TAG_NAME,"td")
              if len(td_sinca)>1:
                txtYYMM="";txtTitle="";txtCnt="0";txtTotsum="0";txtSincaSale="0";txtPurchase="0";txtFreetax='0';stnd_Gb=0
                for i, tag_sinca in enumerate(td_sinca):
                  id_td = tag_sinca.get_attribute('id');#print(id_td)
                  txt_td_sinca = tag_sinca.get_attribute("innerText");#print(txt_td)      
                  if i==1:txtYYMM  = txt_td_sinca.replace("-","")#승인연월      
                  if i==2:txtTitle = txt_td_sinca#자료구분 : [신용카드 자료]
                  if i==3:txtCnt = txt_td_sinca.replace(",","")#건수
                  if i==5:txtTotsum = txt_td_sinca.replace(",","")#매출액계
                  if i==6:txtSincaSale = txt_td_sinca.replace(",","")#신용카드
                  if i==7:txtPurchase = txt_td_sinca.replace(",","")#구매전용
                  if i==8:txtFreetax = txt_td_sinca.replace(",","")#비과세
                if len(txtYYMM)>=6:  
                  if int(txtYYMM[4:6])<=3 :					stnd_Gb = '1'
                  elif int(txtYYMM[4:6])<=6 :				stnd_Gb = '2'
                  elif int(txtYYMM[4:6])<=9 :				stnd_Gb = '3'
                  elif int(txtYYMM[4:6])<=12 :				stnd_Gb = '4'
                  str_del = "delete from Tbl_HomeTax_SaleCard where seq_no="+ result[0]  + " and Tran_MM='" + txtYYMM  + "' and SaleGubun='" + txtTitle  + "'"   
                  print(str_del);cursor.execute(str_del)
                  str_ins = "INSERT INTO Tbl_HomeTax_SaleCard VALUES ('"+txtYYMM[:4]+"', '"+result[0] + "', '" + txtTitle + "', '" + stnd_Gb + "', '"+txtYYMM+"', "+txtCnt+","+ txtTotsum +","+txtSincaSale+","+txtPurchase + "," + txtFreetax + ", GETDATE(),"
                  str_ins +="isnull((SELECT TOP 1 acnt_cd  FROM ds_slipledgr2 where seq_no="+ result[0]  + " and work_yy>"+str(int(flagYear)-5)+" and acnt_cd>=401 and acnt_cd<430 GROUP BY acnt_cd ORDER BY COUNT(acnt_cd) DESC),'401'),'','')"
                  print(str_ins);cursor.execute(str_ins) 
            table_sinca = None;tbody_sinca = None;rows_sinca = None;tr_sinca = None;

            #판매대행 조회
            totalCnt = int(driver.find_element(By.ID,'txtTotal0').text);  print('총건수:'+str(totalCnt))

            if totalCnt>0:
              print("조회페이지수:"+str(math.ceil(totalCnt/50)))
              for j in range(1,math.ceil(totalCnt/50)+1):
                tmpPaging = "pglNavi0_page_" + str(j)
                print("페이지:"+str(j)+",ID"+tmpPaging)
                pagingBtn = driver.find_element(By.ID,tmpPaging);   driver.implicitly_wait(10);      pagingBtn.click();                driver.implicitly_wait(30)    
                time.sleep(2) #필수
                table_pmdh = driver.find_element(By.ID,'mf_txppWframe_grdSleVcexSlsMateInqrDVOList_body_table')
                tbody_pmdh = table_pmdh.find_element(By.ID,'mf_txppWframe_grdSleVcexSlsMateInqrDVOList_body_tbody') 
                rows_pmdh = tbody_pmdh.find_elements(By.TAG_NAME, 'tr')
                print(result[1]+' 판매대행조회결과 개수 :'+ str(len(rows_pmdh)))
                for j, tr_pmdh in enumerate(rows_pmdh):
                  td_pmdh = tr_pmdh.find_elements(By.TAG_NAME,"td")
                  if len(td_pmdh)>1:
                    txtYYMM="";txtTitle="";txtCnt="0";txtTotsum="0";txtSincaSale="0";txtPurchase="0";txtFreetax='0';stnd_Gb=0
                    for i, tag_pmdh in enumerate(td_pmdh):
                      id_td = tag_pmdh.get_attribute('id');#print(id_td)
                      txt_td_pmdh = tag_pmdh.get_attribute("innerText");#print(txt_td)      
                      if i==1:txtYYMM  = txt_td_pmdh.replace("-","")#승인연월      
                      if i==2:txtCnt = txt_td_pmdh.replace(",","")#건수
                      if i==3:txtTotsum = txt_td_pmdh.replace(",","")#매출액계
                      if i==4:txtTitle = txt_td_pmdh#자료구분 : 판매대행사

                    if len(txtYYMM)>=6:  
                      if int(txtYYMM[4:6])<=3 :					stnd_Gb = '1'
                      elif int(txtYYMM[4:6])<=6 :				stnd_Gb = '2'
                      elif int(txtYYMM[4:6])<=9 :				stnd_Gb = '3'
                      elif int(txtYYMM[4:6])<=12 :				stnd_Gb = '4'
                      str_del = "delete from Tbl_HomeTax_SaleCard where seq_no="+ result[0]  + " and Tran_MM='" + txtYYMM  + "' and SaleGubun='" + txtTitle  + "'"  
                      print(str_del);cursor.execute(str_del)
                      str_ins = "INSERT INTO Tbl_HomeTax_SaleCard VALUES ('"+txtYYMM[:4]+"', '"+result[0] + "', '" + txtTitle + "', '" + stnd_Gb + "', '"+txtYYMM+"', "+txtCnt+","+ txtTotsum +","+txtSincaSale+","+txtPurchase + "," + txtFreetax + ", GETDATE(),"
                      str_ins +="isnull((SELECT TOP 1 acnt_cd  FROM ds_slipledgr2 where seq_no="+ result[0]  + " and work_yy>"+str(int(flagYear)-5)+" and acnt_cd>=401 and acnt_cd<430 GROUP BY acnt_cd ORDER BY COUNT(acnt_cd) DESC),'401'),'','')"
                      print(str_ins);cursor.execute(str_ins) 
                table_pmdh = None;tbody_pmdh = None;rows_pmdh = None;tr_pmdh = None
                strsql = "insert into 스크래핑관리 values('"+today+"','"+result[0]+"','"+result[1]+"','신용카드판매대행매출조회')"       
                cursor.execute(strsql)
  elif flag=='2':#2.전자세금계산서 저장
    # 합계표 결과와 다른 경우
    strsql = " SELECT a.seq_no ,biz_name,biz_manager "
    strsql += " ,CASE WHEN  "
    strsql += " isnull((select sum(합계금액) from 전자세금계산서 where seq_no=a.seq_no and " + workPeriod2 + "),0) =  "
    strsql += " isnull((select sum(합계금액) from 전자세금계산서합계표 where seq_no=a.seq_no and work_yy="+flagYear+" and work_qt="+str(work_qt)+"),0)  "
    strsql += " THEN  'True' else 'False'  END AS result "
    strsql += " from mem_user a, mem_deal b where a.seq_no=b.seq_no and a.biz_no<>'' and keeping_yn='Y' and biz_manager<>'' "
    if flagFirst=="관리자별":           strsql += " and biz_manager ='" + flagManagerName + "' "  
    elif flagFirst=="전체(기장만)" :    strsql += " and biz_manager  in ("+flagGroup+") " 
    elif flagFirst=="단일사업자":       strsql += " and a.seq_no ='"+flagSeqNo+"'   "  
    else:                              strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    # if flagFirst=="관리자별" or flagFirst[:2]=="전체":                                    #23.10.19 단체로 할때는 예정대 법인만 / 단일사업자는 예정신고가능
    #   if flagPeriod2=='1분기' or flagPeriod2=='3분기':strsql += " and biz_type<=3 "   
    strsql += " order by a.seq_no/1"
    print(strsql)
    cursor = connection.cursor()
    cursor.execute(strsql)
    members = cursor.fetchall()
    connection.commit()
    connection.close()    
    if members:
      driver = utils.conHometaxLogin(htxLoginID,False) ;time.sleep(1)
      driver.find_element(By.ID,'mf_wfHeader_textbox81212967').click();print('전체메뉴');time.sleep(1)
      driver.find_element(By.ID,'mf_txppWframe_tabControl8_tab_tabs_08_tabHTML').click();time.sleep(0.5);print('세무대리')
      driver.find_element(By.ID,'grpMenuAtag_48_4801090000').click();print('전자(세금)계산서 조회')   ;time.sleep(1) 
      WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID, 'mf_txppWframe_btnChg'))).click();time.sleep(2);print('월분기별 목록조회')
      for member in members:
        if member[3]=='False':    
          try:          
            print('수임사업자전환:'+str(member[0]))
            trigger501=""
            try:
              trigger501 = driver.find_element(By.XPATH, '//*[@id="mf_txppWframe_wf01_trigger501"]')
              
              # receivebtn = driver.find_element(By.XPATH,'//*[@id="mf_txppWframe_wf01_trigger501"]'); driver.implicitly_wait(10)
              if trigger501:
                driver.execute_script("arguments[0].click();", trigger501)
              else:
                print('@@@!!!---수임처버튼 전환 버튼 사라짐---!!!@@@')   
                continue           
            except NoSuchElementException:
              print("수임처버튼 엘리먼트 선택 실패")
              continue
            print('수임처버튼 전환 버튼 클릭')
            time.sleep(1)
            driver.find_element(By.ID,'mf_txppWframe_wf01_UTEETZZA21_wframe_selectbox5').click()
            driver.find_element(By.XPATH,r'/html/body/div[4]/div[2]/div[1]/div/div[1]/div[1]/div[2]/div[2]/div/div/dl[2]/dd/select/option[2]').click();print('사업자번호 선택')
            memuser = MemUser.objects.get(seq_no=member[0])
            bizno = memuser.biz_no
            if flagFirst=="전체":
              strsql = "Merge 스크래핑관리 as A Using (select '"+today+"' as crt_dt,'"+member[0]+"' as seq_no) as B "
              strsql += "On A.crt_dt=B.crt_dt  and A.seqno_lastTry=B.seq_no "
              strsql += "when matched then update set seqno_lastTry='"+member[0]+"',bigo='전자세금계산서' "
              strsql += "When Not Matched Then insert values('"+today+"','"+member[0]+"','"+memuser.biz_name+"','전자세금계산서');"
              print(strsql)
              cursor = connection.cursor();cursor.execute(strsql); cursor.commit()
            time.sleep(2)
            print('사업자번호전달:'+bizno)
            driver.find_element(By.ID,'mf_txppWframe_wf01_UTEETZZA21_wframe_txprDscmNoA').send_keys(bizno)
            bizChange = driver.find_element(By.ID,'mf_txppWframe_wf01_UTEETZZA21_wframe_trigger85')            
            try:
              bizChange.click(); driver.implicitly_wait(10)
              try:
                  # 조회결과가 없으면 다음 사업자로 넘어간다
                  noResultMessage = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'mf_txppWframe_UTEETZZA21_wframe_grdResult_noResultMessageFrame_wq_uuid_14779')))
                  print(f'{memuser.biz_name}은 수임동의 되어있지 않아 조회하지 않습니다')
                  driver.find_element(By.ID,'mf_txppWframe_UTEETZZA21_wframe_btnClose').click()
                  continue
              except Exception as e:              
                driver.find_element(By.XPATH,r'/html/body/div[4]/div[2]/div[1]/div/div[1]/div[1]/div[2]/div[4]/div/div[1]/div/table/tbody/tr[1]/td[1]/label').click();print('업체선택라디오버튼')
                print('조회버튼')
                driver.implicitly_wait(10)
                driver.find_element(By.ID,'mf_txppWframe_wf01_UTEETZZA21_wframe_btnProcess').click()
                time.sleep(1)
                for i in [1, 2, 3, 4]:
                  time.sleep(1)
                  #  전자세금계산서 radioEtxivClsfCd_input_0/전자계산서 radioEtxivClsfCd_input_1 
                  if i in [1,2]:
                    vat = driver.find_element(By.CSS_SELECTOR,'#mf_txppWframe_wf01_radioEtxivClsfCd > div.w2radio_item.w2radio_item_0 > label');   driver.implicitly_wait(5);        print('전자세금계산서 조회시작')
                    driver.execute_script("arguments[0].click();", vat);        print('전자세금계산서 조회버튼클릭')
                    # vat.click() 
                  else:   
                    print('전자계산서 조회시작')
                    nonVat = driver.find_element(By.CSS_SELECTOR,'#mf_txppWframe_wf01_radioEtxivClsfCd > div.w2radio_item.w2radio_item_1 > label');     print('전자계산서 조회 드라이버 선택');     driver.implicitly_wait(10)
                    driver.execute_script("arguments[0].click();", nonVat);     print('전자계산서 조회버튼 클릭')
                    # nonVat.click()
                  driver.implicitly_wait(3)

                  #  매입 매출 선택( 매출:radio3_input_0  매입:radio3_input_1)
                  if i in [1,3]:
                    print("매출조회시작")
                    try:
                      saleBtn = driver.find_element(By.CSS_SELECTOR,'#mf_txppWframe_radio3 > div.w2radio_item.w2radio_item_0 > label');      time.sleep(1.5)
                      if saleBtn: driver.execute_script("arguments[0].click();", saleBtn)
                        # saleBtn.click() 
                    except NoSuchElementException:
                      time.sleep(2)
                      saleBtn = driver.find_element(By.CSS_SELECTOR,'#mf_txppWframe_radio3 > div.w2radio_item.w2radio_item_0 > label');      driver.implicitly_wait(10)
                      if saleBtn: driver.execute_script("arguments[0].click();", saleBtn)
                        # saleBtn.click() 
                  else:
                    print("매입조회시작")
                    costBtn = driver.find_element(By.CSS_SELECTOR,'#mf_txppWframe_radio3 > div.w2radio_item.w2radio_item_1 > label'); driver.implicitly_wait(10);time.sleep(1);  
                    driver.execute_script("arguments[0].click();", costBtn)     
                    # costBtn.click() 
                  driver.implicitly_wait(10)

                  #  조회기간 구분 선택(분기별:radio4_input_1      월별:radio4_input_0        반기별:radio4_input_2
                  tmpflagPeriod="";tmpflagPeriod2 = "";print("조회기간:" + flagPeriod)
                  if flagPeriod=="월별":
                    tmpflagPeriod = "#mf_txppWframe_radio4 > div.w2radio_item.w2radio_item_0 > label";  tmpflagPeriod2 = "mf_txppWframe_selectboxMonth"
                  elif flagPeriod=="분기별":
                    tmpflagPeriod = "#mf_txppWframe_radio4 > div.w2radio_item.w2radio_item_1 > label";  tmpflagPeriod2 = "mf_txppWframe_selectboxQrt"
                  elif flagPeriod=="반기별":
                    tmpflagPeriod = "#mf_txppWframe_radio4 > div.w2radio_item.w2radio_item_2 > label";  tmpflagPeriod2 = "mf_txppWframe_selectboxQrt2"
                  periodBtn = driver.find_element(By.CSS_SELECTOR,tmpflagPeriod); driver.implicitly_wait(10);  periodBtn.click() ;  driver.implicitly_wait(5)
                  select = Select(driver.find_element(By.ID,tmpflagPeriod2)); select.select_by_visible_text(workPeriod);  driver.implicitly_wait(5)                

                  # 발급구분 작성일자 / 발급일자   selectbox11
                  select = Select(driver.find_element(By.ID,'mf_txppWframe_selectbox11'))
                  if flagPeriod=="월별":select.select_by_visible_text('발급일자')
                  else                 :select.select_by_visible_text('작성일자')
                  driver.implicitly_wait(5)

                  # 조회기간 선택(연)
                  select = Select(driver.find_element(By.ID,'mf_txppWframe_selectboxYear'))
                  select.select_by_visible_text('{}'.format(flagYear))
                  driver.implicitly_wait(5)

                  #  정렬구분
                  select = Select(driver.find_element(By.ID,'mf_txppWframe_srtOpt'))
                  select.select_by_visible_text('내림차순')
                  driver.implicitly_wait(5)

                  print('조회하기 버튼 클릭')
                  searchBtn = driver.find_element(By.ID,'mf_txppWframe_trigger50'); driver.implicitly_wait(10); searchBtn.click(); time.sleep(1)
                  totalCnt = int(driver.find_element(By.ID,'mf_txppWframe_txtTotal').text);  print('총건수:'+str(totalCnt))

                  if totalCnt>0:
                    print("조회페이지수:"+str(math.ceil(totalCnt/10)))
                    for j in range(1,math.ceil(totalCnt/10)+1):
                      tmpPaging = "mf_txppWframe_pglNavi_page_" + str(j)
                      print("페이지:"+str(j)+",ID"+tmpPaging)
                      pagingBtn = driver.find_element(By.ID,tmpPaging)
                      driver.implicitly_wait(10)
                      pagingBtn.click()
                      driver.implicitly_wait(30)    
                      time.sleep(2) #필수
                      print(tmpPaging+' 테이블 가져오기')
                      table = driver.find_element(By.ID,'mf_txppWframe_resultGrid_body_table')
                      print("테이블획득")
                      driver.implicitly_wait(30)    

                      tbody = table.find_element(By.ID,'mf_txppWframe_resultGrid_body_tbody');    print("티바디획득")
                      print('조회할테이블수:'+str(len(tbody.find_elements(By.TAG_NAME,'tr'))))
                      tr=[]
                      for tr in tbody.find_elements(By.TAG_NAME,'tr'):
                        td = tr.find_elements(By.TAG_NAME,"td")
                        approveNo="";kkdk="";kkke="";kkkt="";TIsep="";TIuh="";receiveYN="";eml_sup="";eml_buyer1=""
                        if i in [1,2]: 
                          approveNo = td[14].get_attribute("innerText")
                          kkdk      = td[11].get_attribute("innerText").replace(",","")
                          kkke      = td[12].get_attribute("innerText").replace(",","")
                          kkkt      = td[13].get_attribute("innerText").replace(",","")
                          TIsep     = td[15].get_attribute("innerText").replace(",","")
                          TIuh      = td[16].get_attribute("innerText").replace(",","")
                          receiveYN = td[18].get_attribute("innerText")
                          eml_sup   = td[19].get_attribute("innerText")
                          eml_buyer1=td[20].get_attribute("innerText")
                        else:          
                          approveNo = td[12].get_attribute("innerText")
                          kkdk      = td[11].get_attribute("innerText").replace(",","")
                          kkke      = td[11].get_attribute("innerText").replace(",","")
                          kkkt      = "0"
                          TIsep     = td[13].get_attribute("innerText").replace(",","")
                          TIuh      = td[14].get_attribute("innerText")
                          receiveYN = ""
                          eml_sup   = td[18].get_attribute("innerText")
                          eml_buyer1= ""
                        supplier_name="";buyer_name="";supplier_ceoname="";buyer_ceoname=""
                        if i in [1,3]:
                          supplier_name = memuser.biz_name
                          supplier_ceoname = memuser.ceo_name
                          buyer_name=td[8].get_attribute("innerText").replace("\'","")
                          buyer_ceoname = td[9].get_attribute("innerText")
                        else:
                          supplier_name=td[8].get_attribute("innerText").replace("\'","")
                          supplier_ceoname = td[9].get_attribute("innerText")
                          buyer_name =  memuser.biz_name
                          buyer_ceoname = memuser.ceo_name
                        sql = "Merge 전자세금계산서 as A Using (select  '"+memuser.seq_no+"' as seq_no,'"+approveNo+"' as 승인번호 ) as B "
                        sql += "On A.승인번호 = B.승인번호 and A.seq_no=B.seq_no when not matched then "
                        sql += "insert values('"
                        sql += memuser.seq_no+"','"
                        sql += memuser.biz_no+"','"
                        sql += str(i)+"','"
                        sql += td[2].get_attribute("innerText")+"','"
                        sql += approveNo+"','"
                        sql += td[3].get_attribute("innerText")+"','"
                        sql += td[4].get_attribute("innerText")+"','"
                        sql += td[6].get_attribute("innerText")[:12]+"','"
                        sql += supplier_name + "','"    #공급자상호
                        sql += supplier_ceoname + "','" #공급자성명
                        sql += "','"  #공급자주소
                        sql += td[5].get_attribute("innerText")[:15]+"','"
                        sql += buyer_name +"','"
                        sql += buyer_ceoname +"','"
                        sql += "','"
                        sql += kkdk +"','"  #합계
                        sql += kkke +"','"  #공급가액
                        sql += kkkt +"','"  #세액
                        sql += TIsep +"','" #전자세금계산서분류
                        sql += "','"        #전자세금계산서종류
                        sql += TIuh +"','"  #발급유형  인터넷/asp
                        sql += "','"        #비고
                        sql += receiveYN +"','"  #영수/청구
                        sql += eml_sup +"','"  #공급자이메일
                        sql += eml_buyer1 +"','"  #이메일
                        sql += "','"                                    #이메일
                        sql += "','"                                    #품목일자
                        sql += td[10].get_attribute("innerText").replace("\'","")+"','"   #품목명
                        sql += "','"                                    #품목규격
                        sql += "','"                                    #품목수량
                        sql += "0','"                                   #품목단가
                        sql += "0','"                                   #품목공급가액
                        sql += "0','"                                   #품목세액
                        sql += "','"                                   #품목비고
                        sql += today + "','"
                        sql += "','"
                        sql += "','"
                        sql += "N','','','');"
                        print(sql)
                        cursor = connection.cursor()
                        cursor.execute(sql)
                        cursor.commit()
                        print('디비저장성공')
                        #time.sleep(1)
                        strsql = "exec SP_세금계산서_Scrap '" + member[0] + "','" + today + "','"+approveNo+"' "
                        cursor.execute(strsql)    ;print('계정세팅완료')      

                      print("J:"+str(j)+", divmod:"+str(divmod(j,10)[1]))
                      if totalCnt>100 and divmod(j,10)[1]==0:
                        print(str(j)+"개 이후 넥스트버튼 ")
                        nextBtn = driver.find_element(By.XPATH,'//*[@id="mf_txppWframe_pglNavi_next_btn"]/a')
                        driver.implicitly_wait(10)
                        nextBtn.click()
                        print("넥스트 클릭")
                        driver.implicitly_wait(10)
                        time.sleep(1)
                      table=[];tbody=[]
                    driver.implicitly_wait(10)
                  else:
                    print('조회테이블이 없는 경우')
                    time.sleep(1)
                  # strsql = ""
                  driver.implicitly_wait(10)
            except Exception as e:
              print(e)
              strsql = "Merge 스크래핑관리 as A Using (select '"+today+"' as crt_dt,'"+member[0]+"' as seq_no) as B "
              strsql += "On A.crt_dt=B.crt_dt  and A.seqno_lastTry=B.seq_no "
              strsql += "when matched then update set seqno_lastTry='"+member[0]+"',bigo='수임해제됨' "
              strsql += "When Not Matched Then insert values('"+today+"','"+member[0]+"','"+memuser.biz_name+"','수임해제됨');"
              print(strsql)
              cursor = connection.cursor()
              cursor.execute(strsql)
              cursor.commit()   
            time.sleep(1)
          except Exception as e:
            print(e)
            return JsonResponse({'data':'실패:홈택스 내부조회 중 에러'},safe=False)  
        else:
          print(f'{member[1]} : 합계표와 세금계산서 일치')
  elif flag=='1':#1.전자합계표조회 
    # 세금계산서 합계와 다른 경우 => 무조건 조회
    # 홈택스 합계표는 당일 발행분은 하루 뒤에 조회된다!!
    strsql = " SELECT a.seq_no ,biz_name,biz_manager "
    strsql = " select a.seq_no,biz_Name,biz_manager,duzon_id,biz_no "
    strsql += " ,YN_1,YN_2,YN_3,YN_4,YN_5,YN_6,YN_7,YN_8,YN_9,YN_10,YN_11,YN_12,YN_13,YN_14,bigo, fiscalMM,biz_manager,ceo_name,biz_type "    #[6]->YN_1,[15]->YN_10, 22-> biz_manager        
    #strsql += " ,CASE WHEN  "
    #strsql += " isnull((select sum(합계금액) from 전자세금계산서 where seq_no=a.seq_no and " + workPeriod2 + "),0) =  "
    #strsql += " isnull((select sum(합계금액) from 전자세금계산서합계표 where seq_no=a.seq_no and work_yy="+flagYear+" and work_qt="+str(work_qt)+"),1)  "
    #strsql += " THEN  'True' else 'False'  END AS result "
    strsql += " from mem_user a, mem_deal b, tbl_mng_jaroe d where a.seq_no=b.seq_no and d.seq_no=b.seq_no and a.biz_no<>''  and biz_no<>'--' and keeping_yn='Y' and biz_manager<>'' "
    strsql += " and work_yy="+flagYear+" and work_mm='"+str(13+int(work_qt))+"' "
    if flagFirst=="관리자별":           strsql += " and biz_manager ='" + flagManagerName + "' "  
    elif flagFirst=="전체(기장만)" :    strsql += " and biz_manager  in ("+flagGroup+") " 
    elif flagFirst=="단일사업자":       strsql += " and a.seq_no ='"+flagSeqNo+"'   "  
    else:                              strsql += " and biz_manager not in ('종소세','종소세2','종소세3','오피스텔','환급1') " 
    # if flagFirst=="관리자별" or flagFirst[:2]=="전체":                                    #23.10.19 단체로 할때는 예정대 법인만 / 단일사업자는 예정신고가능
    #   if flagPeriod2=='1분기' or flagPeriod2=='3분기':strsql += " and biz_type<=3 "   

    print(strsql)
    cursor = connection.cursor()
    cursor.execute(strsql)
    members = cursor.fetchall()
    connection.commit()
    connection.close()   
    if members:
      driver = utils.conHometaxLogin(htxLoginID,False)
      driver.find_element(By.ID,'mf_wfHeader_textbox81212967').click();print('전체메뉴');time.sleep(1)
      driver.find_element(By.ID,'mf_txppWframe_tabControl8_tab_tabs_08_tabHTML').click();time.sleep(0.5);print('세무대리')
      driver.find_element(By.ID,'grpMenuAtag_48_4801100000').click();print('전자(세금)계산서 합계표 조회')   ;time.sleep(1)       
      for member in members:
        print('수임사업자전환:'+str(member[0]))
        time.sleep(3)#필수
        receivebtn=""
        try:
          receivebtn = driver.find_element(By.XPATH,'//*[@id="mf_txppWframe_trigger501"]')
          driver.implicitly_wait(10)
          if receivebtn:
            receivebtn.click();print('수임처버튼 전환 버튼 클릭')
          else:
            print('@@@!!!---수임처버튼 전환 버튼 사라짐---!!!@@@')   
            continue       
        except NoSuchElementException:
          print("수임처버튼 엘리먼트 선택 실패")
          continue
        print('수임처버튼 전환')

        time.sleep(1)
        print('수임납세자 사업자번호 선택')
        driver.find_element(By.ID,'mf_txppWframe_UTEETZZA21_wframe_selectbox5').click()
        driver.find_element(By.XPATH,r'/html/body/div[4]/div[2]/div[1]/div/div[1]/div[1]/div[2]/div[2]/div/div/dl[2]/dd/select/option[2]').click();print('사업자번호 선택')
        time.sleep(1.5)
        memuser = MemUser.objects.get(seq_no=member[0])
        bizno = memuser.biz_no
        strsql = "Merge 스크래핑관리 as A Using (select '"+today+"' as crt_dt,'"+member[0]+"' as seq_no) as B "
        strsql += "On A.crt_dt=B.crt_dt  and A.seqno_lastTry=B.seq_no "
        strsql += "when matched then update set seqno_lastTry='"+member[0]+"' "
        strsql += "When Not Matched Then insert values('"+today+"','"+member[0]+"','"+memuser.biz_name+"','');"
        print(strsql)
        cursor = connection.cursor()
        cursor.execute(strsql)
        cursor.commit()
        time.sleep(2)
        print('사업자번호전달:'+bizno)
        driver.find_element(By.ID,'mf_txppWframe_UTEETZZA21_wframe_txprDscmNoA').send_keys(bizno)
        bizChange = driver.find_element(By.ID,'mf_txppWframe_UTEETZZA21_wframe_trigger85');print('사업자번호선택 조회버튼')
        driver.implicitly_wait(10)
        try:
          bizChange.click()
          driver.implicitly_wait(10)
          try:
              # 조회결과가 없으면 다음 사업자로 넘어간다
              noResultMessage = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'mf_txppWframe_UTEETZZA21_wframe_grdResult_noResultMessageFrame_wq_uuid_14779')))
              print(f'{memuser.biz_name}은 수임동의 되어있지 않아 조회하지 않습니다')
              driver.find_element(By.ID,'mf_txppWframe_UTEETZZA21_wframe_btnClose').click()
              continue
          except Exception as e:
              driver.find_element(By.XPATH,r'/html/body/div[4]/div[2]/div[1]/div/div[1]/div[1]/div[2]/div[4]/div/div[1]/div/table/tbody/tr[1]/td[1]/label').click();print('업체선택라디오버튼')
              print('조회버튼')
              driver.implicitly_wait(10)
              driver.find_element(By.ID,'mf_txppWframe_UTEETZZA21_wframe_btnProcess').click()
              driver.implicitly_wait(10)
              print('조회버튼2')
              time.sleep(2)
              strdel = "delete from 전자세금계산서합계표 where seq_no='"+memuser.seq_no+"' and  work_yy= '"+flagYear+"' and work_qt='"+str(work_qt)+"'"
              print(strdel)
              cursor = connection.cursor()
              cursor.execute(strdel)
              cursor.commit()
              for i in [1, 2, 3, 4]:
                time.sleep(1)
                if i in [1,2]:  #  전자세금계산서 radioEtxivClsfCd_input_0/전자계산서 radioEtxivClsfCd_input_1 
                  vat = driver.find_element(By.CSS_SELECTOR,'#mf_txppWframe_radioEtxivClsfCd > div.w2radio_item.w2radio_item_0 > label');  driver.implicitly_wait(5);  print('전자세금계산서 조회시작');  vat.click(); print('전자세금계산서 조회버튼클릭')
                else:   #전자계산서 조회시작
                  nonVat = driver.find_element(By.CSS_SELECTOR,'#mf_txppWframe_radioEtxivClsfCd > div.w2radio_item.w2radio_item_1 > label');print('전자계산서 조회 드라이버 선택');driver.implicitly_wait(10); nonVat.click(); print('전자계산서 조회버튼 클릭')
                driver.implicitly_wait(3)
                if i in [1,3]:  #  매입 매출 선택( 매출:radio7_input_0  매입:radio7_input_1) ==> 세금계산서 조회와 다른 부분
                  try:
                    saleBtn = driver.find_element(By.CSS_SELECTOR,'#mf_txppWframe_radio7 > div.w2radio_item.w2radio_item_0 > label'); time.sleep(1.5)
                    if saleBtn: saleBtn.click() 
                  except NoSuchElementException:
                    time.sleep(2)
                    saleBtn = driver.find_element(By.CSS_SELECTOR,'#mf_txppWframe_radio7 > div.w2radio_item.w2radio_item_0 > label'); driver.implicitly_wait(10)
                    if saleBtn: saleBtn.click() 
                else:  #매입조회시작
                  costBtn = driver.find_element(By.CSS_SELECTOR,'#mf_txppWframe_radio7 > div.w2radio_item.w2radio_item_1 > label'); driver.implicitly_wait(10); time.sleep(1);  costBtn.click() 
                driver.implicitly_wait(10)

                #  조회기간 구분 선택(일자별:radio5_input_0      월별:radio5_input_1        분기별:radio5_input_2
                flagPeriod = "#mf_txppWframe_radio5 > div.w2radio_item.w2radio_item_2 > label"
                print("조회기간:" + flagPeriod)
                periodBtn = driver.find_element(By.CSS_SELECTOR,flagPeriod);driver.implicitly_wait(10);periodBtn.click();driver.implicitly_wait(5)

                print('조회기간 선택(연) : ' + flagYear)
                select = Select(driver.find_element(By.ID,'mf_txppWframe_selectbox6'));select.select_by_visible_text('{}'.format(workyear)); driver.implicitly_wait(5)
                #  조회기간 선택(과세기간)   /1기 예정/1기 확정/1기 (예정+확정)  /2기 예정/2기 확정/2기 (예정+확정)   
                print('조회기간 선택(분기) : ' + str(workPeriod)) 
                select = Select(driver.find_element(By.ID,'mf_txppWframe_selectbox8'));  select.select_by_visible_text('{}'.format(workPeriod));  driver.implicitly_wait(5)
                print('조회하기 버튼 클릭')
                searchBtn = driver.find_element(By.ID,'mf_txppWframe_trigger23');driver.implicitly_wait(10); searchBtn.click(); driver.implicitly_wait(10); time.sleep(1)

                table = driver.find_element(By.ID,'mf_txppWframe_group2215'); print("테이블획득");   driver.implicitly_wait(30)    
                tbody = table.find_element(By.ID,'mf_txppWframe_group2221');  print("티바디획득")
                tr = tbody.find_elements(By.TAG_NAME,'tr')[0]
                td = tr.find_elements(By.TAG_NAME,"td")
                trderCnt="0";totCnt="0";kkke="0";kkkt="0";kkdk="0"
                if i in [1,2]: 
                  trderCnt = td[0].get_attribute("innerText").replace(",","")
                  totCnt = td[1].get_attribute("innerText").replace(",","")
                  kkke = td[2].get_attribute("innerText").replace(",","")
                  kkkt = td[3].get_attribute("innerText").replace(",","")
                  kkdk = td[4].get_attribute("innerText").replace(",","")
                else:    
                  trderCt = td[0].get_attribute("innerText").replace(",","")    
                  totCnt = td[1].get_attribute("innerText").replace(",","")
                  kkdk = td[2].get_attribute("innerText").replace(",","")
                if trderCnt=="":trderCnt="0"
                if kkdk=="":kkdk="0"
                if kkke=="":kkke="0"
                if kkkt=="":kkkt="0"
                if totCnt=="":totCnt="0"            
                sql = "insert into 전자세금계산서합계표 values('"
                sql += memuser.seq_no+"','"
                sql += flagYear+"','"
                sql += str(work_qt)+"','"
                sql += str(i)+"',"  #매입매출구분
                sql += str(trderCnt)+","
                sql += str(kkdk)+","
                sql += str(kkke)+","
                sql += str(kkkt)+","
                sql += str(totCnt)+",'"
                sql += today+"');"
                print(sql)
                cursor = connection.cursor()
                cursor.execute(sql)
                cursor.commit()
                strsql = "Merge 스크래핑관리 as A Using (select '"+today+"' as crt_dt,'"+member[0]+"' as seq_no) as B "
                strsql += "On A.crt_dt=B.crt_dt  and A.seqno_lastTry=B.seq_no "
                strsql += "when matched then update set seqno_lastTry='"+member[0]+"' "
                strsql += "When Not Matched Then insert values('"+today+"','"+member[0]+"','"+memuser.biz_name+"','전자세금계산서합계표');"
                print(strsql)
                cursor = connection.cursor()
                cursor.execute(strsql)       
                print('디비저장성공')
                if i==1:# 매출세금계산서인경우
                  strMNG = "Merge  tbl_mng_jaroe as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(13+int(work_qt))+"' as work_mm) as B  "
                  strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_mm=B.work_mm "
                  strMNG += " when matched then update set YN_11='1'  "
                  strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','0','0','0','0','0','0','0','0','0','0','1','0','0','0',''); "  
                elif i==2:#매입세금계산서인경우  
                  strMNG = "Merge  tbl_mng_jaroe as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(13+int(work_qt))+"' as work_mm) as B  "
                  strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_mm=B.work_mm "
                  strMNG += " when matched then update set YN_12='1'  "
                  strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','0','0','0','0','0','0','0','0','0','0','0','1','0','0',''); "
                elif i==3:#매출계산서인경우  
                  strMNG = "Merge  tbl_mng_jaroe as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(13+int(work_qt))+"' as work_mm) as B  "
                  strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_mm=B.work_mm "
                  strMNG += " when matched then update set YN_13='1'  "
                  strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','0','0','0','0','0','0','0','0','0','0','0','0','1','0',''); "
                elif i==4:#매입계산서인경우  
                  strMNG = "Merge  tbl_mng_jaroe as A using(select '"+member[0]+"' as seq_no,'"+flagYear+"' as work_yy, '"+str(13+int(work_qt))+"' as work_mm) as B  "
                  strMNG += " on A.seq_no=B.seq_no  and A.work_YY=B.work_YY and A.work_mm=B.work_mm "
                  strMNG += " when matched then update set YN_14='1'  "
                  strMNG += " when not matched then insert values('"+member[0]+"','"+flagYear+"','"+str(13+int(work_qt))+"','0','0','0','0','0','0','0','0','0','0','0','0','0','1',''); " 
                print(strMNG)
                if member[16] !="2" and member[17]!="2" and member[18]!="2" and member[19]!="2":# tbl_mng_jaroe : member[16](세금계산서) 1은 조회여부 2는 데이터 유무 
                  cursor.execute(strMNG)   
                else:
                  print("전자세금계산서/계산서가 세무사랑에 업로드되어 tbl_mng_jaroe의 상태를 변경하지 않습니다.")                    
        except Exception:
          strsql = "Merge 스크래핑관리 as A Using (select '"+today+"' as crt_dt,'"+member[0]+"' as seq_no) as B "
          strsql += "On A.crt_dt=B.crt_dt  and A.seqno_lastTry=B.seq_no "
          strsql += "when matched then update set seqno_lastTry='"+member[0]+"',bigo='전자세금계산서합계표-수임해제됨' "
          strsql += "When Not Matched Then insert values('"+today+"','"+member[0]+"','"+memuser.biz_name+"','전자세금계산서합계표-수임해제됨');"
          print(strsql)
          cursor = connection.cursor()
          cursor.execute(strsql)
          # driver.find_element(By.ID,'btnClose').click();time.sleep(0.5)      
  print(f'정상종료 : {flagFirst}:{flagManagerName} - {flag} ' )
  return JsonResponse({'data':'성공'},safe=False)




# def Bank_Scrap(request):
#   bankCode = request.GET.get('bankCode',False)
#   # 1.크롬접속
#   driver = utils.ChromeDriver(False)  
#   strsql = "select seq_no,manage_no,bank_code,trim(bank_acct),trim(bank_pw),scrap_YN from scrap_bank_trader where bank_code='"+bankCode+"' and scrap_YN='Y'"
#   cursor = connection.cursor()
#   cursor.execute(strsql)
#   result = cursor.fetchall()
#   connection.commit()
#   connection.close()
#   for bank in result:
#     if bankCode=='0003':
#       driver.get('http://mybank.ibk.co.kr/uib/jsp/guest/qcs/qcs10/qcs1010/PQCS101000_i.jsp')
#       driver.implicitly_wait(10)
#       driver.find_element(By.ID,'in_cus_acn').send_keys(bank[3])
#       time.sleep(0.5)
#       acnt_pwd = driver.find_element(By.ID,'acnt_pwd')
#       time.sleep(0.5)
#       if acnt_pwd:
#         acnt_pwd.send_keys(' ')
#         time.sleep(0.3)
#         acnt_pwd.send_keys(bank[4]) 
          
#       time.sleep(0.3)
#       rnno = driver.find_element(By.ID,'rnno')
#       time.sleep(0.5)
#       if rnno:
#         rnno.send_keys(' ')
#         time.sleep(0.3)
#         rnno.send_keys(bank[1].replace("-","")[3:10])
#         time.sleep(0.3)
#       driver.execute_script("javascript:setDay('30');")
#       time.sleep(1)
#       driver.execute_script("javascript:uf_SaveToFile();")
#       time.sleep(1)
#       iframe = driver.find_element(By.ID ,'gridSaveFileiframe')
#       time.sleep(1)
#       driver.execute_script("javascript:com_saveExcelDataFile();")
#       print('저장버튼클릭')
#       time.sleep(1)
#       utils.DBSave_Downloaded_xlsx(driver,bank,'기업은행') 
#       time.sleep(1)
#     elif bankCode=='0081':
#       driver.get('http://www.kebhana.com/flex/quick/quickService.do?subMenu=1')
#       driver.implicitly_wait(10)
#       # acctNo = driver.find_element(By.ID,'acctNo')#계좌번호      document.querySelector("#acctNo")
#       time.sleep(1)
#       center = pyautogui.locateCenterOnScreen('K:/noa-django/Noa/static/assets/images/autowork/hanabank_title.png')
#       pyautogui.click(center)
#       pyautogui.press('tab',presses=14,interval=0.1)
#       pyautogui.write('')
#       pyautogui.write(bank[3],interval=0.1)
#       time.sleep(1)
#       pyautogui.press('tab')
#       pyautogui.write(bank[4])  #비번
#       time.sleep(0.1)
#       pyautogui.press('tab')
#       pyautogui.write(bank[1].replace("-","")[5:10]) #사업자번호뒷자리5개
#       time.sleep(0.1)      
#       driver.find_element(By.XPATH,'//*[@id="frmQuickInquiry"]/div[1]/table/tbody/tr[4]/td/div[1]/span[6]/a').click()#3개월버튼
#       time.sleep(0.5)
#       pyautogui.press('tab',presses=12,interval=0.1)
#       pyautogui.press('enter')
#       # driver.execute_script("javascript:pbk.quickService.acctInquiry.submitInquiry(document.forms['frmQuickInquiry']);")  #조회버튼
#       print('조회버튼클릭')
#       time.sleep(1)
#       driver.execute_script("javascript:pbk.quickService.acctInquiry.downloadExcel('01', '', 'deposit');") # 엑셀다운로드
#       print('엑셀다운클릭')
#       time.sleep(1)
#       utils.DBSave_Downloaded_xlsx(driver,bank,'은행') 
#       time.sleep(1)      
#   driver.quit()
#   print('정상종료')
#   return JsonResponse({'data':'성공 : 은행 빠른조회'},safe=False)