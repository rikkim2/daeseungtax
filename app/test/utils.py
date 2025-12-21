import pyodbc
import requests
import json
import urllib.parse
import http.client
import http.cookiejar
import urllib.request

import os,glob
import difflib
import time
import pyautogui
import pyperclip
import shutil
import PyPDF2
import img2pdf
import subprocess
import zipfile
import xlwt
import chromedriver_autoinstaller
import tkinter as tk
import re
import math
from tkinter import filedialog
from openpyxl import Workbook

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains

import natsort ## 숫자 정렬용 라이브러리
import pandas as pd

from datetime import datetime
from django.db import connection
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support import expected_conditions as EC
from pywinauto.keyboard import send_keys
from pywinauto import keyboard    # 단축키 기능을 이용하기 위해서
import pywinauto
from pywinauto import Desktop, Application

from selenium.common.exceptions import (
    TimeoutException, WebDriverException, NoSuchWindowException,NoSuchElementException,
    StaleElementReferenceException, UnexpectedAlertPresentException,    StaleElementReferenceException,
    ElementClickInterceptedException
)

from selenium.webdriver.common.keys import Keys
from django.core.exceptions import ObjectDoesNotExist
from app.models import MemUser
from app.models import MemAdmin
from app.models import MemDeal

from app.models import DsSlipledgr2
from django.db.models import Q
from selenium.webdriver.chrome.service import Service as ChromeService

def ChromeDriver(isInVisable):
    chrome_ver = chromedriver_autoinstaller.get_chrome_version()
    print(f'크롬 현재버전:{chrome_ver}')
    driver_path = ChromeDriverManager().install()
    print(f"driver_path:{driver_path}")

    option = Options()
    option.add_argument("--disable-proxy-certificate-handler")

    # ✅ 여기서 각종 '허용/차단' 알람을 안 뜨게 세팅
    prefs = {
        # 알림(Notification) 허용: 1=허용, 2=차단
        "profile.default_content_setting_values.notifications": 1,

        # 팝업/리다이렉트 허용
        "profile.default_content_setting_values.popups": 1,

        # 여러 파일 자동 다운로드 허용
        "profile.default_content_setting_values.automatic_downloads": 1,

        # 다운로드 시 매번 '저장' 물어보는 창 없애기
        "download.prompt_for_download": False,

        # 기타 안전 다운로드 옵션
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    }
    option.add_experimental_option("prefs", prefs)

    # 팝업 차단 기능 끄기
    option.add_argument("--disable-popup-blocking")

    if isInVisable:  # 창 없이 실행
        option.add_argument("--headless=new")
        option.add_argument("--no-sandbox")
        option.add_argument("--disable-dev-shm-usage")
        option.add_argument("--disable-gpu")

    correct_driver_path = os.path.join(os.path.dirname(driver_path), "chromedriver.exe")
    driver = webdriver.Chrome(service=Service(executable_path=correct_driver_path), options=option)

    # 마우스 이동
    width, height = pyautogui.size()
    pyautogui.moveTo(width/2, height/2)
    return driver



# 홈택스 인증서로그인  사용자 로그인
def conHometaxLogin_Personal(id,password,biz_no,biz_type,biz_name,ssn,isNotShowChrome):
  driver = ChromeDriver(isNotShowChrome)
  driver.get('http://www.hometax.go.kr/websquare/websquare_cdn.html?w2xPath=/ui/pp/index.xml');driver.implicitly_wait(1)
  win = pyautogui.getActiveWindow();win.maximize(); #윈도우 최대화
  #팝업제거
  # removeHometaxPopup(driver)  
  start_time = time.time()
  timeout = 61  # Set your timeout limit
  while True:
      try:
          # 타임아웃 확인 로그 출력
          print("남은시간:", time.time() - start_time)
  
          # 로그인 버튼 클릭
          WebDriverWait(driver, 61).until(
              EC.element_to_be_clickable((By.XPATH, r'//*[@id="mf_wfHeader_group1503"]'))
          ).click()
          print(f"{biz_name} - 로그인 버튼 클릭 성공")
          # 페이지가 열릴 때까지 기다림
          time.sleep(1)  # 짧은 지연을 추가하여 로딩을 확인
          iframe = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'mf_txppWframe')))
          driver.execute_script("arguments[0].style.display = 'block';", iframe)
          
          # 성공 시 while 루프 탈출
          print("페이지가 성공적으로 열렸습니다.")
          # '아이디 로그인' li 클릭
          time.sleep(1)
          try:
              # '로그인' 버튼을 클릭하고, 성공 시에만 login_with_retry 호출
              login_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'mf_txppWframe_anchor15')))
              login_button.click()
              # login_with_retry 함수 호출
              rstLogin = login_with_retry_NoCert(driver, id,password,ssn)
              if rstLogin:
                break
              else:#로그인 실패로 false 리턴
                return rstLogin
          except Exception as e:
              print(f"로그인 시도가 실패했습니다: {e}")   
              return False         
          
      except Exception as e:
          # 예외 발생 시 1초 대기 후 재시도
          time.sleep(1)
          # 타임아웃 체크
          if time.time() - start_time > timeout:
              print("타임아웃이 발생했습니다:", e)
              break

  print('개인 로그인 성공')
  popupCheck(driver)#==> 전자세금계산서용 팝업이 뜨는 경우에만 실행에러발생
  time.sleep(1)#필수
  if  biz_type>=4:#개인사업자인 경우
    try:
      driver.find_element(By.ID,'mf_wfHeader_group1508').click();time.sleep(2);print('최상단 사업장전환버튼')
      try:
        driver.find_element(By.XPATH,'/html/body/div[7]/div[2]/div[1]/div/div/div/div[4]/input[2]').click();time.sleep(0.5);print('확인버튼을 클릭하면 초기화면으로 이동합니다')
        popupCheck(driver)
      except:   
        driver.find_element(By.ID,'mf_wfHeader_UTXPPAAA24_wframe_iptBsno').clear();
        driver.find_element(By.ID,'mf_wfHeader_UTXPPAAA24_wframe_iptBsno').send_keys(biz_no.replace('-',''));time.sleep(0.5)
        pyautogui.press('tab',presses=2,interval=0.2);pyautogui.press('enter');time.sleep(0.5);print('조회하기 버튼')
        table = driver.find_element(By.ID,'mf_wfHeader_UTXPPAAA24_wframe_grid1_body_table')
        tbody = table.find_element(By.ID,'mf_wfHeader_UTXPPAAA24_wframe_grid1_body_tbody')  
        if len(tbody.find_elements(By.TAG_NAME,'tr')) >0 : 
          driver.find_element(By.XPATH,'//*[@id="mf_wfHeader_UTXPPAAA24_wframe_grid1_cell_0_4"]/nobr/button').click();time.sleep(0.5) 
          before = driver.window_handles
          print(f'사업장 변경 버튼 누르기전 창 개수 : {before}')
          pyautogui.press('enter');time.sleep(0.5);print('사업장 변경 확인알람 버튼')
          # 보안카드 팝업
          time.sleep(1)
          main = driver.window_handles
          print(f'사업장 변경 버튼 누른 후 창 개수 : {main}')
          if len(main) > len(before):
            driver.switch_to.window(main[1]); time.sleep(1)

            # mf_btnCncl 존재 여부 확인 후 클릭
            btns = driver.find_elements(By.ID, 'mf_btnCncl')
            if btns:
              try:
                WebDriverWait(driver, 5).until(
                  EC.element_to_be_clickable((By.ID, 'mf_btnCncl'))
                ).click()
                time.sleep(1.0); print('취소버튼')

                # 알럿이 뜨는 경우에만 처리
                try:
                  WebDriverWait(driver, 3).until(EC.alert_is_present())
                  Alert(driver).accept(); time.sleep(0.5)
                  print('알람 - 보안카드 인증을 취소하시겠습니까')
                except TimeoutException:
                  print('알럿 없음 (패스)')
              except Exception as e:
                print(f'취소버튼 클릭 실패: {e}')
            else:
              print('mf_btnCncl 버튼 없음 — 팝업 패스')

            main = driver.window_handles
            print(main)
            driver.switch_to.window(main[0])        
    except:
      print(biz_name+' : 사업장 고유아이디 ERROR')
  removeHometaxPopup(driver);time.sleep(1)
  return driver   

def popupCheck(driver, max_loops=5, click_wait=5, alert_wait=2):
    """
    전자세금계산서 팝업(새 창/탭) 정리 루틴
    - 다양한 버튼 셀렉터/텍스트 매칭/JS 클릭/ESC/강제 close까지 순차 시도
    - 알럿은 연속적으로 모두 수락
    - 메인 윈도 복귀 로직 보강
    """
    time.sleep(0.5)

    try:
        main = driver.current_window_handle
    except NoSuchWindowException:
        print("메인 창이 존재하지 않습니다.")
        return False

    def accept_all_alerts():
        """연속 알럿 모두 수락"""
        accepted = False
        for _ in range(5):
            try:
                WebDriverWait(driver, alert_wait).until(EC.alert_is_present())
                Alert(driver).accept()
                accepted = True
                time.sleep(0.15)
            except TimeoutException:
                break
            except WebDriverException:
                break
        return accepted

    def safe_switch(handle):
        """창 전환 안전 처리"""
        try:
            driver.switch_to.window(handle)
            return True
        except (NoSuchWindowException, StaleElementReferenceException, WebDriverException):
            return False

    def try_click_variants():
        """
        취소/닫기 류 버튼을 다양한 방법으로 클릭 시도.
        성공하면 True
        """
        # 1) 대표 id
        candidates = [
            (By.ID, 'mf_btnCncl'),
            # 2) 기존 CSS 후보
            (By.CSS_SELECTOR, 'body div.btn_wrap.border span input[type=button]'),
            (By.CSS_SELECTOR, 'input[type="button"][value*="취소"]'),
            (By.ID, 'trigger4'),
            # 3) 텍스트 기반 XPath (취소/닫기/확인 류)
            (By.XPATH, "//input[translate(@value,' ','')][contains(@value,'취소') or contains(@value,'닫기') or contains(@value,'Close')]"),
            (By.XPATH, "//button[contains(.,'취소') or contains(.,'닫기') or contains(.,'Close')]"),
            (By.XPATH, "//a[contains(.,'취소') or contains(.,'닫기')]"),
        ]

        # 클릭 가능 대기 후 일반 클릭 → JS 클릭 순서
        for by, sel in candidates:
            try:
                els = driver.find_elements(by, sel)
                if not els:
                    continue

                # element_to_be_clickable로 첫 요소를 기다려 클릭
                try:
                    btn = WebDriverWait(driver, click_wait).until(
                        EC.element_to_be_clickable((by, sel))
                    )
                    btn.click()
                    return True
                except (TimeoutException, WebDriverException, UnexpectedAlertPresentException):
                    # JS 클릭 백업
                    try:
                        driver.execute_script("arguments[0].click();", els[0])
                        return True
                    except WebDriverException:
                        continue
            except (StaleElementReferenceException, NoSuchWindowException):
                continue

        # ESC 키로 닫히는 유형
        try:
            driver.switch_to.active_element.send_keys(Keys.ESCAPE)
            time.sleep(0.15)
            return True
        except WebDriverException:
            pass

        return False

    for loop in range(max_loops):
        try:
            handles = [h for h in driver.window_handles if h != main]
        except WebDriverException:
            handles = []

        print(f'전자세금계산서 팝업개수: {len(handles)} (loop {loop+1}/{max_loops})')

        if not handles:
            print("팝업이 없습니다. 루프를 종료합니다.")
            return True

        before_count = len(handles)

        for h in list(handles):
            # 창 전환
            if not safe_switch(h):
                continue

            # 알럿이 먼저 떠 있는 경우 선 처리
            accept_all_alerts()

            clicked = False
            try:
                clicked = try_click_variants()
                if not clicked:
                    # 버튼 클릭이 전혀 안되면 강제 close 시도 (JS → window.close → 드라이버 close 순)
                    try:
                        driver.execute_script("window.close();")
                        time.sleep(0.1)
                    except WebDriverException:
                        try:
                            driver.close()
                        except WebDriverException:
                            pass
            except UnexpectedAlertPresentException:
                accept_all_alerts()
            except WebDriverException as e:
                print(f"팝업 처리 중 오류: {e}")

            # 클릭/닫기 이후에 추가 알럿 정리
            accept_all_alerts()

            # 항상 메인 복귀 시도
            try:
                driver.switch_to.window(main)
            except WebDriverException:
                try:
                    # 메인이 사라졌다면 남아 있는 첫 창으로 복귀
                    remain = driver.window_handles
                    if remain:
                        driver.switch_to.window(remain[0])
                        main = remain[0]  # 새로운 메인을 지정
                except Exception:
                    pass

        # 팝업 수 감소 체크: 줄었거나 0이면 일찍 종료
        try:
            after_count = len([h for h in driver.window_handles if h != main])
        except WebDriverException:
            after_count = 0

        if after_count == 0:
            print("모든 팝업이 닫혔습니다.")
            return True

        if after_count < before_count:
            # 닫히는 애니메이션/리로딩 여유
            time.sleep(0.4)
            continue

        # 변화 없을 때도 한 템포 두고 재시도
        time.sleep(0.4)

    print("경고: 최대 시도 후에도 팝업이 남아 있을 수 있습니다.")
    return False

#세무법인대승 인증서로그인
def conHometaxLogin(loginID,isNotShowChrome):
    context = {}
    # 1.크롬접속
    driver = ChromeDriver(isNotShowChrome)
  #try:
    driver.get('http://www.hometax.go.kr/websquare/websquare_cdn.html?w2xPath=/ui/pp/index.xml');driver.implicitly_wait(1)
    win = pyautogui.getActiveWindow();win.maximize(); #윈도우 최대화

    # 신고기간 신고안내 임시페이지 확인 - 바로가기
    # removeHometaxPopup(driver)   

    today = str(datetime.now())[:10].replace("-","")  
    today_mmdd = today[-4:]
    #종소세 신고기간  임시페이지 확인 - 바로가기
    if int(today_mmdd)>=501 and int(today_mmdd)<=531:
      try:
          baroGo_button = WebDriverWait(driver, 5).until(
              # EC.element_to_be_clickable((By.ID, 'TH4BOX_flag'))
              EC.element_to_be_clickable((By.XPATH, '//*[@id="TH4BOX"]/a/span[2]'))
          )
          baroGo_button.click()
          print('로그인 버튼 클릭 성공')
      except Exception as e:
          print(f"로그인 시도가 실패했습니다: {e}")    

    start_time = time.time()
    timeout = 61  # Set your timeout limit
    while True:
        try:
            # 타임아웃 확인 로그 출력
            print("남은시간:", time.time() - start_time) 

            # 로그인 버튼 클릭
            WebDriverWait(driver, 61).until(
                EC.element_to_be_clickable((By.ID, 'mf_wfHeader_group1503'))
            ).click()
            print("로그인 버튼 클릭 성공")

            # 페이지가 열릴 때까지 기다림
            time.sleep(1)  # 짧은 지연을 추가하여 로딩을 확인
            iframe = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, 'mf_txppWframe'))
            )
            driver.execute_script("arguments[0].style.display = 'block';", iframe)
            
            # 성공 시 while 루프 탈출
            print("페이지가 성공적으로 열렸습니다.")

            # '아이디 로그인' li 클릭
            time.sleep(1)
            try:
                # '로그인' 버튼을 클릭하고, 성공 시에만 login_with_retry 호출
                login_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, 'mf_txppWframe_anchor15'))
                )
                login_button.click()
                print('로그인 버튼 클릭 성공')

                # login_with_retry 함수 호출
                login_with_retry(driver, loginID)
                break
            except Exception as e:
                print(f"로그인 시도가 실패했습니다: {e}")            
            
        except Exception as e:
            # 예외 발생 시 1초 대기 후 재시도
            time.sleep(1)
            
            # 타임아웃 체크
            if time.time() - start_time > timeout:
                print("타임아웃이 발생했습니다:", e)
                break
    print('세무대리인 로그인 성공')
    time.sleep(2)#필수
    # 신고기간 신고안내 임시페이지 확인 - 바로가기
    removeHometaxPopup(driver);time.sleep(1)
   
    return driver  

# 홈택스 진입시 팝업제거
def removeHometaxPopup(driver):
  time.sleep(1.5)#필수
  main_window_handle = driver.current_window_handle
  popup_window_handle = None
  
  while not popup_window_handle:
    if len(driver.window_handles)>1:
      for handle in driver.window_handles:
          if handle != main_window_handle:
              popup_window_handle = handle
              # 팝업 윈도우로 전환
              driver.switch_to.window(popup_window_handle)
              # 팝업 윈도우에서 필요한 동작 수행
              try:
                group2152 = driver.find_element(By.CSS_SELECTOR,'body > div > div > div.btn_wrap.border > span > input[type=button]')
                group2152.click(); time.sleep(0.1)
              except:
                trigger4 = driver.find_elements(By.CSS_SELECTOR, '#mf_trigger4')
                if trigger4:
                    trigger4[0].click()
                    time.sleep(0.1)
                else:
                    # trigger4가 없으면 trigger1 클릭
                    trigger1 = driver.find_element(By.CSS_SELECTOR, '#mf_trigger1')
                    trigger1.click()
                    time.sleep(0.1)
              # 메인 윈도우로 전환
              driver.switch_to.window(main_window_handle)                
              break
    else:
      popup_window_handle = True 


def login_with_retry(driver, loginID):
    try:
        # 로그인 시도
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'mf_txppWframe_iptUserId'))).send_keys(loginID)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'mf_txppWframe_iptUserPw'))).send_keys('daeseung@1128')
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'mf_txppWframe_group91882177'))).click()
        print('로그인 버튼 클릭')
        time.sleep(1)
        driver.find_element(By.ID, 'mf_txppWframe_UTXPPABC12_wframe_iptUserJuminNo1').send_keys('751123')
        driver.find_element(By.ID, 'mf_txppWframe_UTXPPABC12_wframe_iptUserJuminNo2').send_keys('1')
        driver.find_element(By.ID, 'mf_txppWframe_UTXPPABC12_wframe_trigger46').click()

        # 인증서 선택을 위한 iframe 전환
        iframe = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#dscert')))
        driver.execute_script("arguments[0].style.display = 'block';", iframe)
        driver.switch_to.frame(iframe)
        main_window_handle = driver.current_window_handle
        print('원격조정 허용/차단')
        time.sleep(1)
        pyautogui.press('tab',presses=2,interval=0.25)
        pyautogui.press('enter');time.sleep(0.15)
        # 인증서 선택 및 비밀번호 입력
        time.sleep(2)
        pyautogui.press('tab',presses=4,interval=0.25)
        pyautogui.write('daeseung@1203');time.sleep(0.25)
        pyautogui.press('enter');time.sleep(0.15)
        print('공인인증서 로그인')
        time.sleep(1)
        # pyautogui.press('tab',presses=4,interval=0.2)
        pyautogui.press('enter');time.sleep(0.1)
        driver.switch_to.window(main_window_handle) 
        driver.find_element(By.ID,'mf_txppWframe_input1').send_keys('P30447');time.sleep(0.25)
        driver.find_element(By.ID,'mf_txppWframe_input2').send_keys('daeseung@1128');time.sleep(0.25)
        WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.ID, 'mf_txppWframe_group1669'))).click()
    except UnexpectedAlertPresentException as e:
        # 예외 발생 시 알림창 닫기
        alert = driver.switch_to.alert
        alert_text = alert.text
        print(f"Unexpected alert present: {alert_text}")
        alert.accept()  # 알림창 확인(닫기)
        
        # 재시도: ID, 비밀번호 다시 입력
        print("재시도: ID 및 비밀번호 재입력")
        login_with_retry(driver, loginID)

def _debug_dump(driver, tag=""):
    """현재 컨텍스트(창/URL/iframe 등) 덤프"""
    try:
        print(f"\n[DEBUG]{'['+tag+']' if tag else ''}")
        try:
            print("  handles:", driver.window_handles)
            print("  current:", driver.current_window_handle)
        except Exception:
            pass
        try:
            print("  title  :", driver.title)
        except Exception:
            pass
        try:
            print("  url    :", driver.current_url)
        except Exception:
            pass
        try:
            iframes = driver.find_elements(By.TAG_NAME, "iframe")
            print("  iframes:", len(iframes))
        except Exception:
            pass
    except Exception:
        pass


def _accept_alert_if_present(driver, timeout=2, tag=""):
    """알럿 있으면 accept"""
    try:
        WebDriverWait(driver, timeout).until(EC.alert_is_present())
        al = driver.switch_to.alert
        txt = al.text
        print(f"[ALERT]{'['+tag+']' if tag else ''} {txt}")
        al.accept()
        return True
    except Exception:
        return False


def _click_in_popup_with_iframe_fallback(driver, locators, timeout=6, tag=""):
    """
    팝업 현재 window에서:
    1) default content에서 locators 순서대로 클릭 시도
    2) 실패하면 iframe들을 순회하며 클릭 시도
    locators: [(By.ID,"..."), (By.CSS_SELECTOR,"..."), ...]
    """
    # 1) 기본 DOM
    driver.switch_to.default_content()
    for by, sel in locators:
        try:
            el = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, sel)))
            el.click()
            print(f"[POPUP]{'['+tag+']' if tag else ''} click OK -> ({by}, {sel}) in default_content")
            return True
        except Exception:
            pass

    # 2) iframe 순회
    driver.switch_to.default_content()
    frames = driver.find_elements(By.TAG_NAME, "iframe")
    for i, fr in enumerate(frames):
        try:
            driver.switch_to.default_content()
            driver.switch_to.frame(fr)
            for by, sel in locators:
                try:
                    el = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, sel)))
                    el.click()
                    print(f"[POPUP]{'['+tag+']' if tag else ''} click OK -> ({by}, {sel}) in iframe[{i}]")
                    driver.switch_to.default_content()
                    return True
                except Exception:
                    pass
        except Exception:
            continue

    driver.switch_to.default_content()
    print(f"[POPUP]{'['+tag+']' if tag else ''} click FAIL: no locator matched")
    return False

def fast_click_cancel(driver):
    driver.switch_to.default_content()

    # 1️⃣ 가장 유력한 iframe 하나만 바로 시도
    frames = driver.find_elements(By.TAG_NAME, "iframe")
    if frames:
        try:
            driver.switch_to.frame(frames[0])
            driver.find_element(By.ID, "mf_btnCncl").click()
            driver.switch_to.default_content()
            return True
        except Exception:
            driver.switch_to.default_content()

    # 2️⃣ default DOM 시도
    try:
        driver.find_element(By.ID, "mf_btnCncl").click()
        return True
    except Exception:
        return False
def login_with_retry_NoCert(driver, loginID, loginPW, ssn):
    """
    - 로그인 시도
    - 보안인증 팝업이 뜨면: 팝업 핸들 diff로 정확히 잡고, 취소 버튼 클릭(iframe fallback 포함), 알럿 accept, 원래 창으로 복귀
    - 팝업이 안 뜨면: 주민번호 입력 후 로그인 트리거 클릭(있을 때만)
    - 성공 True / 실패 None
    """
    try:
        wait = WebDriverWait(driver, 10)

        # 1) ID/PW 입력
        wait.until(EC.element_to_be_clickable((By.ID, "mf_txppWframe_iptUserId"))).clear()
        wait.until(EC.element_to_be_clickable((By.ID, "mf_txppWframe_iptUserId"))).send_keys(loginID)

        wait.until(EC.element_to_be_clickable((By.ID, "mf_txppWframe_iptUserPw"))).clear()
        wait.until(EC.element_to_be_clickable((By.ID, "mf_txppWframe_iptUserPw"))).send_keys(loginPW)

        # 2) 로그인 버튼 클릭 전 window handle 스냅샷
        before_handles = set(driver.window_handles)
        main_handle = driver.current_window_handle

        # 로그인 버튼 클릭
        wait.until(EC.element_to_be_clickable((By.ID, "mf_txppWframe_group91882177"))).click()
        time.sleep(0.3)

        # 3) 팝업 발생 여부 체크 (짧게)
        popup_handle = None
        try:
            WebDriverWait(driver, 4).until(lambda d: len(set(d.window_handles) - before_handles) > 0)
            diff = list(set(driver.window_handles) - before_handles)
            popup_handle = diff[0] if diff else None
        except TimeoutException:
            popup_handle = None

        if popup_handle:
            print("보안인증 팝업이 열렸습니다.")
            driver.switch_to.window(popup_handle)
            _debug_dump(driver, "POPUP_OPENED")

            # (A) 취소 버튼 후보들 (id가 바뀌는 경우 대비해서 추가)
            cancel_locators = [
                (By.ID, "mf_btnCncl"),
                (By.CSS_SELECTOR, "button[id='mf_btnCncl']"),
                # 텍스트 기반(사이트마다 다름): 필요시 아래 XPath를 커스터마이즈
                (By.XPATH, "//button[contains(., '취소') or contains(., '닫기')]"),
                (By.XPATH, "//*[self::a or self::button][contains(., '취소') or contains(., '닫기')]"),
            ]

            
            if not fast_click_cancel(driver):
                _click_in_popup_with_iframe_fallback(
                    driver,
                    cancel_locators,
                    timeout=4,   # 6 → 4로 축소
                    tag="CANCEL"
                )            
                # 디버깅 정보 더 출력
                _debug_dump(driver, "POPUP_CANCEL_NOT_FOUND")

            # 취소 후 알럿(보안카드 인증 취소?) 처리
            _accept_alert_if_present(driver, timeout=4, tag="SECURITY_CANCEL")

            # 팝업이 닫혔을 수도 있으니 핸들 정리
            try:
                # 팝업 닫기 시도(닫혀 있으면 예외)
                driver.close()
            except Exception:
                pass

            # 메인으로 복귀
            try:
                driver.switch_to.window(main_handle)
            except Exception:
                # 혹시 main_handle이 유효하지 않으면 첫번째로 복귀
                handles = driver.window_handles
                if handles:
                    driver.switch_to.window(handles[0])

            _debug_dump(driver, "BACK_TO_MAIN")
            return True  # ✅ 팝업 처리까지 완료했으면 성공 플래그

        else:
            print("팝업이 열리지 않았습니다.")
            print("로그인 버튼 클릭 -> 주민번호 입력 시도")
            _debug_dump(driver, "NO_POPUP")

            # 4) 주민번호 입력(있으면)
            try:
                j1 = driver.find_element(By.ID, "mf_txppWframe_UTXPPABC12_wframe_iptUserJuminNo1")
                j2 = driver.find_element(By.ID, "mf_txppWframe_UTXPPABC12_wframe_iptUserJuminNo2")
                j1.clear(); j2.clear()
                j1.send_keys(ssn[:6])
                j2.send_keys(ssn[6:])
                driver.find_element(By.ID, "mf_txppWframe_UTXPPABC12_wframe_trigger46").click()
            except NoSuchElementException:
                print("주민번호 입력창이 없어 실행하지 않음")

            return True  # ✅ 여기까지 오면 일단 성공 플래그

    except UnexpectedAlertPresentException:
        # 예기치 않은 알럿 닫기
        try:
            alert = driver.switch_to.alert
            alert_text = alert.text
            print(f"Unexpected alert present: {alert_text}")
            alert.accept()
        except Exception:
            pass
        _debug_dump(driver, "EX_UNEXPECTED_ALERT")
        print("로그인 실패(UnexpectedAlertPresentException): None 리턴")
        return None

    except TimeoutException as e:
        _debug_dump(driver, "EX_TIMEOUT")
        print(f"로그인 실패(Timeout): {e}")
        return None

    except Exception as e:
        _debug_dump(driver, "EX_OTHER")
        print(f"로그인 실패(기타 예외): {e}")
        return None

def conWetaxLogin(isNotShowChrome):
    context = {}

    driver = ChromeDriver(False)   
    driver.get('https://www.wetax.go.kr/login.do');time.sleep(1)
    win = pyautogui.getActiveWindow();win.maximize();time.sleep(0.5) #윈도우 최대화
    driver.find_element(By.XPATH,'//*[@id="contentsWrap"]/div[2]/div[1]/a[1]/p').click();time.sleep(0.5)
    # driver.find_element(By.XPATH,r'//*[@id="contents"]/div[2]/div[1]/a[1]').click();    time.sleep(1)
    # idLoginBtn = driver.find_element(By.XPATH,'//*[@id="tab-1"]/div[1]/button[2]');    time.sleep(1);    idLoginBtn.click();    time.sleep(3)
    loginPass = 'daeseung@1203'
    #인증서버튼
    # certBtn = driver.find_element(By.XPATH,r'//*[@id="tab-1_2"]/div/div[1]/a[1]').click();    time.sleep(3)
    certBtn = driver.find_element(By.ID,'btnCertLogin').click();    time.sleep(3)
    passwd = 'daeseung@1203'
    print('공인인증서 로그인')
    handle = find_window_with_retry('전자 서명 작성', '#32770', 10, 1) 
    # handle = pywinauto.findwindows.find_windows(title='전자 서명 작성', class_name='#32770')[0]
    app = Application().connect(handle=handle)
    w_open = app.window(handle=handle)
    w_open.Edit.type_keys(passwd)
    w_open['확인Button'].click()
    time.sleep(2.5)
    driver.find_element(By.CSS_SELECTOR,'#header > div.header-menu > div > div.access-change > div > button').click();time.sleep(0.5)
    driver.find_element(By.CSS_SELECTOR,'#header > div.header-menu > div > div.access-change.on > div > ul > li > button:nth-child(2)').click();time.sleep(0.5)

    return driver  

#세무포털 로그인
def conSemuportalLogin():
  driver = ChromeDriver(False)
  driver.get('https://www.semuportal.com/bizon/index.do');  driver.implicitly_wait(3)	
                                    
  # 팝업뜨는 경우 없애기
  # 현재 윈도우 핸들 가져오기
  main_window_handle = driver.current_window_handle

  # 버튼 클릭 등으로 새로운 팝업 윈도우가 열리는 이벤트 발생
  # 팝업 윈도우 핸들 가져오기
  popup_window_handle = None
  while not popup_window_handle:
    if len(driver.window_handles)>1:
      for handle in driver.window_handles:
          if handle != main_window_handle:
              popup_window_handle = handle
              # 팝업 윈도우로 전환
              driver.switch_to.window(popup_window_handle)
              # 팝업 윈도우에서 필요한 동작 수행
              try:
                group2152 = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '#boardType01 > tbody > tr > td > a > img'))
                )
                group2152.click()
              except Exception as e:              
                print("Group2152 element not found:", e)
                try:
                    trigger4 = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, 'body > div > form > div > div.popup_close > ul.pop_close_btn.fr > a > span'))
                    )
                    trigger4.click()
                except Exception as e:
                    print("Trigger4 element not found:", e)
              # 메인 윈도우로 전환
              driver.switch_to.window(main_window_handle)                
              break
    else:
      popup_window_handle = True  
  driver.find_element(By.XPATH,'/html/body/div/div[1]/div[1]/ul[2]/li[1]/a').click();time.sleep(0.5)
  driver.find_element(By.ID,'userId').send_keys('simplebook');time.sleep(0.5)
  driver.find_element(By.ID,'userPassword').send_keys('just1928');time.sleep(0.5)
  driver.find_element(By.XPATH,'//*[@id="form1"]/ul[4]/button[1]').click();time.sleep(1.5)

  # 팝업뜨는 경우 없애기
  # 현재 윈도우 핸들 가져오기
  main_window_handle = driver.current_window_handle
  # 팝업 윈도우 핸들 가져오기
  popup_window_handle = None
  while not popup_window_handle:
    if len(driver.window_handles)>1:
      for handle in driver.window_handles:
          if handle != main_window_handle:
              popup_window_handle = handle
              driver.switch_to.window(popup_window_handle)
              # x 버튼 클릭
              driver.find_element(By.XPATH,'/html/body/div/form/div/div[2]/ul[2]/a/span').click();time.sleep(0.5) 
              driver.switch_to.window(main_window_handle)                
              break
    else:
      popup_window_handle = True

  #포커스 가져오기
  procs = pywinauto.findwindows.find_elements();handle=''
  for proc in procs: 
    # print(proc.name)
    if proc.name[:10]== '한국세무사회전산법인':handle = proc.handle;break
  app = Application().connect(handle=handle)
  w_open = app.window(handle=handle)
  w_open.set_focus()  

  return driver
	

# 고용토탈 사무대행기관 인증서로그인까지
def conTotalComwelLogin(bizno):
  driver = ChromeDriver(False)
  driver.get('http://total.comwel.or.kr');  time.sleep(3)  
  procs = pywinauto.findwindows.find_elements();handle=''
  for proc in procs: 
    if proc.name[:6]== '근로복지공단':handle = proc.handle;break
  app = Application().connect(handle=handle)
  w_open = app.window(handle=handle)
  w_open.set_focus()    
  loginbtn = driver.find_element(By.ID,'mf_wfm_header_btn_login').is_displayed();  time.sleep(0.5)
  if loginbtn:
    driver.find_element(By.ID,'mf_wfm_header_btn_login').click();time.sleep(0.5);print('id클릭')
    
  if bizno=='220-85-33586':
    bizno = bizno.split('-')
    driver.find_element(By.ID,'mf_wfm_content_TabContent1_tab_tabs3_tabHTML').click();    time.sleep(0.5);    print('사무대행탭클릭')
    loginbtn = driver.find_element(By.XPATH,'//*[@id="mf_wfm_content_u_group_fg3"]/span[2]/label').click();    time.sleep(0.5);    print('사업장명의인증서 클릭')  
    driver.find_element(By.ID,'mf_wfm_content_drno3_1').send_keys(bizno[0])
    driver.find_element(By.ID,'mf_wfm_content_drno3_2').send_keys(bizno[1])
    driver.find_element(By.ID,'mf_wfm_content_drno3_3').send_keys(bizno[2])
    driver.implicitly_wait(5);time.sleep(1)
    driver.find_element(By.ID,'mf_wfm_content_wq_uuid_638').click() ;    driver.implicitly_wait(5);time.sleep(1)
    driver.find_element(By.XPATH,'//*[@id="xwup_media_hdd"]').click();    driver.implicitly_wait(5);time.sleep(1);    print('하드디스크선택')
    # 테이블내 인증서찾기 개선필요
    driver.find_element(By.XPATH,'//*[@id="xwup_cert_table"]/table/tbody/tr[1]/td[2]/div').click()
    passwd = 'daeseung@1203'
    loginbtn = driver.find_element(By.XPATH,'//*[@id="xwup_certselect_tek_input1"]').is_displayed()
    if loginbtn:
      print('인증서암호')
      driver.find_element(By.XPATH,'//*[@id="xwup_certselect_tek_input1"]').send_keys(passwd)
    driver.implicitly_wait(5)
    driver.find_element(By.ID,'xwup_OkButton').click()
    print('공인인증서 로그인')
    time.sleep(1)  
    try:  
      notToday = driver.find_element(By.ID,'mf_wfm_content_samuInfoPopup_wframe_chk_today').is_displayed()
      driver.find_element(By.ID,'mf_wfm_content_samuInfoPopup_wframe_chk_today').click()  ;time.sleep(1) ;print('오늘 그만열기 해제')
    except:
      print('오늘 그만열기 없음')
  return driver

def find_window_with_retry(title, class_name, max_retries=10, retry_interval=1):
    for _ in range(max_retries):
        handle_list = pywinauto.findwindows.find_windows(title=title, class_name=class_name)
        if handle_list:
            return handle_list[0]
        print("재시도 : "+title)
        print(_)
        time.sleep(retry_interval)
    return None  # 창이 나타나지 않을 경우 None 반환

# 홈택스등 팝업(새 창/새 탭)이 떠도 메인으로 다시 포커스 돌리는 함수
def focus_main_window(driver, main_handle):
    # 새로 열린 팝업 창이 있으면 모두 닫고 메인으로
    for handle in driver.window_handles:
        if handle != main_handle:
            driver.switch_to.window(handle)
            try:
                driver.close()
            except Exception:
                pass

    # 메인 핸들로 복귀
    driver.switch_to.window(main_handle)
    driver.switch_to.default_content()  # iframe 안에 들어갔다면 기본 컨텐츠로

#웹상 버튼 클릭 3회시도하고

#웹상 버튼 클릭 3회시도하고
def click_button_3attempt(driver,BTN_ID, timeout=30):
    wait = WebDriverWait(driver, timeout)

    for attempt in range(3):
        try:
            # 1) 버튼이 클릭 가능해질 때까지 대기
            btn = wait.until(
                EC.element_to_be_clickable((By.ID, BTN_ID))
            )

            # 2) 화면 중앙으로 스크롤
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", btn
            )

            # 3) 먼저 일반 click 시도
            try:
                btn.click()
            except (ElementClickInterceptedException) as e:
                # 가려져 있을 때 JS click fallback
                print(f"{BTN_ID} 클릭 실패 → JS 클릭 시도: {e}")
                driver.execute_script("arguments[0].click();", btn)

            print("[{BTN_ID}] 버튼 클릭 완료")
            break

        except StaleElementReferenceException as e:
            print(f"[{BTN_ID}] StaleElementReference, 재시도 {attempt+1}/3: {e}")
            if attempt == 2:
                raise

        except TimeoutException:
            raise TimeoutException("[{BTN_ID}] 버튼을 찾거나 클릭할 수 없습니다.")

    # 4) 클릭 후 후속 상태 대기 (예: 로딩 레이어 사라짐, 다음 요소 등장 등)
    #    홈택스에서 클릭 후 뜨는 특정 요소가 있다면 여기에 맞춰서 변경
    try:
        wait.until(
            EC.invisibility_of_element_located(
                (By.ID, "mf_wq_uuid_24_mf_wq_uuid_24_wq_proessMsgModal")
            )
        )
        print("[{BTN_ID}] 처리 완료 (로딩 레이어 종료 확인)")
    except TimeoutException:
        print("[{BTN_ID}] 로딩 레이어 종료 대기 타임아웃 (무시 가능)")

#    OS 파일 선택창이 실제로 떴는지 pyautogui로 대략 감지
#    (Windows 기준: 창 제목에 '열기'가 포함되는 경우가 많음)
def wait_file_dialog_open(title_keyword="열기", timeout=10):
    start = time.time()
    while time.time() - start < timeout:
        win = pyautogui.getActiveWindow()
        if win and title_keyword in win.title:
            print("[FILE] 파일 선택창 활성화 확인:", win.title)
            return True
        time.sleep(0.2)
    print("[FILE] 파일 선택창이 열리지 않음")
    return False

#이미 검증이 완료된 자료가 존재합니다.' 웹스퀘어 모달을 감지해서 닫아주고 닫았으면 True, 안 뜬 경우 False 반환
def Htx_Modal_Clear(driver, modalTxt,timeout=5):
    wait = WebDriverWait(driver, timeout)

    try:
        # 1) 해당 문구를 포함하는 요소 찾기
        msg_el = wait.until(
            EC.visibility_of_element_located((
                By.XPATH,
                f"//*[contains(text(),'{modalTxt}')]"
            ))
        )
        print("[MODAL] 이미 검증이 완료된 자료 모달 발견")

        # 2) 같은 모달 영역 안의 '확인' / '예' 버튼 추정해서 클릭
        #    (홈택스 웹스퀘어 팝업 구조에 맞춰 약간 조정 가능)
        ok_btn = msg_el.find_element(
            By.XPATH,
            ".//ancestor::div[contains(@class,'w2window') or contains(@class,'w2group')]"
            "//input[@type='button' or self::a][contains(@value,'확인') or contains(.,'확인') or contains(@value,'예')]"
        )
        ok_btn.click()
        print("[MODAL] 확인 버튼 클릭")

        # 3) 모달이 사라질 때까지 잠깐 대기
        wait.until(
            EC.invisibility_of_element_located((
                By.XPATH,
                "//*[contains(text(),'이미 검증이 완료된 자료가 존재합니다')]"
            ))
        )
        print("[MODAL] 모달 사라짐")
        return True

    except TimeoutException:
        print("[MODAL] '이미 검증이 완료된...' 모달이 안 보임")
        return False

# 홈택스 출력팝업 - 민원서류
def Htx_Popup_Print_Minwon(driver,printBtnName,fileName,directory,directPrint):
  driver.find_element(By.XPATH,printBtnName).click();time.sleep(0.25)
  WebDriverWait(driver, 30).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(0.25)
  time.sleep(3);print('일괄출력 후 6초대기'+printBtnName)
  main = driver.window_handles;        print(main);        driver.switch_to.window(main[1]) 
  driver.find_element(By.XPATH,'/html/body/div/div/div[1]/table/tbody/tr/td/div/nobr/button[6]').click();time.sleep(3);print('최초인쇄')
  # driver.find_element(By.XPATH,"/html/body/div/div[2]/div/button[1]").click();time.sleep(3);print('최종인쇄')
  # main = driver.window_handles;    print(main);    driver.switch_to.window(main[2]); time.sleep(5);print('필수대기시간')
  # pyautogui.moveTo(0,0)
  # pyautogui.click(x=580,y=100);time.sleep(0.5);    
  pyautogui.press('enter');time.sleep(1);print('인쇄버튼')
  createDirectory(directory)    
  fullFileName = directory +"/"+ fileName +".pdf"
  if  os.path.isfile(fullFileName.replace('/',"\\")):os.remove(fullFileName.replace('/',"\\"));print(fullFileName+" 삭제완료") ;time.sleep(1)
  Report_save(fileName,directory);print('파일저장완료');time.sleep(1)
  pyautogui.hotkey('alt','f4')
  try:
    main = driver.window_handles;    print(main);    driver.switch_to.window(main[0]); 
  except Exception as e:
    print('에러 발생:', e)
  # try:
  #   iframe = driver.find_element(By.CSS_SELECTOR,'#txppIframe');      driver.switch_to.frame(iframe);      time.sleep(1); print('iframe 전환txppIframe')
  # except Exception as e:
  #   print('에러 발생:', e)

  return 1


# 홈택스 출력팝업1
def Htx_Popup_Print(driver,printBtnName,fileName,directory,directPrint):
  if printBtnName[0:3]=='//*':
    driver.find_element(By.XPATH,printBtnName).click();time.sleep(0.25)
    if printBtnName!='//*[@id="mf_txppWframe_grdList_cell_0_16"]/nobr/button' :
      try:
        WebDriverWait(driver, 3).until(EC.alert_is_present());al = Alert(driver);al.accept();time.sleep(0.25)
      except:
        print('알람 발생하지 않음')
  else:
    try:
        print('관련버튼ID:'+printBtnName)
        saleBtn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, printBtnName)))
        print('Htx_Popup_Print '+printBtnName + " trying")
        # driver.find_element(By.ID,printBtnName).click()
        element = driver.find_element(By.ID, printBtnName)
        driver.execute_script("arguments[0].click();", element)
    except NoSuchElementException:
        print(f"{printBtnName} 요소를 찾을 수 없습니다.")
        return
    except TimeoutException:
        print(f"{printBtnName} 요소가 클릭 가능한 상태가 되지 않았습니다.")
        return
  print('일괄출력 후 6초대기'+printBtnName);time.sleep(6)
  if not directPrint:
    main = driver.window_handles;        print(main);        driver.switch_to.window(main[1]) 
    # 출력버튼 클릭
    button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '/html/body/div/div/div[1]/table/tbody/tr/td/div/nobr/button[6]'))
    ).click();time.sleep(1.5);print('최초인쇄')
    driver.find_element(By.XPATH,"/html/body/div/div[2]/div/button[1]").click();time.sleep(3);print('최종인쇄')
    start_time = time.time()
    while True:
        window_handles = driver.window_handles
        if len(window_handles) >= 3: break
        elif time.time() - start_time > 30:
            raise Exception("인쇄창이 열리는데 30초 초과 에러")
        time.sleep(0.5)  # Wait for 0.5 seconds before checking again
    # main = driver.window_handles;    print(main);    driver.switch_to.window(main[2]); 
    time.sleep(3);print('필수대기시간')
    pyautogui.moveTo(0,0)
    pyautogui.click(x=580,y=100);time.sleep(0.5);    
    pyautogui.press('enter');time.sleep(1);print('인쇄버튼')
    createDirectory(directory)    
    Report_save(fileName,directory);print('파일저장완료')

  else:#일반출력창이 아닌 경우(프로세스가 멈춰버린다)
    main = driver.window_handles;       driver.switch_to.window(main[1]);   print(main);time.sleep(0.5)
    # 캡처할 팝업창의 좌표와 크기 설정
    x = 10
    y = 100
    width = 830
    height = 900
    # 스크린샷 캡처
    pyautogui.screenshot('popup.png', region=(x, y, width, height))
    # 이미지를 PDF로 변환하여 저장
    with open(directory+"/"+fileName+".pdf", "wb") as f:
        f.write(img2pdf.convert('popup.png'))
    print("변환완료")

  # 창 닫기
  main = driver.window_handles
  driver.switch_to.window(main[1])  # 닫을 팝업 창으로 전환
  driver.execute_script("window.close();")
  time.sleep(1)

  # 원래 창으로 다시 전환
  driver.switch_to.window(main[0])
  print("✅ 팝업 창 닫기 완료 및 메인 창으로 복귀")

  return 1


#다운된 파일 이름바꿔저장하기
def FileSave_Downloaded_PDF(directory, file_Origin_name, file_Purpose_name, seq_no):
    try:
        # 사용자 및 거래 정보 가져오기
        memuser = MemUser.objects.get(seq_no=seq_no)
        memdeal = MemDeal.objects.get(seq_no=str(memuser.seq_no).strip())

        # 다운로드 폴더에서 가장 최근 파일 찾기 (file_Origin_name이 비어 있을 경우)
        if not file_Origin_name:
            downloads = "C:\\Users\\Administrator\\Downloads"
            try:
                file_Origin_name = max(
                    [os.path.join(downloads, f) for f in os.listdir(downloads) if os.path.isfile(os.path.join(downloads, f))],
                    key=os.path.getctime
                )
                print(f'가장 최근 다운로드된 파일: {file_Origin_name}')
            except ValueError:
                print("다운로드 폴더에 파일이 없습니다.")
                return 0  # 파일이 없으면 종료

        # 저장 경로 설정
        savePath = directory.replace('/', "\\")
        if memdeal.biz_manager == '화물':
            savePath = directory.replace('D:\\', "D:\\화물연대\\")

        print(f"저장 경로: {savePath}")

        # 디렉토리 존재 여부 확인 및 생성
        if not os.path.isdir(savePath):
            os.makedirs(savePath, exist_ok=True)

        # 같은 파일이 존재하는 경우 새로운 이름 설정
        destination_path = os.path.join(savePath, file_Purpose_name)
        if os.path.exists(destination_path):
            base, ext = os.path.splitext(file_Purpose_name)
            counter = 1
            while os.path.exists(destination_path):
                destination_path = os.path.join(savePath, f"{base}_{counter}{ext}")
                counter += 1

        # 파일 이동
        shutil.move(file_Origin_name, destination_path)
        print(f'파일 저장 성공: {destination_path}')
        return 1

    except ObjectDoesNotExist as e:
        print(f"데이터베이스 조회 오류: {e}")
        return 0
    except Exception as e:
        print(f"예외 발생: {e}")
        return 0

# 디렉토리생성
def createDirectory(directory):
  try:
    if not os.path.exists(directory):
      os.makedirs(directory)
      return "none"
    else:
      file_list = os.listdir(directory)
      return file_list
  except OSError:
    print("Error: 폴더생성실패 - "+directory)

def Report_save(fileName,directory):  
  pyautogui.hotkey('alt','d')
  pyperclip.copy(directory.replace('/',"\\"));   time.sleep(0.25);pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  # pyperclip.copy("c:/".replace('/',"\\"));    time.sleep(0.25);pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  pyautogui.press('enter');time.sleep(0.5);pyautogui.press('tab',presses=3,interval=0.5)
  pyautogui.hotkey('alt','n')
  pyperclip.copy(fileName);    time.sleep(0.5)
  pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  pyautogui.hotkey('alt','s');    time.sleep(5)  
  return 1    

def delete_CorpSheet_998_files(flag,directory):
  delFile = ['1','3','4','3-1','3-2','3-4','15','47-갑','48','50-갑','92-지-43-2','92-지-43','101','102','103','104','105','106','300']
  if flag=="2": delFile = ['1','40-6','40-7','40-8','40-9','46','47','69','71','71-1','73','91-1','91-2','91-2-2','101','102','103','104','105','106','300']
  if os.path.exists(directory):
    for file in os.scandir(directory):
      title = file.name.split(".")[0]
      if title  in delFile: os.remove(file.path);print('지운파일:'+file.path)
    print('법인세 결산서 삭제완료')
  else:
    print('Directory Not Found')  

def delete_CorpSheet_files(flag,directory):
  notDel = ['0','19-갑','58','98','99','100']
  if os.path.exists(directory):
    if os.path.isfile(directory+"/Thumbs.db"): os.remove(directory+"/Thumbs.db")
    for file in os.scandir(directory):
      print(file.name)
      title = file.name.split(".")[0]
      titleDetail = "110"
      if title.find("-")!=-1: #신고서외 보조파일인 경우
        titleDetail = title.split("-")[0]
        if title not in notDel and int(titleDetail)<110: os.remove(file.path);print('지운파일:'+file.path)
      elif title not in notDel and int(title)<110: os.remove(file.path);print('지운파일:'+file.path)
      if title=='300':os.remove(file.path)
    print('법인세 신고서 삭제완료')
  else:
    print('Directory Not Found')  
# 폴더내 파일 전체 삭제
def delete_all_files(download_path):
  if os.path.exists(download_path):
    for file in os.scandir(download_path):
      os.remove(file.path)
    print('Remove All File')
  else:
    print('Directory Not Found')

# 경로의 파일 갯수 체크
def dir_search(self, download_path):
    file_len = len(os.listdir(download_path))
    print('파일갯수: ' + str(file_len))
    return file_len
# 경로내의 지정된 단어를 가진 파일리스트 리턴
def find_target_files(folder_path,txtTarget):
    if not os.path.exists(folder_path):
        raise ValueError(f"{folder_path} : 경로가 존재하지 않습니다")

    target_files = []
    search_pattern = os.path.join(folder_path, f'*{txtTarget}*')
    for root, _, files in os.walk(folder_path):
        for file_name in files:
            if txtTarget in file_name:
              target_files.append([root, file_name])
                # target_files.append(os.path.join(root, file_name))

    return target_files


# 다운로드
def download_excel(self, download_path):
    driver = self.driver
    # 폴더가 없으면 생성
    if os.path.isdir(download_path) == False:
        os.mkdir(download_path)
    else:  print(f'이미 폴더가 있습니다.')
    # 폴더안에 이미 존재하는 파일 삭제
    print(f'{self.delete_all_files(download_path)}')
    # 폴더안의 파일 갯수 (다운로드 확인에 사용)
    before_file_len = self.dir_search(download_path)
    # 다운로드 버튼 클릭 (실패 시 무한반복)
    downloaded = False
    while downloaded == False:
        try:
            # 다운로드 클릭
            download_button = driver.find_element(By.CSS_SELECTOR, 'div[onclick*="javascript:doExcelDownload()"]')
            download_button.click()
            # 확인 클릭
            ok_button = driver.find_element(By.CSS_SELECTOR, '#button-1006')
            ok_button.click()
            downloaded = True
        except Exception as e:
            downloaded = False
    # 다운로드 완료 여부를 5초에 한번씩 체크
    after_file_len = self.dir_search(download_path)
    while before_file_len == after_file_len:
        print('대기중...')
        after_file_len = self.dir_search(download_path)
        time.sleep(5)
    try: # 가장 최근에 다운받은 파일 이름을 변경
        new_file_name = max([download_path + "\\" + f for f in os.listdir(download_path)],key=os.path.getctime)
        shutil.move(new_file_name,os.path.join(download_path,r"myfile.xlsx"))
        new_file_name = download_path + "\\" + r"myfile.xlsx"
        print(f'{new_file_name}')
    except Exception as e:
        print(e)
        self.log_append(f'{download_path} 파일 찾기 실패')
        print('파일명 변경 실패')

    return new_file_name
def check_alert(driver):
    from selenium.common.exceptions import NoAlertPresentException
    try:
        driver.switch_to.alert
        return True
    except NoAlertPresentException:
        return False
def PDF_Merge(fileName,directory):
  merger = PyPDF2.PdfMerger()
  filst = os.listdir(directory)
  for file in natsort.natsorted(filst):
    merger.append(os.path.join(directory,file))
  merger.write(f"{directory}/{fileName}.pdf")
  merger.close()
  print(directory+" 폴더에 "+fileName+".PDF로 병합 성공")
  return 1

def safe_list_assign(lst, index, value):
    """Expands the list and assigns value at index, safely avoiding IndexError."""
    while len(lst) <= index:
        lst.append(None)  # Or use a default value suitable for your logic
    lst[index] = value
def unicode_slice(s, start, end):
    """Slices a Unicode string from start to end, considering Hangul character width."""
    sliced = ""
    count = 0
    adding = False

    for char in s:
        prev_count = count
        if not is_hangul(char):
            count += 1
        else:
            count += 2

        if not adding and count >= start:
            adding = True
            # Adjust if start is within the "visual space" of a Hangul character
            if prev_count < start and is_hangul(char):
                continue

        if adding:
            sliced += char
            if count >= end:
                break

    return sliced

def is_hangul(char):
    """Determines if a character is a Hangul character."""
    cp = ord(char)
    return 0xAC00 <= cp <= 0xD7AF


def find_similar_strings(matched_files, target_string):
  for file_name in matched_files:
    if file_name.find(target_string) != -1:
      print(file_name)
      print(target_string)      
      return True
  return False
def get_quarter(strmonth):
  month = int(strmonth)
  if month in [1, 2, 3]:
      return '1'
  elif month in [4, 5, 6]:
      return '2'
  elif month in [7, 8, 9]:
      return '3'
  elif month in [10, 11, 12]:
      return '4'
  else:
      return None
  

  
def is_program_running(process_name):
  windows = Desktop(backend='uia').windows()
  for window in windows:
    # print(window.window_text())
    if window.window_text()==process_name:
      return True
  return False

def semusarang_LaunchProgram_App(strID):
  global isSemusarangOn
  print('semusarang_LaunchProgram')
  app = None
  program_path = "C:\\NewGen\\Rebirth\\Rebirth.exe"

  # 프로그램 실행 여부 확인
  if is_program_running("세무사랑 Pro"):
    print('이미 실행 중인 경우, 프로그램에 연결')
    app = Application(backend='uia').connect(path=program_path)
    subprocess.Popen(r'C:/NewGen/Rebirth/Rebirth.exe');  time.sleep(2)
    dig = app['세무사랑 Pro']
  else:
    print('실행되지 않은 경우, 프로그램을 실행')
    app = Application(backend='uia').start(program_path)  ;time.sleep(5)
    dig = app['세무사랑 Pro']
    # Edit 컨트롤이 활성화될 때까지 대기 (예: 최대 10초 동안 대기)
    timeout = 10  # 대기 시간 초 단위로 설정
    start_time = time.time()
    while time.time() - start_time < timeout:
      if dig.Edit.exists():  break
      print(time.time() - start_time)
      time.sleep(1)  # 1초 대기

    # Edit 컨트롤이 활성화되었는지 확인
    if dig.Edit.exists():
      # Edit 컨트롤에 엔터 키 입력
      dig.Edit.type_keys('{ENTER}')  #Edit 속성은 맨위로 가게한다(서버-1)
      # dig.type_keys(strID)  
      dig.type_keys('{ENTER}')    
      dig.type_keys('1122')  ;dig.type_keys('{ENTER}') 
      dig.type_keys('102')  ;dig.type_keys('{ENTER}') ;  time.sleep(1)
      dig.type_keys('112837');dig.type_keys('{ENTER}')  ;time.sleep(2)   # dig['Button4'].click()  이것도 가능
      if pyautogui.locateOnScreen('K:/noa-django/Noa/static/assets/images/autowork/semusarang_diff_Update.png',confidence=0.8): 
        pyautogui.press('enter');print('서버와 업데이트일자가 다릅니다') 
      time.sleep(3)
      isSemusarangOn = True
    else:
      print("Edit 컨트롤이 활성화되지 않았습니다.")
      isSemusarangOn = False
  return dig

def semusarang_LaunchProgram(strID):
  global isSemusarangOn
  print('semusarang_LaunchProgram')
  subprocess.Popen(r'C:/NewGen/Rebirth/Rebirth.exe');  time.sleep(3)
  print('세무사랑ID가 이미 등록된 경우는 아래 3개를 지워야해')
  pyautogui.write(' ');  time.sleep(1)
  pyautogui.write(strID);  time.sleep(1)
  pyautogui.press('enter');  time.sleep(1)
  pyautogui.write('1122');  time.sleep(0.25)
  pyautogui.press('enter');  time.sleep(0.25) 
  pyautogui.press('f2')
  isSemusarangOn = True
  return isSemusarangOn
def semusarang_ChangeCompany_ID(DuzonID):
  print('회사변경')
  time.sleep(0.5) 
  pyautogui.press('f3');  time.sleep(1)  
  print(DuzonID)
  pyautogui.write(str(DuzonID));  time.sleep(1)  
  if len(str(DuzonID))==3:pyautogui.press('enter');  time.sleep(1.5)#필수 
  if str(DuzonID)=='102' or str(DuzonID)=='2078' or str(DuzonID)=='1507' or str(DuzonID)=='1508':
    print('비번입력')      
    pyautogui.write('112837');    time.sleep(0.1)
    pyautogui.press('enter');    time.sleep(2) #필수
  pyautogui.press('enter');  time.sleep(1.5)#필수 
  return 1  
def semusarang_ChangeCompany_ID_App(dig,DuzonID):
  print('회사변경')
  time.sleep(0.5) 
  # pyautogui.press('f5');  time.sleep(1)  
  dig['사업자등록번호'].click();time.sleep(0.5)
  keyboard.send_keys(DuzonID) 
  if len(str(DuzonID))==3:keyboard.send_keys('{ENTER}');  time.sleep(1.5)#필수 
  if str(DuzonID)=='102' or str(DuzonID)=='2078' or str(DuzonID)=='1507' or str(DuzonID)=='1508':
    print('비번입력')      
    keyboard.send_keys('112837');    time.sleep(0.1)
    keyboard.send_keys('{ENTER}');    time.sleep(2) #필수
  keyboard.send_keys('{ENTER}');  time.sleep(2)#필수 
  keyboard.send_keys('{ENTER}');print('업데이트일자가 다릅니다. 팝업제거')
  return 1  

def semusarang_ChangeCompany(bizNo):
  print('회사변경')
  time.sleep(0.5) 
  pyautogui.press('f3');  time.sleep(1)  
  pyautogui.press('f2');  time.sleep(0.5)  
  bizNo = bizNo.replace('-','')
  pyautogui.write(str(bizNo));  time.sleep(1)  
  pyautogui.press('enter');  time.sleep(1.5)#필수  
  if str(bizNo)=='2208533586' or str(bizNo)=='1448127161':
    print('비번입력')      
    pyautogui.write('112837');    time.sleep(0.1)
    pyautogui.press('enter');    time.sleep(2) #필수
  return 1
def semusarang_ChangeFiscalYearAll(Kisu,strYear):
  print('회계 기수변경')
  time.sleep(1)
  pyautogui.press('enter')
  pyautogui.press('f4') ; print('f4')
  time.sleep(0.5)
  print('기수 : '+ Kisu)
  pyautogui.write(Kisu);  time.sleep(0.5) 
  pyautogui.press('enter');time.sleep(0.25)
  pyautogui.press('down');print('연도세팅 : '+strYear)
  pyautogui.write(strYear);    time.sleep(2)#필수 
  pyautogui.press('tab')
  time.sleep(1.5)     
  return 1

def semusarang_ChangeFiscalYear(flag,strYear):
  time.sleep(2)
  if flag=='insa':
    print('인사급여 기수변경')
    pyautogui.press('enter')
    pyautogui.press('f4');time.sleep(1); print('f4')
    print('다운버튼 1회')
    pyautogui.press('down');    time.sleep(0.25) 
    print('연도세팅 : '+strYear)
    pyautogui.write(strYear);    time.sleep(2)#필수 
  else:
    print('회계 기수변경')
    time.sleep(1)
    pyautogui.press('enter')
    pyautogui.press('f4') ; print('f4')
    time.sleep(1)
    print('기수 : '+ strYear)
    pyautogui.write(strYear)
    time.sleep(0.5) 
    pyautogui.press('enter')
    time.sleep(2)#필수 
  pyautogui.press('tab')
  time.sleep(1.5)     
  return 1

def semusarang_ChangeFiscalYear_App(flag, finalKi, fiscalMM):
    # finalKi를 정수로 변환(문자/숫자 혼용 대비)
    try:
        ki_int = int(str(finalKi).strip())
    except Exception:
        ki_int = int(finalKi)

    # 회계기간 종료월(fiscalMM) 초과 시 +1 (단, 인사급여(ins a) 는 제외)
    if datetime.now().month > int(fiscalMM) and flag != 'insa':
        ki_int += 1
    finalKi_str = str(ki_int)

    time.sleep(2)
    if flag == 'insa':
        print('인사급여 기수변경'); time.sleep(1)
        keyboard.send_keys('{ENTER}')
        keyboard.send_keys('{F4}'); time.sleep(1); print('다운버튼 1회')
        keyboard.send_keys('{DOWN}'); time.sleep(0.25); print('연도세팅 : ' + finalKi_str)
        keyboard.send_keys(finalKi_str); time.sleep(2)  # 필수
    else:
        print('회계 기수변경'); time.sleep(1)
        keyboard.send_keys('{F2}')
        keyboard.send_keys('{ENTER}')
        keyboard.send_keys('{F4}'); time.sleep(1)

        print('기수 : ' + finalKi_str)
        keyboard.send_keys(finalKi_str); time.sleep(0.5)
        keyboard.send_keys('{ENTER}'); time.sleep(2)  # 필수

    keyboard.send_keys('{TAB}'); time.sleep(1.5)
    return 1

def semusarang_Login(semusarangID,duzon_id,biz_no,reg_date,workyear,fiscalMM):
  dig = semusarang_LaunchProgram_App(semusarangID)
  if duzon_id.strip()=='1':  semusarang_ChangeCompany(biz_no.replace("-",""))
  else:               semusarang_ChangeCompany_ID_App(dig,duzon_id.strip())          
  finalKi = workyear - reg_date.year + 1    
  semusarang_ChangeFiscalYear_App('vat',str(finalKi),fiscalMM)   
  return True

def semusarang_Print(fileName,directory):
  createDirectory(directory)    
  fullFileName = directory+"/"+fileName +".pdf"
  if  os.path.isfile(fullFileName.replace('/',"\\")):os.remove(fullFileName.replace('/',"\\"));print(fullFileName+" 삭제완료") 
  pyautogui.press('f9');time.sleep(3); print(fileName + '출력버튼') #낮은사양에서는 필수
  pyautogui.hotkey('alt','1');time.sleep(3)    
  pyperclip.copy(fileName);    time.sleep(0.5)
  pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  pyautogui.hotkey('alt','d');    time.sleep(0.25)
  pyperclip.copy(directory.replace('/',"\\"))
  pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  pyautogui.hotkey('alt','s');    time.sleep(2)
  pyautogui.press('esc',presses=4,interval=0.25);print('열린 세무사랑 하부메뉴 닫기');    time.sleep(0.5)        
  return 1 
def semusarang_Print_Sheet(AppTitle,txtSelect,fileName_Month,directory):
  fileName = fileName_Month + txtSelect[:5]+".pdf"
  fullFileName = directory+"/"+fileName_Month + txtSelect[:5]+".pdf"
  if  os.path.isfile(fullFileName.replace('/',"\\")):os.remove(fullFileName.replace('/',"\\"));print(fullFileName+" 삭제완료") 

  pyautogui.press('f9');time.sleep(3); print(fileName + ' 출력') #낮은사양에서는 필수
  try:
    handle = pywinauto.findwindows.find_windows(title=AppTitle, class_name='#32770')[0]
  except:
    SET_FOCUS(AppTitle)
    pyautogui.press('f9');time.sleep(3); print(fileName + ' 출력') #낮은사양에서는 필수
    handle = pywinauto.findwindows.find_windows(title=AppTitle, class_name='#32770')[0]
  app = Application().connect(handle=handle)
  w_open = app.window(handle=handle)
  w_open.ComboBox3.Select(txtSelect);time.sleep(1)    

  pyautogui.hotkey('alt','1');time.sleep(3)    
  pyperclip.copy(fileName);    time.sleep(0.5)
  pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  pyautogui.hotkey('alt','d');    time.sleep(0.25)
  pyperclip.copy(directory.replace('/',"\\"))
  pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
  pyautogui.hotkey('alt','s');    time.sleep(2)
  #pyautogui.press('esc');print('열린 세무사랑 하부메뉴 닫기');    => 저장후 출력창은 바로 닫힌다        
  return 1 


def semusarang_Print_F11(fileName,directory):
    pyautogui.press('f11');time.sleep(3); print(fileName + '출력버튼') #낮은사양에서는 필수
    pyautogui.hotkey('alt','1');time.sleep(3)    
    pyperclip.copy(fileName);    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
    pyautogui.hotkey('alt','d');    time.sleep(0.25)
    pyperclip.copy(directory.replace('/',"\\"))
    pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
    pyautogui.hotkey('alt','s');    time.sleep(2)
    pyautogui.press('esc',presses=4,interval=0.25);print('열린 세무사랑 하부메뉴 닫기');    time.sleep(0.5)        
    return 1  
def semusarang_Print_TI_Summary(fileName,directory):
    createDirectory(directory)
    if fileName=='2.pdf' or fileName=='4.pdf':
      center = pyautogui.locateCenterOnScreen('K:/noa-django/Noa/static/assets/images/autowork/vatCostSummation.png',confidence=0.7)
      #center = pyautogui.locateCenterOnScreen('K:/noa-django/Noa/static/assets/images/autowork/vatCostSummation.png',confidence=0.7)
      if center:pyautogui.click(center);    time.sleep(1) 
    pyautogui.hotkey('ctrl','f11');time.sleep(3); print(fileName + '출력버튼') #낮은사양에서는 필수
    pyautogui.hotkey('alt','1');time.sleep(3)    
    pyperclip.copy(fileName);    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
    pyautogui.hotkey('alt','d');    time.sleep(0.25)
    pyperclip.copy(directory.replace('/',"\\"))
    pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
    pyautogui.hotkey('alt','s');    time.sleep(5)     
    return 1 
def semusarang_Print_NotExit(fileName,directory):
    pyautogui.press('f9');time.sleep(3); print('출력버튼') #낮은사양에서는 필수
    center = pyautogui.locateCenterOnScreen('K:/noa-django/Noa/static/assets/images/autowork/semusarang_PDF_save.png')
    pyautogui.click(center)
    time.sleep(0.5) 
    pyautogui.press('tab',presses=2,interval=0.25)
    pyautogui.press('enter');print('찾기버튼');    time.sleep(0.5)      
    pyperclip.copy(fileName);    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
    pyautogui.hotkey('alt','d');    time.sleep(0.25)
    pyperclip.copy(directory.replace('/',"\\"))
    pyautogui.hotkey('ctrl', 'v');    time.sleep(0.5)
    pyautogui.hotkey('alt','s');    time.sleep(2)
    pyautogui.press('tab',presses=2,interval=0.25)
    pyautogui.press('enter');print('확인버튼')  ;    time.sleep(2)   
    pyautogui.hotkey('alt','f4');print('열린 PDF닫기');    time.sleep(1)     
    return 1 
def semusarang_Excelclose_Openmenuclose():
  time.sleep(2)
  pyautogui.hotkey('alt','f4')
  time.sleep(1)
  pyautogui.press('esc',presses=3, interval=1) 
def semusarang_Menu_Popup_App(menuName):
  time.sleep(1)
  keyboard.send_keys('^{ENTER}');time.sleep(0.5);pyperclip.copy(menuName)    #메뉴검색
  keyboard.send_keys('^v');time.sleep(0.5);keyboard.send_keys('{ENTER}');  time.sleep(1)
  return 1

#세무사랑 메뉴띄우기
def semusarang_Menu_Popup(menuName):
  time.sleep(1)
  keyboard.send_keys('^{ENTER}');time.sleep(0.5) #메뉴검색
  pyperclip.copy(menuName);  keyboard.send_keys('^v');time.sleep(1.5)
  pyautogui.press('enter');time.sleep(2)   
  if menuName in ('간이지급명세서전자신고','신용카드매출전표등수령명세서(갑)(을)','지급명세서전자신고(연말정산)','원천징수이행상황신고서','거래처등록','접대비(기업업무추진비)조정명세','지출증명서류수취검토서식','지방소득세과세표준및세액조정계','재고자산(유가증권)평가조정','미상각분감가상각비'):
    pyautogui.press('enter');time.sleep(5)  
  if menuName not in ('이익잉여금처분계산서','표준원가명세서'):
    intlen = len(menuName)
    procs = pywinauto.findwindows.find_elements();handle=''
    for proc in procs: 
      if proc.name[:intlen]== menuName:handle = proc.handle;break
    app = Application().connect(handle=handle)
    w_open = app.window(handle=handle)
    w_open.set_focus()  
  return 1

#세무사랑 메뉴띄우기 - 메뉴코드 - 원본
# def semusarang_MenuCode_Popup(menuCode,menuName):
#   time.sleep(2)
#   keyboard.send_keys('^{ENTER}');time.sleep(0.5) #메뉴검색
#   pyautogui.write(menuCode)
#   pyautogui.press('enter');time.sleep(2)   

#   intlen = len(menuName)
#   procs = pywinauto.findwindows.find_elements();handle=''
#   for proc in procs: 
#     if proc.name[:intlen]== menuName:handle = proc.handle;break
#   app = Application().connect(handle=handle)
#   w_open = app.window(handle=handle)
#   w_open.set_focus()  
#   return 1

def semusarang_MenuCode_Popup(menuCode, menuName):
    def open_menu():
        keyboard.send_keys('^{ENTER}')  # 메뉴 검색창 열기
        time.sleep(0.5)
        pyautogui.write(menuCode)
        pyautogui.press('enter')
        time.sleep(2)

    def find_window_handle():
        intlen = len(menuName)
        for proc in pywinauto.findwindows.find_elements():
            if proc.name[:intlen] == menuName:
                return proc.handle
        return None

    time.sleep(1)
    open_menu()

    handle = find_window_handle()
    if handle is None:
        print(f"🔄 '{menuName}' 창을 찾을 수 없어 메뉴 재시도...")
        open_menu()
        handle = find_window_handle()

    if handle is None:
        raise Exception(f"❌ '{menuName}' 창을 찾을 수 없습니다.")

    if isinstance(handle, str):
        handle = int(handle, 16) if handle.startswith('0x') else int(handle)

    app = Application(backend="uia").connect(handle=handle)
    w_open = app.window(handle=handle)
    w_open.set_focus()

    return 1

def SET_FOCUS(menuName):
  intlen = len(menuName)
  procs = pywinauto.findwindows.find_elements();handle=''
  for proc in procs: 
    if proc.name[:intlen]== menuName:handle = proc.handle;break
  app = Application().connect(handle=handle)
  w_open = app.window(handle=handle)
  w_open.set_focus()  
  return True

import psutil
def is_App_running(exe_name):
    """해당 exe가 실행 중인지 확인"""
    for proc in psutil.process_iter(['name']):
        if proc.info['name'] and proc.info['name'].lower() == exe_name.lower():
            return True
    return False

#디비데이터로 엑셀파일 만들어 driver에 저장하기
def MakeExcel_FromDB(driver,manageNo,flag):
  totalPath = driver
  wb = Workbook()  # 새 워크북 만들기  
  ws = wb.active   # 워크북의 첫 번째 워크시트 가져오기
  fileName = f"{flag}.xlsx" 
  if flag=="엑셀자료 일반전표전송":
    ws['A1'] = flag
    ws['O3'] = manageNo[1]
    ws['M4'] = 'v1.2'           
    MIDTITLE = ["년도월일","구분","코드","계정과목","코드","거래처명","코드","적요명","금액"]   
    for i, value in enumerate(MIDTITLE):
      ws.cell(row=10, column=i+1, value=value)    
    strsql = "select Work_YY+Tran_Dt YYMMDD,CrDr,Acnt_cd,Acnt_Nm,Trader_Code,Trader_Name,'' remk_cd,Remk,TranAmt "
    strsql += f"from DS_SlipLedgr_Ecount a where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and left(Tran_Stat,4)<>'매입전표' and left(Tran_Stat,4)<>'매출전표' order by Tran_Dt,slip_no,prt_ord "
    print(strsql)
    cursor = connection.cursor()
    cursor.execute(strsql)
    results = cursor.fetchall()
    connection.commit()
    if results:
      for i, row in enumerate(results):
        for j, value in enumerate(row):
          ws.cell(row=i+11, column=j+1, value=value)
  elif flag=="신용카드 매입 자료 입력":
    ws['A1'] = flag
    ws['c4'] = manageNo[1]       
    MIDTITLE = ["카드종류","신용카드사명","신용카드번호","승인일자","사업자등록번호","거래처명","거래처유형","공급가액","세액","봉사료","합계금액","부가세공제여부","부가세유형","품목"]   
    for i, value in enumerate(MIDTITLE):
      ws.cell(row=8, column=i+1, value=value) 
    #--카드전표: 매입전표(상대계정이 하나인 경우) 
    strsql = f"select '3', trader_name,trader_bizno,work_yy+tran_dt,"
    strsql +=f"(select case when trader_bizno='' then Trader_Code_ORG else Trader_Bizno end from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and acnt_cd<>253 and Acnt_cd<>135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"(select case when trader_name='' then Trader_name_ORG else Trader_name end from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and acnt_cd<>253 and Acnt_cd<>135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"'일반',"
    strsql +=f"(select tranamt_cr from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"(select tranamt from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd=135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"'0',TranAmt,'공제','57',"
    strsql +=f"(select acnt_cd from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"(select remk from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ) "
    strsql +=f"from DS_SlipLedgr_Ecount a where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and left(Tran_Stat,4)='매입전표' and acnt_cd=253"
    strsql +=f"and (select COUNT(slip_no) from DS_SlipLedgr_Ecount where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt )>2"
    strsql +=f" union all "
    #--카드전표: 일반전표(상대계정이 하나인 경우)
    strsql +=f"select '3', trader_name,trader_bizno,work_yy+tran_dt,"
    strsql +=f"(select case when trader_bizno='' then Trader_Code_ORG else Trader_Bizno end from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and acnt_cd<>253 and Acnt_cd<>135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"(select case when trader_name='' then Trader_name_ORG else Trader_name end from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and acnt_cd<>253 and Acnt_cd<>135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"'일반',"
    strsql +=f"(select tranamt_cr from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"'0','0',TranAmt,'불공제','',"
    strsql +=f"(select acnt_cd from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"(select remk from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt )"
    strsql +=f"from DS_SlipLedgr_Ecount a where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and left(Tran_Stat,4)='매입전표' and acnt_cd=253 "
    strsql +=f"and (select COUNT(slip_no) from DS_SlipLedgr_Ecount where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt)=2    "
    strsql +=f" union all "
    #--카드전표: 일반전표(상대계정이 둘 이상인 경우)
    strsql +=f"select '3', trader_name,trader_bizno,work_yy+tran_dt,"
    strsql +=f"(select case when trader_bizno='' then Trader_Code_ORG else Trader_Bizno end from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and acnt_cd<>253 and Acnt_cd<>135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"(select case when trader_name='' then Trader_name_ORG else Trader_name end from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and acnt_cd<>253 and Acnt_cd<>135 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ),"
    strsql +=f"'일반',"
    strsql +=f"(select tranamt_cr from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ), "
    strsql +=f"'0','0',TranAmt,'불공제','',"
    strsql +=f"(select acnt_cd from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ), "
    strsql +=f"(select remk from  DS_SlipLedgr_Ecount where left(Tran_Stat,4)='매입전표' and  Acnt_cd<>135 and Acnt_cd<>253 and Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt ) "
    strsql +=f"from DS_SlipLedgr_Ecount a where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and left(Tran_Stat,4)='매입전표' and acnt_cd=253 "
    strsql +=f"and (select COUNT(slip_no) from DS_SlipLedgr_Ecount where Seq_No={manageNo[0]} and Work_YY={manageNo[2]} and work_qt={manageNo[3]} and slip_no=a.slip_no and Tran_Dt=a.tran_dt)>3"
    print(strsql)
    cursor = connection.cursor()
    cursor.execute(strsql)
    results = cursor.fetchall()
    connection.commit()
    if results:
      for i, row in enumerate(results):
        for j, value in enumerate(row):
          ws.cell(row=i+10, column=j+1, value=value)    
  ##########################
  wholeName = totalPath + fileName
  if os.path.exists(wholeName):        os.remove(wholeName)
  wb.save(wholeName)
  # openpyxl로 읽어온 xlsx 데이터를 xlwt로 생성한 xls 파일에 쓰기
  wb_xls = xlwt.Workbook()
  for sheet_name in wb.sheetnames:
    ws_xlsx = wb[sheet_name]
    ws_xls = wb_xls.add_sheet(sheet_name)
    for row in ws_xlsx.rows:
      for cell in row:
        ws_xls.write(cell.row-1, cell.col_idx-1, cell.value) 
  if os.path.exists(wholeName.split('.')[0]+'.xls'):os.remove(wholeName.split('.')[0]+'.xls')         
  wb_xls.save(wholeName.split('.')[0]+'.xls')    #xlsx를 xls로 변환한다  
  os.remove(wholeName)      
  return True

def BankTrade_xlsx(flagSeqNo,flagBank):
  arrflagBank = flagBank.split(":")  
  # 파일 대화상자를 표시하여 파일 선택
  file_name = select_file()
  memuser = MemUser.objects.get(seq_no=flagSeqNo);print(memuser.biz_no)
  # 파일이 선택된 경우 PyAutoGUI를 사용하여 파일 경로를 입력
  if file_name:
    pyautogui.write(file_name)
    pyautogui.press('enter')
    # 첫 번째 시도: 헤더 없이 전체 데이터를 읽어옴
    temp_df = pd.read_excel(file_name, header=None)

    # 데이터가 시작되는 행 탐지 (예를 들어, 특정 키워드 '거래일시'를 포함하는 첫 번째 행을 찾는 경우)
    df = None;date_column = None;totXlsArr = [] 
    header_row = None
    for i, row in temp_df.iterrows():
      if '거래일시' in row.values:
        header_row = i
        break

    # 실제 데이터가 시작되는 행을 찾았다면, 해당 행을 헤더로 사용하여 다시 파일을 읽음
    if header_row is not None:
      df = pd.read_excel(file_name, header=header_row)
      df = df.sort_values(by='거래일시')
    else:
      print("[거래일시]가 포함된 헤더 행을 찾을 수 없음, [거래일자]로 재시도")
      for i, row in temp_df.iterrows():
        if '거래일자' in row.values:
          header_row = i
          break      
      if header_row is not None:
        df = pd.read_excel(file_name, header=header_row)
        df = df.sort_values(by='거래일자')
      else:
        print("[거래일시,거래일자]가 포함된 헤더 행을 찾을 수 없음, [적요]로 재시도")
        for i, row in temp_df.iterrows():
          if '적요' in row.values:
            header_row = i
            break      
        if header_row is not None:
          df = pd.read_excel(file_name, header=header_row)  
      # 날짜 형식의 정규 표현식 정의
      date_pattern = re.compile(r'\d{4}-\d{2}-\d{2}')  # 기본적인 YYYY-MM-DD 형식
      # 날짜 열 식별
      date_column = identify_date_column(df, date_pattern)
      # 날짜 열을 기준으로 정렬
      if date_column:
        df = df.sort_values(by=date_column)
      else:
        date_pattern = re.compile(r'\d{4}/\d{2}/\d{2}')  # 기본적인 YYYY-MM-DD 형식
        date_column = identify_date_column(df, date_pattern)
        if date_column:
          df = df.sort_values(by=date_column)
        else:        
          print("날짜 형식의 열을 식별할 수 없습니다. 엑셀에서 날짜헤더를 [거래일시]로 수정해주세요")
          return False
    if df is not None and not df.empty:
      df = df.fillna('')  # Excel의 nil 값을 ''로 변환
      for row in df.itertuples():    
        rowDate = []
        try:
          rowDate = row.거래일시
        except:
          rowDate = row.거래일자
        if date_column:rowDate = getattr(row, date_column)
        if rowDate:#거래일시가 있는 경우
          # 거래일시, 입금액, 출금액, 적요를 변수에 할당
          tran_dt=  firstRemk=secondRemk=originRemk = ""
          tranamt_cr= tranamt_dr=tranamt_rm = 0
          if arrflagBank[1][:2]=="하나":
            tran_dt, tranamt_cr, tranamt_dr,tranamt_rm, firstRemk,secondRemk,originRemk = row.거래일시, row.입금, row.출금,row.잔액,row.구분,row[4], row.적요 #하나은행기준
          elif arrflagBank[1][:2]=="농협":
            tran_dt, tranamt_cr, tranamt_dr,tranamt_rm,firstRemk,secondRemk,originRemk = row.거래일자.replace("/",""), row[4], row[3],row[5],row.거래내용, row.거래기록사항, row[7]
          elif arrflagBank[1][:2]=="국민":
            tran_dt, tranamt_cr, tranamt_dr,tranamt_rm,firstRemk,secondRemk,originRemk = row.거래일시[:10].replace(".",""), row[5], row[4],row[6],row.적요, row[3], row[3]
          elif arrflagBank[1][:2]=="기업":
            tran_dt, tranamt_cr, tranamt_dr,tranamt_rm,firstRemk,secondRemk,originRemk = row.거래일시[:10].replace(".",""), row.입금, row.출금,row[5],row.거래구분, row.거래내용, row[6]
          elif arrflagBank[1][:2]=="우체":
            tran_dt, tranamt_cr, tranamt_dr,tranamt_rm,firstRemk,secondRemk,originRemk = row.거래일시[:10].replace(".",""), row.입금액, row.출금액,row[5],row.적요, row.내역, row[6]
          elif arrflagBank[1][:2]=="우리":
            tran_dt =row.거래일시[:10].replace("-","")
            tran_dt, tranamt_cr, tranamt_dr,tranamt_rm,firstRemk,secondRemk,originRemk = row.거래일시[:10].replace("-",""), row[6], row[5],row[7],row.적요, row.기재내용, row.메모
          crdr = "3"; crdr_op = "4";  tranAmt = tranamt_cr
          if tranamt_cr=='':tranamt_cr = 0
          if tranamt_dr=='':tranamt_dr = 0
          if tranamt_dr>0:crdr = "4"; crdr_op = "3";tranAmt = tranamt_dr
          if tran_dt!="합계":
            bankLine = [tran_dt[:10].replace("-",""),crdr,"103","보통예금",arrflagBank[0],arrflagBank[1]," ",originRemk,tranAmt]  #은행 거래내역 처리먼저
            totXlsArr.append(bankLine) 
            if firstRemk and ("체크" in firstRemk or "실시간이체" in firstRemk ):
              strsql = "select top 1 Work_YY,Tran_Dt,Slip_No "
              strsql += f" from DS_slipledgr2  with(index(IDX1_ds_slipledgr2), nolock)  "
              strsql += f" where acnt_cd = '253'"
              strsql += f" AND TranAmt_Cr > 0   "
              strsql += f" AND seq_no = {flagSeqNo}   "
              strsql += f" and slip_No in (SELECT slip_No "
              strsql += f"      FROM DS_slipledgr2 WITH(NOLOCK) "
              strsql += f"      WHERE seq_no = {flagSeqNo} "
              strsql += f"        AND acnt_cd = '103' "
              strsql += f"        AND TranAmt_Dr > 0 "
              strsql += f"        AND trader_code = '{arrflagBank[0]}') "
              strsql += f" ORDER BY work_yy DESC, tran_dt DESC"
              # strsql += f" where seq_no={flagSeqNo}   and acnt_cd='103'  and TranAmt_Dr>0 and trader_code='{arrflagBank[0]}' order by work_yy desc,tran_dt desc"
              # print(strsql)
              cursor = connection.cursor()
              result = cursor.execute(strsql)
              result = cursor.fetchone()
              connection.commit()
              connection.close()            
              if result:
                strsql2 = "select top 1 Trader_Code,Trader_Name "
                strsql2 += " from DS_slipledgr2  with(index(IDX1_ds_slipledgr2), nolock)  "
                strsql2 += f" where seq_no={flagSeqNo}  and Work_YY='{result[0]}' and tran_dt='{result[1]}'  and slip_no='{result[2]}' and acnt_cd<>'103' order by acnt_cd desc"       
                # print(strsql2)
                cursor = connection.cursor()
                result2 = cursor.execute(strsql2)
                result2 = cursor.fetchone()
                connection.commit()
                connection.close()               
              if  tranamt_dr!=0:
                bankLine = [tran_dt[:10].replace("-",""),crdr_op,"253","미지급금",result2[0],result2[1]," ",originRemk,tranAmt] 
                totXlsArr.append(bankLine) 
              else:
                bankLine = [tran_dt[:10].replace("-",""),crdr_op,"253","미지급금",result2[0],result2[1]," ",originRemk,tranAmt] 
                totXlsArr.append(bankLine)               
            else:#체크카드가 아닌 경우
              tranRemk = originRemk
              if tranamt_dr>0 and originRemk:
                tranRemk = remove_dates_and_extract_text(originRemk) #출금거래의 적요 중 숫자가 포함된 경우 잘라낸다  예)"건강보험2304월" ==> "건강보험"
              if originRemk=='':
                if originRemk: tranRemk=originRemk
                else:tranRemk=firstRemk
              # 괄호를 기준으로 문자열을 분리합니다.
              tranRemk_parts = re.split(r'[()]', tranRemk)
              # 빈 문자열 제거
              tranRemk_parts = [part for part in tranRemk_parts if part.strip()]
              # 검색 조건을 생성합니다.
              # 주의: SQL 인젝션을 방지하기 위해 파라미터화된 쿼리를 사용하는 것이 좋습니다.
              if len(tranRemk_parts) >= 2:
                  tranRemk_like = f"%{tranRemk_parts[0]}%{tranRemk_parts[1]}%"
              elif len(tranRemk_parts) == 1:
                  tranRemk_like = f"%{tranRemk_parts[0]}%"
              else:
                  tranRemk_like = "%%"
              strsql = "select top 1 Work_YY,Tran_Dt,crdr,Acnt_Nm,Trader_Name,TranAmt_Cr,TranAmt_Dr,Remk,Tran_Stat,Slip_No "
              strsql += f" from DS_slipledgr2  with(index(IDX1_ds_slipledgr2), nolock)  "
              strsql += f" where seq_no={flagSeqNo}   and acnt_cd='103' and Remk like '{tranRemk_like}' order by work_yy desc,tran_dt desc"
              # print(strsql)  적요와 비슷한 최근의 거래를 찾는다
              cursor = connection.cursor()
              result = cursor.execute(strsql)
              result = cursor.fetchall()
              connection.commit()
              connection.close()
                  
              if result:
                print(f"{tranRemk} : 적요와 비슷한 최근의 거래가 있는 경우")
                column_names = [desc[0] for desc in cursor.description]
                result_dicts = [dict(zip(column_names, row)) for row in result]
                for r in result_dicts:
                  rWork_YY,rTran_Dt,rSlip_no = r['Work_YY'],r['Tran_Dt'],r['Slip_No']
                  # print(f"유사한 거래 발견: {r['Work_YY']} {r['Tran_Dt']} {r['Acnt_Nm']} {r['Remk']} {r['Slip_No']}")

                  strsql2 = "select work_yy,Tran_Dt,crdr,acnt_cd,acnt_nm,Trader_Code,Trader_Name,TranAmt_Cr,TranAmt_Dr,Remk,Tran_Stat "
                  strsql2 += " from DS_slipledgr2  with(index(IDX1_ds_slipledgr2), nolock)  "
                  strsql2 += f" where seq_no={flagSeqNo}  and Work_YY='{rWork_YY}' and tran_dt='{rTran_Dt}'  and slip_no='{rSlip_no}' "
                  if tranamt_cr>0:    strsql2 += " AND NOT  Acnt_cd = '103' order by acnt_cd desc"   
                  elif tranamt_dr>0:  strsql2 += " AND NOT  Acnt_cd = '103' order by acnt_cd desc"   
                        
                  # print(strsql2)
                  cursor = connection.cursor()
                  result2 = cursor.execute(strsql2)
                  result2 = cursor.fetchall()
                  connection.commit()
                  connection.close() 
                  # print(len(result2))
                  column_names = [desc[0] for desc in cursor.description]
                  result_dicts = [dict(zip(column_names, row)) for row in result2]    
                  tmpAmt = 0#임시
                  for r in result_dicts:#유사거래 내에서 루핑
                    if (int(r['acnt_cd'])>400 and int(r['acnt_cd'])<830) or  tranRemk=='국민연금' or tranRemk=='건강보험' or tranRemk=='국민건강': #판관비중 지급수수료가 아닌 경우
                      if int(r['acnt_cd'])!=813:
                        tmpAmt = tranAmt/2#유사거래 아닌 현재 거래 차변
                      else:
                        print('접대비')
                    elif int(r['acnt_cd'])==831 and int(r['TranAmt_Cr'])<1000: #지급수수료인 경우
                      tmpAmt = int(r['TranAmt_Cr'])
                    elif int(r['acnt_cd'])>901 : #잡이익
                      tmpAmt =  tranamt_dr % 10# 유사거래 아닌 현재 거래 대변의 원단위

                    if tmpAmt==0:
                      #판관비가 아닌 경우
                      bankLine = [tran_dt[:10].replace("-",""),crdr_op,r['acnt_cd'],r['acnt_nm'],r['Trader_Code'],r['Trader_Name']," ",tranRemk,tranAmt-tmpAmt]
                    else:
                      if r['acnt_cd']:
                        if int(r['acnt_cd'])>400:#자산부채가 아닌 경우
                          bankLine = [tran_dt[:10].replace("-",""),crdr_op,r['acnt_cd'],r['acnt_nm'],r['Trader_Code'],r['Trader_Name']," ",tranRemk,tmpAmt]
                        else:
                          bankLine = [tran_dt[:10].replace("-",""),crdr_op,r['acnt_cd'],r['acnt_nm'],r['Trader_Code'],r['Trader_Name']," ",tranRemk,tranAmt-tmpAmt]
                      else:#계정코드가 없는 경우
                        print('계정코드가 없는 경우')
                    totXlsArr.append(bankLine)                    
              else:
                print(f"{tranRemk} : 적요와 비슷한 최근의 거래가 없는 경우")

                strsql = "select * from "
                if '국민연금' in tranRemk:
                  tranAmt1 = math.floor(tranAmt/2)
                  bankLine = [tran_dt[:10].replace("-",""),crdr_op,"254","예수금",arrflagBank[0],arrflagBank[1]," ",tranRemk,tranAmt1]
                  totXlsArr.append(bankLine)    
                  bankLine = [tran_dt[:10].replace("-",""),crdr_op,"817","세금과공과",arrflagBank[0],arrflagBank[1]," ",tranRemk,tranAmt-tranAmt1]
                  totXlsArr.append(bankLine)          
                elif '건강보험' in tranRemk or '국민건강' in tranRemk :
                  tranAmt1 = math.floor(tranAmt/2)
                  bankLine = [tran_dt[:10].replace("-",""),crdr_op,"254","예수금",arrflagBank[0],arrflagBank[1]," ",tranRemk,tranAmt1]
                  totXlsArr.append(bankLine)    
                  bankLine = [tran_dt[:10].replace("-",""),crdr_op,"811","복리후생비",arrflagBank[0],arrflagBank[1]," ",tranRemk,tranAmt-tranAmt1]
                  totXlsArr.append(bankLine)      
                elif  '사회보험' in tranRemk:   
                  bankLine = [tran_dt[:10].replace("-",""),crdr_op,"254","예수금",arrflagBank[0],arrflagBank[1]," ",tranRemk,tranAmt]
                  totXlsArr.append(bankLine)                   
                else:  
                  if flagSeqNo=='3639': 
                    print(f'유사거래 없음 - {tran_dt} : {rowDate} 입금 : {tranamt_cr}, 출금 : {tranamt_dr}, 적요 : {tranRemk}')
                    query = "SELECT semu_name FROM cms_semusarang_trdst"
                    cursor = connection.cursor()
                    cursor.execute(query)
                    # 유사도 체크
                    for row in cursor:
                      semu_name = row.semu_name
                      similarity = difflib.SequenceMatcher(None, semu_name, tranRemk).ratio()
                      if similarity*100>3.5:
                        print(f"유사도 {similarity*100:.2f}%: {semu_name} / {tranRemk}")
                    cursor.close()
                  else:
                    print(f'유사거래 없음 - {tran_dt} : {rowDate} 입금 : {tranamt_cr}, 출금 : {tranamt_dr}, 적요 : {tranRemk}')
                    query = f"SELECT 공급자상호 FROM 전자세금계산서 where SEQ_NO={flagSeqNo} "
                    cursor = connection.cursor()
                    cursor.execute(query)
                    # 유사도 체크
                    max_similarity = 0
                    most_similar_name = ""
                    for row in cursor:
                      semu_name = row.공급자상호
                      similarity = difflib.SequenceMatcher(None, semu_name, tranRemk).ratio()
                      if similarity * 100 > 40 and similarity > max_similarity:
                        max_similarity = similarity
                        most_similar_name = semu_name
                        #print(f"유사도 {similarity*100:.2f}%: {semu_name} / {tranRemk}")
                    cursor.close()            
                    if most_similar_name:
                      print(f"가장 유사한 상호: {most_similar_name} (유사도: {max_similarity * 100:.2f}%)")
                      if crdr_op == "3":
                        bankLine = [tran_dt[:10].replace("-",""),crdr_op,"251","외상매입금","",most_similar_name," ",tranRemk,tranAmt]
                      else:
                        bankLine = [tran_dt[:10].replace("-",""),crdr_op,"108","외상매출금","",most_similar_name," ",tranRemk,tranAmt]
                    else:
                      #  print("40% 이상의 유사도를 갖는 상호가 없습니다.")                            
                      if "년결산" in tranRemk:
                        bankLine = [tran_dt[:10].replace("-",""),crdr_op,"901","이자수익",arrflagBank[0],arrflagBank[1]," ",tranRemk,tranAmt]
                      elif "사업세" in tranRemk :
                        bankLine = [tran_dt[:10].replace("-",""),crdr_op,"254","예수금","",tranRemk," ",tranRemk,tranAmt]
                      elif "갑근세" in tranRemk:
                        bankLine = [tran_dt[:10].replace("-",""),crdr_op,"254","예수금","",tranRemk," ",tranRemk,tranAmt]
                      elif "국세" in tranRemk or  "서울시세" in tranRemk:
                        bankLine = [tran_dt[:10].replace("-",""),crdr_op,"261","미지급세금","",tranRemk," ",tranRemk,tranAmt]
                      else:
                        bankLine = [tran_dt[:10].replace("-",""),crdr_op,"138","전도금",arrflagBank[0],arrflagBank[1]," ",tranRemk,tranAmt]
                    totXlsArr.append(bankLine)                                  
      wb = Workbook()  # 새 워크북 만들기  
      ws = wb.active   # 워크북의 첫 번째 워크시트 가져오기
      fileName = f"엑셀자료 일반전표전송({memuser.biz_name}-{arrflagBank[0]}{arrflagBank[1]}).xlsx" 
      ws['A1'] = "엑셀자료 일반전표전송"
      ws['O3'] = memuser.biz_no
      ws['M4'] = 'v1.2'           
      MIDTITLE = ["년도월일","구분","코드","계정과목","코드","거래처명","코드","적요명","금액"]   
      for i, value in enumerate(MIDTITLE):
        ws.cell(row=10, column=i+1, value=value)      
      for i, row in enumerate(totXlsArr):
        for j, value in enumerate(row):
          ws.cell(row=i+11, column=j+1, value=value)
      ##########################
      totalPath = "C:\\Users\\Administrator\\Documents\\"
      wholeName = totalPath + fileName
      if os.path.exists(wholeName):        os.remove(wholeName)
      wb.save(wholeName)          
      # openpyxl로 읽어온 xlsx 데이터를 xlwt로 생성한 xls 파일에 쓰기
      wb_xls = xlwt.Workbook()
      for sheet_name in wb.sheetnames:
        ws_xlsx = wb[sheet_name]
        ws_xls = wb_xls.add_sheet(sheet_name)
        for row in ws_xlsx.rows:
          for cell in row:
            ws_xls.write(cell.row-1, cell.col_idx-1, cell.value)           
      if os.path.exists(wholeName.split('.')[0]+'.xls'):os.remove(wholeName.split('.')[0]+'.xls')         
      wb_xls.save(wholeName.split('.')[0]+'.xls')    #xlsx를 xls로 변환한다  
      os.remove(wholeName) 
  return True

def Make98002(totXlsArr):
  wb = Workbook()  # 새 워크북 만들기  
  ws = wb.active   # 워크북의 첫 번째 워크시트 가져오기
  fileName = f"엑셀자료 일반전표전송(98002).xlsx" 
  ws['A1'] = "엑셀자료 일반전표전송"
  ws['O3'] = "2208533586"
  ws['M4'] = 'v1.2'           
  MIDTITLE = ["년도월일","구분","코드","계정과목","코드","거래처명","코드","적요명","금액"]   
  for i, value in enumerate(MIDTITLE):
    ws.cell(row=10, column=i+1, value=value)      
  for i, row in enumerate(totXlsArr):
    for j, value in enumerate(row):
      ws.cell(row=i+11, column=j+1, value=value)
  ##########################
  totalPath = "C:\\Users\\Administrator\\Documents\\"
  wholeName = totalPath + fileName
  if os.path.exists(wholeName):        os.remove(wholeName)
  wb.save(wholeName)          
  # openpyxl로 읽어온 xlsx 데이터를 xlwt로 생성한 xls 파일에 쓰기
  wb_xls = xlwt.Workbook()
  for sheet_name in wb.sheetnames:
    ws_xlsx = wb[sheet_name]
    ws_xls = wb_xls.add_sheet(sheet_name)
    for row in ws_xlsx.rows:
      for cell in row:
        ws_xls.write(cell.row-1, cell.col_idx-1, cell.value)           
  if os.path.exists(wholeName.split('.')[0]+'.xls'):os.remove(wholeName.split('.')[0]+'.xls')         
  wb_xls.save(wholeName.split('.')[0]+'.xls')    #xlsx를 xls로 변환한다  
  os.remove(wholeName) 
    
  return True

#다운된 파일 열고 디비로 저장하기
def DBSave_Downloaded_xlsx(driver,manageNo,flag):
  workyear = datetime.now().year-1
  today = str(datetime.now())[:10].replace("-","")
  file_name=""
  if flag in ("보수총액신고대상자업데이트",'거래처등록'):
    downloads = 'C:\\NewGen\\Rebirth\\리버스문서보관함'
    file_name = max([downloads + "\\" + f for f in os.listdir(downloads)],key=os.path.getctime)    
  else:
    if flag!="이카운트분개장업로드":
      downloads = 'C:\\Users\\Administrator\\Downloads'
      try:
        file_name = max([downloads + "\\" + f for f in os.listdir(downloads)],key=os.path.getctime)
        extension = os.path.splitext(file_name)[1].lower()  # Get the extension and convert it to lower case
        if extension in ['.xls', '.xlsx']:
            print("This is an Excel file.")
        else:
            print(f"{extension} - 엑셀파일이 아닙니다.")      
            return False
      except:
        print(manageNo[6] + "의 엑셀파일 없음")
        return False
  print(f'{file_name}')
  if flag=='부가세예정고지(확정)':
    df = pd.read_excel(file_name,header=1, engine='openpyxl')
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
    for row in df.itertuples():
      bizno=str(row[5])
      try:
        memuser = MemUser.objects.get(biz_no=bizno[:3]+"-"+bizno[3:5]+"-"+bizno[5:10])
        # if int(row[8])>0 and not (row[2]=='예정고지 취소' or row[2]=='예정고지 기준금액 미만' or row[2]=='예정고지 기준금액 미만 및 폐업'):
        if int(row[8]) > 0 and row[2].strip() not in {'예정고지 취소', '예정고지 기준금액 미만', '예정고지 기준금액 미만 및 폐업'}:
          sql = "merge tbl_vat  as A Using (select '"+memuser.seq_no+"' as seq_no , '"+manageNo[0]+"' as work_yy, '"+manageNo[1]+"' as work_qt) as B "
          sql += "On A.seq_no = B.seq_no and A.work_yy = B.work_yy  and A.work_qt = B.work_qt "
          sql += "WHEN Matched Then  	Update set YN_15=" + str(row[8])
          sql += "	When Not Matched Then  "
          sql += "	INSERT 	values('"+memuser.seq_no+"','"+manageNo[0]+"','"+manageNo[1]+"',0,0,0,0,0,0,0,0,0,0,0,0,'',-1,'"+str(row[8])+"',0,0);"
          print(sql)
          cursor = connection.cursor()
          cursor.execute(sql)
          cursor.commit()
          print(str(row[0])+" : "+row[6]+" : "+'디비저장성공')
        else:
          print(str(row[0])+" : "+row[6]+" : "+row[2]+" : "+str(row[8]))
      except:
        print(str(row[0])+" : "+row[6]+" : "+str(row[5])+" : Mem_User DB 사업자번호 중복 또는 없음")
  elif flag=='사무수임사업장':
    strdel = "delete from EDI_COMWEL"
    cursor = connection.cursor()
    cursor.execute(strdel)    
    df = pd.read_excel(file_name,header=1, engine='openpyxl')
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
    for row in df.itertuples():  
      sql = f"insert into EDI_COMWEL values('{row[1]}','{row[2]}','{row[3]}','{row[4]}','{row[5]}','{row[6]}','{row[7]}','{row[8]}','{row[9]}','{row[10]}'"
      sql += f",'{row[11]}','{row[12]}','{row[13]}','{row[14]}','{row[15]}','{row[16]}','{row[17]}','{row[18]}')"
      print(sql)
      cursor = connection.cursor()
      cursor.execute(sql)
      cursor.commit()
    print('디비저장성공')  
  elif flag=='근로자고용정보현황조회':
    df = pd.read_excel(file_name,header=1, engine='openpyxl')
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
    strdel = "delete from tbl_Employ where seq_no='"+ manageNo[2]  + "'"
    cursor = connection.cursor()
    cursor.execute(strdel)
    for row in df.itertuples():
      if row[5]!="" and row[11]!="":
        Birth_Dt=""
        if    int(row[3][7:8])<3 : Birth_Dt	= "19" + row[3][:2] + "-" + row[3][2:4] +"-"+ row[3][4:6]
        elif  int(row[3][7:8])<5 :	Birth_Dt	= "20"+ row[3][:2] + "-" + row[3][2:4] +"-"+ row[3][4:6]
        elif int(row[3][7:8])<7 : Birth_Dt	= "19"+ row[3][:2] + "-" + row[3][2:4] +"-"+ row[3][4:6]
        else:			                Birth_Dt	= "20"+ row[3][:2] + "-" + row[3][2:4] +"-"+ row[3][4:6]
        # sql = "Merge tbl_employ as A Using (select '"+manageNo[2]+"' as seqno , '"+row[3][:8]+"' as juNum, '"+row[5]+"' as EnterDate) as B "
        # sql += "On A.seq_no = B.seqno and left(A.empJumin,8)=B.juNum and A.empRegDate=B.EnterDate "
        # sql += "when not matched then insert values('"+manageNo[2]

        sql = f"insert into Tbl_Employ values('{manageNo[2]}','','{row[3]}','{row[4]}','{Birth_Dt}','{row[5]}','{row[6]}','" 
        sql += "','','','','','','0');"
        print(sql)
        cursor = connection.cursor()
        cursor.execute(sql)
        cursor.commit()
        print('디비저장성공')
    time.sleep(1)
    driver.find_element(By.XPATH,'//*[@id="mf_wfm_content_edtGwanriNo"]').clear()
    time.sleep(0.5)
  elif flag=='보수총액신고대상자':
    df = pd.read_excel(file_name, header=0,dtype = {'주민(외국인)등록번호':str})
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환    c.관리번호, trim(a.biz_name), a.seq_no
    strdel = f"delete from 보수총액신고_고용 where seq_no={manageNo[2]} and work_yy={workyear}"
    cursor = connection.cursor()
    cursor.execute(strdel)       
    for row in df.itertuples():
      # sql = "Merge 보수총액신고_고용 as A Using (select '"+str(manageNo[2])+"' as seqno , '"+str(workyear)+"' as work_yy, '"+str(row[2])+"' as ssn, '"+str(row[5])[:8]+"' as 산재취득일 , '"+str(row[9])[:8]+"' as 고용취득일) as B "
      # sql += "On A.seq_no = B.seqno and A.seq_no=B.seqno and A.work_yy=B.work_yy and A.주민번호=B.ssn and A.산재취득일=B.산재취득일 and A.고용취득일=B.고용취득일 when not matched then "
      sql = "insert into 보수총액신고_고용 values('"+str(manageNo[2])
      sql += "','" + str(workyear)  #귀속년
      sql += "','" + str(row[1])  #성명
      sql += "','" + str(row[2]) #주번
      sql += "','" + str(row[3]) #부과구분
      sql += "','" + str(row[4]).replace(".0","") #건설업
      sql += "','" + str(row[5]).replace(".0","") #산재취득일
      sql += "','" + str(row[6]).replace(".0","") #산재싱실일
      sql += "',0"  #산재보수총액
      sql += ",'" + str(row[8]).replace(".0","") #근무지코드
      sql += "','" + str(row[9])[:8] #고용취득일
      sql += "','" + str(row[10])[:8] #고용상
      sql += "',0"  #총액상반기
      sql += ",0"  #총액하반기
      sql += "); "
      print(sql)
      cursor = connection.cursor()
      cursor.execute(sql)
      cursor.commit()
      print('디비저장성공')      
  elif flag=='보수총액신고대상자업데이트':
    df = None
    if file_name.endswith('.xls'):
        df = pd.read_excel(file_name, header=2, engine='xlrd', dtype={'주민(외국인)등록번호': str})
    elif file_name.endswith('.xlsx'):
        df = pd.read_excel(file_name, header=2, engine='openpyxl', dtype={'주민(외국인)등록번호': str})
    else:
        raise ValueError("지원하지 않는 파일 형식입니다.")    
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환    c.관리번호, trim(a.biz_name), a.seq_no
    print(df)    
    for row in df.itertuples():
      if str(row[6])!='':
        sql = "Merge 보수총액신고_고용 as A Using (select '"+str(manageNo)+"' as seqno , '"+str(workyear)+"' as work_yy, '"+str(row[2]).replace("-","")+"' as ssn,"
        sql += "'"+str(row[4]).replace("-","")+"' as 산재취득일 , '"+str(row[8]).replace("-","")+"' as 고용취득일) as B "
        sql += "On A.seq_no = B.seqno and A.seq_no=B.seqno and A.work_yy=B.work_yy and A.주민번호=B.ssn and left(A.산재취득일,6)=left(B.산재취득일,6)  "
        sql += " when  matched then update set "        
        sql += " 산재연간보수총액="+str(row[6])+";"
        print(sql)
        cursor = connection.cursor()
        cursor.execute(sql)
        cursor.commit()
      if str(row[10])!='':
        sql = "Merge 보수총액신고_고용 as A Using (select '"+str(manageNo)+"' as seqno , '"+str(workyear)+"' as work_yy, '"+str(row[2]).replace("-","")+"' as ssn, '"+str(row[4]).replace("-","")+"' as 산재취득일 , '"+str(row[8]).replace("-","")+"' as 고용취득일) as B "
        sql += "On A.seq_no = B.seqno and A.seq_no=B.seqno and A.work_yy=B.work_yy and A.주민번호=B.ssn  and left(A.고용취득일,6)=left(B.고용취득일,6) "
        sql += " when  matched then update set "                
        sql += " 고용연간보수총액_상반기 = " + str(row[10])+";"
        print(sql)
        cursor = connection.cursor()
        cursor.execute(sql)
        cursor.commit()        
      if str(row[11])!='':
        sql = "Merge 보수총액신고_고용 as A Using (select '"+str(manageNo)+"' as seqno , '"+str(workyear)+"' as work_yy, '"+str(row[2]).replace("-","")+"' as ssn, '"+str(row[4]).replace("-","")+"' as 산재취득일 , '"+str(row[8]).replace("-","")+"' as 고용취득일) as B "
        sql += "On A.seq_no = B.seqno and A.seq_no=B.seqno and A.work_yy=B.work_yy and A.주민번호=B.ssn  and left(A.고용취득일,6)=left(B.고용취득일,6)  "
        sql += "when  matched then update set "                
        sql += " 고용연간보수총액_하반기 = " + str(row[11])+";"
        print(sql)
        cursor = connection.cursor()
        cursor.execute(sql)
        cursor.commit()
      print('디비저장성공')            
  elif flag=='은행':
    if manageNo[2]=='0003':   #기업은행
      df = pd.read_excel(file_name, header=None)
      df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
      print(df)
      for row in df.itertuples():
        sql = "Merge 은행거래내역 as A Using (select '"+str(manageNo[0])+"' as seqno , '"+str(manageNo[2])+"' as bankCode, '"+str(manageNo[3].replace(" ",""))+"' as bankNum, '"+str(row[1])+"' as tradeDatetime) as B "
        sql += "On A.seq_no = B.seqno and A.은행코드=B.bankCode and A.계좌번호=B.bankcode and A.거래일시=B.tradeDatetime when not matched then "
        sql += "insert values('"+str(manageNo[0])
        sql += "','" + str(manageNo[2])  #은행코드
        sql += "','" + str(manageNo[3].replace(" ",""))  #계좌번호
        sql += "'," + "(SELECT ISNULL(MAX(일련번호), 0) + 1 FROM 은행거래내역  Where Seq_No = '"+str(manageNo[0])+"' AND 은행코드 = '"+str(manageNo[2].replace(" ",""))+"' AND 계좌번호  = '"+str(manageNo[3].replace(" ",""))+"')"  #일련번호
        sql += ",'" + str(row[1]) #거래일시
        sql += "'," + str(row[2]) #출금
        sql += "," + str(row[3]) #입금
        sql += "," + str(row[4]) #잔액
        sql += ",'" + str(row[5]) #거래내용1
        sql += "','" + "" #거래내용2
        sql += "','" + "" #거래내용3
        sql += "','" + str(row[6]) #상대계좌
        sql += "','" + str(row[7]) #상대은행
        sql += "','" + str(row[8]) #cms코드
        sql += "','" + str(row[9]) #거래구분
        sql += "',0"  #미결재
        sql += ",'" + str(datetime.now())[:10].replace("-","") #crtdt
        sql += "','" + "" #slipdt
        sql += "','" + "" #slipAcntCd
        sql += "','" + "" #slipAcntCd
        sql += "N');" #slipYN
        print(sql)
        cursor = connection.cursor()
        cursor.execute(sql)
        cursor.commit()
      print('디비저장성공')    
    elif manageNo[2]=='0081':   #하나은행
      df = pd.read_excel(file_name, header=5)
      df = df.fillna('0')  # 3. excel 의 nill 값을 '' 로 변환
      print(df)      
      for row in df.itertuples():
        row5 = str(row[5]).replace(" ","")
        if row5=='':row5="0"
        row4 = str(row[4]).replace(" ","")
        if row4=='':row4="0"
        row6 = str(row[6]).replace(" ","")
        if row6=='':row6="0"
        sql = "Merge 은행거래내역 as A Using (select '"+str(manageNo[0])+"' as seqno , '"+str(manageNo[2])+"' as bankCode, '"+str(manageNo[3].replace(" ",""))+"' as bankNum, '"+str(row[1])+" "+str(row[7])+"' as tradeDatetime) as B "
        sql += "On A.seq_no = B.seqno and A.은행코드=B.bankCode and A.계좌번호=B.bankcode and A.거래일시=B.tradeDatetime when not matched then "
        sql += "insert values('"+str(manageNo[0])
        sql += "','" + str(manageNo[2])  #은행코드
        sql += "','" + str(manageNo[3].replace(" ",""))  #계좌번호
        sql += "'," + "(SELECT ISNULL(MAX(일련번호), 0) + 1 FROM 은행거래내역  Where Seq_No = '"+str(manageNo[0])+"' AND 은행코드 = '"+str(manageNo[2].replace(" ",""))+"' AND 계좌번호  = '"+str(manageNo[3].replace(" ",""))+"')"  #일련번호
        sql += ",'" + str(row[1])+" "+str(row[7]) #거래일시
        sql += "'," + row5 #출금
        sql += "," + row4 #입금
        sql += "," + row6.replace(".0","") #잔액
        sql += ",'" + str(row[3]) #거래내용1
        sql += "','" + str(row[2]) #거래내용2
        sql += "','" + str(row[8]) #거래내용3
        sql += "','" + "" #상대계좌
        sql += "','" + "" #상대은행
        sql += "','" + "" #cms코드
        sql += "','" + "" #거래구분
        sql += "',0"  #미결재
        sql += ",'" + str(datetime.now())[:10].replace("-","") #crtdt
        sql += "','" + "" #slipdt
        sql += "','" + "" #slipAcntCd
        sql += "','" + "" #slipAcntCd
        sql += "N');" #slipYN
        print(sql)
        cursor = connection.cursor()
        cursor.execute(sql)
        cursor.commit()      
  elif flag=='거래처등록':
    strsql = f"select trim(TrdCd_Semu) from Financial_Trdst_Ecnt where seq_no={manageNo} order by TrdCd_Semu/1"
    cursor = connection.cursor()
    cursor.execute(strsql);print(strsql)
    TrdCd_Semus = cursor.fetchall()
    connection.commit()
     
    df = pd.read_excel(file_name,header=0)
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
    for row in df.itertuples():
      TrdCd_Semu = str(row[1])
      if len(TrdCd_Semu)==3:TrdCd_Semu = "00"+TrdCd_Semu 
      elif len(TrdCd_Semu)==4:TrdCd_Semu = "0"+TrdCd_Semu 
      if TrdCd_Semu not in [row[0] for row in TrdCd_Semus]:
        strIns = f"insert into Financial_Trdst_Ecnt values({TrdCd_Semu},'{manageNo}','{TrdCd_Semu}','{row[2]}','{row[3].replace('-','')}','{row[2]}','{row[3]}','{driver}')"
        print(strIns)
        cursor.execute(strIns)
  elif flag=='CMS회원목록':
    if f'{today}_회원목록' in file_name :
      strsql = f"select trim(seq_CMS),trim(seq_SEMU) from CMS_MEMBER  order by seq_CMS"
      cursor = connection.cursor()
      cursor.execute(strsql);print(strsql)
      seq_CMSs = cursor.fetchall()
      connection.commit()

      df = pd.read_excel(file_name,header=0)
      df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
      
      for row in df.itertuples():  
        seq_MEMBER = seq_BIZNO = seq_SEMU = ''
        strsql = f"select a.seq_no,biz_no,biz_type from mem_user a,mem_deal b  where a.seq_no=b.seq_no and trim(biz_name) = '{row[2].strip()}' and biz_no<>'--' "
        cursor = connection.cursor()
        cursor.execute(strsql);print(strsql)
        mem_user = cursor.fetchone()
        connection.commit()
        if mem_user:
          seq_MEMBER = mem_user[0];seq_BIZNO = mem_user[1]
          # strsql = f"select trim(seq_SEMU) from cms_semusarang_trdst where semu_bizno='{seq_BIZNO}'"
          strsql = f"select trim(TrdCd_Semu) from Financial_Trdst_Ecnt where Trdst_Num='{seq_BIZNO}'"
          cursor = connection.cursor()
          cursor.execute(strsql);print(strsql)
          seq_SEMUs = cursor.fetchone()
          connection.commit()
          if seq_SEMUs:seq_SEMU = seq_SEMUs[0]

        seq_CMS_set = {row[0] for row in seq_CMSs}
        if row[1].strip() not in seq_CMS_set:
          strIns = f"insert into CMS_MEMBER values('{row[1]}','{row[2]}','{row[3]}','{row[4]}','{row[7]}','{row[8]}','{row[9]}','{row[10]}','{seq_MEMBER}','{seq_SEMU}')"
          print(strIns)
          cursor.execute(strIns)        
        else:#CMS_MEMBER에 있는데
          # seq_CMS 값을 키로 하고, 해당하는 seq_SEMU 값을 값으로 하는 딕셔너리 생성
          seq_CMS_SEMU_dict = {row[0]: row[1] for row in seq_CMSs}          
          if seq_CMS_SEMU_dict[row[1].strip()] == '' and seq_SEMU:
            strUdt = f"update CMS_MEMBER set seq_semu='{seq_SEMU}' where seq_cms='{row[1].strip()}'"
            print(strUdt)
            cursor.execute(strUdt) 
  elif flag=='CMS출금내역업로드':
    if f'{today}_CMS출금조회내역' in file_name :
      df = pd.read_excel(file_name,header=0)
      df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
      for row in df.itertuples():
        strsql = f"select trim(seq_SEMU),trim(seq_MEMUSER) from CMS_MEMBER where trim(seq_CMS)='{row[3].strip()}'"
        cursor = connection.cursor()
        cursor.execute(strsql);print(strsql)
        CMS_MEMBER = cursor.fetchone()
        connection.commit()
        if CMS_MEMBER and CMS_MEMBER[1]:
          if row[9] == '출금실패' or  row[9] == '재출금실패':
            str_mg = f"Merge tbl_mng_jaroe as A Using (select '{CMS_MEMBER[1]}' as seq_no, '{row[1][:4]}' as work_yy, '{int(row[1][4:6])}' as work_mm  ) as B "
            str_mg += "On A.seq_no = B.seq_no and A.work_yy = B.work_yy  and A.work_mm = B.work_mm  "
            str_mg += f"WHEN Matched Then  Update set YN_10=1 "
            str_mg += "	When Not Matched Then  "
            str_mg += "	INSERT (seq_no,work_YY,work_MM,YN_1,YN_2,YN_3,YN_4,YN_5,YN_6,YN_7,YN_8,YN_9,YN_10,YN_11,YN_12,YN_13,YN_14,bigo) "
            str_mg += f"	values({CMS_MEMBER[1]},{row[1][:4]},{int(row[1][4:6])},0,0,0,0,0,0,0,0,0,1,0,0,0,0,'');"          
            print(str_mg)
            cursor.execute(str_mg)  
      for row in df.itertuples():
        strsql = f"select trim(seq_SEMU),trim(seq_MEMUSER) from CMS_MEMBER where trim(seq_CMS)='{row[3].strip()}'"
        cursor = connection.cursor()
        cursor.execute(strsql);print(strsql)
        CMS_MEMBER = cursor.fetchone()
        connection.commit()
        if CMS_MEMBER and CMS_MEMBER[1]:
          if '성공' in row[9]:    
            str_mg = f"Merge tbl_mng_jaroe as A Using (select '{CMS_MEMBER[1]}' as seq_no, '{row[1][:4]}' as work_yy, '{int(row[1][4:6])}' as work_mm  ) as B "
            str_mg += "On A.seq_no = B.seq_no and A.work_yy = B.work_yy  and A.work_mm = B.work_mm  "
            str_mg += f"WHEN Matched Then  Update set YN_10=0 ;"         
            print(str_mg)
            cursor.execute(str_mg)             
  elif flag=='이카운트분개장업로드':
    str_del = f"delete from Ds_slipledgr_Ecount where seq_no='{manageNo[0]}' and work_YY='{manageNo[1]}' and work_QT = '{manageNo[2]}'"
    cursor = connection.cursor()
    cursor.execute(str_del)    
    connection.close()  
    file_name = driver;print(file_name)
    temp_df = pd.read_excel(file_name, header=None)
    header_row = None
    for i, row in temp_df.iterrows():
      if '계정코드' in row.values:
        header_row = i
        break    
    df = pd.read_excel(file_name,header=header_row)
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환
    for i,row in enumerate(df.itertuples()):
      if row[7]!='':#마지막 공란에러를 발생시켜야 원본파일이 삭제되지 않는다
        strsql_trd = "select AcntCd_Semu,AcntNm_Semu, "
        if row[5].strip()=="" :
          strsql_trd += " '' TrdCd_Semu ,'' TrdNm_Semu ,'' Trdst_Num  "
        else:
          strsql_trd += f" isnull((select max(trdcd_semu) from financial_Trdst_Ecnt where seq_no={manageNo[0]} and trdCd_ecnt=rtrim('{row[5].strip()}')) ,'') TrdCd_Semu, "
          strsql_trd += f" isnull((select max(trdnm_semu) from financial_Trdst_Ecnt where seq_no={manageNo[0]} and trdCd_ecnt=rtrim('{row[5].strip()}')) ,'') TrdNm_Semu,"
          strsql_trd += f" isnull((select max(trdst_Num) from financial_Trdst_Ecnt where seq_no={manageNo[0]} and trdCd_ecnt=rtrim('{row[5].strip()}')) ,'') Trdst_Num "			
        strsql_trd += f" from Financial_AcntCd_Ecnt where AcntCd_Ecnt like '{int(row[7])}%'"#;print(strsql_trd)
        cursor = connection.cursor()
        cursor.execute(strsql_trd)
        result = cursor.fetchone()
        columns = [col[0] for col in cursor.description]
        connection.commit()
        if result:
          rs2 = dict(zip(columns, result))
          TrdCd_Semu = rs2["TrdCd_Semu"].strip()
          TrdNm_Semu = rs2["TrdNm_Semu"].strip()
          Trn_Num = rs2["Trdst_Num"].strip()
          Trn_Date = row[2].replace("/","").strip()
          if TrdCd_Semu == "" :
            tmpBizNo = row[5].strip()
            if len(tmpBizNo)<=12:tmpBizNo = f"{row[5][:3]}-{row[5][3:5]}-{row[5][5:10]}"
            strsql = f"select trdcd_semu,trdNm_semu from financial_trdst_ecnt where seq_no={manageNo[0]} and trdst_num='{tmpBizNo}'"#;print(strsql)
            cursor.execute(strsql)
            result2 = cursor.fetchone()
            connection.commit()
            if result2:
              TrdCd_Semu=result2[0]
              TrdNm_Semu=result2[1]
              strUpt = f"update  Financial_Trdst_Ecnt set TrdCd_Ecnt='{row[5].strip()}' where seq_no='{manageNo[0]}' and trdcd_semu='{TrdCd_Semu}'"#;print(strUpt)
              cursor.execute(strUpt)
          AcntCd_Semu = rs2["AcntCd_Semu"].strip()
          AcntNm_Semu = rs2["AcntNm_Semu"].strip()
          if TrdCd_Semu=="" : TrdNm_Semu=row[6].strip()
          CrDr = "";TranAmt_Cr = TranAmt_Dr = 0
          if row[9] is not None and row[9]!='':
            strtmp = str(row[9]).split(".");  str_amt = strtmp[0].replace(",","") 
            CrDr = "3";TranAmt_Dr = 0
            # 안전하게 숫자 부분만 추출
            clean_amt = str(str_amt).split('\n', 1)[0].strip()   # '13533874\n[9772]' -> '13533874'
            TranAmt_Cr = int(clean_amt) if clean_amt else 0
          if row[10] is not None and row[10]!='':
            strtmp = str(row[10]).split(".");  str_amt = strtmp[0].replace(",","")
            CrDr = "4";TranAmt_Cr = 0
            # 안전하게 숫자 부분만 추출
            clean_amt = str(str_amt).split('\n', 1)[0].strip()   # '13533874\n[9772]' -> '13533874'
            TranAmt_Dr = int(clean_amt) if clean_amt else 0
          row11 = row[11].replace("'","")
          str_ins = f"insert into Ds_slipledgr_Ecount values('{manageNo[0]}','{manageNo[1]}','{manageNo[2]}','{str(row[1]).strip()}"
          str_ins += f"','{AcntCd_Semu}','{AcntNm_Semu}','{Trn_Date}','{row11.strip()}','{TrdCd_Semu}','{TrdNm_Semu}','{Trn_Num}'"
          str_ins += f",'{int(row[3])}','{row[4].strip()}','{CrDr}',{int(TranAmt_Cr)+int(TranAmt_Dr)},'{TranAmt_Cr}','{TranAmt_Dr}'" 
          str_ins += f",'{int(row[7])}','{row[8].strip()}','{row[5].strip()}','{row[6].strip()}','{today}')" ;#print(str_ins)
          cursor.execute(str_ins)
        else:
          print(f"이카운트 계정코드 미등록으로 업로드 안됨 : {i},  sql :{strsql_trd}")
      else:
        print(f'{i}: 계정코드가 공란입니다 END OF FILE')
  elif flag=='연말정산대상근로자':
    try:
      df = pd.read_excel(file_name, header=0,dtype = {'주민등록번호':str})
    except:
      return False
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환    c.관리번호, trim(a.biz_name), a.seq_no
    for row in df.itertuples():
      sql = f"Merge 연말정산일괄제공근로자 as A Using (select '{manageNo[2]}' as seqno , '{manageNo[10]}' as work_yy, '{row[2]}' as ssn) as B "
      sql += "On A.seq_no = B.seqno and A.work_yy=B.work_yy and A.주민등록번호=B.ssn   "
      sql += f" when matched then update set 동의일자='{row[6]}',동의취소일자='{row[7]}',비고='조회일 : " + str(datetime.now())[:10].replace("-","") +"'"
      sql += f"when not matched then insert values('{manageNo[2]}','{manageNo[11]}','{row[1]}','{row[2]}','{row[3]}','{row[4]}','{row[5]}','{row[6]}','{row[7]}','0"   #PDFDownYN
      sql += "','" + "조회일 : " + str(datetime.now())[:10].replace("-","")  #비고 - 조회일
      sql += "'); "
      print(sql)
      cursor = connection.cursor()
      cursor.execute(sql)
      cursor.commit()
      print('연말정산일괄제공근로자 디비저장성공') 
  elif flag=='간이지급명세서':
    try:
      df = pd.read_excel(file_name, header=0)
    except:
      print('간이지급명세서 전자신고 결과 엑셀파일 읽기 실패')
      return False
    df = df.fillna('')  # 3. excel 의 nill 값을 '' 로 변환    c.관리번호, trim(a.biz_name), a.seq_no
    for row in df.itertuples():
      sql = "Merge 지급조서간이소득 as A Using (select '"+manageNo[0]+"' as 과세년도 , '"+manageNo[1]+"' as 신고서종류, '"+manageNo[2]+"' as 사업자번호) as B "
      sql += "On A.과세년도 = B.과세년도 and A.신고서종류=B.신고서종류 and A.사업자번호=B.사업자번호  "
      sql += "when matched then update set 신고구분='"+ str(row[8])+"',접수자='"+str(row[2])+"',제출건수='"+str(row[9])+"',제출금액='"+str(row[10])+"',접수번호='"+str(row[1])+"',접수일시='"+str(row[11])+"',작성일자=getdate() "
      sql += "when not matched then insert values('"+str(row[3])  #신고서종류
      sql += "','" + str(row[4])  #사업자번호
      sql += "','" + str(row[5])  #상호
      sql += "','" + str(row[6]) #과세년월
      sql += "','" + str(row[8]) #정기/기한후
      sql += "','" + str(row[2]) #접수자
      sql += "','" + str(row[9]) #제출건수
      sql += "','" + str(row[10]) #제출금액
      sql += "','" + str(row[11]) #접수번호
      sql += "','" + str(row[1]) #접수번호
      sql += "',getdate()" #접수일
      sql += "); "
      print(sql)
      cursor = connection.cursor()
      cursor.execute(sql)
      cursor.commit()
      print('간이지급명세서 전자신고 결과 디비저장성공') 

  if flag!='이카운트분개장업로드':
    # EXCEL.EXE 프로세스 강제 종료 (이미 없으면 조용히 무시)
    subprocess.run(
        ["taskkill", "/IM", "EXCEL.EXE", "/F"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL
    )
    os.remove(file_name)
  return True

#zip파일 압축해제
def Extract_Download_File(zip_path, destination_folder, password=None):
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    # Open the zip file
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        # If a password was provided, use it to extract the file
        if password:
            zip_ref.setpassword(password.encode())

        # Extract all the contents into the destination folder
        zip_ref.extractall(destination_folder)

    return True

# 파일 대화상자를 표시하는 함수
def select_file():
  root = tk.Tk()
  root.withdraw()  # Tkinter 창 숨기기
  file_path = filedialog.askopenfilename()  # 파일 대화상자 표시
  return file_path


# 날짜 열을 식별하기 위한 함수
def identify_date_column(df, pattern):
    """
    DataFrame에서 날짜 형식에 맞는 첫 번째 열의 이름을 반환합니다.
    :param df: 검색할 pandas DataFrame
    :param pattern: 날짜 형식의 정규 표현식
    :return: 날짜 형식에 맞는 첫 번째 열의 이름 또는 None
    """
    for col in df.columns:
        # 현재 열의 첫 번째 값만 검사하여 성능을 최적화할 수 있습니다.
        if pd.notna(df[col].iloc[0]) and re.match(pattern, str(df[col].iloc[0])):
            return col
    return None

# 먼저 연월을 나타내는 숫자 패턴(\d{4} 또는 \d{1,2}월)을 제거하고, 남은 문자열에서 원하는 텍스트만을 반환
def remove_dates_and_extract_text(input_string):
    # 먼저 "월"을 포함하는 패턴을 찾아 제거합니다. 예: "03월"
    input_string = re.sub(r'\d{1,2}월', '', input_string)

    # 다음으로 4자리 숫자 패턴을 찾아 제거합니다. 예: "2203" 또는 "2204"
    input_string = re.sub(r'\d{4}', '', input_string)

    # 최종 결과에서 좌우 공백을 제거합니다.
    result = input_string.strip()

    return result

#문장에서 검색할 word부터 30까지의 단어를 반환
def retrieve_words_from_end(sentence, matching_word):
    start_pos = sentence.find(matching_word)
    
    # Check if the matching word was found
    if start_pos == -1:
        return "Matching word not found in the sentence."
    
    # Calculate the end position of the matching word
    end_pos = start_pos + len(matching_word)
    
    # Extract up to 30 characters from the end of the matching word
    result_string = sentence[end_pos:end_pos + 30]
    
    # Convert the list of words back into a string

    retrieved_words = result_string.strip()
    
    return retrieved_words

def fnStrLength(s):
    """Calculate the byte length of a string assuming UTF-8 encoding."""
    return len(s.encode('utf-8'))

def fnStrLengthCut(s, x):
    """Cut a string to a specific byte length."""
    byte_count = 0
    cut_str = ""
    for char in s:
        char_byte_len = len(char.encode('utf-8'))
        if byte_count + char_byte_len > x:
            break
        cut_str += char
        byte_count += char_byte_len
    return cut_str

def mid_union(s, start, end):
    """Extract a substring based on byte length, similar to Mid in VBScript but based on bytes."""
    total_str = fnStrLengthCut(s, start - 1 + end)
    left_str = fnStrLengthCut(s, start - 1)
    right_str = total_str[len(left_str):]
    return right_str

def count_korean_characters(sentence):
    """Counts the number of Korean characters in a given sentence."""
    def is_korean(char):
        """Checks if a character is Korean."""
        return ord('가') <= ord(char) <= ord('힣') or \
               ord('ㄱ') <= ord(char) <= ord('ㅎ') or \
               ord('ㅏ') <= ord(char) <= ord('ㅣ')
    
    return sum(1 for char in sentence if is_korean(char))


def contains_only_numbers(proposition):
    # Use regular expression to match only numbers
    return re.match(r'^\d+$', proposition) is not None


# Pyautogui로 결과가 나올때까지 대기하는 함수
def wait_for_value(driver, xpath,  timeout=30):
    start_time = time.time()
    while time.time() - start_time < timeout:
        # Try to get the element and check its text
        try:
          element = WebDriverWait(driver, 10).until(
              EC.visibility_of_element_located((By.XPATH, xpath))
          )
          if  element.text != "0":
            print("결과 :", element.text)
            return True
        except Exception as e:
            print(f"{time.time() - start_time} 결과대기중:", e)
        time.sleep(1)  # wait a bit before trying again to avoid high CPU usage
    return False

# 쿼리 실행 함수 (재시도 로직 포함)
def execute_query_with_retry(query):
  delayTime=0.1
  if query[:4]=="exec":
     delayTime=0.5
  try:
      with connection.cursor() as cursor:
          # 트랜잭션 격리 수준 설정
          cursor.execute("SET TRANSACTION ISOLATION LEVEL READ COMMITTED")
          connection.autocommit = False
          cursor.execute(query)
          time.sleep(delayTime)#중요
          connection.commit()
          print(f"프로시저 성공 : {query}")
  except Exception as e:
      connection.rollback()
      print(f"트랜잭션 롤백 에러 : {e}")
